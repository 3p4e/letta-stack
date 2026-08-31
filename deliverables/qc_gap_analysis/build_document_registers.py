#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""The three QC document registers in one workbook: eCoA receipt, iCoA issuance, CoQ issuance.

    python3 deliverables/qc_gap_analysis/build_document_registers.py [OUT.xlsx]

**eCoA receipt** — every certificate received from an outside laboratory, in issue-date
order, with its provenance, what it reports, whether it has been read against its own page,
and a live link. This is the document-control view of the release register: it answers *what
did we receive, from whom, when, and has anyone checked it*.

**iCoA issuance** — the in-house CoA owed for each batch and what it must cover.

**CoQ issuance** — every Certificate of Quality predicted, initial release and reissue, with
the iCoA it references and what still blocks it.

## The cultivation-batch / packaged-lot linkage

`PP_Potency_MASTER_Spec.xlsx` carries a column the release register does not: **Cultiv. Batch
No.** beside **PP Batch No.** It shows that seven register entries named by P-number are
packaged lots of cultivation batches, and for four of them **the cultivation batch is a
separate register entry of its own**:

| Packaged lot | ref | drawn from | ref |
|---|---|---|---|
| `P060152` | 72 | `J31102501` | 43 |
| `P060212` | 73 | `JD112501` | 58 |
| `P060242` | 74 | `OPM122501` | 51 |
| `P060402` | 78 | `GG012603` | 62 |

Two more — `P060352` from `FB012602`, `P060382` from `SCR012603` — are packaged lots whose
cultivation batch is not in the release register at all. `P060332` appears in neither.

This is what those seven were: they carry only the Farmahem 197-series because that is the
analysis performed **at packaging**, and their earlier testing sits under the cultivation
batch's own entry. It also means the register counts the same material twice for four of
them, which the CoQ count turns on: whether a CoQ is issued per cultivation batch or per
released lot is a QC decision, and both totals are given rather than one chosen.

The linkage is carried in a `packaged lot of` / `same material as ref` column on every sheet
so it is visible wherever a count is read.
"""
import csv
import json
import os
import re
import sys

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from collections import Counter

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

REG_X = "deliverables/qc_gap_analysis/PP_Batch_Release_QC_Register_FHM3_2026-08-31.xlsx"
ICOA = "deliverables/qc_gap_analysis/icoa_issuance_register_2026-08-31.csv"
MASTER = "deliverables/qc_gap_analysis/potency_master_batch_map.json"
SHEET = "Batch Release QC"

INK, SUB = "FF16232B", "FF5A6E75"
BAND, SUBBAND = "FF16232B", "FFE7ECEA"
ZEBRA, AMBER, RED, GREEN, WHITE = "FFF6F8F7", "FFFAEDD4", "FFF9DEDB", "FFE6F1EB", "FFFFFFFF"
THIN = Side(style="thin", color="FFD6DEDB")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def key(d):
    m = re.match(r"(\d{2})\.(\d{2})\.(\d{4})", str(d or ""))
    return (m.group(3) + m.group(2) + m.group(1)) if m else "99999999"


def nk(s):
    return re.sub(r"[^A-Z0-9]", "", str(s).upper())


def band(ws, cols, title, subtitle):
    ws.cell(row=1, column=1, value=title).font = Font("Arial", 15, bold=True, color=INK)
    ws.cell(row=2, column=1, value=subtitle).font = Font("Arial", 9, color=SUB)
    for i, (name, w, note) in enumerate(cols, 1):
        h = ws.cell(row=4, column=i, value=name)
        h.font = Font("Arial", 9, bold=True, color="FFFFFFFF")
        h.fill = PatternFill("solid", fgColor=BAND)
        h.alignment = Alignment("center", "center", wrap_text=True)
        s = ws.cell(row=5, column=i, value=note)
        s.font = Font("Arial", 8, bold=True, color=INK)
        s.fill = PatternFill("solid", fgColor=SUBBAND)
        s.alignment = Alignment("center", "center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 30
    ws.row_dimensions[5].height = 26


def put(ws, row, vals, zebra, centre=(), wrap=(), fills=None):
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = Font("Arial", 10, color=INK)
        cell.border = BOX
        cell.alignment = Alignment(vertical="center",
                                   horizontal="center" if c in centre else "left",
                                   wrap_text=c in wrap)
        f = (fills or {}).get(c)
        cell.fill = PatternFill("solid", fgColor=f if f else (ZEBRA if zebra else WHITE))


def load_register():
    ws = load_workbook(REG_X)[SHEET]
    cols = {c: (ws.cell(row=4, column=c).value or "", ws.cell(row=5, column=c).value or "")
            for c in range(5, 23)}
    rows, cur = [], None
    for r in range(6, 292):
        ref, b = ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value
        if ref not in (None, "") and b:
            cur = dict(ref=str(ref), batch=str(b),
                       pn=str(ws.cell(row=r, column=3).value or ""),
                       strain=str(ws.cell(row=r, column=4).value or ""))
        code = ws.cell(row=r, column=23).value
        if not code or not cur:
            continue
        reported = [cols[c][0] for c in range(5, 23)
                    if str(ws.cell(row=r, column=c).value or "").strip() not in ("", "/", "None")]
        flags = []
        for c in [2] + list(range(5, 23)) + [23]:
            rgb = getattr(getattr(ws.cell(row=r, column=c).fill, "start_color", None), "rgb", None)
            if rgb == AMBER and "amber" not in flags:
                flags.append("amber")
            if rgb == RED and "red" not in flags:
                flags.append("red")
        rows.append(dict(row=r, **cur, code=str(code).strip(),
                         date=str(ws.cell(row=r, column=24).value or ""),
                         lab=str(ws.cell(row=r, column=25).value or ""),
                         pdf=getattr(ws.cell(row=r, column=26).hyperlink, "target", None) or "",
                         reported=reported, flags=flags))
    # The CNP full Ph. Eur. form certifies identity and foreign matter too,
    # but those determinations have no register column, so a column-derived
    # "parameters reported" list silently drops them. Append them from the
    # documents' own table text.
    ff = full_form_cnp()
    if ff:
        for r in rows:
            if any(nk(x) in ff for x in re.findall(r"ППК\s*\d+", r["code"])):
                r["reported"] = r["reported"] + ["Ident A + B", "Foreign matter"]
    return rows


def full_form_cnp():
    """Certificate codes of the CNP full Ph. Eur. form — the certificates whose
    own results table prints Идентификација (Макроскопија · Микроскопија) and
    Страни материи. The release register has no column for identity or foreign
    matter, so a receipt row derived from columns alone reads these documents
    as potency-only — "not noticed", as the owner put it. Detected from each
    document's own table text in the vendored corpus cache; 12 as of
    31.08.2026 (ППК26110–26115, 26116–26119 for P160012/22/32 and SCR022601,
    26127, 26128)."""
    import re as _re
    path = os.path.join(ROOT, "ingestion", "ragflow", "cache",
                        "all_cert_texts_2026-08-30.json")
    if not os.path.exists(path):
        return set()
    out = set()
    for rec in json.load(open(path, encoding="utf-8")):
        if "CNP" not in rec["name"]:
            continue
        t = rec["text"]
        if _re.search(r"дентифика", t) and _re.search(r"акроскоп", t):
            m = _re.search(r"ППК\s*\d+", rec["name"])
            if m:
                out.add(nk(m.group(0)))
    return out


def verified_map():
    """folded certificate code -> verification block, from the page-read records.

    Keyed on verification_coverage.fold(), never nk(): nk() strips every
    non-ASCII character, so the Cyrillic К and М that distinguish a Farmahem
    cannabinoid certificate from its mycotoxin sibling vanish, ГС never folds
    to LOD, and a trailing bracketed note defeats the match. That produced 59
    false "not verified" receipts (56 Farmahem + 3 IPH) out of 74."""
    import glob
    import verification_coverage as VC
    out = {}
    for p in sorted(glob.glob("review/*_page_reads_*.json")):
        blk = p.split("/")[-1].split("_page_reads")[0]
        for c in json.load(open(p, encoding="utf-8")):
            out[VC.fold(c)] = blk
    return out


def page_verified(code, ver):
    """The verification block whose page reads cover this certificate, or ''."""
    import verification_coverage as VC
    for c in VC.candidates(code):
        if c in ver:
            return ver[c]
    return ""


def main(out):
    master = json.load(open(MASTER, encoding="utf-8"))
    pp2cult = {nk(b): a for a, b in master["pairs"]}
    rows = load_register()
    ver = verified_map()
    ic = list(csv.DictReader(open(ICOA, encoding="utf-8")))
    refof = {nk(r["batch"]): r["register_ref"] for r in ic}

    def lot(batch):
        """'packaged lot of X (ref n)' — only where the material was actually repackaged.

        The potency master maps every batch to a PP number, and for material never
        repackaged that number IS the batch (BG1024 -> BG1024). Reporting those as a
        linkage would put "packaged lot of BG1024" on BG1024 and bury the seven rows
        that say something.
        """
        c = pp2cult.get(nk(batch))
        if not c or nk(c) == nk(batch):
            return ""
        r = refof.get(nk(c))
        return f"packaged lot of {c}" + (f" — also register ref {r}" if r else " — cultivation batch not in the register")

    wb = Workbook()

    # ---------------------------------------------------------------- 1 · eCoA receipt
    ws = wb.active
    ws.title = "eCoA Receipt"
    COLS = [("Seq", 6, "issue order"), ("Date of issue", 13, "as printed on the document"),
            ("Laboratory", 26, ""), ("Certificate code", 22, "as printed"),
            ("Batch", 16, ""), ("Reg. ref", 8, ""), ("Strain", 19, ""), ("P-number", 11, ""),
            ("Parameters reported", 44, "columns this document fills in the release register"),
            ("Page-verified", 14, "read off its own page, 31.08.2026"),
            ("Flag", 9, "amber / red per the register legend"),
            ("Packaged-lot linkage", 34, "from PP_Potency_MASTER_Spec"),
            ("Register row", 11, ""), ("Document link", 15, "")]
    band(ws, COLS, "Purely Plant GmbH — certificate receipt register",
         "Every certificate the release register cites, in issue-date order — outsourced "
         "and, where the batch has one, Purely Plant's own in-house CoA. Document control: "
         "what was received, from whom, when, and whether it has been read off its page.")
    docs = [r for r in rows if not r["code"].lower().startswith(("n/a", "(not numbered)"))]
    docs.sort(key=lambda r: (key(r["date"]), r["row"]))
    # One document, one row. The register cites PP CoA #033 on two rows of GP092501's
    # block, one dated 20.02.2026 and one undated; a receipt register that lists the
    # same certificate twice with different dates cannot answer its own question.
    # Keep the dated row and record the duplicate rather than dropping it silently.
    # Keyed on the code EXACTLY as printed, never on nk(): nk() strips every
    # non-ASCII character, and the Farmahem series distinguishes its cannabinoid
    # certificate from its mycotoxin certificate by a single Cyrillic letter —
    # 197-1-К/26 against 197-1-М/26. Normalising merged all 21 pairs into 21 rows
    # and lost 21 real documents before this was caught.
    seen, dupes, uniq = {}, [], []
    for r in docs:
        k = r["code"].strip()
        if k in seen:
            dupes.append((r, seen[k]))
            continue
        seen[k] = r
        uniq.append(r)
    docs = uniq
    for n, r in enumerate(docs):
        i = 6 + n
        v = page_verified(r["code"], ver)
        fl = "red" if "red" in r["flags"] else ("amber" if r["flags"] else "")
        put(ws, i, [n + 1, r["date"] or "no date", r["lab"], r["code"], r["batch"], r["ref"],
                    r["strain"], r["pn"], ", ".join(r["reported"]) or "—",
                    v or "not verified", fl or "—", lot(r["batch"]), r["row"],
                    "Open" if r["pdf"] else "—"],
            n % 2 == 0, centre=(1, 6, 10, 11, 13, 14), wrap=(9, 12),
            fills={10: GREEN if v else WHITE,
                   11: RED if fl == "red" else (AMBER if fl else WHITE)})
        if r["pdf"]:
            c = ws.cell(row=i, column=14)
            c.hyperlink, c.font = r["pdf"], Font("Arial", 10, color="FF2B4899", underline="single")
    ws.freeze_panes = "E6"
    ws.auto_filter.ref = f"A5:N{5+len(docs)}"

    # ---------------------------------------------------------------- 2 · iCoA issuance
    ws = wb.create_sheet("iCoA Issuance")
    COLS = [("Seq", 6, "release order"), ("Release date", 13, "latest non-197 certificate"),
            ("Reg. ref", 8, ""), ("Batch", 16, ""), ("Strain", 19, ""),
            ("iCoA scope", 14, ""), ("Ident A", 14, "macroscopy = appearance, 2.8.23"),
            ("Ident B", 14, "microscopy, 2.8.23"), ("Ident C", 14, "TLC, 2.8.23"),
            ("Foreign matter", 15, "2.8.2"), ("CoQ initial", 10, ""), ("CoQ reissue", 10, ""),
            ("Outsourced outstanding", 24, "CoQ waits on this too"),
            ("Packaged-lot linkage", 34, "from PP_Potency_MASTER_Spec"),
            ("iCoA number", 15, "fill in on issue"), ("iCoA date", 12, "fill in on issue"),
            ("Issued by", 15, "fill in on issue"), ("Reference in CoQ", 18, "fill in when added")]
    band(ws, COLS, "Purely Plant GmbH — in-house CoA issuance register",
         "One iCoA per batch, in release order. Each row names the CoQ documents that will "
         "reference it and the parameters it must carry.")
    for n, r in enumerate(ic):
        i = 6 + n
        failed = r["batch"] == "FB032601"
        fm = r["foreign_matter"] + (" — FAILED" if failed else "")
        put(ws, i, [int(r["seq"]), r["release_date"] or "no date", r["register_ref"], r["batch"],
                    r["strain"], "Ident C only" if r["ident_A"] != "required" else "Full panel",
                    r["ident_A"], r["ident_B"], r["ident_C"], fm, "yes", r["coq_reissue"],
                    r["outsourced_outstanding"], lot(r["batch"]), "", "", "", ""],
            n % 2 == 0, centre=(1, 3, 7, 8, 9, 10, 11, 12), wrap=(13, 14),
            fills={c: (RED if (failed and c == 10) else
                       (AMBER if [r["ident_A"], r["ident_B"], r["ident_C"], fm][c - 7] == "required"
                        else GREEN)) for c in (7, 8, 9, 10)}
                  | {c: WHITE for c in (15, 16, 17, 18)})
        if failed:
            ws.cell(row=i, column=10).font = Font("Arial", 10, bold=True, color="FF9E2A2A")
    ws.freeze_panes = "E6"
    ws.auto_filter.ref = f"A5:R{5+len(ic)}"

    # ---------------------------------------------------------------- 3 · CoQ issuance
    #
    # The CoQ universe (owner, 31.08.2026): one initial CoQ per batch on record and
    # a 12-month cannabinoid + mycotoxin reissue for every batch — 61 carry numbers
    # from the ISSUE_COQ plan (48 initial for Tranche 01 + 02, 13 additional
    # testing), the rest are predicted and numbered on issue. No CoQ prints an issue
    # date before the SOP's in-use date, 11.05.2026. The coverage columns come from
    # build_coq_schedule.schedule() so this sheet and the parameter schedule cannot
    # drift apart.
    import build_coq_schedule as CQ
    sched, cov, dets = CQ.schedule()
    scount = Counter(r["Status"] for r in sched)

    ws = wb.create_sheet("CoQ Issuance")
    COLS = [("CoQ number", 17, "ISSUE_COQ numbering"), ("CoQ type", 20, ""),
            ("Issue date", 13,
             "set on the day of issue — no earlier than shown, never post-dated"),
            ("Packaged lot", 12, ""), ("Cultivation batch", 14, ""),
            ("Strain", 19, ""), ("Grade", 7, ""), ("Class", 8, ""),
            ("Banner THC", 11, "actual assay, per the plan"),
            ("Acceptance range", 28, "grade nominal ± tolerance"),
            ("iCoA reference", 16, "printed on the CoQ"),
            ("Covered", 9, "of 21 on an initial CoQ, 10 in the retest scope"),
            ("To perform", 10, ""),
            ("Not tested", 10, ""), ("Out of spec", 10, ""),
            ("Documents", 10, ""),
            ("Source documents", 66, "every report this CoQ must cite"),
            ("Who performs the rest", 52, ""), ("Note", 40, "")]
    band(ws, COLS, "Purely Plant GmbH — Certificate of Quality issuance register",
         "One initial CoQ per batch on record and a 12-month reissue for every batch "
         "(owner, 31.08.2026). 61 carry ISSUE_COQ numbers — 48 initial release for "
         "Tranche 01 + 02, 13 additional testing — and the rest are predicted, "
         "numbered on issue. Issue dates are set at issue, no earlier than the CoQ "
         "SOP's in-use date of 11.05.2026. Detail per determination in "
         "PP_CoQ_Parameter_Schedule_2026-08-31.xlsx.")
    for n, c in enumerate(cov):
        i = 6 + n
        k = c["counts"]
        covered = (k[CQ.ST_OK] + k[CQ.ST_FINDING] + k[CQ.ST_OOS] + k[CQ.ST_UNDET]
                   + k[CQ.ST_BLOCK] + k[CQ.ST_OFFREG])
        oos = k[CQ.ST_OOS] + k[CQ.ST_BLOCK]
        note = c["spec_conflict"]
        if not c["in_register"]:
            note = ("No eCoA on file for this cultivation batch — locate the physical "
                    "certificates, scan and upload. " + note).strip()
        put(ws, i, [c["number"], c["type"], c["date"], c["pp"], c["cb"], c["strain"],
                    c["grade"], f"THC {c['cls']}", c["banner_thc"], c["thc"],
                    c["icoa_ref"], covered,
                    k[CQ.ST_ICOA] + k[CQ.ST_SCAN]
                    + k[CQ.ST_AWAIT_K] + k[CQ.ST_AWAIT_M],
                    k[CQ.ST_NONE] + k[CQ.ST_NOSPEC], oos, len(c["codes"]),
                    "; ".join(c["codes"]),
                    "; ".join(dict.fromkeys(c["route"].values())) or "—", note],
            n % 2 == 0, centre=(7, 8, 9, 12, 13, 14, 15, 16), wrap=(10, 17, 18, 19),
            fills=({15: RED} if oos else {})
                  | ({19: RED} if not c["in_register"] else
                     ({19: AMBER} if note else {})))
    ws.freeze_panes = "D6"
    ws.auto_filter.ref = f"A5:S{5+len(cov)}"
    coq = cov  # the summary below counts over the plan's CoQs

    # ---------------------------------------------------------------- 4 · summary
    s = wb.create_sheet("Summary & Notes")
    s.cell(row=1, column=1, value="The three registers, and what they count").font = Font("Arial", 15, bold=True, color=INK)
    s.cell(row=2, column=1, value="Generated 31.08.2026 from the corrected release register, "
           "the iCoA gap analysis, and PP_Potency_MASTER_Spec.xlsx").font = Font("Arial", 9, color=SUB)
    lines = [
        ("Certificate receipt register", len(docs),
         "distinct certificates the release register cites — %d of them Purely Plant's "
         "own in-house CoA%s" % (
             sum(1 for r in docs if r["lab"].startswith("Purely")),
             (" · %d duplicate row%s removed: %s" % (
                 len(dupes), "" if len(dupes) == 1 else "s",
                 ", ".join(sorted({d[0]["code"] for d in dupes})))) if dupes else "")),
        ("  page-verified against their own page", sum(1 for r in docs if nk(r["code"]) in ver or
            any(nk(m) in ver for m in re.findall(r"ППК\s*\d+", r["code"]))), "31.08.2026 campaign"),
        ("  carrying a live document link", sum(1 for r in docs if r["pdf"]), ""),
        ("", "", ""),
        ("iCoA issuance register", len(ic),
         "one row per register entry — NOT one per material: six rows (P060152, "
         "P060212, P060242, P060352, P060382, P060402) are the Farmahem 12-month "
         "re-analyses of plan lots that also appear under their cultivation batch"),
        ("  full panel — Ident A + B + C + foreign matter", sum(1 for r in ic if r["ident_A"] == "required"), ""),
        ("  Ident C only — CNP Ph. Eur. 11.5 form covers the rest", sum(1 for r in ic if r["ident_A"] != "required"), ""),
        ("", "", ""),
        ("CoQ issuance register", len(coq), "one initial CoQ per batch on record + "
         "a 12-month reissue for every batch; 61 numbered by the ISSUE_COQ plan, "
         "the rest predicted — issue dates from 11.05.2026 (SOP in use) onward"),
        ("  issued — numbered by the ISSUE_COQ plan",
         sum(1 for x in coq if x["issued"]),
         "48 initial (Tranche 01: 19, Tranche 02: 29) + 13 additional testing"),
        ("  predicted initial — batches past Tranche 02",
         sum(1 for x in coq if x["type"] == "initial release — predicted"),
         "no packaged lot or grade assigned yet; number copied at issue"),
        ("  predicted reissue — the remaining Tranche 01/02 lots",
         sum(1 for x in coq if x["type"].startswith("additional")
             and not x["issued"] and x["pp"]),
         "every one of the 48 retests cannabinoids and mycotoxins"),
        ("  predicted reissue — batches past Tranche 02",
         sum(1 for x in coq if x["type"].startswith("additional")
             and not x["issued"] and not x["pp"]),
         "release + 12 months"),
        ("  carrying an out-of-specification determination",
         sum(1 for x in coq if x["counts"][CQ.ST_OOS] + x["counts"][CQ.ST_BLOCK]), ""),
        ("  no eCoA on file for the cultivation batch",
         sum(1 for x in coq if not x["in_register"]),
         "locate the physical certificates, scan and upload"),
        ("  re-analysed ahead of the 12-month date",
         sum(1 for x in coq if x["type"].startswith("additional") and not x["issued"]
             and x["counts"][CQ.ST_OK] + x["counts"][CQ.ST_FINDING]
             + x["counts"][CQ.ST_OOS] + x["counts"][CQ.ST_UNDET]),
         "the 197-series pair is on file — cannabinoids and mycotoxins certified; "
         "identity and foreign matter still to perform before any of them can issue"),
        ("  initial CoQs whose banner potency is the re-analysis, not the release "
         "assay", 19,
         "the master spec carries the newest assay; defensible only because no CoQ "
         "issues before 11.05.2026 — a QC decision to take knowingly"),
        ("", "", ""),
        ("In-house certificates the routing requires", len(CQ.icoa_plan(cov)),
         "one Ident A + B and one foreign-matter iCoA per initial CoQ where no "
         "outsourced certificate covers them; one foreign-matter iCoA per reissue"),
        ("", "", ""),
        ("", "", ""),
        ("CoQ parameter schedule", len(sched),
         f"{len(coq)} CoQs x 23 determinations"),
        ("  covered by a certificate on file", scount[CQ.ST_OK] + scount[CQ.ST_FINDING]
         + scount[CQ.ST_OOS] + scount[CQ.ST_UNDET] + scount[CQ.ST_BLOCK]
         + scount[CQ.ST_OFFREG], "conforming, findings, OOS, undetermined and blocked alike"),
        ("  still to be performed", scount[CQ.ST_ICOA],
         "identity A, B, C and foreign matter — routed per the Sourcing routes sheet"),
        ("  awaiting the 12-month re-analysis — Farmahem",
         scount[CQ.ST_AWAIT_K] + scount[CQ.ST_AWAIT_M],
         "cannabinoids + mycotoxins on every not-yet-re-analysed batch"),
        ("  outside the retest scope",
         sum(v for k2, v in scount.items()
             if k2.startswith("outside the retest scope")),
         "on a reissue, the release determination stands on the initial CoQ"),
        ("  not tested by anyone", scount[CQ.ST_NONE], ""),
        ("  criterion unstatable — no product specification", scount[CQ.ST_NOSPEC],
         "parameter 4's criterion is 'per target grade as per Section 01'"),
        ("  upon request — not required for release", scount[CQ.ST_REQ],
         "P. aeruginosa and S. aureus"),
        ("  in-house CoA only — no outsourced eCoA on file", scount[CQ.ST_SCAN],
         "GG1024: locate the physical certificates, scan and upload"),
        ("     of the covered, out of specification", scount[CQ.ST_OOS], ""),
        ("     of the covered, undetermined pending QCSP 001", scount[CQ.ST_UNDET], ""),
        ("     of the covered, a conforming laboratory finding", scount[CQ.ST_FINDING], ""),
        ("     of the covered, declared out of specification by the laboratory",
         scount[CQ.ST_BLOCK], ""),
        ("     of the covered, batch not in the release register",
         scount[CQ.ST_OFFREG], ""),
        ("", "", ""),
        ("Determinations no laboratory has ever performed", "",
         f"per CoQ, over {len(coq)} CoQs"),
    ] + [("  " + d["en"], sum(1 for r in sched if r["Parameter"] == d["en"]
                             and r["Status"] in (CQ.ST_NONE, CQ.ST_ICOA)),
          "QCSP 001 no. " + d["no"])
         for d in dets if sum(1 for r in sched if r["Parameter"] == d["en"]
                              and r["Status"] in (CQ.ST_NONE, CQ.ST_ICOA)) >= 60] + [
        ("", "", ""),
        ("CoQs whose grade range disagrees with the QCSP 001 PDF",
         sum(1 for c in cov if c["spec_conflict"]),
         "two owner documents, two grade designs — recorded, not resolved"),
    ]
    for i, (a, b, c) in enumerate(lines, 4):
        s.cell(row=i, column=1, value=a).font = Font("Arial", 10, bold=not a.startswith("  ") and bool(a), color=INK)
        cb = s.cell(row=i, column=2, value=b)
        cb.font = Font("Arial", 11, bold=True, color=INK)
        cb.alignment = Alignment(horizontal="center")
        s.cell(row=i, column=3, value=c).font = Font("Arial", 9, color=SUB)
    s.column_dimensions["A"].width = 52
    s.column_dimensions["B"].width = 9
    s.column_dimensions["C"].width = 74
    notes = [
        "THE CULTIVATION-BATCH / PACKAGED-LOT LINKAGE, and why the CoQ total depends on it.",
        "PP_Potency_MASTER_Spec.xlsx carries a column the release register does not: Cultiv. Batch No. "
        "beside PP Batch No. Seven register entries named by a P-number are packaged lots, and for four "
        "of them the cultivation batch is a separate register entry of its own: P060152 (ref 72) from "
        "J31102501 (ref 43); P060212 (73) from JD112501 (58); P060242 (74) from OPM122501 (51); "
        "P060402 (78) from GG012603 (62).",
        "That is what those seven were. They carry only the Farmahem 197-series because that is the "
        "analysis performed at packaging; their earlier testing sits under the cultivation batch's own "
        "entry. Two more — P060352 from FB012602, P060382 from SCR012603 — are packaged lots whose "
        "cultivation batch is not in the release register at all. P060332 appears in neither document.",
        "So for four batches the register counts the same material twice. THAT QUESTION IS NOW "
        "SETTLED: the owner's ruling of 31.08.2026 issues one CoQ per released material, and the six "
        "P-number blocks fold into their cultivation batch rather than standing as batches of their "
        "own. The CoQ sheet counts 166 documents on that basis — one initial per batch on record and "
        "one 12-month reissue per batch. The linkage is still carried on every sheet, because the "
        "register itself has not been renumbered.",
        "",
        "WHY THIS SHEET STILL LISTS 81 ROWS WHILE THE ROUTING REQUIRES 224 IN-HOUSE CERTIFICATES. "
        "The 81 rows predate the plan and are kept as the per-register-entry view — and six of them "
        "(P060152, P060212, P060242, P060352, P060382, P060402) are the Farmahem re-analysis rows of "
        "plan lots, so 81 rows are not 81 materials. What the routing actually requires is on the CoQ "
        "sheets and in PP_CoQ_Parameter_Schedule_2026-08-31.xlsx: 71 Ident A + B certificates and 153 "
        "foreign-matter certificates, 224 in all.",
        "AN EARLIER NOTE HERE SAID THE OPPOSITE AND IS WITHDRAWN. It read: \"Identity and foreign "
        "matter are properties of the material, determined once. A CoQ reissued because the "
        "cannabinoids and mycotoxins were re-analysed covers the same batch, so it references the "
        "iCoA already issued.\" The owner's routing of 31.08.2026 rules otherwise: at a 12-month "
        "reissue Farmahem performs identity together with the assay, and Purely Plant's laboratory "
        "issues a fresh foreign-matter iCoA. A certificate from the release round cannot stand behind "
        "a determination on a document dated a year later — the same rule that stops an initial CoQ "
        "citing the 197-series re-analysis, running the other way.",
        "WHERE THE SCOPE SPLIT COMES FROM. CNP changed its certificate form in mid-2026: the older DAB "
        "form reports loss on drying and cannabinoids only; the Ph. Eur. 11.5 form adds identification "
        "(macroscopy and microscopy) and foreign matter. Twelve certificates are on the newer form — "
        "ППК26110-26119, ППК26127, ППК26128 — and all 73 CNP certificates were read off their own pages "
        "on 31.08.2026, so this is not inferred from a filename or a parse.",
        "IDENTIFICATION C is on every iCoA because no laboratory has ever performed it. Those twelve "
        "Ph. Eur. pages print Идентификација — Макроскопија, Микроскопија and stop. It is discharged "
        "in-house by risk analysis with scientific justification, from the HPLC assay.",
        "FB032601 needs Ident C only, like its eleven siblings, and its CoQ still cannot be issued: the "
        "foreign matter CNP does cover reads 0.08 % against a 2.00 % maximum and is marked Не одговара, "
        "cannabis seed present. A blocker, not a gap.",
        "THE BLANK COLUMNS on the right of each sheet are where issuance is recorded — number, date, "
        "issuer, and the reference the CoQ gives it.",
    ]
    r0 = 4 + len(lines) + 2
    s.cell(row=r0, column=1, value="NOTES").font = Font("Arial", 9, bold=True, color=INK)
    for i, t in enumerate(notes, 1):
        cell = s.cell(row=r0 + i, column=1, value=t)
        cell.font = Font("Arial", 9, bold=t.isupper() and len(t) < 90, color=INK if t else SUB)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        s.merge_cells(start_row=r0 + i, start_column=1, end_row=r0 + i, end_column=3)
        if t:
            s.row_dimensions[r0 + i].height = max(14, 12 * (len(t) // 116 + 1))

    wb.save(out)
    print(f"{out}")
    print(f"  eCoA Receipt    {len(docs):>4} certificates")
    print(f"  iCoA Issuance   {len(ic):>4} in-house CoAs")
    print(f"  CoQ Issuance    {len(coq):>4} CoQ documents")
    print(f"  Summary & Notes")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1] if len(sys.argv) > 1 else
                  "deliverables/qc_gap_analysis/PP_QC_Document_Registers_2026-08-31.xlsx"))
