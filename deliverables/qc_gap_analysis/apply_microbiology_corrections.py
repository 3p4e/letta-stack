#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Apply the ten microbiology corrections found by reading all forty pages.

Every change is backed by a rendered IJZ-MB certificate page at 300 DPI, recorded
in `review/MICROBIOLOGY_PAGE_VERIFICATION_2026-08-31.md` with the raw reads in
`review/microbiology_page_reads_2026-08-31.json`. Nothing here comes from the RAG
corpus or from a text layer: Drive's extraction strips every exponent from these
scans and `pdftotext` returns nothing at all, so the render is the only admissible
source and a single superscript decides pass from fail.

    python3 deliverables/qc_gap_analysis/apply_microbiology_corrections.py IN.xlsx OUT.xlsx

Idempotent, and it refuses to touch a workbook it does not recognise: every change
compares the current cell against the value it was verified against and stops on any
mismatch rather than guessing. Row numbers are those of the LINKS workbook of
31.08.2026 — i.e. **after** `add_ppk25139_and_codes.py` has run. That script inserts a
row and shifts everything below 46, so it must be run before this one, never after.

Three kinds of change, and the first is the only one that moves a number:

**Row 26 was carrying the wrong certificate.** Two microbiology certificates exist for
`GP0824-01`, both issued 14.04.2025 — `318/0585/25` (TAMC 4,2×10³, TYMC 6×10³) and
`319/0586/25` (TAMC 5,2×10³, TYMC 8,8×10³). The row's TAMC matches `318` exactly, so
the row holds `318`'s results under `319`'s code, and `318`'s TYMC lost a decade on the
way in. Both halves are fixed here, and the PDF link is repointed with them: it opened
`319`, which after the code fix would be a live link to the wrong document — the failure
mode this repository has already had to clean up once across 190 rows.

`319/0586/25` is then a real certificate for a released batch with no row of its own.
Adding it is an addition rather than a correction, and it stays an open question for QC;
it is deliberately **not** done here.

**Four bile-tolerant GNB entries.** Two of them, rows 92 and 142, are impossible as
written — nothing is both under and over 10¹, and nothing is both under 10 and over 10³.
A range check at data entry would have caught both. The laboratory's own bracket is
restored, in each row's existing caret or superscript notation.

**Four E. coli entries recorded as counts.** Rows 148, 156, 171 and 177 record `< 10`
where every certificate reports `Одговара` against a specification of `отсутна/g` —
absent in 1 g, a presence/absence determination that cannot produce a count. Not a
safety finding, since the batches complied, but the register states something the
laboratory did not, in a column a reviewer reads as a measurement. The replacement
matches each row's own Salmonella cell so the row stays internally consistent.

Two comments are attached rather than columns added. `1220/2171/25` and `1221/2172/25`
run a tighter manufacturer specification than the column header (TYMC 10² against 10⁴)
and test *S. aureus* and *P. aeruginosa*, for which the register has no columns. Row 122
is already amber from the 29.08 pass; the comment records **why** its 200 is a finding
when the column says the limit is 10⁴.
"""
import sys

from openpyxl import load_workbook
from openpyxl.comments import Comment

SHEET = "Batch Release QC"
TYMC, GNB, ECOLI = 11, 12, 14    # columns K, L, N
CODE, PDF = 23, 26               # columns W, Z

AUTHOR = "QC page verification 31.08.2026"

# --- row 26: the wrong certificate, and a lost decade ------------------------
R26 = 26
R26_CODE = ("319/0586/25", "318/0585/25")
R26_TYMC = ("6×10²", "6×10³")
R26_LINK = ("https://drive.google.com/file/d/1Z-MuH4YV4bNogxnkjetCL791Vaz2Xgsz/view",
            "https://drive.google.com/file/d/1ilPweUdwUCaop3lfeutJnQYKbdrQY5Tc/view")
R26_NOTE = (
    "Certificate corrected 31.08.2026 from a rendered page.\n\n"
    "Two IJZ-MB certificates exist for GP0824-01, both 14.04.2025:\n"
    "  318/0585/25 — TAMC 4,2 x 10^3, TYMC 6 x 10^3\n"
    "  319/0586/25 — TAMC 5,2 x 10^3, TYMC 8,8 x 10^3\n\n"
    "This row's TAMC matches 318, so it was carrying 318's results under 319's "
    "code, with 318's TYMC recorded a decade low. Code, TYMC and PDF link now "
    "all address 318/0585/25.\n\n"
    "OPEN FOR QC: 319/0586/25 is a real certificate for this batch and has no "
    "row of its own."
)

# row, column, current value, page value, certificate, what the page prints
CELLS = [
    (88,  GNB, "<10¹",            "< 10",
     "948/1686/25  HPA052501",    "page prints < 10"),
    (92,  GNB, "<10^1 and >10^1", "< 10^3 and > 10^2",
     "1009/1813/25  GP062501",    "page prints < 10³ и >10² — as written the cell "
                                  "was impossible: nothing is both under and over 10¹"),
    (142, GNB, "< 10 and > 10^3", "< 10^4 and > 10^3",
     "1227/2193/25  GP082501-2",  "page prints < 10⁴ и > 10³ — as written the cell "
                                  "was impossible: nothing is both under 10 and over 10³"),
    (156, GNB, "< 10^2 and > 10", "< 10",
     "4/0007/26  CJ082501-2",     "page prints a plain < 10, not a bracket"),
    (148, ECOLI, "< 10", "Odgovara (Absent)",
     "3/0006/26  WC082501",       "page reports Одговара against отсутна/g"),
    (156, ECOLI, "< 10", "Odgovara (Absent)",
     "4/0007/26  CJ082501-2",     "page reports Одговара against отсутна/g"),
    (171, ECOLI, "< 10", "Odgovara (Absent)",
     "6/0009/26  PM092501",       "page reports Одговара against отсутна/g"),
    (177, ECOLI, "< 10", "Odgovara (Absent)",
     "10/0013/26  CJ092501",      "page reports Одговара against отсутна/g"),
]

# The two certificates whose printed limits are tighter than the column header, and
# which test two parameters the register has no columns for. Recorded as comments
# because adding columns for two of 291 rows would leave 289 rows of blanks.
SPEC_NOTES = [
    (122, TYMC, "1220/2171/25", "PM072501",
     "TYMC 200 CFU/g against a PRINTED limit of 10^2 — 2x over.\n\n"
     "This certificate cites Ph.Eur. 5.1.8 Kat. C AND производителска спецификација, "
     "which is tighter than this column's header (TAMC 10^4 not 10^5, TYMC 10^2 not "
     "10^4, GNB <= 10^2). The result therefore PASSES the column and FAILS its own "
     "paper. Confirmed on the rendered page 31.08.2026.\n\n"
     "Also tested, with no column here: S. aureus Одговара, P. aeruginosa Одговара.\n\n"
     "The certificate concludes ОДГОВАРА."),
    (118, TYMC, "1221/2172/25", "WC072501",
     "Runs the same tighter производителска спецификација as 1220/2171/25 "
     "(TAMC 10^4, TYMC 10^2, GNB <= 10^2), not this column's 10^4.\n\n"
     "TYMC 90 against that 10^2 limit — compliant, and compliant against the "
     "tighter limit rather than the column's.\n\n"
     "Also tested, with no column here: S. aureus Одговара, P. aeruginosa Одговара."),
]


def _cur(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return "" if v is None else str(v).strip()


def main(src, dst):
    wb = load_workbook(src)
    ws = wb[SHEET]
    log, skipped = [], []

    # --- 1. row 26 ------------------------------------------------------------
    old_code, new_code = R26_CODE
    old_tymc, new_tymc = R26_TYMC
    cur_code, cur_tymc = _cur(ws, R26, CODE), _cur(ws, R26, TYMC)
    if cur_code == new_code and cur_tymc == new_tymc:
        skipped.append(f"r{R26} already {new_code} / {new_tymc}")
    else:
        # Both halves are one finding, so they move together or not at all: a row
        # holding 318's code beside 319's TYMC would be a new kind of wrong.
        if (cur_code, cur_tymc) != (old_code, old_tymc):
            raise SystemExit(
                f"REFUSING r{R26}: expected ({old_code!r}, {old_tymc!r}), found "
                f"({cur_code!r}, {cur_tymc!r}). Not the revision this was verified "
                f"against — re-read the page first.")
        ws.cell(row=R26, column=CODE).value = new_code
        ws.cell(row=R26, column=TYMC).value = new_tymc
        cell = ws.cell(row=R26, column=PDF)
        if getattr(cell.hyperlink, "target", None) == R26_LINK[0]:
            cell.value, cell.hyperlink = "Open", R26_LINK[1]
            log.append(f"Z{R26}  link repointed 319/0586/25 -> 318/0585/25")
        ws.cell(row=R26, column=CODE).comment = Comment(R26_NOTE, AUTHOR)
        log.append(f"W{R26}  {old_code!r} -> {new_code!r}   TAMC 4.2×10³ matches 318")
        log.append(f"K{R26}  {old_tymc!r} -> {new_tymc!r}   page prints 6 x 10³")

    # --- 2. the eight cell corrections ---------------------------------------
    for row, col, old, new, cert, why in CELLS:
        letter = ws.cell(row=row, column=col).column_letter
        cur = _cur(ws, row, col)
        if cur == new:
            skipped.append(f"r{row} {letter} already {new!r}")
            continue
        if cur != old:
            raise SystemExit(
                f"REFUSING {letter}{row}: expected {old!r}, found {cur!r}. "
                f"Re-verify against {cert} first.")
        ws.cell(row=row, column=col).value = new
        log.append(f"{letter}{row:<4} {old!r} -> {new!r}   {cert} — {why}")

    # --- 3. the two per-certificate specification notes -----------------------
    for row, col, cert, batch, note in SPEC_NOTES:
        if _cur(ws, row, CODE) != cert:
            raise SystemExit(
                f"REFUSING r{row}: expected code {cert!r}, found "
                f"{_cur(ws, row, CODE)!r}.")
        cell = ws.cell(row=row, column=col)
        if cell.comment is not None:
            skipped.append(f"r{row} {cert} note already present")
            continue
        cell.comment = Comment(note, AUTHOR)
        log.append(f"K{row:<4} note — {cert} {batch}: printed limit is 10², "
                   f"not the column's 10⁴")

    wb.save(dst)
    print(f"in : {src}\nout: {dst}\n")
    for line in log:
        print("  CHANGED  " + line)
    for s in skipped:
        print("  skipped  " + s)
    print(f"\n{len(log)} change(s), {len(skipped)} already applied.")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 3:
        raise SystemExit(__doc__)
    sys.exit(main(sys.argv[1], sys.argv[2]))
