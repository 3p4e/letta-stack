#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Close the last six certificates the four family passes did not cover.

    python3 deliverables/qc_gap_analysis/apply_residual_corrections.py IN.xlsx OUT.xlsx

Once the four families were done, `verification_coverage.py` counted what was left rather
than estimating it: 90 populated result cells over 16 rows. Twenty-four of those turned out
to be already verified — the six `PP CoA #nnn / ППКnnnnn` rows, whose folded code matches no
certificate because the cell names two documents, hold the CNP values and agree 18 for 18.

Six certificates were genuinely unread, and they are the residue of two family passes rather
than a family of their own: five IJZ-MB microbiology reports that carry no IJZ-MB laboratory
number in the shape the earlier sweep matched on, and one State Phytosanitary pesticide
report. All six were fetched and read. **Twenty-five of the twenty-six values agree.**

## The one correction · row 109, pesticides

| | |
|---|---|
| Certificate | `10802_2845/2`, ДФЛ, 17.11.2025 |
| Page | `Забелешка: Анализата е направена врз основа на барањето за испитување на пестициди според **cPh Eur.**` |
| Register held | `… — COMPLIES with **USP and Ph.Eur.**` |
| Corrected to | `… — COMPLIES with **Ph. Eur.**` |

The finding itself is right and is not in doubt: no pesticide above the 0,01 mg/kg limit of
quantification, out of **471 residues** screened — 265 by LC/MS/MS and 206 by GC/MS/MS, against
IPH's fixed 25-row panel. Ph. Eur. is right too. **USP is the half with no support: it appears
nowhere on the page.** The certificate names one pharmacopoeia; the register names two.

So this correction removes an unsupported claim rather than replacing a wrong one — a smaller
defect than the U column or the receipt date, and a different kind. A cell that credits a
conformity to two authorities where the laboratory named one is not false about the result; it
is false about what was certified, which is what a release register exists to record.

**The guard caught me writing this script.** A first draft had the old value as
`… COMPLIES with USP`, because the survey that found the row printed cells truncated to 60
characters and `USP` is where the cut fell. The refuse-on-mismatch check rejected the workbook
and printed the full string, which is how the "and Ph.Eur." came to light — and how a
correction that would have silently dropped a true half of the cell was stopped. Third time
this guard has caught a misreading of mine rather than a corrupt workbook.

Two further facts about that certificate are recorded in the cell comment rather than changed,
because neither is this register's to decide. Its MRL column reads `/` — the laboratory states
no maximum residue level, so "complies" here means "nothing detected", not "below a limit".
And the commissioning client is **New Garden Pharma**, not Purely Plant, with the sample
identified as `Blue Sunset Sherbet BSS 052501 NGP`.

## What the five microbiology reports settle

Nothing needed correcting — all 25 values agree, and every difference is the register glossing
the page more fully (`Одговара (absent)`, `< 10^3 and > 10^2` for `< 10³ и >10²`). Two things
are worth having on record.

**`J311122501` is confirmed a typo.** The IPH pass found `1625/2026` and `1628/2026`, two
certificates for one Jokerz 31 batch, printing serials one digit apart, and could not say which
was right. `230/0393/26` and `231/0394/26` — trimmed and hand-trimmed, sampled the same day —
both print **`J31122501`**. Two independent laboratories now agree against `1628/2026`, and the
register's `J31122501` is right.

**The IJZ-MB result notation changed in 2026.** The three April-2026 reports print counts as
plain integers — `1900`, `840`, `120`, `100` — where every 2025 report prints `1,9 x 10³`. Same
laboratory, same form family, form version 2 to version 3. A parser that reads counts by
matching a coefficient and a power of ten returns nothing at all on the newer reports: not a
wrong value, an absent one. That is the microbiology family's own version of the wording change
the IPH pass found on `305/0549/26`.

## And what is left, which is nothing that can be read

With these six done, every external certificate the register cites has been read. The 40
result cells still unverified sit on **four in-house Purely Plant rows** — two cross-checks
(`In-house HPLC cross-check NPCCC/SCP-02`, `In-house GC cross-check NGP/QCG/SOP-024`) and two
in-house CoAs whose own code cells say `n/a … no certificate/report`. They are unverifiable
because they *are* the record, not because anybody skipped them.

One of them carries a defect that can be seen without a source. Row 288's bile-tolerant GNB
reads **`<10²>10³`** — below 100 and above 1000 at once. Nothing can be both. The microbiology
pass found the same impossible construction twice (rows 92 and 142) and could correct it,
because a certificate said what the laboratory meant; here there is no document to consult.

That cell was **already amber** before this script ran — it is one of the register's original
18 flags, alongside `L285` on the sibling in-house row, so somebody had noticed something here
before. What it did not carry was any statement of what is wrong with it. This adds the
diagnosis: the two bounds are the wrong way round, both plausible readings are set out, and
neither can be chosen without the underlying in-house record. Guessing a value into a release
register is worse than leaving a visible contradiction in it.

Idempotent, and it refuses a workbook whose cells differ from what was verified.
"""
import sys

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

SHEET = "Batch Release QC"
PEST, CODE = 22, 23               # columns V and W
AUTHOR = "QC page verification 31.08.2026"
AMBER = PatternFill("solid", fgColor="FFFAEDD4")

ROW = 109
CERT = "10802_2845/2"
OLD = "Not found any pesticide above LOQ (≤LOQ) — COMPLIES with USP and Ph.Eur."
NEW = "Not found any pesticide above LOQ (≤LOQ) — COMPLIES with Ph. Eur."

NOTE = (
    "Corrected 31.08.2026 from the rendered page of 10802_2845/2.\n\n"
    "The certificate's own footnote reads: \"Забелешка: Анализата е направена врз "
    "основа на барањето за испитување на пестициди според cPh Eur.\" — one "
    "pharmacopoeia. The cell credited two: \"COMPLIES with USP and Ph.Eur.\". "
    "Ph. Eur. is supported; USP appears nowhere on the page, and has been removed. "
    "Confirmed at 2.4x magnification.\n\n"
    "The result is unchanged and is not in doubt: no pesticide above the 0,01 mg/kg "
    "limit of quantification, across 471 residues (265 by LC/MS/MS, 206 by GC/MS/MS, "
    "both МКС EN 15662:2011). Measurement uncertainty 50 % at k=2 and 95 %.\n\n"
    "Two things on this certificate are left for QC rather than changed here:\n\n"
    "  · The MRL column reads \"/\" — the laboratory states no maximum residue level. "
    "\"Complies\" on this row means nothing was detected, not that a result sits below "
    "a limit.\n\n"
    "  · The commissioning client is ЊУ ГАРДЕН ФАРМА (New Garden Pharma), not Purely "
    "Plant, and the sample is identified as \"Blue Sunset Sherbet BSS 052501 NGP\". "
    "The header numbers the report 10802_2845/2 while \"Бр. на барање за анализа\" on "
    "the same page reads 10802-2845/1."
)

GNB288_NOTE = (
    "Data-integrity flag, 31.08.2026.\n\n"
    "This cell reads \"<10²>10³\" — below 10² AND above 10³. No count can be both; "
    "the two bounds are the wrong way round whichever reading is intended.\n\n"
    "The microbiology pass found the same impossible construction on rows 92 and 142 "
    "and corrected both, because an IJZ-MB certificate said what the laboratory "
    "actually reported. This row has no such document: its own code cell reads \"n/a — "
    "Purely Plant in-house CoA (Batch OPM1024, no certificate/report)\".\n\n"
    "Two readings are plausible and this register cannot choose between them:\n"
    "  · \"< 10³ and > 10²\", the standard IJZ-MB phrasing for a count between 100 and "
    "1000 — the most likely intent, and what row 285's sibling batch would suggest;\n"
    "  · \"< 10² and > 10\", a decade lower.\n\n"
    "Both comply with the <= 10⁴ limit, so the batch's disposition does not turn on it. "
    "Resolving the cell needs the underlying in-house record, not a guess."
)

JOKERZ_NOTE = (
    "Page-verified 31.08.2026 against 230/0393/26.\n\n"
    "This report and its hand-trimmed sibling 231/0394/26 both print серија "
    "J31122501, matching this register and IPH 1625/2026.\n\n"
    "It settles the question the IPH pass left open: 1628/2026 prints J311122501, one "
    "digit longer, for the same batch sampled the same day. Two laboratories and three "
    "certificates say J31122501; one certificate says otherwise. The register is right."
)


def _cur(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return "" if v is None else str(v).strip()


def main(src, dst):
    wb = load_workbook(src)
    ws = wb[SHEET]
    log, skipped = [], []

    got = _cur(ws, ROW, CODE)
    if got != CERT:
        raise SystemExit(f"REFUSING W{ROW}: expected {CERT!r}, found {got!r}.")

    cur = _cur(ws, ROW, PEST)
    if cur == NEW:
        skipped.append(f"r{ROW} pesticides already cites Ph. Eur.")
    elif cur != OLD:
        raise SystemExit(
            f"REFUSING V{ROW}: expected {OLD!r}, found {cur!r}. Not the revision this "
            f"was verified against — re-read {CERT} first.")
    else:
        ws.cell(row=ROW, column=PEST).value = NEW
        log.append(f"V{ROW}  '…USP and Ph.Eur.' -> '…Ph. Eur.'   {CERT} names one "
                   f"pharmacopoeia, not two")

    cell = ws.cell(row=ROW, column=PEST)
    if cell.comment is None:
        cell.comment = Comment(NOTE, AUTHOR)

    gnb = ws.cell(row=288, column=12)
    if str(gnb.value or "").strip() == "<10²>10³":
        gnb.fill = AMBER
        if gnb.comment is None:
            gnb.comment = Comment(GNB288_NOTE, AUTHOR)

    j = ws.cell(row=217, column=CODE)
    if str(j.value or "").startswith("230/0393/26") and j.comment is None:
        j.comment = Comment(JOKERZ_NOTE, AUTHOR)

    wb.save(dst)
    print(f"in : {src}\nout: {dst}\n")
    for line in log:
        print("  CHANGED  " + line)
    for s in skipped:
        print("  skipped  " + s)
    print(f"\n{len(log)} change(s), {len(skipped)} already applied.")
    print("row 288 bile-tolerant GNB: '<10²>10³' is impossible as written. The cell "
          "was already amber; a comment now says why, and why it cannot be resolved.")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 3:
        raise SystemExit(__doc__)
    sys.exit(main(sys.argv[1], sys.argv[2]))
