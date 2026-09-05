#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Which testing does each batch actually have recorded?

    python3 deliverables/qc_gap_analysis/batch_test_coverage.py REGISTER.xlsx [--csv OUT.csv]

The register is organised by document: one row per certificate, grouped into a block per
batch. That answers "what does this certificate say" very well and "is this batch fully
tested" not at all — the answer to the second question is the *shape* of a block, and a
shape is not something a column can hold.

This reads the shape. For every batch it reports which of five test families appears,
identified from the certificate code and the issuing institution:

| Family | Recognised by | Covers |
|---|---|---|
| potency | `ППКnnnnn` (CNP) | Δ9-THC, CBD, CBN, loss on drying |
| cannabinoids | Farmahem `-К/26`, `-ГС/26`, `-LoD-26` | cannabinoids, loss on drying |
| mycotoxins | Farmahem `-М/26` | aflatoxins Σ, B1, ochratoxin A |
| microbiology | IJZ-MB `nnn/nnnn/yy` | TAMC, TYMC, GNB, E. coli, Salmonella |
| physico-chemical | IPH `nnnn/yyyy` | Pb, Cd, As, Hg, aflatoxins, pesticides |

**This reports what the register records, not what a laboratory did.** A batch with no
microbiology row may have been tested and not entered, may be mid-testing, or may not have
been sent. The register cannot tell those apart and neither can this script; that is the
point of printing the gap rather than a verdict.

Two strings mark the same absence and are worth knowing about together, because a reader
filtering on either one sees half the picture: eleven blocks write `(not numbered)` in the
code column and twelve write `n/a`. Both mean the same thing — a batch whose block holds
only potency results — and they sit in adjacent rows of the same register.
"""
import csv
import re
import sys

from openpyxl import load_workbook

SHEET = "Batch Release QC"
REF, BATCH, CODE, DATE, INST = 1, 2, 23, 24, 25
FIRST_DATA = 6

FAMILIES = ["potency", "cannabinoids", "mycotoxins", "microbiology", "phys-chem"]


def family(code, inst):
    """Which test family a row's certificate belongs to, or None for a placeholder.

    The trailing bracketed note has to come off first, exactly as in `fold()`. Rows 213
    and 217 are IJZ-MB microbiology reports written `231/0394/26 (Racno trimiran cvet)`,
    and a first version of this script read the brackets as part of the code, failed to
    match, and reported two batches as having no microbiology when they have it. The
    annotation is a note about the sample, not part of the certificate number.
    """
    c = re.sub(r"\s*\([^)]*\)\s*$", "", str(code or "")).strip()
    i = str(inst or "")
    if not c or c.lower() in ("n/a", "(not numbered)", "none"):
        return None
    if "ППК" in c:
        return "potency"
    if "Farmahem" in i:
        # Farmahem issues three report types and the suffix names them: -К cannabinoids,
        # -М mycotoxins, -ГС / -LoD loss on drying (which travels with the cannabinoids).
        if re.search(r"-\s*[МM]\s*/", c):
            return "mycotoxins"
        return "cannabinoids"
    if re.match(r"^\d+/\d+/\d+", c):
        return "microbiology"
    if re.match(r"^\d+/\d{4}$", c):
        return "phys-chem"
    if "10802" in c:
        return "phys-chem"          # State Phytosanitary pesticide screen
    return None                     # in-house CoAs and cross-checks: not a test family


def blocks(ws):
    out, cur = [], None
    for r in range(FIRST_DATA, ws.max_row + 1):
        if ws.cell(row=r, column=REF).value not in (None, ""):
            if cur:
                out.append(cur)
            cur = {"hdr": r, "batch": ws.cell(row=r, column=BATCH).value, "rows": []}
        if cur:
            cur["rows"].append(r)
    if cur:
        out.append(cur)
    return [b for b in out if b["batch"]]


def main(src, csv_out=None):
    ws = load_workbook(src)[SHEET]
    table = []
    for b in blocks(ws):
        have = set()
        marker = ""
        dates = []
        for r in b["rows"]:
            c = str(ws.cell(row=r, column=CODE).value or "").strip()
            f = family(c, ws.cell(row=r, column=INST).value)
            if f:
                have.add(f)
                d = str(ws.cell(row=r, column=DATE).value or "")
                if re.match(r"\d{2}\.\d{2}\.\d{4}", d):
                    dates.append(d)
            elif c.lower() in ("n/a", "(not numbered)"):
                marker = c
        latest = max(dates, key=lambda d: d[6:] + d[3:5] + d[:2]) if dates else ""
        table.append((b["hdr"], str(b["batch"]), have, marker, latest))

    full = [t for t in table if len(t[2]) >= 3 and "microbiology" in t[2] and "phys-chem" in t[2]]
    nomicro = [t for t in table if t[2] and "microbiology" not in t[2]]

    print(f"register: {src}\n")
    print(f"  batches                                  : {len(table)}")
    print(f"  with microbiology AND physico-chemical   : {len(full)}")
    print(f"  with NO microbiology recorded            : {len(nomicro)}")
    print(f"  with NO physico-chemical recorded        : "
          f"{len([t for t in table if t[2] and 'phys-chem' not in t[2]])}\n")

    hdr = f"  {'row':<6}{'batch':<15}" + "".join(f"{f:<15}" for f in FAMILIES) + "latest   marker"
    print(hdr)
    print("  " + "-" * (len(hdr) - 2))
    for r, batch, have, marker, latest in table:
        cells = "".join(("yes" if f in have else "—").ljust(15) for f in FAMILIES)
        print(f"  r{r:<5}{batch:<15}{cells}{latest:<9}{marker}")

    if csv_out:
        with open(csv_out, "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(["row", "batch"] + FAMILIES + ["latest_certificate", "marker"])
            for r, batch, have, marker, latest in table:
                w.writerow([r, batch] + ["yes" if f in have else "" for f in FAMILIES]
                           + [latest, marker])
        print(f"\n  written: {csv_out}")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 2:
        raise SystemExit(__doc__)
    out = sys.argv[sys.argv.index("--csv") + 1] if "--csv" in sys.argv else None
    sys.exit(main(sys.argv[1], out))
