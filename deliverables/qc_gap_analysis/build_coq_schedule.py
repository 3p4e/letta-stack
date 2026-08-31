#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Fill in every CoQ of the owner's issue plan: all determinations, criteria, documents.

    python3 deliverables/qc_gap_analysis/build_coq_schedule.py

The numbered CoQs are the owner's, not a prediction. The ISSUE_COQ folder in Drive
holds the approved master template, 48 rendered initial-release CoQs numbered
`CoQ-PP-{year}-{NNNN}` sequentially by packaging date, and 13 additional-testing CoQs
(`CoQ-PP-2026-0027 … 0039`) for the lots whose 12-month retest fell due — one CoQ per
**packaged lot**, each carrying an `iCoA-PP-{year}-{NNNN}` reference. Its issue plan is
committed beside this file as `coq_issue_plan.json`; the conventions are in
`ISSUE_COQ_CONVENTIONS.md`.

The universe those 61 sit in is the owner's ruling of 31.08.2026: **one initial CoQ
for every batch on record, first to last** — the 48 numbered lots are Tranche 01 (19)
and Tranche 02 (29) only, and every batch past them gets a PREDICTED initial CoQ —
**and a 12-month cannabinoid + mycotoxin reissue for every batch**, starting from the
beginning of Tranche 01/02: 13 reissues carry numbers (0027…0039), the other 35 of
the 48 are predicted at packaging + 12 months, and every later batch is predicted at
release + 12 months. A predicted CoQ carries no number — numbers are copied from the
issuance record at issue, never computed in advance.

Dating: **the CoQ SOP was put in use on 11.05.2026** (owner, 31.08.2026). No CoQ may
print an earlier issue date. The plan's per-CoQ dates are packaging dates — the basis
of the numbering series (`CoQ-PP-2025-…` is the 2025 packaging series), never issue
dates. Each schedule row therefore shows the basis date and the earliest permissible
issue date: the SOP date, the newest document the CoQ cites, or the 12-month due
date, whichever is latest. QC sets the real date at issue.

The plan deliberately leaves controlled blanks — every result the master spec does not
hold, and every eCoA code: *"A CoQ must never carry a result or a conformity assertion
that has not been certified — QC transcribes these from the actual iCoA/eCoA at issue."*
This build is that transcription source: **61 CoQs × 23 determinations = 1 403 rows**,
each with the parameter, method and acceptance criterion verbatim from QCSP 001 v.03,
the result and document code from the release register, a status, and — for identity
and foreign matter — the owner's routing decision of 31.08.2026:

  * where an outsourced laboratory did not cover them at release, Purely Plant's
    laboratory performs them and issues **two iCoAs — Ident A + B in one, foreign
    matter in another**; Identity C needs an outside laboratory;
  * at retest, **Farmahem performs Ident A, B and C together with the assay**, and the
    in-house laboratory issues one iCoA for foreign matter;
  * an additional-testing CoQ cites only post-release certificates — the 197-series
    re-analysis where it exists, blanks owed where it does not.

`GG1024` (CoQ-PP-2025-0003 / 0029) is the case the whole schedule exists to expose: its
outsourced testing was performed — the in-house CoA attributes it to accredited
laboratories — but **not one eCoA is on file**. Its values come from the in-house CoA
transcription and every one is flagged: locate the physical certificates, scan, upload.
Nine more lots' cultivation batches are in the same state.

Cross-checks the build runs and reports rather than resolves: the plan's banner THC
against the register's assays (**22 lots disagree** — the master spec carries analyses
whose certificates the register has never received), the plan's grade range against the
QCSP 001 PDF (**27 lots differ** beyond the endpoint convention — two grade designs —
which the schedule shows on 54 CoQ documents, each lot's conflict appearing on both its
initial CoQ and its reissue), and every citation's date against the CoQ's earliest
permissible issue date.

Conformity is judged by `ingestion/ragflow/validate_ecoa_limits.py`, imported rather
than reimplemented — that rule has been got wrong once already, in two directions, and
a second copy would be a third chance.
"""
import csv
import importlib.util
import json
import os
import re
import sys
from collections import OrderedDict, defaultdict

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
REG_X = os.path.join(HERE, "PP_Batch_Release_QC_Register_FHM2_2026-08-31.xlsx")
SPEC_J = os.path.join(HERE, "product_specifications_QCSP001.json")
ICOA_C = os.path.join(HERE, "icoa_issuance_register_2026-08-31.csv")
MASTER_J = os.path.join(HERE, "potency_master_batch_map.json")
PLAN_J = os.path.join(HERE, "coq_issue_plan.json")
INHOUSE_TSV = os.path.join(ROOT, "ingestion", "coa_track", "letta-imb-coas",
                           "exports", "master_coa_table.tsv")
BATCH_ID = os.path.join(ROOT, "ingestion", "common", "batch_id.py")
OUT_X = os.path.join(HERE, "PP_CoQ_Parameter_Schedule_2026-08-31.xlsx")
OUT_C = os.path.join(HERE, "coq_parameter_schedule_2026-08-31.csv")
VALIDATOR = os.path.join(ROOT, "ingestion", "ragflow", "validate_ecoa_limits.py")

SHEET, STABILITY_SHEET = "Batch Release QC", "Stability Testing Programme"
HEADER_ROW, SPEC_ROW, FIRST_DATA = 4, 5, 6
BATCH, PNUM, STRAIN, CODE, DATE, LAB = 2, 3, 4, 23, 24, 25
NOT_A_RESULT = {"", "/", "n/a", "na", "-", "—", "not tested", "none"}

FONT = "Arial"
INK, SUB, ACCENT = "16232B", "5A6472", "0E6E6E"
BAND, REFBG = "F6F8F7", "E7ECEA"
RED, AMBER, GREEN, BLUE = "F9DEDB", "FAEDD4", "E3F0E8", "E6ECF7"

# Status vocabulary. One string per state, so the column can be filtered.
ST_OK = "covered"
ST_FINDING = "covered — laboratory finding recorded"
ST_OOS = "OUT OF SPECIFICATION"
ST_UNDET = "UNDETERMINED — pending the QCSP 001 reading"
ST_ICOA = "to be performed — see route"
ST_NONE = "not tested — no certificate covers it"
ST_REQ = "upon request — not required for release"
ST_BLOCK = "BLOCKED — declared out of specification by the laboratory"
ST_NOSPEC = "no product specification on file — criterion cannot be stated"

# ---------------------------------------------------------------------------------
# Identity and foreign matter: what each laboratory CAN do, and what WILL be done.
#
# Four determinations have no column in the release register, and none of them is
# "in-house only" — each can be commissioned in more than one way. Two constraints in
# the menu shape everything downstream, and both come from the owner, 31.08.2026:
#
#   · **CNP does not sell foreign matter separately.** It comes only inside the full
#     identity-plus-assay package. A batch that needs foreign matter and nothing else
#     has one outside route and it is the expensive one.
#   · **Purely Plant's own laboratory can do Ident A, Ident B and foreign matter, but
#     not Ident C.** Identity C is HPLC/HPTLC and needs an outside laboratory.
CNP_FULL = "CNP — Ident A + B + C with Assay and Foreign matter"
CNP_IDENT_ASSAY = "CNP — Ident A + B + C with Assay"
CNP_IDENT_C = "CNP — Ident C only"
CNP_IDENT_AB = "CNP — Ident A + B"
FARMAHEM_IDENT = "Farmahem — Ident A + B + C"
PP_LAB = "Purely Plant in-house laboratory"

CAN_PERFORM = {
    "1": [CNP_FULL, CNP_IDENT_ASSAY, CNP_IDENT_AB, FARMAHEM_IDENT, PP_LAB],
    "2": [CNP_FULL, CNP_IDENT_ASSAY, CNP_IDENT_AB, FARMAHEM_IDENT, PP_LAB],
    "3": [CNP_FULL, CNP_IDENT_ASSAY, CNP_IDENT_C, FARMAHEM_IDENT],
    "7": [CNP_FULL, PP_LAB],
}

# And what will actually be done, which is a decision rather than an option. Where no
# outsourced laboratory covered a determination:
#
#   at initial release  · Purely Plant's laboratory issues TWO iCoAs — one covering
#                         Ident A and Ident B together, one covering foreign matter.
#                         Identity C it cannot do, so that waits for an outside lab.
#   at retest           · Farmahem performs Ident A, B and C together with the assay on
#                         the re-analysis, and Purely Plant's laboratory issues ONE iCoA
#                         for foreign matter.
#
# The two iCoAs at initial release are separate documents on purpose: identity and
# foreign matter are different determinations on different samples, and one document
# certifying both would have to be reissued whenever either is repeated.
ICOA_AB = "Purely Plant laboratory — iCoA covering Ident A + B"
ICOA_FM = "Purely Plant laboratory — iCoA covering Foreign matter"
FARMAHEM_RETEST = "Farmahem — Ident A + B + C with the Assay, at retest"
AWAIT_C = "outside laboratory — Ident C (CNP Ident C only, or Farmahem with the retest)"

ROUTE = {
    ("initial release", "1"): ICOA_AB,
    ("initial release", "2"): ICOA_AB,
    ("initial release", "3"): AWAIT_C,
    ("initial release", "7"): ICOA_FM,
    ("reissue", "1"): FARMAHEM_RETEST,
    ("reissue", "2"): FARMAHEM_RETEST,
    ("reissue", "3"): FARMAHEM_RETEST,
    ("reissue", "7"): ICOA_FM,
}

# The iCoA register already records, per batch, which of the four a CNP certificate
# already covers — `covered by CNP` against `required`. A first draft of this builder
# ignored those four columns and reported all 102 CoQs as owing all four determinations.
# Twelve batches do not: their CNP Ph. Eur. 11.5 certificate carries Ident A, Ident B
# and foreign matter, and only Ident C is outstanding.
ICOA_FIELD = {"1": "ident_A", "2": "ident_B", "3": "ident_C", "7": "foreign_matter"}

# The master spec writes OMP1024_01 where the register writes OPM1024_01. The register's
# own P-number column settles that they are one lot — its OPM1024_01 block carries
# P050042, the very lot the issue plan files under OMP1024_01. One transposition, on the
# master. No other alias is evidence-backed: the plan's FB012601/1 (assay 17.99) is NOT
# the register's FB012601 block (assay 14.68, ППК26067), and JD012603/01 (21.01) is not
# JD012601 (18.16) — those sub-lots genuinely have no register block.
PLAN_CB_ALIASES = {"OMP1024_01": "OPM1024_01"}


def route_for(coq_type, no):
    """Who performs an outstanding identity or foreign-matter determination.

    >>> route_for("initial release", "1")
    'Purely Plant laboratory — iCoA covering Ident A + B'
    >>> route_for("initial release", "7")
    'Purely Plant laboratory — iCoA covering Foreign matter'
    >>> route_for("initial release", "3")
    'outside laboratory — Ident C (CNP Ident C only, or Farmahem with the retest)'
    >>> route_for("reissue", "3")
    'Farmahem — Ident A + B + C with the Assay, at retest'
    >>> route_for("reissue", "7")
    'Purely Plant laboratory — iCoA covering Foreign matter'
    """
    return ROUTE[(coq_type, no)]


def icoa_plan(per_coq):
    """The in-house certificates the routing generates, one row each.

    Not the same as the 81-row iCoA issuance register, and the difference is the point:
    that register carried one iCoA per batch of mixed scope. Under the routing an
    initial release owing both identity and foreign matter needs **two** documents, and
    a retest needs one — because Farmahem covers identity at retest and Purely Plant's
    laboratory only has foreign matter left to certify.

    One CoQ is deliberately absent from the foreign-matter side: **FB032601's reissue**.
    Its foreign matter was determined at release and DECLARED OUT OF SPECIFICATION by
    the laboratory (ППК26127). Foreign matter does not improve in storage, so buying the
    determination again twelve months on is not what that batch needs; its reissue CoQ
    carries the finding, not a blank. So 83 reissues yield 82 foreign-matter
    certificates, and the one missing is missing on purpose.
    """
    plan, seen = [], set()
    for p in per_coq:
        for scope, nos in (("Ident A + B", ("1", "2")), ("Foreign matter", ("7",))):
            # A CoQ is a reissue when it is additional testing — NOT merely when its
            # type differs from the string "initial release". A first version tested
            # `p["type"] != "initial release"`, which sent all 35 *predicted* initial
            # releases down the retest route, where Farmahem covers identity, and
            # silently dropped 26 Ident A + B certificates the schedule's own rows
            # say the in-house laboratory owes.
            typ = "reissue" if p["type"].startswith("additional") else "initial release"
            owed = [n for n in nos if n in p["outstanding"]
                    and route_for(typ, n).startswith("Purely Plant laboratory")]
            if not owed:
                continue
            # Keyed on the CoQ, not on its packaged lot: a predicted CoQ has no lot
            # number yet, and keying on one collapsed all 70 of them into two rows.
            key = (p["coq"], scope)
            if key in seen:
                continue
            seen.add(key)
            plan.append({"pp": p["pp"], "cb": p["cb"], "strain": p["strain"],
                         "number": p["number"], "icoa_ref": p["icoa_ref"],
                         "coq_type": p["type"], "date": p["date"],
                         "scope": scope, "determinations": ", ".join(owed)})
    return plan


def _validator():
    """Import the acceptance-criterion rule rather than writing it twice.

    It has been got wrong once already, in two directions: against the literal power of
    ten, and against the register's unsourced x5 parenthetical. A second copy here would
    be a third chance.
    """
    spec = importlib.util.spec_from_file_location("validate_ecoa_limits", VALIDATOR)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


V = _validator()


def _batch_id():
    """The one batch-identity rule, imported rather than re-approximated.

    The master spec writes `CJ052501/01` where the register writes `CJ052501-1`, and
    matching them by string fails on every separator the corpus uses. `batch_key` is
    the campaign's single definition of when two spellings name the same batch.
    """
    spec = importlib.util.spec_from_file_location("batch_id", BATCH_ID)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


BI = _batch_id()


def clean(x):
    return "" if x is None else str(x).strip()


def is_result(x):
    return bool(clean(x)) and clean(x).lower() not in NOT_A_RESULT


def family(code):
    """Which report series a certificate code belongs to.

    The two IPH series are told apart by shape, not by content: `nnn/nnnn/nn` is a
    microbiology report, `nnnn/yyyy` a mycotoxin, metals and pesticide report. Measured
    over the register: 44 of the first carry exactly TAMC/TYMC/GNB/Salmonella/E. coli,
    and every one of the second carries the metals panel.

    >>> family("320/0587/25")
    'IPH microbiology'
    >>> family("752/2025")
    'IPH mycotoxins, metals, pesticides'
    >>> family("ППК25050")
    'UKIM CNP potency'
    >>> family("197-1-К/26")
    'Farmahem re-analysis — cannabinoids'
    >>> family("197-1-М/26")
    'Farmahem re-analysis — mycotoxins'
    >>> family("051-6-GS/26")
    'Farmahem — loss on drying'
    >>> family("QCCoA 001v02 — 033")
    'Purely Plant in-house'
    """
    c = clean(code)
    if re.match(r"^\d+/\d+/\d+", c):
        return "IPH microbiology"
    if re.match(r"^\d+/\d{4}$", c):
        return "IPH mycotoxins, metals, pesticides"
    if c.startswith("ППК") or c.startswith("PP CoA"):
        return "UKIM CNP potency"
    if re.match(r"^197-.*[КK]/26$", c):
        return "Farmahem re-analysis — cannabinoids"
    if re.match(r"^197-.*[МM]/26$", c):
        return "Farmahem re-analysis — mycotoxins"
    if re.search(r"(ГС|GS)/\d\d$", c):
        return "Farmahem — loss on drying"
    if re.match(r"^\d+-\d+-[КK]/\d\d$", c):
        return "Farmahem — cannabinoids"
    if re.match(r"^\d+-\d+-[МM]/\d\d$", c):
        return "Farmahem — mycotoxins"
    if "QCCoA" in c or c.lower().startswith("n/a") or "in-house" in c.lower():
        return "Purely Plant in-house"
    return "other"


def is_reanalysis(code):
    """True for a Farmahem 197-series certificate — the re-analysis a reissue rests on.

    >>> is_reanalysis("197-11-К/26"), is_reanalysis("ППК25174")
    (True, False)
    """
    return clean(code).startswith("197-")


def sort_date(d):
    """DD.MM.YYYY to a sortable key; anything unparseable sorts last.

    >>> sort_date("13.06.2025") < sort_date("01.12.2025")
    True
    >>> sort_date("Completed Nov-Dec 2025")
    '9999'
    """
    m = re.match(r"^(\d{2})\.(\d{2})\.(\d{4})$", clean(d))
    return f"{m.group(3)}{m.group(2)}{m.group(1)}" if m else "9999"


# The in-house CoA's parameter names, keyed to the determination each answers. Used
# only for a cultivation batch the release register does not carry — today that is
# exactly GG1024, whose outsourced testing was performed (the in-house CoA attributes
# it to accredited laboratories in its own footnote) but whose eCoAs are not on file
# anywhere: not in Drive, not in the corpus, not in the register. The owner is
# locating the physical certificates to scan and upload; until then the in-house CoA
# transcription is the only record, and every value from it is flagged as resting on a
# document that cites certificates nobody can currently produce.
INHOUSE_MAP = [
    ("appearance", "1"), ("identification", "3"), ("foreign matter", "7"),
    ("total δ9-tetrahydrocannabinol", "4"), ("total cannabidiol", "5"),
    ("cannabinol", "6"), ("loss on drying", "8"),
    ("total aerobic microbial", "9.1"), ("total combined yeasts", "9.2"),
    ("bile-tolerant", "9.3"), ("salmonella", "9.4"), ("escherichia coli", "9.5"),
    ("total aflatoxines", "10.2"),
    ("cadmium", "11.2"), ("lead", "11.1"), ("mercury", "11.4"), ("arsenic", "11.3"),
]

ST_SCAN = ("in-house CoA only — underlying eCoA NOT on file: locate the physical "
           "certificate, scan and upload")
ST_OFFREG = ("covered — certificate on file, but the release register has no block "
             "for this cultivation batch")

# The retest programme is universal — every batch has a 12-month re-analysis of the
# CANNABINOIDS and the MYCOTOXINS and gets a CoQ reissue for it (owner, 31.08.2026).
# Identity is redone by Farmahem together with the assay, foreign matter by the
# in-house laboratory. Everything else is not retested: on a reissue those
# determinations stand on the initial CoQ, and saying "owed" for them — as an earlier
# revision of this file did — invented testing nobody ordered.
RETEST_K = {"4", "5", "6"}
RETEST_M = {"10.1", "10.2", "10.3"}
ST_AWAIT_K = "awaiting the cannabinoid re-analysis — Farmahem, with Ident A + B + C"
ST_AWAIT_M = "awaiting the mycotoxin re-analysis — Farmahem"
ST_OUTSIDE = "outside the retest scope — the release determination stands on {initial}"
NO_NUMBER = "(assigned on issue)"

# The CoQ SOP was put in use on 11.05.2026 (owner, 31.08.2026): no CoQ may print an
# issue date before it. The plan's per-CoQ dates are packaging dates — the numbering
# series basis — never issue dates.
SOP_EFFECTIVE = "11.05.2026"


IN_HOUSE_CODE = ("n/a — Purely Plant", "PP CoA #", "In-house ")


def inhouse_cells(cb):
    """The certificate transcription for a cultivation batch, keyed by determination.

    Used only where the release register has no block for the batch at all, to put the
    values that exist on the CoQ rather than leaving it blank.

    **Not everything it returns is an in-house document, and it must not say so.** The
    transcription table holds whatever certificate covers the batch — for GG1024 that is
    genuinely Purely Plant's own unnumbered CoA, but for FB012601/1 it is `ППК26067`, a
    UKIM CNP certificate that is on file and page-verified. A first version stamped every
    record `Purely Plant in-house` and the caller then flagged it *"underlying eCoA NOT on
    file: locate the physical certificate"* — three false assertions about a document
    anyone can open. The record now carries `family(code)`, the same rule the register
    uses, and an `inhouse` flag the caller reads instead of assuming.

    Pesticides are 13 residue rows on the in-house CoA; the schedule carries the panel
    as one determination, so they collapse to N.D. only when every row is N.D.
    """
    try:
        with open(INHOUSE_TSV, encoding="utf-8") as fh:
            rows = [r for r in csv.DictReader(fh, delimiter="\t")
                    if BI.batch_key(r["Cultivation Batch"] or "") == BI.batch_key(cb)]
    except OSError:
        return {}
    if not rows:
        return {}
    out, pest = {}, []
    for r in rows:
        pname = (r["Parameter"] or "").lower()
        code = clean(r["Certificate Code"])
        rec = {"value": clean(r["Result"]), "code": code,
               "date": clean(r["Issue Date"]), "lab": clean(r["Issuing Institution"]),
               "family": family(code), "flag": None, "stability": False,
               "note": "", "row": None,
               "inhouse": code.startswith(IN_HOUSE_CODE)}
        if "ph. eur. 2.8.13" in (r["Parameter"] or "").lower():
            pest.append(rec)
            continue
        for key, no in INHOUSE_MAP:
            if pname.startswith(key):
                out[no] = rec
                break
    if pest and all(x["value"].upper().startswith("N.D") for x in pest):
        out["12"] = dict(pest[0], value=f"N.D. — all {len(pest)} residues")
    return out


def read_register():
    """Every result in the release register, grouped by batch and column letter."""
    wb = load_workbook(REG_X)
    ws = wb[SHEET]

    stability = set()
    if STABILITY_SHEET in wb.sheetnames:
        for row in wb[STABILITY_SHEET].iter_rows(min_row=6):
            if row[5].value:
                stability |= V.codes_in(clean(row[5].value)) if hasattr(V, "codes_in") \
                    else {V.fold(clean(row[5].value))}

    limits = {}
    for c in range(5, 23):
        letter = get_column_letter(c)
        limits[letter] = V.acceptance_limit(clean(ws.cell(row=SPEC_ROW, column=c).value),
                                            clean(ws.cell(row=HEADER_ROW, column=c).value))

    batches, order, cur = {}, [], None
    for r in range(FIRST_DATA, ws.max_row + 1):
        if clean(ws.cell(row=r, column=1).value).upper() == "LEGEND":
            break
        b = clean(ws.cell(row=r, column=BATCH).value)
        if b:
            cur = b
            order.append(b)
            batches[b] = {"pnumber": clean(ws.cell(row=r, column=PNUM).value),
                          "strain": clean(ws.cell(row=r, column=STRAIN).value),
                          "cells": defaultdict(list)}
        code = clean(ws.cell(row=r, column=CODE).value)
        if not code or cur is None:
            continue
        for c in range(5, 23):
            cell = ws.cell(row=r, column=c)
            if not is_result(cell.value):
                continue
            rgb = getattr(cell.fill.fgColor, "rgb", None)
            batches[cur]["cells"][get_column_letter(c)].append({
                "value": clean(cell.value), "code": code,
                "date": clean(ws.cell(row=r, column=DATE).value),
                "lab": clean(ws.cell(row=r, column=LAB).value),
                "family": family(code), "row": r,
                "flag": {"FFFAEDD4": "amber", "FFF9DEDB": "red"}.get(
                    rgb if cell.fill.patternType == "solid" else None),
                "stability": V.fold(code) in stability,
                "note": cell.comment.text.strip() if cell.comment else "",
            })
    return order, batches, limits


def pick(candidates, reissue):
    """Which certificate this CoQ cites for this determination, and what else exists.

    Three rules, and the middle one is the one that matters.

    **A stability timepoint is never a release value.** The register's own Stability
    Testing Programme sheet says so in its subtitle.

    **An initial-release CoQ never cites the 197-series re-analysis.** That analysis
    happened in August 2026; a CoQ anchored on a February 2025 release cannot rest on
    it. A first draft of this function fell back to the re-analysis whenever no release
    certificate carried the determination, and produced a February 2025 CoQ citing
    `197-1-К/26` for CBN and `197-1-М/26` for two mycotoxins — back-dating evidence by
    seventeen months. Where no release certificate covers a determination the honest
    answer is that it was not determined at release, and the reissue CoQ is where the
    later result belongs.

    The rule names the 197 series specifically rather than filtering on date, because
    the register itself identifies that series as the re-analysis basis (`coq_reissue`,
    `reissue_basis_date`). Farmahem's `051-n` and `100-n` certificates are 2026-dated
    too and are the *release* testing for their batches, which no date rule would tell
    apart.

    **A reissue CoQ** cites the re-analysis where it covered the determination and the
    original release certificate where it did not — a reissued CoQ still certifies the
    whole panel.
    """
    live = [c for c in candidates if not c["stability"]]
    if not live:
        return None, candidates
    if reissue:
        re_an = [c for c in live if is_reanalysis(c["code"])]
        if re_an:
            chosen = sorted(re_an, key=lambda c: sort_date(c["date"]))[-1]
            return chosen, [c for c in candidates if c is not chosen]
    plain = [c for c in live if not is_reanalysis(c["code"])]
    if not plain:
        return None, candidates
    chosen = sorted(plain, key=lambda c: sort_date(c["date"]))[0]
    return chosen, [c for c in candidates if c is not chosen]


def outstanding_of(det, icoa_row, reg, reissue, blocked):
    """True when this identity or foreign-matter determination still has to be bought.

    Blocked is not outstanding: FB032601's foreign matter was determined, and failed.
    Buying it again is not what that batch needs.

    **Release-time coverage does not carry to a reissue.** Twelve batches have identity
    and foreign matter covered by their CNP Ph. Eur. 11.5 certificate *at release*. A
    12-month CoQ certifies the material as it is at that date: the owner's routing of
    31.08.2026 sends identity to Farmahem with the assay and leaves foreign matter to
    the in-house laboratory, on every batch. A certificate from the release round
    cannot stand behind a determination on a document dated a year later — the same
    rule that keeps an initial CoQ from citing the 197-series re-analysis, running the
    other way.
    """
    if blocked and det["no"] == "7":
        return False
    if reissue:
        return True
    return icoa_row.get(ICOA_FIELD[det["no"]], "required") == "required"


def status_of(det, chosen, limit, batch, blocked):
    """The status a CoQ row carries for one determination."""
    if blocked and det["no"] == "7":
        return ST_BLOCK
    if det["source"] == "upon_request":
        return ST_REQ
    if chosen is None:
        return ST_ICOA if det["source"] == "in_house_icoa" else ST_NONE
    got = V.magnitude(chosen["value"])
    if limit and limit.value and got is not None and not chosen["value"].startswith(("<", "≤")):
        if got > limit.value:
            return ST_OOS
        if limit.power and got > limit.power:
            return ST_UNDET
    return ST_FINDING if chosen["flag"] == "amber" else ST_OK


def schedule():
    """The whole schedule, without writing anything.

    The CoQ universe is the owner's, not a prediction. `coq_issue_plan.json` — copied
    verbatim from the ISSUE_COQ folder, which generates it from
    PP_Potency_MASTER_Spec.xlsx — lists **48 packaged lots with issued CoQ numbers**
    (`CoQ-PP-{year}-{NNNN}`, sequential by packaging date) **plus 13 additional-testing
    CoQs** (`CoQ-PP-2026-0027 … 0039`) for the lots whose 12-month retest fell due.
    Each carries an `iCoA-PP-{year}-{NNNN}` reference, the lot's cultivation batch, the
    grade acceptance range and the issue date. See ISSUE_COQ_CONVENTIONS.md.

    This settles a question three registers left open: **a CoQ is issued per packaged
    lot**, and 61 carry numbers. The universe around them is the owner's ruling of
    31.08.2026: one initial CoQ per batch on record, a reissue for all 48 Tranche
    01 + 02 lots (19 + 29), and a predicted reissue a year after release for every
    batch past them — the 12-month cannabinoid + mycotoxin retest programme is
    universal.

    What the ISSUE_COQ folder leaves as controlled blanks — every result the master
    spec does not hold, and every eCoA code — is exactly what this fills in, from the
    release register, so QC transcribes from a schedule instead of hunting 248
    certificates.
    """
    spec = json.load(open(SPEC_J, encoding="utf-8"))
    dets, specs = spec["determinations"], spec["specifications"]
    order, batches, limits = read_register()

    with open(ICOA_C, encoding="utf-8") as fh:
        icoa = {BI.batch_key(r["batch"]): r for r in csv.DictReader(fh)}
    reg_by_key = {BI.batch_key(b): b for b in batches}

    plan = json.load(open(PLAN_J, encoding="utf-8"))
    plan.sort(key=lambda x: (sort_date(x["pk"]), x["id"]))

    # Six register blocks are keyed by a packaged-lot P-number and hold the
    # Farmahem 12-month re-analyses OF plan lots (the register's "197-only"
    # rows: P060152 is J31102501, P060212 is JD112501, P060242 is OPM122501,
    # P060352 is FB012602, P060382 is SCR012603, P060402 is GG012603 — matched
    # by the plan's own packaged-lot number). Same material, second analysis:
    # their cells fold into the plan lot's block and they get no CoQ of their
    # own. Where the lot's cultivation batch has no register block of its own,
    # the P-number block IS its block. P160012/22/32 and P060332 match no plan
    # lot and stand as their own entries on the record.
    pp_alias = {}
    for x in plan:
        ppk = BI.batch_key(x["pp"]) if x["pp"] else None
        cbk = BI.batch_key(PLAN_CB_ALIASES.get(x["cb"], x["cb"]))
        if ppk and ppk != cbk and ppk in reg_by_key:
            pp_alias[ppk] = cbk
    for ppk, cbk in pp_alias.items():
        if cbk in reg_by_key:
            dst, srcb = batches[reg_by_key[cbk]], batches[reg_by_key[ppk]]
            for col, lst in srcb["cells"].items():
                dst["cells"][col] += lst
        else:
            reg_by_key[cbk] = reg_by_key[ppk]

    # One initial CoQ for every batch on record, first to last (owner,
    # 31.08.2026): the 48 numbered lots of the plan, then a PREDICTED initial
    # CoQ for every register batch the plan does not cover. No packaged lot,
    # grade or number exists for these yet; the schedule holds their place and
    # their documents — a CoQ number is copied from the issuance record at
    # issue, never computed in advance.
    coqs = [{"date": x["issue"], "type": "initial release", "plan": x,
             "number": x["id"], "issued": True} for x in plan]
    covered = {BI.batch_key(PLAN_CB_ALIASES.get(x["cb"], x["cb"])) for x in plan}
    covered |= set(pp_alias)
    stubs = []
    for key, r in icoa.items():
        if key in covered:
            continue
        stub = {"pp": "", "cb": r["batch"], "nm": r["strain"], "grade": "",
                "cls": "", "nom": "", "tol": "", "lo": "", "hi": "", "thc": "",
                "md": "", "pk": "", "id": NO_NUMBER, "ic": NO_NUMBER,
                "issue": r["release_date"], "retest": ""}
        stubs.append(stub)
        coqs.append({"date": r["release_date"], "type": "initial release — predicted",
                     "plan": stub, "number": NO_NUMBER, "issued": False})

    def plus_year(d):
        m = re.match(r"^(\d{2})\.(\d{2})\.(\d{4})$", clean(d))
        return f"{m.group(1)}.{m.group(2)}.{int(m.group(3)) + 1}" if m else ""

    # Then the reissue series, starting from the beginning of Tranche 01 and 02
    # — all 48 lots (19 + 29) retest cannabinoids and mycotoxins and get a CoQ
    # reissue. The 13 already due carry ordinary sequential numbers, 0027…0039,
    # in packaging-date order — confirmed against the rendered files in
    # ISSUE_COQ (CoQ-PP-2026-0034 is P050052, 0038 is P050092, 0039 is
    # P050112). The other 35 are PREDICTED at packaging + 12 months, in plan
    # order — from the beginning.
    for i, x in enumerate([x for x in plan if x.get("retest")]):
        coqs.append({"date": x["retest"], "type": "additional testing (12-month)",
                     "plan": x, "number": f"CoQ-PP-2026-{27 + i:04d}", "issued": True})
    for x in plan:
        if not x.get("retest"):
            coqs.append({"date": plus_year(x["issue"]),
                         "type": "additional testing (12-month) — predicted",
                         "plan": x, "number": NO_NUMBER, "issued": False})

    # The retest programme is universal: every batch past Tranche 02 gets a
    # predicted reissue a year after release, in record order.
    for stub in stubs:
        coqs.append({"date": plus_year(stub["issue"]),
                     "type": "additional testing (12-month) — predicted",
                     "plan": stub, "number": NO_NUMBER, "issued": False})

    rows, per_coq = [], []
    for n, coq in enumerate(coqs, 1):
        x = coq["plan"]
        cb = reg_by_key.get(BI.batch_key(
            PLAN_CB_ALIASES.get(x["cb"], x["cb"])))
        reg = batches.get(cb, {"pnumber": "", "strain": x["nm"],
                               "cells": defaultdict(list)})
        ic_row = icoa.get(BI.batch_key(x["cb"]), {})
        sp = specs.get(x["pp"]) if x["pp"] else None
        additional = coq["type"].startswith("additional")
        blocked = cb == "FB032601"

        thc_criterion = (f"{x['lo']} – {x['hi']} %  (grade {x['grade']}, class "
                         f"THC {x['cls']}, nominal {x['nom']} ± {x['tol']})"
                         if x["lo"] else
                         "Per target grade — no packaged lot or grade assigned in "
                         "the master spec yet")
        # The plan writes the top of the range as nominal + tolerance − 0.01 (an
        # inclusive endpoint: 24.00 ± 2.40 → 21.60–26.39) where the QCSP PDF writes
        # 21.60–26.40. That is one range in two conventions, not a conflict; only a
        # difference beyond 0.01 on either endpoint is one.
        conflict = ""
        if sp:
            m = re.findall(r"[\d.]+", sp["thc_criterion"])
            if len(m) >= 2:
                lo_s, hi_s = float(m[0]), float(m[1])
                if abs(lo_s - float(x["lo"])) > 0.011 or \
                        abs(hi_s - float(x["hi"])) > 0.011:
                    conflict = (f"QCSP 001 prints {sp['thc_criterion']} for this lot; "
                                f"the issue plan's grade range is {x['lo']} – {x['hi']} %. "
                                f"Recorded, not resolved.")

        inhouse = {} if cb else inhouse_cells(x["cb"])
        codes, counts = OrderedDict(), defaultdict(int)
        start = len(rows)
        assay = pick(reg["cells"].get("E", []), additional)[0]
        cnp = assay if assay and assay["family"] == "UKIM CNP potency" else \
            (pick([c for c in reg["cells"].get("E", [])
                   if c["family"] == "UKIM CNP potency"], False)[0])

        for det in dets:
            col = det["column"]
            cands = reg["cells"].get(col, []) if col else []
            if additional:
                # an additional-testing CoQ certifies the additional testing: only a
                # post-release (197-series) certificate may stand behind a result on it
                cands = [c for c in cands if is_reanalysis(c["code"])]
                chosen, others = pick(cands, True)
            else:
                chosen, others = pick(cands, False)
            lim = limits.get(col)

            if det["no"] in ICOA_FIELD and not additional:
                if ic_row.get(ICOA_FIELD[det["no"]], "required") != "required" and cnp:
                    chosen, others = dict(cnp, value="Conforms | Соодветствува"), []

            criterion = det["criterion"]
            if det.get("per_batch_criterion"):
                criterion = thc_criterion

            st = status_of(det, chosen, lim, cb, blocked)
            if chosen is None and not additional and det["no"] in inhouse:
                chosen, others = inhouse[det["no"]], []
                # An outsourced certificate that happens to be the only source for a
                # batch the register has no block for is still an outsourced
                # certificate: judge it, do not declare it missing.
                st = ST_SCAN if chosen.get("inhouse") else \
                    (ST_OFFREG if status_of(det, chosen, lim, cb, blocked) == ST_OK
                     else status_of(det, chosen, lim, cb, blocked))
            if additional and chosen is None:
                if det["no"] in ICOA_FIELD:
                    st = ST_ICOA
                elif det["no"] in RETEST_K:
                    st = ST_AWAIT_K
                elif det["no"] in RETEST_M:
                    st = ST_AWAIT_M
                elif det["no"] in ("9.6", "9.7"):
                    st = ST_REQ
                else:
                    st = ST_OUTSIDE.format(
                        initial=x["id"] if x["id"] != NO_NUMBER else
                        "the batch's initial CoQ (number assigned on issue)")
            counts[st] += 1
            if chosen:
                codes.setdefault(chosen["code"], chosen["lab"])

            route = ""
            if det["no"] in ICOA_FIELD and st == ST_ICOA:
                route = route_for("reissue" if additional else "initial release",
                                  det["no"])

            rows.append(OrderedDict([
                ("Seq", len(rows) + 1), ("CoQ number", coq["number"]),
                ("CoQ type", coq["type"]), ("Basis date", coq["date"] or "—"),
                ("Issue date", ""),
                ("Packaged lot", x["pp"]), ("Cultivation batch", x["cb"]),
                ("Strain", x["nm"]), ("Grade", x["grade"]),
                ("iCoA reference", x["ic"]),
                ("№", det["no"]), ("Group", det["group"]),
                ("Parameter", det["en"]), ("Параметар", det["mk"]),
                ("Method / reference", det["method"]),
                ("Acceptance criterion", criterion),
                ("Result", chosen["value"] if chosen else "—"),
                ("Source document", chosen["code"] if chosen else "—"),
                ("Document date", chosen["date"] if chosen else ""),
                ("Issuing institution", chosen["lab"] if chosen else ""),
                ("Report series", chosen["family"] if chosen else ""),
                ("Status", st),
                ("Performed by", route),
                ("Also on file", "; ".join(f"{o['value']} ({o['code']})"
                                           for o in others)),
            ]))

        # The printed issue date is set at issue, and may be no earlier than the
        # latest of the SOP in-use date (11.05.2026) and the newest document the
        # CoQ cites. An additional-testing CoQ that cites nothing yet cannot
        # precede its 12-month due date either — but once the re-analysis is on
        # file, the due date stops being a floor (Farmahem ran several lots
        # early: J31102501's 197-16 pair is dated 10 months after packaging).
        bound = (sort_date(SOP_EFFECTIVE), SOP_EFFECTIVE)
        cited_dates = [r2["Document date"] for r2 in rows[start:]
                       if r2["Source document"] not in ("", "—")
                       and r2["Document date"]
                       and sort_date(r2["Document date"]) != "9999"]
        for dd in cited_dates:
            bound = max(bound, (sort_date(dd), dd))
        if additional and not cited_dates and coq["date"] \
                and sort_date(coq["date"]) != "9999":
            bound = max(bound, (sort_date(coq["date"]), coq["date"]))
        for r2 in rows[start:]:
            r2["Issue date"] = "≥ " + bound[1]

        per_coq.append({
            "coq": n, "number": coq["number"], "type": coq["type"],
            "date": "≥ " + bound[1], "basis": coq["date"] or "—",
            "issued": coq["issued"], "pp": x["pp"], "cb": x["cb"],
            "in_register": bool(cb), "strain": x["nm"], "grade": x["grade"],
            "cls": x["cls"], "icoa_ref": x["ic"], "banner_thc": x["thc"],
            "md": x["md"], "pk": x["pk"], "thc": thc_criterion,
            "spec_conflict": conflict,
            "spec_doc": sp["spec_doc_code"] if sp else "",
            "codes": list(codes), "counts": counts,
            "outstanding": sorted(
                d["no"] for d in dets if d["no"] in ICOA_FIELD
                and outstanding_of(d, ic_row, reg, additional, blocked)),
        })
        per_coq[-1]["route"] = {n2: route_for(
            "reissue" if additional else "initial release", n2)
            for n2 in per_coq[-1]["outstanding"]}


    return rows, per_coq, dets


def main():
    rows, per_coq, dets = schedule()
    write_workbook(rows, per_coq, dets)
    write_csv(rows)
    report(rows, per_coq, dets)
    plan = icoa_plan(per_coq)
    print()
    print(f"  in-house certificates the routing requires: {len(plan)}")
    for scope in ("Ident A + B", "Foreign matter"):
        n = sum(1 for x in plan if x["scope"] == scope)
        print(f"    {n:>5}  {scope}")
    return 0


def _band(ws, cols, title, sub):
    ws.cell(row=1, column=1, value=title).font = Font(FONT, 15, bold=True, color=INK)
    ws.cell(row=2, column=1, value=sub).font = Font(FONT, 9, color=SUB)
    ws.row_dimensions[1].height, ws.row_dimensions[2].height = 22, 15
    thin = Side("thin", color="D6DBD9")
    for i, (name, width, hint) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
        c = ws.cell(row=4, column=i, value=name)
        c.font = Font(FONT, 9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=ACCENT)
        c.alignment = Alignment("left", "bottom", wrap_text=True)
        c.border = Border(bottom=thin)
        h = ws.cell(row=5, column=i, value=hint)
        h.font = Font(FONT, 8, bold=True, color=INK)
        h.fill = PatternFill("solid", fgColor=REFBG)
        h.alignment = Alignment("left", "center", wrap_text=True)
    ws.row_dimensions[4].height, ws.row_dimensions[5].height = 30, 22


def write_workbook(rows, per_coq, dets):
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    # ------------------------------------------------------------- 1 · schedule
    ws = wb.create_sheet("CoQ Parameter Schedule")
    COLS = [("Seq", 6, ""), ("CoQ number", 17, "ISSUE_COQ numbering"),
            ("CoQ type", 20, ""),
            ("Basis date", 12, "packaging / 12-month due"),
            ("Issue date", 13, "set at issue — no earlier than shown"),
            ("Packaged lot", 12, ""), ("Cultivation batch", 14, ""),
            ("Strain", 18, ""), ("Grade", 7, ""),
            ("iCoA reference", 17, "printed on the CoQ"),
            ("№", 6, "QCSP 001"), ("Group", 26, ""), ("Parameter", 30, ""),
            ("Параметар", 30, ""), ("Method / reference", 34, "QCSP 001 v.03, verbatim"),
            ("Acceptance criterion", 38, ""),
            ("Result", 20, "from the release register"), ("Source document", 20, ""),
            ("Document date", 13, ""), ("Issuing institution", 30, ""),
            ("Report series", 30, ""), ("Status", 40, ""),
            ("Performed by", 54, "the decided route for identity and foreign matter"),
            ("Also on file", 40, "other documents carrying this determination")]
    _band(ws, COLS, "Purely Plant GmbH — Certificate of Quality parameter schedule",
          "The full CoQ universe of 31.08.2026: one initial CoQ per batch on record "
          "(48 numbered by the ISSUE_COQ plan — Tranche 01 + 02 — the rest predicted) "
          "and a 12-month cannabinoid + mycotoxin reissue for every batch, starting "
          "from the beginning of Tranche 01/02 (13 numbered, 35 predicted; later "
          "batches at release + 12 months). Every determination QCSP 001 v.03 "
          "requires, with the document that certifies it. The CoQ SOP is in use "
          "since 11.05.2026: no CoQ prints an earlier issue date, and a predicted "
          "CoQ's number is copied from the issuance record at issue. This is the "
          "transcription source for the controlled blanks on each CoQ.")
    fills = {ST_OOS: RED, ST_BLOCK: RED, ST_UNDET: AMBER, ST_FINDING: AMBER,
             ST_ICOA: BLUE, ST_NONE: AMBER, ST_NOSPEC: AMBER, ST_OK: GREEN,
             ST_SCAN: RED}
    for i, row in enumerate(rows):
        rr = 6 + i
        band = PatternFill("solid", fgColor=BAND) if (i // len(dets)) % 2 else None
        for j, (k, v) in enumerate(row.items(), 1):
            c = ws.cell(row=rr, column=j, value=v)
            c.font = Font(FONT, 9, color=INK)
            c.alignment = Alignment("left", "top", wrap_text=k in
                                    ("Method / reference", "Acceptance criterion",
                                     "Status", "Performed by", "Also on file", "Group"))
            if band:
                c.fill = band
        st = row["Status"]
        cell = ws.cell(row=rr, column=22)
        if st in fills:
            cell.fill = PatternFill("solid", fgColor=fills[st])
            cell.font = Font(FONT, 9, bold=st in (ST_OOS, ST_BLOCK, ST_SCAN), color=INK)
        elif st in (ST_AWAIT_K, ST_AWAIT_M):
            cell.fill = PatternFill("solid", fgColor=AMBER)
    ws.freeze_panes = "K6"
    ws.auto_filter.ref = f"A5:X{5 + len(rows)}"

    # ------------------------------------------------------------- 2 · per CoQ
    ws = wb.create_sheet("CoQ Coverage")
    COLS = [("CoQ number", 17, ""), ("Type", 20, ""),
            ("Issue date", 13, "no earlier than — set at issue"),
            ("Packaged lot", 12, ""), ("Cultivation batch", 14, ""),
            ("Strain", 18, ""), ("Grade", 7, ""), ("Class", 8, ""),
            ("Banner THC", 11, "actual assay"),
            ("Acceptance range", 30, "issue plan, grade nominal ± tolerance"),
            ("Spec doc", 22, "QCSP 001"),
            ("iCoA reference", 17, ""),
            ("Covered", 9, "of 21 on an initial CoQ, 10 in the retest scope"),
            ("To perform", 10, ""),
            ("Not tested", 10, ""), ("Out of spec", 10, ""), ("Undetermined", 12, ""),
            ("Documents", 10, "count"),
            ("Source documents", 70, "every report this CoQ must cite"),
            ("Who performs the rest", 56, ""),
            ("Note", 42, "")]
    _band(ws, COLS, "Purely Plant GmbH — CoQ coverage and document references",
          "One row per CoQ — issued and predicted. Covered counts determinations a "
          "certificate on file stands behind; the identity and foreign-matter routing "
          "is the owner's decision of 31.08.2026, and no issue date precedes the CoQ "
          "SOP's in-use date of 11.05.2026.")
    for i, p2 in enumerate(per_coq):
        rr = 6 + i
        k = p2["counts"]
        # BLOCKED is covered: a certificate on file stands behind it, and it failed.
        # Leaving it out of every bucket made FB032601's row sum to 20 of 21.
        covered = k[ST_OK] + k[ST_FINDING] + k[ST_OOS] + k[ST_UNDET] + k[ST_BLOCK] \
            + k[ST_OFFREG]
        note = p2["spec_conflict"]
        if not p2["in_register"]:
            note = ("Cultivation batch not in the release register — no eCoA on file "
                    "for any determination. The in-house CoA attributes the testing to "
                    "accredited laboratories: locate the physical certificates, scan "
                    "and upload. " + note).strip()
        vals = [p2["number"], p2["type"], p2["date"], p2["pp"], p2["cb"],
                p2["strain"], p2["grade"], f"THC {p2['cls']}", p2["banner_thc"],
                p2["thc"], p2["spec_doc"] or "no QCSP PDF on file", p2["icoa_ref"],
                covered, k[ST_ICOA] + k[ST_SCAN]
                + k[ST_AWAIT_K] + k[ST_AWAIT_M],
                k[ST_NONE] + k[ST_NOSPEC], k[ST_OOS] + k[ST_BLOCK], k[ST_UNDET],
                len(p2["codes"]), "; ".join(p2["codes"]),
                "; ".join(dict.fromkeys(p2["route"].values())) or "—", note]
        band = PatternFill("solid", fgColor=BAND) if i % 2 else None
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=rr, column=j, value=v)
            cell.font = Font(FONT, 9, color=INK)
            cell.alignment = Alignment(
                horizontal="center" if j in (13, 14, 15, 16, 17, 18) else "left",
                vertical="top", wrap_text=j in (10, 19, 20, 21))
            if band:
                cell.fill = band
        if k[ST_OOS] + k[ST_BLOCK]:
            ws.cell(row=rr, column=16).fill = PatternFill("solid", fgColor=RED)
        if k[ST_UNDET]:
            ws.cell(row=rr, column=17).fill = PatternFill("solid", fgColor=AMBER)
        if not p2["in_register"]:
            ws.cell(row=rr, column=21).fill = PatternFill("solid", fgColor=RED)
        elif p2["spec_conflict"]:
            ws.cell(row=rr, column=21).fill = PatternFill("solid", fgColor=AMBER)
    ws.freeze_panes = "D6"
    ws.auto_filter.ref = f"A5:U{5 + len(per_coq)}"

    # ------------------------------------------------------------- 3 · the iCoA plan
    plan = icoa_plan(per_coq)
    ws = wb.create_sheet("iCoA Plan")
    COLS = [("Seq", 6, ""), ("For CoQ", 18, ""), ("CoQ type", 20, ""),
            ("Issue date", 12, ""), ("Packaged lot", 12, ""),
            ("Cultivation batch", 14, ""),
            ("Strain", 20, ""), ("iCoA scope", 22, "one document per scope"),
            ("QCSP 001 nos.", 14, ""), ("Issued by", 34, ""),
            ("Reference on the CoQ", 18, "iCoA-PP-YYYY-NNNN"),
            ("Date", 12, "fill in on issue"),
            ("Analyst", 18, "fill in on issue")]
    _band(ws, COLS,
          "Purely Plant GmbH — in-house certificates the routing requires",
          "Identity A and B in one document, foreign matter in another. Two determinations "
          "on two samples do not belong on one certificate: either would then have to be "
          "reissued whenever the other was repeated. At retest Farmahem covers identity "
          "with the assay, so only foreign matter is left in house.")
    for i, p in enumerate(plan):
        rr = 6 + i
        vals = [i + 1, p["number"], p["coq_type"], p["date"], p["pp"], p["cb"],
                p["strain"], p["scope"], p["determinations"],
                "Purely Plant laboratory", p["icoa_ref"], "", ""]
        band = PatternFill("solid", fgColor=BAND) if i % 2 else None
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=rr, column=j, value=v)
            c.font = Font(FONT, 9, color=INK)
            c.alignment = Alignment(horizontal="center" if j in (1, 2, 9) else "left",
                                    vertical="top")
            if band:
                c.fill = band
            if j >= 11:
                c.fill = PatternFill("solid", fgColor="FFFFFF")
    ws.freeze_panes = "F6"
    ws.auto_filter.ref = f"A5:M{5 + len(plan)}"

    # ------------------------------------------------------------- 4 · sourcing routes
    ws = wb.create_sheet("Sourcing routes")
    COLS = [("№", 6, "QCSP 001"), ("Determination", 32, ""),
            ("Can be performed by", 62, "every route"),
            ("At initial release", 46, "the decided route"),
            ("At retest", 46, "the decided route")]
    _band(ws, COLS, "Who can perform identity and foreign matter, and who will",
          "The four determinations no outsourced release certificate carries a column "
          "for. CNP does not sell foreign matter separately — it comes only inside the "
          "full identity-plus-assay package — and Purely Plant's own laboratory cannot "
          "perform Identity C, which is HPLC/HPTLC.")
    for i, no in enumerate(["1", "2", "3", "7"]):
        d = next(x for x in dets if x["no"] == no)
        rr = 6 + i
        vals = [no, d["en"], " · ".join(CAN_PERFORM[no]),
                route_for("initial release", no), route_for("reissue", no)]
        band = PatternFill("solid", fgColor=BAND) if i % 2 else None
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=rr, column=j, value=v)
            c.font = Font(FONT, 9, color=INK)
            c.alignment = Alignment("left", "top", wrap_text=j >= 3)
            if band:
                c.fill = band
        ws.row_dimensions[rr].height = 46

    # ------------------------------------------------------------- 5 · the spec
    ws = wb.create_sheet("QCSP 001 v.03")
    COLS = [("№", 6, ""), ("Group", 30, ""), ("Parameter", 32, ""),
            ("Параметар", 32, ""), ("Method / reference", 36, "verbatim"),
            ("Acceptance criterion", 44, "verbatim"),
            ("Register column", 14, "which column carries it"),
            ("Supplied by", 24, "")]
    _band(ws, COLS, "QCSP 001 v.03 — the release specification a CoQ certifies against",
          "Section 02 of the signed product specification, identical in all 48 documents "
          "under deliverables/imb_spec_pdfs/. Transcribed verbatim: a paraphrased "
          "acceptance criterion on a release document is a defect.")
    SUPPLY = {"outsourced_certificate": "outsourced laboratory",
              "in_house_icoa": "in-house iCoA — none issued yet",
              "upon_request": "upon request only"}
    for i, d in enumerate(dets):
        rr = 6 + i
        vals = [d["no"], d["group"], d["en"], d["mk"], d["method"], d["criterion"],
                d["column"] or "—", SUPPLY[d["source"]]]
        band = PatternFill("solid", fgColor=BAND) if i % 2 else None
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=rr, column=j, value=v)
            c.font = Font(FONT, 9, color=INK)
            c.alignment = Alignment("left", "top", wrap_text=j in (2, 5, 6))
            if band:
                c.fill = band
            if j == 8 and d["source"] == "in_house_icoa":
                c.fill = PatternFill("solid", fgColor=BLUE)
        ws.row_dimensions[rr].height = 26
    ws.freeze_panes = "C6"
    wb.save(OUT_X)


def write_csv(rows):
    with open(OUT_C, "w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=list(rows[0]))
        w.writeheader()
        w.writerows(rows)


def report(rows, per_coq, dets):
    from collections import Counter
    st = Counter(r["Status"] for r in rows)
    print(os.path.relpath(OUT_X, ROOT))
    print(f"  CoQ documents            {len(per_coq):>5}")
    by_type = Counter(p["type"] for p in per_coq)
    for t in ("initial release", "initial release — predicted",
              "additional testing (12-month)",
              "additional testing (12-month) — predicted"):
        if by_type.get(t):
            print(f"    {by_type[t]:>5}  {t}")
    print(f"  determinations each      {len(dets):>5}  "
          f"({sum(1 for d in dets if d['source'] == 'upon_request')} upon request)")
    print(f"  schedule rows            {len(rows):>5}")
    print()
    for k, v in st.most_common():
        print(f"  {v:>5}  {k}")
    print()
    # Dating rule: the CoQ SOP is in use since 11.05.2026 — no CoQ prints an
    # earlier issue date, and none may print a date earlier than the newest
    # document it cites (or its 12-month due date). Every row shows the bound.
    viol = [r for r in rows if r["Source document"] not in ("", "—")
            and sort_date(r["Document date"]) != "9999"
            and sort_date(r["Document date"]) >
            sort_date(r["Issue date"].lstrip("≥ "))]
    if viol:
        print(f"  {len(viol):>5}  rows cite a document dated AFTER the CoQ's "
              f"earliest-issue bound — MUST NOT HAPPEN, check the build")
        print()
    later = sorted({(p["number"], p["cb"], p["date"]) for p in per_coq
                    if p["date"] != "≥ " + SOP_EFFECTIVE})
    print(f"  issue dates: the CoQ SOP is in use since {SOP_EFFECTIVE}; the plan's "
          f"2025/2026 dates are packaging dates (the numbering series), not issue "
          f"dates. {len(per_coq) - len(later)} CoQs have no date constraint beyond "
          f"the SOP floor; {len(later)} are bound later by a cited document or a "
          f"12-month due date.")
    ready = [p for p in per_coq
             if not (p["counts"][ST_ICOA] + p["counts"][ST_NONE] + p["counts"][ST_SCAN]
                     + p["counts"][ST_AWAIT_K] + p["counts"][ST_AWAIT_M]
                     + p["counts"][ST_NOSPEC])]
    print(f"         A date floor is not a clearance. {len(ready)} of {len(per_coq)} "
          f"CoQs have every determination in scope certified — the rest carry at "
          f"least one result no document stands behind, and a CoQ must never carry a "
          f"conformity assertion that has not been certified.")
    print()
    # The plan's banner THC is supposed to be the lot's actual assay. Where a register
    # block exists, at least one register THC value should equal it; a disagreement
    # means the master spec and the register tell different stories about one lot.
    mism = []
    for p in per_coq:
        if not p["in_register"] or p["type"] != "initial release":
            continue
        got = {r["Result"] for r in rows
               if r["CoQ number"] == p["number"] and r["№"] == "4"}
        alts = {a.split(" (")[0] for r in rows if r["CoQ number"] == p["number"]
                and r["№"] == "4" for a in r["Also on file"].split("; ") if a}
        if p["banner_thc"] not in got | alts:
            mism.append((p["number"], p["cb"], p["banner_thc"],
                         ", ".join(sorted(got | alts))))
    # Where the plan's banner is found ONLY in a 197-series re-analysis, the lot does
    # not mismatch — but it does not agree with a release assay either. On an
    # initial-release CoQ that is the banner quoting a result the batch did not have
    # at release, and it is only defensible because no CoQ may now be issued before
    # 11.05.2026: a certificate of quality speaks as of its date of issue.
    reanalysed = []
    for p in per_coq:
        if not p["in_register"] or p["type"] != "initial release":
            continue
        for r in rows:
            if r["CoQ number"] != p["number"] or r["№"] != "4":
                continue
            if r["Result"] == p["banner_thc"]:
                break
            for a in r["Also on file"].split("; "):
                if a.startswith(p["banner_thc"] + " (197-"):
                    reanalysed.append((p["number"], p["cb"], p["banner_thc"],
                                       r["Result"], a.split("(")[1].rstrip(")")))
                    break
    if mism:
        print(f"  {len(mism):>5}  CoQs whose issue-plan banner THC matches NO register "
              f"assay for the lot — master spec vs register, unresolved")
        for m in mism:
            print(f"         {m[0]} {m[1]}: plan {m[2]} vs register {m[3]}")
        print()
    if reanalysed:
        print(f"  {len(reanalysed):>5}  INITIAL-release CoQs whose banner THC is the "
              f"12-month RE-ANALYSIS, not a release assay")
        for m in reanalysed:
            print(f"         {m[0]} {m[1]}: banner {m[2]} ({m[4]}) vs release assay "
                  f"{m[3]}")
        print("         Not a mismatch and not an error — the master spec carries the "
              "newest assay, and no CoQ may now print an issue date before "
              f"{SOP_EFFECTIVE}, so the certificate speaks as of a date on which the "
              "re-analysis exists. It IS a decision QC should make knowingly: an "
              "initial-release CoQ banner that is not the release result.")
        print()
    # An ISSUED additional-testing CoQ with nothing to cite is a controlled
    # blank on a signed document: the plan records the retest date, but the
    # Farmahem certificate never reached the file.
    ghost = [p for p in per_coq
             if p["type"] == "additional testing (12-month)"
             and not any(c.startswith("197-") for c in p["codes"])]
    if ghost:
        print(f"  {len(ghost):>5}  ISSUED additional-testing CoQs cite NO re-analysis "
              f"certificate on file — the plan records the retest, the 197-series "
              f"certificate is missing: locate and scan")
        for p in ghost:
            print(f"         {p['number']} {p['cb']} — retest per plan {p['basis']}")
        print()
    early = [p for p in per_coq
             if p["type"].endswith("predicted") and p["type"].startswith("additional")
             and any(c.startswith("197-") for c in p["codes"])]
    # P060332 is re-analysed but cannot be queued: its cultivation batch, CC012601/1
    # per the certificate table, is in no register and in no issue plan. A certificate
    # of quality cannot be issued for material whose identity is unresolved.
    unresolved = [p for p in early if not p["pp"] and p["cb"].startswith("P0")]
    early = [p for p in early if p not in unresolved]
    print(f"  {len(early):>5}  batches re-analysed ahead of their 12-month date — the "
          f"197-series pair is on file, so the cannabinoid and mycotoxin half of the "
          f"reissue is already certified.")
    print("         NOT issuable on that alone. The 197 series carries the assay, "
          "total CBD and CBN on its К certificate and the mycotoxins on its М — "
          "nothing else. Identity A, B and C and foreign matter are outstanding on "
          "every one of them, and a CoQ must not carry a conformity assertion that "
          "has not been certified. What they are is FIRST IN THE QUEUE: order "
          "Farmahem's identity, issue the in-house foreign-matter iCoA, then number "
          "and issue.")
    for p in unresolved:
        print(f"         {p['cb']} is re-analysed too and is NOT in that queue: its "
              f"cultivation batch is in no register and no issue plan. Resolve the "
              f"identity first.")
    print()
    # One lot, one conflict — but a lot has two CoQs, its initial and its reissue,
    # and the note belongs on both. Report both granularities: counting the documents
    # alone silently doubles a lot-level finding.
    nospec = sum(1 for p in per_coq if p["spec_conflict"])
    nospec_lots = len({p["cb"] for p in per_coq if p["spec_conflict"]})
    print(f"  {nospec_lots:>5}  lots whose issue-plan grade range disagrees with the "
          f"QCSP 001 PDF, shown on {nospec} CoQ documents (initial + reissue each)")
    print(f"  {sum(1 for p in per_coq if not p['in_register']):>5}  CoQs whose "
          f"cultivation batch has no eCoA on file at all — physical certificates to "
          f"locate and scan")
    print(f"  {sum(1 for p in per_coq if p['counts'][ST_OOS]):>5}  CoQs carrying an "
          f"out-of-specification determination")
    print(f"  {sum(len(p['codes']) for p in per_coq):>5}  document references in total, "
          f"{len({c for p in per_coq for c in p['codes']})} distinct")
    print()
    print(os.path.relpath(OUT_C, ROOT))


if __name__ == "__main__":
    sys.exit(main())
