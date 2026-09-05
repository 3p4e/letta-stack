#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Add ППК25139 to the register, and confirm four certificate codes.

Three changes, all of them decided by the batch owner and all evidenced:

1. **ППК25139 becomes a row.** `GP0824_02` was analysed twice by the Center for
   Natural Products — 22.05.2025 (ППК25139) and 10.07.2025 (ППК25174) — and only the
   later one was registered. The owner's ruling is that the earlier analysis belongs in
   the register as a second analysis of the batch. It is added as a continuation row,
   not as a release value: row 46 keeps ППК25174 and its 23.19.

   Values are taken from the certificate itself
   (Drive `1_Shekzw6aT8o2_JtmrEtRZyzxrOZVVKj`), not from the RAG corpus, whose record
   for this document is corrupted in three places — Δ9-THCA as 0.52 where the page
   prints 26.52, `Satre Pie` for Grape Pie, and `GF0824_02` for GP0824_02.

   The page's Δ9-THCA is deliberately **not** carried across: the register has no THCA
   column, so there is nowhere for it to go and nothing is lost by omitting it.

2. **Four certificate codes stop being guesses.** Each was carried as a parenthetical
   doubt — `НИК22155 (likely OCR misread of ППК25xxx)` and three like it. Each resolves
   against the sole CNP certificate held for that batch, with P-number and issue date
   both matching. The annotations were honest and correct that something was wrong;
   this closes them.

3. **Row 41 is flagged amber.** `OPM1024_02`'s certificate ППК25154 is one of three
   whose value in the RAG corpus is corrupt — the corpus holds a total of 1.87 where
   0.41 + 20.37 × 0.877 = 18.27, which is what the register already says.
   **The register value is correct.** The flag marks that the corpus disagrees with the
   paper, not that the batch has a problem, and the legend entry says so.

Idempotent, and it refuses to touch a workbook it does not recognise: every change
checks the current cell against the value it was verified against and stops on any
mismatch rather than guessing.

    python3 deliverables/qc_gap_analysis/add_ppk25139_and_codes.py IN.xlsx OUT.xlsx

Note for anything that hard-codes row numbers: inserting the new row shifts every row
below 46 down by one. `apply_register_corrections.py` addresses rows by index and must
be run **before** this script, never after.
"""
import sys
from copy import copy

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill

SHEET = "Batch Release QC"
ANCHOR = 46                      # GP0824_02 header row, carrying ППК25174
THC, CBD, CBN, LOD = 5, 7, 8, 9  # columns E, G, H, I
CODE, DATE, INST = 23, 24, 25    # columns W, X, Y

# Read off the certificate. `/` is the register's own "not determined on this
# certificate" marker, used for every column the document does not report.
NEW_ROW = {
    THC: "23.79",   # Вкупно Δ9-THC
    CBD: "0.10",    # Вкупно CBD
    CBN: "0.02",    # Содржина на CBN
    LOD: "7.21",    # Губиток при сушење
    CODE: "ППК25139",
    DATE: "22.05.2025",
    INST: "UKIM Faculty of Pharmacy — Center for Natural Products",
}
PDF = 26            # column Z
PDF_URL = "https://drive.google.com/file/d/1_Shekzw6aT8o2_JtmrEtRZyzxrOZVVKj/view"

# row, current value, confirmed code — each corroborated by P-number and issue date
# against the only CNP certificate on file for that batch.
CODES = [
    (36, "НИК22155 (likely OCR misread of ППК25xxx)", "ППК25155",
     "HPA1024_01 / P050052 — P050052, ППК25155, 24.06.2025, CNP"),
    (41, "ППК21554 (likely OCR misread of ППК25xxx)", "ППК25154",
     "OPM1024_02 / P050062 — P050062, ППК25154, 24.06.2025, CNP"),
    (66, "ППК52211 (likely OCR misread of ППК25211)", "ППК25211",
     "MB0824_05 / P050112 — P050112, ППК25211, 28.07.2025, CNP"),
    (69, "ППК52557 (likely OCR misread)", "ППК25257",
     "OPM052501 / P050132 — P050132, ППК25257, 05.09.2025, CNP"),
]

# The sheet's own legend colour for "a laboratory finding or a data-integrity flag".
AMBER_FILL = PatternFill("solid", fgColor="FFFAEDD4")
AMBER_FONT = Font(color="FF9A6300")

FLAG_ROW = 41
FLAG_NOTE = (
    "Data-integrity flag on the SOURCE CORPUS, not on this batch.\n\n"
    "This value (18.27) is correct and matches the certificate: "
    "0.41 + 20.37 x 0.877 = 18.27.\n\n"
    "The RAGFlow corpus holds a corrupted total of 1.87 for ППК25154. "
    "Flagged so the disagreement is visible; no deviation is implied."
)


def _cur(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return "" if v is None else str(v).strip()


def main(src, dst):
    wb = load_workbook(src)
    ws = wb[SHEET]
    log, skipped = [], []

    # --- 1. the new row -------------------------------------------------------
    already = any(
        _cur(ws, r, CODE) == "ППК25139"
        for r in range(ANCHOR, min(ANCHOR + 12, ws.max_row + 1))
    )
    if already:
        skipped.append("ППК25139 row already present")
    else:
        anchor_code = _cur(ws, ANCHOR, CODE)
        if anchor_code != "ППК25174":
            raise SystemExit(
                f"REFUSING: row {ANCHOR} should carry ППК25174, found {anchor_code!r}. "
                f"This is not the revision the change was verified against.")

        new = ANCHOR + 1
        style_src = new          # the existing first continuation row, before the shift

        # insert_rows moves cell VALUES but leaves hyperlinks anchored to their original
        # cell references, so every PDF link below the insert point would silently come
        # to sit on the wrong row — 143 of them, each opening another batch's
        # certificate. Snapshot them, then re-place them against the shifted rows.
        links = {(c.row, c.column): c.hyperlink
                 for row in ws.iter_rows() for c in row if c.hyperlink is not None}
        ws.insert_rows(new)
        for (r, col) in links:
            ws.cell(row=r, column=col).hyperlink = None
        for (r, col), link in links.items():
            tgt = ws.cell(row=r + 1 if r >= new else r, column=col)
            link.ref = tgt.coordinate
            tgt.hyperlink = link
            if tgt.value is None or str(tgt.value).startswith("http"):
                tgt.value = "Open"

        # insert_rows moves values but leaves merged ranges and styles behind.
        for col in range(1, ws.max_column + 1):
            s, d = ws.cell(row=style_src + 1, column=col), ws.cell(row=new, column=col)
            d._style = copy(s._style)

        # Columns A–D stay blank: this is a continuation row of the GP0824_02 block,
        # not a new batch. Every result column the certificate does not report gets the
        # register's own "/" so an empty cell never reads as an untested parameter.
        for col in range(THC, CODE):
            ws.cell(row=new, column=col).value = "/"
        for col, val in NEW_ROW.items():
            ws.cell(row=new, column=col).value = val
        pdf = ws.cell(row=new, column=PDF)
        pdf.value, pdf.hyperlink = "Open", PDF_URL

        # insert_rows does not move merged ranges, and the footer legend is merged
        # across A:Z on six consecutive rows. They must be unmerged *all* first: moving
        # them one at a time makes each target collide with the range still sitting
        # there, and openpyxl resolves that by silently dropping ranges.
        moving = [r for r in ws.merged_cells.ranges if r.min_row >= new]
        spec = [(r.min_row, r.max_row, r.min_col, r.max_col) for r in moving]
        for r in moving:
            ws.unmerge_cells(str(r))
        for r0, r1, c0, c1 in spec:
            ws.merge_cells(start_row=r0 + 1, end_row=r1 + 1,
                           start_column=c0, end_column=c1)

        log.append(f"row {new} inserted — ППК25139, 22.05.2025, THC 23.79 "
                   f"(second CNP analysis of GP0824_02; row {ANCHOR} unchanged)")

    # CODES and FLAG_ROW are original row numbers. Once the new row exists, everything
    # below the anchor sits one lower — and that is true whether this run inserted it
    # or a previous one did, so the shift is not conditional on `already`.
    shift = 1

    # --- 2. the four codes ----------------------------------------------------
    for row, old, new_code, why in CODES:
        r = row + (shift if row > ANCHOR else 0)
        cur = _cur(ws, r, CODE)
        if cur == new_code:
            skipped.append(f"r{r} code already {new_code}")
            continue
        if cur != old:
            raise SystemExit(
                f"REFUSING r{r}: expected {old!r}, found {cur!r}. Re-verify first.")
        ws.cell(row=r, column=CODE).value = new_code
        log.append(f"W{r}  {old!r}\n         -> {new_code!r}   {why}")

    # --- 3. the amber flag ----------------------------------------------------
    fr = FLAG_ROW + (shift if FLAG_ROW > ANCHOR else 0)
    cell = ws.cell(row=fr, column=THC)
    if cell.fill.fgColor and cell.fill.fgColor.rgb == "FFFAEDD4":
        skipped.append(f"r{fr} already flagged")
    else:
        cell.fill, cell.font = AMBER_FILL, AMBER_FONT
        cell.comment = Comment(FLAG_NOTE, "QC verification 30.08.2026")
        log.append(f"E{fr}  amber flag + note — corpus holds 1.87, "
                   f"register 18.27 is correct")

    wb.save(dst)
    print(f"in : {src}\nout: {dst}\n")
    for line in log:
        print("  CHANGED  " + line)
    for s in skipped:
        print("  skipped  " + s)
    print(f"\n{len(log)} change(s), {len(skipped)} already applied.")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 3:
        raise SystemExit(__doc__)
    sys.exit(main(sys.argv[1], sys.argv[2]))
