#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Correct the eleven dates of issue that the page disagrees with.

Evidence in `review/DATE_OF_ISSUE_VERIFICATION_2026-08-31.md`. Every replacement value
was read off the certificate, not inferred from the filename.

    python3 deliverables/qc_gap_analysis/apply_date_corrections.py IN.xlsx OUT.xlsx

**How this block was found without reading 266 documents.** The certificate folder names
each file `<batch>_<code>, <DD.MM.YYYY>_<lab>.pdf`. That date is not the document, but it
is an independent transcription of it — nobody copied it from the register. Comparing all
266 register dates against it made 231 checkable and left 16 disagreements, and only
those 16 needed a page. 215 agreed, and agreed with something the register did not
produce.

Three kinds of defect came out of it.

**Two rows hold the wrong date field.** These IPH certificates print *two* dates:
`Дата на прием` in the head and `Дата: ... година` above the signatures. The column is
labelled Date of issue and holds the receipt date:

| Row | Certificate | Page: receipt | Page: issue | Register held |
|---|---|---|---|---|
| 43 | `588/1067/25` | 13.06.2025 | **18.06.2025** | 13.06.2025 |
| 126 | `1219/2170/25` | 26.11.2025 | **01.12.2025** | 26.11.2025 |

Right document, wrong field — the same shape as the Farmahem U-column and the register's
metal limits: nothing here is a misread character.

**Eight rows hold prose instead of a date** — `Completed Nov-Dec 2025`, `Issued ~Dec
2025`, `Issued Dec 2025`. Placeholders that were never resolved. All eight documents were
read: the four CNP certificates print `Скопје, 12.12.2025 год.` and the four IPH ones
print `Дата: 05.12.2025 година`.

**One row carries an annotation** — `20.01.2026 [see source]`. The date is right; the
note is scaffolding left in a released register.

Not corrected, and deliberately: rows 123, 127, 131, 135 and 139 hold `21.01.2026`
against a code cell that names **two** documents, `PP CoA #027 / ППК25370`. The CNP
certificate is dated 28.11.2025 or 12.12.2025; 21.01.2026 is plausibly the in-house CoA's
own date. The cell cannot be corrected without first deciding which document the row is
for — a question for QC, not a transcription fix.

Idempotent, and it refuses a workbook whose cells differ from what was verified.
"""
import sys

from openpyxl import load_workbook
from openpyxl.comments import Comment

SHEET = "Batch Release QC"
DATE = 24                         # column X
AUTHOR = "QC page verification 31.08.2026"

# row, current value, page value, certificate, what the page shows
DATES = [
    (43,  "13.06.2025", "18.06.2025", "588/1067/25",
     "page: Дата на прием 13.06.2025, Дата 18.06.2025 — the register held the receipt date"),
    (126, "26.11.2025", "01.12.2025", "1219/2170/25",
     "page: Дата на прием 26.11.2025, Дата 01.12.2025 — the register held the receipt date"),
    (128, "Completed Nov-Dec 2025", "12.12.2025", "ППК25378", "page: Скопје, 12.12.2025 год."),
    (132, "Completed Nov-Dec 2025", "12.12.2025", "ППК25379", "page: Скопје, 12.12.2025 год."),
    (136, "Completed Nov-Dec 2025", "12.12.2025", "ППК25380", "page: Скопје, 12.12.2025 год."),
    (140, "Completed Nov-Dec 2025", "12.12.2025", "ППК25381", "page: Скопје, 12.12.2025 год."),
    (130, "Issued ~Dec 2025", "05.12.2025", "1228/2194/25", "page: Дата 05.12.2025 година"),
    (134, "Issued ~Dec 2025", "05.12.2025", "1229/2195/25", "page: Дата 05.12.2025 година"),
    (138, "Issued Dec 2025",  "05.12.2025", "1226/2192/25", "page: Дата 05.12.2025 година"),
    (142, "Issued Dec 2025",  "05.12.2025", "1227/2193/25", "page: Дата 05.12.2025 година"),
    (185, "20.01.2026 [see source]", "20.01.2026", "5/0008/26",
     "the date was already right; the bracketed note is scaffolding"),
]

NOTES = {
    43: ("Corrected 31.08.2026 from the rendered page of 588/1067/25.\n\n"
         "The certificate prints two dates: Дата на прием 13.06.2025 (receipt) in the "
         "head, and Дата: 18.06.2025 година above the signatures (issue). This column "
         "is Date of issue; it held the receipt date."),
    126: ("Corrected 31.08.2026 from the rendered page of 1219/2170/25.\n\n"
          "Дата на прием 26.11.2025 (receipt), Дата: 01.12.2025 година (issue). This "
          "column is Date of issue; it held the receipt date."),
}


def _cur(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return "" if v is None else str(v).strip()


def main(src, dst):
    wb = load_workbook(src)
    ws = wb[SHEET]
    log, skipped = [], []

    for row, old, new, cert, why in DATES:
        cur = _cur(ws, row, DATE)
        if cur == new:
            skipped.append(f"r{row} already {new}")
            continue
        if cur != old:
            raise SystemExit(
                f"REFUSING X{row}: expected {old!r}, found {cur!r}. Not the revision "
                f"this was verified against — re-read {cert} first.")
        ws.cell(row=row, column=DATE).value = new
        log.append(f"X{row:<4} {old!r} -> {new!r}   {cert} — {why}")

    for row, note in NOTES.items():
        cell = ws.cell(row=row, column=DATE)
        if cell.comment is None:
            cell.comment = Comment(note, AUTHOR)

    wb.save(dst)
    print(f"in : {src}\nout: {dst}\n")
    for line in log:
        print("  CHANGED  " + line)
    for s in skipped:
        print("  skipped  " + s)
    print(f"\n{len(log)} change(s), {len(skipped)} already applied.")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 3:
        raise SystemExit(__doc__)
    sys.exit(main(sys.argv[1], sys.argv[2]))
