#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Purely Plant has its own release specification, and it changes four dispositions back.

    python3 deliverables/qc_gap_analysis/apply_qcsp001_reconciliation.py IN.xlsx OUT.xlsx

Evidence in `review/OOS_RECTIFICATION_2026-08-31.md`, section "Reopened".

## What was missed

`apply_acceptance_criterion_corrections.py` withdrew four TYMC amber flags on the ground
that Ph. Eur. 5.1.4 reads an enumeration criterion of `10⁴ CFU/g` as a maximum acceptable
count of 20 000. That reading of the **pharmacopoeia** is right and is not in question here.

What it missed is that the pharmacopoeia is not the only specification in force. The same
authority that gives the ×2 rule also scopes it: *a manufacturer's own specification means
what its own document says, and the interpretation note does not automatically extend to
it.* **Purely Plant has such a document, it is in this repository, and no pass had opened
it.**

`deliverables/imb_spec_pdfs/SPC_FINAL_ImB_PDF/**` holds 48 product specifications,
`QCSP 001 v.03`, signed by the QC and QA Managers. **All 48 state the same thing**, section
09 *Microbiological Purity | Микробиолошка Чистота*:

| Parameter | Reference printed | Criterion printed |
|---|---|---|
| TAMC | `Ph. Eur. 2.6.12 cat. C` | `≤ 10⁵ CFU/g` |
| TYMC | `Ph. Eur. 2.6.12 cat. C` | `≤ 10⁴ CFU/g` |
| Bile-tolerant gram-neg. | `Ph. Eur. 2.6.31 cat. C` | `≤ 10⁴ CFU/g` |

**Two readings, and this register cannot choose between them.**

* The criterion cites **Ph. Eur. 2.6.12**, and 2.6.12's own *Interpretation of results* is
  precisely where the ×2 note lives. A specification that adopts a chapter arguably adopts
  how that chapter says to read it: `≤ 10⁴` then means 20 000, and the four results conform.
* The criterion is written `≤ 10⁴ CFU/g` and states **no maximum acceptable count**. Purely
  Plant documents distinguish the two ideas when they mean to — the in-house CoA form prints
  `<10^4, max 50 000 CFU/g`. QCSP 001 prints no max. Read as a plain in-house ceiling,
  `≤ 10⁴` means 10 000 and the four results **fail by 20 % to 90 %**.

So this script does not restore an out-of-specification determination and it does not leave
the results cleared. It puts them back to **undetermined**, which is what they are:

| Row | Batch | Certificate | TYMC | vs Ph. Eur. 20 000 | vs QCSP 001 as written, 10 000 |
|---|---|---|---|---|---|
| 35 | GG1024_02 | `472/0863/25` | 19 000 | conforms | **1.90× over** |
| 75 | CJ052501-1 | `949/1687/25` | 17 000 | conforms | **1.70× over** |
| 38 | HPA1024_01 | `587/1066/25` | 15 000 | conforms | **1.50× over** |
| 57 | GP0824_03 | `628/1129/25` | 12 000 | conforms | **1.20× over** |

The five confirmed exceedances are untouched: 26 000 to 49 000 fail under **both** readings
and under the register's superseded ×5 parenthetical is the only way they ever passed.

**Two further defects in QCSP 001, recorded and not resolved here.** Both bear on which
reading governs, and neither is this register's to settle:

1. It labels every microbiology row **`cat. C`** while printing 10⁵ / 10⁴. Ph. Eur. 5.1.8
   Category C is TAMC 10⁴ and TYMC 10² — which is exactly what certificates `1220/2171/25`
   and `1221/2172/25` print. If Category C is meant, the TYMC criterion is 10² and these
   results fail by a further two decades.
2. **Ph. Eur. 2.6.12 has no categories at all.** It is the enumeration *method*; the
   categories live in 5.1.4 and 5.1.8. `Ph. Eur. 2.6.12 cat. C` cites a chapter for
   something that chapter does not contain.

And the version in force on the test date is unestablished: the header reads
`QCSP_001_HPA-I_v.01`, the footer `QCSP 001 v.03`, and the signature date is 01.06.2026 —
a year after the June-2025 testing these releases rest on.

## The stability sheet disagrees with two of its own certificates

Checked while reopening this. The **Stability Testing Programme** sheet has never been
page-verified — `CNP_PAGE_VERIFICATION_2026-08-31.md`'s "202 of 202" covers the Batch
Release QC sheet only. Against the CNP page reads, **eight of its ten rows match exactly
and two do not**:

| Row | Certificate | Sheet held | Page reads |
|---|---|---|---|
| 16 | `ППК26037` | CBN **0.19**, no remark | CBN **1.09**, over the certificate's own `≤ 1.00 %` |
| 18 | `ППК26058` | Δ⁹-THC **0.29**, Δ⁹-THCA **0.97**, Total **1.17** | **16.99**, **0.07**, **17.05** |

Row 16 is the worse of the two: a result over its printed limit sits on the stability sheet
as a comfortably passing 0.19, with the remark column blank where its two siblings carry
`CBN … exceeds the ≤ 1.00 % limit`. The register's own Batch Release QC row 95 holds 1.09,
page-read and amber-flagged, so the workbook contradicts itself across two sheets.

**Row 18 is the one worth understanding.** `0.29 + 0.97 × 0.877 = 1.14`, and the sheet
prints 1.17 — **within the R4 tolerance.** The page's `16.99 + 0.07 × 0.877 = 17.05` is
equally consistent. So a set of three corrupted values reproduced the certificate's own
arithmetic proof. **R4 cannot catch this**, and that is worth knowing about R4: it proves a
document self-consistent, not correct. Only the page settles it.

Both are corrected here from the page reads, on the same standard as every other correction
in this campaign, and the remark on row 16 is written to match its siblings.

## And the verdicts were never read

The four CBN cell comments said the stability certificates "carry no verdict of any kind".
**That is not established.** `CNP_PAGE_VERIFICATION_2026-08-31.md` records that the CNP
pages were rendered *cropped to 0.10–0.82 of page height*, and that `ППК26127`'s failure was
found only on a **second, uncropped** render because "the ЗАКЛУЧОК heading ran off the
bottom edge". The ЗАКЛУЧОК block sits below the results table — below the crop. So on 72 of
73 CNP certificates the region where a verdict appears was never in the image.

The comments are amended to say that instead. An absence nobody looked for is not evidence.

Idempotent, and it refuses a workbook whose cells differ from what was verified.
"""
import sys

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

REL, STAB = "Batch Release QC", "Stability Testing Programme"
TYMC, CBN = 11, 8                        # Batch Release QC columns K and H
S_CODE, S_CBN, S_D9, S_THCA, S_TOTAL, S_REMARK = 6, 12, 13, 14, 16, 17
AUTHOR = "QC acceptance-criterion rectification 31.08.2026"
AMBER = PatternFill("solid", fgColor="FFFAEDD4")

PH_EUR, QCSP = 20000, 10000

QCSP_TEXT = (
    "Purely Plant's own release specification, QCSP 001 v.03, section 09 "
    "\"Microbiological Purity\", prints:\n\n"
    "    TYMC | Вкупен број квасци/мувли    Ph. Eur. 2.6.12 cat. C    <= 10^4 CFU/g\n\n"
    "All 48 product specifications in deliverables/imb_spec_pdfs/ state it identically, "
    "signed by the QC and QA Managers.\n\n"
    "TWO READINGS, and this register cannot choose between them:\n\n"
    "  · The criterion cites Ph. Eur. 2.6.12, whose own \"Interpretation of results\" is "
    "where the x2 note lives. A specification that adopts a chapter arguably adopts how "
    "that chapter says to read it — 10^4 then means 20 000 and this result conforms.\n\n"
    "  · The criterion states NO maximum acceptable count. Purely Plant documents "
    "distinguish the two ideas when they mean to: the in-house CoA form prints "
    "\"<10^4, max 50 000 CFU/g\". QCSP 001 prints no max. Read as a plain in-house "
    "ceiling, 10^4 means 10 000 and this result is over it.\n\n"
    "Two further defects in QCSP 001 bear on the answer and are not resolved here:\n"
    "  · it labels the row \"cat. C\" while printing 10^4. Ph. Eur. 5.1.8 Category C is "
    "TYMC 10^2 — which is what certificates 1220/2171/25 and 1221/2172/25 print. On that "
    "reading this result fails by two further decades.\n"
    "  · Ph. Eur. 2.6.12 has no categories. It is the enumeration method; the categories "
    "live in 5.1.4 and 5.1.8.\n\n"
    "And the version in force on the test date is unestablished: header QCSP_001_..._v.01, "
    "footer QCSP 001 v.03, signed 01.06.2026 — a year after this testing.\n\n"
    "DECISION OWED BY QA: whether QCSP 001's \"<= 10^4 CFU/g\" means 10 000 or 20 000, and "
    "which version was in force. Four release results turn on it."
)

# row, batch, certificate, printed value, counted value
UNDETERMINED = [
    (35, "GG1024_02",  "472/0863/25", "1,9 x 10^4", 19000),
    (38, "HPA1024_01", "587/1066/25", "1,5 x 10^4", 15000),
    (57, "GP0824_03",  "628/1129/25", "1,2 x 10^4", 12000),
    (75, "CJ052501-1", "949/1687/25", "1,7 x 10^4", 17000),
]

WITHDRAWN_OPENER = "Amber flag WITHDRAWN 31.08.2026. This result conforms."

# sheet row, column, old value, new value, what the page reads
STABILITY = [
    (16, S_CBN,   0.19, 1.09, "ППК26037", "CBN 1.09 %"),
    (18, S_D9,    0.29, 16.99, "ППК26058", "Δ9-THC 16.99 %"),
    (18, S_THCA,  0.97, 0.07, "ППК26058", "Δ9-THCA 0.07 %"),
    (18, S_TOTAL, 1.17, 17.05, "ППК26058", "Вкупно Δ9-THC 17.05 %"),
]

REMARK_16 = "CBN 1.09 % exceeds the ≤ 1.00 % limit stated on the certificate"

STAB_NOTE = (
    "Corrected 31.08.2026 from the CNP page read of {code} "
    "(review/cnp_page_reads_2026-08-31.json).\n\n"
    "This sheet has never been page-verified — CNP_PAGE_VERIFICATION_2026-08-31.md's "
    "\"202 of 202\" covers the Batch Release QC sheet only. Checked against the page reads, "
    "eight of its ten rows match exactly and two did not: this one and its sibling.\n\n"
    "The page reads {reads}.\n\n"
    "{extra}"
)

STAB_16_EXTRA = (
    "The sheet held 0.19 with the remark column blank, so a result over the certificate's "
    "own <= 1.00 % limit read as comfortably inside it — while Batch Release QC row 95 "
    "holds 1.09, page-read and amber-flagged. The workbook contradicted itself across two "
    "sheets. The remark now matches rows 8, 12 and 18."
)

STAB_18_EXTRA = (
    "This row is worth understanding rather than just fixing. The sheet held Δ9-THC 0.29, "
    "Δ9-THCA 0.97 and Total 1.17, and 0.29 + 0.97 x 0.877 = 1.14 — WITHIN the R4 tolerance "
    "of 0.06. The page's 16.99 + 0.07 x 0.877 = 17.05 is equally consistent.\n\n"
    "So three corrupted values together reproduced the certificate's own arithmetic proof. "
    "R4 shows a document self-consistent, not correct. Only the page settles it."
)

CBN_OLD_TAIL = (
    "Amber, not red: the certificate carries no verdict of any kind. Whether an "
    "accelerated-condition sample is expected to meet the release limit is a question for "
    "the stability protocol, not something this register can settle."
)

CBN_NEW_TAIL = (
    "Amber, not red, for two reasons.\n\n"
    "Whether an accelerated-condition sample is expected to meet the release limit is a "
    "question for the stability protocol, not something this register can settle.\n\n"
    "And the certificate's verdict was never read. An earlier version of this comment said "
    "the certificate \"carries no verdict of any kind\"; that is not established. The CNP "
    "pages were rendered cropped to 0.10-0.82 of page height, and ППК26127's failure was "
    "found only on a second, uncropped render because its ЗАКЛУЧОК heading ran off the "
    "bottom edge. The ЗАКЛУЧОК block sits below the results table, so on 72 of the 73 CNP "
    "certificates the region where a verdict appears was never in the image. An absence "
    "nobody looked for is not evidence."
)


def _num(cell):
    return None if cell.value is None else float(cell.value)


def main(src, dst):
    wb = load_workbook(src)
    rel, stab = wb[REL], wb[STAB]
    log, skipped = [], []

    for row, batch, cert, printed, counted in UNDETERMINED:
        cell = rel.cell(row=row, column=TYMC)
        rgb = getattr(cell.fill.fgColor, "rgb", None)
        if cell.fill.patternType == "solid" and rgb == "FFFAEDD4" and \
                cell.comment and cell.comment.text.startswith("Result UNDETERMINED"):
            skipped.append(f"K{row} already undetermined ({batch})")
            continue
        if cell.comment is None or not cell.comment.text.startswith(WITHDRAWN_OPENER):
            raise SystemExit(
                f"REFUSING K{row}: expected the withdrawal comment from "
                f"apply_acceptance_criterion_corrections.py, found "
                f"{(cell.comment.text[:60] + '…') if cell.comment else 'no comment'!r}.")
        cell.fill = AMBER
        cell.comment = Comment(
            "Result UNDETERMINED, 31.08.2026. Not a confirmed exceedance and not cleared.\n\n"
            f"Certificate {cert} reports TYMC {printed} CFU/g = {counted:,} CFU/g."
            .replace(",", " ") + "\n\n"
            f"  · Against the Ph. Eur. 5.1.4 maximum acceptable count of {PH_EUR:,} it "
            f"CONFORMS.\n"
            f"  · Against QCSP 001's \"<= 10^4 CFU/g\" read as written, {QCSP:,}, it is "
            f"{counted / QCSP:.2f}x OVER.\n\n".replace(",", " ")
            + "An amber flag withdrawn earlier on 31.08.2026 is reinstated here. The "
            "withdrawal read the pharmacopoeia correctly and stopped there; it did not "
            "open Purely Plant's own release specification, which had never been consulted "
            "by any pass in this campaign.\n\n" + QCSP_TEXT, AUTHOR)
        log.append(f"K{row}  amber REINSTATED as undetermined  {batch:<11} {printed} = "
                   f"{counted:,} — conforms vs {PH_EUR:,}, over vs {QCSP:,}"
                   .replace(",", " "))

    for row, col, old, new, code, reads in STABILITY:
        cell = stab.cell(row=row, column=col)
        letter = cell.column_letter
        got = _num(cell)
        if got is not None and abs(got - new) < 1e-9:
            skipped.append(f"{STAB}!{letter}{row} already {new}")
            continue
        if got is None or abs(got - old) > 1e-9:
            raise SystemExit(
                f"REFUSING {STAB}!{letter}{row}: expected {old}, found {got!r}.")
        cell.value = new
        extra = STAB_16_EXTRA if row == 16 else STAB_18_EXTRA
        if cell.comment is None:
            cell.comment = Comment(
                STAB_NOTE.format(code=code, reads=reads, extra=extra), AUTHOR)
        log.append(f"{STAB}!{letter}{row}  {old} -> {new}   {code} — the page reads {reads}")

    rem = stab.cell(row=16, column=S_REMARK)
    if (rem.value or "") == REMARK_16:
        skipped.append(f"{STAB}!Q16 remark already present")
    elif rem.value:
        raise SystemExit(f"REFUSING {STAB}!Q16: expected empty, found {rem.value!r}.")
    else:
        rem.value = REMARK_16
        log.append(f"{STAB}!Q16  remark added, matching rows 8, 12 and 18")

    for row in (53, 59, 95, 97):
        cell = rel.cell(row=row, column=CBN)
        if cell.comment is None:
            raise SystemExit(f"REFUSING H{row}: expected a CBN comment, found none.")
        text = cell.comment.text
        if CBN_NEW_TAIL in text:
            skipped.append(f"H{row} verdict caveat already recorded")
            continue
        if CBN_OLD_TAIL not in text:
            raise SystemExit(
                f"REFUSING H{row}: the comment does not carry the verdict claim this "
                f"amends.")
        cell.comment = Comment(text.replace(CBN_OLD_TAIL, CBN_NEW_TAIL), AUTHOR)
        log.append(f"H{row}  CBN comment amended: the certificate's verdict was never read")

    wb.save(dst)
    print(f"in : {src}\nout: {dst}\n")
    for line in log:
        print("  CHANGED  " + line)
    for s in skipped:
        print("  skipped  " + s)
    print(f"\n{len(log)} change(s), {len(skipped)} already applied.")
    print("\nTYMC: 5 confirmed out of specification, 4 UNDETERMINED pending QA's reading "
          "of QCSP 001 v.03.")
    print("QCSP 001 also labels its microbiology rows 'Ph. Eur. 2.6.12 cat. C' — a chapter "
          "that has no categories — while printing Category B figures. Both are QA items.")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 3:
        raise SystemExit(__doc__)
    sys.exit(main(sys.argv[1], sys.argv[2]))
