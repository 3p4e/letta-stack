#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Apply what reading the Farmahem cannabinoid reports found.

All 31 reports were read off their rendered pages; evidence is in
`review/FARMAHEM_PAGE_VERIFICATION_2026-08-31.md`, raw readings in
`review/farmahem_page_reads_2026-08-31.json`. 93 register values compared, 89 agree.

    python3 deliverables/qc_gap_analysis/apply_farmahem_corrections.py IN.xlsx OUT.xlsx

**Every one of the four defects is the same mistake in a different disguise: the cell
holds a number from the certificate, but not the one under `Резултат` for that
component.** Nothing here is a misread digit.

Each Farmahem report prints one table with four columns —
`Име на компонента | Кратенка | Резултат [%w/w] | U [%w/w]` — and three rows, CBD, CBN
and Δ9-THC. So next to every result sits its own uncertainty, and immediately above or
below sits another component's result. Three of the four ways to pick the wrong cell
have happened:

| Row | Certificate | Register held | Page prints | What went wrong |
|---|---|---|---|---|
| 9 | `197-1-К/26` | CBN `0.02` | CBN `0.28`, U `0.02` | the **U column** |
| 276 | `197-7-К/26` | CBD `<LOQ`, CBN `0.22` | CBD `0.22`, CBN `< LOQ` | the **two rows swapped** |
| 286 | `197-13-К/26` | CBN `N.D.` | CBN `0.36` | a real result recorded as **not detected** |

The shape of the error matters more than any of the numbers. **Δ9-THC is correct on all
31 reports** — it is the release-critical parameter and it is perfect. All four defects
are in the CBD/CBN pair: the two minor cannabinoids, adjacent to each other and to the
uncertainty column, where a value looks plausible whatever cell it came from. A range
check cannot see any of this. Only reading the column header can.

Row 286 is the one a reviewer should care about: `N.D.` states the laboratory looked and
found nothing, when it in fact reported 0.36 %. That is a difference in kind, not
degree.

Idempotent, and it refuses a workbook whose cells differ from what was verified.
"""
import sys

from openpyxl import load_workbook
from openpyxl.comments import Comment

SHEET = "Batch Release QC"
CBD, CBN = 7, 8                   # columns G and H
AUTHOR = "QC page verification 31.08.2026"

# row, column, current value, page value, certificate, what went wrong
CELLS = [
    (9, CBN, "0.02", "0.28", "197-1-К/26  BG1024",
     "the register took the U (uncertainty) column; the Резултат column reads 0.28"),
    (276, CBD, "<LOQ", "0.22", "197-7-К/26  P060352",
     "CBD and CBN are transposed: 0.22 is the CBD result"),
    (276, CBN, "0.22", "< LOQ", "197-7-К/26  P060352",
     "CBD and CBN are transposed: CBN is < LOQ"),
    (286, CBN, "N.D.", "0.36", "197-13-К/26  HPA1024",
     "a reported result of 0.36 % was recorded as not detected"),
]

NOTES = {
    9: ("Corrected 31.08.2026 from the rendered page of 197-1-К/26.\n\n"
        "The certificate's table is Име | Кратенка | Резултат | U. Total CBN reads "
        "0.28 under Резултат and 0.02 under U. The register held 0.02 — the "
        "uncertainty, not the result. Confirmed at 2.6x magnification."),
    276: ("Corrected 31.08.2026 from the rendered page of 197-7-К/26.\n\n"
          "The page reads Total CBD 0.22 and Total CBN < LOQ. The register had the two "
          "the other way round. Confirmed at 2.6x magnification."),
    286: ("Corrected 31.08.2026 from the rendered page of 197-13-К/26.\n\n"
          "The page reads Total CBN 0.36 (U 0.02). The register held N.D., which "
          "states the laboratory found nothing — it did not. Confirmed at 2.6x "
          "magnification."),
}


def _cur(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return "" if v is None else str(v).strip()


def main(src, dst):
    wb = load_workbook(src)
    ws = wb[SHEET]
    log, skipped = [], []

    for row, col, old, new, cert, why in CELLS:
        letter = ws.cell(row=row, column=col).column_letter
        cur = _cur(ws, row, col)
        if cur == new:
            skipped.append(f"r{row} {letter} already {new!r}")
            continue
        if cur != old:
            raise SystemExit(
                f"REFUSING {letter}{row}: expected {old!r}, found {cur!r}. Not the "
                f"revision this was verified against — re-read {cert} first.")
        ws.cell(row=row, column=col).value = new
        log.append(f"{letter}{row:<4} {old!r} -> {new!r}   {cert} — {why}")

    for row, note in NOTES.items():
        cell = ws.cell(row=row, column=CBN)
        if cell.comment is None:
            cell.comment = Comment(note, AUTHOR)

    wb.save(dst)
    print(f"in : {src}\nout: {dst}\n")
    for line in log:
        print("  CHANGED  " + line)
    for s in skipped:
        print("  skipped  " + s)
    print(f"\n{len(log)} change(s), {len(skipped)} already applied.")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 3:
        raise SystemExit(__doc__)
    sys.exit(main(sys.argv[1], sys.argv[2]))
