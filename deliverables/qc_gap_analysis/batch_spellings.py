#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Every spelling of every batch, across every source we hold.

The same batch is written differently in the register, on the certificate and in
the scan filename, and the difference is never meaningful — `GG1024_01`,
`GG1024-01` and `GG 1024_01` are one batch. `ingestion/common/batch_id.py` states
the rule; this script measures how often documents actually depart from it.

That matters because it is the failure mode of any pipeline keyed on the string
rather than the batch: one batch silently becomes three records, its certificates
scattered across them, and nothing looks wrong. The output names the batches where
that would happen.

    python3 deliverables/qc_gap_analysis/batch_spellings.py
"""
import collections, csv, io, os, re, sys

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
sys.path.insert(0, os.path.join(ROOT, "ingestion", "common"))
from batch_id import batch_key                                  # noqa: E402

import openpyxl                                                 # noqa: E402

REGISTER = os.path.join(HERE, "PP_Batch_Release_QC_Register_CORRECTED.xlsx")
SHARED = os.path.join(HERE, "QC_eCoA_Database_ProductCoA_2026-08-22.xlsx")
GAP = os.path.join(HERE, "batch_gap_analysis.csv")
OUT = os.path.join(HERE, "batch_spellings.csv")

# a batch code is a short alphabetic stem followed by at least four digits
LOOKS_LIKE_BATCH = re.compile(r"^[A-Z]{1,4}\d{4,9}")


def collect():
    seen = collections.defaultdict(lambda: collections.defaultdict(set))

    def add(source, raw):
        key = batch_key(raw)
        if key and LOOKS_LIKE_BATCH.match(key):
            seen[key][source].add(str(raw).strip().replace("\n", " "))

    for row in csv.DictReader(io.open(GAP, encoding="utf-8")):
        add("gap analysis", row["batch"])

    ws = openpyxl.load_workbook(REGISTER, data_only=True)["Batch Release QC"]
    rows = [[c.value for c in r] for r in ws.iter_rows()]
    for r in rows[5:]:
        if r and r[1] and str(r[1]).strip() not in ("", "/", "None", "-", "—"):
            add("register", str(r[1]).strip())

    wb = openpyxl.load_workbook(SHARED, data_only=True)

    reg = wb["1_Certificate_Register"]
    head = [c.value for c in reg[1]]
    col = next(i for i, h in enumerate(head)
               if h and "Batch / serial on certificate" in str(h))
    printed = re.compile(r"[A-Za-zА-Яа-я]{1,4}\s?\d{4,9}(?:\s?[/_-]\s?\d{1,2})?V?")
    for r in reg.iter_rows(min_row=2, values_only=True):
        if r[0] and r[col]:
            # a cell often reads "Strain name / CODE" — the code is the last match
            found = printed.findall(str(r[col]).replace("\n", " ").strip())
            if found:
                add("certificate", found[-1])

    stem_code = re.compile(r"\b[A-Z]{2,4}\s?\d{4,7}(?:[/_-]\d{1,2})?\b")
    for r in wb["6_Source_File_Index"].iter_rows(min_row=2, values_only=True):
        if not r[0]:
            continue
        stem = os.path.basename(str(r[0])).rsplit(".", 1)[0]
        stem = re.sub(r"\s*\(\d\)$|_CoA|_bulk", "", stem)
        for c in stem_code.findall(stem.upper()):
            add("scan filename", c)

    return seen


def main():
    seen = collect()
    out = []
    for key in sorted(seen):
        spellings = sorted({s for d in seen[key].values() for s in d})
        out.append({
            "batch_key": key,
            "spellings": len(spellings),
            "as_written": " | ".join(spellings),
            "sources": " | ".join(sorted(seen[key])),
        })
    with io.open(OUT, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["batch_key", "spellings", "as_written", "sources"])
        w.writeheader()
        w.writerows(out)

    multi = [r for r in out if r["spellings"] > 1]
    worst = max((r["spellings"] for r in out), default=0)
    print(f"batch keys observed          : {len(out)}")
    print(f"recorded more than one way   : {len(multi)}")
    print(f"most spellings of one batch  : {worst}")
    print(f"written to                   : {os.path.relpath(OUT, ROOT)}")
    return len(out), len(multi), worst


if __name__ == "__main__":
    main()
