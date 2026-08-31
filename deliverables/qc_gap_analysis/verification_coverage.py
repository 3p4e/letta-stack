#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""How much of the register has actually been checked against a document.

    python3 deliverables/qc_gap_analysis/verification_coverage.py REGISTER.xlsx

Every previous statement of coverage in this repository was an estimate carried
forward by addition — 146, then 403, then 571, then 797. This counts it instead, by
folding each row's CoA code and looking for it among the certificates whose pages were
actually read on 30-31.08.2026, recorded in `review/*_page_reads_*.json`.

The unit is the **populated result cell**, not the row: a row whose certificate was read
had every value on that certificate compared, because the readings are of whole pages.
A cell holding `/` or `n/a` is not a result and is not counted either way — the register
uses `/` for "parameter not covered by that certificate".

The output is deliberately a list, not a number. What is left unverified is only useful
if you can see which rows it is.
"""
import glob
import json
import re
import sys
from collections import Counter

from openpyxl import load_workbook

SHEET = "Batch Release QC"
CODE, PDF = 23, 26
FIRST_DATA = 6
FIRST_RESULT, LAST_RESULT = 5, 22        # columns E..V, the result columns
NOT_A_RESULT = {"", "/", "n/a", "na", "-", "—"}

_CYR = "АВЕКМНОРСТУХЈЅІ"
_LAT = "ABEKMHOPCTYXJSI"


def fold(s):
    """Same comparison key as repair_register_pdf_links.py: homoglyphs onto Latin,
    a trailing bracketed note dropped, then everything but A-Z0-9 removed."""
    u = re.sub(r"\s*\([^)]*\)\s*$", "", str(s or "")).upper()
    u = re.sub(r"(?<=[^A-Za-z0-9])(GS|ГС)(?=[^A-Za-z0-9]|$)", "LOD", u)
    for a, b in zip(_CYR, _LAT):
        u = u.replace(a, b)
    return re.sub(r"[^A-Z0-9]", "", u)


def candidates(cell):
    """Every certificate this code cell might be naming, folded.

    Six rows name two documents — `PP CoA #027 / ППК25370` — because the in-house CoA
    and the report its numbers came from are both recorded there. Folding the whole cell
    produces a string that matches no certificate, so those rows counted as unverified
    while their values had in fact been read. Splitting on `/` is not an option: half the
    certificate codes in this corpus contain one (`627/1128/25`). So the whole cell is
    tried first, and then any control-book number embedded in it.
    """
    out = {fold(cell)}
    out.update(fold(m) for m in re.findall(r"ППК\s*\d+", str(cell)))
    return {c for c in out if c}


def read_codes():
    """Every certificate whose page was read, folded, with the block it came from."""
    seen = {}
    for path in sorted(glob.glob("review/*_page_reads_*.json")):
        block = path.split("/")[-1].split("_page_reads")[0]
        for code in json.load(open(path, encoding="utf-8")):
            seen[fold(code)] = block
    return seen


def main(src):
    read = read_codes()
    wb = load_workbook(src)
    ws = wb[SHEET]
    hdr = {c: ws.cell(row=5, column=c).value for c in range(FIRST_RESULT, LAST_RESULT + 1)}

    done = todo = 0
    by_block = Counter()
    unread = []
    for row in ws.iter_rows(min_row=FIRST_DATA):
        r = row[0].row
        code = ws.cell(row=r, column=CODE).value
        if not code or str(code).strip().lower() in NOT_A_RESULT:
            continue
        vals = sum(1 for c in range(FIRST_RESULT, LAST_RESULT + 1)
                   if str(ws.cell(row=r, column=c).value or "").strip().lower()
                   not in NOT_A_RESULT)
        block = next((read[c] for c in candidates(code) if c in read), None)
        if block:
            done += vals
            by_block[block] += vals
        else:
            todo += vals
            if vals:
                unread.append((r, str(code)[:30], vals,
                               str(ws.cell(row=r, column=25).value or "")[:34]))

    tot = done + todo
    print(f"register: {src}\n")
    print(f"  populated result cells       : {tot}")
    print(f"  on a page-verified certificate: {done}  ({100*done/tot:.1f}%)")
    print(f"  never checked against a page : {todo}  ({100*todo/tot:.1f}%)\n")
    print("  by block:")
    for b, n in by_block.most_common():
        print(f"     {b:<14} {n:>5}")

    print(f"\n  rows still unverified ({len(unread)} carrying results):")
    inst = Counter(u[3] for u in unread)
    for i, n in inst.most_common():
        print(f"     {n:>4} cells over {sum(1 for u in unread if u[3]==i):>3} rows   {i or '(no institution)'}")
    print()
    for r, code, n, i in unread[:60]:
        print(f"     r{r:<4} {code:<30} {n:>2} values   {i}")
    if len(unread) > 60:
        print(f"     … and {len(unread)-60} more")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 2:
        raise SystemExit(__doc__)
    sys.exit(main(sys.argv[1]))
