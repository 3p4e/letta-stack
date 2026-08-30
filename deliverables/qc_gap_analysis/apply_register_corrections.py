#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Apply the verified corrections to the Batch Release QC register.

Every change here is backed by a rendered certificate page, cited in CORRECTIONS
below and in review/ECOA_REGISTER_VERIFICATION_2026-08-29.md. Nothing is changed
on inference: where the page was ambiguous or the two sources disagreed for a
reason that could not be settled from the paper, the row is left alone and
raised as a question instead.

Why the pages and not the text: the mould counts are superscripts, and neither
machine-readable path survives them. Drive's text extraction drops the exponent
entirely ("10 CFU/g"), and the RAGFlow parse renders it but gets it wrong. The
register agrees with the parse and disagrees with the paper, which is how a
result over the Ph. Eur. limit came to be recorded as one under it.

    python3 deliverables/qc_gap_analysis/apply_register_corrections.py IN.xlsx OUT.xlsx

Idempotent: re-running on an already-corrected file changes nothing and says so.
"""
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

SHEET = "Batch Release QC"
TYMC = 11                      # column K
CODE = 23                      # column W — CoA code

# The sheet's own legend: "Amber — a laboratory finding or a data-integrity flag
# on the certificate". A count above the printed limit that the laboratory still
# passed is exactly that.
AMBER_FILL = PatternFill("solid", fgColor="FFFAEDD4")
AMBER_FONT = Font(color="FF9A6300")

# row, current value, corrected value, certificate, what the page shows
CORRECTIONS = [
    (8,   "1×10³",   "1×10⁴",   "163/0271/25  BG1024",
     "page reads 1 x 10⁴; at the 10⁴ limit, so the disposition does not change"),
    (21,  "4.2×10³", "4.2×10⁴", "320/0587/25  GG1024_01",
     "page reads 4,2 x 10⁴ — 4.2x the 10⁴ limit"),
    (56,  "1.2×10³", "1.2×10⁴", "628/1129/25  GP0824_03",
     "page reads 1,2 x 10⁴ — over the 10⁴ limit"),
    (71,  "3.3×10³", "3.3×10⁴", "904/1589/25  OPM052501",
     "page reads 3,3 x 10⁴ — 3.3x the 10⁴ limit"),
    (82,  "3.6×10³", "3.6×10⁴", "946/1684/25  GP052501",
     "page reads 3,6 x 10⁴ — 3.6x the 10⁴ limit"),
    (100, "4.9×10^3", "4.9×10^4", "1032/1851/25  CJ062501-2",
     "page reads 4,9 x 10⁴ — 4.9x the 10⁴ limit, the largest exceedance found"),
]

# An omission rather than an error: the row carries results but no code. The
# certificate is identified beyond doubt — GP082501-2 is P050322, the row's own
# TAMC 3.2×10^3 and TYMC 5.8×10^3 match the page exactly, and the institution
# and month agree.
CODE_FILLS = [
    (141, "Microbiology report (Dec 2025)", "1227/2193/25",
     "P050322, 1227-2193-25, 05.12.2025, IJZ-MB.pdf"),
]

# Every row whose TYMC exceeds its printed limit, after the corrections above.
# Only two of the ten carried the legend's amber today, which is part of why
# this went unnoticed; flagging them consistently is what makes them visible.
FLAG_ROWS = [21, 35, 38, 56, 71, 74, 82, 87, 100, 121]


def main(src, dst, flag=True):
    wb = load_workbook(src)
    ws = wb[SHEET]
    log, skipped = [], []

    for row, old, new, cert, why in CORRECTIONS:
        cell = ws.cell(row=row, column=TYMC)
        cur = "" if cell.value is None else str(cell.value).strip()
        if cur == new:
            skipped.append(f"r{row} already {new}")
            continue
        if cur != old:
            raise SystemExit(
                f"REFUSING r{row}: expected {old!r}, found {cur!r}. The sheet is "
                f"not the revision this was verified against — re-verify first.")
        cell.value = new
        log.append(f"K{row}  {old:>9} -> {new:<9}  {cert:<26} {why}")

    for row, old, new, cert in CODE_FILLS:
        cell = ws.cell(row=row, column=CODE)
        cur = "" if cell.value is None else str(cell.value).strip()
        if cur == new:
            skipped.append(f"r{row} code already {new}"); continue
        if cur != old:
            raise SystemExit(f"REFUSING r{row}: expected {old!r}, found {cur!r}")
        cell.value = new
        log.append(f"W{row}  {old!r} -> {new!r}  identified as {cert}")

    if flag:
        for row in FLAG_ROWS:
            c = ws.cell(row=row, column=TYMC)
            if c.fill.fgColor and c.fill.fgColor.rgb == "FFFAEDD4":
                continue
            c.fill, c.font = AMBER_FILL, AMBER_FONT
            log.append(f"K{row}  amber flag applied (legend: laboratory finding)")

    wb.save(dst)
    print(f"in : {src}\nout: {dst}\n")
    for line in log:
        print("  CHANGED ", line)
    for s in skipped:
        print("  skipped ", s)
    print(f"\n{len(log)} change(s), {len(skipped)} already correct.")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        raise SystemExit(__doc__)
    main(sys.argv[1], sys.argv[2], flag="--no-flag" not in sys.argv)
