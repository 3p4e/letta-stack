#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Rebuild the corrected register from the owner's original and check it lands.

    python3 deliverables/qc_gap_analysis/replay_correction_chain.py [WORKDIR]

Eleven scripts corrected this register, each written to refuse a workbook whose cells
differ from what it was verified against. That guard proves a script cannot be applied to
the wrong revision. It does not prove the *chain* is intact — that the committed workbooks
really are the outputs of those scripts in that order, with nothing edited by hand along
the way.

This proves it, by replaying every step from `..._CORRECTED.xlsx`, the file the owner
supplied on 18.08.2026, and comparing each result against the workbook committed for that
step. For a release register the reproducibility is the point: a correction nobody can
re-derive is an assertion, not a record.

**The order matters and was not what the file listing suggested.** Reading the directory
alphabetically, or by date, gives `MICRO → LINKS → LINKS2`. The true order is
`LINKS → MICRO → LINKS2`, and the register itself proves it: the committed `LINKS`
workbook does not carry the ten microbiology corrections, and `LINKS2` carries them plus a
different link on row 26 — because correcting that row's code from `319/0586/25` to
`318/0585/25` changed which certificate it resolves to. The links moved because the code
moved. A chain documented from a file listing would have got this backwards, and this
script is what caught it.

## Two divergences are expected, and both are improvements landing early

The replay uses **today's** scripts against **historical** intermediate outputs, so any
improvement made after an intermediate was committed shows up as a difference at that step
and disappears once the step that introduced it is reached.

1. **58 Farmahem links at steps 3-4.** The committed `LINKS` predates the `ГС` → `LoD`
   fold fix in `repair_register_pdf_links.py`. Today's script resolves those rows on the
   first pass; historically they needed the second. `LINKS2` matches, which is the fix
   working.
2. **`Z109` from step 3 onward.** The block-header inheritance added on 31.08 resolves the
   phytosanitary report's link as soon as the first link pass runs. Historically it stayed
   dead until the final pass.

Neither changes a value. **Every value at every step matches**, and the final workbook
matches exactly on both values and links — which is the assertion this script exists to
make.
"""
import os
import subprocess
import sys
import tempfile

from openpyxl import load_workbook

D = "deliverables/qc_gap_analysis"
MAP = f"{D}/drive_certificate_ids_2026-08-31.json"
SHEET = "Batch Release QC"


def W(name):
    return f"{D}/PP_Batch_Release_QC_Register_{name}.xlsx"


# script, committed output to compare against (None = intermediate, not committed), extra args
CHAIN = [
    ("apply_register_corrections.py",     "CORRECTED_2026-08-30", []),
    ("add_ppk25139_and_codes.py",         "CORRECTED_2026-08-31", []),
    ("repair_register_pdf_links.py",      "LINKS_2026-08-31",     [MAP]),
    ("apply_microbiology_corrections.py", "MICRO_2026-08-31",     []),
    ("repair_register_pdf_links.py",      "LINKS2_2026-08-31",    [MAP]),
    ("apply_iph_corrections.py",          "IPH_2026-08-31",       []),
    ("apply_farmahem_corrections.py",     "FHM_2026-08-31",       []),
    ("apply_date_corrections.py",         "DATES_2026-08-31",     []),
    ("apply_cnp_corrections.py",          "CNP_2026-08-31",       []),
    ("apply_residual_corrections.py",     "FINAL_2026-08-31",     []),
    ("repair_ppcoa_pdf_links.py",         None,                   [MAP]),
    ("repair_register_pdf_links.py",      "LINKED_2026-08-31",    [MAP]),
    ("apply_acceptance_criterion_corrections.py", "AC_2026-08-31",  []),
]

FINAL = "AC_2026-08-31"


def cells(path):
    ws = load_workbook(path)[SHEET]
    values = {c.coordinate: str(c.value)
              for row in ws.iter_rows() for c in row if c.value is not None}
    links = {c.coordinate: c.hyperlink.target
             for row in ws.iter_rows() for c in row if c.hyperlink}
    return values, links


def diff(expected, got):
    ev, eh = cells(expected)
    gv, gh = cells(got)
    return ({k for k in set(ev) | set(gv) if ev.get(k) != gv.get(k)},
            {k for k in set(eh) | set(gh) if eh.get(k) != gh.get(k)})


def main(workdir):
    cur = W("CORRECTED")
    if not os.path.exists(cur):
        raise SystemExit(f"missing the owner's original: {cur}")

    value_drift, final_ok = 0, False
    print(f"replaying from {cur}\n")
    for i, (script, expect, extra) in enumerate(CHAIN, 1):
        dst = os.path.join(workdir, f"step{i:02d}.xlsx")
        p = subprocess.run(["python3", f"{D}/{script}", cur, dst] + extra,
                           capture_output=True, text=True)
        if p.returncode != 0:
            print(f"  {i:>2}. {script:<36} FAILED to run")
            print((p.stdout + p.stderr)[-600:])
            return 1
        if expect is None:
            print(f"  {i:>2}. {script:<36} (intermediate, not committed)")
        else:
            dv, dh = diff(W(expect), dst)
            value_drift += len(dv)
            note = "match" if not dv and not dh else \
                   f"values {len(dv)}, links {len(dh)}"
            print(f"  {i:>2}. {script:<36} -> {expect:<22} {note}")
            if dv:
                print(f"        VALUE DRIFT: {sorted(dv)[:8]}")
            if dh and expect != FINAL:
                print(f"        links (see docstring): {sorted(dh)[:8]}")
            if expect == FINAL:
                final_ok = not dv and not dh
        cur = dst

    print()
    if value_drift:
        print(f"FAIL  {value_drift} value(s) differ from the committed chain — a workbook "
              f"was edited outside these scripts, or the order is wrong.")
        return 1
    print("ok    every value at every step matches the committed workbook")
    if not final_ok:
        print("FAIL  the final workbook does not reproduce exactly")
        return 1
    print(f"ok    {FINAL} reproduces exactly, values and links")
    print("\nThe corrected register is re-derivable from the owner's original.")
    return 0


if __name__ == "__main__":
    if len(sys.argv) > 1:
        sys.exit(main(sys.argv[1]))
    with tempfile.TemporaryDirectory() as tmp:
        sys.exit(main(tmp))
