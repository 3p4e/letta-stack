#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Apply what reading the IPH physico-chemical certificates found.

Evidence is in `review/IPH_PHYSCHEM_PAGE_VERIFICATION_2026-08-31.md`, with the raw
readings in `review/iph_physchem_page_reads_2026-08-31.json`. Every value was taken
from the rendered page at 200 DPI, cropped to keep the `Број: NNNN/YYYY` line so the
image names its own certificate.

    python3 deliverables/qc_gap_analysis/apply_iph_corrections.py IN.xlsx OUT.xlsx

Two changes, both on row 31, and both against `2156/2025`:

**Mercury.** The page prints `жива 0,001 mg/kg(l)`; the register holds `0.011`. An
inserted digit. Confirmed by re-reading the metals block at 2.2x magnification, since
one digit is exactly the kind of difference a first read can invent. Both values are
far below the 0,1 limit, so the batch's disposition does not change — but a release
register that states a number the laboratory did not is wrong whether or not the
number matters.

One change only. An earlier draft of this script also stripped the annotation from
the certificate code, on the reading that `2156/2025 (microbiology sub-report lab-ref
not distinctly captured in OCR text)` was describing the physico-chemical report. It
is not: that annotation sits on **row 32**, a separate microbiology row for the same
batch, and it is honest — it says the microbiology report's own laboratory reference
could not be read. Row 31 carries the clean code. The edit was dropped; the script's
refuse-on-mismatch guard is what surfaced the mistake rather than quietly applying it.

What row 32 does expose belongs in the microbiology work, not here: it holds a full
set of microbiological results whose source document has never been identified.

Idempotent, and it refuses a workbook whose cells differ from what was verified.
"""
import sys

from openpyxl import load_workbook
from openpyxl.comments import Comment

SHEET = "Batch Release QC"
HG = 21                           # column U
AUTHOR = "QC page verification 31.08.2026"

ROW = 31
HG_OLD, HG_NEW = "0.011", "0.001"
NOTE = (
    "Corrected 31.08.2026 from the rendered page of 2156/2025.\n\n"
    "The certificate prints жива 0,001 mg/kg(l). The register held 0.011 — an "
    "inserted digit. Re-read at 2.2x magnification to be sure of the single "
    "digit.\n\n"
    "MaxDK on this certificate is 0,1, so both the old and the new value comply "
    "and the batch's disposition is unchanged."
)


def _cur(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return "" if v is None else str(v).strip()


def main(src, dst):
    wb = load_workbook(src)
    ws = wb[SHEET]
    log, skipped = [], []

    cur = _cur(ws, ROW, HG)
    if cur == HG_NEW:
        skipped.append(f"r{ROW} Hg already {HG_NEW}")
    elif cur != HG_OLD:
        raise SystemExit(
            f"REFUSING U{ROW}: expected {HG_OLD!r}, found {cur!r}. Not the revision "
            f"this was verified against — re-read the page first.")
    else:
        ws.cell(row=ROW, column=HG).value = HG_NEW
        ws.cell(row=ROW, column=HG).comment = Comment(NOTE, AUTHOR)
        log.append(f"U{ROW}  {HG_OLD!r} -> {HG_NEW!r}   2156/2025 prints жива 0,001")

    wb.save(dst)
    print(f"in : {src}\nout: {dst}\n")
    for line in log:
        print("  CHANGED  " + line)
    for s in skipped:
        print("  skipped  " + s)
    print(f"\n{len(log)} change(s), {len(skipped)} already applied.")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 3:
        raise SystemExit(__doc__)
    sys.exit(main(sys.argv[1], sys.argv[2]))
