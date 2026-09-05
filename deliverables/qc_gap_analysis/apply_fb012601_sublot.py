#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Chain step 19 — the batch that lost its sub-lot index.

    python3 deliverables/qc_gap_analysis/apply_fb012601_sublot.py IN.xlsx OUT.xlsx

Register ref 53 reads `FB012601`. Every document says `FB012601/1`:

- certificate ППК26067 (UKIM CNP, 11.05.2026) prints the batch as `FB012601/1` —
  read off its own page on 31.08.2026, `review/cnp_page_reads_2026-08-31.json`,
  key `ППК26067`, field `batch_page`; an earlier independent transcription
  (`deliverables/qc_register/extracted_params.json`, `batch_printed`) read the same;
- the IPH microbiology report 308/0552/26 (28.04.2026) prints `Серија: FB012601/1`
  in its own sample line; the IPH chemical-safety report 2362/2026 (30.04.2026) is
  filed under the same `FB012601_1`;
- the potency master (`potency_master_batch_map.json`) pairs `FB012601/1` with
  packaged lot P060322.

Under the batch-identity rule the index nests and carries meaning: `FB012601/1` is
a record distinct from `FB012601`, and a bare `FB012601` cannot be folded onto it
the way `GRC102501-2` folds onto `GRC102501/2`. The consequence in the shipped
deliverables was concrete — the register's `FB012601` and the plan's `FB012601/1`
never joined, so the CoQ schedule invented a second, predicted CoQ pair for a batch
that already had one (`CoQ Issuance` rows 74 and 157), and the two IPH certificates
filed under `FB012601_1` had no register row to attach to.

The question of fact was settled on 22.08.2026 (`CROSS_CHECK_2026-08-22.md`) and
carried as open decision **B5**. The owner ruled on 01.09.2026: the digit is part
of the batch number — `x…x_1`. This step applies that ruling.

Two cells move, both from `FB012601` to `FB012601/1`: `Batch Release QC!B225`
(ref 53) and `THC by Strain!B34`, the same batch in the per-strain sheet. The code,
date, laboratory and the three CNP values on the row are untouched. **The two IPH
certificates are deliberately not added here**: a certificate enters the release
register as a QC act performed against the document, not as a rectification; they
are listed, with 23 others in the same position, in
`review/TRACKER_TRUTH_CHECK_2026-09-01.md`.

Refuses to run unless the row is the one it was written against (code ППК26067,
date 11.05.2026). A second run on its own output reports both cells as already
applied.
"""
import sys

from openpyxl import load_workbook
from openpyxl.comments import Comment

SHEET, ROW, COL = "Batch Release QC", 225, 2
STRAIN_SHEET, STRAIN_ROW = "THC by Strain", 34
OLD, NEW = "FB012601", "FB012601/1"
SENTINEL = "01.09.2026, owner's ruling B5"

NOTE = (
    "Corrected 01.09.2026, owner's ruling B5. The cell read 'FB012601'; the batch "
    "is the sub-lot 'FB012601/1'. Certificate ППК26067 prints 'FB012601/1' on two "
    "independent readings (review/cnp_page_reads_2026-08-31.json key 'ППК26067', "
    "field batch_page; deliverables/qc_register/extracted_params.json, "
    "batch_printed). The IPH microbiology report 308/0552/26 prints "
    "'Серија: FB012601/1'; the potency master pairs FB012601/1 with lot P060322.\n\n"
    "Under the batch-identity rule the index nests and carries meaning, so the "
    "bare row could not join the plan's FB012601/1: the CoQ schedule was inventing "
    "a second CoQ pair for this batch, and the two IPH certificates filed under "
    "FB012601_1 (2362/2026, 308/0552/26) had no row to attach to.\n\n"
    "Those two certificates are NOT added here — a certificate enters the release "
    "register as a QC act against the document. See "
    "review/TRACKER_TRUTH_CHECK_2026-09-01.md."
)


def fix(ws, row, col, what):
    c = ws.cell(row=row, column=col)
    v = str(c.value or "").strip()
    if v == NEW and c.comment and SENTINEL in c.comment.text:
        print(f"already applied: {what} = {NEW!r}")
        return 1
    if v != OLD:
        print(f"REFUSED: {what} reads {v!r}, expected {OLD!r}")
        return -1
    c.value = NEW
    c.comment = Comment(NOTE, "QC data verification", height=300, width=460)
    print(f"applied: {what} {OLD!r} -> {NEW!r}")
    return 0


def main(inp, outp):
    wb = load_workbook(inp)
    ws = wb[SHEET]
    code = str(ws.cell(row=ROW, column=23).value or "").strip()
    date = str(ws.cell(row=ROW, column=24).value or "").strip()
    if code != "ППК26067" or date != "11.05.2026":
        print(f"REFUSED: row {ROW} reads code {code!r} / date {date!r}; "
              f"expected 'ППК26067' / '11.05.2026'")
        return 2
    st = wb[STRAIN_SHEET]
    strain_code = str(st.cell(row=STRAIN_ROW, column=6).value or "").strip()
    if strain_code != "ППК26067":
        print(f"REFUSED: {STRAIN_SHEET}!F{STRAIN_ROW} reads {strain_code!r}, expected 'ППК26067'")
        return 2
    r1 = fix(ws, ROW, COL, f"{SHEET}!B{ROW} (ref 53)")
    r2 = fix(st, STRAIN_ROW, 2, f"{STRAIN_SHEET}!B{STRAIN_ROW}")
    if r1 < 0 or r2 < 0:
        return 2
    wb.save(outp)
    if r1 == 1 and r2 == 1:
        print("both cells already present")
    return 0


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
