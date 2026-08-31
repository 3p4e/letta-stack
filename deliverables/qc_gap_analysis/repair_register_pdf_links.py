#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Point every register row at its own certificate.

The register's PDF column has two independent defects, and the second one would
have been wrong even if the first did not exist.

**Every link is dead.** All 285 links address Drive files that no longer resolve.
Every PDF in the certificate folder carries `createdTime = 2026-08-22`: the folder was
rebuilt that day, Drive issues a new id to a re-uploaded file, and the register still
addresses the originals. The documents are intact and correctly named — only the
addresses are stale.

**And 190 rows never pointed at their own document.** There are 285 links but only 136
distinct ids among them; 62 ids are shared by two or more rows. One link per *batch* was
copied down across all of that batch's certificate rows, so `BSS052501` has six rows —
an in-house CoA, a phytosanitary report, two IPH reports and a Farmahem report — all
addressing one file. Only 74 rows carry an id no other row uses.

That second defect is the more serious of the two. A dead link fails visibly, and
someone eventually notices. A live link to the wrong laboratory's certificate does not:
it opens a real document with real numbers on it, and nothing about it says it belongs
to a different row.

Matching is on the certificate code plus the P-number or batch, both of which the Drive
filename already carries — `P050022_ППК25174, 10.07.2025_CNP.pdf`. Never on the
register's own stale ids. Codes are folded through the same Cyrillic/Latin homoglyph
rule used everywhere else in this repository, because the register writes `197-18-К/26`
with a Cyrillic К where the filename has a Latin `197-18-K-26`.

    python3 deliverables/qc_gap_analysis/repair_register_pdf_links.py IN.xlsx OUT.xlsx MAP.json

MAP.json is `{"<drive filename>": "<live file id>"}`, taken from a listing of the
certificate folder. A row whose certificate is absent from the map is **left exactly as
it is** and reported: a wrong link is not improved by replacing it with a guess.
"""
import json
import re
import sys

from openpyxl import load_workbook

SHEET = "Batch Release QC"
CODE, PDF = 23, 26          # columns W and Z
FIRST_DATA = 6

_CYR = "АВЕКМНОРСТУХЈЅІ"
_LAT = "ABEKMHOPCTYXJSI"


def fold(s):
    """Comparison key: homoglyphs onto Latin, separators dropped, upper case.

    Three conventions differ between the register and the filenames, and all three
    were found the hard way, by a search returning nothing:

    - `197-18-К/26` (Cyrillic К, slash) and `197-18-K-26` (Latin K, dash) are the
      same certificate written two ways, and both appear in this corpus.
    - Farmahem's loss-on-drying reports are `-GS-` / `-ГС-` in the register and
      `-LoD-` in the filename — the same suffix in two languages, not two reports.
    - The register annotates some codes with the sample description in brackets,
      `230/0393/26 (Trimiran cvet)`. The brackets are a note, not part of the code.
    """
    u = re.sub(r"\s*\([^)]*\)\s*$", "", str(s or "")).upper()
    # Before the homoglyph fold, not after: Г has no Latin lookalike and is left
    # alone by it, so `ГС` would arrive here as `ГC` and never match the rule.
    u = re.sub(r"(?<=[^A-Za-z0-9])(GS|ГС)(?=[^A-Za-z0-9]|$)", "LOD", u)
    for a, b in zip(_CYR, _LAT):
        u = u.replace(a, b)
    return re.sub(r"[^A-Z0-9]", "", u)


def build_index(mapping):
    """filename map -> {folded code: [{key, fid, title, lang}, ...]}.

    The filename is `<batch or P-number>_<certificate code>, <date>_<lab>.pdf`, and
    the underscore is not a reliable boundary in either direction: the batch can
    carry one (`GRC102501_2_051-6-LoD-26`) and so can the code
    (`P050192_10802_2845-2 MK`). Splitting at the first underscore loses the former,
    splitting at the last loses the latter.

    So every underscore is treated as a possible boundary and each right-hand side
    is indexed as a candidate code. A register code matches whichever split is the
    real one; the wrong splits index strings no certificate code looks like, and the
    P-number check behind this guards the rest.
    """
    idx = {}
    for title, fid in mapping.items():
        head = title.rsplit(".", 1)[0].split(",")[0]
        parts = head.split("_")
        for i in range(1, len(parts)):
            key, code = "_".join(parts[:i]), "_".join(parts[i:]).strip()
            # A report issued in both languages is one report: `10802_2845-2 MK` and
            # `... EN` are the Macedonian original and its translation, and the
            # register has a single row for them.
            m = re.search(r"\s(MK|EN)$", code)
            lang = m.group(1) if m else ""
            if m:
                code = code[:m.start()]
            idx.setdefault(fold(code), []).append(
                {"key": fold(key), "fid": fid, "title": title, "lang": lang})
    return idx


def main(src, dst, mapfile):
    idx = build_index(json.load(open(mapfile, encoding="utf-8")))
    wb = load_workbook(src)
    ws = wb[SHEET]

    fixed, already, unmatched, ambiguous = [], [], [], []
    for row in ws.iter_rows(min_row=FIRST_DATA):
        r = row[0].row
        code = ws.cell(row=r, column=CODE).value
        if not code:
            continue
        cands = idx.get(fold(code), [])
        if not cands:
            unmatched.append((r, str(code)[:34]))
            continue
        if len(cands) > 1:
            # Same code issued for more than one batch: disambiguate on the row's own
            # P-number, then its batch code.
            want = fold(ws.cell(row=r, column=3).value) or fold(ws.cell(row=r, column=2).value)
            narrowed = [c for c in cands if c["key"] == want]
            # Two files that are one report in two languages: link the Macedonian,
            # which is the issued original — the English is a translation of it.
            if len(narrowed) == 2 and {c["lang"] for c in narrowed} == {"MK", "EN"}:
                narrowed = [c for c in narrowed if c["lang"] == "MK"]
            if len(narrowed) != 1:
                ambiguous.append((r, str(code)[:34], len(cands)))
                continue
            cands = narrowed

        cell = ws.cell(row=r, column=PDF)
        target = f"https://drive.google.com/file/d/{cands[0]['fid']}/view"
        cur = getattr(cell.hyperlink, "target", None)
        if cur == target:
            already.append(r)
            continue
        cell.value, cell.hyperlink = "Open", target
        fixed.append((r, str(code)[:22], cands[0]["title"][:46]))

    wb.save(dst)
    print(f"in : {src}\nout: {dst}\nmap: {mapfile}\n")
    print(f"  repaired            : {len(fixed)}")
    print(f"  already correct     : {len(already)}")
    print(f"  no file in the map  : {len(unmatched)}  (left untouched)")
    print(f"  ambiguous           : {len(ambiguous)}  (left untouched)")
    if fixed:
        print("\n  sample of repairs:")
        for r, code, title in fixed[:10]:
            print(f"     r{r:<4} {code:<22} -> {title}")
    if unmatched:
        print("\n  rows still carrying a dead link, by certificate:")
        for r, code in unmatched[:40]:
            print(f"     r{r:<4} {code}")
        if len(unmatched) > 40:
            print(f"     … and {len(unmatched)-40} more")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 4:
        raise SystemExit(__doc__)
    sys.exit(main(sys.argv[1], sys.argv[2], sys.argv[3]))
