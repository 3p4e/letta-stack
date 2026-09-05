#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Point the in-house CoA rows at their own documents, and fix one impossible date.

    python3 deliverables/qc_gap_analysis/repair_ppcoa_pdf_links.py IN.xlsx OUT.xlsx MAP.json

`audit_register_pdf_links.py` checks the opposite direction from the repair script: it
takes the file id a cell actually links to and asks what document that is. On the
corrected register it reports **225 links opening the right document and none opening the
wrong one** — the serious defect is gone. What it also reports is 21 rows linking to a
file id that is in no map, and every one of those ids is **dead**: `get_file_metadata`
returns "Requested entity was not found" for all 17 distinct ids among them.

Twenty of the 21 are the register's own in-house certificates, and they were never
missing. `repair_register_pdf_links.py` left them alone because it matches on the
certificate code, and here the code does not match by design:

| Register | Drive filename |
|---|---|
| `PP CoA #016` | `P050202_QCCoA 001v02, 04.12.2025_PP.pdf` |
| `PP CoA #028` | `P050212_QCCoA 001v02, 21.01.2026_PP.pdf` |

**The register numbers these documents individually; Drive files them all under one form
number.** `QCCoA 001v02` is the form, not the document — 26 files share it. The
distinguishing key is the P-number, which the filename carries as its prefix and the
register carries in column C. So the repair script's fallback key is the only key here,
and its primary key never had a chance.

Matching on the P-number is not a guess, and four documents were opened to prove it. Each
prints its own **Certificate number** in the head, and each equals the register's code:

| File | Prints | Register row |
|---|---|---|
| `P050202_QCCoA 001v02` | `Certificate number: 016` | r89 `PP CoA #016` |
| `P050212_QCCoA 001v02` | `Certificate number: 028` | r98 `PP CoA #028` |
| `P060052_QCCoA 001v02` | `Certificate number: 037` | r163/164 `PP CoA #037` |
| `P060092_QCCoA 001v02` | `Certificate number: 033` | r182/183 `PP CoA #033` |

The document's own content names the register's code. That is a stronger check than the
filename match used everywhere else in this repository.

## The date, and the two that cannot be settled

Fifteen of the nineteen P-numbered in-house rows carry a date identical to their file's.
Four do not, and only one of the four can be corrected.

**Row 182 holds `20.02.2020 [see source]`.** The year is impossible: the certificate
gives `Manufacturing date: December 2025` and `Retest date: July 2026`, so a CoA cannot
have issued in February 2020. The filename says 20.02.2026, and all seven sibling
`P0600xx` rows say 20.02.2026. It also carries the same `[see source]` scaffolding that
`apply_date_corrections.py` stripped from row 185. Corrected.

**Rows 89 and 98 cannot be settled, because these documents print no issue date at all.**
The QCCoA form has a certificate number, a strain, a batch, a manufacturing month and a
retest month — and no date of issue. Row 89 holds 04.10.2025 against a filename saying
04.12.2025; row 98 holds 24.01.2026 against 21.01.2026. The filename is an independent
transcription and it is the only outside evidence there is; it was good enough to
*corroborate* 215 dates in the earlier pass, and good enough to *flag* these two, but a
filename alone is not a document. Both are reported and left. Correcting a release
register from a filename when the document is silent would be exactly the substitution
this campaign has spent four passes catching other people making.

**Row 163 is a different question again.** See below.

## A correction to what this campaign said yesterday

`review/RESIDUAL_PAGE_VERIFICATION_2026-08-31.md` states that the six `PP CoA #nnn /
ППКnnnnn` rows each sit above a bare `ППКnnnnn` row and that "the pattern holds on all six
pairs". **It holds on five.** Block 37 has three rows, not two:

| Row | Code | Date |
|---|---|---|
| 163 | `PP CoA #037 / ППК26005` | 21.01.2026 |
| 164 | `PP CoA #037` | 20.02.2026 |
| 165 | `ППК26005` | 21.01.2026 |

Both documents already have a row of their own, so the header's combined code cell is
redundant — and its date, 21.01.2026, is `ППК26005`'s, while `PP CoA #037`'s own row
correctly holds the 20.02.2026 that the file confirms. Nothing here is a wrong value; row
163 duplicates a document that row 165 already covers. Which of the two should carry the
batch's results is a structural question for QC, and it is left alone.

Idempotent, and it refuses a workbook whose cells differ from what was verified.
"""
import json
import re
import sys

from openpyxl import load_workbook
from openpyxl.comments import Comment

SHEET = "Batch Release QC"
PNUM, DATE, CODE, PDF = 3, 24, 23, 26
FIRST_DATA = 6

# Continuation rows carry no P-number of their own: they are further rows of the batch
# block above, naming the same in-house certificate. The block header's P-number is the
# right key for them — same batch, same document, same file.
INHERIT = {164: 163, 183: 182}

DATE_ROW = 182
DATE_OLD, DATE_NEW = "20.02.2020 [see source]", "20.02.2026"
DATE_NOTE = (
    "Corrected 31.08.2026 from P060092_QCCoA 001v02.\n\n"
    "The cell held \"20.02.2020 [see source]\". February 2020 is impossible for this "
    "document: the certificate prints Manufacturing date December 2025 and Retest date "
    "July 2026, and its own head reads Certificate number 033, Batch No P060092.\n\n"
    "20.02.2026 is corroborated three ways: the Drive filename, all seven sibling "
    "P0600xx rows, and the manufacturing and retest months on the page. The bracketed "
    "note was scaffolding, as on row 185.\n\n"
    "Note that this form prints no date of issue anywhere, so 20.02.2026 rests on those "
    "three corroborations rather than on a date read off the page."
)

UNRESOLVED = {
    89: ("04.10.2025", "04.12.2025"),
    98: ("24.01.2026", "21.01.2026"),
}
UNRESOLVED_NOTE = (
    "Flagged 31.08.2026, not corrected.\n\n"
    "The Drive filename for this row's certificate carries {file}; this cell holds "
    "{reg}. The filename is an independent transcription — nobody copied it from this "
    "register — and it corroborated 215 dates in the 31.08 date pass.\n\n"
    "It cannot settle this one. The QCCoA 001v02 form prints a certificate number, a "
    "strain, a batch, a manufacturing month and a retest month, and NO date of issue. "
    "The page was rendered and checked; there is nothing there to read.\n\n"
    "So the disagreement is real and unresolvable from the document. Correcting a "
    "release register from a filename while the document itself is silent would be the "
    "same substitution this verification campaign exists to catch. Resolving it needs "
    "the issuing record, not the file listing."
)


def main(src, dst, mapfile):
    by_pnum = {}
    for title, fid in json.load(open(mapfile, encoding="utf-8")).items():
        if "QCCoA" not in title:
            continue
        by_pnum[title.split("_", 1)[0].strip().upper()] = (fid, title)

    wb = load_workbook(src)
    ws = wb[SHEET]
    fixed, already, unmatched = [], [], []

    for row in ws.iter_rows(min_row=FIRST_DATA):
        r = row[0].row
        code = str(ws.cell(row=r, column=CODE).value or "")
        if not code.startswith("PP CoA"):
            continue
        src_row = INHERIT.get(r, r)
        pn = str(ws.cell(row=src_row, column=PNUM).value or "").strip().upper()
        hit = by_pnum.get(pn)
        if not hit:
            unmatched.append((r, code[:30], pn or "(no P-number)"))
            continue
        fid, title = hit
        cell = ws.cell(row=r, column=PDF)
        target = f"https://drive.google.com/file/d/{fid}/view"
        if getattr(cell.hyperlink, "target", None) == target:
            already.append(r)
            continue
        cell.value, cell.hyperlink = "Open", target
        fixed.append((r, code[:24], title[:44]))

    cur = str(ws.cell(row=DATE_ROW, column=DATE).value or "").strip()
    if cur == DATE_NEW:
        date_done = "already"
    elif cur != DATE_OLD:
        raise SystemExit(
            f"REFUSING X{DATE_ROW}: expected {DATE_OLD!r}, found {cur!r}.")
    else:
        ws.cell(row=DATE_ROW, column=DATE).value = DATE_NEW
        ws.cell(row=DATE_ROW, column=DATE).comment = Comment(DATE_NOTE, AUTHOR := "QC page verification 31.08.2026")
        date_done = "corrected"

    for r, (reg, filed) in UNRESOLVED.items():
        cell = ws.cell(row=r, column=DATE)
        if str(cell.value or "").strip() != reg:
            raise SystemExit(f"REFUSING X{r}: expected {reg!r}, found {cell.value!r}.")
        if cell.comment is None:
            cell.comment = Comment(
                UNRESOLVED_NOTE.format(reg=reg, file=filed),
                "QC page verification 31.08.2026")

    wb.save(dst)
    print(f"in : {src}\nout: {dst}\nmap: {mapfile}\n")
    for r, code, title in fixed:
        print(f"  RELINKED  r{r:<4} {code:<24} -> {title}")
    if already:
        print(f"  {len(already)} row(s) already linked correctly")
    for r, code, pn in unmatched:
        print(f"  UNMATCHED r{r:<4} {code:<30} {pn}")
    print(f"\n{len(fixed)} link(s) repaired.")
    print(f"X{DATE_ROW} date {date_done}: {DATE_OLD!r} -> {DATE_NEW!r}")
    print(f"{len(UNRESOLVED)} date(s) flagged as unresolvable from the document: "
          f"{sorted(UNRESOLVED)}")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 4:
        raise SystemExit(__doc__)
    sys.exit(main(sys.argv[1], sys.argv[2], sys.argv[3]))
