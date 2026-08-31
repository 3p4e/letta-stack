#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Chain step 18 — an impossible microbiology range, and a certificate with no number.

    python3 deliverables/qc_gap_analysis/apply_gnb_range_and_number.py IN.xlsx OUT.xlsx

Both defects predate the correction chain and were found by the after-intervention
review of 31.08.2026, which compared every page-read value against the register
column by column — a comparison the earlier Farmahem-only cross-check could not
make.

**The impossible range.** Row 57 (`628/1129/25`, IPH, GP0824_03) carries
bile-tolerant gram-negative bacteria as

    <10¹ and >10³

which asserts a count below 10 and above 1 000 at the same time. No count can be
both. The page reads `< 10^4 и > 10^3` — a plate-count bracket between 1 000 and
10 000, which is what the neighbouring TAMC 1.1×10⁴ and TYMC 1.2×10⁴ on the same
certificate lead you to expect. A single superscript, ¹ for ⁴, transcribed wrong.
The corrected cell keeps the register's own English form: `<10⁴ and >10³`.

Nothing about the batch's disposition changes: the criterion for this column is
`≤ 10⁴ CFU/g` (max acceptable 20 000 under Ph. Eur. 5.1.4), and a bracket whose
upper bound is 10⁴ conforms — as the certificate itself concludes (ОДГОВАРА).
What changes is that the register stops printing an impossibility.

**The certificate with no number.** Row 239 is the continuation row of the
`SCR112501` block and its code cell reads `(not numbered)`. The certificate is not
missing: `305/0549/26` was read off its own page on 31.08.2026 and is recorded in
`review/microbiology_page_reads_2026-08-31.json` against batch `SCR112501` — TAMC
2,3×10³, TYMC 4,3×10³, BTGNB < 10³ и > 10², Salmonella and E. coli absent,
verdict СЕ ВО СОГЛАСНОСТ. Only the number was never written into the register, so
the receipt register — which drops `(not numbered)` rows — could not carry the
document at all.

This step writes the **number and the issuing laboratory only**. The number is
transcribed from the page read; the laboratory follows from it — `nnn/nnnn/nn` is
the IPH microbiology series, and the page read sits in the microbiology campaign
whose other forty records are all IPH. **The five result values are deliberately
not written**: transcribing results into the release register is a QC act on the
physical certificate, not a rectification, and the artifact already shows them as
page-read provenance beside the receipt. The date of issue stays blank for the
same reason — the page read did not capture it.

One count moves downstream: the receipt register goes from 247 to 248 documents,
232 to 233 page-verified.

Refuses to run if either cell differs from the value it was verified against, so
it cannot corrupt a different revision. A second run on its own output reports
both changes as already applied.
"""
import sys

from openpyxl import load_workbook
from openpyxl.comments import Comment

SHEET = "Batch Release QC"
GNB_ROW, GNB_COL = 57, 12          # L57 — bile-tolerant gram-negative bacteria
NUM_ROW, CODE_COL, LAB_COL = 239, 23, 25

GNB_OLD, GNB_NEW = "<10¹ and >10³", "<10⁴ and >10³"
NUM_OLD, NUM_NEW = "(not numbered)", "305/0549/26"
LAB_NEW = "IPH — Institute of Public Health"

SENTINEL = "31.08.2026, after-intervention review"

GNB_NOTE = (
    "Corrected 31.08.2026, after-intervention review. The cell read '<10¹ and "
    ">10³' — below 10 and above 1 000 at once, which no count can be. The page "
    "reads '< 10^4 и > 10^3': a bracket between 1 000 and 10 000, consistent with "
    "TAMC 1,1×10⁴ and TYMC 1,2×10⁴ on the same certificate. One superscript "
    "transcribed wrong, ¹ for ⁴.\n\n"
    "Source: review/microbiology_page_reads_2026-08-31.json, key '628-1129-25'.\n\n"
    "No disposition changes: the column criterion is ≤ 10⁴ CFU/g (maximum "
    "acceptable count 20 000, Ph. Eur. 5.1.4) and the certificate concludes "
    "ОДГОВАРА."
)

NUM_NOTE = (
    "Number written 31.08.2026, after-intervention review. The cell read '(not "
    "numbered)'. The certificate exists and was read off its own page: "
    "'305/0549/26', batch SCR112501 — TAMC 2,3×10³, TYMC 4,3×10³, "
    "bile-tolerant GNB < 10³ и > 10², Salmonella отсустна/25, E. coli "
    "отсустна/g, verdict СЕ ВО СОГЛАСНОСТ.\n\n"
    "Source: review/microbiology_page_reads_2026-08-31.json, key '305-0549-26'. "
    "The laboratory follows from the number: nnn/nnnn/nn is the IPH microbiology "
    "series and every other record in that campaign file is IPH.\n\n"
    "The five results are NOT written here. Transcribing results into the release "
    "register is a QC act performed against the physical certificate, not a "
    "rectification; the receipt register carries them as page-read provenance. The "
    "date of issue stays blank — the page read did not capture it. QC to complete "
    "both from the document."
)


def main(inp, outp):
    wb = load_workbook(inp)
    ws = wb[SHEET]
    gnb = ws.cell(row=GNB_ROW, column=GNB_COL)
    code = ws.cell(row=NUM_ROW, column=CODE_COL)
    lab = ws.cell(row=NUM_ROW, column=LAB_COL)

    # identity guards: the rows must be the ones this step was written against
    w57 = str(ws.cell(row=GNB_ROW, column=23).value or "").strip()
    if w57 != "628/1129/25":
        print(f"REFUSED: W{GNB_ROW} reads {w57!r}, expected '628/1129/25'")
        return 2
    hdr = str(ws.cell(row=238, column=2).value or "").strip()
    if hdr != "SCR112501":
        print(f"REFUSED: B238 reads {hdr!r}, expected 'SCR112501' "
              f"(row {NUM_ROW} must be that block's continuation row)")
        return 2

    done = 0
    if str(gnb.value) == GNB_NEW and gnb.comment and SENTINEL in gnb.comment.text:
        print(f"already applied: L{GNB_ROW} = {GNB_NEW!r}")
        done += 1
    elif str(gnb.value) == GNB_OLD:
        gnb.value = GNB_NEW
        gnb.comment = Comment(GNB_NOTE, "QC data verification", height=250, width=430)
        print(f"applied: L{GNB_ROW} {GNB_OLD!r} -> {GNB_NEW!r}")
    else:
        print(f"REFUSED: L{GNB_ROW} reads {gnb.value!r}, expected {GNB_OLD!r}")
        return 2

    if str(code.value) == NUM_NEW and code.comment and SENTINEL in code.comment.text:
        print(f"already applied: W{NUM_ROW} = {NUM_NEW!r}")
        done += 1
    elif str(code.value) == NUM_OLD:
        code.value = NUM_NEW
        code.comment = Comment(NUM_NOTE, "QC data verification", height=280, width=440)
        lab.value = LAB_NEW
        print(f"applied: W{NUM_ROW} {NUM_OLD!r} -> {NUM_NEW!r}, Y{NUM_ROW} = laboratory")
    else:
        print(f"REFUSED: W{NUM_ROW} reads {code.value!r}, expected {NUM_OLD!r}")
        return 2

    wb.save(outp)
    if done == 2:
        print("both changes already present")
    return 0


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
