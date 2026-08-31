#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Chain step 16 — one transposed pair on 197-7-К/26, and the batch behind P060332.

    python3 deliverables/qc_gap_analysis/apply_farmahem_transposition.py IN.xlsx OUT.xlsx

**The transposition.** `197-7-К/26` (Fat Bastard, cultivation batch FB012602, packaged
lot P060352, Farmahem, 07.08.2026) reports

    Total Δ9-THC   18.86 %w/w  (U 1.16, k=2)
    Total CBD      <LOQ %w/w
    Total CBN       0.22 %w/w  (U 0.01, k=2)

and the register carried CBD `0.22` with CBN `< LOQ` — the two swapped. Total Δ9-THC
was right, which is why no potency cross-check caught it: 18.86 is what the plan, the
register and the certificate all say.

The whole 197 series was swept cell by cell against the certificate transcriptions in
`ingestion/coa_track/letta-imb-coas/exports/master_coa_table.tsv` — 42 certificates,
every Total Δ9-THC, Total CBD, Total CBN, Aflatoxin B₁, aflatoxin sum and Ochratoxin A
they report. **This is the only disagreement in the series**, and it is a transposition,
not a misreading: both values are present and correct, in each other's columns.

Neither value changes a disposition. CBD `< LOQ` and CBN `0.22 %` both conform to
`≤ 1.00 %`; so did the transposed pair. What it would have changed is the reissue CoQ
for P060352, which would have printed 0.22 % against Total CBD and `< LOQ` against
Total CBN — two wrong results on a signed certificate of quality.

**The batch behind P060332.** The register carries `P060332` as a block keyed by its own
packaged-lot number, with no cultivation batch: it appears in neither the release
register's batch column nor the owner's issue plan, and its Cash Cow potency (17.67 %)
matches neither CC112501's 13.35 % nor the plan's CC012603 at 14.76 %. The certificate
table resolves it: **`197-6-К/26` and `197-6-М/26` are issued for cultivation batch
`CC012601/1`**, a Cash Cow lot that appears nowhere else in this repository.

That is recorded in the block's own cells rather than invented into the batch column:
naming a cultivation batch there would create a 81st register batch out of a
certificate transcription. What the register can say truthfully is *which* batch the
laboratory certified, and that its earlier testing is missing — the certificate table
notes for P060332 that no identity, foreign matter, loss on drying, heavy metal,
pesticide or microbiological record exists for it anywhere, only the two Farmahem
panels. Same defect class as GG1024, one lot further on.
"""
import sys

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

SHEET = "Batch Release QC"
AMBER = PatternFill("solid", fgColor="FFFAEDD4")

ROW = 276
NOTE = (
    "Corrected 31.08.2026. Total CBD and Total CBN were transposed against their own "
    "certificate.\n\n"
    "197-7-К/26 (Farmahem, 07.08.2026, cultivation batch FB012602) reports:\n"
    "    Total D9-THC   18.86 %w/w (U 1.16, k=2)\n"
    "    Total CBD      <LOQ %w/w\n"
    "    Total CBN       0.22 %w/w (U 0.01, k=2)\n\n"
    "The register held CBD 0.22 and CBN < LOQ. Total D9-THC was correct, so no potency "
    "cross-check could see it.\n\n"
    "Found by sweeping all 42 certificates of the 197 series cell by cell against "
    "ingestion/coa_track/letta-imb-coas/exports/master_coa_table.tsv. It is the only "
    "disagreement in the series. Neither value changes a disposition — both conform to "
    "<= 1.00 % — but the reissue CoQ for P060352 would have printed both results in the "
    "wrong rows."
)

P060332_ROW = 273
P060332_NOTE = (
    "Cultivation batch resolved 31.08.2026: CC012601/1 (Cash Cow).\n\n"
    "This block is keyed by its packaged-lot number and carries no cultivation batch. "
    "Its 17.67 % matches neither CC112501 (13.35 %, the only Cash Cow cultivation batch "
    "in this register) nor the issue plan's CC012603 (14.76 %) — it is a third Cash Cow "
    "lot. Both its certificates, 197-6-К/26 and 197-6-М/26, are issued for cultivation "
    "batch CC012601/1, which appears nowhere else in this repository.\n\n"
    "The batch column is deliberately NOT filled in: doing so would create an 81st "
    "register batch out of a certificate transcription. QC to confirm CC012601/1 "
    "against the paper and open the block properly.\n\n"
    "OPEN, and the same class as GG1024: the certificate record holds NOTHING ELSE for "
    "this lot — no identity, no foreign matter, no loss on drying, no heavy metals, no "
    "pesticides, no microbiology. Only the two Farmahem panels of August 2026. Its "
    "release testing is either unfiled or was never performed."
)


def main(src, dst):
    wb = load_workbook(src)
    ws = wb[SHEET]
    applied = again = 0

    g, h = str(ws[f"G{ROW}"].value or "").strip(), str(ws[f"H{ROW}"].value or "").strip()
    if (g, h) == ("< LOQ", "0.22"):
        again += 1
        print(f"  = G{ROW}/H{ROW}  already ('< LOQ', '0.22')")
    elif (g, h) == ("0.22", "< LOQ"):
        if str(ws[f"W{ROW}"].value or "").strip() != "197-7-К/26":
            sys.exit(f"REFUSED: row {ROW} does not carry 197-7-К/26")
        ws[f"G{ROW}"], ws[f"H{ROW}"] = "< LOQ", "0.22"
        for c in (f"G{ROW}", f"H{ROW}"):
            ws[c].fill = AMBER
            ws[c].comment = Comment(NOTE, "QC review")
        applied += 1
        print(f"  + G{ROW}/H{ROW}  CBD/CBN transposition corrected -> ('< LOQ', '0.22')")
    else:
        sys.exit(f"REFUSED: row {ROW} reads CBD={g!r} CBN={h!r}, expected the "
                 f"transposed pair ('0.22', '< LOQ') or the corrected one")

    cell = ws[f"B{P060332_ROW}"]
    if str(cell.value or "").strip() != "P060332":
        sys.exit(f"REFUSED: B{P060332_ROW} is not P060332")
    if cell.comment and "CC012601/1" in cell.comment.text:
        again += 1
    else:
        cell.comment = Comment(P060332_NOTE, "QC review")
        cell.fill = AMBER
        applied += 1
        print(f"  + B{P060332_ROW}  cultivation batch recorded as CC012601/1 (comment)")

    wb.save(dst)
    print(f"\n{applied} applied, {again} already in place -> {dst}")
    return 0


if __name__ == "__main__":
    if len(sys.argv) != 3:
        sys.exit(__doc__)
    sys.exit(main(sys.argv[1], sys.argv[2]))
