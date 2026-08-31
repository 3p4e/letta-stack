#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""One JSON for the QC issuance artifact: CoQ schedule, iCoA plan, eCoA receipt.

    python3 deliverables/qc_gap_analysis/export_coq_artifact_data.py OUT.json

Everything is derived from the same computations the workbooks use —
build_coq_schedule.schedule() / icoa_plan() and build_document_registers
load_register() / verified_map() — so the artifact cannot drift from the
deliverables. No value is retyped here.
"""
import json
import os
import sys
from collections import OrderedDict

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
os.chdir(ROOT)
sys.path.insert(0, HERE)

import build_coq_schedule as CQ            # noqa: E402
import build_document_registers as DR      # noqa: E402


# Which fields of each page-reads record are analytical values, and the
# register-facing label each renders under. Everything else on a record
# (batch identity, uncertainties, spec strings, notes) stays in the file.
PR_LABELS = {
    "cnp": [("thc", "Total Δ9-THC %"), ("d9", "Δ9-THC %"), ("thca", "Δ9-THCA %"),
            ("cbd", "Total CBD %"), ("cbd_raw", "CBD %"), ("cbda", "CBDA %"),
            ("cbn", "CBN %"), ("lod", "Loss on drying %"),
            ("foreign_matter", "Foreign matter")],
    "farmahem": [("thc", "Total Δ9-THC %"), ("cbd", "Total CBD %"),
                 ("cbn", "Total CBN %"), ("afla", "Aflatoxins Σ µg/kg"),
                 ("aflab1", "Aflatoxin B1 µg/kg"), ("afla_b2", "Aflatoxin B2 µg/kg"),
                 ("afla_g1", "Aflatoxin G1 µg/kg"), ("afla_g2", "Aflatoxin G2 µg/kg"),
                 ("ota", "Ochratoxin A µg/kg"), ("lod", "Loss on drying %")],
    "iph_physchem": [("pb", "Pb mg/kg"), ("cd", "Cd mg/kg"), ("as", "As mg/kg"),
                     ("hg", "Hg mg/kg"), ("cu", "Cu mg/kg"),
                     ("afla", "Aflatoxins µg/kg"), ("pest", "Pesticides")],
    "microbiology": [("tamc", "TAMC CFU/g"), ("tymc", "TYMC CFU/g"),
                     ("gnb", "Bile-tolerant GNB CFU/g"), ("ecoli", "E. coli /1 g"),
                     ("salm", "Salmonella /25 g"), ("verdict", "Verdict")],
    "residual": [("tamc", "TAMC CFU/g"), ("tymc", "TYMC CFU/g"),
                 ("gnb", "Bile-tolerant GNB CFU/g"), ("ecoli", "E. coli /1 g"),
                 ("salmonella", "Salmonella /25 g"), ("pesticides", "Pesticides"),
                 ("verdict", "Verdict")],
}


def page_rect_map():
    """fold(code) -> the certificate's values as read off its own page (tier T1)."""
    import glob
    import verification_coverage as VC
    out = {}
    for path in sorted(glob.glob("review/*_page_reads_*.json")):
        blk = path.split("/")[-1].split("_page_reads")[0]
        for code, rec in json.load(open(path, encoding="utf-8")).items():
            vals = [{"k": lbl, "v": str(rec[f]), "src": "page-read 31.08.2026 (" + blk + ")",
                     "t": "T1"}
                    for f, lbl in PR_LABELS[blk]
                    if rec.get(f) not in (None, "", {})]
            if vals:
                out[VC.fold(code)] = vals
    return out


def corpus_rect_map():
    """code (exactly as printed) -> corpus-derived values (tiers T2/T3), from
    the rectification sweep's output when it exists. Corpus values never mark
    a certificate page-verified — the remediation desk is the promotion path."""
    path = os.path.join(HERE, "ecoa_rectification_2026-08-31.json")
    if not os.path.exists(path):
        return {}
    data = json.load(open(path, encoding="utf-8"))
    return {code: rec["params"] for code, rec in data.get("certs", {}).items()
            if rec.get("params")}


# Every flagged receipt, with its documented cause — assembled from the
# correction chain's own comments and the page-verification campaigns.
# Four baseline ambers carried no recorded reason anywhere; their causes are
# documented here from the page reads, marked as such.
_TYMC_STANDS = ("TYMC %s CFU/g against the cat. C criterion 10\u2074 — over even the "
                "Ph. Eur. 5.1.4 maximum acceptable count 2\u00d710\u2074 (%s\u00d7); the certificate "
                "still concludes ОДГОВАРА. Flag stands; deviation record open.")
_TYMC_UNDET = ("TYMC %s CFU/g — conforms against the Ph. Eur. 5.1.4 maximum acceptable "
               "count 2\u00d710\u2074, over QCSP 001's literal 10\u2074. UNDETERMINED until QC rules "
               "how QCSP 001 reads its maximum.")
FLAG_DOSSIER = {
    "320/0587/25": _TYMC_STANDS % ("4.2\u00d710\u2074", "2.10"),
    "904/1589/25": _TYMC_STANDS % ("3.3\u00d710\u2074", "1.65"),
    "946/1684/25": _TYMC_STANDS % ("3.6\u00d710\u2074", "1.80") +
        " Value read 10\u00b3 in the first transcription; corrected to 10\u2074 on the page read.",
    "948/1686/25": _TYMC_STANDS % ("2.6\u00d710\u2074", "1.30"),
    "1032/1851/25": _TYMC_STANDS % ("4.9\u00d710\u2074", "2.45") + " Largest count on the register.",
    "472/0863/25": _TYMC_UNDET % "1.9\u00d710\u2074",
    "587/1066/25": _TYMC_UNDET % "1.5\u00d710\u2074",
    "628/1129/25": _TYMC_UNDET % "1.2\u00d710\u2074",
    "949/1687/25": _TYMC_UNDET % "1.7\u00d710\u2074" + " The closest call on the register.",
    "1220/2171/25": "TYMC 200 CFU/g against the certificate's own printed limit 10\u00b2 — "
        "2\u00d7 over on its own paper, conforming against the register column's 10\u2074. "
        "A per-certificate manufacturer spec, outside the pharmacopoeial rule.",
    "ППК25154": "Data-integrity flag on the source corpus, not on the batch: the RAGflow "
        "corpus holds a corrupted total 1.87; the register's 18.27 is page-confirmed correct.",
    "ППК25155": "Baseline amber on CBD 0.1 % with no recorded reason anywhere in the "
        "chain. Page read confirms CBD 0.10 %, conforming (< 1.00). Documented on "
        "rectification, 31.08.2026 — reads as a transcription-check mark, not a finding.",
    "ППК26033": "CBN above the \u2264 1.00 % criterion on a 40\u00b0C/75%RH accelerated-stability "
        "arm — a stability finding, not a release failure (R5).",
    "ППК26035": "CBN above the \u2264 1.00 % criterion on a 40\u00b0C/75%RH accelerated-stability "
        "arm — a stability finding, not a release failure (R5).",
    "ППК26037": "CBN 1.09 % above the \u2264 1.00 % criterion on a stability arm — a stability "
        "finding, not a release failure (R5).",
    "ППК26058": "CBN 2.05 % above the \u2264 1.00 % criterion on an accelerated-stability "
        "timepoint — a stability finding, not a release failure (R5).",
    "ППК26127": "The certificate itself concludes НЕ ОДГОВАРА — foreign matter: cannabis "
        "seed present. Every numeric value on the page is in specification; the failure "
        "lives only in the conclusion line.",
    "197-7-К/26": "The register's CBD/CBN pair has a history: transposed in the "
        "first transcription, corrected by the page campaign (chain step 7), swapped "
        "back by a misreading in chain step 16, and restored to the certificate's "
        "values — CBD 0.22, CBN < LOQ — by chain step 17 after the rectification "
        "cross-check caught the regression against two primary sources. Both values "
        "conform either way; the amber records the history.",
    "197-14-М/26": "Ochratoxin A 2.06 \u00b5g/kg — DETECTED above LOQ, conforming against "
        "the 20 \u00b5g/kg criterion. Baseline amber; cause documented on rectification, "
        "31.08.2026.",
    "100-2-ГС/26": "Loss on drying 10.3 % — above the 10.00 % limit CNP applies. The page "
        "also prints twin-batch label J31112501 where its siblings print J31122501 "
        "(see FARMAHEM_PAGE_VERIFICATION_2026-08-31).",
    "100-3-ГС/26": "Loss on drying 10.3 % — above the 10.00 % limit CNP applies. Same "
        "twin-batch label defect as 100-2-ГС/26, and the scan is bound backwards "
        "(results page first).",
    "197-6-К/26": "The batch cell is flagged, not a value: cultivation batch CC012601/1 "
        "appears in no register and no issue plan — identity unresolved (P060332). "
        "No certificate of quality can issue until it is.",
}


def main(out):
    rows, per_coq, dets = CQ.schedule()
    plan = CQ.icoa_plan(per_coq)
    specj = json.load(open(os.path.join(HERE, "product_specifications_QCSP001.json"),
                           encoding="utf-8"))["specifications"]

    coqs, by_n = [], {}
    for r in rows:
        k = (r["CoQ number"], r["Cultivation batch"], r["CoQ type"])
        if k not in by_n:
            p = per_coq[len(coqs)]
            assert p["number"] == r["CoQ number"] and p["cb"] == r["Cultivation batch"]
            by_n[k] = {
                "n": p["number"], "t": p["type"], "basis": p["basis"],
                "issue": p["date"], "pp": p["pp"], "cb": p["cb"],
                "strain": p["strain"], "grade": p["grade"], "cls": p["cls"],
                "ic": p["icoa_ref"], "thc": p["banner_thc"],
                "spec": p["spec_doc"], "conflict": p["spec_conflict"],
                "reg": p["in_register"], "issued": p["issued"],
                "md": p.get("md", ""), "pk": p.get("pk", ""),
                "pcode": (specj.get(p["pp"]) or {}).get("product_code", ""),
                "rows": [],
            }
            coqs.append(by_n[k])
        by_n[k]["rows"].append({
            "no": r["№"], "crit": r["Acceptance criterion"],
            "res": r["Result"], "doc": r["Source document"],
            "dd": r["Document date"], "lab": r["Issuing institution"],
            "fam": r["Report series"], "st": r["Status"],
            "route": r["Performed by"], "also": r["Also on file"],
        })

    docs = [r for r in DR.load_register()
            if not r["code"].lower().startswith(("n/a", "(not numbered)"))]
    docs.sort(key=lambda r: (DR.key(r["date"]), r["row"]))
    seen = set()
    docs = [r for r in docs
            if not (r["code"].strip() in seen or seen.add(r["code"].strip()))]
    ver = DR.verified_map()
    rectmap = page_rect_map()
    corpus = corpus_rect_map()
    import verification_coverage as VC

    def rect_for(code):
        vals = next((rectmap[c] for c in VC.candidates(code) if c in rectmap), [])
        extra = corpus.get(code.strip(), [])
        have = {x["k"] for x in vals}
        return vals + [x for x in extra if x["k"] not in have]

    ecoa = [{
        "date": r["date"], "lab": r["lab"], "code": r["code"], "batch": r["batch"],
        "ref": r["ref"], "strain": r["strain"], "pn": r["pn"],
        "reported": r["reported"], "flag": ("red" if "red" in r["flags"] else
                                            ("amber" if r["flags"] else "")),
        "verified": bool(DR.page_verified(r["code"], ver)),
        "rect": rect_for(r["code"]),
        "why": FLAG_DOSSIER.get(r["code"].strip(), ""),
        "pdf": r["pdf"],
    } for r in docs]

    # The release-register view: every certificate row of every batch block,
    # with the column criteria, so the page can show the register the way the
    # published register artifact does — and judge values with the same
    # acceptance-limit rule.
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    wb = load_workbook(CQ.REG_X)
    ws = wb[CQ.SHEET]
    columns = {}
    for cidx in range(5, 23):
        L = get_column_letter(cidx)
        columns[L] = {"name": str(ws.cell(row=4, column=cidx).value or ""),
                      "crit": str(ws.cell(row=5, column=cidx).value or "")}
    order, batches, _limits = CQ.read_register()
    reg = []
    for cb in order:
        b = batches[cb]
        by_code = OrderedDict()
        for L, lst in b["cells"].items():
            for c in lst:
                r = by_code.setdefault(c["code"], {
                    "code": c["code"], "date": c["date"], "lab": c["lab"],
                    "fam": c["family"], "stab": c["stability"], "vals": {},
                    "flags": {}})
                r["vals"][L] = c["value"]
                if c["flag"]:
                    r["flags"][L] = c["flag"]
        reg.append({"cb": cb, "pn": b["pnumber"], "strain": b["strain"],
                    "certs": list(by_code.values())})

    data = {
        "generated": "31.08.2026",
        "sop_effective": CQ.SOP_EFFECTIVE,
        "reg_columns": columns,
        "reg": reg,
        "dets": [{"no": d["no"], "group": d["group"], "en": d["en"],
                  "mk": d["mk"], "method": d["method"], "crit": d["criterion"],
                  "src": d["source"], "col": d["column"]} for d in dets],
        "coqs": coqs,
        "icoa_plan": plan,
        "ecoa": ecoa,
    }
    with open(out, "w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False, separators=(",", ":"))
    print(f"{out}: {len(coqs)} CoQs, {sum(len(c['rows']) for c in coqs)} rows, "
          f"{len(plan)} iCoAs, {len(ecoa)} eCoA documents, "
          f"{os.path.getsize(out)//1024} KiB")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1] if len(sys.argv) > 1 else
                  os.path.join(HERE, "coq_artifact_data.json")))
