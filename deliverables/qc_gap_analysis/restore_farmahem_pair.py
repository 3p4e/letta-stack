#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Chain step 17 — restore 197-7-К/26's CBD/CBN pair to the certificate's values.

    python3 deliverables/qc_gap_analysis/restore_farmahem_pair.py IN.xlsx OUT.xlsx

Step 16 (`apply_farmahem_transposition.py`) was wrong, and this step corrects
it — as a new step, because an applied correction is never rewritten in place.

What actually happened, in order:

1. The page-verification campaign of 31.08.2026 (morning) read all 63 Farmahem
   reports and found row 276 transposed: the register held CBD `< LOQ` with
   CBN `0.22` where the certificate prints

       Total CBD   0.22 %w/w  (U 0.01, k=2)
       Total CBN   < LOQ      (U /)

   `apply_farmahem_corrections.py` (chain step 7) corrected it to
   G=`0.22`, H=`< LOQ` — matching the page.
2. Step 16, written later the same day, swapped the corrected pair back to
   G=`< LOQ`, H=`0.22`, un-fixing the fix. It was not a misreading: step 16
   cited `ingestion/coa_track/letta-imb-coas/exports/master_coa_table.tsv`,
   an older derived export, and that file genuinely pairs the two values the
   other way round (`Total CBN … 0.22`, `Total CBD … <LOQ`) — on this one
   certificate only. So the register held a **source conflict**, not a slip:
   two direct readings of the certificate against one derived export.
3. The rectification cross-check of 31.08.2026 (evening) compared every
   register value against the page reads and caught the regression. Two
   independent primary sources agree against step 16:
   `review/farmahem_page_reads_2026-08-31.json` ("1977K26": cbd 0.22,
   cbn < LOQ) and the certificate's own results table in the corpus chunk
   text (`P060352, 197-7-K-26, 07.08.2026, FHM.pdf`):

       | Вкупен Cannabidiol  | Total CBD | 0.22 | 0.01 |
       | Вкупен Cannabinol   | Total CBN | <LOQ | /    |

The direct readings win: a visual read of the rendered page and a verbatim
transcription of the certificate's own results table both outrank an export
derived from an earlier parse. The export was swept against all 31 Farmahem
page reads during the after-intervention review — 93 values compared, and this
pair is its only disagreement, so it is not a systematic ordering difference
but one transposed record. **`master_coa_table.tsv` still holds the transposed
pair**: anything re-derived from that export reintroduces this defect, and a
warning to that effect sits beside it.

Neither value changes a disposition — both conform to ≤ 1.00 % — but a CoQ
compiled for P060352's reissue would have printed them under the wrong
parameters. The amber fill stays: the cell's history is the finding.

Refuses to run if the cells differ from the state step 16 left them in, so it
cannot corrupt a different revision. Second run on its own output reports the
change as already applied.
"""
import sys

from openpyxl import load_workbook
from openpyxl.comments import Comment

SHEET = "Batch Release QC"
ROW = 276
SENTINEL = "estored 31.08.2026"

NOTE = (
    "Restored 31.08.2026. Step 16 of the correction chain transposed this pair "
    "the wrong way: the certificate prints\n\n"
    "    Total CBD       0.22 %w/w (U 0.01, k=2)\n"
    "    Total CBN       < LOQ (U /)\n\n"
    "confirmed against two primary sources — the page read of 31.08.2026 and the "
    "certificate's own results table in the corpus chunk text. The morning "
    "page-verification fix (chain step 7) had it right; step 16 misread the "
    "certificate and un-fixed it; this step restores the certificate's values. "
    "Both values conform to the 1.00 % criterion either way — the register "
    "reports what the laboratory reports.\n\n"
    "History of this cell: transposed in the first transcription; corrected by "
    "the page-verification campaign of 31.08.2026 (chain step 7, whose note read "
    "'Corrected 31.08.2026 against the page: the certificate reports Total CBD "
    "0.22 and Total CBN < LOQ'); swapped back by chain step 16 on the authority "
    "of an older derived export; restored here as chain step 17."
)


def main(inp, outp):
    wb = load_workbook(inp)
    ws = wb[SHEET]
    code = str(ws.cell(row=ROW, column=23).value or "").strip()
    if code != "197-7-К/26":
        print(f"REFUSED: W{ROW} reads {code!r}, expected '197-7-К/26'")
        return 2
    g, h = ws.cell(row=ROW, column=7), ws.cell(row=ROW, column=8)
    if (str(g.value), str(h.value)) == ("0.22", "< LOQ") and \
            g.comment and SENTINEL in g.comment.text:
        print("already applied: G276='0.22', H276='< LOQ' with restore note")
        wb.save(outp)
        return 0
    if (str(g.value), str(h.value)) != ("< LOQ", "0.22"):
        print(f"REFUSED: G{ROW}/H{ROW} read ({g.value!r}, {h.value!r}), "
              "expected ('< LOQ', '0.22') — a different revision of the workbook")
        return 2
    g.value, h.value = "0.22", "< LOQ"
    for c in (g, h):
        c.comment = Comment(NOTE, "QC data verification", height=270, width=430)
    wb.save(outp)
    print(f"applied: G{ROW}='0.22', H{ROW}='< LOQ' — certificate order restored")
    return 0


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
