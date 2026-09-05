#!/usr/bin/env python3
"""Build CoQ_Analysis_Master_v8.xlsx — v8's readings in the block layout, criteria enforced.

This is the convergence of the two lines of work:

  from v8 (PR #17, built from the eCoA database on the ingestion host)
      the readings themselves — verbatim from the certificate, two independent reads,
      derived cannabinoid totals, the per-compound pesticide panel read as one result,
      and the vocabulary that never calls an accredited certificate silent:
      "held for review", "not on this certificate", "not ingested".
  from v7 (this line)
      the layout the Head of QC specified — one two-row block per TESTING INSTANCE —
      the acceptance criteria in header row 3 and enforced per Ph. Eur. 5.1.4, the lot
      join on the P batch, the merge of an original and a re-analysis row into one lot,
      and the Credit Audit.

Merge rule, and it matters: v8's reading wins where it has one, because it is verbatim
from the page. Where v8 reports no value but the desk holds one from the release register
or a page read, THE DESK VALUE STANDS and is marked ᴿ — "not ingested" is a statement
about v8's corpus, not about the certificate, and a verified result is never dropped by a
rebuild. Checked before adopting: of the values both hold, 754 agree once decimal commas,
unit suffixes and Cyrillic connectives are normalised, and none contradict.

The block rule, per the specification:

  * one TESTING INSTANCE = one block of two rows, and a batch holds as many blocks as it
    has testing instances (the largest number of certificates credited to any one of its
    parameters), so a parameter tested twice on two dates gets two blocks, never two text
    lines inside one; the batch identity and STATUS are merged down all of its blocks and
    the whole batch is boxed with a thick border;
  * single-value parameters (#1–#8, #12) occupy Result | eCOA ref | ✓/✗, each cell
    merged vertically across the two rows;
  * #9 Microbiology, #10 Mycotoxins and #11 Heavy metals give each sub-determination
    its own column: the top row carries the sub-results, the bottom row carries the
    certificate reference(s) merged across them;
  * a parameter's certificates are taken in ascending date order: the n-th certificate
    fills the n-th block, so a row of blocks reads across as "everything known from the
    n-th round of testing";
  * acceptance criteria live in header row 3 and are enforced: a result that provably
    exceeds its criterion is printed red and bold and is named in the batch's STATUS.

Sources: `tracker_data.py` (the desk's values, the owner's certificate credits).
"""
import collections, importlib.util, math, os, re, sys

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter as L

HERE = os.path.dirname(os.path.abspath(__file__))
spec = importlib.util.spec_from_file_location("tracker_data", os.path.join(HERE, "tracker_data.py"))
T = importlib.util.module_from_spec(spec)
spec.loader.exec_module(T)

V9 = "--v9" in sys.argv
# --version=N names the build: the file, the tracker sheet and the Read Me carry vN.
VER = next((a.split("=", 1)[1] for a in sys.argv if a.startswith("--version=")), "9" if V9 else "8")
V9 = V9 or VER not in ("8",)
SRC = os.path.join(HERE, "CoQ_Analysis_Master_v6.xlsx")
V8VALS = os.path.join(HERE, "v8_values.json")
OUT = os.path.join(HERE, f"CoQ_Analysis_Master_v{VER}.xlsx")
SHEET = f"CoQ Parameter Tracker v{VER}"
BUILT = "02.09.2026"
import json  # noqa: E402
STATUS_PARTIAL_MAX = 3          # 1–3 without a release result → ⚠ PARTIAL; ≥4 → ✗

# --------------------------------------------------------------------------- palette
NAVY = "1F3864"; SUBHDR = "FFF2CC"; GREY = "EFEFEF"
FILL = {"green": "C6EFCE", "orange": "FCE5CD", "amber": "FDE9D9", "red": "F4CCCC", "extra": "EDEDED"}
GLYPHFILL = {"green": "6AA84F", "orange": "E69138", "amber": "F6B26B", "red": "CC0000", "extra": "A6A6A6"}
STATUSFILL = {"green": "38761D", "orange": "E69138", "red": "CC0000"}
RED = "9C0006"
F7 = Font(name="Calibri", size=7)
F7B = Font(name="Calibri", size=7, bold=True)
F7R = Font(name="Calibri", size=7, bold=True, color=RED)
F7U = Font(name="Calibri", size=7, bold=True, color="B45F06")
FBAD = Font(name="Calibri", size=7, bold=True, color=RED)
F6 = Font(name="Calibri", size=6)
F6I = Font(name="Calibri", size=6.5, italic=True, color="595959")
FW = Font(name="Calibri", size=8, bold=True, color="FFFFFF")
FWS = Font(name="Calibri", size=7, bold=True, color="FFFFFF")
FSUB = Font(name="Calibri", size=7, bold=True)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
TOPC = Alignment(horizontal="center", vertical="top", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
THICK = Side(style="medium", color="000000")
MED = Side(style="thin", color="404040")


def put(ws, r, c, v, font=F7, fill=None, al=CEN):
    cell = ws.cell(r, c, v)
    cell.font, cell.alignment, cell.border = font, al, BOX
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    return cell


def fill_range(ws, r1, c1, r2, c2, colour):
    """A merged range takes its format from the anchor cell in Excel and Google Sheets
    alike, so only the anchor is styled — writing the covered cells doubled the file."""
    ws.cell(r1, c1).fill = PatternFill("solid", fgColor=colour)


def outline(ws, r1, c1, r2, c2, side):
    """Draw one rectangle without disturbing the inner borders."""
    for c in range(c1, c2 + 1):
        for r, edge in ((r1, "top"), (r2, "bottom")):
            b = ws.cell(r, c).border
            ws.cell(r, c).border = Border(left=b.left, right=b.right,
                                          top=side if edge == "top" else b.top,
                                          bottom=side if edge == "bottom" else b.bottom)
    for r in range(r1, r2 + 1):
        for c, edge in ((c1, "left"), (c2, "right")):
            b = ws.cell(r, c).border
            ws.cell(r, c).border = Border(top=b.top, bottom=b.bottom,
                                          left=side if edge == "left" else b.left,
                                          right=side if edge == "right" else b.right)


def nlines(text, width):
    per = max(1, int(width * 1.25))
    return sum(max(1, math.ceil(len(ln) / per)) for ln in str(text or "").split("\n")) or 1


# --------------------------------------------------------------------------- data
VAL, INH, STAB, STRAIN, DET = T.load_desk()
batches, index_rows = T.load_owner()

# Every document the index holds for a batch, with the parameters it covers. A document
# that covers a parameter but is not credited to it on the owner's tracker is still a
# testing instance and is shown — greyed, marked "•", and labelled "not credited" — but it
# is NOT counted as coverage: what discharges a parameter stays the owner's judgement.
# The index is joined on the P BATCH, never on the CU code: four CU codes carry two
# tracker rows (an original and the August re-analysis) and three lots share a CU with
# no code of their own, so a CU join pulls another lot's certificates into the batch.
COVERS = {}
INDEX_DOCS = collections.defaultdict(list)
for _r in index_rows:
    _cov = {int(x[1:]) for x in re.findall(r"#\d+", _r["params"])}
    _k = _r["p"] or re.sub(r"[＊*]", "", _r["cu"])
    INDEX_DOCS[_k].append((_r["code"], _r["date"], _r["lab"], _cov))
    COVERS[(_k, T.nkey(_r["code"]))] = _cov


# v8's readings, keyed on the certificate. Extracted from the v8 workbook of PR #17,
# which was built from the eCoA database; the map is committed so this build is
# reproducible without the database, which lives on the ingestion host.
def tidy(v):
    """v8 prints the page verbatim, which mixes decimal commas with decimal points and
    repeats the unit the column header already states. The separator and the unit are
    rendering, not measurement: normalise those two and leave everything else as printed."""
    t = str(v or "").strip()
    t = re.sub(r"(?<=\d),(?=\d)", ".", t)
    t = re.sub(r"\s*[xх×]\s*10\s*\^?\s*(\d+)",
               lambda m: "×10" + "".join("⁰¹²³⁴⁵⁶⁷⁸⁹"[int(c)] for c in m.group(1)), t)
    t = re.sub(r"\s+(CFU/g|µg/kg|mg/kg|%|w/w)\s*$", "", t, flags=re.I)
    t = re.sub(r"([<>≤≥])\s+(?=[\d.])", r"\1", t)
    return t.strip()


_V8 = json.load(open(V8VALS))
SILENT = ("held for review", "not on this certificate", "not ingested", "n.r.")
V8VAL, V8SILENT = collections.defaultdict(dict), collections.defaultdict(dict)
_bycode = collections.defaultdict(set)
for _k in _V8:
    _bycode[_k.split("|", 1)[1]].add(_k)
for _k, _d in _V8.items():
    _cu, _ck = _k.split("|", 1)
    for _no, _v in _d.items():
        _val = _v if _v in SILENT else tidy(_v)
        (V8SILENT if _v in SILENT else V8VAL)[(_cu, _ck)][_no] = _val
        if len(_bycode[_ck]) == 1:                    # the code identifies one lot only
            (V8SILENT if _v in SILENT else V8VAL)[("*", _ck)][_no] = _val


def v8_of(code, cu, table):
    ck = T.nkey(code)
    return table.get((cu, ck)) or table.get(("*", ck)) or {}


def was_read(code, lab, cu):
    """True when the desk holds any value from this document. A credited document with no
    value at all was never read into the corpus — a different problem from a document that
    was read and simply does not report the parameter."""
    if VAL.get(T.nkey(code)):
        return True
    return T.kind_of(code, lab, STAB) == "In-house" and bool(INH.get(T.cu_key(cu)))


def silence_reason(code, lab, cu, no=None):
    """v8's own words when it has them; otherwise derived from the desk."""
    if no is not None:
        said = silent_as(code, cu, no)
        if said:
            return said
    said = {v for d in (v8_of(code, cu, V8SILENT),) for v in d.values()}
    if said and len(said) == 1:
        return said.pop()
    return "not on this certificate" if was_read(code, lab, cu) else "not ingested"


def scope_of(b_or_key, code):
    key = b_or_key if isinstance(b_or_key, str) else join_key(b_or_key)
    return COVERS.get((key, T.nkey(code)))


def join_key(b):
    """The lot: its P batch, or its CU code when no P batch is assigned. The owner marks
    some CU codes with an asterisk; the index does not, so the mark is folded away."""
    return re.sub(r"[＊*]", "", b["cu"]) if b["p"].startswith("N/A") else b["p"]


# One batch is one lot: tracker rows that share a CU and a P batch are the same lot split
# by the owner across an original row and a re-analysis row, which the flat layout forced.
# The block layout carries both rounds, so they are merged back into a single batch.
_merged, _by = [], {}
for _b in batches:
    _k = (_b["cu"], _b["p"])
    if _k in _by:
        _t = _by[_k]
        for _n in range(1, 13):
            _have = {T.nkey(_x[0]) for _x in _t["docs"][_n]}
            for _d in _b["docs"][_n]:
                if T.nkey(_d[0]) not in _have:
                    _t["docs"][_n].append(_d)
                    _have.add(T.nkey(_d[0]))
        _t["labs"] = sorted(set(_t["labs"]) | set(_b["labs"]))
        _t["rows"] = _t.get("rows", 1) + 1
    else:
        _by[_k] = _b
        _merged.append(_b)
print(f"tracker rows {len(batches)} → batches {len(_merged)} "
      f"({sum(1 for b in _merged if b.get('rows', 1) > 1)} lots had an original and a re-analysis row)")
batches = _merged


# --------------------------------------------------------------------------- new testing instances
# Certificates ingested after the owner's tracker was read (eCOA_DB, 04.09.2026). Each is a
# testing instance credited to the parameter it reports, carrying the values the two
# independent reads agreed on; a value the reads disagreed on is 'held for review' until a
# person rules on the page. A certificate naming a P batch the tracker does not carry opens a
# new lot row with no CU code, and a Work Order task to record it.
NEW_INSTANCES = next((a.split("=", 1)[1] for a in sys.argv if a.startswith("--new-instances=")),
                     os.path.join(HERE, "new_instances.json"))
NEW = json.load(open(NEW_INSTANCES)) if os.path.exists(NEW_INSTANCES) else []
NEW_TOUCHED, NEW_LOTS, NEW_HELD = [], [], []


def _lot_of(inst):
    for b in batches:
        if inst["p"]:
            if inst["p"] in [x.strip() for x in b["p"].split("/")]:
                return b
        elif b["p"].startswith("N/A") and re.sub(r"[＊*]", "", b["cu"]) == inst["cu"]:
            return b
    return None


for inst in NEW:
    b = _lot_of(inst)
    if b is None:
        b = {"cu": "— not recorded —", "p": inst["p"] or "N/A — no P batch assigned", "status": "",
             "labs": [], "docs": {n: [] for n in range(1, 13)}, "certs": [],
             "strain": inst.get("strain", ""), "new_lot": True}
        batches.append(b)
        NEW_LOTS.append(b)
    b.setdefault("strain", inst.get("strain", ""))
    key = join_key(b)
    ck = T.nkey(inst["code"])
    for n in inst["params"]:
        if ck not in {T.nkey(c) for c, d, l in b["docs"][n]}:
            b["docs"][n].append((inst["code"], inst["date"], inst["lab"]))
    INDEX_DOCS[key].append((inst["code"], inst["date"], inst["lab"], set(inst["params"])))
    COVERS[(key, ck)] = set(inst["params"])
    V8VAL[("*", ck)] = dict(inst["vals"])        # read from the eCoA database: no ᴿ mark
    VAL[ck] = dict(inst["vals"])
    for no in inst.get("held", []):
        NEW_HELD.append((b["cu"], b["p"], inst["code"], inst["date"], inst["lab"], int(no.split(".")[0])))
    if b not in NEW_TOUCHED:
        NEW_TOUCHED.append(b)
if NEW:
    print(f"new instances: {len(NEW)} on {len(NEW_TOUCHED)} lot(s); "
          f"{len(NEW_LOTS)} lot(s) not on the owner's tracker; {len(NEW_HELD)} value(s) held for review")
    if V9 and VER == "9":
        VER = "9.1"
        OUT = os.path.join(HERE, "CoQ_Analysis_Master_v9.1.xlsx")

# --------------------------------------------------------------------------- the iCoA rule (04.09.2026)
# Head of QC, 04.09.2026: identification A (appearance) and B (microscopy) are tested at
# Purely Plant together with foreign matter, at the date of packaging, and ONE iCoA per batch
# carries the three results for release. Identification C conforms to the ImB specification
# and is referenced to the certificate that carries the cannabinoid assay (#4), which the
# desk already credits. Harvest and packaging dates per batch are the Head of QC's list of
# 04.09.2026 (batch_dates.csv, --dates=): the in-house instance is dated on the FIRST day of
# packaging — the day the issuance plan already uses as the CoQ basis — and a lot the list
# does not carry keeps "packaging date — to record" and sorts last. The iCoA Issuance sheet
# lists what each batch's iCoA must carry, Batch Dates the list itself. Foreign matter is "Conforms" by the declaration of 13.08.2026,
# except where an outsourced certificate reports otherwise (FB032601, ППК26127: 0.08 %,
# Не одговара) — that lot's foreign matter is held for the Head of QC.
ICOA_RULE = "--icoa" in sys.argv
ICOA_ROWS = []
NEW_NONCONF = []
ADDED = collections.defaultdict(list)        # lot -> documents added by this build (coverage recount)
for _b in NEW_TOUCHED:
    ADDED[id(_b)] += [(x["code"], x["date"], x["lab"]) for x in NEW if _lot_of(x) is _b]
if ICOA_RULE:
    _D = json.load(open(T.DESK))
    _plan = {}
    for _x in _D["icoa_plan"]:
        _plan.setdefault(T.batch_key(re.sub(r"[＊*]", "", _x["cb"])), _x["icoa_ref"])
        if _x.get("pp"):
            _plan.setdefault("P:" + _x["pp"].strip(), _x["icoa_ref"])
    ICOA_DATE = "packaging date — to record"
    RETEST_DATE = "retest sampling date — to record"
    _bd = importlib.util.spec_from_file_location("batch_dates", os.path.join(HERE, "batch_dates.py"))
    BD = importlib.util.module_from_spec(_bd)
    _bd.loader.exec_module(BD)
    _dates_path = next((a.split("=", 1)[1] for a in sys.argv if a.startswith("--dates=")), os.path.join(HERE, "batch_dates.csv"))
    DATES = BD.load_dates(_dates_path)
    DATE_ROWS = sorted({r["seq"]: r for r in DATES.values()}.values(), key=lambda r: int(r["seq"]))
    DATES_CU = {T.batch_key(r["cu_batch"]): r for r in DATE_ROWS}
    DATE_USED = {}                                # seq -> the tracker lot it dated

    def _dates_of(b, cu0):
        """The list's rows for a lot: by P-number (one CU row may hold several P lots), else by CU code."""
        rows = []
        for p in b["p"].split("/"):
            r = DATES.get(p.strip())
            if r and r not in rows:
                rows.append(r)
        if not rows and DATES_CU.get(T.batch_key(cu0)):
            rows.append(DATES_CU[T.batch_key(cu0)])
        for r in rows:
            DATE_USED[r["seq"]] = b["cu"] if b["cu"] and not b["cu"].startswith("—") else b["p"]
        return rows

    def _span_of(rows, a, z, many):
        out = []
        for r in rows:
            t = BD.span(r[a], r[z]) or "— not given —"
            out.append(f"{r['p_batch']}: {t}" if many and r["p_batch"] else t)
        return " | ".join(dict.fromkeys(out))
    print(f"batch dates: {len(DATE_ROWS)} row(s) from {os.path.basename(_dates_path)}")
    _COQ, _RETEST = {}, {}
    for _c in _D["coqs"]:
        _tbl = _COQ if _c["t"].startswith("initial release") else _RETEST
        _tbl.setdefault(T.batch_key(re.sub(r"[＊*]", "", _c["cb"])), _c)
        if _c.get("pp"):
            _tbl.setdefault("P:" + _c["pp"].strip(), _c)

    def _lookup(tbl, b, cu0):
        return tbl.get(T.batch_key(cu0)) or next((tbl.get("P:" + p.strip()) for p in b["p"].split("/") if tbl.get("P:" + p.strip())), None)

    def _ident_c(b, after=None):
        """The cannabinoid-assay certificate that covers identification C: the assay
        certificate itself, never a loss-on-drying one; `after` restricts it to the retest."""
        pool = list(b["docs"][4]) or list(b["docs"][3])
        pool = [x for x in pool if not re.search(r"LoD|ГС", x[0], re.I)] or pool
        if after:
            pool = [x for x in pool if str(T.date_key(x[1])) >= str(T.date_key(after))]
        pool = sorted(pool, key=lambda x: (T.date_key(x[1]), x[0]))
        return f"{pool[0][0]}, ({pool[0][1]}) [{pool[0][2]}]" if pool else None

    CNP_COVERED = {}

    def _vals_of(code, cu):
        """v8's reading first, the desk's beneath it (values_of is defined further down)."""
        ck = T.nkey(code)
        out = dict(VAL.get(ck) or {})
        out.update(V8VAL.get(("*", ck)) or {})
        out.update(V8VAL.get((cu, ck)) or {})
        return out

    _pending_inst = []                            # (lot, key, scope, values, date, row): the in-house instance, keyed after the register
    for b in batches:
        cu0 = re.sub(r"[＊*]", "", b["cu"])
        key = join_key(b)
        _drows = _dates_of(b, cu0)
        # --- CNP first: a CNP certificate that reports identification A, B or foreign matter
        #     is the reference for them (its document code goes on the CoQ)
        cnp = {}
        seen = set()
        for c, d, l, cov in list(INDEX_DOCS.get(key, [])) + [(c, d, l, set()) for n in (1, 2, 3, 4, 7) for c, d, l in b["docs"][n]]:
            if l != "CNP" or T.nkey(c) in seen:
                continue
            seen.add(T.nkey(c))
            vals = _vals_of(c, b["cu"])
            for n in (1, 2, 7):
                if vals.get(str(n)) and n not in cnp:
                    cnp[n] = (c, d, l)
        for n, (c, d, l) in cnp.items():
            if T.nkey(c) not in {T.nkey(x[0]) for x in b["docs"][n]}:
                b["docs"][n].append((c, d, l))
            COVERS[(key, T.nkey(c))] = set(COVERS.get((key, T.nkey(c)), set())) | {n}
            ADDED[id(b)]
        CNP_COVERED[id(b)] = cnp
        if cnp and b not in NEW_TOUCHED:
            NEW_TOUCHED.append(b)
        fm_cnp = cnp.get(7)
        if fm_cnp and "Не одговара" in str(_vals_of(fm_cnp[0], b["cu"]).get("7", "")):
            NEW_NONCONF.append((b["cu"], b["p"], fm_cnp[0], fm_cnp[1], "CNP", 7))
        # --- TWO iCoAs PER P LOT (Head of QC, 05.09.2026): identification A and B on the P01-02
        #     master, foreign matter on the P07 master — one of each per P lot, unless a CNP
        #     certificate reports that scope (then the certificate is the reference and no iCoA is
        #     needed for it). A tracker row that holds several P lots (GRC102501: two, JD012603:
        #     three) is one lot per P number; a row with one P lot (or none, the R&D lots) is one
        #     lot. The CNP references are the row's. Every certificate row carries a KEY
        #     (lot|AB or FM|I initial or R retest) that the sheets' lookup formulas use.
        _plist = [p.strip() for p in b["p"].split("/") if p.strip().upper().startswith("P0")]
        _lots = _plist if len(_plist) > 1 else [b["p"]]
        _single = len(_lots) == 1
        scope = [n for n in (1, 2, 7) if n not in cnp]
        fm = (INH.get(T.cu_key(b["cu"])) or {}).get("7") or "Conforms"
        vals = {"1": "Conforms", "2": "Conforms", "7": fm}
        cell = lambda n: (f"CNP {cnp[n][0]}" if n in cnp else vals[str(n)])
        for _lot in _lots:
            _lot_id = _lot.strip() if re.match(r"^P0\d{5}$", _lot.strip()) else cu0
            _lrows = [r for r in _drows if _single or r["p_batch"] == _lot]
            _lpk = sorted((r["packaging_from"] for r in _lrows if r["packaging_from"]), key=lambda d: str(T.date_key(d)))
            _lend = sorted((r["packaging_to"] for r in _lrows if r["packaging_to"]), key=lambda d: str(T.date_key(d)))
            _ldate = _lpk[0] if _lpk else ICOA_DATE
            _complete = _lend[-1] if _lend else "— packaging date to record —"
            _lharvest = _span_of(_lrows, "harvest_from", "harvest_to", False) if _lrows else "— not on the list —"
            _lpackaging = _span_of(_lrows, "packaging_from", "packaging_to", False) if _lrows else "— not on the list —"
            if _single:
                _pref = _plan.get(T.batch_key(cu0)) or next((_plan["P:" + p.strip()] for p in b["p"].split("/") if _plan.get("P:" + p.strip())), None)
            else:
                _pref = _plan.get("P:" + _lot)
            _pref = _pref if (_pref and _pref.startswith("iCoA-")) else ""
            _coq = _lookup(_COQ, b, cu0) if _single else _COQ.get("P:" + _lot)
            if _coq and _coq.get("basis") and _lpk and _coq["basis"] != _lpk[0]:
                _lpackaging += f" (the issuance plan's basis: {_coq['basis']})"
            _strain = b.get("strain") or STRAIN.get(T.batch_key(cu0), "")
            # a lot the owner's tracker does not carry has no CU code there; the Head of QC's list names it
            _cu_show = b["cu"] if not b["cu"].startswith("—") else (f"{_lrows[0]['cu_batch']} (from the list)" if _lrows else b["cu"])
            # --- RETEST SERIES: the QP's retesting campaign (medical use, GACP product / API) —
            #     the same two certificates again per P lot, dated at the retest sampling; a
            #     cannabinoid assay dated more than 60 days after the release basis is the retest
            #     (the Farmahem 197-series of August 2026 is the QP's campaign, whatever the plan's
            #     nominal 12-month date says); release-time CNP coverage does not excuse a reissue
            _rt = _lookup(_RETEST, b, cu0) if _single else _RETEST.get("P:" + _lot)
            _after = None
            if _rt and _coq and _coq.get("basis"):
                import datetime as _dt
                try:
                    _after = (_dt.datetime.strptime(_coq["basis"], "%d.%m.%Y") + _dt.timedelta(days=60)).strftime("%d.%m.%Y")
                except ValueError:
                    _after = None
            c_rt = _ident_c(b, after=_after) if _after else None

            def _retest_doc(pno):
                if not _after:
                    return None
                pool = [x for x in b["docs"][pno] if str(T.date_key(x[1])) >= str(T.date_key(_after))
                        and not re.search(r"LoD|ГС", x[0], re.I)]
                pool = sorted(pool, key=lambda x: (T.date_key(x[1]), x[0]))
                return f"{pool[0][0]}, ({pool[0][1]}) [{pool[0][2]}]" if pool else None
            m_rt = _retest_doc(10) if _rt else None
            _st_rt = ("due — retest assay and mycotoxins on file" if (c_rt and m_rt) else
                      "due — retest assay on file, mycotoxins pending" if c_rt else
                      "pending — retest assay not yet on file")
            for _sc, _scn, _master in (("AB", [n for n in scope if n in (1, 2)], "iCoA_P01-02 · Appearance & Identification A/B"),
                                       ("FM", [7] if 7 in scope else [], "iCoA_P07 · Foreign matter")):
                _names = {1: "Ident A", 2: "Ident B", 7: "Foreign matter"}
                _full = " + ".join(_names[n] for n in ((1, 2) if _sc == "AB" else (7,)))
                row = {"series": "initial release", "icoa": "" if _scn else "not needed", "plan_ref": _pref,
                       "coq": _coq["n"] if _coq else "— not in the issuance plan —",
                       "basis": (_coq["basis"] if _coq and _coq.get("basis") else (_lpk[0] if _lpk else "")),
                       "issue": ("≥ 11.05.2026 · tests at packaging " + _ldate) if _lpk else "≥ 11.05.2026 · at packaging",
                       "cu": _cu_show, "p": _lot, "strain": _strain,
                       "harvest": _lharvest, "packaging": _lpackaging, "complete": _complete,
                       "sortdate": (_lpk[0] if _lpk else ""), "sc": _sc, "sc_order": 0 if _sc == "AB" else 1,
                       "master": _master, "key": f"{_lot_id}|{_sc}|I",
                       "scope": (" + ".join(_names[n] for n in _scn) if _scn else f"— {_full} on the CNP certificate —"),
                       "a": cell(1) if _sc == "AB" else "n/a", "b": cell(2) if _sc == "AB" else "n/a",
                       "fm": cell(7) if _sc == "FM" else "n/a",
                       "c": _ident_c(b) or "— no cannabinoid certificate —",
                       "assay_rt": "", "myco_rt": "", "carry": "",
                       "status": (f"not needed — CNP covers {'A and B' if _sc == 'AB' else 'foreign matter'}" if not _scn else "to register")}
                ICOA_ROWS.append(row)
                if _scn:
                    _pending_inst.append((b, key, _scn, vals, _ldate, row))
                if _rt:
                    ICOA_ROWS.append({"series": "retest — QP campaign", "icoa": "", "plan_ref": "",
                                      "coq": _rt["n"] if not _rt["n"].startswith("(") else "CoQ reissue (assigned on issue)",
                                      "basis": _rt.get("basis", ""), "issue": "at retest sampling · QP campaign",
                                      "cu": _cu_show, "p": _lot, "strain": _strain,
                                      "harvest": _lharvest, "packaging": _lpackaging, "complete": _complete,
                                      "sortdate": "", "sc": _sc, "sc_order": 0 if _sc == "AB" else 1,
                                      "master": _master, "key": f"{_lot_id}|{_sc}|R", "scope": _full,
                                      "a": "to test" if _sc == "AB" else "n/a", "b": "to test" if _sc == "AB" else "n/a",
                                      "fm": "to test" if _sc == "FM" else "n/a",
                                      "c": c_rt or "— retest assay not yet on file —",
                                      "assay_rt": c_rt or "— pending —", "myco_rt": m_rt or "— pending —",
                                      "carry": "#8, #9, #11, #12 from the initial testing",
                                      "status": _st_rt})

    def _bkey(r):
        d = r.get("sortdate") or r["basis"]
        return (0 if r["series"] == "initial release" else 1,
                str(T.date_key(d)) if d else "99999998", r["cu"], r["p"], r["sc_order"])
    ICOA_ROWS.sort(key=_bkey)
    for _i, _row in enumerate(ICOA_ROWS, 1):
        _row["seq"] = _i
    print(f"iCoA rule: {sum(1 for r in ICOA_ROWS if r['series'] == 'initial release')} initial certificate rows "
          f"({sum(1 for r in ICOA_ROWS if r['status'].startswith('not needed'))} covered by CNP), "
          f"{sum(1 for r in ICOA_ROWS if r['series'] != 'initial release')} retest rows "
          f"({sum(1 for r in ICOA_ROWS if r['status'].startswith('due'))} with the retest assay on file)")
    print(f"batch dates: {len(DATE_USED)} of {len(DATE_ROWS)} list rows date a tracker lot; "
          f"{len({r['p'] for r in ICOA_ROWS if r['series'] == 'initial release' and r['packaging'].startswith('— not on')})} lot(s) not on the list")

    # ------------------------------------------------------------------ the preliminary iCoA issuance register
    # Head of QC, 05.09.2026: document codes iCoA-PP_26-nnn (nnn = 001 … 999, one series for the
    # year of issue), assigned in the order the certificates can be issued. Every printed issue
    # date lies in [11.05.2026 … the day of signing] (ISSUE_COQ_CONVENTIONS: the SOP floor, never
    # post-dated), so the preliminary issue date is the earliest permissible one — the SOP floor,
    # or the last day of packaging where that is later; QC sets the real date at issue. A row that
    # cannot be issued yet carries no number: a lot without a packaging date, a held result, and
    # every retest iCoA (tested at the retest sampling, which is not on the desk). Numbers are
    # never reserved ahead of issue. The plan's references of 31.08.2026 (iCoA-PP-YYYY-NNNN) are
    # superseded and kept beside the code. On the sheet the number, the code and the dates are
    # FORMULAS (see _fill_register), so a row inserted between two certificates renumbers every
    # row beneath it; the values computed here are the same numbers, for the page and the checks.
    SOP_FLOOR = "11.05.2026"
    _dk = lambda d: str(T.date_key(d))
    _issuable, _later = [], []
    for r in ICOA_ROWS:
        if r["icoa"] == "not needed":
            r["code"], r["earliest"], r["issue_date"], r["issuable"] = "", "", "", "no"
            r["reg_status"] = r["status"]
            continue
        comp = None if str(r["complete"]).startswith("—") else r["complete"]
        if r["series"] != "initial release":
            r["earliest"] = "retest sampling date — to record"
            r["why"] = ("retest assay on file; " + ("identification A and B" if r["sc"] == "AB" else "foreign matter")
                        + " to test at the retest sampling" if r["status"].startswith("due") else "retest assay not yet on file")
            _later.append(r)
        elif comp is None:
            r["earliest"], r["why"] = "— packaging date to record —", "no packaging date on the list"
            _later.append(r)
        elif r["sc"] == "FM" and r["fm"] == "held for review":
            r["earliest"], r["why"] = max(SOP_FLOOR, comp, key=_dk), "foreign matter held for the Head of QC"
            _later.append(r)
        else:
            r["earliest"] = max(SOP_FLOOR, comp, key=_dk)
            _issuable.append(r)
    _issuable.sort(key=lambda r: (_dk(r["earliest"]), _dk(r["sortdate"] or r["basis"]) if (r["sortdate"] or r["basis"]) else "9",
                                  r["cu"], r["p"], r["sc_order"]))
    for _i, r in enumerate(_issuable, 1):
        r["code"] = f"iCoA-PP_26-{_i:03d}"
        r["issue_date"] = r["earliest"]
        r["issuable"] = "yes"
        r["reg_status"] = "registered — preliminary date, QC sets the real date at issue"
        r["icoa"], r["status"], r["issue"] = r["code"], r["reg_status"], r["issue_date"]
    for r in _later:
        r["code"] = "— at issue —"
        r["issue_date"] = r["earliest"]
        r["issuable"] = "no"
        r["reg_status"] = "not yet issuable — " + r["why"]
        if r["series"] == "initial release":
            r["icoa"], r["status"], r["issue"] = r["code"], r["reg_status"], r["earliest"]
        else:
            r["icoa"] = r["code"]
    REGISTER = _issuable + sorted(_later, key=lambda r: (0 if r["series"] == "initial release" else 1 if r["status"].startswith("due") else 2,
                                                          _dk(r.get("sortdate") or r["basis"]) if (r.get("sortdate") or r["basis"]) else "9",
                                                          r["cu"], r["p"], r["sc_order"]))
    print(f"iCoA register: {len(_issuable)} numbered (iCoA-PP_26-001 … {_issuable[-1]['code'][-3:] if _issuable else '—'}: "
          f"{sum(1 for r in _issuable if r['sc'] == 'AB')} identification A/B, {sum(1 for r in _issuable if r['sc'] == 'FM')} foreign matter), "
          f"{len(_later)} not yet issuable ({sum(1 for r in _later if r['series'] == 'initial release')} initial, "
          f"{sum(1 for r in _later if r['series'] != 'initial release')} retest); "
          f"issue dates {min(_dk(r['issue_date']) for r in _issuable) if _issuable else '—'} … {max(_dk(r['issue_date']) for r in _issuable) if _issuable else '—'}")
    # the in-house instance now carries the register code (or the at-issue placeholder); the
    # tracker cell that cites it is a lookup into the register by KEY (INST_KEY), so it follows a
    # renumbering on the sheet
    INST_KEY = {}
    for b, key, scope, vals, _ldate, row in _pending_inst:
        ref = row["code"] if row["code"].startswith("iCoA-PP_") else f"iCoA — at issue ({row['p'] if row['p'].startswith('P0') else row['cu']} {row['sc']})"
        ck = T.nkey(ref)
        INST_KEY[ck] = row["key"]
        for n in scope:
            if ck not in {T.nkey(c) for c, d, l in b["docs"][n]}:
                b["docs"][n].append((ref, _ldate, "PP"))
        INDEX_DOCS[key].append((ref, _ldate, "PP", set(scope)))
        COVERS[(key, ck)] = set(scope)
        V8VAL[(b["cu"], ck)] = {str(n): vals[str(n)] for n in scope}
        ADDED[id(b)].append((ref, _ldate, "PP"))
        if b not in NEW_TOUCHED:
            NEW_TOUCHED.append(b)


# --------------------------------------------------------------------------- credit corrections
# Two corrections to the owner's credit table, each applied only where the evidence is
# explicit, each recorded on the "Credit Corrections" sheet, and neither written back to
# the owner's workbook. A removed credit does not remove the document: it still appears as
# a testing instance, marked "on file, not credited".
LOD_RX = re.compile(r"LoD|ГС", re.I)
K_RX = re.compile(r"(^|\D)\d+-\d+-[KК]-\d+")
corrections = []


def _reports(b, code, pno):
    lab = next((l for c, d, l in b["docs"][pno] if c == code), "")
    vals = values_of(code, lab, b["cu"], scope_of(b, code))
    return any(vals.get(no) for no in T.GROUPS[pno])


def apply_credit_corrections(batches):
    """R1  The Farmahem pair is credited jointly for #3–#6 and #8, but the "K" certificate
           reports identification C, THC, CBD and CBN, and the "LoD" certificate reports
           loss on drying alone. Each keeps only what it reports.
       R2  Identification B is credited to CNP certificates that carry no microscopy row.
           The credit is removed there, and kept on the certificates that do report it
           (ППК26110–26119, ППК26127–26128), where the laboratory's newer report format
           carries the determination."""
    for b in batches:
        for p in T.PARAMS:
            keep = []
            for c, d, l in b["docs"][p["n"]]:
                rule = why = None
                if LOD_RX.search(c) and p["n"] != 8 and not _reports(b, c, p["n"]):
                    rule, why = "R1 Farmahem pair", "loss-on-drying certificate; reports no such determination"
                elif K_RX.search(c) and p["n"] == 8 and not _reports(b, c, 8):
                    rule, why = "R1 Farmahem pair", "cannabinoid certificate; loss on drying is on the LoD certificate"
                elif p["n"] == 2 and l == "CNP" and not _reports(b, c, 2):
                    rule, why = "R2 CNP identification B", "no microscopy row on this certificate"
                if rule:
                    corrections.append((b["cu"], b["p"], c, d, l, p["n"], p["title"], rule, why))
                else:
                    keep.append((c, d, l))
            b["docs"][p["n"]] = keep
    return batches


def instances(b, pno):
    """(code, date, lab, credited) for every testing instance of this parameter, in date order."""
    credited = [(c, d, l) for c, d, l in b["docs"][pno]]
    seen = {T.nkey(c) for c, d, l in credited}          # fold: the tracker and the index
    out = [(c, d, l, True) for c, d, l in credited]     # spell the same code differently
    for c, d, l, cov in INDEX_DOCS.get(join_key(b), []):
        if T.nkey(c) in seen:
            continue
        reports = any((VAL.get(T.nkey(c)) or {}).get(no) for no in T.GROUPS[pno])
        if pno in cov or reports:
            out.append((c, d, l, False))
            seen.add(T.nkey(c))
    return sorted(out, key=lambda x: (T.date_key(x[1]), x[0]))


def desk_values(code, lab, cu, scope=None):
    """What this document reports. For an in-house record the desk keeps its values under
    the batch, not the document, so the fallback is restricted to the parameters the index
    says the document covers — otherwise one Report of Analysis would claim every in-house
    determination the batch holds."""
    src = VAL.get(T.nkey(code))
    if not src and T.kind_of(code, lab, STAB) == "In-house":
        src = INH.get(T.cu_key(cu))
        if src is not None and scope is not None:
            src = {no: v for no, v in src.items() if int(str(no).split(".")[0]) in scope}
    return src or {}


REGISTER_ONLY = set()          # (code, det) the desk holds and v8 does not — marked ᴿ


def values_of(code, lab, cu, scope=None):
    """v8's reading where it has one; the desk's verified value where it does not."""
    out = dict(v8_of(code, cu, V8VAL))
    for no, v in desk_values(code, lab, cu, scope).items():
        if no not in out:
            out[no] = f"{v} ᴿ"        # ᴿ: the release register or a page read, not v8's extraction
            REGISTER_ONLY.add((T.nkey(code), no))
    return out


def silent_as(code, cu, no):
    """What v8 says about a determination it holds no value for."""
    return v8_of(code, cu, V8SILENT).get(no)


def ecoa_line(code, date, lab):
    return f"{code}, ({date}) [{lab}]"


# --------------------------------------------------------------------------- layout
cols, col = [], 4
for p in T.PARAMS:
    p["start"] = col
    subs = T.GROUPS[p["n"]]
    if len(subs) > 1:
        p["subs"] = subs
        for no in subs:
            cols.append((col, p, no)); col += 1
    else:
        p["subs"] = None
        cols.append((col, p, "result")); col += 1
        cols.append((col, p, "ecoa")); col += 1
    cols.append((col, p, "check")); col += 1
    p["end"] = col - 1
    p["width"] = p["end"] - p["start"] + 1
LAST = col - 1

batches = apply_credit_corrections(batches)
print(f"credit corrections applied: {len(corrections)} "
      f"({collections.Counter(c[7] for c in corrections)})")

wb = openpyxl.load_workbook(SRC)
if SHEET in wb.sheetnames:
    wb.remove(wb[SHEET])
flat = wb["CoQ Parameter Tracker"]
flat.title = "CoQ Parameter Tracker (flat)"
ws = wb.create_sheet(SHEET, wb.sheetnames.index("Batch Coverage") + 1)

# ---- header rows 1–4
put(ws, 1, 1, "BATCH IDENTIFICATION", FW, NAVY)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
g, gstart = None, 0
for i, p in enumerate(T.PARAMS):
    if p["group"] != g:
        if g:
            ws.merge_cells(start_row=1, start_column=gstart, end_row=1, end_column=p["start"] - 1)
            put(ws, 1, gstart, g, FW, NAVY)
        g, gstart = p["group"], p["start"]
    if i == len(T.PARAMS) - 1:
        ws.merge_cells(start_row=1, start_column=gstart, end_row=1, end_column=p["end"])
        put(ws, 1, gstart, g, FW, NAVY)
for c, t in ((1, "CU Batch"), (2, "P Batch"), (3, "STATUS")):
    ws.merge_cells(start_row=2, start_column=c, end_row=4, end_column=c)
    put(ws, 2, c, t, FWS, NAVY)
for p in T.PARAMS:
    ws.merge_cells(start_row=2, start_column=p["start"], end_row=2, end_column=p["end"])
    put(ws, 2, p["start"], p["title"] + "\n" + p["method"], FW, NAVY)
    if p["subs"]:
        for i, no in enumerate(p["subs"]):
            put(ws, 3, p["start"] + i, "A.C.: " + T.CRIT[no], F6I, GREY)
            put(ws, 4, p["start"] + i, T.SUB[no], FSUB, SUBHDR)
        put(ws, 3, p["end"], "", F6I, GREY)
    else:
        put(ws, 3, p["start"], "A.C.: " + T.CRIT[str(p["n"])], F6I, GREY)
        put(ws, 3, p["start"] + 1, "", F6I, GREY)
        ws.merge_cells(start_row=3, start_column=p["start"], end_row=3, end_column=p["start"] + 1)
        put(ws, 4, p["start"], "Result (as reported)", FSUB, SUBHDR)
        put(ws, 4, p["start"] + 1, "eCOA ref, (date) [Lab] — one certificate per line", FSUB, SUBHDR)
    put(ws, 3, p["end"], "", F6I, GREY)
    put(ws, 4, p["end"], "✓/✗", FSUB, SUBHDR)
for h, height in ((1, 16), (2, 30), (3, 26), (4, 24)):
    ws.row_dimensions[h].height = height

# ---- body: one block of two rows per testing instance, batches boxed together
row = 5
stats = collections.Counter()
oos_rows = []
audit = []
SPAN = {}
for _cu, _pb, _code, _date, _lab, _pno in NEW_NONCONF:
    audit.append((_cu, _pb, _code, _date, _lab, _pno,
                  next(p["title"] for p in T.PARAMS if p["n"] == _pno), "non-conformance reported"))
for _cu, _pb, _code, _date, _lab, _pno in NEW_HELD:
    audit.append((_cu, _pb, _code, _date, _lab, _pno,
                  next(p["title"] for p in T.PARAMS if p["n"] == _pno), "held for review"))
BLANK = "FFFFFF"
for b in batches:
    docs = {p["n"]: instances(b, p["n"]) for p in T.PARAMS}
    K = max([len(v) for v in docs.values()] + [1])
    extra = len({(x[0], x[1]) for v in docs.values() for x in v if not x[3]})
    first, last = row, row + 2 * K - 1
    SPAN[id(b)] = (first, last)

    for cidx, v, font in ((1, b["cu"], FWS), (2, b["p"], FWS), (3, None, FWS)):
        if cidx != 3:
            ws.merge_cells(start_row=first, start_column=cidx, end_row=last, end_column=cidx)
            put(ws, first, cidx, v, font, NAVY)
            fill_range(ws, first, cidx, last, cidx, NAVY)

    no_cert = cert_no_result = missing = 0
    oos_list, und_list, stab_list = [], [], []

    def verdict(det_no, release_vals, stability_vals, label):
        """Judge release results only; a stability exceedance is reported apart."""
        bad = any(T.over_limit(det_no, x) for x in release_vals)
        und = (not bad) and any(T.undetermined(det_no, x) for x in release_vals)
        if bad and label not in oos_list:
            oos_list.append(label)
        elif und and label not in und_list:
            und_list.append(label)
        if any(T.over_limit(det_no, x) for x in stability_vals) and label not in stab_list:
            stab_list.append(label)
        return F7R if bad else (F7U if und else F7B)

    # the batch-level state of each parameter, from all of its certificates together
    pstate = {}
    for p in T.PARAMS:
        rel = stab = credited = 0
        for code, date, lab, is_cred in docs[p["n"]]:
            if not is_cred:
                continue
            vals = values_of(code, lab, b["cu"], scope_of(b, code))
            if any(vals.get(no) for no in T.GROUPS[p["n"]]):
                if T.nkey(code) in STAB:
                    stab += 1
                else:
                    rel += 1
            else:
                credited += 1
        if rel:
            st = "green"
        elif stab:
            st = "orange"
        elif credited:
            st = "amber"; cert_no_result += 1
        else:
            st = "red"; no_cert += 1
        if st != "green":
            missing += 1
        stats[st] += 1
        pstate[p["n"]] = st
        # the verdict is the batch's, judged over every certificate it holds
        for no in T.GROUPS[p["n"]]:
            relv, stabv = [], []
            for code, date, lab, is_cred in docs[p["n"]]:
                if not is_cred:
                    continue
                v = values_of(code, lab, b["cu"], scope_of(b, code)).get(no)
                if v:
                    (stabv if T.nkey(code) in STAB else relv).append(v)
            verdict(no, relv, stabv, f"#{p['n']}" + (" " + T.SUB[no] if len(T.GROUPS[p['n']]) > 1 else ""))

    for i in range(K):
        top, bot = first + 2 * i, first + 2 * i + 1
        top_lines = bot_lines = 1
        for p in T.PARAMS:
            dlist = docs[p["n"]]
            here = dlist[i] if i < len(dlist) else None
            credited_here = [x for x in dlist if x[3]]
            if here is None:
                state = "red" if (not credited_here and i == 0) else "none"
            else:
                code, date, lab, is_cred = here
                vals = values_of(code, lab, b["cu"], scope_of(b, code))
                has = any(vals.get(no) for no in T.GROUPS[p["n"]])
                if not is_cred:
                    state = "extra"
                else:
                    state = ("orange" if T.nkey(code) in STAB else "green") if has else "amber"
                    if not has:
                        audit.append((b["cu"], b["p"], code, date, lab, p["n"], p["title"],
                                      silence_reason(code, lab, b["cu"], T.GROUPS[p["n"]][0])))
            if state == "none":
                continue                       # an empty later block: no cells, no styles
            fill = FILL[state] if state in FILL else BLANK
            glyph = {"green": "✓", "orange": "✓", "amber": "✗", "red": "✗", "extra": "•"}.get(state, "")
            gfill = GLYPHFILL[state] if state in GLYPHFILL else BLANK
            ref = (f"{code}, ({date}) [{lab}]" + ("" if here[3] else " · on file, not credited")) if here \
                else ("— no certificate —" if state == "red" else "")
            ref_disp = ref
            if here and ICOA_RULE and T.nkey(code) in INST_KEY:
                ref = ("=IFERROR(INDEX('iCoA Register'!$B:$B,MATCH(\"%s\",'iCoA Register'!$R:$R,0)),\"iCoA — at issue\")&\", (%s) [PP]\""
                       % (INST_KEY[T.nkey(code)], date))

            if p["subs"]:
                for j, no in enumerate(p["subs"]):
                    if here:
                        v = values_of(here[0], here[2], b["cu"], scope_of(b, here[0])).get(no)
                        blank = not any(values_of(here[0], here[2], b["cu"], scope_of(b, here[0])).get(x)
                                        for x in p["subs"])
                        cell_v = v or (silence_reason(here[0], here[2], b["cu"], no) if blank and j == 0
                                       else ("" if blank else "n.r."))
                        font = F7R if T.over_limit(no, v or "") else (F7U if T.undetermined(no, v or "") else F7B)
                    else:
                        cell_v = "— MISSING —" if state == "red" else ""
                        font = FBAD if state == "red" else F7
                    put(ws, top, p["start"] + j, cell_v, font, fill)
                    top_lines = max(top_lines, nlines(cell_v, 8.5))
                ws.merge_cells(start_row=bot, start_column=p["start"], end_row=bot, end_column=p["end"] - 1)
                put(ws, bot, p["start"], ref, F6, fill, TOPC)
                fill_range(ws, bot, p["start"], bot, p["end"] - 1, fill)
                bot_lines = max(bot_lines, nlines(ref_disp, 8.6 * (len(p["subs"]) - 1) + 21))
            else:
                no = str(p["n"])
                if here:
                    v = values_of(here[0], here[2], b["cu"], scope_of(b, here[0])).get(no)
                    cell_v = v or silence_reason(here[0], here[2], b["cu"], no)
                    if ICOA_RULE and p["n"] == 3 and str(cell_v).startswith("Conforms"):
                        cell_v = str(cell_v).replace("Conforms", "Conforms (ImB spec.)", 1)
                    font = F7R if T.over_limit(no, v or "") else (F7U if T.undetermined(no, v or "") else F7B)
                else:
                    cell_v = "— MISSING —" if state == "red" else ""
                    font = FBAD if state == "red" else F7
                ws.merge_cells(start_row=top, start_column=p["start"], end_row=bot, end_column=p["start"])
                put(ws, top, p["start"], cell_v, font, fill)
                ws.merge_cells(start_row=top, start_column=p["start"] + 1, end_row=bot, end_column=p["start"] + 1)
                put(ws, top, p["start"] + 1, ref, F6, fill, TOPC)
                fill_range(ws, top, p["start"], bot, p["start"] + 1, fill)
                need = max(nlines(cell_v, 10), nlines(ref_disp, 21))
                if need > top_lines + bot_lines:
                    bot_lines = need - top_lines

            ws.merge_cells(start_row=top, start_column=p["end"], end_row=bot, end_column=p["end"])
            put(ws, top, p["end"], glyph, FWS, gfill, CEN)
            fill_range(ws, top, p["end"], bot, p["end"], gfill)

        ws.row_dimensions[top].height = max(14, 9.0 * top_lines + 3)
        ws.row_dimensions[bot].height = max(13, 9.0 * bot_lines + 3)
        if i:                                   # a hairline between testing instances
            outline(ws, top, 1, bot, LAST, MED)

    if missing == 0:
        st, colour = "✓ COMPLETE", STATUSFILL["green"]
    else:
        glyph = "⚠" if missing <= STATUS_PARTIAL_MAX else "✗"
        st = f"{glyph} {missing} NO RESULT\n({no_cert} no cert / {cert_no_result} cert w/o result)"
        colour = STATUSFILL["orange"] if missing <= STATUS_PARTIAL_MAX else STATUSFILL["red"]
    if K > 1:
        st += f"\n{K} testing instances"
    if extra:
        st += f"\n• {extra} document(s) on file, not credited"
        stats["uncredited instances"] += extra
    if oos_list:
        st += "\n✗ OUT OF SPECIFICATION: " + ", ".join(oos_list)
        colour = STATUSFILL["red"]
        oos_rows.append((b["cu"], "OOS", oos_list))
    if und_list:
        st += "\n◐ UNDETERMINED (Ph. Eur. band): " + ", ".join(und_list)
        if not oos_list:
            colour = STATUSFILL["orange"]
        oos_rows.append((b["cu"], "UNDETERMINED", und_list))
    if stab_list:
        st += "\n· stability above A.C.: " + ", ".join(stab_list)
        oos_rows.append((b["cu"], "stability", stab_list))
    ws.merge_cells(start_row=first, start_column=3, end_row=last, end_column=3)
    put(ws, first, 3, st, FWS, colour)
    fill_range(ws, first, 3, last, 3, colour)

    outline(ws, first, 1, last, LAST, THICK)
    for p in T.PARAMS:
        outline(ws, first, p["start"], last, p["end"], MED)
    row = last + 1

LASTROW = row - 1

# ---- key
key = ("KEY — ✓ green: certificate on file AND its result on the desk (release or re-test). "
       "✓ orange: stability-timepoint certificate — the result is NOT a release result. "
       "✗ amber: the certificate is credited for this parameter but the desk holds no result from it. "
       "✗ red — MISSING —: no certificate covers this parameter for this batch. "
       "BLOCK RULE: one TESTING INSTANCE = one block of two rows — result(s) on the top row, the certificate that reports them on "
       "the bottom row. A batch holds as many blocks as it has testing instances, and a parameter's certificates are taken in "
       "ascending date order, so the n-th block is the n-th round of testing; a parameter tested once has an empty cell in the "
       "later blocks. For #9, #10 and #11 each sub-determination has its own column on the top row. n.r. = that sub-determination "
       "is not reported on that certificate; \"no result on file\" = the certificate is credited here but the desk holds no result "
       "from it. RED BOLD result = OUT OF SPECIFICATION against the criterion in row 3; AMBER BOLD result = UNDETERMINED, in the "
       "Ph. Eur. band between a printed count limit and twice it. The check follows the Quality Desk exactly: a counted "
       "microbiological limit printed as ≤ 10ⁿ CFU/g is judged against 2 × 10ⁿ (Ph. Eur. 5.1.4); ND, <LOQ, <10, absent, a range "
       "written with \"and\", and any prose annotation are never judged; only release results are judged, and a stability "
       "timepoint above the criterion is named separately in STATUS. Every out-of-specification result needs an investigation "
       "record, not just a red cell. Acceptance criteria in row 3 are the global criteria of the CoQ parameter "
       f"schedule. v7 — built {BUILT} from the desk's record and the owner's certificate credits.")
krow = LASTROW + 2
ws.merge_cells(start_row=krow, start_column=1, end_row=krow, end_column=LAST)
put(ws, krow, 1, key, F6I, GREY, Alignment(horizontal="left", vertical="top", wrap_text=True))
ws.row_dimensions[krow].height = 62

# ---- widths, panes, print
for c, w in ((1, 13), (2, 15), (3, 20)):
    ws.column_dimensions[L(c)].width = w
for c, p, kindc in cols:
    if kindc == "check":
        w = 3.4
    elif kindc == "result":
        w = 10
    elif kindc == "ecoa":
        w = 21
    else:
        w = 8 if kindc in ("9.4", "9.5") else 8.6
    ws.column_dimensions[L(c)].width = w
ws.freeze_panes = "D5"
ws.print_title_rows = "1:4"
ws.page_setup.orientation = "landscape"
ws.page_setup.paperSize = ws.PAPERSIZE_A3
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = ws.page_margins.right = 0.3
ws.page_margins.top = ws.page_margins.bottom = 0.4

# --------------------------------------------------------------------------- Credit Audit
# Every certificate the owner's tracker credits to a parameter that it does not report.
# Two different problems, and they need different action:
#   not on this certificate — the document was read and carries no such row; the credit
#                             belongs on the certificate that does report it;
#   not ingested            — the desk holds no value from this document at all; it has
#                             never been read into the corpus, so nothing can be credited
#                             until it is.
aud = wb.create_sheet("Credit Audit", wb.sheetnames.index(SHEET) + 1)
acols = [("CU Batch", 14), ("P Batch", 14), ("Certificate", 24), ("Date", 11), ("Lab", 8),
         ("#", 5), ("Parameter", 30), ("Finding", 22), ("Action", 52)]
for _i, (_t, _w) in enumerate(acols, 1):
    put(aud, 1, _i, _t, FW, NAVY, CEN)
    aud.column_dimensions[L(_i)].width = _w
aud.row_dimensions[1].height = 22
ACTION = {
    "not on this certificate": "Move the credit to the certificate that reports this parameter, or record why it stands.",
    "not ingested": "Re-extract the document into the corpus; the batch cannot reach a CoQ on this parameter until then.",
    "held for review": "The two independent reads disagreed. A person must confirm the figure from the page.",
    "n.r.": "Not reported on this certificate.",
}
_r = 2
for cu, pb, code, date, lab, pno, title, why in sorted(audit, key=lambda x: (x[7], x[0], x[5])):
    put(aud, _r, 1, cu, F7B, IDF if False else None, CEN)
    put(aud, _r, 2, pb, F7, None, CEN)
    put(aud, _r, 3, code, F7, None, CEN)
    put(aud, _r, 4, T.as_date(date), F7, None, CEN)
    aud.cell(_r, 4).number_format = "DD.MM.YYYY"
    put(aud, _r, 5, lab, F7, None, CEN)
    put(aud, _r, 6, f"#{pno}", F7, None, CEN)
    put(aud, _r, 7, title, F7, None, Alignment(horizontal="left", vertical="center", wrap_text=True))
    put(aud, _r, 8, why, F7B, FILL["amber"] if why in ("not on this certificate", "n.r.") else FILL["red"], CEN)
    put(aud, _r, 9, ACTION.get(why, ""), F6I, None, Alignment(horizontal="left", vertical="center", wrap_text=True))
    _r += 1
aud.auto_filter.ref = f"A1:{L(len(acols))}{_r - 1}"
aud.freeze_panes = "A2"
aud.print_title_rows = "1:1"
aud.page_setup.orientation = "landscape"
aud.page_setup.fitToWidth = 1
aud.page_setup.fitToHeight = 0
aud.sheet_properties.pageSetUpPr.fitToPage = True
print("credit audit rows:", _r - 2,
      dict(collections.Counter(x[7] for x in audit)))

# --------------------------------------------------------------------------- Credit Corrections
cor = wb.create_sheet("Credit Corrections", wb.sheetnames.index("Credit Audit") + 1)
ccols = [("CU Batch", 14), ("P Batch", 14), ("Certificate", 24), ("Date", 11), ("Lab", 8),
         ("#", 5), ("Parameter", 30), ("Correction", 24), ("Evidence", 58)]
for _i, (_t, _w) in enumerate(ccols, 1):
    put(cor, 1, _i, _t, FW, NAVY, CEN)
    cor.column_dimensions[L(_i)].width = _w
cor.row_dimensions[1].height = 22
_r = 2
for cu, pb, code, date, lab, pno, title, rule, why in sorted(corrections, key=lambda x: (x[7], x[0], x[5])):
    put(cor, _r, 1, cu, F7B, None, CEN); put(cor, _r, 2, pb, F7, None, CEN)
    put(cor, _r, 3, code, F7, None, CEN)
    put(cor, _r, 4, T.as_date(date), F7, None, CEN); cor.cell(_r, 4).number_format = "DD.MM.YYYY"
    put(cor, _r, 5, lab, F7, None, CEN); put(cor, _r, 6, f"#{pno}", F7, None, CEN)
    put(cor, _r, 7, title, F7, None, Alignment(horizontal="left", vertical="center", wrap_text=True))
    put(cor, _r, 8, "credit removed", F7B, FILL["amber"], CEN)
    put(cor, _r, 9, f"{rule} — {why}", F6I, None, Alignment(horizontal="left", vertical="center", wrap_text=True))
    _r += 1
cor.auto_filter.ref = f"A1:{L(len(ccols))}{_r - 1}"
cor.freeze_panes = "A2"; cor.print_title_rows = "1:1"
cor.page_setup.orientation = "landscape"; cor.page_setup.fitToWidth = 1; cor.page_setup.fitToHeight = 0
cor.sheet_properties.pageSetUpPr.fitToPage = True
note = ("These corrections are applied when the workbook is built and are NOT written back to the owner's tracker. "
        "A removed credit does not remove the document: it still appears as a testing instance, marked • and "
        "\"on file, not credited\". R1 — the Farmahem pair was credited jointly for #3–#6 and #8, while the K "
        "certificate reports identification C, THC, CBD and CBN and the LoD certificate reports loss on drying alone; "
        "each now keeps only what it reports. R2 — identification B was credited to CNP certificates carrying no "
        "microscopy row; the credit is removed there and kept on ППК26110–26119 and ППК26127–26128, whose newer report "
        "format does carry it. Removing R2 leaves 51 lots with no evidence for identification B at all: those lots need "
        "an in-house iCoA for identification A and B, which is what the issuance plan already foresees.")
cor.merge_cells(start_row=_r + 1, start_column=1, end_row=_r + 1, end_column=len(ccols))
put(cor, _r + 1, 1, note, F6I, GREY, Alignment(horizontal="left", vertical="top", wrap_text=True))
cor.row_dimensions[_r + 1].height = 58
print("credit corrections sheet rows:", _r - 2)

# --------------------------------------------------------------------------- Work Order
# What no rebuild can fix: documents the database holds no read of, and figures the two
# independent reads disagreed on. Each row is a task for the ingestion queue or for a person.
wo = wb.create_sheet("Work Order", wb.sheetnames.index("Credit Corrections") + 1)
wcols = [("Task", 22), ("CU Batch", 14), ("P Batch", 14), ("Certificate", 24), ("Date", 11),
         ("Lab", 8), ("Parameters affected", 26), ("What is needed", 62)]
for _i, (_t, _w) in enumerate(wcols, 1):
    put(wo, 1, _i, _t, FW, NAVY, CEN)
    wo.column_dimensions[L(_i)].width = _w
wo.row_dimensions[1].height = 22
tasks = collections.defaultdict(lambda: {"params": set(), "meta": None})
for cu, pb, code, date, lab, pno, title, why in audit:
    if why not in ("not ingested", "held for review", "non-conformance reported"):
        continue
    k = (why, cu, code)
    tasks[k]["params"].add(pno)
    tasks[k]["meta"] = (pb, date, lab)
for _b in NEW_LOTS:
    for _c, _d, _l in _b["docs"][9]:
        _k = ("lot not on tracker", _b["p"], _c)
        tasks[_k]["params"].add(9)
        tasks[_k]["meta"] = (_b["p"], _d, _l)
NEED = {
    "not ingested": "Re-extract the document into the eCoA database (two independent reads, 300 DPI). "
                    "Until then the lot cannot reach a CoQ on these parameters.",
    "held for review": "The two independent reads disagreed. A person must read the figure from the page and confirm it.",
    "non-conformance reported": "The laboratory reports the result as not conforming (Не одговара). A deviation / OOS "
                                "record is needed before this lot's CoQ can issue.",
    "lot not on tracker": "The certificate names a P batch the owner's tracker does not carry. Record the lot "
                          "(cultivation batch code, strain) on the tracker so the certificate is credited under its batch.",
}
_r = 2
for (why, cu, code), d in sorted(tasks.items(), key=lambda x: (x[0][0], x[0][1])):
    pb, date, lab = d["meta"]
    put(wo, _r, 1, why, F7B, FILL["red"] if why == "not ingested" else FILL["amber"], CEN)
    put(wo, _r, 2, cu, F7B, None, CEN); put(wo, _r, 3, pb, F7, None, CEN)
    put(wo, _r, 4, code, F7, None, CEN)
    put(wo, _r, 5, T.as_date(date), F7, None, CEN); wo.cell(_r, 5).number_format = "DD.MM.YYYY"
    put(wo, _r, 6, lab, F7, None, CEN)
    put(wo, _r, 7, ", ".join(f"#{n}" for n in sorted(d["params"])), F7, None, CEN)
    put(wo, _r, 8, NEED[why], F6I, None, Alignment(horizontal="left", vertical="center", wrap_text=True))
    _r += 1
wo.auto_filter.ref = f"A1:{L(len(wcols))}{max(_r - 1, 2)}"
wo.freeze_panes = "A2"; wo.print_title_rows = "1:1"
wo.page_setup.orientation = "landscape"; wo.page_setup.fitToWidth = 1; wo.page_setup.fitToHeight = 0
wo.sheet_properties.pageSetUpPr.fitToPage = True
print("work order rows:", _r - 2)

# --------------------------------------------------------------------------- index: values + batch key
ix = wb["eCOA Document Index"]
hdr = [str(c.value or "") for c in ix[1]]
cv, ck, cc = len(hdr) + 1, len(hdr) + 2, len(hdr) + 3
put(ix, 1, cv, "PARAMETER VALUES", FW, NAVY)
put(ix, 1, ck, "BATCH KEY", FW, NAVY)
put(ix, 1, cc, "CREDITED FOR", FW, NAVY)
ix.column_dimensions[L(cv)].width = 60
ix.column_dimensions[L(ck)].width = 14
ix.column_dimensions[L(cc)].width = 22

# which parameters the owner's tracker credits each document with, per lot
CREDITED = collections.defaultdict(set)
for _b in batches:
    for _p in T.PARAMS:
        for _c, _d, _l in _b["docs"][_p["n"]]:
            CREDITED[(join_key(_b), T.nkey(_c))].add(_p["n"])
for r in range(2, ix.max_row + 1):
    code = str(ix.cell(r, 6).value or "").strip()
    cu = str(ix.cell(r, 2).value or "").strip()
    lab = str(ix.cell(r, 3).value or "").strip()
    if not code:
        continue
    pb = str(ix.cell(r, 1).value or "").strip()
    vals = values_of(code, lab, cu, COVERS.get((pb or re.sub(r"[＊*]", "", cu), T.nkey(code))))
    parts = []
    for p in T.PARAMS:
        subs = T.GROUPS[p["n"]]
        got = [(T.SUB[no] + " " + vals[no]) if len(subs) > 1 else vals[no] for no in subs if vals.get(no)]
        if got:
            parts.append(f"#{p['n']} " + " · ".join(got))
    stab = " · stability timepoint" if T.nkey(code) in STAB else ""
    cell = ix.cell(r, cv, (" · ".join(parts) + stab) if parts else "no result on the desk for this document")
    cell.font, cell.alignment, cell.border = F7, Alignment(vertical="top", wrap_text=True), BOX
    key = pb or re.sub(r"[＊*]", "", cu)
    cell = ix.cell(r, ck, key)
    cell.font, cell.alignment, cell.border = F7, CEN, BOX
    cred = sorted(CREDITED.get((key, T.nkey(code)), set()))
    cell = ix.cell(r, cc, ", ".join(f"#{n}" for n in cred) if cred else "— not credited —")
    cell.font, cell.alignment, cell.border = (F7 if cred else F6I), CEN, BOX
    if not cred:
        cell.fill = PatternFill("solid", fgColor=FILL["extra"])
ix.auto_filter.ref = f"A1:{L(cc)}{ix.max_row}"

# ---- Read Me: the v7 block rule
rm = wb["Read Me"]
r = rm.max_row + 2
for label, text in (("v7", "This workbook adds the sheet 'CoQ Parameter Tracker v7' — one batch per two-row block, "
                           "certificates stacked as lines in date order, sub-determinations in their own columns, acceptance "
                           "criteria in header row 3 and enforced, out-of-specification results printed red and named in STATUS. "
                           "The v6 flat table is kept as 'CoQ Parameter Tracker (flat)' for comparison."),
                    ("Index", "'eCOA Document Index' now also carries PARAMETER VALUES and BATCH KEY, so the tracker can be "
                              "rebuilt from the index alone (see CoQ_Tracker_v7_rebuild.gs)."),
                    ("OOS check", "A result is flagged only when it provably exceeds its criterion. ND, <LOQ, <10 and absent pass; "
                                  "'<10² and >10' is judged by its upper bound; a value that cannot be parsed is never flagged.")):
    c0 = rm.cell(r, 1, label); c0.font = Font(name="Calibri", size=9, bold=True); c0.alignment = Alignment(vertical="top")
    c1 = rm.cell(r, 2, text); c1.font = Font(name="Calibri", size=9); c1.alignment = Alignment(vertical="top", wrap_text=True)
    rm.row_dimensions[r].height = 13 * (len(text) // 118 + 1)
    r += 1

if V9:
    for _n in ("Results Register", "CoQ Parameter Tracker (flat)", "eCOA Document Index"):
        if _n in wb.sheetnames:
            wb.remove(wb[_n])
    _rm = wb["Read Me"]
    _r = _rm.max_row + 2
    for _label, _text in (
        (f"v{VER}" if VER == "9" else "v9", "The v8 build, verified and slimmed to live in Drive. Every decision-bearing value was checked against "
               "the filed certificate page (review/V8_TRUTH_CHECK_2026-09-02.md): the five out-of-specification TYMC "
               "results, the four undetermined ones and the stability CBN exceedance are real. The Results Register, "
               "the flat tracker and the eCOA Document Index are left out here; they stay in v8 in the repository."),
        ("Do not", "read counts from the eCOA_DB text layer. On five of six certificates checked it read the exponent "
                   "or a digit too low, every time in the direction of a conforming result."),
    ):
        _c = _rm.cell(_r, 1, _label); _c.font = Font(name="Calibri", size=9, bold=True); _c.alignment = Alignment(vertical="top")
        _c = _rm.cell(_r, 2, _text); _c.font = Font(name="Calibri", size=9); _c.alignment = Alignment(vertical="top", wrap_text=True)
        _rm.row_dimensions[_r].height = 13 * (len(_text) // 118 + 1); _r += 1
    wb["Read Me"]["A1"] = f"CoQ Analysis Master — v{VER}"

# --------------------------------------------------------------------------- new instances: the other sheets
from copy import copy as _copy


def _style_from(dst, src):
    dst.font, dst.fill, dst.border, dst.alignment, dst.number_format = \
        _copy(src.font), _copy(src.fill), _copy(src.border), _copy(src.alignment), src.number_format


def patch_coverage(wb):
    """Batch Coverage is inherited from v6: mark the newly covered parameters, recount the
    missing list, the certificate count and the laboratories, and add the lots the owner's
    tracker did not carry."""
    cov = wb["Batch Coverage"]
    last = max(r for r in range(2, cov.max_row + 1) if cov.cell(r, 1).value) if cov.max_row > 1 else 1
    short = {}
    for r in range(2, last + 1):
        for tok in str(cov.cell(r, 18).value or "").split(";"):
            m = re.match(r"\s*#(\d+)\s+(.+)", tok)
            if m:
                short[int(m.group(1))] = m.group(2).strip()
    for p in T.PARAMS:
        short.setdefault(p["n"], p["title"].split(" ", 1)[1])
    tick = next((cov.cell(r, c) for r in range(2, last + 1) for c in range(5, 17)
                 if cov.cell(r, c).value == "✓" and cov.cell(r, c).fill.fgColor.rgb not in (None, "00000000")), None)
    cross = next((cov.cell(r, c) for r in range(2, last + 1) for c in range(5, 17) if cov.cell(r, c).value == "✗"), None)
    status_style = {}
    for r in range(2, last + 1):
        st = str(cov.cell(r, 4).value or "")
        status_style.setdefault(st[:1], cov.cell(r, 4))

    def rowkey(cu, p):
        return (re.sub(r"[＊*]", "", cu).strip(), "— not assigned —" if p.startswith("N/A") else p.strip())

    rows = {rowkey(str(cov.cell(r, 1).value or ""), str(cov.cell(r, 2).value or "")): r for r in range(2, last + 1)}

    def recount(r):
        miss = [n for n in range(1, 13) if cov.cell(r, 4 + n).value == "✗"]
        cov.cell(r, 17).value = len(miss)
        cov.cell(r, 18).value = "; ".join(f"#{n} {short[n]}" for n in miss) or "—"
        st = "✓ COMPLETE" if not miss else (f"⚠ {len(miss)} MISSING" if len(miss) <= 3 else f"❌ {len(miss)} MISSING")
        cell = cov.cell(r, 4)
        cell.value = st
        if st[:1] in status_style:
            _style_from(cell, status_style[st[:1]])

    for b in NEW_TOUCHED:
        new_docs = sorted(set(ADDED.get(id(b), [])))
        if b.get("new_lot"):
            r = last + 1
            last = r
            src = next((rr for rr in range(2, r) if str(cov.cell(rr, 4).value or "").startswith("❌")), 2)
            for c in range(1, 21):
                _style_from(cov.cell(r, c), cov.cell(src, c))
            cov.cell(r, 1).value = b["cu"]
            cov.cell(r, 2).value = "— not assigned —" if b["p"].startswith("N/A") else b["p"]
            cov.cell(r, 3).value = b.get("strain", "")
            for n in range(1, 13):
                cell = cov.cell(r, 4 + n)
                cell.value = "✓" if b["docs"][n] else "✗"
                _style_from(cell, tick if b["docs"][n] else cross)
                if b["docs"][n] and all(T.kind_of(c, l, STAB) in ("In-house", "iCoA") for c, d, l in b["docs"][n]):
                    cell.fill = PatternFill("solid", fgColor="E7E6E6")
                    cell.font = Font(name="Calibri", size=9, color="595959")
            cov.cell(r, 19).value = len(new_docs)
            cov.cell(r, 20).value = f"[{new_docs[0][2]}] {len(new_docs)}" if new_docs else ""
            recount(r)
            continue
        r = rows.get(rowkey(b["cu"], b["p"]))
        if r is None:                       # the owner's row names the lot by its P batch only
            _kp = rowkey(b["cu"], b["p"])[1]
            r = next((rr for (kcu, kp), rr in rows.items() if kp == _kp and _kp != "— not assigned —"), None)
        if r is None:
            print("coverage: no row for", b["cu"], b["p"])
            continue
        for n in range(1, 13):
            if b["docs"][n] and cov.cell(r, 4 + n).value != "✓":
                cov.cell(r, 4 + n).value = "✓"
                _style_from(cov.cell(r, 4 + n), tick)
                if all(T.kind_of(c, l, STAB) in ("In-house", "iCoA") for c, d, l in b["docs"][n]):
                    cov.cell(r, 4 + n).fill = PatternFill("solid", fgColor="E7E6E6")
                    cov.cell(r, 4 + n).font = Font(name="Calibri", size=9, color="595959")
        cov.cell(r, 19).value = int(cov.cell(r, 19).value or 0) + len(new_docs)
        labs = collections.OrderedDict()
        for tok in str(cov.cell(r, 20).value or "").split(";"):
            m = re.match(r"\s*\[([^\]]+)\]\s*(\d+)", tok)
            if m:
                labs[m.group(1)] = int(m.group(2))
        for c, d, l in new_docs:
            labs[l] = labs.get(l, 0) + 1
        cov.cell(r, 20).value = "; ".join(f"[{k}] {v}" for k, v in labs.items())
        recount(r)
    if cov.auto_filter.ref:
        cov.auto_filter.ref = f"A1:{L(20)}{last}"
    return last


def patch_dashboard(wb, last):
    cov, dash = wb["Batch Coverage"], wb["Summary Dashboard"]
    n = last - 1
    marks = {p: [cov.cell(r, 4 + p).value for r in range(2, last + 1)] for p in range(1, 13)}
    missing = [sum(1 for p in range(1, 13) if cov.cell(r, 4 + p).value == "✗") for r in range(2, last + 1)]
    complete = sum(1 for m in missing if m == 0)
    partial = sum(1 for m in missing if 1 <= m <= 3)
    incomplete = sum(1 for m in missing if m >= 4)
    for r in range(1, dash.max_row + 1):
        label = str(dash.cell(r, 1).value or "")
        if label == "Total Batches":
            dash.cell(r, 2).value = n
        elif label == "Total eCOA Documents":
            dash.cell(r, 2).value = int(dash.cell(r, 2).value or 0) + len({x["code"] for x in NEW})
        elif label.startswith("✅"):
            dash.cell(r, 2).value = complete; dash.cell(r, 3).value = f"{round(100 * complete / n)}% of batches"
        elif label.startswith("⚠"):
            dash.cell(r, 2).value = partial; dash.cell(r, 3).value = f"{round(100 * partial / n)}% of batches"
        elif label.startswith("❌"):
            dash.cell(r, 2).value = incomplete; dash.cell(r, 3).value = f"{round(100 * incomplete / n)}% of batches"
        else:
            m = re.match(r"#(\d+)\s", label)
            if m:
                p = int(m.group(1))
                k = sum(1 for v in marks[p] if v == "✗")
                dash.cell(r, 2).value = k
                dash.cell(r, 3).value = f"({round(100 * k / n)}% of batches missing this)"


def add_mikro(wb, src_path):
    """Rebuild the owner's 'Mikro CoQ Parameter' sheet from the tracker: the same lots, the
    identity columns and the #7–#12 blocks, copied cell for cell from the rebuilt tracker."""
    src = openpyxl.load_workbook(src_path, read_only=True)
    if "Mikro CoQ Parameter" not in src.sheetnames:
        return 0
    want = []
    for row in src["Mikro CoQ Parameter"].iter_rows(min_row=1, max_row=400, values_only=True):
        cu, p = str(row[0] or "").strip(), str(row[1] or "").strip()
        if cu and cu not in ("BATCH IDENTIFICATION", "CU Batch #") and not cu.startswith("KEY"):
            want.append((cu, p))
    src.close()
    lots = []
    for cu, p in want:
        cu0 = re.sub(r"[＊*]", "", cu)
        for b in batches:
            bp = "/" if b["p"].startswith("N/A") else b["p"]
            if (cu0 == re.sub(r"[＊*]", "", b["cu"]) and (p in ("", "/", bp) or p == "N/A — no P batch assigned" and bp == "/")) \
               or (cu0 == "— not recorded —" and p == bp) or (cu0 == "— not recorded —" and p and p in b["p"]):
                if b not in lots:
                    lots.append(b)
                break
    if "Mikro CoQ Parameter" in wb.sheetnames:
        wb.remove(wb["Mikro CoQ Parameter"])
    dst = wb.create_sheet("Mikro CoQ Parameter", wb.sheetnames.index(SHEET) + 1)
    p7, p12 = next(p for p in T.PARAMS if p["n"] == 7), next(p for p in T.PARAMS if p["n"] == 12)
    cols = [1, 2, 3] + list(range(p7["start"], p12["end"] + 1))
    cmap = {c: i + 1 for i, c in enumerate(cols)}
    rows = [1, 2, 3, 4]
    for b in lots:
        if id(b) in SPAN:
            rows += list(range(SPAN[id(b)][0], SPAN[id(b)][1] + 1))
    key_row = next((r for r in range(ws.max_row, 4, -1) if str(ws.cell(r, 1).value or "").startswith("KEY")), None)
    if key_row:
        rows.append(key_row)                 # the legend travels with the sheet
    rmap = {r: i + 1 for i, r in enumerate(rows)}
    for r in rows:
        for c in cols:
            sc, dc = ws.cell(r, c), dst.cell(rmap[r], cmap[c])
            dc.value = sc.value
            _style_from(dc, sc)
        if ws.row_dimensions[r].height:
            dst.row_dimensions[rmap[r]].height = ws.row_dimensions[r].height
    for rng in ws.merged_cells.ranges:
        if rng.min_row in rmap and rng.max_row in rmap and rng.min_col in cmap:
            dst.merge_cells(start_row=rmap[rng.min_row], start_column=cmap[rng.min_col],
                            end_row=rmap[rng.max_row], end_column=cmap.get(rng.max_col, len(cols)))
    for c in cols:
        w = ws.column_dimensions[L(c)].width
        if w:
            dst.column_dimensions[L(cmap[c])].width = w
    dst.freeze_panes = "D5"
    dst.sheet_view.zoomScale = ws.sheet_view.zoomScale
    matched = {(re.sub(r"[＊*]", "", b["cu"]), b["p"]) for b in lots}
    print(f"Mikro CoQ Parameter: {len(lots)} of {len(want)} lot(s) matched, {len(rows) - 4} row(s)")
    for cu, p in want:
        if not any(re.sub(r"[＊*]", "", cu) == mcu or (cu == "— not recorded —" and p in mp) for mcu, mp in matched):
            print("   not matched:", repr(cu), repr(p))
    return len(lots)


def add_icoa_sheet(wb):
    sh = wb.create_sheet("iCoA Issuance", wb.sheetnames.index("Work Order") + 1)
    icols = [("Seq", 6), ("Series", 20), ("iCoA", 22), ("Plan reference (31.08.2026)", 22), ("CoQ", 30), ("Basis date", 12), ("Issue date (preliminary)", 24), ("CU Batch", 14),
             ("P Batch", 22), ("Strain", 20), ("Harvest", 24), ("Packaging", 28), ("Packaging complete", 18), ("iCoA scope", 30), ("Master", 34), ("#1 Ident. A", 14), ("#2 Ident. B", 14),
             ("#7 Foreign matter", 16), ("Ident C — covered by (eCoA)", 34),
             ("Retest assay #4–#6 (eCoA)", 34), ("Retest mycotoxins #10 (eCoA)", 34), ("Carried forward", 30), ("Status", 44)]
    keys = ("seq", "series", "icoa", "plan_ref", "coq", "basis", "issue", "cu", "p", "strain", "harvest", "packaging", "complete", "scope", "master", "a", "b", "fm", "c",
            "assay_rt", "myco_rt", "carry", "status")
    for _i, (_t, _w) in enumerate(icols, 1):
        put(sh, 1, _i, _t, FW, NAVY, CEN)
        sh.column_dimensions[L(_i)].width = _w
    sh.row_dimensions[1].height = 22
    _r = 2
    for row in ICOA_ROWS:
        for _i, k in enumerate(keys, 1):
            v = row[k]
            if k == "icoa" and row.get("key") and v != "not needed":
                v = REG_LOOKUP("B", row["key"], "— at issue —")           # the code follows the register
            if k == "issue" and row.get("issuable") == "yes":
                v = REG_LOOKUP("D", row["key"], "")                        # so does the preliminary date
            c = put(sh, _r, _i, v, F7B if k in ("icoa", "cu") else F7,
                FILL["amber"] if (k == "fm" and v == "held for review") or (k in ("c", "assay_rt", "myco_rt", "harvest", "packaging", "complete") and str(v).startswith("—"))
                or (k == "status" and str(v).startswith(("pending", "not yet"))) or (k == "icoa" and str(v).startswith("—")) else
                (FILL["green"] if k == "status" and str(v).startswith(("due", "registered")) else
                 (FILL["extra"] if k in ("a", "b", "fm") and str(v).startswith("CNP ") else None)), CEN)
            if k == "issue" and str(v).startswith("="):
                c.number_format = "DD.MM.YYYY"
        _r += 1
    note = ("Chronological issuance: one iCoA per batch, in the order of the release basis date its initial-release "
            "CoQ follows (the issuance plan of 31.08.2026; numbers iCoA-PP-YYYY-NNNN, one series per calendar year), "
            "then the RETEST SERIES: the same scope again for every batch, for the QP's retesting campaign (medical "
            "use, GACP product / API), dated at the retest sampling, in the order of the additional-testing CoQ each "
            "one belongs to. The reissued CoQ carries a NEW cannabinoid assay (#4–#6) and NEW mycotoxins (#10); #8, #9, "
            "#11 and #12 are carried forward from the initial testing; identification C is the new Farmahem K "
            "certificate. NUMBERING: a retest iCoA is a new document — a new number in the year of issue, never the "
            "initial iCoA's number (the plan of 31.08.2026 reused it; that is corrected here). 'due' where the retest "
            "certificates are on file, 'pending' where they are not. "
            "Where a CNP certificate reports identification A, B or foreign matter, its document code is the "
            "reference for them (grey cells) and the iCoA covers only the rest; where CNP reports all three, no "
            "iCoA is needed. Farmahem: identification C is the K (potency) certificate. "
            "Head of QC, 04.09.2026: identification A (appearance) and B (microscopy) are tested at Purely Plant "
            "together with foreign matter, at the date of packaging, and ONE iCoA per batch carries the three "
            "results for release. Numbers follow the issuance plan (iCoA-PP-YYYY-NNNN); a lot without a planned "
            "number takes its number at issue. Harvest and packaging dates are the Head of QC's list of 04.09.2026 "
            "(sheet Batch Dates): the iCoA is dated on the FIRST day of packaging, the day the issuance plan already "
            "uses as the CoQ basis, and a packaging that ran over several days shows its whole span; a lot the list "
            "does not carry keeps 'packaging date — to record'. ONE iCoA PER P LOT: a tracker row that holds "
            "several P lots (GRC102501, JD012603) has one row per P lot, each with its own number, CoQ and dates. "
            "PACKAGING COMPLETE is the last day of packaging — the earliest day the iCoA can be issued: the sample "
            "is taken before primary packaging (QCSOP 005), the certificate attests the complete packaged lot. "
            "Foreign matter is 'Conforms' by the declaration of 13.08.2026, except FB032601, where "
            "ППК26127 reports 0.08 % (Не одговара): held for the Head of QC. Identification C is not on the iCoA: "
            "it conforms to the ImB specification on the certificate that carries the cannabinoid assay (#4), "
            "named per batch in the last column.")
    sh.merge_cells(start_row=_r + 1, start_column=1, end_row=_r + 1, end_column=len(icols))
    put(sh, _r + 1, 1, note, F6I, GREY, Alignment(horizontal="left", vertical="top", wrap_text=True))
    sh.row_dimensions[_r + 1].height = 64
    sh.auto_filter.ref = f"A1:{L(len(icols))}{_r - 1}"
    sh.freeze_panes = "A2"
    print("iCoA issuance rows:", _r - 2)


REG_COLS = [("No.", 6), ("iCoA code", 18), ("Issuable", 9), ("Issue date (preliminary)", 20), ("Earliest permissible", 18),
            ("Basis — sampling at packaging", 16), ("Packaging complete", 16), ("Series", 20), ("CU Batch", 16), ("P Batch", 12),
            ("Strain", 20), ("iCoA scope", 22), ("Master", 34), ("CNP reference", 24), ("Ident C — eCoA", 34), ("CoQ", 28),
            ("Plan reference (31.08.2026)", 22), ("Key", 18), ("Status", 60)]
REG_SHEET = "'iCoA Register'"


def REG_LOOKUP(col, key, default):
    """A cell that follows the register: the column `col` of the row whose Key is `key`."""
    return f'=IFERROR(INDEX({REG_SHEET}!${col}:${col},MATCH("{key}",{REG_SHEET}!$R:$R,0)),"{default}")'


def _date(v):
    import datetime as _dt
    try:
        return _dt.datetime.strptime(str(v), "%d.%m.%Y").date()
    except ValueError:
        return None
REG_NOTE = ("Head of QC, 05.09.2026: preliminary iCoA issuance register. Document codes iCoA-PP_26-nnn (nnn = 001 … 999), "
            "one series for the year of issue, assigned in the order the certificates can be issued: by the earliest "
            "permissible issue date, then by the basis date (the first day of packaging, when the sample is taken). "
            "Every printed issue date lies in [11.05.2026 … the day of signing] — the CoQ SOP came into use on "
            "11.05.2026 and no document is post-dated — so the preliminary date is the SOP floor, or the last day of "
            "packaging where that is later; QC sets the real date at issue (documents may be issued in batches: one "
            "date, sequential numbers). No number is reserved for a row that cannot be issued yet: a lot without a "
            "packaging date, a held result, and every retest iCoA (identification A, B and foreign matter are tested "
            "at the retest sampling, whose date is not on the desk). Where a CNP certificate reports all three, no iCoA "
            "is needed. The plan's references of 31.08.2026 (iCoA-PP-YYYY-NNNN) are superseded by these codes and kept "
            "beside them. Identification C is not on the iCoA: the CoQ cites the cannabinoid-assay eCoA named here. "
            "TWO CERTIFICATES PER P LOT (05.09.2026): identification A and B on the P01-02 master, foreign matter on "
            "the P07 master. FORMULAS: No. counts the issuable rows above it, the code is built from No., the earliest "
            "and preliminary dates from the SOP floor and the packaging dates looked up on Batch Dates by P batch (or "
            "CU batch); insert a row, set Issuable to yes, and every code beneath moves by one — the iCoA Issuance "
            "sheet and the tracker cite the register by KEY, so they follow. Rows are not re-sorted by a formula: a "
            "changed date is a manual move.")


def _register_rows():
    """The register's static columns per row (the formula columns are written by _fill_register)."""
    out = []
    for r in REGISTER:
        cnp = " / ".join(f"{k}: {v[4:]}" for k, v in (("A", r["a"]), ("B", r["b"]), ("FM", r["fm"])) if str(v).startswith("CNP ")) or "—"
        out.append({"issuable": r["issuable"], "series": r["series"], "cu": r["cu"], "p": r["p"], "strain": r["strain"],
                    "scope": r["scope"], "master": r["master"], "cnp": cnp, "c": r["c"], "coq": r["coq"],
                    "plan_ref": r["plan_ref"] or "—", "key": r["key"], "status": r["reg_status"],
                    "code": r["code"], "issue_date": r["issue_date"]})
    return out


def _fill_register(sh):
    """The register as an Excel table whose number, code and dates are formulas:
         A No.       =IF(C="yes", COUNT(A$1:A[above])+1, "")     — counts the issuable rows above it
         B code      ="iCoA-PP_26-" & TEXT(A, "000")
         D issue     =E (the preliminary date; QC overtypes the real one)
         E earliest  =MAX(DATE(2026,5,11), G)                     — the SOP floor or the last day of packaging
         F, G        packaging first / last day, looked up on Batch Dates by P batch, else by CU batch
       Insert a row inside the table and set Issuable to yes: the table copies the formulas into it
       (Excel; Google Sheets and LibreOffice: fill the formulas down) and every code beneath moves by one."""
    from openpyxl.worksheet.table import Table, TableStyleInfo
    for _i, (_t, _w) in enumerate(REG_COLS, 1):
        put(sh, 1, _i, _t, FW, NAVY, CEN)
        sh.column_dimensions[L(_i)].width = _w
    sh.row_dimensions[1].height = 22
    BD_ = "'Batch Dates'"
    _r = 2
    for row in _register_rows():
        f_no = f'=IF(C{_r}="yes",COUNT(A$1:A{_r - 1})+1,"")'
        f_code = f'=IF(A{_r}="","— at issue —","iCoA-PP_26-"&TEXT(A{_r},"000"))'
        f_issue = f'=IF(A{_r}="","",E{_r})'
        f_earliest = f'=IF(AND(C{_r}="yes",ISNUMBER(G{_r})),MAX(DATE(2026,5,11),G{_r}),"")'
        f_from = (f'=IFERROR(INDEX({BD_}!$F:$F,MATCH(J{_r},{BD_}!$C:$C,0)),'
                  f'IFERROR(INDEX({BD_}!$F:$F,MATCH(I{_r},{BD_}!$B:$B,0)),""))')
        f_to = (f'=IFERROR(INDEX({BD_}!$G:$G,MATCH(J{_r},{BD_}!$C:$C,0)),'
                f'IFERROR(INDEX({BD_}!$G:$G,MATCH(I{_r},{BD_}!$B:$B,0)),""))')
        cells = (f_no, f_code, row["issuable"], f_issue, f_earliest, f_from, f_to, row["series"], row["cu"], row["p"],
                 row["strain"], row["scope"], row["master"], row["cnp"], row["c"], row["coq"], row["plan_ref"], row["key"], row["status"])
        for _i, v in enumerate(cells, 1):
            c = put(sh, _r, _i, v, F7B if _i in (2, 9) else F7,
                    FILL["green"] if (_i == 19 and str(v).startswith("registered")) or (_i == 3 and v == "yes") else
                    FILL["amber"] if (_i == 19 and str(v).startswith("not yet")) or (_i == 3 and v == "no")
                    or (_i in (14, 15, 17) and str(v).startswith("—")) else None,
                    CEN if _i != 19 else Alignment(horizontal="left", vertical="center", wrap_text=True))
            if _i in (4, 5, 6, 7):
                c.number_format = "DD.MM.YYYY"
        _r += 1
    tab = Table(displayName="iCoA_Register", ref=f"A1:{L(len(REG_COLS))}{_r - 1}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False, showLastColumn=False,
                                        showRowStripes=False, showColumnStripes=False)
    sh.add_table(tab)
    sh.merge_cells(start_row=_r + 1, start_column=1, end_row=_r + 1, end_column=len(REG_COLS))
    put(sh, _r + 1, 1, REG_NOTE, F6I, GREY, Alignment(horizontal="left", vertical="top", wrap_text=True))
    sh.row_dimensions[_r + 1].height = 96
    sh.freeze_panes = "C2"
    return _r - 2


def add_register_sheet(wb):
    sh = wb.create_sheet("iCoA Register", wb.sheetnames.index("iCoA Issuance") + 1)
    print("iCoA register rows:", _fill_register(sh))


def write_register_file(path):
    """The register on its own, for the person issuing the certificates (with the Batch Dates
    sheet its date formulas look up)."""
    w = openpyxl.Workbook()
    sh = w.active
    sh.title = "iCoA Register"
    _fill_register(sh)
    _fill_dates(w.create_sheet("Batch Dates"))
    rm = w.create_sheet("Read Me")
    rm.column_dimensions["A"].width = 120
    for _i, t in enumerate((f"iCoA Issuance Register — preliminary — built with CoQ Analysis Master v{VER} on 05.09.2026",
                            REG_NOTE,
                            "Source: CoQ_Analysis_Master_v" + VER + ".xlsx, sheets iCoA Issuance and Batch Dates; "
                            "dates from the Head of QC's list of 04.09.2026; the issuance plan of 31.08.2026 for the CoQ numbers."), 1):
        c = rm.cell(_i, 1, t)
        c.font = Font(name="Calibri", size=9, bold=(_i == 1))
        c.alignment = Alignment(vertical="top", wrap_text=True)
    rm.row_dimensions[2].height = 120
    w.save(path)
    print("saved", path)


def _fill_dates(sh):
    """The Head of QC's list of 04.09.2026, normalised (batch_dates.py), as DATE cells — the register's
    lookups read columns B (batch as listed), C (P batch), F (packaging from) and G (packaging to)."""
    cols = [("Seq", 6), ("Batch (as listed)", 18), ("P Batch", 12), ("Harvest from", 13), ("Harvest to", 13),
            ("Packaging from", 14), ("Packaging to", 14), ("iCoA basis (first day of packaging)", 20), ("Tracker lot", 16), ("Note", 60)]
    for _i, (_t, _w) in enumerate(cols, 1):
        put(sh, 1, _i, _t, FW, NAVY, CEN)
        sh.column_dimensions[L(_i)].width = _w
    sh.row_dimensions[1].height = 22
    _r = 2
    for r in DATE_ROWS:
        lot = DATE_USED.get(r["seq"])
        row = (int(r["seq"]), r["cu_batch"], r["p_batch"], _date(r["harvest_from"]) or "— not given —", _date(r["harvest_to"]) or "— not given —",
               _date(r["packaging_from"]) or "— not given —", _date(r["packaging_to"]) or "— not given —", f"=F{_r}",
               lot or "— not on the tracker —", r["note"])
        for _i, v in enumerate(row, 1):
            c = put(sh, _r, _i, v, F7B if _i in (2, 3) else F7,
                    FILL["amber"] if (_i in (4, 5, 6, 7, 9) and str(v).startswith("—")) else None,
                    CEN if _i != 10 else Alignment(horizontal="left", vertical="center", wrap_text=True))
            if _i in (4, 5, 6, 7, 8):
                c.number_format = "DD.MM.YYYY"
        _r += 1
    return _r, cols


def add_dates_sheet(wb):
    sh = wb.create_sheet("Batch Dates", wb.sheetnames.index("iCoA Issuance") + 1)
    _r, cols = _fill_dates(sh)
    note = ("Head of QC, 04.09.2026: date of harvest and packaging date per batch, as sent (batch_dates_raw_2026-09-04.tsv), "
            "normalised by batch_dates.py. A year the list does not print is the harvest year of the same row, else the "
            "year of the row above (the list is chronological); '11-13-11.2025' is read as 11-13.11.2025; '0' and ']' are "
            "no date. The iCoA for identification A, B and foreign matter is dated on the first day of packaging, the "
            "day the issuance plan of 31.08.2026 already uses as the CoQ basis. Rows without a tracker lot are batches "
            "the owner's tracker does not carry; tracker lots without a row keep 'packaging date — to record'. The "
            "iCoA Register looks its packaging dates up here by P batch (else by the batch as listed): a date "
            "corrected on this sheet moves the register's earliest and preliminary issue dates with it.")
    sh.merge_cells(start_row=_r + 1, start_column=1, end_row=_r + 1, end_column=len(cols))
    put(sh, _r + 1, 1, note, F6I, GREY, Alignment(horizontal="left", vertical="top", wrap_text=True))
    sh.row_dimensions[_r + 1].height = 52
    sh.auto_filter.ref = f"A1:{L(len(cols))}{_r - 1}"
    sh.freeze_panes = "A2"
    print("batch dates rows:", _r - 2)


if NEW:
    _last = patch_coverage(wb)
    patch_dashboard(wb, _last)
    _mikro = next((a.split("=", 1)[1] for a in sys.argv if a.startswith("--mikro=")), None)
    if _mikro and os.path.exists(_mikro):
        add_mikro(wb, _mikro)
    if ICOA_RULE:
        add_icoa_sheet(wb)
        add_register_sheet(wb)
        add_dates_sheet(wb)
        write_register_file(os.path.join(HERE, "iCoA_Issuance_Register_prelim.xlsx"))
    _rm = wb["Read Me"]
    _r = _rm.max_row + 2
    _held = ", ".join(sorted({f"{c} (#{n})" for cu, pb, c, d, l, n in NEW_HELD}))
    _lots = ", ".join(b["p"] for b in NEW_LOTS)
    _text = (f"v{VER} — {len(NEW)} certificates ingested into eCOA_DB on 04.09.2026 (IJZ-MB microbiology, issued 31.08 and "
             f"01.09.2026) are added as testing instances credited to #9, with the values the two independent reads "
             f"agreed on. Batch Coverage, the tracker, the Credit Audit, the Work Order and the Summary Dashboard are "
             f"recomputed with them."
             + (f" Held for a person's read: {_held}." if _held else "")
             + (f" Lots the owner's tracker does not carry, opened without a CU code: {_lots}." if _lots else "")
             + " The laboratory prints the zero of a P-number as a letter O (PO60052); it is folded to P060052 here."
             + (" Head of QC, 04.09.2026: identification A and B are tested at Purely Plant together with foreign "
                "matter at the date of packaging, and one iCoA per batch carries the three results — every lot "
                "now holds that in-house instance for #1, #2 and #7 (see iCoA Issuance); identification C "
                "conforms to the ImB specification on the certificate that carries the cannabinoid assay. "
                "Harvest and packaging dates per batch are the Head of QC's list of 04.09.2026 (sheet Batch Dates): "
                "the iCoA instance is dated on the first day of packaging (sampling before primary packaging), the "
                "day the issuance plan uses as the CoQ basis, and issued no earlier than the last day of packaging; "
                "one iCoA per P lot where a tracker row holds several. iCoA Register (05.09.2026): the preliminary "
                "issuance register — codes iCoA-PP_26-nnn in the order the certificates can be issued, preliminary "
                "issue dates at the earliest permissible day (the SOP floor 11.05.2026 or the last day of packaging), "
                "no number for a row that cannot be issued yet; the tracker's in-house instances cite these codes. Rulings of 04.09.2026 on the two bile-tolerant gram-negative rows the reads disagreed on: "
                "P060262 < 10³ и > 10² CFU/g, P060432 < 10² и > 10 CFU/g (decisions_2026-09-04.tsv)."
                if ICOA_RULE else ""))
    for _label, _t in ((f"v{VER}", _text),):
        _c = _rm.cell(_r, 1, _label); _c.font = Font(name="Calibri", size=9, bold=True); _c.alignment = Alignment(vertical="top")
        _c = _rm.cell(_r, 2, _t); _c.font = Font(name="Calibri", size=9); _c.alignment = Alignment(vertical="top", wrap_text=True)
        _rm.row_dimensions[_r].height = 13 * (len(_t) // 118 + 1); _r += 1
    if V9:
        wb["Read Me"]["A1"] = f"CoQ Analysis Master — v{VER}"
wb.save(OUT)
print("saved", OUT)
print("batches:", len(batches), "two-row blocks:", (LASTROW - 4) // 2, "rows:", LASTROW - 4, "columns:", LAST)
print("parameter cells by state:", dict(stats))
import collections as _c
by = _c.Counter(k for _, k, _ in oos_rows)
print("verdicts:", dict(by))
for cu, kind, lst in oos_rows:
    print(f"    {kind:14s} {cu:14s} → {', '.join(lst)}")
