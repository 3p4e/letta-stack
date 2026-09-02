#!/usr/bin/env python3
"""Shared reader for the CoQ tracker builders — no side effects on import.

Two sources, kept apart on purpose:

  the desk    ../coq_artifact_data.json — the release register (correction chain
              step 19), the page reads of 31.08.2026 and the 12-month re-analyses.
              Supplies every reported value, the laboratory and the stability flag.
  the owner   CoQ_Analysis_Master_v3.xlsx — which certificate is credited to which
              parameter, on the tracker sheet and on the eCOA Document Index.

Nothing here invents a value: a determination the desk does not hold is absent
from the mapping, and the caller decides how to show that.
"""
import collections, datetime, json, os, re, sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, "..", "..", ".."))
sys.path.insert(0, os.path.join(ROOT, "ingestion", "common"))
from batch_id import batch_key  # noqa: E402

SRC_V3 = os.path.join(HERE, "CoQ_Analysis_Master_v3.xlsx")
DESK = os.path.join(ROOT, "deliverables", "qc_gap_analysis", "coq_artifact_data.json")

CYR = str.maketrans("АВЕКМНОРСТУХЈЅІавекмнорстухјѕі", "ABEKMHOPCTYXJSIabekmhopctyxjsi")

# --------------------------------------------------------------------------- parameters
GROUPS = {1: ["1"], 2: ["2"], 3: ["3"], 4: ["4"], 5: ["5"], 6: ["6"], 7: ["7"], 8: ["8"],
          9: ["9.1", "9.2", "9.3", "9.4", "9.5"], 10: ["10.1", "10.2", "10.3"],
          11: ["11.1", "11.2", "11.3", "11.4"], 12: ["12"]}
SUB = {"9.1": "TAMC", "9.2": "TYMC", "9.3": "GNB", "9.4": "Salm.", "9.5": "E. coli",
       "10.1": "AfB₁", "10.2": "ΣAf", "10.3": "OTA",
       "11.1": "Pb", "11.2": "Cd", "11.3": "As", "11.4": "Hg"}
# Global acceptance criteria, transcribed from the CoQ parameter schedule.
CRIT = {"1": "Conforms to Ph. Eur. monograph 3028 (Cannabis flos)",
        "2": "Conforms to Ph. Eur. monograph 3028 (Cannabis flos)",
        "3": "Conforms to Ph. Eur. monograph 3028 (Cannabis flos)",
        "4": "Per target grade (CoQ §01)", "5": "≤ 1.0 % w/w", "6": "≤ 1.0 % w/w",
        "7": "≤ 2.0 % (25–50 g); leaves < 1 cm; no seeds", "8": "≤ 12.0 %",
        "9.1": "≤ 10⁵ CFU/g", "9.2": "≤ 10⁴ CFU/g", "9.3": "≤ 10⁴ CFU/g",
        "9.4": "Absence / 25 g", "9.5": "Absence / 1 g",
        "10.1": "≤ 2 µg/kg", "10.2": "≤ 4 µg/kg", "10.3": "≤ 20 µg/kg",
        "11.1": "≤ 0.5 mg/kg", "11.2": "≤ 0.3 mg/kg", "11.3": "≤ 0.2 mg/kg", "11.4": "≤ 0.1 mg/kg",
        "12": "≤ LOQ (Ph. Eur. 2.8.13 / CUMCS equivalency)"}
# Numeric upper limits for the automatic conformance check. A result is flagged only
# when it parses to a number above the limit; ND, <LOQ, <10, absent pass.
LIMIT = {"5": 1.0, "6": 1.0, "8": 12.0,
         "9.1": 1e5, "9.2": 1e4, "9.3": 1e4,
         "10.1": 2.0, "10.2": 4.0, "10.3": 20.0,
         "11.1": 0.5, "11.2": 0.3, "11.3": 0.2, "11.4": 0.1}
ABSENCE = {"9.4", "9.5"}
PARAMS = [
    dict(n=1, group="IDENTIFICATION  1–3", title="#1 Identification A", method="Appearance · Ph. Eur. mon. 3028"),
    dict(n=2, group="IDENTIFICATION  1–3", title="#2 Identification B", method="Microscopy · Ph. Eur. 2.8.23"),
    dict(n=3, group="IDENTIFICATION  1–3", title="#3 Identification C", method="HPLC · Ph. Eur. 2.2.29"),
    dict(n=4, group="CANNABINOID ASSAY  4–6", title="#4 Assay — Total Δ⁹-THC*", method="Ph. Eur. 2.2.29 (HPLC)"),
    dict(n=5, group="CANNABINOID ASSAY  4–6", title="#5 Assay — Total CBD", method="Ph. Eur. 2.2.29 · CBD + CBDA×0.877"),
    dict(n=6, group="CANNABINOID ASSAY  4–6", title="#6 Total CBN", method="Ph. Eur. 2.2.29 · CBN + CBNA×0.876"),
    dict(n=7, group="PHYSICAL  7–8", title="#7 Foreign Matter", method="Ph. Eur. 2.8.2 / in-house"),
    dict(n=8, group="PHYSICAL  7–8", title="#8 Loss on Drying", method="Ph. Eur. 2.2.32 · 40 °C, 24 h, 15–25 mbar"),
    dict(n=9, group="MICROBIOLOGY  9", title="#9 Microbiological Purity", method="Ph. Eur. 2.6.12 / 2.6.13 / 2.6.31 · cat. C"),
    dict(n=10, group="CONTAMINANTS  10–12", title="#10 Mycotoxins", method="Ph. Eur. 2.8.18 / 2.8.22 (HPLC-FLD)"),
    dict(n=11, group="CONTAMINANTS  10–12", title="#11 Heavy Metals", method="Ph. Eur. 2.4.27 (ICP-MS)"),
    dict(n=12, group="CONTAMINANTS  10–12", title="#12 Pesticide Residues", method="Ph. Eur. 2.8.13 (LC-MS/MS) · CUMCS equivalency"),
]
LABNAME = {"CNP": "UKIM Faculty of Pharmacy — Center for Natural Products", "FHM-K": "Farmahem — cannabinoids",
           "FHM-M": "Farmahem — mycotoxins", "IJZ": "Institute of Public Health",
           "IJZ-MB": "Institute of Public Health — microbiology", "NGP": "Purely Plant — in-house form",
           "PP": "Purely Plant — in-house", "DFL": "State Phytosanitary Laboratory"}


# --------------------------------------------------------------------------- keys and cleaning
def nkey(code):
    """Fold a certificate code to a comparison key (Cyrillic homoglyphs, separators, LoD)."""
    c = str(code or "").strip()
    c = re.sub(r"\s*\([^)]*\)\s*$", "", c)
    c = re.sub(r"\s+(EN|MK)$", "", c)
    c = re.sub(r"[-/](GS|ГС|LOD|LoD)[-/]", "/LOD/", c)
    c = c.translate(CYR)
    return re.sub(r"[\s\-/_.]+", "/", c).strip("/").upper()


def cu_key(cu):
    cu = re.sub(r"[＊*]", "", cu).replace("OMP1024", "OPM1024").replace("SJ0925021", "SJ092501").replace("_01/1", "_01")
    return batch_key(cu)


def clean(v):
    """The value as the certificate prints it, with the desk's own annotations removed."""
    v = str(v or "").strip()
    if not v or v in ("—", "/"):
        return ""
    if " | " in v:
        v = v.split(" | ")[0]
    v = re.sub(r"\s*[xх]\s*10\^?", "×10", v)
    v = re.sub(r"10\^(\d)", lambda m: "10" + "⁰¹²³⁴⁵⁶⁷⁸⁹"[int(m.group(1))], v)
    v = re.sub(r"([<>≤≥])\s+", r"\1", v)
    if re.search(r"одговара|отсутн|отсуств|absent", v, re.I):
        v = "absent"
    if re.match(r"^conforms", v, re.I):
        v = "Conforms"
    if re.match(r"^complies", v, re.I):
        v = "Complies"
    if re.match(r"^not found any pesticide", v, re.I):
        v = "≤LOQ"
    if re.match(r"^н\.\s*д\.", v, re.I):
        v = "N.D."
    if re.match(r"^BLQ\b", v):
        v = "BLQ"
    m = re.match(r"^(-?\d+(?:\.\d+)?)\s+—\s+DETECTED", v)
    if m:
        v = m.group(1)
    return v


def as_date(s):
    try:
        return datetime.datetime.strptime(s, "%d.%m.%Y").date()
    except Exception:
        return s


def date_key(s):
    m = re.match(r"(\d{2})\.(\d{2})\.(\d{4})", str(s or ""))
    return (m.group(3) + m.group(2) + m.group(1)) if m else "99999999"


def as_number(v):
    if isinstance(v, str) and re.fullmatch(r"-?\d+(\.\d+)?", v):
        dec = len(v.split(".")[1]) if "." in v else 0
        return (float(v) if dec else int(v)), ("0." + "0" * dec if dec else "0")
    return v, None


# --------------------------------------------------------------------------- conformance
# Ported verbatim in behaviour from the Quality Desk (live_instrument/script.js:
# magnitude, acceptanceLimit, isProse, overLimit, undetBand) so the workbook and the
# desk cannot disagree about what conforms.
#
# The rule that matters: a counted microbiological criterion printed as "≤ 10ⁿ CFU/g" is
# judged against 2 × 10ⁿ (Ph. Eur. 5.1.4 — the printed limit is a maximum acceptable
# count, and a result is only out of specification above twice it). Between the printed
# limit and twice it the result is UNDETERMINED, not failing.
PH_EUR_FACTOR = 2
COUNTED = re.compile(r"tamc|tymc|cfu|gnb|gram-neg|aerobic|yeast|mould|mold", re.I)
SUPS = {"⁰": "0", "¹": "1", "²": "2", "³": "3", "⁴": "4", "⁵": "5", "⁶": "6", "⁷": "7", "⁸": "8", "⁹": "9"}


def _sup(t):
    return re.sub(r"[⁰¹²³⁴⁵⁶⁷⁸⁹]", lambda m: "^" + SUPS[m.group(0)], str(t))


def magnitude(v):
    """The numeric magnitude of a reported value, or None when there is none."""
    s = str("" if v is None else v).strip()
    if not s or s in ("/", "—"):
        return None
    s = _sup(s).replace("х", "x").replace("Х", "x").replace("×", "x").replace("·", "x")
    s = s.replace("≤", "<=").replace("≥", ">=")
    s = re.sub(r"(\d),(\d)", r"\1.\2", s)
    m = re.search(r"(\d+(?:\.\d+)?)\s*x\s*10\s*\^?\s*(\d+)", s, re.I)
    if m:
        return float(m.group(1)) * 10 ** int(m.group(2))
    m = re.search(r"(?:^|[^\d.])10\s*\^\s*(\d+)", s)
    if m:
        return 10.0 ** int(m.group(1))
    m = re.search(r"(\d+(?:\.\d+)?)", s)
    if m and not re.search(r"and|и", s, re.I):
        return float(m.group(1))
    return None


def acceptance_limit(lim, name=""):
    """The number a result is actually judged against — twice the printed count limit."""
    s = str(lim or "").strip()
    if not s:
        return None
    norm = _sup(s).replace("х", "x").replace("Х", "x").replace("×", "x")
    norm = re.sub(r"\s+", "", norm)
    m = re.match(r"^[<≤]?10\^(\d)(?!\d)", norm)
    if m and COUNTED.search(str(name) + " " + s):
        return PH_EUR_FACTOR * 10 ** int(m.group(1))
    return magnitude(s)


def is_prose(s):
    """Four letters running means an annotation, and an annotation is never judged."""
    return bool(re.search(r"[A-Za-zА-Яа-я]{4}", str(s)))


def _judgeable(v, lim):
    s = str(v or "").strip()
    if not s or not lim:
        return None
    if re.match(r"^[<≤]", s) or re.search(r"and|и", s, re.I):
        return None
    if re.match(r"^(n\.?d\.?|blq|absent|одговара|н\.д)", s, re.I) or is_prose(s):
        return None
    return s


def over_limit(det_no, v):
    """True only when the value provably exceeds its acceptance criterion."""
    lim = CRIT.get(det_no)
    s = _judgeable(v, lim)
    if s is None:
        return False
    if not re.search(r"[<≤]", lim) and not re.search(r"max", lim, re.I):
        return False
    a, b = magnitude(s), acceptance_limit(lim, name_of(det_no))
    return a is not None and b is not None and a > b * 1.0000001


def undetermined(det_no, v):
    """True in the band between a printed count limit and twice it — not a failure."""
    lim = CRIT.get(det_no)
    s = _judgeable(v, lim)
    if s is None:
        return False
    norm = re.sub(r"\s+", "", _sup(lim).replace("×", "x"))
    m = re.match(r"^[<≤]?10\^(\d)(?!\d)", norm)
    if not m or not COUNTED.search(name_of(det_no) + " " + lim):
        return False
    printed = 10 ** int(m.group(1))
    a = magnitude(s)
    return a is not None and a > printed * 1.0000001 and a <= printed * PH_EUR_FACTOR * 1.0000001


def name_of(det_no):
    return SUB.get(det_no, next((p["title"] for p in PARAMS if str(p["n"]) == str(det_no).split(".")[0]), ""))


# --------------------------------------------------------------------------- the desk
def load_desk():
    """Returns (VAL, INH, STAB, STRAIN, DET): reported values keyed on certificate code."""
    D = json.load(open(DESK))
    col2det = {d["col"]: d["no"] for d in D["dets"] if d.get("col")}
    DET = {d["no"]: d for d in D["dets"]}
    VAL, INH, STAB, STRAIN = collections.defaultdict(dict), collections.defaultdict(dict), set(), {}
    for b in D["reg"]:
        STRAIN.setdefault(batch_key(b["cb"]), b["strain"])
        for ct in b["certs"]:
            k = nkey(ct["code"])
            if ct.get("stab"):
                STAB.add(k)
            for colL, v in ct["vals"].items():
                no, v = col2det.get(colL), clean(v)
                if no and v:
                    VAL[k].setdefault(no, v)
            if ct["code"].startswith("n/a"):
                for colL, v in ct["vals"].items():
                    no, v = col2det.get(colL), clean(v)
                    if no and v:
                        INH[batch_key(b["cb"])].setdefault(no, v)
    for c in D["coqs"]:
        cbk = batch_key(re.sub(r"[＊*]", "", c["cb"]))
        STRAIN.setdefault(cbk, c["strain"])
        for r in c["rows"]:
            v = clean(r["res"])
            if not v or not r["doc"] or r["doc"] == "—":
                continue
            if r["doc"].startswith("n/a"):
                INH[cbk][r["no"]] = v
            else:
                VAL[nkey(r["doc"])][r["no"]] = v
    return VAL, INH, STAB, STRAIN, DET


def kind_of(code, lab, stab):
    if str(code).startswith("iCoA-"):
        return "iCoA"
    if str(code).startswith("NO-DOC-CODE") or str(code).startswith("NGP-") or lab in ("PP", "NGP"):
        return "In-house"
    if nkey(code) in stab:
        return "Stability"
    return "eCoA"


# --------------------------------------------------------------------------- the owner's credits
ENTRY = re.compile(r"([^\n;]+?)\s*,\s*\((\d{2}\.\d{2}\.\d{4})\)\s*\[([^\]]+)\]")


def load_owner(path=SRC_V3):
    """The owner's tracker: per batch, the certificates credited to each parameter.

    Returns a list of {cu, p, status, labs, docs{param -> [(code, date, lab)]},
    certs[((code,date,lab), {params})]} in the owner's row order, and the index rows.
    """
    wb = openpyxl.load_workbook(path)
    old = wb["CoQ Parameter Tracker"]
    pcol = {int(m.group(1)): i for i, v in enumerate([c.value for c in old[3]], 1)
            for m in [re.match(r"#(\d+)", str(v or ""))] if m}
    batches = []
    for r in range(4, old.max_row + 1):
        cu = str(old.cell(r, 1).value or "").strip()
        if not cu:
            continue
        p = str(old.cell(r, 2).value or "").strip()
        rec = {"cu": "— not recorded —" if "NOT ASSIGNED" in cu else cu,
               "p": "N/A — no P batch assigned" if ("NOT ASSIGNED" in p or not p) else p,
               "status": str(old.cell(r, 3).value or "").strip(),
               "labs": [x.strip() for x in str(old.cell(r, 28).value or "").split("|") if x.strip()],
               "docs": {}}
        for pn, cidx in pcol.items():
            rec["docs"][pn] = [(m.group(1).strip(), m.group(2), m.group(3))
                               for m in ENTRY.finditer(str(old.cell(r, cidx).value or ""))]
        certs = {}
        for pn, docs in rec["docs"].items():
            for d in docs:
                certs.setdefault(d, set()).add(pn)
        rec["certs"] = sorted(certs.items(), key=lambda kv: (date_key(kv[0][1]), kv[0][0]))
        batches.append(rec)

    ix = wb["eCOA Document Index"]
    rows, p_last, cu_last = [], "", ""
    for rr in range(2, ix.max_row + 1):
        vals = [ix.cell(rr, c).value for c in range(1, 9)]
        if vals[0]:
            p_last = str(vals[0]).strip()
        if vals[1]:
            cu_last = str(vals[1]).strip()
        if not vals[5]:
            continue
        rows.append(dict(p=p_last, cu=cu_last, lab=str(vals[2] or "").strip(), params=str(vals[3] or "").strip(),
                         dtype=str(vals[4] or "").strip(), code=str(vals[5] or "").strip(),
                         date=str(vals[6] or "").strip(), fname=str(vals[7] or "").strip()))
    return batches, rows
