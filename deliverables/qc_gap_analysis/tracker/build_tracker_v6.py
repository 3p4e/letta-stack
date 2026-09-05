#!/usr/bin/env python3
"""Build CoQ_Analysis_Master_v6.xlsx — the owner's CoQ tracker as flat tables.

Design rules (v6):
  * no merged cells anywhere in a data region; one value per cell; one header row
    per sheet with an autofilter, panes frozen under it;
  * every table row is a complete record — identity columns are repeated, never merged
    down a block;
  * dates are real dates (DD.MM.YYYY), numeric results are real numbers;
  * the state of a cell is carried by the standard Good / Neutral / Bad fills (legend on
    the Read Me sheet), the mark is always a plain ✓ or ✗.

Inputs
  CoQ_Analysis_Master_v3.xlsx  — the owner's tracker (which certificate is credited to
                                 which parameter), reflowed; the content is theirs.
  ../coq_artifact_data.json    — the desk's record: release register (chain step 19),
                                 page reads of 31.08.2026, 12-month re-analyses.
Output
  CoQ_Analysis_Master_v6.xlsx
"""
import collections, datetime, json, os, re, sys
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter as L

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, "..", "..", ".."))
sys.path.insert(0, os.path.join(ROOT, "ingestion", "common"))
from batch_id import batch_key  # noqa: E402

SRC = os.path.join(HERE, "CoQ_Analysis_Master_v3.xlsx")
DATA = os.path.join(ROOT, "deliverables", "qc_gap_analysis", "coq_artifact_data.json")
OUT = os.path.join(HERE, "CoQ_Analysis_Master_v6.xlsx")
BUILT = "02.09.2026"

# ----------------------------------------------------------------------------- keys
CYR = str.maketrans("АВЕКМНОРСТУХЈЅІавекмнорстухјѕі", "ABEKMHOPCTYXJSIabekmhopctyxjsi")


def nkey(code):
    c = str(code or "").strip()
    c = re.sub(r"\s*\([^)]*\)\s*$", "", c)
    c = re.sub(r"\s+(EN|MK)$", "", c)
    c = re.sub(r"[-/](GS|ГС|LOD|LoD)[-/]", "/LOD/", c)
    c = c.translate(CYR)
    return re.sub(r"[\s\-/_.]+", "/", c).strip("/").upper()


def cu_key(cu):
    cu = re.sub(r"[＊*]", "", cu).replace("OMP1024", "OPM1024").replace("SJ0925021", "SJ092501").replace("_01/1", "_01")
    return batch_key(cu)


def clean(v):
    v = str(v or "").strip()
    if not v or v in ("—", "/"):
        return ""
    if " | " in v:
        v = v.split(" | ")[0]
    v = re.sub(r"\s*[xх]\s*10\^?", "×10", v)
    v = re.sub(r"10\^(\d)", lambda m: "10" + "⁰¹²³⁴⁵⁶⁷⁸⁹"[int(m.group(1))], v)
    v = re.sub(r"([<>≤≥])\s+", r"\1", v)
    if re.search(r"одговара|отсутн|отсуств|absent", v, re.I):
        v = "absent"
    if re.match(r"^conforms", v, re.I):
        v = "Conforms"
    if re.match(r"^complies", v, re.I):
        v = "Complies"
    if re.match(r"^not found any pesticide", v, re.I):
        v = "≤LOQ"
    if re.match(r"^н\.\s*д\.", v, re.I):
        v = "N.D."
    if re.match(r"^BLQ\b", v):
        v = "BLQ"
    m = re.match(r"^(-?\d+(?:\.\d+)?)\s+—\s+DETECTED", v)
    if m:
        v = m.group(1)
    return v


def as_date(s):
    try:
        return datetime.datetime.strptime(s, "%d.%m.%Y").date()
    except Exception:
        return s


def as_number(v):
    """A plain decimal result becomes a number (keeps its printed precision)."""
    if isinstance(v, str) and re.fullmatch(r"-?\d+(\.\d+)?", v):
        dec = len(v.split(".")[1]) if "." in v else 0
        return (float(v) if dec else int(v)), ("0." + "0" * dec if dec else "0")
    return v, None


# ----------------------------------------------------------------------------- the desk's record
D = json.load(open(DATA))
col2det = {d["col"]: d["no"] for d in D["dets"] if d.get("col")}
DET = {d["no"]: d for d in D["dets"]}
VAL = collections.defaultdict(dict)   # nkey(certificate) -> det no -> value
INH = collections.defaultdict(dict)   # batch key -> det no -> value (in-house documents)
STAB = set()
STRAIN = {}
for b in D["reg"]:
    STRAIN.setdefault(batch_key(b["cb"]), b["strain"])
    for ct in b["certs"]:
        k = nkey(ct["code"])
        if ct.get("stab"):
            STAB.add(k)
        for colL, v in ct["vals"].items():
            no, v = col2det.get(colL), clean(v)
            if no and v:
                VAL[k].setdefault(no, v)
        if ct["code"].startswith("n/a"):
            for colL, v in ct["vals"].items():
                no, v = col2det.get(colL), clean(v)
                if no and v:
                    INH[batch_key(b["cb"])].setdefault(no, v)
for c in D["coqs"]:
    cbk = batch_key(re.sub(r"[＊*]", "", c["cb"]))
    STRAIN.setdefault(cbk, c["strain"])
    for r in c["rows"]:
        v = clean(r["res"])
        if not v or not r["doc"] or r["doc"] == "—":
            continue
        if r["doc"].startswith("n/a"):
            INH[cbk][r["no"]] = v
        else:
            VAL[nkey(r["doc"])][r["no"]] = v

# ----------------------------------------------------------------------------- parameters
GROUPS = {1: ["1"], 2: ["2"], 3: ["3"], 4: ["4"], 5: ["5"], 6: ["6"], 7: ["7"], 8: ["8"],
          9: ["9.1", "9.2", "9.3", "9.4", "9.5"], 10: ["10.1", "10.2", "10.3"],
          11: ["11.1", "11.2", "11.3", "11.4"], 12: ["12"]}
DETS = [no for p in range(1, 13) for no in GROUPS[p]]
PNAME = {1: "Identification A — Appearance", 2: "Identification B — Microscopy", 3: "Identification C — HPLC",
         4: "Assay — Total Δ⁹-THC", 5: "Assay — Total CBD", 6: "Total CBN", 7: "Foreign Matter", 8: "Loss on Drying",
         9: "Microbiological Purity", 10: "Mycotoxins", 11: "Heavy Metals", 12: "Pesticide Residues"}
DNAME = {"1": "", "2": "", "3": "", "4": "", "5": "", "6": "", "7": "", "8": "", "12": "",
         "9.1": "TAMC", "9.2": "TYMC", "9.3": "Bile-tolerant gram-negative bacteria", "9.4": "Salmonella", "9.5": "Escherichia coli",
         "10.1": "Aflatoxin B₁", "10.2": "Aflatoxins ∑ (B₁+B₂+G₁+G₂)", "10.3": "Ochratoxin A",
         "11.1": "Lead (Pb)", "11.2": "Cadmium (Cd)", "11.3": "Arsenic (As)", "11.4": "Mercury (Hg)"}
HEAD = {"1": "1 Ident. A", "2": "2 Ident. B", "3": "3 Ident. C", "4": "4 Total Δ⁹-THC (%)", "5": "5 Total CBD (%)",
        "6": "6 Total CBN (%)", "7": "7 Foreign matter", "8": "8 Loss on drying (%)",
        "9.1": "9.1 TAMC (CFU/g)", "9.2": "9.2 TYMC (CFU/g)", "9.3": "9.3 Bile-tol. GNB (CFU/g)", "9.4": "9.4 Salmonella", "9.5": "9.5 E. coli",
        "10.1": "10.1 Aflatoxin B₁ (µg/kg)", "10.2": "10.2 ΣAflatoxins (µg/kg)", "10.3": "10.3 Ochratoxin A (µg/kg)",
        "11.1": "11.1 Pb (mg/kg)", "11.2": "11.2 Cd (mg/kg)", "11.3": "11.3 As (mg/kg)", "11.4": "11.4 Hg (mg/kg)", "12": "12 Pesticides"}
CRIT = {"1": "Conforms to Ph. Eur. monograph 3028 (Cannabis flos)", "2": "Conforms to Ph. Eur. monograph 3028 (Cannabis flos)",
        "3": "Conforms to Ph. Eur. monograph 3028 (Cannabis flos)", "4": "Per target grade (CoQ §01)", "5": "≤ 1.0 % w/w", "6": "≤ 1.0 % w/w",
        "7": "≤ 2.0 % (25–50 g); leaves < 1 cm; no seeds", "8": "≤ 12.0 %",
        "9.1": "≤ 10⁵ CFU/g", "9.2": "≤ 10⁴ CFU/g", "9.3": "≤ 10⁴ CFU/g", "9.4": "Absence / 25 g", "9.5": "Absence / 1 g",
        "10.1": "≤ 2 µg/kg", "10.2": "≤ 4 µg/kg", "10.3": "≤ 20 µg/kg",
        "11.1": "≤ 0.5 mg/kg", "11.2": "≤ 0.3 mg/kg", "11.3": "≤ 0.2 mg/kg", "11.4": "≤ 0.1 mg/kg",
        "12": "≤ LOQ (Ph. Eur. 2.8.13 / CUMCS equivalency)"}
SHORT = {1: "Ident. A", 2: "Ident. B", 3: "Ident. C", 4: "Δ⁹-THC", 5: "CBD", 6: "CBN", 7: "Foreign matter", 8: "LoD",
         9: "Microbiology", 10: "Mycotoxins", 11: "Heavy metals", 12: "Pesticides"}
SOURCE = {"in_house_icoa": "In-house (iCoA)", "outsourced_certificate": "Outsourced certificate (eCoA)", "upon_request": "Upon request"}
LABNAME = {"CNP": "UKIM Faculty of Pharmacy — Center for Natural Products", "FHM-K": "Farmahem — cannabinoids",
           "FHM-M": "Farmahem — mycotoxins", "IJZ": "Institute of Public Health", "IJZ-MB": "Institute of Public Health — microbiology",
           "NGP": "Purely Plant — in-house form", "PP": "Purely Plant — in-house", "DFL": "State Phytosanitary Laboratory"}


def kind(code, lab):
    if code.startswith("iCoA-"):
        return "iCoA"
    if code.startswith("NO-DOC-CODE") or code.startswith("NGP-") or lab in ("PP", "NGP"):
        return "In-house"
    if nkey(code) in STAB:
        return "Stability"
    return "eCoA"


def values_of(code, lab, cu):
    src = VAL.get(nkey(code))
    if not src and kind(code, lab) == "In-house":
        src = INH.get(cu_key(cu))
    return src or {}


# ----------------------------------------------------------------------------- the owner's tracker
wb = openpyxl.load_workbook(SRC)
old = wb["CoQ Parameter Tracker"]
PCOL = {int(m.group(1)): i for i, v in enumerate([c.value for c in old[3]], 1) for m in [re.match(r"#(\d+)", str(v or ""))] if m}
ENTRY = re.compile(r"([^\n;]+?)\s*,\s*\((\d{2}\.\d{2}\.\d{4})\)\s*\[([^\]]+)\]")
batches = []
for r in range(4, old.max_row + 1):
    cu = str(old.cell(r, 1).value or "").strip()
    if not cu:
        continue
    p = str(old.cell(r, 2).value or "").strip()
    if "NOT ASSIGNED" in p:
        p = "— not assigned —"
    if "NOT ASSIGNED" in cu:
        cu = "— not assigned —"
    rec = {"cu": cu, "p": p, "status": str(old.cell(r, 3).value or "").strip(),
           "labs": [x.strip() for x in str(old.cell(r, 28).value or "").split("|") if x.strip()],
           "missing": [x.strip() for x in str(old.cell(r, 29).value or "").split(",") if x.strip()], "docs": {}}
    for pn, cidx in PCOL.items():
        rec["docs"][pn] = [(m.group(1).strip(), m.group(2), m.group(3)) for m in ENTRY.finditer(str(old.cell(r, cidx).value or ""))]
    certs = {}
    for pn, docs in rec["docs"].items():
        for code, date, lab in docs:
            certs.setdefault((code, date, lab), set()).add(pn)
    rec["certs"] = sorted(certs.items(), key=lambda kv: (as_date(kv[0][1]).isoformat() if not isinstance(as_date(kv[0][1]), str) else "9", kv[0][0]))
    rec["strain"] = STRAIN.get(cu_key(cu), "")
    batches.append(rec)

# ----------------------------------------------------------------------------- styles
FONT = "Calibri"
F = Font(name=FONT, size=9)
FB = Font(name=FONT, size=9, bold=True)
FH = Font(name=FONT, size=9, bold=True, color="FFFFFF")
FT = Font(name=FONT, size=14, bold=True, color="1F3864")
FS = Font(name=FONT, size=9, italic=True, color="595959")
HEADER = PatternFill("solid", fgColor="1F3864")
GOOD = PatternFill("solid", fgColor="C6EFCE")     # certificate and value on the desk
NEUTRAL = PatternFill("solid", fgColor="FFEB9C")  # certificate credited, no value on the desk
BAD = PatternFill("solid", fgColor="FFC7CE")      # no certificate
STABF = PatternFill("solid", fgColor="F8CBAD")    # stability timepoint
INHF = PatternFill("solid", fgColor="E7E6E6")     # in-house document only
IDF = PatternFill("solid", fgColor="F2F2F2")      # identity columns
FGOOD = Font(name=FONT, size=9, color="006100")
FNEUT = Font(name=FONT, size=9, color="9C5700")
FBAD = Font(name=FONT, size=9, color="9C0006", bold=True)
FSTAB = Font(name=FONT, size=9, color="833C0B")
FINH = Font(name=FONT, size=9, color="595959")
thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
LEFT = Alignment(vertical="center", horizontal="left")
CENTER = Alignment(vertical="center", horizontal="center")
HDR = Alignment(vertical="center", horizontal="center", wrap_text=True)
WRAP = Alignment(vertical="top", horizontal="left", wrap_text=True)


def put(ws, r, c, v, font=F, fill=None, al=LEFT, fmt=None):
    cell = ws.cell(r, c, v)
    cell.font, cell.alignment, cell.border = font, al, BOX
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    return cell


def header(ws, cols, height=30):
    """cols: list of (title, width). One header row, autofilter, frozen."""
    for i, (t, w) in enumerate(cols, 1):
        put(ws, 1, i, t, FH, HEADER, HDR)
        ws.column_dimensions[L(i)].width = w
    ws.row_dimensions[1].height = height


def finish(ws, ncols, nrows, freeze, fit_width=1):
    ws.auto_filter.ref = f"A1:{L(ncols)}{nrows}"
    ws.freeze_panes = freeze
    ws.print_title_rows = "1:1"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth = fit_width
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.4
    ws.sheet_view.zoomScale = 100


def state(has_doc, has_val, kd):
    """(mark, fill, font) for one credited certificate × determination."""
    if not has_doc:
        return "✗", BAD, FBAD
    if not has_val:
        return "✗", NEUTRAL, FNEUT
    if kd == "Stability":
        return "✓", STABF, FSTAB
    if kd == "In-house":
        return "✓", INHF, FINH
    return "✓", GOOD, FGOOD


def result_cell(ws, r, c, v, fill, font):
    num, fmt = as_number(v)
    put(ws, r, c, num, font, fill, CENTER, fmt)


stats = collections.Counter()

# ============================================================================= 1. Batch Coverage
cov = wb.create_sheet("Batch Coverage", 0)
cols = [("CU Batch", 12), ("P Batch", 24), ("Strain", 18), ("Status", 14)] + [(f"{p} {PNAME[p]}", 11) for p in range(1, 13)] + \
       [("Missing (n)", 8), ("Missing parameters", 48), ("Certificates (n)", 8), ("Labs present", 44)]
header(cov, cols, 42)
for i, b in enumerate(batches, 2):
    put(cov, i, 1, b["cu"], FB, IDF); put(cov, i, 2, b["p"], F, IDF); put(cov, i, 3, b["strain"], F, IDF)
    put(cov, i, 4, b["status"], FBAD if b["status"].startswith("❌") else FNEUT, BAD if b["status"].startswith("❌") else NEUTRAL, CENTER)
    for p in range(1, 13):
        docs = b["docs"][p]
        kinds = {kind(code, lab) for code, date, lab in docs}
        cert_val = any(values_of(code, lab, b["cu"]).get(no) for code, date, lab in docs for no in GROUPS[p] if kind(code, lab) in ("eCoA", "iCoA"))
        if not docs:
            m, fill, font = "✗", BAD, FBAD; stats["cov ✗ none"] += 1
        elif kinds & {"eCoA", "iCoA"}:
            m, fill, font = ("✓", GOOD, FGOOD) if cert_val else ("✓", NEUTRAL, FNEUT); stats["cov ✓ value" if cert_val else "cov ✓ no value"] += 1
        elif "Stability" in kinds:
            m, fill, font = "✓", STABF, FSTAB; stats["cov ✓ stability only"] += 1
        else:
            m, fill, font = "✓", INHF, FINH; stats["cov ✓ in-house only"] += 1
        put(cov, i, 4 + p, m, font, fill, CENTER)
    put(cov, i, 17, len(b["missing"]), FB if b["missing"] else F, None, CENTER)
    miss = "; ".join(f"#{p} {SHORT[p]}" for p in range(1, 13) if not b["docs"][p])
    put(cov, i, 18, miss or "—", F, None, WRAP if len(miss) > 60 else LEFT)
    if len(miss) > 60:
        cov.row_dimensions[i].height = 13 * (len(miss) // 60 + 1)
    put(cov, i, 19, len(b["certs"]), F, None, CENTER)
    put(cov, i, 20, "; ".join(x.replace(" doc", "").replace("  ", " ") for x in b["labs"]), F)
finish(cov, len(cols), len(batches) + 1, "E2")

# ============================================================================= 2. CoQ Parameter Tracker (one row per batch × certificate)
name = "CoQ Parameter Tracker"; wb.remove(old); trk = wb.create_sheet(name, 1)
cols = [("CU Batch", 12), ("P Batch", 24), ("Certificate", 28), ("Date", 11), ("Lab", 8), ("Kind", 9)] + \
       [(HEAD[no], 10 if no in ("4", "5", "6", "8", "9.4", "9.5", "11.1", "11.2", "11.3", "11.4") else 11) for no in DETS]
header(trk, cols, 42)
r = 2
for b in batches:
    for (code, date, lab), params in b["certs"]:
        kd = kind(code, lab); vals = values_of(code, lab, b["cu"])
        put(trk, r, 1, b["cu"], FB, IDF); put(trk, r, 2, b["p"], F, IDF); put(trk, r, 3, code, F, IDF)
        put(trk, r, 4, as_date(date), F, IDF, CENTER, "DD.MM.YYYY"); put(trk, r, 5, lab, F, IDF, CENTER); put(trk, r, 6, kd, F, IDF, CENTER)
        for j, no in enumerate(DETS):
            p = int(no.split(".")[0])
            if p not in params:
                put(trk, r, 7 + j, None, F, None, CENTER); continue
            v = vals.get(no, "")
            _, fill, font = state(True, bool(v), kd)
            result_cell(trk, r, 7 + j, v or "—", fill, font); stats["trk value" if v else "trk credited, no value"] += 1
        r += 1
finish(trk, len(cols), r - 1, "G2")
print("tracker rows:", r - 2)

# ============================================================================= 3. Results Register (long table)
reg = wb.create_sheet("Results Register", 2)
cols = [("CU Batch", 12), ("P Batch", 24), ("#", 6), ("Parameter", 26), ("Determination", 26), ("Mark", 6), ("Result", 14),
        ("Acceptance criterion", 36), ("Certificate", 28), ("Date", 11), ("Lab", 8), ("Kind", 9), ("Note", 60)]
header(reg, cols, 24)
r = 2
NOTE = {"none": "no certificate credited on the tracker", "noval": "certificate credited; the desk holds no value for this determination from it",
        "Stability": "stability-timepoint certificate — the value is not a release result",
        "In-house": "in-house document — not an eCoA or iCoA; not coverage for a release certificate", "ok": ""}
for b in batches:
    for p in range(1, 13):
        docs = b["docs"][p]
        if not docs:
            put(reg, r, 1, b["cu"], FB, IDF); put(reg, r, 2, b["p"], F, IDF); put(reg, r, 3, str(p), F, IDF, CENTER)
            put(reg, r, 4, PNAME[p], F, IDF); put(reg, r, 5, "", F, IDF); put(reg, r, 6, "✗", FBAD, BAD, CENTER); put(reg, r, 7, None, F, BAD, CENTER)
            put(reg, r, 8, CRIT[GROUPS[p][0]] if len(GROUPS[p]) == 1 else "see determinations", F)
            for c in (9, 10, 11, 12): put(reg, r, c, None, F, None, CENTER)
            put(reg, r, 13, NOTE["none"], FS); r += 1; stats["reg ✗ none"] += 1; continue
        for code, date, lab in docs:
            kd = kind(code, lab); vals = values_of(code, lab, b["cu"])
            for no in GROUPS[p]:
                v = vals.get(no, "")
                m, fill, font = state(True, bool(v), kd)
                put(reg, r, 1, b["cu"], FB, IDF); put(reg, r, 2, b["p"], F, IDF); put(reg, r, 3, no, F, IDF, CENTER)
                put(reg, r, 4, PNAME[p], F, IDF); put(reg, r, 5, DNAME[no], F, IDF); put(reg, r, 6, m, font, fill, CENTER)
                result_cell(reg, r, 7, v or "—", fill, font); put(reg, r, 8, CRIT[no], F)
                put(reg, r, 9, code, F); put(reg, r, 10, as_date(date), F, None, CENTER, "DD.MM.YYYY"); put(reg, r, 11, lab, F, None, CENTER); put(reg, r, 12, kd, F, None, CENTER)
                put(reg, r, 13, NOTE["noval"] if not v else NOTE.get(kd, ""), FS); r += 1; stats["reg ✓" if v else "reg ✗ no value"] += 1
finish(reg, len(cols), r - 1, "C2")
print("register rows:", r - 2)

# ============================================================================= 4. eCOA Document Index (flat)
ixo = wb["eCOA Document Index"]
rows = []
p_last = cu_last = ""
for rr in range(2, ixo.max_row + 1):
    vals = [ixo.cell(rr, c).value for c in range(1, 9)]
    if vals[0]: p_last = str(vals[0]).strip()
    if vals[1]: cu_last = str(vals[1]).strip()
    if not vals[5]: continue
    rows.append([p_last, cu_last] + [str(x).strip() if x is not None else "" for x in vals[2:8]])
wb.remove(ixo); ix = wb.create_sheet("eCOA Document Index", 3)
cols = [("P Batch", 24), ("CU Batch", 12), ("Lab", 8), ("Laboratory", 40), ("Kind", 9), ("Certificate", 28), ("Date", 11),
        ("Document type", 52), ("Parameters covered", 22), ("Values on desk", 9), ("Filename", 44)]
header(ix, cols, 24)
for i, (pb, cu, lab, params, dtype, code, date, fname) in enumerate(rows, 2):
    kd = kind(code, lab); vals = values_of(code, lab, cu); n = sum(1 for no in DETS if vals.get(no))
    put(ix, i, 1, pb, F, IDF); put(ix, i, 2, cu, FB, IDF); put(ix, i, 3, lab, F, None, CENTER); put(ix, i, 4, LABNAME.get(lab, ""), F)
    put(ix, i, 5, kd, F, None, CENTER); put(ix, i, 6, code, F); put(ix, i, 7, as_date(date), F, None, CENTER, "DD.MM.YYYY")
    put(ix, i, 8, dtype, F); put(ix, i, 9, params, F)
    m, fill, font = state(True, n > 0, kd); put(ix, i, 10, m, font, fill, CENTER); put(ix, i, 11, fname, F)
    stats["index ✓" if n else "index ✗"] += 1
finish(ix, len(cols), len(rows) + 1, "C2")
print("index rows:", len(rows))

# ============================================================================= 5. Parameters
par = wb.create_sheet("Parameters", 4)
cols = [("#", 6), ("Parameter", 28), ("Determination", 34), ("Method", 42), ("Acceptance criterion (global)", 40), ("Source", 28), ("Tracker column", 12)]
header(par, cols, 24)
i = 2
for p in range(1, 13):
    for no in GROUPS[p]:
        d = DET[no]
        put(par, i, 1, no, F, IDF, CENTER); put(par, i, 2, PNAME[p], F, IDF); put(par, i, 3, DNAME[no] or d["en"], F)
        put(par, i, 4, d["method"], F); put(par, i, 5, CRIT[no], F); put(par, i, 6, SOURCE.get(d["src"], d["src"]), F)
        put(par, i, 7, L(7 + DETS.index(no)), F, None, CENTER); i += 1
finish(par, len(cols), i - 1, "A2")

# ============================================================================= 0. Read Me
rm = wb.create_sheet("Read Me", 0)
rm.column_dimensions["A"].width = 16; rm.column_dimensions["B"].width = 120
rm.sheet_view.showGridLines = False
lines = [
    ("title", "CoQ Analysis Master — v6", ""),
    ("sub", f"Built {BUILT} on the owner's CoQ_Analysis_Master (v3 reflow). Which certificate is credited to which parameter is the owner's; results, dates and labs are the desk's record: release register (correction chain step 19), page reads of 31.08.2026 and the 12-month re-analyses. Nothing is invented; a value the desk does not hold is shown as —.", ""),
    ("", "", ""),
    ("h", "SHEETS", ""),
    ("row", "Batch Coverage", "One row per batch. ✓/✗ for each of the 12 parameters, the missing list, the number of certificates and the laboratories present."),
    ("row", "CoQ Parameter Tracker", "One row per batch × certificate. One column per determination (21): the value that certificate reports, as the desk holds it. A blank cell means the certificate is not credited for that parameter."),
    ("row", "Results Register", "One row per batch × determination × certificate — the same facts as a flat, filterable register: mark, result, acceptance criterion, certificate, date, lab, kind, note."),
    ("row", "eCOA Document Index", "One row per document: batch, laboratory, kind, code, date, type, parameters covered, whether the desk holds a value from it, filename."),
    ("row", "Parameters", "The 21 determinations with method, global acceptance criterion and source."),
    ("row", "Summary Dashboard", "The owner's aggregate, unchanged."),
    ("", "", ""),
    ("h", "LEGEND", ""),
    ("leg", ("✓", GOOD, FGOOD), "Certificate (eCoA or iCoA) on file and its value on the desk."),
    ("leg", ("—", NEUTRAL, FNEUT), "Certificate credited on the tracker, but the desk holds no value for this determination from it — the determination is not on that certificate, or the certificate never entered the release register."),
    ("leg", ("✓", STABF, FSTAB), "Stability-timepoint certificate — its value is not a release result."),
    ("leg", ("✓", INHF, FINH), "In-house document only (Report of Analysis, NGP form, PP) — not an eCoA or iCoA; not coverage for a release certificate."),
    ("leg", ("✗", BAD, FBAD), "No certificate credited for the parameter."),
    ("", "", ""),
    ("h", "CONVENTIONS", ""),
    ("row", "Batch names", "CU batch as the owner names it; a sub-lot digit belongs to the batch (FB012601_1), never to the certificate code."),
    ("row", "Certificate codes", "As printed on the certificate (Cyrillic ППК codes kept); the Kind column says eCoA / iCoA / Stability / In-house."),
    ("row", "Dates", "Real dates, DD.MM.YYYY — sortable and filterable."),
    ("row", "Results", "Numeric results are numbers with the certificate's printed precision; qualitative results are text (Conforms, absent, <LOQ, ND)."),
    ("row", "Print", "Every table: A3 landscape, one page wide, header row repeated, panes frozen under the header."),
]
for i, (k, a, b) in enumerate(lines, 1):
    if k == "title":
        rm.cell(i, 1, a).font = FT; rm.row_dimensions[i].height = 24
    elif k == "sub":
        c = rm.cell(i, 1, a); c.font = FS; c.alignment = WRAP; rm.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2); rm.row_dimensions[i].height = 40
    elif k == "h":
        rm.cell(i, 1, a).font = FB
    elif k == "row":
        c0 = rm.cell(i, 1, a); c0.font = FB; c0.alignment = Alignment(vertical="top"); c = rm.cell(i, 2, b); c.font = F; c.alignment = WRAP
        rm.row_dimensions[i].height = 13 if len(b) < 120 else 26
    elif k == "leg":
        mark, fill, font = a
        c = rm.cell(i, 1, mark); c.font = font; c.fill = fill; c.alignment = Alignment(vertical="top", horizontal="center"); c.border = BOX
        c = rm.cell(i, 2, b); c.font = F; c.alignment = WRAP; rm.row_dimensions[i].height = 13 if len(b) < 120 else 26
rm.page_setup.orientation = "portrait"; rm.page_setup.fitToWidth = 1; rm.page_setup.fitToHeight = 0; rm.sheet_properties.pageSetUpPr.fitToPage = True

# ----------------------------------------------------------------------------- order, save
order = ["Read Me", "Batch Coverage", "CoQ Parameter Tracker", "Results Register", "eCOA Document Index", "Parameters", "Summary Dashboard"]
if "Missing Parameters Report" in wb.sheetnames:
    wb.remove(wb["Missing Parameters Report"])
wb._sheets = [wb[n] for n in order]
wb.active = 1
wb.save(OUT)
print("saved", OUT); print(dict(stats))
