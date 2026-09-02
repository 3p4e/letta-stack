#!/usr/bin/env python3
"""Build CoQ_Analysis_Master_v7.xlsx — the block layout the Head of QC specified.

v7 adds the sheet `CoQ Parameter Tracker v7` to the v6 workbook and keeps everything
else, exactly as the specification asks (the v6 flat tracker is kept alongside, renamed
`CoQ Parameter Tracker (flat)`, so the two can be compared).

The block rule, per the specification:

  * one batch = one block of TWO rows, boxed with a thick border;
  * single-value parameters (#1–#8, #12) occupy Result | eCOA ref | ✓/✗, each cell
    merged vertically across the two rows;
  * #9 Microbiology, #10 Mycotoxins and #11 Heavy metals give each sub-determination
    its own column: the top row carries the sub-results, the bottom row carries the
    certificate reference(s) merged across them;
  * several certificates for the same parameter are stacked as lines in ascending date
    order — line n of Result matches line n of eCOA ref matches glyph n of ✓/✗, and the
    glyphs are coloured individually;
  * acceptance criteria live in header row 3 and are enforced: a result that provably
    exceeds its criterion is printed red and bold and is named in the batch's STATUS.

Sources: `tracker_data.py` (the desk's values, the owner's certificate credits).
"""
import collections, importlib.util, math, os, sys

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter as L

HERE = os.path.dirname(os.path.abspath(__file__))
spec = importlib.util.spec_from_file_location("tracker_data", os.path.join(HERE, "tracker_data.py"))
T = importlib.util.module_from_spec(spec)
spec.loader.exec_module(T)

SRC = os.path.join(HERE, "CoQ_Analysis_Master_v6.xlsx")
OUT = os.path.join(HERE, "CoQ_Analysis_Master_v7.xlsx")
SHEET = "CoQ Parameter Tracker v7"
BUILT = "02.09.2026"
STATUS_PARTIAL_MAX = 3          # 1–3 without a release result → ⚠ PARTIAL; ≥4 → ✗

# --------------------------------------------------------------------------- palette
NAVY = "1F3864"; SUBHDR = "FFF2CC"; GREY = "EFEFEF"
FILL = {"green": "C6EFCE", "orange": "FCE5CD", "amber": "FDE9D9", "red": "F4CCCC"}
GLYPHFILL = {"green": "6AA84F", "orange": "E69138", "amber": "F6B26B", "red": "CC0000"}
STATUSFILL = {"green": "38761D", "orange": "E69138", "red": "CC0000"}
RED = "9C0006"
F7 = Font(name="Calibri", size=7)
F7B = Font(name="Calibri", size=7, bold=True)
F7R = Font(name="Calibri", size=7, bold=True, color=RED)
F7U = Font(name="Calibri", size=7, bold=True, color="B45F06")
F6 = Font(name="Calibri", size=6)
F6I = Font(name="Calibri", size=6.5, italic=True, color="595959")
FW = Font(name="Calibri", size=8, bold=True, color="FFFFFF")
FWS = Font(name="Calibri", size=7, bold=True, color="FFFFFF")
FSUB = Font(name="Calibri", size=7, bold=True)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
TOPC = Alignment(horizontal="center", vertical="top", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
THICK = Side(style="medium", color="000000")
MED = Side(style="thin", color="404040")


def put(ws, r, c, v, font=F7, fill=None, al=CEN):
    cell = ws.cell(r, c, v)
    cell.font, cell.alignment, cell.border = font, al, BOX
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    return cell


def fill_range(ws, r1, c1, r2, c2, colour):
    pf = PatternFill("solid", fgColor=colour)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).fill = pf


def outline(ws, r1, c1, r2, c2, side):
    """Draw one rectangle without disturbing the inner borders."""
    for c in range(c1, c2 + 1):
        for r, edge in ((r1, "top"), (r2, "bottom")):
            b = ws.cell(r, c).border
            ws.cell(r, c).border = Border(left=b.left, right=b.right,
                                          top=side if edge == "top" else b.top,
                                          bottom=side if edge == "bottom" else b.bottom)
    for r in range(r1, r2 + 1):
        for c, edge in ((c1, "left"), (c2, "right")):
            b = ws.cell(r, c).border
            ws.cell(r, c).border = Border(top=b.top, bottom=b.bottom,
                                          left=side if edge == "left" else b.left,
                                          right=side if edge == "right" else b.right)


def nlines(text, width):
    per = max(1, int(width * 1.25))
    return sum(max(1, math.ceil(len(ln) / per)) for ln in str(text or "").split("\n")) or 1


# --------------------------------------------------------------------------- data
VAL, INH, STAB, STRAIN, DET = T.load_desk()
batches, index_rows = T.load_owner()


def values_of(code, lab, cu):
    src = VAL.get(T.nkey(code))
    if not src and T.kind_of(code, lab, STAB) == "In-house":
        src = INH.get(T.cu_key(cu))
    return src or {}


def ecoa_line(code, date, lab):
    return f"{code}, ({date}) [{lab}]"


# --------------------------------------------------------------------------- layout
cols, col = [], 4
for p in T.PARAMS:
    p["start"] = col
    subs = T.GROUPS[p["n"]]
    if len(subs) > 1:
        p["subs"] = subs
        for no in subs:
            cols.append((col, p, no)); col += 1
    else:
        p["subs"] = None
        cols.append((col, p, "result")); col += 1
        cols.append((col, p, "ecoa")); col += 1
    cols.append((col, p, "check")); col += 1
    p["end"] = col - 1
    p["width"] = p["end"] - p["start"] + 1
LAST = col - 1

wb = openpyxl.load_workbook(SRC)
if SHEET in wb.sheetnames:
    wb.remove(wb[SHEET])
flat = wb["CoQ Parameter Tracker"]
flat.title = "CoQ Parameter Tracker (flat)"
ws = wb.create_sheet(SHEET, wb.sheetnames.index("Batch Coverage") + 1)

# ---- header rows 1–4
put(ws, 1, 1, "BATCH IDENTIFICATION", FW, NAVY)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
g, gstart = None, 0
for i, p in enumerate(T.PARAMS):
    if p["group"] != g:
        if g:
            ws.merge_cells(start_row=1, start_column=gstart, end_row=1, end_column=p["start"] - 1)
            put(ws, 1, gstart, g, FW, NAVY)
        g, gstart = p["group"], p["start"]
    if i == len(T.PARAMS) - 1:
        ws.merge_cells(start_row=1, start_column=gstart, end_row=1, end_column=p["end"])
        put(ws, 1, gstart, g, FW, NAVY)
for c, t in ((1, "CU Batch"), (2, "P Batch"), (3, "STATUS")):
    ws.merge_cells(start_row=2, start_column=c, end_row=4, end_column=c)
    put(ws, 2, c, t, FWS, NAVY)
for p in T.PARAMS:
    ws.merge_cells(start_row=2, start_column=p["start"], end_row=2, end_column=p["end"])
    put(ws, 2, p["start"], p["title"] + "\n" + p["method"], FW, NAVY)
    if p["subs"]:
        for i, no in enumerate(p["subs"]):
            put(ws, 3, p["start"] + i, "A.C.: " + T.CRIT[no], F6I, GREY)
            put(ws, 4, p["start"] + i, T.SUB[no], FSUB, SUBHDR)
        put(ws, 3, p["end"], "", F6I, GREY)
    else:
        put(ws, 3, p["start"], "A.C.: " + T.CRIT[str(p["n"])], F6I, GREY)
        put(ws, 3, p["start"] + 1, "", F6I, GREY)
        ws.merge_cells(start_row=3, start_column=p["start"], end_row=3, end_column=p["start"] + 1)
        put(ws, 4, p["start"], "Result (as reported)", FSUB, SUBHDR)
        put(ws, 4, p["start"] + 1, "eCOA ref, (date) [Lab] — one certificate per line", FSUB, SUBHDR)
    put(ws, 3, p["end"], "", F6I, GREY)
    put(ws, 4, p["end"], "✓/✗", FSUB, SUBHDR)
for h, height in ((1, 16), (2, 30), (3, 26), (4, 24)):
    ws.row_dimensions[h].height = height

# ---- body: one block of two rows per batch
row = 5
stats = collections.Counter()
oos_rows = []
for b in batches:
    top, bot = row, row + 1
    ws.merge_cells(start_row=top, start_column=1, end_row=bot, end_column=1)
    put(ws, top, 1, b["cu"], FWS, NAVY)
    ws.merge_cells(start_row=top, start_column=2, end_row=bot, end_column=2)
    put(ws, top, 2, b["p"], FWS, NAVY)
    for r in (top, bot):
        for c in (1, 2):
            ws.cell(r, c).fill = PatternFill("solid", fgColor=NAVY)
            ws.cell(r, c).border = BOX

    no_cert = cert_no_result = missing = 0
    oos_list, und_list, stab_list = [], [], []

    def verdict(det_no, release_vals, stability_vals, label):
        """Judge release results only; a stability exceedance is reported apart."""
        bad = any(T.over_limit(det_no, x) for x in release_vals)
        und = (not bad) and any(T.undetermined(det_no, x) for x in release_vals)
        if bad:
            oos_list.append(label)
        elif und:
            und_list.append(label)
        if any(T.over_limit(det_no, x) for x in stability_vals):
            stab_list.append(label)
        return F7R if bad else (F7U if und else F7B)
    top_lines = bot_lines = 1
    for p in T.PARAMS:
        docs = b["docs"][p["n"]]
        rel, stab, credited = [], [], []
        for code, date, lab in docs:
            vals = values_of(code, lab, b["cu"])
            has = any(vals.get(no) for no in T.GROUPS[p["n"]])
            if has:
                (stab if T.nkey(code) in STAB else rel).append(((code, date, lab), vals))
            else:
                credited.append(((code, date, lab), {}))
        lines = rel + stab
        if rel:
            state = "green"
        elif stab:
            state = "orange"
        elif credited:
            state = "amber"; cert_no_result += 1
        else:
            state = "red"; no_cert += 1
        if state != "green":
            missing += 1
        stats[state] += 1

        ecoa = [ecoa_line(*d) for d, _ in lines] + [ecoa_line(*d) for d, _ in credited]
        glyphs = [("✓", "FFFFFF") for d, v in rel] + [("✓", "FFF2CC") for d, v in stab] + \
                 [("✗", "FFFFFF") for _ in credited]
        if not glyphs:
            glyphs = [("✗", "FFFFFF")]

        if p["subs"]:
            for i, no in enumerate(p["subs"]):
                txt = [(v.get(no) or "n.r.") for _, v in lines] + ["—" for _ in credited]
                cell_v = "\n".join(txt) if txt else "— MISSING —"
                font = verdict(no, [v.get(no) for _, v in rel], [v.get(no) for _, v in stab], f"#{p['n']} {T.SUB[no]}")
                put(ws, top, p["start"] + i, cell_v, font, FILL[state])
                top_lines = max(top_lines, nlines(cell_v, 8.5))
            ws.merge_cells(start_row=bot, start_column=p["start"], end_row=bot, end_column=p["end"] - 1)
            put(ws, bot, p["start"], "\n".join(ecoa), F6, FILL[state], TOPC)
            fill_range(ws, bot, p["start"], bot, p["end"] - 1, FILL[state])
            bot_lines = max(bot_lines, nlines("\n".join(ecoa), 8.5 * (len(p["subs"]) - 1) + 21))
        else:
            res = [str(v.get(str(p["n"])) or "") for _, v in lines] + ["—" for _ in credited]
            cell_v = "\n".join(res) if res else "— MISSING —"
            font = verdict(str(p["n"]), [v.get(str(p["n"])) for _, v in rel], [v.get(str(p["n"])) for _, v in stab], f"#{p['n']}")
            ws.merge_cells(start_row=top, start_column=p["start"], end_row=bot, end_column=p["start"])
            put(ws, top, p["start"], cell_v, font, FILL[state])
            ws.merge_cells(start_row=top, start_column=p["start"] + 1, end_row=bot, end_column=p["start"] + 1)
            put(ws, top, p["start"] + 1, "\n".join(ecoa), F6, FILL[state], TOPC)
            fill_range(ws, top, p["start"], bot, p["start"] + 1, FILL[state])
            need = max(nlines(cell_v, 10), nlines("\n".join(ecoa), 21))
            if need > top_lines + bot_lines:
                bot_lines = need - top_lines

        ws.merge_cells(start_row=top, start_column=p["end"], end_row=bot, end_column=p["end"])
        rich = CellRichText()
        for i, (gl, colour) in enumerate(glyphs):
            rich.append(TextBlock(InlineFont(rFont="Calibri", sz=7, b=True, color=colour), gl + ("\n" if i < len(glyphs) - 1 else "")))
        cell = ws.cell(top, p["end"], rich)
        cell.alignment = CEN
        fill_range(ws, top, p["end"], bot, p["end"], GLYPHFILL[state])

    if missing == 0:
        st, colour = "✓ COMPLETE", STATUSFILL["green"]
    else:
        glyph = "⚠" if missing <= STATUS_PARTIAL_MAX else "✗"
        st = f"{glyph} {missing} NO RESULT\n({no_cert} no cert / {cert_no_result} cert w/o result)"
        colour = STATUSFILL["orange"] if missing <= STATUS_PARTIAL_MAX else STATUSFILL["red"]
    if oos_list:
        st += "\n✗ OUT OF SPECIFICATION: " + ", ".join(oos_list)
        colour = STATUSFILL["red"]
        oos_rows.append((b["cu"], "OOS", oos_list))
    if und_list:
        st += "\n◐ UNDETERMINED (Ph. Eur. band): " + ", ".join(und_list)
        if not oos_list:
            colour = STATUSFILL["orange"]
        oos_rows.append((b["cu"], "UNDETERMINED", und_list))
    if stab_list:
        st += "\n· stability above A.C.: " + ", ".join(stab_list)
        oos_rows.append((b["cu"], "stability", stab_list))
    ws.merge_cells(start_row=top, start_column=3, end_row=bot, end_column=3)
    put(ws, top, 3, st, FWS, colour)
    fill_range(ws, top, 3, bot, 3, colour)

    ws.row_dimensions[top].height = max(15, 9.0 * top_lines + 3)
    ws.row_dimensions[bot].height = max(15, 9.0 * bot_lines + 3)
    outline(ws, top, 1, bot, LAST, THICK)
    for p in T.PARAMS:
        outline(ws, top, p["start"], bot, p["end"], MED)
    row = bot + 1

LASTROW = row - 1

# ---- key
key = ("KEY — ✓ green: certificate on file AND its result on the desk (release or re-test). "
       "✓ orange: stability-timepoint certificate — the result is NOT a release result. "
       "✗ amber: the certificate is credited for this parameter but the desk holds no result from it. "
       "✗ red — MISSING —: no certificate covers this parameter for this batch. "
       "BLOCK RULE: one batch = one block of two rows; several certificates for the same parameter are stacked as lines in "
       "ascending date order, line n of Result ↔ line n of eCOA ref ↔ glyph n of ✓/✗. For #9, #10 and #11 the top row holds the "
       "sub-determination results and the bottom row the certificate(s). n.r. = that sub-determination is not reported on that "
       "certificate; — on a result line means that certificate is credited here but holds no result. RED BOLD result = OUT OF SPECIFICATION against the criterion in row 3; AMBER BOLD result = UNDETERMINED, in the "
       "Ph. Eur. band between a printed count limit and twice it. The check follows the Quality Desk exactly: a counted "
       "microbiological limit printed as ≤ 10ⁿ CFU/g is judged against 2 × 10ⁿ (Ph. Eur. 5.1.4); ND, <LOQ, <10, absent, a range "
       "written with \"and\", and any prose annotation are never judged; only release results are judged, and a stability "
       "timepoint above the criterion is named separately in STATUS. Every out-of-specification result needs an investigation "
       "record, not just a red cell. Acceptance criteria in row 3 are the global criteria of the CoQ parameter "
       f"schedule. v7 — built {BUILT} from the desk's record and the owner's certificate credits.")
krow = LASTROW + 2
ws.merge_cells(start_row=krow, start_column=1, end_row=krow, end_column=LAST)
put(ws, krow, 1, key, F6I, GREY, Alignment(horizontal="left", vertical="top", wrap_text=True))
ws.row_dimensions[krow].height = 62

# ---- widths, panes, print
for c, w in ((1, 13), (2, 15), (3, 20)):
    ws.column_dimensions[L(c)].width = w
for c, p, kindc in cols:
    if kindc == "check":
        w = 3.4
    elif kindc == "result":
        w = 10
    elif kindc == "ecoa":
        w = 21
    else:
        w = 8 if kindc in ("9.4", "9.5") else 8.6
    ws.column_dimensions[L(c)].width = w
ws.freeze_panes = "D5"
ws.print_title_rows = "1:4"
ws.page_setup.orientation = "landscape"
ws.page_setup.paperSize = ws.PAPERSIZE_A3
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = ws.page_margins.right = 0.3
ws.page_margins.top = ws.page_margins.bottom = 0.4

# --------------------------------------------------------------------------- index: values + batch key
ix = wb["eCOA Document Index"]
hdr = [str(c.value or "") for c in ix[1]]
cv, ck = len(hdr) + 1, len(hdr) + 2
put(ix, 1, cv, "PARAMETER VALUES", FW, NAVY)
put(ix, 1, ck, "BATCH KEY", FW, NAVY)
ix.column_dimensions[L(cv)].width = 60
ix.column_dimensions[L(ck)].width = 14
for r in range(2, ix.max_row + 1):
    code = str(ix.cell(r, 6).value or "").strip()
    cu = str(ix.cell(r, 2).value or "").strip()
    lab = str(ix.cell(r, 3).value or "").strip()
    if not code:
        continue
    vals = values_of(code, lab, cu)
    parts = []
    for p in T.PARAMS:
        subs = T.GROUPS[p["n"]]
        got = [(T.SUB[no] + " " + vals[no]) if len(subs) > 1 else vals[no] for no in subs if vals.get(no)]
        if got:
            parts.append(f"#{p['n']} " + " · ".join(got))
    stab = " · stability timepoint" if T.nkey(code) in STAB else ""
    cell = ix.cell(r, cv, (" · ".join(parts) + stab) if parts else "no result on the desk for this document")
    cell.font, cell.alignment, cell.border = F7, Alignment(vertical="top", wrap_text=True), BOX
    cell = ix.cell(r, ck, cu)
    cell.font, cell.alignment, cell.border = F7, CEN, BOX
ix.auto_filter.ref = f"A1:{L(ck)}{ix.max_row}"

# ---- Read Me: the v7 block rule
rm = wb["Read Me"]
r = rm.max_row + 2
for label, text in (("v7", "This workbook adds the sheet 'CoQ Parameter Tracker v7' — one batch per two-row block, "
                           "certificates stacked as lines in date order, sub-determinations in their own columns, acceptance "
                           "criteria in header row 3 and enforced, out-of-specification results printed red and named in STATUS. "
                           "The v6 flat table is kept as 'CoQ Parameter Tracker (flat)' for comparison."),
                    ("Index", "'eCOA Document Index' now also carries PARAMETER VALUES and BATCH KEY, so the tracker can be "
                              "rebuilt from the index alone (see CoQ_Tracker_v7_rebuild.gs)."),
                    ("OOS check", "A result is flagged only when it provably exceeds its criterion. ND, <LOQ, <10 and absent pass; "
                                  "'<10² and >10' is judged by its upper bound; a value that cannot be parsed is never flagged.")):
    c0 = rm.cell(r, 1, label); c0.font = Font(name="Calibri", size=9, bold=True); c0.alignment = Alignment(vertical="top")
    c1 = rm.cell(r, 2, text); c1.font = Font(name="Calibri", size=9); c1.alignment = Alignment(vertical="top", wrap_text=True)
    rm.row_dimensions[r].height = 13 * (len(text) // 118 + 1)
    r += 1

wb.save(OUT)
print("saved", OUT)
print("blocks:", len(batches), "rows:", LASTROW - 4, "columns:", LAST)
print("parameter cells by state:", dict(stats))
import collections as _c
by = _c.Counter(k for _, k, _ in oos_rows)
print("verdicts:", dict(by))
for cu, kind, lst in oos_rows:
    print(f"    {kind:14s} {cu:14s} → {', '.join(lst)}")
