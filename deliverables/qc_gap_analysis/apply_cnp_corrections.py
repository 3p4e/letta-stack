#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Apply what reading the 73 CNP potency certificates found.

Evidence is in `review/CNP_PAGE_VERIFICATION_2026-08-31.md`, raw readings in
`review/cnp_page_reads_2026-08-31.json`.

    python3 deliverables/qc_gap_analysis/apply_cnp_corrections.py IN.xlsx OUT.xlsx

**Not one of the 202 values the register already held is wrong.** This family is the
first to come back perfect, and it is the family where perfection was cheapest to
check: every CNP certificate prints Δ9-THC, Δ9-THCA and Вкупно Δ9-THC together with
the relation between them, so each page proves its own arithmetic —
`total = Δ9-THC + Δ9-THCA x 0.877` held on all 60 certificates that print the
breakdown, to within 0.01.

So there is nothing here to correct. What there is instead is **results the register
does not hold at all**, and **two facts about it that a reader cannot see**.

## 1. Ten stability certificates whose results are recorded nowhere

Rows 52-54, 58-60 and 94-97 name a certificate, an issue date and an institution, and
then hold `/` in every result cell. The legend defines `/` as *parameter not covered by
that certificate* — and that is false: all ten certificates print a full set of
results. The rows assert an absence the documents contradict.

They are the Grape Pie stability studies, three sample lots at 3, 6 and 9 months under
two conditions. Filling them is a correction, not an addition: the values were read off
the page and every one passes R4.

## 2. Four of them are over the certificate's own CBN limit

Every one of the four samples held at **40 °C / 75 % RH** exceeds the `<= 1.00 %` CBN
limit printed on its own page:

| Row | Certificate | Sample | CBN | Δ9-THCA |
|---|---|---|---|---|
| 95 | `ППК26037` | month 3 | **1.09** | 0.17 |
| 59 | `ППК26035` | month 6 | **2.15** | BLQ |
| 53 | `ППК26033` | month 6 | **2.35** | 0.01 |
| 97 | `ППК26058` | month 6 | **2.05** | 0.07 |

Four for four, and the THCA column says why: under accelerated conditions the acid has
decarboxylated away and the neutral cannabinoids have oxidised on to CBN. This is the
study working, not a batch problem — the matched 25 °C / 60 % RH samples all read 0.04
to 0.30 with THCA intact. But the numbers are over a printed limit, **no certificate in
the set carries a verdict**, and until now none of them was in the register at all.
Amber, per the legend: a laboratory finding.

## 3. The one batch that failed release · row 260, `ППК26127`

`FB032601`, 21.07.2026. Foreign matter **0.08 %** against a maximum of **2.00 %** — 25
times inside the limit — and the result cell reads **(Не одговара)**. It is the only
CNP certificate in the corpus that carries a ЗАКЛУЧОК, and the only one that concludes:

> во однос на параметарот „Страни материи", **НЕ ОДГОВАРА** на барањата пропишани во
> Ph. Eur. 11.5, монографија *Cannabis flos* (07/2024:3028), **поради утврдено
> присуство на семе од канабис**.

The operative half of the limit is its parenthetical — *без присуство на семе и листови
подолги од 1 cm* — and it is not a number. Seed was found, so the sample fails,
whatever the percentage.

**Every numeric value on this page is in specification.** R1 passes it. The register's
column checks pass it. A validator that compares a value against a limit passes it. The
only thing that fails it is a sentence. And the register has no foreign-matter column
and no verdict column, so row 260 is today indistinguishable from row 262, which
passed.

That cannot be fixed by writing a value into a cell — a column is a decision for QC.
What this script does is refuse to let it stay invisible: the certificate code goes
red, which the legend already defines as *result the issuing laboratory declared out of
specification*, and the comment carries the conclusion verbatim.

Idempotent, and it refuses a workbook whose cells differ from what was verified.
"""
import sys

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

SHEET = "Batch Release QC"
THC, CBD, CBN, LOD, CODE = 5, 7, 8, 9, 23      # columns E, G, H, I, W
AUTHOR = "QC page verification 31.08.2026"
AMBER = PatternFill("solid", fgColor="FFFAEDD4")
RED = PatternFill("solid", fgColor="FFF9DEDB")

# The ten stability certificates, read off their pages. Every cell being written
# currently holds "/" — the legend's "parameter not covered by that certificate".
# row, code, sample, THC, CBD, CBN, LoD
STABILITY = [
    (52, "ППК26032", "month 6, 25C/60% RH", "21.31", "0.03", "0.23", "6.32"),
    (53, "ППК26033", "month 6, 40C/75% RH", "13.16", "0.05", "2.35", "5.84"),
    (54, "ППК26059", "month 9, 25C/60% RH", "23.08", "0.03", "0.30", "6.62"),
    (58, "ППК26034", "month 6, 25C/60% RH", "24.62", "0.04", "0.05", "5.92"),
    (59, "ППК26035", "month 6, 40C/75% RH", "14.99", "0.02", "2.15", "5.52"),
    (60, "ППК26060", "month 9, 25C/60% RH", "25.98", "0.04", "0.04", "7.07"),
    (94, "ППК26036", "month 3, 25C/60% RH", "22.83", "0.05", "0.05", "5.83"),
    (95, "ППК26037", "month 3, 40C/75% RH", "18.62", "0.04", "1.09", "5.38"),
    (96, "ППК26057", "month 6, 25C/60% RH", "24.51", "0.05", "0.04", "6.85"),
    (97, "ППК26058", "month 6, 40C/75% RH", "17.05", "0.03", "2.05", "6.60"),
]

# The four CBN results over the <= 1.00 % limit printed on their own certificates.
OVER_CBN = {
    53: ("ППК26033", "2.35"),
    59: ("ППК26035", "2.15"),
    95: ("ППК26037", "1.09"),
    97: ("ППК26058", "2.05"),
}

CBN_NOTE = (
    "Read 31.08.2026 from the rendered page of {code}.\n\n"
    "CBN {val} % against the <= 1.00 % limit printed on this certificate. The sample "
    "was held at 40 C / 75 % RH; its Δ9-THCA has decarboxylated to near zero and the "
    "neutral cannabinoids have oxidised on to CBN. All four 40 C samples in the study "
    "are over the limit; every matched 25 C / 60 % RH sample reads 0.04-0.30 with THCA "
    "intact.\n\n"
    "Amber, not red: the certificate carries no verdict of any kind. Whether an "
    "accelerated-condition sample is expected to meet the release limit is a question "
    "for the stability protocol, not something this register can settle."
)

STAB_NOTE = (
    "Filled 31.08.2026 from the rendered page of {code} ({sample}).\n\n"
    "These four cells held \"/\", which this register's legend defines as \"parameter "
    "not covered by that certificate\". The certificate covers all four: it prints "
    "loss on drying, CBDA, CBD, CBN, Δ9-THC, Δ9-THCA, Вкупно CBD and Вкупно Δ9-THC.\n\n"
    "The total checks against the certificate's own footnote formula, "
    "Вкупно Δ9-THC = Δ9-THC + Δ9-THCA x 0.877."
)

R260_NOTE = (
    "Read 31.08.2026 from the rendered page of ППК26127.\n\n"
    "THIS BATCH FAILED. The certificate is the only one in the CNP set that carries a "
    "ЗАКЛУЧОК, and it concludes:\n\n"
    "  \"во однос на параметарот „Страни материи“, НЕ ОДГОВАРА на барањата пропишани "
    "во Ph. Eur. 11.5, монографија Cannabis flos (07/2024:3028), поради утврдено "
    "присуство на семе од канабис.\"\n\n"
    "Foreign matter reads 0.08 % against a maximum of 2.00 % — twenty-five times "
    "inside the limit — and is still marked (Не одговара). The operative half of the "
    "limit is its parenthetical, без присуство на семе и листови подолги од 1 cm, and "
    "it is not a number: seed was found, so the sample fails whatever the "
    "percentage.\n\n"
    "Every numeric value on this page is in specification. This register has no "
    "foreign-matter column and no verdict column, so nothing else on row 260 "
    "distinguishes it from a batch that passed. Adding those columns is a QC decision; "
    "this flag is here so the failure is not silent in the meantime."
)

LOD_HEADER_NOTE = (
    "Scope note, 31.08.2026, from reading all 73 CNP certificates.\n\n"
    "This header states the Ph. Eur. monograph 07/2024:3028 limit of 12.00 %. Only 12 "
    "of the 73 CNP certificates are issued against that monograph. The other 61 print "
    "\"Губиток при сушење <= 10.00 %\" and were judged against 10.00 %.\n\n"
    "So for 61 of 73 rows the limit shown here is looser than the one the issuing "
    "laboratory applied. Nothing in the set is affected in fact — the highest reading "
    "is 9.68 % on ППК25370, under both — but 9.52 % and 9.68 % sit close to the limit "
    "that actually applied and comfortably inside the one shown. A check against this "
    "column is not a check against the certificate."
)


def _cur(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return "" if v is None else str(v).strip()


def main(src, dst):
    wb = load_workbook(src)
    ws = wb[SHEET]
    log, skipped = [], []

    for row, code, sample, thc, cbd, cbn, lod in STABILITY:
        got = _cur(ws, row, CODE)
        if got != code:
            raise SystemExit(
                f"REFUSING row {row}: expected certificate {code!r} in W{row}, found "
                f"{got!r}. Not the revision this was verified against.")
        done = 0
        for col, new in ((THC, thc), (CBD, cbd), (CBN, cbn), (LOD, lod)):
            letter = ws.cell(row=row, column=col).column_letter
            cur = _cur(ws, row, col)
            if cur == new:
                done += 1
                continue
            if cur != "/":
                raise SystemExit(
                    f"REFUSING {letter}{row}: expected '/', found {cur!r}. This cell "
                    f"already holds something — re-read {code} before overwriting it.")
            ws.cell(row=row, column=col).value = new
            log.append(f"{letter}{row:<4} '/' -> {new!r}   {code}  {sample}")
        if done == 4:
            skipped.append(f"r{row} {code} already filled")
        cell = ws.cell(row=row, column=CODE)
        if cell.comment is None:
            cell.comment = Comment(STAB_NOTE.format(code=code, sample=sample), AUTHOR)

    for row, (code, val) in OVER_CBN.items():
        cell = ws.cell(row=row, column=CBN)
        cell.fill = AMBER
        if cell.comment is None:
            cell.comment = Comment(CBN_NOTE.format(code=code, val=val), AUTHOR)

    r260 = ws.cell(row=260, column=CODE)
    got = _cur(ws, 260, CODE)
    if got != "ППК26127":
        raise SystemExit(f"REFUSING W260: expected 'ППК26127', found {got!r}.")
    r260.fill = RED
    if r260.comment is None:
        r260.comment = Comment(R260_NOTE, AUTHOR)

    hdr = ws.cell(row=5, column=LOD)
    if hdr.comment is None:
        hdr.comment = Comment(LOD_HEADER_NOTE, AUTHOR)

    wb.save(dst)
    print(f"in : {src}\nout: {dst}\n")
    for line in log:
        print("  FILLED   " + line)
    for s in skipped:
        print("  skipped  " + s)
    print(f"\n{len(log)} cell(s) filled, {len(skipped)} row(s) already applied.")
    print(f"{len(OVER_CBN)} CBN result(s) amber-flagged over the certificate's own limit.")
    print("row 260 (ППК26127) flagged red: the laboratory declared it НЕ ОДГОВАРА.")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 3:
        raise SystemExit(__doc__)
    sys.exit(main(sys.argv[1], sys.argv[2]))
