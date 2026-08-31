#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Which in-house CoA has to be issued for which CoQ, in the order they fall due.

    python3 deliverables/qc_gap_analysis/icoa_issuance_register.py [--csv OUT.csv]

Four release parameters are never covered by an outsourced certificate for most
batches, and the CoQ cannot be compiled without them:

| | Parameter | Ph. Eur. 07/2024:3028 | Who can report it |
|---|---|---|---|
| **A** | Identification — macroscopy | 2.8.23 | CNP, on the Ph. Eur. 11.5 form only |
| **B** | Identification — microscopy | 2.8.23 | CNP, on the Ph. Eur. 11.5 form only |
| **C** | Identification — TLC | 2.8.23 | **no laboratory, ever** |
| **FM** | Foreign matter | 2.8.2 | CNP, on the Ph. Eur. 11.5 form only |

**Identification A is the appearance examination.** The monograph's macroscopic
examination of *Cannabis flos* is the same test as appearance; they are not two
parameters and are not counted twice here.

## Where the scope comes from, and why it can be trusted

CNP changed its certificate form in mid-2026. The older DAB-monograph form reports
loss on drying and cannabinoids and nothing else; the Ph. Eur. 11.5 form adds
identification (macroscopy and microscopy) and foreign matter. So a batch's scope
depends entirely on which form its CNP certificate uses — and **all 73 CNP
certificates were read off their own pages on 31.08.2026**, so this is not inferred
from a filename or a parse.

Twelve certificates are on the Ph. Eur. form. The gap analysis independently marks
exactly twelve batches `IdentC only`, and the two sets agree with **no difference in
either direction** — the page reads confirm the model rather than replacing it.

## What this produces

One iCoA per batch, not one per CoQ. Identity and foreign matter are properties of
the material: a CoQ reissued because the cannabinoids or mycotoxins were re-analysed
covers the same batch, so it references the same iCoA rather than triggering a new
one. The register therefore lists 81 iCoAs against 102 CoQ documents, and names for
each which CoQ or CoQs will reference it.

Chronology is by the batch's own release testing — the latest certificate that is not
a 197-series re-analysis — because that is the point from which the CoQ can be
compiled once its iCoA exists.

**Seven batches cannot be classified from the register, and are marked.** The gap
analysis flags 21 batches for a reissued CoQ, on the basis that they carry a Farmahem
197-series analysis. Fourteen of those also carry earlier testing, so the 197 series
really is a re-analysis and a reissue follows. The other seven — `P060152`, `P060212`,
`P060242`, `P060332`, `P060352`, `P060382`, `P060402` — carry **nothing but** the 197
pair: no potency, no microbiology, no contaminants, no earlier certificate of any kind.

Calling that a *reissue* asserts a first CoQ with nothing behind it. Their names are
P-numbers rather than cultivation batch codes, which reads like packaged lots drawn
from bulk already released — in which case the 197 analysis is their initial release
testing and their CoQ should reference the bulk batch's earlier certificates. The
register does not say which, so this script does not decide: it anchors their
chronology on the 197 date so they can be worked through in order, marks them
`197-only`, and leaves the initial-versus-reissue question where it belongs.
"""
import csv
import re
import sys

from openpyxl import load_workbook

GAP = "deliverables/qc_gap_analysis/batch_gap_analysis.csv"
REG = "deliverables/qc_gap_analysis/PP_Batch_Release_QC_Register_FHM4_2026-08-31.xlsx"
SHEET = "Batch Release QC"

PARAMS = [("A", "identA_appearance"), ("B", "identB_microscopy"),
          ("C", "identC_tlc"), ("FM", "foreign_matter")]


def dkey(d):
    m = re.match(r"(\d{2})\.(\d{2})\.(\d{4})", str(d or ""))
    return (m.group(3) + m.group(2) + m.group(1)) if m else ""


def norm(b):
    return re.sub(r"[^A-Z0-9]", "", str(b).upper())


def register():
    """Per batch: its rows split into initial-release testing and 197-series re-analysis,
    plus the families recorded — a CoQ needs the outsourced panel complete as well as
    its iCoA, and the register is where that is visible."""
    ws = load_workbook(REG)[SHEET]
    out, cur = {}, None
    for r in range(6, 292):
        ref, b = ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value
        if ref not in (None, "") and b:
            cur = str(b)
            out[cur] = {"ref": str(ref), "init": [], "re": [], "fam": set()}
        code = ws.cell(row=r, column=23).value
        if not code or not cur:
            continue
        c = str(code).replace(" ", " ").strip()
        c = re.sub(r"\s*\([^)]*\)\s*$", "", c)
        d = str(ws.cell(row=r, column=24).value or "")
        lab = str(ws.cell(row=r, column=25).value or "")
        if "ППК" in c:
            out[cur]["fam"].add("potency")
        elif "Farmahem" in lab:
            out[cur]["fam"].add("mycotoxins" if re.search(r"-\s*[МM]\s*/", c) else "cannabinoids")
        elif re.match(r"^\d+/\d+/\d+", c):
            out[cur]["fam"].add("microbiology")
        elif re.match(r"^\d+/\d{4}$", c) or "10802" in c:
            out[cur]["fam"].add("physchem")
        if dkey(d):
            (out[cur]["re"] if re.match(r"^197-", c) else out[cur]["init"]).append((dkey(d), d, c))
    return out


def main(csv_out=None):
    gap = list(csv.DictReader(open(GAP, encoding="utf-8")))
    reg = register()
    byk = {norm(k): v for k, v in reg.items()}

    table = []
    for g in gap:
        b = byk.get(norm(g["batch"]))
        missing = [p for p, col in PARAMS if not g[col].strip().upper().startswith("Y")]
        init = max(b["init"]) if b and b["init"] else None
        rea = max(b["re"]) if b and b["re"] else None
        # A batch whose only certificates are the 197 pair has no earlier testing to
        # anchor on. Use the 197 date so it still falls in sequence, and mark it.
        only197 = bool(rea) and not init
        if only197:
            init = rea
        fam = b["fam"] if b else set()
        blocked = [f for f in ("microbiology", "physchem") if f not in fam]
        coq = ["initial"] + (["reissue"] if g["needs_CoQ_reissue"] == "Y" else [])
        table.append({
            "no": int(g["no"]), "batch": g["batch"], "strain": g["strain"],
            "ref": b["ref"] if b else "—",
            "missing": missing, "scope": g["iCoA_scope"].strip(),
            "init_date": init[1] if init else "", "init_key": init[0] if init else "",
            "init_cert": init[2] if init else "",
            "re_date": rea[1] if rea else "", "re_cert": rea[2] if rea else "",
            "coq": coq, "blocked": blocked, "in_register": g["in_register"],
            "only197": only197,
        })
    table.sort(key=lambda r: (r["init_key"] or "99999999", r["no"]))

    n_coq = sum(len(r["coq"]) for r in table)
    full = [r for r in table if len(r["missing"]) == 4]
    conly = [r for r in table if r["missing"] == ["C"]]
    print("CoQ documents predicted")
    print(f"  initial release, one per batch            : {len(table)}")
    reis = [r for r in table if "reissue" in r["coq"]]
    o197 = [r for r in table if r["only197"]]
    print(f"  reissue after 197-series re-analysis      : {len(reis)}")
    print(f"     of which carry earlier testing too     : {len(reis)-len(o197)}  (a real re-analysis)")
    print(f"     of which carry the 197 pair and nothing else : {len(o197)}  (see the docstring)")
    print(f"  total CoQ documents                       : {n_coq}\n")
    print("iCoA documents required — one per batch, referenced by that batch's CoQ(s)")
    print(f"  full panel   A + B + C + foreign matter   : {len(full)}")
    print(f"  Ident C only (CNP Ph. Eur. form covers the rest): {len(conly)}")
    print(f"  total iCoA documents                      : {len(table)}\n")
    print("Missing parameter, counted over the CoQ documents that will reference it")
    for p, label in [("A", "Ident A — macroscopy / appearance"),
                     ("B", "Ident B — microscopy"),
                     ("C", "Ident C — TLC"),
                     ("FM", "Foreign matter")]:
        rs = [r for r in table if p in r["missing"]]
        print(f"  {label:<40} {len(rs):>3} batches, {sum(len(r['coq']) for r in rs):>3} CoQ documents")

    print("\n" + "=" * 118)
    print("CHRONOLOGICAL iCoA ISSUANCE REGISTER")
    print("=" * 118)
    hdr = f"{'#':>3}  {'release':<11} {'batch':<15} {'strain':<22} {'iCoA covers':<22} {'CoQ':<17} outstanding"
    print(hdr); print("-" * 118)
    for i, r in enumerate(table, 1):
        cov = " + ".join(r["missing"]) if r["missing"] else "—"
        coq = "initial" + (" + reissue" if "reissue" in r["coq"] else "")
        out = ", ".join(r["blocked"]) if r["blocked"] else ""
        if r["only197"]: out = ("197-only · " + out) if out else "197-only"
        if not r["in_register"].strip().upper().startswith("Y"): out = ("not in register · " + out) if out else "not in register"
        print(f"{i:>3}  {r['init_date'] or '(no date)':<11} {r['batch']:<15} "
              f"{r['strain'][:22]:<22} {cov:<22} {coq:<17} {out}")

    if csv_out:
        with open(csv_out, "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(["seq", "release_date", "register_ref", "batch", "strain",
                        "icoa_scope", "ident_A", "ident_B", "ident_C", "foreign_matter",
                        "coq_initial", "coq_reissue", "reissue_basis_date",
                        "outsourced_outstanding", "in_register", "note"])
            for i, r in enumerate(table, 1):
                w.writerow([i, r["init_date"], r["ref"], r["batch"], r["strain"], r["scope"],
                            "required" if "A" in r["missing"] else "covered by CNP",
                            "required" if "B" in r["missing"] else "covered by CNP",
                            "required" if "C" in r["missing"] else "covered by CNP",
                            "required" if "FM" in r["missing"] else "covered by CNP",
                            "yes", "yes" if "reissue" in r["coq"] else "", r["re_date"],
                            "; ".join(r["blocked"]), r["in_register"],
                            "197-only" if r["only197"] else ""])
        print(f"\nwritten: {csv_out}")
    return 0


if __name__ == "__main__":
    out = sys.argv[sys.argv.index("--csv") + 1] if "--csv" in sys.argv else None
    sys.exit(main(out))
