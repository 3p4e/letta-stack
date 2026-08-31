#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Read the microbiology acceptance criteria the way the pharmacopoeia says to.

    python3 deliverables/qc_gap_analysis/apply_acceptance_criterion_corrections.py IN.xlsx OUT.xlsx

Evidence in `review/OOS_RECTIFICATION_2026-08-31.md`.

Nine TYMC cells in this register were amber-flagged as exceeding their limit.
**Five of them do. Four of them do not, and they should never have been flagged.**

## The defect is this campaign's own recurring one, turned on the campaign

Everything on this register that has needed correcting has had the same shape: a
value that is right, sitting in a column that asks a different question. The
Farmahem U column held mercury under arsenic. Two date cells held the receipt
date under a heading that says date of issue. Row 260 holds a foreign-matter
percentage twenty-five times inside its limit on a batch the laboratory failed,
because the operative half of that limit is not a number.

This is the same defect in the code that checks the register. `magnitude()`
answers *what number is this measurement* — and `4,2 × 10⁴ CFU/g` is 42 000,
correctly. It was then asked for the **acceptance criterion** as well, and an
acceptance criterion answers a different question: *what is the largest result
that still conforms.* For a microbial enumeration criterion written as a bare
power of ten, the pharmacopoeia says those two numbers are not the same.

    Ph. Eur. 5.1.4, in identical PDG-harmonised wording USP <1111>, and again in
    the "Interpretation of results" text of Ph. Eur. 2.6.12 / USP <61>:

        "The following interpretation should be applied:
         10¹ CFU: maximum acceptable count = 20;
         10² CFU: maximum acceptable count = 200;
         10³ CFU: maximum acceptable count = 2000, and so forth."

The series is **×2 per decade**. So `10⁴ CFU/g` means a maximum acceptable count
of **20 000**, and `10⁵` means **200 000**. The factor exists because the
criteria are order-of-magnitude limits and plate counting is intrinsically
imprecise; it is a rounding convention on the log scale, not a safety margin to
be stacked with anything else. It governs enumeration criteria only — TAMC, TYMC
and bile-tolerant Gram-negative bacteria in CFU/g — and never the "absence in
1 g / 25 g" criteria for specified micro-organisms, which are absolute.

## Three different multipliers were in play at once

| Where | Reads 10⁴ as | Authority |
|---|---|---|
| This register's specification row | 50 000 (×5) | none found |
| Ph. Eur. 5.1.4 / 2.6.12, USP <1111> | **20 000 (×2)** | the harmonised interpretation note |
| `validate_ecoa_limits.py`, and every review document | 10 000 (×1) | the literal power of ten |

**The ×5 is not mine and not a transcription slip.** `≤ 10⁵ (max 500 000)` and
`≤ 10⁴ (max 50 000)` are byte-identical in every one of the twelve workbooks in
the correction chain, including the owner-supplied baseline
`PP_Batch_Release_QC_Register_CORRECTED.xlsx`; no correction script has ever
touched row 5. The same phrasing is printed on the **Purely Plant in-house CoA
form itself** — `<10^5, max 500 000 CFU/g` and `<10^4, max 50 000 CFU/g`,
transcribed verbatim from the GG1024, HPA1024 and OPM1024 in-house release CoAs
in `ingestion/coa_track/letta-imb-coas/add_gg1024_rows.py`.

So the register faithfully copied a company document, and the company document
is wrong. **Correcting the in-house CoA form is a QA action and is not in this
script's gift** — it is recorded here so it does not get lost. What this script
does is stop the register from repeating it.

## What changes, and what does not

The specification row is rewritten to state the maximum acceptable count the
pharmacopoeia gives, with the authority named in the cell comment and the
superseded figure quoted there rather than erased.

Four amber flags are **withdrawn** — the values conform, and each cell gets a
comment saying so and why the flag was there:

| Row | Batch | Certificate | TYMC | vs 20 000 |
|---|---|---|---|---|
| 35 | GG1024_02 | `472/0863/25` | 1,9 × 10⁴ | 19 000 — inside, by 5 % |
| 38 | HPA1024_01 | `587/1066/25` | 1,5 × 10⁴ | 15 000 |
| 57 | GP0824_03 | `628/1129/25` | 1,2 × 10⁴ | 12 000 |
| 75 | CJ052501-1 | `949/1687/25` | 1,7 × 10⁴ | 17 000 |

Five **stand**, and for the first time say what is wrong with them. They were
amber with no comment at all — a flag that states nothing is only slightly
better than no flag:

| Row | Batch | Certificate | TYMC | vs 20 000 | Laboratory concluded |
|---|---|---|---|---|---|
| 21 | GG1024_01 | `320/0587/25` | 4,2 × 10⁴ | 2.10× over | ОДГОВАРА |
| 72 | OPM052501 | `904/1589/25` | 3,3 × 10⁴ | 1.65× over | ОДГОВАРА |
| 83 | GP052501 | `946/1684/25` | 3,6 × 10⁴ | 1.80× over | ОДГОВАРА |
| 88 | HPA052501 | `948/1686/25` | 2,6 × 10⁴ | 1.30× over | ОДГОВАРА |
| 101 | CJ062501-2 | `1032/1851/25` | 4,9 × 10⁴ | 2.45× over | ОДГОВАРА |

Row 35 is the closest call on the register: 19 000 against a maximum acceptable
count of 20 000 is inside by 5 %, and `1,9 × 10⁴` carries two significant
figures. It conforms as reported. Releasing on that margin is worth doing
against the raw plate count, and the cell comment says so.

**Row 122 is untouched.** Its TYMC of 200 sits against a limit of 10² printed on
its own certificate under `производителска спецификација` — a manufacturer's own
tighter specification, which the compendial ×2 reading does not automatically
govern. If that specification document does not itself invoke the Ph. Eur.
interpretation, 10² there is an absolute 100 and 200 fails. That is a judgment on
the wording of a document this register does not hold, and the existing comment
already sets it out.

**The four CBN cells are untouched.** They are accelerated 40 °C / 75 % RH
stability timepoints, they are amber rather than red, and their comments already
say that whether an accelerated-condition sample must meet a release limit is a
question for the stability protocol. Nothing there needs correcting.

Idempotent, and it refuses a workbook whose cells differ from what was verified.
"""
import sys

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

SHEET = "Batch Release QC"
TYMC, TAMC, GNB = 11, 10, 12          # columns K, J, L
SPEC_ROW = 5
AUTHOR = "QC acceptance-criterion rectification 31.08.2026"
AMBER = PatternFill("solid", fgColor="FFFAEDD4")
NOFILL = PatternFill(fill_type=None)

RULE = (
    "Ph. Eur. 5.1.4, in identical harmonised wording USP <1111>, and again in the "
    "\"Interpretation of results\" text of Ph. Eur. 2.6.12 / USP <61>:\n\n"
    "  \"The following interpretation should be applied: 10^1 CFU: maximum "
    "acceptable count = 20; 10^2 CFU: maximum acceptable count = 200; 10^3 CFU: "
    "maximum acceptable count = 2000, and so forth.\"\n\n"
    "The series is x2 per decade. It is a rounding convention on the log scale — "
    "the criteria are order-of-magnitude limits and plate counting is intrinsically "
    "imprecise — and not a margin to be stacked with anything else. It governs "
    "enumeration criteria only, never the \"absence in 1 g / 25 g\" criteria for "
    "specified micro-organisms."
)

PROVENANCE = (
    "The superseded text is not a transcription error. \"max 500 000\" and \"max "
    "50 000\" are byte-identical in every workbook in the correction chain, "
    "including the owner-supplied baseline, and the same figures are printed on the "
    "Purely Plant in-house CoA form itself (\"<10^5, max 500 000 CFU/g\" on the "
    "GG1024, HPA1024 and OPM1024 in-house release CoAs). No pharmacopoeial text "
    "uses a x5 multiplier for microbial limits.\n\n"
    "ACTION FOR QA, not for this register: the in-house CoA form states a maximum "
    "acceptable count 2.5 times the compendial one and needs correcting at source."
)

# column, old specification text, new specification text, extra note
SPECS = [
    (TAMC, "≤ 10⁵ (max 500 000)", "≤ 10⁵ CFU/g (max 200 000)",
     "TAMC acceptance criterion 10^5 CFU/g -> maximum acceptable count 200 000 "
     "CFU/g. The cell previously read \"max 500 000\"."),
    (TYMC, "≤ 10⁴ (max 50 000)", "≤ 10⁴ CFU/g (max 20 000)",
     "TYMC acceptance criterion 10^4 CFU/g -> maximum acceptable count 20 000 "
     "CFU/g. The cell previously read \"max 50 000\", which reported nine results "
     "over the limit where five are."),
    (GNB, "≤ 10⁴ CFU/g", "≤ 10⁴ CFU/g (max 20 000)",
     "Bile-tolerant Gram-negative bacteria, acceptance criterion 10^4 CFU/g -> "
     "maximum acceptable count 20 000 CFU/g. The cell stated no maximum, so the "
     "same workbook used two conventions in adjacent columns.\n\nSeparately, and "
     "not resolved here: certificates 1220/2171/25 and 1221/2172/25 print <= 10^2 "
     "for this parameter under a manufacturer specification, and the Purely Plant "
     "in-house CoA prints <10^2 CFU/g. This column is looser than both. Which "
     "applies to a release is a QC decision."),
]

# row, batch, certificate, printed value, counted value
WITHDRAW = [
    (35, "GG1024_02",  "472/0863/25", "1,9 x 10^4", 19000),
    (38, "HPA1024_01", "587/1066/25", "1,5 x 10^4", 15000),
    (57, "GP0824_03",  "628/1129/25", "1,2 x 10^4", 12000),
    (75, "CJ052501-1", "949/1687/25", "1,7 x 10^4", 17000),
]

STANDS = [
    (21,  "GG1024_01",  "320/0587/25",  "4,2 x 10^4", 42000),
    (72,  "OPM052501",  "904/1589/25",  "3,3 x 10^4", 33000),
    (83,  "GP052501",   "946/1684/25",  "3,6 x 10^4", 36000),
    (88,  "HPA052501",  "948/1686/25",  "2,6 x 10^4", 26000),
    (101, "CJ062501-2", "1032/1851/25", "4,9 x 10^4", 49000),
]

MAX_TYMC = 20000


def sp(n):
    """Thousands separated by a space, the way the certificates print counts.

    Applied to the number and to nothing else. An earlier draft ran
    `.replace(",", " ")` over the whole formatted string, which also ate the
    decimal comma out of "1,9 x 10^4" and the commas out of the prose around it
    — the same defect this script exists to correct, committed while correcting
    it. The guard did not catch it because the workbook was valid; only reading
    the output did.

    >>> sp(20000)
    '20 000'
    >>> sp(500)
    '500'
    """
    return f"{n:,}".replace(",", "\u00a0").replace("\u00a0", " ")

# Row 101 alone writes its exponent with a caret where the other eight use a
# superscript. Same number, and nothing turns on it — but a register that spells
# one value differently from its eight siblings invites the question of whether
# it came from somewhere else, and it did not.
NOTATION = (101, "4.9×10^4", "4.9×10⁴")

LEGEND = [
    "Microbial enumeration criteria (TAMC, TYMC, bile-tolerant GNB) are read under "
    "Ph. Eur. 5.1.4 / 2.6.12 and USP <1111>: an acceptance criterion of 10ⁿ CFU/g "
    "means a maximum acceptable count of 2 × 10ⁿ. So ≤ 10⁴ conforms up to 20 000 "
    "CFU/g and ≤ 10⁵ up to 200 000. This does not apply to Salmonella or E. coli, "
    "whose criteria are absolute.",
    "Amber on a microbiology cell means the count exceeds that maximum acceptable "
    "count. Five do, on certificates that all concluded ОДГОВАРА; each says so in "
    "its own comment.",
]


def _cur(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return "" if v is None else str(v).strip()


def _is_amber(cell):
    rgb = getattr(cell.fill.fgColor, "rgb", None)
    return cell.fill.patternType == "solid" and rgb == "FFFAEDD4"


def main(src, dst):
    wb = load_workbook(src)
    ws = wb[SHEET]
    log, skipped = [], []

    for col, old, new, note in SPECS:
        cur = _cur(ws, SPEC_ROW, col)
        letter = ws.cell(row=SPEC_ROW, column=col).column_letter
        if cur == new:
            skipped.append(f"{letter}{SPEC_ROW} already states the compendial maximum")
        elif cur != old:
            raise SystemExit(
                f"REFUSING {letter}{SPEC_ROW}: expected {old!r}, found {cur!r}. Not "
                f"the revision this was verified against.")
        else:
            ws.cell(row=SPEC_ROW, column=col).value = new
            log.append(f"{letter}{SPEC_ROW}  {old!r} -> {new!r}")
        cell = ws.cell(row=SPEC_ROW, column=col)
        if cell.comment is None:
            cell.comment = Comment(
                f"Corrected 31.08.2026.\n\n{note}\n\n{RULE}\n\n{PROVENANCE}", AUTHOR)

    row, old, new = NOTATION
    cur = _cur(ws, row, TYMC)
    if cur == new:
        skipped.append(f"K{row} already uses a superscript")
    elif cur != old:
        raise SystemExit(f"REFUSING K{row}: expected {old!r}, found {cur!r}.")
    else:
        ws.cell(row=row, column=TYMC).value = new
        log.append(f"K{row}  {old!r} -> {new!r}   notation only, same value")

    for row, batch, cert, printed, counted in WITHDRAW:
        cell = ws.cell(row=row, column=TYMC)
        if not _is_amber(cell):
            skipped.append(f"K{row} flag already withdrawn ({batch})")
        else:
            cell.fill = NOFILL
            log.append(f"K{row}  amber WITHDRAWN  {batch:<11} {printed} = "
                       f"{sp(counted)} <= {sp(MAX_TYMC)}")
        if cell.comment is None:
            margin = ""
            if counted > MAX_TYMC * 0.9:
                margin = (
                    "\n\nThis is the closest call on the register: inside by "
                    f"{100 * (1 - counted / MAX_TYMC):.0f} %, on a value reported to "
                    "two significant figures. It conforms as reported; confirming it "
                    "against the raw plate count before releasing on that margin is "
                    "cheap and worth doing.")
            cell.comment = Comment(
                "Amber flag WITHDRAWN 31.08.2026. This result conforms.\n\n"
                f"Certificate {cert} reports TYMC {printed} CFU/g = "
                f"{sp(counted)}, against an acceptance criterion of 10^4 CFU/g, whose "
                f"maximum acceptable count is {sp(MAX_TYMC)} CFU/g. The issuing "
                "laboratory concluded ОДГОВАРА and was right.\n\n"
                "The flag was raised by comparing the count against 10 000 — the "
                "literal power of ten, which is the measurement's reading of the "
                "notation and not the criterion's. The value was never in doubt; "
                "the ceiling it was compared against was.\n\n"
                f"{RULE}{margin}", AUTHOR)

    for row, batch, cert, printed, counted in STANDS:
        cell = ws.cell(row=row, column=TYMC)
        if not _is_amber(cell):
            cell.fill = AMBER
            log.append(f"K{row}  amber RESTORED   {batch}")
        if cell.comment is None:
            cell.comment = Comment(
                "Amber flag CONFIRMED 31.08.2026 against the pharmacopoeial "
                "interpretation. This result is out of specification.\n\n"
                f"Certificate {cert} reports TYMC {printed} CFU/g = "
                f"{sp(counted)}, against a maximum acceptable count of "
                f"{sp(MAX_TYMC)} CFU/g — {counted / MAX_TYMC:.2f}x over. The "
                "certificate concludes "
                "ОДГОВАРА.\n\n"
                "It remains out of specification on every reading that has any "
                "authority: over the literal 10^4, over the compendial 2 x 10^4, and "
                "over the certificate's own printed criterion. Only the register's "
                "superseded \"max 50 000\" made it pass, and that figure has none.\n\n"
                "This needs a deviation record. The laboratory's verdict is not a "
                "substitute for the arithmetic, and five certificates concluding "
                "ОДГОВАРА over their own limit is a finding about the laboratory's "
                "review step as much as about the batches.\n\n"
                f"{RULE}", AUTHOR)

    last = max(r for r in range(280, 320) if _cur(ws, r, 1))
    existing = {_cur(ws, r, 1) for r in range(290, last + 1)}
    for i, line in enumerate(LEGEND):
        if line in existing:
            skipped.append(f"legend line {i + 1} already present")
            continue
        r = last + 1 + i
        c = ws.cell(row=r, column=1, value=line)
        c.font = Font(name="Aptos Narrow", size=8, color="4A5A55")
        c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=26)
        ws.row_dimensions[r].height = 26
        log.append(f"A{r}  legend: acceptance-criterion interpretation")

    wb.save(dst)
    print(f"in : {src}\nout: {dst}\n")
    for line in log:
        print("  CHANGED  " + line)
    for s in skipped:
        print("  skipped  " + s)
    print(f"\n{len(log)} change(s), {len(skipped)} already applied.")
    print(f"\nTYMC: {len(STANDS)} results over the {sp(MAX_TYMC)} CFU/g maximum "
          f"acceptable count, {len(WITHDRAW)} flags withdrawn.")
    print("The Purely Plant in-house CoA form still prints 'max 500 000' and "
          "'max 50 000'. Correcting it is a QA action outside this register.")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 3:
        raise SystemExit(__doc__)
    sys.exit(main(sys.argv[1], sys.argv[2]))
