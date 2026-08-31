#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Check that each row's PDF link opens that row's own certificate.

    python3 deliverables/qc_gap_analysis/audit_register_pdf_links.py REGISTER.xlsx MAP.json

`repair_register_pdf_links.py` pointed every row at a file by matching the certificate
code against the Drive filename. This audits the result the other way round: take the
file id the cell actually links to, look up what document that id is, and ask whether it
is the document the row says it is.

That distinction is the whole point of the exercise, and the repair script's own docstring
said why: a dead link fails visibly and someone eventually notices; **a live link to the
wrong laboratory's certificate opens a real document with real numbers on it, and nothing
about it says it belongs to a different row.**

The audit is worth more now than it would have been before the reading campaign, because
275 of these files have since been opened and read. The link is no longer being checked
against a filename listing — it is being checked against documents whose contents are on
record in `review/*_page_reads_*.json`.

Two rules decide whether a file belongs to a row, because the corpus has two naming
conventions. Most certificates carry a code unique to the document, and the filename
carries the same code. The **in-house CoAs do not**: the register numbers them
`PP CoA #016`, `PP CoA #028` and so on, while every Drive file is named for the *form*,
`QCCoA 001v02` — 26 files share that string. For those, the distinguishing key is the
P-number, which the filename carries as its prefix and the register carries in column C.

An audit that knew only the first rule would report every repaired in-house row as
misdirected. That is worse than useless: a permanent block of known-false alarms is
exactly where a real misdirection goes unnoticed.

MAP.json is `{"<drive filename>": "<file id>"}`; the check inverts it.
"""
import json
import re
import sys
from collections import Counter

from openpyxl import load_workbook

SHEET = "Batch Release QC"
PNUM, CODE, INST, PDF = 3, 23, 25, 26
FIRST_DATA = 6

# Continuation rows carry no P-number: they belong to the batch block above.
INHERIT = {164: 163, 183: 182}

_CYR = "АВЕКМНОРСТУХЈЅІ"
_LAT = "ABEKMHOPCTYXJSI"


def fold(s):
    u = re.sub(r"\s*\([^)]*\)\s*$", "", str(s or "")).upper()
    u = re.sub(r"(?<=[^A-Za-z0-9])(GS|ГС)(?=[^A-Za-z0-9]|$)", "LOD", u)
    for a, b in zip(_CYR, _LAT):
        u = u.replace(a, b)
    return re.sub(r"[^A-Z0-9]", "", u)


def title_codes(title):
    """Every certificate code the filename could be carrying, folded.

    `<batch or P-number>_<code>, <date>_<lab>.pdf`, and the underscore is not a reliable
    boundary in either direction — the batch can contain one (`GRC102501_2_051-6-LoD-26`)
    and so can the code (`P050192_10802_2845-2 MK`). Every split is tried, as in
    repair_register_pdf_links.py; the wrong ones index strings no register code matches.
    """
    head = title.rsplit(".", 1)[0].split(",")[0]
    parts = head.split("_")
    out = set()
    for i in range(1, len(parts)):
        code = "_".join(parts[i:]).strip()
        code = re.sub(r"\s(MK|EN)$", "", code)
        out.add(fold(code))
    return {c for c in out if c}


def row_codes(cell):
    """The register's code cell, and any control-book number embedded in it."""
    out = {fold(cell)}
    out.update(fold(m) for m in re.findall(r"ППК\s*\d+", str(cell)))
    return {c for c in out if c}


def main(src, mapfile):
    by_id = {fid: title for title, fid in
             json.load(open(mapfile, encoding="utf-8")).items()}
    wb = load_workbook(src)
    ws = wb[SHEET]

    ok, unknown, nolink, wrong = 0, [], [], []
    for row in ws.iter_rows(min_row=FIRST_DATA):
        r = row[0].row
        code = ws.cell(row=r, column=CODE).value
        if not code or str(code).strip().lower() in ("", "n/a"):
            continue
        tgt = getattr(ws.cell(row=r, column=PDF).hyperlink, "target", None)
        if not tgt:
            nolink.append((r, str(code)[:40]))
            continue
        m = re.search(r"/d/([A-Za-z0-9_-]+)", tgt)
        fid = m.group(1) if m else None
        title = by_id.get(fid)
        if title is None:
            unknown.append((r, str(code)[:40], fid))
            continue
        if row_codes(code) & title_codes(title):
            ok += 1
        elif str(code).startswith("PP CoA") and "QCCoA" in title:
            # Form-numbered in-house CoA: the code cannot match, so check the P-number.
            pn = str(ws.cell(row=INHERIT.get(r, r), column=PNUM).value or "").strip().upper()
            if pn and title.split("_", 1)[0].strip().upper() == pn:
                ok += 1
            else:
                wrong.append((r, str(code)[:34], title[:52]))
        else:
            wrong.append((r, str(code)[:34], title[:52]))

    n = ok + len(unknown) + len(nolink) + len(wrong)
    print(f"register: {src}\nmap     : {mapfile}\n")
    print(f"  rows with a certificate code   : {n}")
    print(f"  link opens this row's document : {ok}")
    print(f"  link opens a DIFFERENT document: {len(wrong)}")
    print(f"  link points outside the map    : {len(unknown)}")
    print(f"  no link at all                 : {len(nolink)}")

    if wrong:
        print("\n  MISDIRECTED — the row says one certificate, the file is another:")
        for r, c, t in wrong:
            print(f"     r{r:<4} {c:<34} -> {t}")
    if unknown:
        print("\n  linked to a file not in the map:")
        for r, c, f in unknown[:20]:
            print(f"     r{r:<4} {c:<40} {f}")
    if nolink:
        print("\n  no link, by code:")
        for c, k in Counter(c for _, c in nolink).most_common(10):
            print(f"     {k:>3}  {c}")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 3:
        raise SystemExit(__doc__)
    sys.exit(main(sys.argv[1], sys.argv[2]))
