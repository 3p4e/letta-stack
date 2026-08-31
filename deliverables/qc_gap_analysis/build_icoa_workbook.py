#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Write the iCoA issuance register as a workbook in the release register's own style.

    python3 deliverables/qc_gap_analysis/build_icoa_workbook.py OUT.xlsx

Deliberately a separate file rather than a fourth sheet in
`PP_Batch_Release_QC_Register_AC_2026-08-31.xlsx`. That workbook is the end of a
chain that `replay_correction_chain.py` re-derives from the owner's original and proves
byte-for-byte; adding a sheet to it would mean the released register and the register
that chain reproduces are no longer the same file. This one is generated from
`icoa_issuance_register_2026-08-31.csv` and can be regenerated at any time.

Typography, header bands and flag colours follow the release register so the two read as
one set: Arial, the dark title band on row 4, the pale sub-header on row 5, data from
row 6, and the same amber and red the register's legend already defines.
"""
import csv
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SRC = "deliverables/qc_gap_analysis/icoa_issuance_register_2026-08-31.csv"

INK, SUB = "FF16232B", "FF5A6E75"
BAND, SUBBAND = "FF16232B", "FFE7ECEA"
ZEBRA, AMBER, RED, GREEN = "FFF6F8F7", "FFFAEDD4", "FFF9DEDB", "FFE6F1EB"
THIN = Side(style="thin", color="FFD6DEDB")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = [
    ("Seq",              6,  "release order"),
    ("Release date",     13, "latest non-197 certificate"),
    ("Reg. ref",         9,  "row in the release register"),
    ("Batch",            17, ""),
    ("Strain",           21, ""),
    ("iCoA scope",       17, "what the in-house CoA must cover"),
    ("Ident A",          15, "macroscopy = appearance, 2.8.23"),
    ("Ident B",          15, "microscopy, 2.8.23"),
    ("Ident C",          15, "TLC, 2.8.23"),
    ("Foreign matter",   15, "2.8.2"),
    ("CoQ initial",      11, "one per batch"),
    ("CoQ reissue",      11, "after the 197-series"),
    ("Reissue basis",    13, "date of the re-analysis"),
    ("Outsourced outstanding", 24, "CoQ cannot be compiled until this arrives"),
    ("In register",      11, ""),
    ("Note",             13, ""),
    ("iCoA number",      16, "fill in on issue"),
    ("iCoA date",        12, "fill in on issue"),
    ("Issued by",        16, "fill in on issue"),
    ("Reference in CoQ", 20, "fill in when added to the CoQ"),
]

SUMMARY = [
    ("CoQ documents predicted", "", ""),
    ("Initial release, one per batch", 81, ""),
    ("Reissue after a Farmahem 197-series re-analysis", 21,
     "14 carry earlier testing; 7 carry the 197 pair and nothing else"),
    ("Total CoQ documents", 102, ""),
    ("", "", ""),
    ("iCoA documents required", 81, "one per batch, not one per CoQ"),
    ("Full panel — Ident A + B + C + foreign matter", 69, ""),
    ("Ident C only — CNP Ph. Eur. 11.5 form covers the rest", 12, ""),
    ("", "", ""),
    ("Missing, counted over the 102 CoQ documents", "", ""),
    ("Ident A — macroscopy / appearance", 90, "69 batches"),
    ("Ident B — microscopy", 90, "69 batches"),
    ("Ident C — TLC", 102, "81 batches — no laboratory has ever performed it"),
    ("Foreign matter", 90, "69 batches"),
    ("", "", ""),
    ("CoQ blocked by outstanding outsourced testing", 39,
     "38 register batches without microbiology, plus GG1024 which has no register block"),
]

NOTES = [
    "Why 81 iCoAs against 102 CoQ documents. Identity and foreign matter are properties of the "
    "material, determined once. A CoQ reissued because the cannabinoids and mycotoxins were "
    "re-analysed covers the same batch, so it references the iCoA already issued rather than "
    "triggering a second one.",
    "",
    "Where the scope split comes from. CNP changed its certificate form in mid-2026. The older "
    "DAB form reports loss on drying and cannabinoids only; the Ph. Eur. 11.5 form adds "
    "identification (macroscopy and microscopy) and foreign matter. Twelve certificates are on "
    "the newer form — ППК26110-26119, ППК26127, ППК26128 — and all 73 CNP certificates were read "
    "off their own pages on 31.08.2026, so this is not inferred from a filename or a parse.",
    "",
    "Identification C is on every iCoA because no laboratory has ever performed it. Not on the "
    "older form and not on the Ph. Eur. form either: those twelve pages print Идентификација — "
    "Макроскопија, Микроскопија and stop. It is discharged in-house by risk analysis with "
    "scientific justification, using the qualitative determination from the HPLC assay.",
    "",
    "FB032601 (seq 72) needs Ident C only, like its eleven siblings, and its CoQ still cannot be "
    "issued. The foreign matter CNP does cover reads 0.08 % against a 2.00 % maximum and is "
    "marked Не одговара — cannabis seed present. A blocker, not a gap; no in-house work changes it.",
    "",
    "Seven batches marked 197-only — P060152, P060212, P060242, P060332, P060352, P060382, "
    "P060402 — carry nothing but the 197 pair: no potency, no microbiology, no contaminants, no "
    "earlier certificate at all. Calling that a reissue asserts a first CoQ with nothing behind "
    "it. Their names are P-numbers rather than cultivation batch codes and their strains all "
    "appear on earlier bulk batches, which reads like packaged lots drawn from released bulk. "
    "The register does not say which, so the question is left open and they are anchored on the "
    "197 date to keep them in sequence.",
    "",
    "GG1024 (seq 81) is not in the release register — one certificate on file, no chronological "
    "anchor.",
    "",
    "An iCoA does not unblock a CoQ on its own. The Outsourced outstanding column names the "
    "panel a batch is still waiting on; 37 of the 38 are 2026 batches, consistent with testing "
    "in flight rather than testing skipped.",
    "",
    "The four columns at the right are empty by design — they are where issuance is recorded: "
    "the iCoA number, its date, who issued it, and the reference it is given in the CoQ.",
]


def style_header(ws, ncol, title, subtitle):
    ws.cell(row=1, column=1, value=title).font = Font("Arial", 15, bold=True, color=INK)
    ws.cell(row=2, column=1, value=subtitle).font = Font("Arial", 9, color=SUB)
    for i, (name, w, note) in enumerate(COLS[:ncol], 1):
        h = ws.cell(row=4, column=i, value=name)
        h.font = Font("Arial", 9, bold=True, color="FFFFFFFF")
        h.fill = PatternFill("solid", fgColor=BAND)
        h.alignment = Alignment("center", "center", wrap_text=True)
        s = ws.cell(row=5, column=i, value=note)
        s.font = Font("Arial", 8, bold=True, color=INK)
        s.fill = PatternFill("solid", fgColor=SUBBAND)
        s.alignment = Alignment("center", "center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 30
    ws.row_dimensions[5].height = 26


def main(out):
    rows = list(csv.DictReader(open(SRC, encoding="utf-8")))
    wb = Workbook()

    ws = wb.active
    ws.title = "iCoA Issuance Register"
    style_header(ws, len(COLS),
                 "Purely Plant GmbH — in-house CoA issuance register",
                 "One iCoA per batch, in release order. Each row names the CoQ documents that "
                 "will reference it and the parameters it must carry.")
    for n, r in enumerate(rows):
        i = 6 + n
        # FB032601's foreign matter IS reported by CNP — and it failed. Leaving the cell
        # saying only "covered by CNP" makes the red fill carry the whole meaning, and a
        # fill does not survive a copy-paste, a CSV export or a printout in black and white.
        failed = r["batch"] == "FB032601"
        fm = r["foreign_matter"] + (" — FAILED" if failed else "")
        vals = [
            int(r["seq"]), r["release_date"] or "no date", r["register_ref"], r["batch"],
            r["strain"], "Ident C only" if r["ident_A"] != "required" else "Full panel",
            r["ident_A"], r["ident_B"], r["ident_C"], fm,
            "yes", r["coq_reissue"], r["reissue_basis_date"],
            r["outsourced_outstanding"], r["in_register"], r["note"], "", "", "", "",
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font = Font("Arial", 10, color=INK)
            cell.border = BOX
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center" if c in (1, 3, 7, 8, 9, 10, 11, 12, 15) else "left",
                                       wrap_text=c in (14, 16))
            if c in (7, 8, 9, 10):
                if failed and c == 10:
                    cell.fill = PatternFill("solid", fgColor=RED)
                    cell.font = Font("Arial", 10, bold=True, color="FF9E2A2A")
                elif v == "required":
                    cell.fill = PatternFill("solid", fgColor=AMBER)
                else:
                    cell.fill = PatternFill("solid", fgColor=GREEN)
            elif c >= 17:
                cell.fill = PatternFill("solid", fgColor="FFFFFFFF")
            elif n % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ZEBRA)
        if failed:
            ws.cell(row=i, column=10).comment = None
            ws.cell(row=i, column=14).value = (
                "FOREIGN MATTER FAILED — 0.08 % against max 2.00 %, marked Не одговара "
                "(cannabis seed present). Nothing is missing; the CoQ cannot be issued as "
                "conforming. " + (r["outsourced_outstanding"] or ""))
    ws.freeze_panes = "F6"
    ws.auto_filter.ref = f"A5:{get_column_letter(len(COLS))}{5+len(rows)}"

    leg = 6 + len(rows) + 2
    ws.cell(row=leg, column=1, value="LEGEND").font = Font("Arial", 9, bold=True, color=INK)
    for k, t in enumerate([
        "Amber — the parameter must be covered by the in-house CoA",
        "Green — the parameter is reported on this batch's CNP Ph. Eur. 11.5 certificate",
        "Red — the parameter is reported and FAILED; no in-house work resolves it",
        "The four right-hand columns are blank by design: iCoA number, date, issuer, and the "
        "reference given to it in the CoQ.",
    ], 1):
        ws.cell(row=leg + k, column=1, value=t).font = Font("Arial", 9, color=SUB)

    s = wb.create_sheet("Summary")
    s.cell(row=1, column=1, value="What this register counts").font = Font("Arial", 15, bold=True, color=INK)
    s.cell(row=2, column=1, value="Generated from icoa_issuance_register_2026-08-31.csv").font = Font("Arial", 9, color=SUB)
    for i, (label, n, note) in enumerate(SUMMARY, 4):
        a = s.cell(row=i, column=1, value=label)
        a.font = Font("Arial", 10, bold=not str(n).isdigit() and bool(label), color=INK)
        b = s.cell(row=i, column=2, value=n)
        b.font = Font("Arial", 11, bold=True, color=INK)
        b.alignment = Alignment(horizontal="center")
        c = s.cell(row=i, column=3, value=note)
        c.font = Font("Arial", 9, color=SUB)
    s.column_dimensions["A"].width = 52
    s.column_dimensions["B"].width = 9
    s.column_dimensions["C"].width = 72
    r0 = 4 + len(SUMMARY) + 2
    s.cell(row=r0, column=1, value="NOTES").font = Font("Arial", 9, bold=True, color=INK)
    for i, t in enumerate(NOTES, 1):
        cell = s.cell(row=r0 + i, column=1, value=t)
        cell.font = Font("Arial", 9, color=INK if t else SUB)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        s.merge_cells(start_row=r0 + i, start_column=1, end_row=r0 + i, end_column=3)
        if t:
            s.row_dimensions[r0 + i].height = max(14, 12 * (len(t) // 118 + 1))

    wb.save(out)
    print(f"{out}\n  {len(rows)} iCoA rows, {len(COLS)} columns, 2 sheets")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1] if len(sys.argv) > 1 else
                  "deliverables/qc_gap_analysis/PP_iCoA_Issuance_Register_2026-08-31.xlsx"))
