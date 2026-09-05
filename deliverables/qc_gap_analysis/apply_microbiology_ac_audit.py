#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Chain step 15 — the microbiology acceptance criteria, audited parameter by parameter.

    python3 deliverables/qc_gap_analysis/apply_microbiology_ac_audit.py IN.xlsx OUT.xlsx

Asked to pinpoint every microbiology acceptance-criterion error and say whether TAMC
should read 10⁵. **It should, and it does.** Every microbiology criterion on the
register's specification row was re-derived from the source documents:

| № | Parameter | Register spec row | Verdict |
|---|---|---|---|
| 9.1 | TAMC | `≤ 10⁵ CFU/g (max 200 000)` | correct — 44 of 47 certificates print `≤ 10⁵` |
| 9.2 | TYMC | `≤ 10⁴ CFU/g (max 20 000)` | correct — 44 of 47 print `≤ 10⁴` |
| 9.3 | Bile-tolerant GNB | `≤ 10⁴ CFU/g (max 20 000)` | value correct — 41 of 48 print `≤ 10⁴`; **the column heading's unit is wrong** |
| 9.4 | Salmonella | `Absent` | correct, and absolute — never a counted criterion |
| 9.5 | E. coli | `Absent` | correct, and absolute |

So **no result is out of specification because of a wrong criterion, and no TAMC result
is over any reading of its criterion** — the highest is 5.1 × 10⁴ against a maximum
acceptable count of 200 000. All nine microbiology exceedances in this register are
TYMC, and TYMC is `≤ 10⁴` on the certificates, in QCSP 001 v.03 and on the CoQ master
template alike. Raising TYMC to 10⁵ would clear five real failures by rewriting the
specification around them, and nothing in any source document supports it.

What this step changes, both purely documentary:

**1 · The bile-tolerant GNB column heading printed the wrong unit.** `Bile-tolerant GNB
/1 g` — but `/1 g` is the *absence* qualifier belonging to E. coli. Bile-tolerant
gram-negative bacteria are enumerated, and every certificate reports them in `CFU/g`.
The heading is what `acceptance_limit()` reads to decide a criterion is a count, so a
unit that says "absence" on an enumerated parameter is not cosmetic.

**2 · Ochratoxin A's criterion was still `per PP spec`.** QCSP 001 v.03 settles it at
`≤ 20 µg/kg` (Ph. Eur. 2.8.22). The register's one detected value, 2.06 µg/kg, conforms
with room to spare; the column simply never carried the number.

What this step **records and does not change**, because each is a QA determination on a
signed document rather than a transcription defect:

**A · Purely Plant's own in-house CoA form multiplies by five.** It prints
`<10^5, max 500 000 CFU/g` and `<10^4, max 50 000 CFU/g`. Ph. Eur. 5.1.4 — and USP
<1111> in identical harmonised wording — give ×2 per decade: 200 000 and 20 000. Three
documents carry the ×5 maxima (the GG1024, HPA1024 and OPM1024 in-house CoAs). No
disposition turns on it: their TAMC values are 1.1 × 10⁴, 5.2 × 10⁴ and 1.3 × 10⁴ and
their TYMC values < 10, 7 × 10² and 7 × 10², all conforming under either factor. **The
form is still wrong and issues on every batch it is used for.**

**B · The HPA1024 in-house CoA prints a criterion with no exponent** — `<10^ CFU/g` for
bile-tolerant GNB. Unreadable as written.

**C · The category is C, and every criterion in this register is the category C
criterion.** The owner confirmed on 31.08.2026 that all microbiological purity testing
is Ph. Eur. 5.1.8 **category C**, and the primary documents say the same thing in their
own words. Seventeen IPH certificates name the category on the parameter line —
`Microbiological quality (Ph.Eur. 5.1.8 cat. C) — Total aerobic microbial count (TAMC)
≤ 10^5 CFU/g`, `… TYMC ≤ 10^4 CFU/g`, `… Bile-tolerant gram-negative bacteria ≤ 10^4
CFU/g` — and Purely Plant's own in-house CoA form writes `Total aerobic microbial count
(TAMC) — Ph. Eur. 5.1.8, category C`. So:

| № | Parameter | Category C | This register |
|---|---|---|---|
| 9.1 | TAMC | `≤ 10⁵ CFU/g` | `≤ 10⁵ CFU/g` ✓ |
| 9.2 | TYMC | `≤ 10⁴ CFU/g` | `≤ 10⁴ CFU/g` ✓ |
| 9.3 | Bile-tolerant gram-negative | `≤ 10⁴ CFU/g` | `≤ 10⁴ CFU/g` ✓ |
| 9.4 | Salmonella | absence / 25 g | `Absent` ✓ |
| 9.5 | E. coli | absence / 1 g | `Absent` ✓ |
| 9.6 | P. aeruginosa | absence / 1 g, upon request | upon request ✓ |
| 9.7 | S. aureus | absence / 1 g, upon request | upon request ✓ |

**This withdraws a finding of my own.** `apply_qcsp001_reconciliation.py` (chain step
14) recorded, in the cell comment on the four undetermined TYMC results, that "Ph. Eur.
5.1.8 Category C is TYMC 10²" and that QCSP 001 therefore contradicted itself by
printing `cat. C` beside `10⁴`. **That was wrong, and it was wrong from memory rather
than from the documents.** The documents in this corpus pair `5.1.8 cat. C` with 10⁵ /
10⁴ / 10⁴ seventeen times and never once with 10². The label and the numbers agree;
there is no conflict to resolve, and no result's disposition ever turned on it. The
comment on step 14 stands as written in the workbook it produced, superseded here.

What survives from that finding is the smaller half: **`1220/2171/25` and `1221/2172/25`
(TAMC `≤ 10⁴`, TYMC `≤ 10²`, GNB `≤ 10²`), `406/0744/25` (TAMC `≤ 10³`, TYMC `≤ 10²`),
`587/1066/25` (GNB `≤ 10²`) and `627/1128/25` (GNB `≤ 10¹`) apply criteria tighter than
category C to a category C product.** Every result on them conforms anyway — PM072501's
TAMC 110 and TYMC 200, WC072501's TAMC 100 and TYMC 90, MB0824_04's <10 — and the
laboratory declared each conforming. It is an inconsistency in the laboratory's own
practice, conservative in direction, and worth one question to IPH rather than a
correction here.

**D · Ph. Eur. 2.6.12 has no categories, and QCSP 001 cites them there anyway.** It
prints `Ph. Eur. 2.6.12 cat. C` for TAMC and TYMC, `Ph. Eur. 2.6.31 cat. C` for
bile-tolerant GNB and Salmonella, and `Ph. Eur. 2.6.13 cat. C` for E. coli, P.
aeruginosa and S. aureus. Those chapters are the *test methods*; the categories and
their acceptance criteria live in **5.1.8** for herbal medicinal products for oral use.
The IPH certificates cite it correctly — `Microbiology (Ph.Eur. 5.1.8 cat C)` for the
criterion, `Ph.Eur. 2.6.12` for the method. QCSP 001's method column should read
`Ph. Eur. 2.6.12 · 5.1.8 cat. C` and its siblings likewise. A citation defect on the
signed specification, not a numeric one: **every acceptance criterion it states for
group 9 is the correct category C criterion.**
"""
import sys

from openpyxl import load_workbook
from openpyxl.comments import Comment

SHEET = "Batch Release QC"
HEADER_ROW, SPEC_ROW = 4, 5

# cell, before, after, comment
EDITS = [
    ("L4", "Bile-tolerant GNB /1 g", "Bile-tolerant GNB CFU/g",
     "Unit corrected 31.08.2026.\n\n"
     "The heading read \"/1 g\", which is the absence qualifier belonging to E. coli "
     "(absence in 1 g) and to Salmonella (absence in 25 g). Bile-tolerant gram-negative "
     "bacteria are ENUMERATED, not tested for absence: all 48 certificates in the corpus "
     "report them in CFU/g, and this column's own criterion is a count.\n\n"
     "Not cosmetic: acceptance_limit() reads the column heading to decide whether a bare "
     "power of ten is an enumeration criterion (and so carries Ph. Eur. 5.1.4's maximum "
     "acceptable count of 2 x 10^n) or an absolute one. A heading that says \"absence\" "
     "on an enumerated parameter argues for the wrong reading of its own limit."),
    ("Q5", "per PP spec", "≤ 20 µg/kg",
     "Criterion supplied 31.08.2026 from QCSP 001 v.03, determination 10.3:\n\n"
     "    Ochratoxin A | Охратоксин A    Ph. Eur. 2.8.22    <= 20 ug/kg\n\n"
     "The column had carried \"per PP spec\" since the register was built, with the "
     "number nowhere on the sheet. The signed product specification has it.\n\n"
     "One value in this column is a detection: 2.06 ug/kg on 163/0271/25, amber-flagged "
     "as DETECTED, >LOQ. Against 20 ug/kg it conforms with room to spare, and the amber "
     "flag stands as a detection notice, not a failure.\n\n"
     "Aflatoxin B1 (<= 2) and aflatoxins sum (<= 4) were already stated and match "
     "QCSP 001 exactly."),
]

# Recorded on the criteria that are RIGHT, so the next reader does not re-open them.
CONFIRMED = {
    "J5": ("TAMC — criterion confirmed 31.08.2026, unchanged.\n\n"
           "Ph. Eur. 5.1.8 CATEGORY C, confirmed by the owner 31.08.2026 and named on "
           "the certificates themselves: 17 IPH certificates print the parameter as "
           "\"Microbiological quality (Ph.Eur. 5.1.8 cat. C) - Total aerobic microbial "
           "count (TAMC)  <= 10^5 CFU/g\", and Purely Plant's in-house CoA form writes "
           "\"TAMC - Ph. Eur. 5.1.8, category C\". 44 of the 47 certificates reporting "
           "TAMC print \"<= 10^5 CFU/g\", as does QCSP 001 v.03 (det. 9.1) and the CoQ "
           "master template. Ph. Eur. 5.1.4 reads an "
           "enumeration criterion of 10^n as a maximum acceptable count of 2 x 10^n, so "
           "10^5 conforms up to 200 000 CFU/g.\n\n"
           "NO TAMC RESULT IN THIS REGISTER IS OVER ANY READING OF THIS CRITERION. The "
           "highest is 5.1 x 10^4 (GG1024_01, 320/0587/25) — half the printed 10^5 and "
           "a quarter of the maximum acceptable count. Where a batch card shows one "
           "determination over limit beside a TAMC and a TYMC value, the TYMC is the "
           "one over.\n\n"
           "Three certificates apply a tighter category (1220/2171/25 and 1221/2172/25 "
           "at <= 10^4, 406/0744/25 at <= 10^3). Their own results — 110, 100 and <10 — "
           "conform to those too."),
    "K5": ("TYMC — criterion confirmed 31.08.2026, unchanged at 10^4.\n\n"
           "Ph. Eur. 5.1.8 CATEGORY C, confirmed by the owner 31.08.2026: 17 IPH "
           "certificates name it on the parameter line (\"Microbiology (Ph.Eur. 5.1.8 "
           "cat C) - Total combined yeasts/moulds count (TYMC)  <= 10^4 CFU/g\"). An "
           "earlier note of mine on the four undetermined results claimed category C "
           "was TYMC 10^2 and that QCSP 001 contradicted itself; that claim came from "
           "memory, the documents say otherwise seventeen times, and it is WITHDRAWN. "
           "No disposition ever turned on it.\n\n"
           "44 of 47 certificates print \"<= 10^4 CFU/g\"; so do QCSP 001 v.03 (det. "
           "9.2) and the CoQ master template. It is NOT 10^5, and raising it there is "
           "the one change this audit refuses: five release results — GG1024_01 4.2 x "
           "10^4, OPM052501 3.3 x 10^4, GP052501 3.6 x 10^4, HPA052501 2.6 x 10^4, "
           "CJ062501-2 4.9 x 10^4 — exceed even the pharmacopoeial maximum acceptable "
           "count of 20 000, and would all be cleared by rewriting the specification "
           "around them. Four more sit between 10 000 and 20 000 and are recorded as "
           "undetermined pending QA's reading of QCSP 001.\n\n"
           "Every microbiological exceedance in this register is in this column. TAMC, "
           "bile-tolerant GNB, Salmonella and E. coli have none."),
    "L5": ("Bile-tolerant gram-negative bacteria — criterion confirmed 31.08.2026 at "
           "10^4 CFU/g: Ph. Eur. 5.1.8 category C, named as such on four IPH "
           "certificates (\"Microbiological quality (Ph.Eur. 5.1.8 cat. C) - "
           "Bile-tolerant gram-negative bacteria  <= 10^4 CFU/g\"), matching 41 of the "
           "48 certificates and QCSP 001 v.03 (det. 9.3). Seven certificates apply a "
           "tighter limit (five at 10^2, one at 10^1, one at 10^3); every result in "
           "this column is bounded (<10, <10^2 and >10, <10^3 and >10^2) and conforms "
           "under all of them.\n\n"
           "The column heading's unit was corrected in the same pass — it read \"/1 g\", "
           "which belongs to E. coli."),
    "M5": ("Salmonella — absence in 25 g, per QCSP 001 v.03 det. 9.4, Ph. Eur. 5.1.8 "
           "category C and every certificate. Confirmed 31.08.2026.\n\n"
           "An absence criterion is absolute: Ph. Eur. 5.1.4's maximum-acceptable-count "
           "series applies to ENUMERATION criteria only and must never be applied here. "
           "acceptance_limit() returns no numeric limit for this column by design."),
    "N5": ("Escherichia coli — absence in 1 g, per QCSP 001 v.03 det. 9.5, Ph. Eur. "
           "5.1.8 category C and every certificate. Confirmed 31.08.2026. Absolute, as "
           "for Salmonella.\n\n"
           "Determinations 9.6 (P. aeruginosa) and 9.7 (S. aureus) carry the same "
           "absence-in-1-g criterion in QCSP 001 and are marked upon request: they are "
           "not category C release criteria and no release turns on them. Two "
           "certificates report them anyway, both absent."),
}


def main(src, dst):
    wb = load_workbook(src)
    ws = wb[SHEET]
    applied = again = 0

    for cell, before, after, note in EDITS:
        got = str(ws[cell].value or "").strip()
        if got == after:
            again += 1
            print(f"  = {cell}  already {after!r}")
            continue
        if got != before:
            sys.exit(f"REFUSED: {cell} reads {got!r}, expected {before!r} or {after!r}")
        ws[cell] = after
        ws[cell].comment = Comment(note, "QC review")
        applied += 1
        print(f"  + {cell}  {before!r} -> {after!r}")

    for cell, note in CONFIRMED.items():
        if ws[cell].comment and "onfirmed 31.08.2026" in ws[cell].comment.text:
            again += 1
            continue
        old = ws[cell].comment.text if ws[cell].comment else ""
        ws[cell].comment = Comment((old + "\n\n" if old else "") + note, "QC review")
        applied += 1
        print(f"  + {cell}  criterion confirmed, note appended")

    wb.save(dst)
    print(f"\n{applied} applied, {again} already in place -> {dst}")
    return 0


if __name__ == "__main__":
    if len(sys.argv) != 3:
        sys.exit(__doc__)
    sys.exit(main(sys.argv[1], sys.argv[2]))
