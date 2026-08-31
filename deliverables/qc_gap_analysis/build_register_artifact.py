#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build the published release-register artifact from the repository.

    python3 deliverables/qc_gap_analysis/build_register_artifact.py REGISTER.xlsx OUT.html

Until now the artifact was assembled from a JSON file that lived only in a
scratch directory, which meant the published page could be *read* but not
*rebuilt* — and a page that cannot be rebuilt cannot be corrected. Everything it
needs is in the repository, so this reads it from there:

| Part | Source |
|---|---|
| columns, criteria, rows, blocks, legend, flags, links | the register workbook |
| which certificates are stability timepoints | the workbook's own `Stability Testing Programme` sheet |
| which rows were page-verified, and in which pass | the five `review/*_page_reads_2026-08-31.json` files |
| the iCoA issuance register | `icoa_issuance_register_2026-08-31.csv` |

Two things the page could not previously say, and now can.

**A stability timepoint is not a release failure.** Four accelerated
40 °C / 75 % RH samples sit above the CBN release limit; the page used to render
them in the same red as a genuine exceedance. The register files them on its own
stability sheet and says there, in its own subtitle, that its results are not
release results. Each row now carries `stability: true` and the page marks those
values *stability timepoint* rather than *over*.

**An acceptance criterion of 10ⁿ CFU/g is not the number 10ⁿ.** Ph. Eur. 5.1.4 /
2.6.12 and USP <1111> read it as a maximum acceptable count of 2 × 10ⁿ. The
page's `acceptanceLimit()` applies that, and prints the resulting count beside
the criterion so a reader can see which number a value was judged against.

The page carries its own shell base64-encoded so that it can republish itself
through the `artifact` capability: `publish()` needs a complete document, while
the Artifact tool takes a body fragment and wraps it, so the page has to wrap
itself. The round trip is proved here before the file is written — decoding the
shell back out and re-substituting must reproduce the page byte for byte.
"""
import base64
import csv
import json
import os
import re
import sys

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
TEMPLATE = os.path.join(HERE, "register_artifact_template.html")
ICOA_CSV = os.path.join(HERE, "icoa_issuance_register_2026-08-31.csv")
REVIEW = os.path.join(ROOT, "review")

SHEET = "Batch Release QC"
STABILITY_SHEET = "Stability Testing Programme"
HEADER_ROW, SPEC_ROW, FIRST_DATA = 4, 5, 6
FIRST_RESULT, LAST_RESULT = 5, 22        # columns E..V
REF, BATCH, PNUM, STRAIN = 1, 2, 3, 4
CODE, DATE, LAB, PDF = 23, 24, 25, 26

AMBER, RED = "FFFAEDD4", "FFF9DEDB"

# Each page-verification pass wrote what it read. A row is verified in a pass if
# its certificate code appears among that pass's keys.
PASSES = [
    ("cnp", "cnp_page_reads_2026-08-31.json"),
    ("farmahem", "farmahem_page_reads_2026-08-31.json"),
    ("iph_physchem", "iph_physchem_page_reads_2026-08-31.json"),
    ("microbiology", "microbiology_page_reads_2026-08-31.json"),
    ("residual", "residual_page_reads_2026-08-31.json"),
]

_CYR, _LAT = "АВЕКМНОРСТУХЈЅІ", "ABEKMHOPCTYXJSI"


def fold(s):
    """Homoglyphs onto Latin, a trailing bracketed note dropped, then everything
    but A-Z0-9 removed — the comparison key used throughout this campaign.

    >>> fold("ППК26033") == fold("ППК 26033 (month 6)")
    True
    >>> fold("320/0587/25") == fold("320-0587-25")
    True

    Farmahem writes its loss-on-drying series with the Macedonian abbreviation
    ГС, and the same series appears elsewhere as GS or LoD. The substitution has
    to run BEFORE the homoglyph pass — after it, ГС has already become GC and no
    longer matches — which is why it is the first thing here and not the last:

    >>> fold("100-2-ГС/26") == fold("100-2-GS/26") == fold("100-2-LoD/26")
    True
    """
    u = re.sub(r"\s*\([^)]*\)\s*$", "", str(s or "")).upper()
    u = re.sub(r"(?<=[^A-Za-z0-9])(GS|ГС)(?=[^A-Za-z0-9]|$)", "LOD", u)
    for a, b in zip(_CYR, _LAT):
        u = u.replace(a, b)
    return re.sub(r"[^A-Z0-9]", "", u)


def codes_in(cell):
    """Every certificate a code cell might be naming, folded.

    Six rows name two documents — `PP CoA #027 / ППК25370` — because the
    in-house CoA and the report its numbers came from are both recorded there.
    Splitting on `/` is not an option: half the certificate codes in this corpus
    contain one.

    Cyrillic П has no Latin homoglyph in the campaign's map and is dropped, so
    `ППК25370` folds to `K25370` — consistently on both sides of every
    comparison, which is all a comparison key has to be.

    >>> sorted(codes_in("PP CoA #027 / ППК25370"))
    ['K25370', 'PPCOA027K25370']
    """
    out = {fold(cell)}
    out.update(fold(m) for m in re.findall(r"ППК\s*\d+", str(cell or "")))
    return {c for c in out if c}


def _v(cell):
    return "" if cell.value is None else str(cell.value).strip()


def _flag(cell):
    rgb = getattr(cell.fill.fgColor, "rgb", None)
    if cell.fill.patternType != "solid":
        return None
    return {AMBER: "amber", RED: "red"}.get(rgb)


def build(path):
    wb = load_workbook(path)
    ws = wb[SHEET]

    stability = set()
    if STABILITY_SHEET in wb.sheetnames:
        for row in wb[STABILITY_SHEET].iter_rows(min_row=6):
            if row[5].value:
                stability |= codes_in(row[5].value)

    verified = {}
    for name, fn in PASSES:
        with open(os.path.join(REVIEW, fn), encoding="utf-8") as fh:
            for key in json.load(fh):
                verified.setdefault(fold(key), name)

    columns = {}
    for c in range(FIRST_RESULT, LAST_RESULT + 1):
        columns[ws.cell(row=HEADER_ROW, column=c).column_letter] = {
            "name": _v(ws.cell(row=HEADER_ROW, column=c)),
            "limit": _v(ws.cell(row=SPEC_ROW, column=c)),
        }

    rows, blocks, cur = [], [], None
    for r in range(FIRST_DATA, ws.max_row + 1):
        code = _v(ws.cell(row=r, column=CODE))
        ref = _v(ws.cell(row=r, column=REF))
        batch = _v(ws.cell(row=r, column=BATCH))
        if not code:
            if ref and ref.upper() == "LEGEND":
                break
            continue
        if batch:
            cur = {"ref": ref, "batch": batch,
                   "pnumber": _v(ws.cell(row=r, column=PNUM)),
                   "strain": _v(ws.cell(row=r, column=STRAIN)),
                   "hdr": r, "rows": []}
            blocks.append(cur)
        if cur is None:                       # a row before any block header
            continue
        cur["rows"].append(r)

        vals, flags, notes = {}, {}, {}
        for c in range(FIRST_RESULT, LAST_RESULT + 1):
            cell = ws.cell(row=r, column=c)
            vals[cell.column_letter] = _v(cell)
            f = _flag(cell)
            if f:
                flags[cell.column_letter] = f
        for c in (CODE, DATE):
            f = _flag(ws.cell(row=r, column=c))
            if f:
                flags[ws.cell(row=r, column=c).column_letter] = f
        # The verification notes are the campaign's audit trail and they live in
        # the workbook's cell comments — what was read, off which page, at what
        # magnification, and what was deliberately not changed. A register that
        # shows a corrected value without saying who corrected it and from what
        # is worth less than one that shows neither.
        for c in range(FIRST_RESULT, DATE + 1):
            cell = ws.cell(row=r, column=c)
            if cell.comment is not None:
                notes[cell.column_letter] = cell.comment.text.strip()

        link = ws.cell(row=r, column=PDF).hyperlink
        keys = codes_in(code)
        rec = {"row": r, "ref": cur["ref"], "batch": cur["batch"],
               "pnumber": cur["pnumber"], "strain": cur["strain"],
               "code": code, "date": _v(ws.cell(row=r, column=DATE)),
               "lab": _v(ws.cell(row=r, column=LAB)),
               "pdf": link.target if link else "", "vals": vals}
        if flags:
            rec["flags"] = flags
        if notes:
            rec["notes"] = notes
        hit = next((verified[k] for k in keys if k in verified), None)
        if hit:
            rec["verified"] = hit
        if keys & stability:
            rec["stability"] = True
        rows.append(rec)

    legend = []
    for r in range(ws.max_row, FIRST_DATA, -1):
        text = _v(ws.cell(row=r, column=1))
        if text:
            legend.append(text)
        if text.upper() == "LEGEND":
            break
    legend.reverse()

    # A batch whose certificate the issuing laboratory declared out of
    # specification cannot have a conforming CoQ issued, however complete its
    # iCoA panel is. That is a different kind of blocker from a missing test,
    # and the iCoA card has to say so or it reads as "one document away from
    # done". Read off the register's own red flag rather than restated here.
    failed = {}
    for rec in rows:
        if "red" in (rec.get("flags") or {}).values():
            failed[rec["batch"]] = (
                f"Certificate {rec['code']} was declared out of specification by "
                f"the issuing laboratory (register row {rec['row']}). Nothing is "
                f"missing here; the CoQ cannot be issued as conforming until that "
                f"is resolved.")

    icoa = []
    with open(ICOA_CSV, encoding="utf-8") as fh:
        for i, row in enumerate(csv.DictReader(fh), 1):
            p = {k: row.get("ident_" + k, "") == "required" for k in "ABC"}
            icoa.append({
                "seq": i, "date": row["release_date"], "ref": row["register_ref"],
                "batch": row["batch"], "strain": row["strain"],
                "scope": "".join(k for k in "ABC" if p[k])
                         + ("FM" if row["foreign_matter"] == "required" else ""),
                "p": p,
                "fm": row["foreign_matter"] == "required",
                "reissue": row["coq_reissue"] == "yes",
                "rdate": row["reissue_basis_date"],
                "block": row["outsourced_outstanding"],
                "inreg": row["in_register"] == "Y",
                "note": row.get("note", ""),
            })
            if row["batch"] in failed:
                icoa[-1]["failed"] = failed[row["batch"]]

    return {
        "source": os.path.basename(path),
        "title": _v(ws.cell(row=1, column=1)),
        "subtitle": _v(ws.cell(row=2, column=1)),
        "columns": columns, "legend": legend,
        "blocks": blocks, "rows": rows, "icoa": icoa,
    }


def assemble(data, out):
    """Substitute the shell into itself so the published page can republish."""
    shell = open(TEMPLATE, encoding="utf-8").read()
    for token in ("{{SHELL_B64}}", "{{REGISTER_JSON}}", "{{OVERLAY_JSON}}"):
        n = shell.count(token)
        if n != 1:
            raise SystemExit(f"{token} appears {n} times in the template — must be 1")

    reg = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    b64 = base64.b64encode(shell.encode("utf-8")).decode("ascii")
    doc = (shell.replace("{{SHELL_B64}}", b64)
                .replace("{{REGISTER_JSON}}", reg)
                .replace("{{OVERLAY_JSON}}", "{}"))

    m = re.search(r'<script type="text/plain" id="shell">([A-Za-z0-9+/=]+)</script>', doc)
    if not m:
        raise SystemExit("shell block not found in the built page")
    back = base64.b64decode(m.group(1)).decode("utf-8")
    rebuilt = (back.replace("{{SHELL_B64}}", m.group(1))
                   .replace("{{REGISTER_JSON}}", reg)
                   .replace("{{OVERLAY_JSON}}", "{}"))
    if rebuilt != doc:
        raise SystemExit("round trip does not reproduce the page")
    open(out, "w", encoding="utf-8").write(doc)
    return doc, reg


def main(src, out):
    data = build(src)
    doc, reg = assemble(data, out)
    json_path = os.path.splitext(out)[0] + ".json"
    with open(json_path, "w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False, indent=1)

    flagged = sum(len(r.get("flags", {})) for r in data["rows"])
    noted = sum(len(r.get("notes", {})) for r in data["rows"])
    print(f"source   {data['source']}")
    print(f"blocks   {len(data['blocks'])}")
    print(f"rows     {len(data['rows'])}  "
          f"({sum(1 for r in data['rows'] if r.get('verified'))} page-verified, "
          f"{sum(1 for r in data['rows'] if r.get('stability'))} stability timepoints)")
    print(f"flags    {flagged}\nnotes    {noted}")
    print(f"iCoA     {len(data['icoa'])}")
    print(f"data     {len(reg):>9,} chars")
    print(f"page     {len(doc):>9,} chars  "
          f"({len(doc.encode('utf-8')) / 1048576:.2f} MB)")
    print("round trip: the page rebuilds itself exactly")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 3:
        raise SystemExit(__doc__)
    sys.exit(main(sys.argv[1], sys.argv[2]))
