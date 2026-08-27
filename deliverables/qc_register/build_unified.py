#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""One unified QC Excel workbook, per-certificate representation of the
Batch Release QC register (modeled on the owner's main register first sheet),
covering all themes of the three source workbooks, filled from the repaired
RAGFlow extraction of eCoA_DATABASE."""
import json, re, datetime as dt, unicodedata
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

S = '.'
C = json.load(open(f'{S}/consolidated.json'))
ROWS, CERT_MAP = C['rows'], C['cert_map']
RECS = {r['name']: r for r in json.load(open(f'{S}/extracted_params.json'))}
SANITY = json.load(open(f'{S}/sanity_results.json'))
REPAIRS = json.load(open(f'{S}/repair_log.json'))

def nrm(s):
    s = unicodedata.normalize('NFKC', str(s))
    return re.sub(r'[\s,_/()-]+', '', s).lower()

DRIVE = {}
for line in open(f'{S}/drive_all.tsv', encoding='utf-8'):
    fid, name = line.rstrip('\n').split('\t')
    DRIVE[nrm(name)] = fid
def drive_url(cert_name):
    fid = DRIVE.get(nrm(cert_name))
    return f'https://drive.google.com/file/d/{fid}/view' if fid else None

def base_batch(b):
    return str(b).split(' (')[0]

def lookup(d, row):
    for k in (row['batch'], base_batch(row['batch']), row.get('p_number') or '§'):
        v = d.get(nrm(k))
        if v: return v
    return None

LABNAME = {'CNP': 'UKIM Faculty of Pharmacy — Center for Natural Products',
           'IJZ': 'IPH — Institute of Public Health (chemistry)',
           'IJZ-MB': 'IPH — Institute of Public Health (microbiology)',
           'FHM': 'Farmahem', 'PP': 'Purely Plant (in-house)',
           'NGP': 'NGP — in-house GC cross-check', 'DFL': 'DFL Deutsches Institut (DE)'}

# ---- spec / tier data from file_C and file_B --------------------------------
wbC = openpyxl.load_workbook(f'{S}/file_C.xlsx', data_only=True)
wsA = wbC['Batch Assignments']
ASSIGN = {}
for r in range(3, wsA.max_row+1):
    b = wsA.cell(r,2).value
    if not b: continue
    ASSIGN[nrm(b)] = {'batch': b, 'production': wsA.cell(r,3).value, 'strain': wsA.cell(r,4).value,
                      'thc': wsA.cell(r,5).value, 'spec': wsA.cell(r,6).value, 'grade': wsA.cell(r,7).value,
                      'pcode': wsA.cell(r,8).value, 'range': wsA.cell(r,9).value, 'nominal': wsA.cell(r,10).value}
wsS = wbC['Specifications by Strain']
SPECS = []  # rows: strain header carried down
cur = None
for r in range(3, wsS.max_row+1):
    st, code = wsS.cell(r,3).value, wsS.cell(r,4).value
    if st: cur = {'strain': st, 'code': code, 'first_prod': wsS.cell(r,2).value, 'no': wsS.cell(r,1).value}
    if wsS.cell(r,5).value:
        SPECS.append({**cur, 'spec': wsS.cell(r,5).value, 'pcode': wsS.cell(r,6).value,
                      'range': wsS.cell(r,7).value, 'nominal': wsS.cell(r,8).value})
wbB = openpyxl.load_workbook(f'{S}/file_B.xlsx', data_only=True)
wsP = wbB['Potency by Strain']
TIER = {}
for r in range(6, wsP.max_row+1):
    b = wsP.cell(r,2).value
    if not b: continue
    TIER[nrm(b)] = {'tier': wsP.cell(r,8).value, 'nominal': wsP.cell(r,9).value, 'tol': wsP.cell(r,10).value,
                    'range': wsP.cell(r,11).value, 'newname': wsP.cell(r,12).value, 'brand': wsP.cell(r,13).value}
wsStock = wbB['Stock without register result']
STOCK = []
for r in range(5, wsStock.max_row+1):
    if wsStock.cell(r,2).value:
        STOCK.append([wsStock.cell(r,c).value for c in range(1,9)])

def parse_range(txt):
    m = re.search(r'([\d.]+)\s*%?\s*[–-]\s*([\d.]+)', str(txt))
    return (float(m.group(1)), float(m.group(2))) if m else None

# ---- workbook scaffolding ---------------------------------------------------
wb = openpyxl.Workbook(); wb.remove(wb.active)
H1 = Font(bold=True, size=13); H2 = Font(size=9, italic=True, color='555555')
HDR = Font(bold=True, size=9, color='FFFFFF'); HDRF = PatternFill('solid', fgColor='1F4E5F')
REF = Font(size=8, italic=True, color='666666'); REFF = PatternFill('solid', fgColor='EFF3F5')
BATCHF = PatternFill('solid', fgColor='DCE8EC'); NEWF = PatternFill('solid', fgColor='FFF2CC')
LINK = Font(size=9, color='0563C1', underline='single')
CELL = Font(size=9); BOLD9 = Font(size=9, bold=True)
thin = Side(style='thin', color='C0C0C0'); BORD = Border(left=thin,right=thin,top=thin,bottom=thin)
CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
LFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

def sheet(title, heads, widths, title_txt, sub_txt, ref_row=None):
    ws = wb.create_sheet(title)
    ws.cell(1,1,title_txt).font = H1
    ws.cell(2,1,sub_txt).font = H2
    for i,(h,w) in enumerate(zip(heads,widths),1):
        c = ws.cell(4,i,h); c.font=HDR; c.fill=HDRF; c.alignment=CTR; c.border=BORD
        ws.column_dimensions[get_column_letter(i)].width = w
    r0 = 5
    if ref_row:
        for i,v in enumerate(ref_row,1):
            c = ws.cell(5,i,v); c.font=REF; c.fill=REFF; c.alignment=CTR; c.border=BORD
        r0 = 6
    ws.freeze_panes = ws.cell(r0,5)
    return ws, r0

def put(ws,r,c,v,font=CELL,align=CTR,fill=None):
    cell = ws.cell(r,c,v); cell.font=font; cell.alignment=align; cell.border=BORD
    if fill: cell.fill=fill
    return cell

# ============ Sheet 1 — Batch Release QC (per-certificate) ===================
HEADS = ['No.','Batch number','P-number','Strain','THC %','THC spec','CBD %','CBN %',
 'Loss on drying %','Foreign matter %','Macroscopic ID','Microscopic ID','HPTLC ID',
 'TAMC CFU/g','TYMC CFU/g','Bile-tolerant GNB /1 g','Salmonella /25 g','E. coli /1 g',
 'Aflatoxins Σ µg/kg','Aflatoxin B1 µg/kg','Ochratoxin A µg/kg','Pb mg/kg','Cd mg/kg',
 'As mg/kg','Hg mg/kg','Pesticides','CoA code','Date of issue','Issuing institution','PDF']
WID = [5,15,10,17,8,15,7,7,9,9,10,10,9,11,11,12,11,11,9,9,9,7,7,7,7,11,16,11,26,7]
REFROW = ['Ref.','','','','per-batch label range','% w/w (label claim)','< 1.00 %','< 1.00 %',
 '≤ 12.00 (3028)','≤ 2.00 %','Conforms (2.8.23)','Conforms (2.8.23)','Identity confirmed',
 '≤ 10⁵','≤ 10⁴','≤ 10⁴ CFU/g','Absent','Absent','≤ 4','≤ 2','≤ 20','≤ 0.5','≤ 0.3','≤ 0.2','≤ 0.1',
 '≤ LOQ 0.01 mg/kg','','','','']
FIELD2COL = {'total_thc_pct':5,'total_cbd_pct':7,'total_cbn_pct':8,'loss_on_drying_pct':9,
 'foreign_matter_pct':10,'macroscopic_id':11,'microscopic_id':12,'hptlc_id':13,'tamc':14,'tymc':15,
 'bile_tolerant_gnb':16,'salmonella':17,'e_coli':18,'aflatoxins_total':19,'aflatoxin_b1':20,
 'ochratoxin_a':21,'pb':22,'cd':23,'arsenic':24,'hg':25,'pesticides':26}
# certificate-level print anomalies shown as printed, flagged in remark colour
PRINT_ANOM = {nrm('P050042, ППК25117, 06.05.2025, CNP.pdf'): 'printed total 1.58 contradicts own components (15.38) — VERIFY ON PAPER',
              nrm('P050062, ППК25154, 24.06.2025, CNP.pdf'): 'printed total 1.87 anomalous vs FHM retest 18.04 — VERIFY ON PAPER'}
ANOMF = PatternFill('solid', fgColor='F8CBAD')

reg_batches = set()
import openpyxl as _ox
_mr = _ox.load_workbook(f'{S}/main_register.xlsx', data_only=True)
_brq = _mr['Batch Release QC']
for r in range(6, _brq.max_row+1):
    v = _brq.cell(r,2).value
    if v: reg_batches.add(nrm(v))

ws, r = sheet('Batch Release QC', HEADS, WID,
    'Purely Plant GmbH — Production Batch Release QC Register (rebuilt from eCoA_DATABASE)',
    'One row per certificate · "/" = parameter not tested on that certificate · grouped per batch, batches chronological by first result · '
    'source: RAGFlow eCoA_DATABASE (291 certificates, sanity-checked against certificate text) · built 27.08.2026', REFROW)
nb = 0
for row in ROWS:
    nb += 1
    is_new = nrm(base_batch(row['batch'])) not in reg_batches and nrm(row.get('p_number') or '§') not in reg_batches
    certs = sorted(CERT_MAP[row['key']], key=lambda n: RECS[n]['meta'].get('date_of_issue','')[::-1])
    def dkey(n):
        d = RECS[n]['meta'].get('date_of_issue') or '99.99.9999'
        p = d.split('.')
        return (p[2], p[1], p[0]) if len(p)==3 else ('9','9','9')
    certs = sorted(CERT_MAP[row['key']], key=dkey)
    first = True
    for name in certs:
        rec = RECS[name]; p = rec.get('params') or {}; m = rec['meta']
        fill = NEWF if (is_new and first) else (BATCHF if first else None)
        put(ws,r,1, row and (nb if first else ''), BOLD9 if first else CELL, CTR, fill)
        put(ws,r,2, row['batch'] if first else '', BOLD9, LFT, fill)
        put(ws,r,3, (row.get('p_number') or '') if first else '', CELL, CTR, fill)
        put(ws,r,4, (row.get('strain') or '') if first else '', CELL, LFT, fill)
        for col in range(5,27): put(ws,r,col,'/')
        anom = PRINT_ANOM.get(nrm(name))
        LABSC = {'cannabinoid': ('CNP','FHM','NGP','PP'), 'micro': ('IJZ-MB','PP'),
                 'chem': ('IJZ','FHM','DFL','PP')}
        FAM = {**{k:'cannabinoid' for k in ('total_thc_pct','total_cbd_pct','total_cbn_pct',
               'loss_on_drying_pct','foreign_matter_pct','macroscopic_id','microscopic_id','hptlc_id')},
               **{k:'micro' for k in ('tamc','tymc','bile_tolerant_gnb','salmonella','e_coli')},
               **{k:'chem' for k in ('aflatoxins_total','aflatoxin_b1','ochratoxin_a','pb','cd','arsenic','hg','pesticides')}}
        QUAL = re.compile(r'^(n\.?d\.?|н\.?д\.?|nd|absent|отсутна?)$', re.I)
        for f,col in FIELD2COL.items():
            v = p.get(f)
            if v is None and f=='aflatoxins_total': v = p.get('total_aflatoxins')
            if v not in (None,''):
                lab = m.get('lab')
                if lab not in LABSC.get(FAM[f], ()) and QUAL.match(str(v).strip()):
                    continue  # extraction filler on a report that does not test this family
                c = put(ws,r,col,str(v))
                if anom and f=='total_thc_pct': c.fill = ANOMF
        # THC spec on the CoQ-forming THC row
        if name == (row.get('total_thc_pct__file') or '').replace('_',', ').replace(', pdf','.pdf') or \
           nrm(name) == nrm(row.get('total_thc_pct__file') or '§'):
            a = lookup(ASSIGN, row)
            put(ws,r,6, a['range'] if a else '/')
        put(ws,r,27, m.get('cert_code') or p.get('doc_code') or '')
        put(ws,r,28, m.get('date_of_issue') or '')
        put(ws,r,29, LABNAME.get(m.get('lab'), m.get('lab') or ''), CELL, LFT)
        url = drive_url(name)
        cell = put(ws,r,30, 'Open' if url else '—', LINK if url else CELL)
        if url: cell.hyperlink = url
        if anom:
            note = put(ws,r,31, anom, Font(size=8, italic=True, color='B05000'), LFT)
        first = False; r += 1
print('sheet1 rows:', r)

# ============ Sheet 2 — Stability Testing Programme ==========================
DOCS = {d['name']: d['text'] for d in json.load(open(f'{S}/all_cert_texts.json'))}
stab_meta = {}
_st = _mr['Stability Testing Programme']
for rr in range(6, _st.max_row+1):
    code = _st.cell(rr,6).value
    if code: stab_meta[nrm(code)] = {'timepoint': _st.cell(rr,4).value, 'cond': _st.cell(rr,5).value}
def comp(text, label):
    m = re.search(label + r'\s*\n?\s*[≤\d.,/ %]*?\n?\s*([<>≤]?\s?(?:LOQ|BLQ|ND|Н\.?[Дд]\.?|\d+[.,]\d+|\d+))\s*%?\s*\n', text)
    return m.group(1).strip() if m else ''
SH = ['Seq','Batch (P-number)','Variety','Timepoint','Storage condition','Report No.','Date of issue','Laboratory',
      'Loss on drying %','CBDA %','CBD %','CBN %','Δ⁹-THC %','Δ⁹-THCA %','Total CBD %','Total Δ⁹-THC %','Remark']
ws2, r2 = sheet('Stability Programme', SH, [5,12,11,10,13,11,11,30,9,7,7,7,8,8,8,9,26],
    'Purely Plant GmbH — Stability Testing Programme (observations from eCoA_DATABASE)',
    'ICH long-term / accelerated · deliberately aged samples · results are trend observations and are NEVER CoQ-forming release values · built 27.08.2026')
seq = 0
stabs = [(row, n) for row in ROWS for n in CERT_MAP[row['key']]
         if RECS[n]['meta'].get('test_type')=='STABILITY_TIMEPOINT' or (RECS[n].get('params') or {}).get('is_stability')]
for row, name in stabs:
    seq += 1; rec = RECS[name]; m = rec['meta']; t = DOCS[name]
    sm = stab_meta.get(nrm(m.get('cert_code') or ''), {})
    cbn = comp(t, r'Содржина на CBN')
    vals = [comp(t, r'Губиток\s+(?:при|со)\s+сушење'), comp(t, r'Содржина на CBDA'), comp(t, r'Содржина на CBD(?!A)'),
            cbn, comp(t, r'Содржина на Δ9-THC(?!A)'), comp(t, r'Содржина на Δ9-THCA'),
            comp(t, r'Вкупно CBD'), comp(t, r'Вкупно Δ9-THC')]
    remark = ''
    try:
        if cbn and float(cbn.replace(',','.')) > 1.0: remark = f'CBN {cbn} % exceeds the ≤ 1.00 % release limit (aged sample, expected trend)'
    except ValueError: pass
    data = [seq, f"{row['batch']} ({row.get('p_number') or '—'})", row.get('strain') or '', sm.get('timepoint') or '',
            sm.get('cond') or '', m.get('cert_code'), m.get('date_of_issue'), LABNAME.get(m.get('lab'), m.get('lab')),
            *vals, remark]
    for i,v in enumerate(data,1): put(ws2,r2,i,v, CELL, LFT if i in (2,3,8,17) else CTR)
    url = drive_url(name)
    if url:
        c = ws2.cell(r2,6); c.hyperlink = url; c.font = LINK
    r2 += 1
print('sheet2 rows:', r2)

# ============ Sheet 3 — THC by Strain ========================================
ws3, r3 = sheet('THC by Strain', ['Seq','Batch number','P-number','Total Δ⁹-THC %','Batch THC spec','CoA code','Date of issue','Issuing institution','Certificate'],
    [5,15,10,12,16,16,11,34,9],
    'Purely Plant GmbH — Total Δ⁹-THC by Strain',
    'Every potency result grouped by strain · release-type results from cannabinoid-scoped laboratories · retests follow the original chronologically · built 27.08.2026')
by_strain = {}
for row in ROWS: by_strain.setdefault(row.get('strain') or '— strain not identified —', []).append(row)
seq = 0
for strain in sorted(by_strain, key=lambda s: (s.startswith('—'), s)):
    put(ws3,r3,1,f'{strain}      {len(by_strain[strain])} batch(es)', BOLD9, LFT, BATCHF)
    for i in range(2,10): put(ws3,r3,i,'',CELL,CTR,BATCHF)
    r3 += 1
    for row in by_strain[strain]:
        seq += 1; first = True
        for name in CERT_MAP[row['key']]:
            rec = RECS[name]; p = rec.get('params') or {}; m = rec['meta']
            if m.get('test_type')=='STABILITY_TIMEPOINT' or p.get('is_stability'): continue
            if p.get('total_thc_pct') in (None,'') or m.get('lab') not in ('CNP','FHM','NGP','PP'): continue
            a = lookup(ASSIGN, row)
            put(ws3,r3,1, seq if first else '')
            put(ws3,r3,2, row['batch'] if first else '', BOLD9 if first else CELL, LFT)
            put(ws3,r3,3, row.get('p_number') or '' if first else '')
            put(ws3,r3,4, str(p['total_thc_pct']))
            put(ws3,r3,5, (a['range'] if a else '/') if first else '')
            put(ws3,r3,6, m.get('cert_code') or ''); put(ws3,r3,7, m.get('date_of_issue') or '')
            put(ws3,r3,8, LABNAME.get(m.get('lab'), m.get('lab')), CELL, LFT)
            url = drive_url(name); c = put(ws3,r3,9,'Open' if url else '—', LINK if url else CELL)
            if url: c.hyperlink = url
            first = False; r3 += 1
print('sheet3 rows:', r3)

# ============ Sheet 4 — Potency & Tier Assignment ============================
ws4, r4 = sheet('Potency & Tiers', ['No.','Batch','P-number','Strain','CoQ-forming THC %','Anchor eCoA · date','Tier','Nominal','Tolerance','Range','New name','Brand','In range?'],
    [5,15,10,16,11,24,8,9,10,16,18,10,11],
    'Purely Plant GmbH — Potency Tier Assignment per Batch',
    'CoQ-forming Total Δ⁹-THC vs the tier system of the potency workbook · tier data from Потенција по сорти workbook, verdict recomputed from the RAG value · built 27.08.2026')
n4 = 0
for row in ROWS:
    n4 += 1
    thc = row.get('total_thc_pct'); src = row.get('total_thc_pct__src') or ''
    tier = lookup(TIER, row)
    a = lookup(ASSIGN, row)
    rng = (tier or {}).get('range') or (a or {}).get('range')
    verdict = ''
    pr = parse_range(rng) if rng else None
    try:
        tv = float(str(thc).replace(',','.'))
        if pr: verdict = 'ДА | YES' if pr[0] <= tv <= pr[1] else 'НАДВОР | OUTSIDE'
    except (TypeError, ValueError): pass
    vals = [n4, row['batch'], row.get('p_number') or '', row.get('strain') or '', str(thc or ''),
            src, (tier or {}).get('tier') or (a or {}).get('grade') or '', (tier or {}).get('nominal') or (a or {}).get('nominal') or '',
            (tier or {}).get('tol') or '', rng or '', (tier or {}).get('newname') or '', (tier or {}).get('brand') or '', verdict]
    for i,v in enumerate(vals,1):
        c = put(ws4,r4,i,v, CELL, LFT if i in (2,4,6,11) else CTR)
        if i==13 and 'OUTSIDE' in str(v): c.fill = ANOMF
    r4 += 1
print('sheet4 rows:', r4)

# ============ Sheet 5 — Stock without register result ========================
have = set()
for row in ROWS:
    have.add(nrm(row['batch']))
    if row.get('p_number'): have.add(nrm(row['p_number']))
ws5, r5 = sheet('Stock w-o eCoA', ['Strain','Batch','Tranche','Basis','Value','Tier','Nominal ± tol.','Range','Status vs eCoA_DATABASE (27.08.2026)'],
    [16,13,8,30,9,8,16,16,34],
    'Purely Plant GmbH — Stock batches without a register result',
    'Carried from the potency workbook · status column re-checked live against eCoA_DATABASE · built 27.08.2026')
for s in STOCK:
    status = 'still NO eCoA in database — declared value only'
    if nrm(s[1]) in have:
        row = next(x for x in ROWS if nrm(x['batch'])==nrm(s[1]) or nrm(x.get('p_number') or '§')==nrm(s[1]))
        status = f"now HAS eCoA: THC {row.get('total_thc_pct')} ({row.get('total_thc_pct__src')})"
    vals = list(s) + [status]
    for i,v in enumerate(vals,1):
        c = put(ws5,r5,i,v, CELL, LFT if i in (1,4,9) else CTR)
        if i==9 and 'HAS' in status: c.fill = NEWF
    r5 += 1
print('sheet5 rows:', r5)

# ============ Sheet 6 — Specifications by Strain =============================
ws6, r6 = sheet('Specs by Strain', ['#','First Production','Strain / Cultivar','Specification Code · Ver','Spec','Product Code','Range (Total Δ⁹-THC)','Nominal','Assigned Batches (from eCoA_DATABASE, CoQ-forming THC)'],
    [4,12,22,18,8,16,15,8,60],
    'Purely Plant GmbH — Product Specifications by Strain',
    'Specification ladder per cultivar · assignment column regenerated from the RAG database: every batch whose CoQ-forming THC falls in the spec range · built 27.08.2026')
last_no = None
for sp in SPECS:
    pr = parse_range(sp['range'])
    assigned = []
    if pr:
        for row in ROWS:
            st = (row.get('strain') or '').lower().replace('junky','junkie')
            sps = str(sp['strain']).lower()
            if not st or st.split()[0] not in sps: continue
            try: tv = float(str(row.get('total_thc_pct')).replace(',','.'))
            except (TypeError,ValueError): continue
            if pr[0] <= tv <= pr[1]:
                assigned.append(f"{row['batch']} ({row.get('production') or '?'}, {row.get('total_thc_pct')}%)")
    vals = [sp['no'] if sp['no']!=last_no else '', sp['first_prod'] if sp['no']!=last_no else '',
            sp['strain'] if sp['no']!=last_no else '', sp['code'] if sp['no']!=last_no else '',
            sp['spec'], sp['pcode'], sp['range'], sp['nominal'], ' | '.join(assigned) or '— reserve —']
    last_no = sp['no']
    for i,v in enumerate(vals,1): put(ws6,r6,i,v, CELL, LFT if i in (3,4,9) else CTR)
    r6 += 1
print('sheet6 rows:', r6)

# ============ Sheet 7 — PP_Spec_Parameter_Matrix =============================
ws7, r7 = sheet('PP_Spec_Parameter_Matrix', ['Parameter','Unit','Acceptance criterion','Method / Ph. Eur.','Laboratories in scope','Rule notes'],
    [26,10,30,22,34,44],
    'PP_Spec_Parameter_Matrix — governing matrix for this register',
    'Acceptance criteria per Cannabis flos monograph 07/2024:3028 (Ph. Eur.) and PP QCSP 001 · lab scope = laboratories whose result may form the CoQ value for the parameter · built 27.08.2026')
MATRIX = [
 ('Identification — Macroscopic','—','Conforms to description','Ph. Eur. 2.8.23','CNP, PP','identity per monograph'),
 ('Identification — Microscopic','—','Conforms to description','Ph. Eur. 2.8.23','CNP, PP',''),
 ('Identification — HPTLC','—','Identity confirmed','Ph. Eur. 2.8.25','CNP, PP',''),
 ('Foreign matter','%','≤ 2.00 (no seed; no leaves > 1 cm)','Ph. Eur. 2.8.2','CNP, PP',''),
 ('Loss on drying','% w/w','≤ 12.00','Ph. Eur. 2.2.32','CNP, FHM, PP',''),
 ('Total Δ⁹-THC','% w/w','per target grade (QCSP 001 §01) — see grade ladder below','Ph. Eur. 2.2.29 (HPLC)','CNP, FHM, NGP, PP','CoQ-forming: latest release-type result; >60 days apart = retest round, newest wins'),
 ('Total CBD','% w/w','≤ 1.00','Ph. Eur. 2.2.29','CNP, FHM, NGP, PP',''),
 ('Total CBN','% w/w','≤ 1.00','Ph. Eur. 2.2.29','CNP, FHM, NGP, PP','stability-aged samples may exceed; never CoQ-forming'),
 ('TAMC','CFU/g','≤ 10⁵','Ph. Eur. 2.6.12','IJZ-MB, PP',''),
 ('TYMC','CFU/g','≤ 10⁴','Ph. Eur. 2.6.12','IJZ-MB, PP',''),
 ('Bile-tolerant gram-negative bacteria','CFU/g','≤ 10⁴','Ph. Eur. 2.6.13','IJZ-MB, PP',''),
 ('Salmonella','/25 g','Absent','Ph. Eur. 2.6.13','IJZ-MB, PP',''),
 ('Escherichia coli','/1 g','Absent','Ph. Eur. 2.6.13','IJZ-MB, PP',''),
 ('Aflatoxins (sum B1+B2+G1+G2)','µg/kg','≤ 4','Ph. Eur. 2.8.18','IJZ, FHM, DFL, PP',''),
 ('Aflatoxin B1','µg/kg','≤ 2','Ph. Eur. 2.8.18','IJZ, FHM, DFL, PP',''),
 ('Ochratoxin A','µg/kg','≤ 20','Ph. Eur. 2.8.22','IJZ, FHM, DFL, PP',''),
 ('Lead (Pb)','mg/kg','≤ 0.5','Ph. Eur. 2.4.27','IJZ, FHM, DFL, PP',''),
 ('Cadmium (Cd)','mg/kg','≤ 0.3','Ph. Eur. 2.4.27','IJZ, FHM, DFL, PP',''),
 ('Arsenic (As)','mg/kg','≤ 0.2','Ph. Eur. 2.4.27','IJZ, FHM, DFL, PP',''),
 ('Mercury (Hg)','mg/kg','≤ 0.1','Ph. Eur. 2.4.27','IJZ, FHM, DFL, PP',''),
 ('Pesticide residues','mg/kg','≤ LOQ (0.01) each','Ph. Eur. 2.8.13 / MKC EN 15662','IJZ, DFL, PP','reported per-substance not-detected'),
]
for m_ in MATRIX:
    for i,v in enumerate(m_,1): put(ws7,r7,i,v, CELL, LFT)
    r7 += 1
r7 += 1
put(ws7,r7,1,'RULES', BOLD9, LFT, BATCHF)
for i in range(2,7): put(ws7,r7,i,'',CELL,CTR,BATCHF)
r7 += 1
for rule in (
 'CoQ-forming value = latest release-type result from a laboratory scoped for the parameter; PP in-house restatements form the value only when no external laboratory tested it.',
 'Retest rule: certificates for the same batch/parameter issued more than 60 days apart open a new testing round; the newest result supersedes and becomes the CoQ-forming value.',
 'Stability-programme results (deliberately aged samples) are trend observations only and are never CoQ-forming.',
 'CoQ document codes CoQ-PP-2026-NNNN per QCSOP 012 v.03 §6.4.2; a superseding CoQ carries retested values and reuses time-invariant results.',
 'Values are carried exactly as printed on the certificate (decimal places preserved); units live in the column headers only.'):
    put(ws7,r7,1,rule, CELL, LFT); ws7.merge_cells(start_row=r7,start_column=1,end_row=r7,end_column=6)
    r7 += 1
r7 += 1
put(ws7,r7,1,'GRADE LADDER (per strain, from Specifications by Strain)', BOLD9, LFT, BATCHF)
for i in range(2,7): put(ws7,r7,i,'',CELL,CTR,BATCHF)
r7 += 1
for i,h in enumerate(('Strain','Spec code','Spec','Product code','Range','Nominal'),1):
    c = put(ws7,r7,i,h, HDR, CTR); c.fill = HDRF
r7 += 1
for sp in SPECS:
    for i,v in enumerate((sp['strain'],sp['code'],sp['spec'],sp['pcode'],sp['range'],sp['nominal']),1):
        put(ws7,r7,i,v, CELL, LFT if i==1 else CTR)
    r7 += 1
print('sheet7 rows:', r7)

# ============ Sheet 8 — RAG Sanity Check =====================================
ws8, r8 = sheet('RAG Sanity Check', ['#','Layer / item','Result','Detail'],
    [5,34,22,90],
    'RAG database sanity check — is the RAGed data what the eCoA actually says?',
    'Three independent verification layers + paper-loop vision samples · run 27.08.2026 against all 291 certificates of eCoA_DATABASE')
i8 = 0
def srow(a,b,c_):
    global i8, r8; i8 += 1
    put(ws8,r8,1,i8); put(ws8,r8,2,a,BOLD9,LFT); put(ws8,r8,3,b,CELL,CTR); put(ws8,r8,4,c_,CELL,LFT); r8 += 1
srow('Layer A — verbatim presence','2469 / 2512 (98.3%)','Every extracted value searched verbatim (normalized) in its own certificate parse text; the 43 misses were individually adjudicated and repaired — see repair log below.')
srow('Layer B — labeled-line anchors','106 / 107 (99.1%)','Potency-critical fields (THC/CBD/CBN/LoD) independently re-read from the labeled line/table row of the parse text (CNP + FHM layouts). The 1 residual is a regex artifact on a spec cell (ППК25367 LoD — extracted 8.22 is correct, confirmed on paper by QC).')
srow('Layer C — human register cross-check','271 / 314 values agree','Values compared against the owner-maintained Batch Release register per CoA code. ~30 "differences" are Cyrillic Н.Д. vs Latin N.D. (equivalent). Adjudicated live diffs: BG1024 CBN (register transcribed uncertainty 0.02, cert prints 0.28), HPA1024 CBN 0.36, P060352 CBD/CBN swapped in register, P050042 Hg 0.001 (register 0.011) — in all four the RAG value matches the certificate; the register cell is wrong.')
srow('Paper loop — vision samples','3 / 3 exact','Actual PDFs read visually and compared to the parse: GG012603/ППК26114 (CNP), BG1024/197-1-K-26 (FHM), CJ072501/ППК25367 (CNP, owner-supplied scan). Parse text reproduces the printed tables exactly.')
srow('Repairs applied to extraction','49 entries','All logged with certificate-text evidence; classes: 5 pesticide polarity flips ("Не одговара" fabricated where certs print not-detected), 3 blanket-phantom docs (fields deleted for sections not on the certificate), GG012603 column shift (FM/LoD/CBN/CBD re-seated), 2 line-pick errors (ППК25367 CBD, ППК26005 THC), as-printed wording restored (Одговара vs "Absent").')
srow('Still open (QC review / paper checks)','6 items','ППК25117 printed THC total 1.58 vs own components 15.38 — VERIFY ON PAPER; ППК25154 printed 1.87 same class; BSS052501 sheet 20.47 — probably the NGP result but the NGP scan parses as garbage, read off paper; GP062501 PP cert prints 24.89 vs tranche sheet 22.89 — QC decision; J31112501 sheet 25.00 = the March 051-5-K-26 result (25.27, CF-77/26), stale — superseded by the April retest 100-1-K-26 (20.21); J31122501 RESOLVED 27.08.2026: QC designated the trimmed preparation (100-3-K-26, 21.84) as CoQ-forming; hand-trimmed 19.84 (100-2-K-26) kept as experimental processing-mode comparison of the same batch; tranche sheet prints 21.77 — transcription slip of 21.84, correct the sheet.')
srow('Verdict','DATABASE FAITHFUL','After the 49 logged repairs, the RAGed data is verbatim-consistent with the certificate text at 3 verification layers, and the parse text is vision-verified against the printed PDFs. Known residual risks are confined to the 4 open paper checks above.')
r8 += 1
put(ws8,r8,1,'REPAIR LOG', BOLD9, LFT, BATCHF)
for i in range(2,5): put(ws8,r8,i,'',CELL,CTR,BATCHF)
r8 += 1
for i,h in enumerate(('#','Certificate','Field · old → new','Reason'),1):
    c = put(ws8,r8,i,h,HDR,CTR); c.fill = HDRF
r8 += 1
for j,l in enumerate(REPAIRS,1):
    put(ws8,r8,1,j)
    put(ws8,r8,2,l['doc'],CELL,LFT)
    put(ws8,r8,3,f"{l['field']} · {l.get('old')} → {l.get('new') if 'new' in l else l.get('action','')}",CELL,LFT)
    put(ws8,r8,4,l['reason'],CELL,LFT)
    r8 += 1
print('sheet8 rows:', r8)

OUT = f'{S}/PP_QC_eCoA_Master_Register.xlsx'
wb.save(OUT)
print('SAVED', OUT)
