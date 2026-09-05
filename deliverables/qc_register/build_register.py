#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Batch QC Testing Results Register — consolidated workbook.

One row per batch; parameter values as numbers with the same decimals the
certificate printed; units in headers only; CoQ-forming value = latest
release-type result from a lab scoped for that parameter (stability excluded);
retest supersessions and full provenance carried on their own sheets.
"""
import json, re, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

S = os.path.dirname(os.path.abspath(__file__))
cons = json.load(open(f'{S}/consolidated.json'))
verify = json.load(open(f'{S}/verify_vs_sheet.json'))
ref = json.load(open(f'{S}/ref_register.json'))
rows, PARAMS, AC = cons['rows'], cons['params'], cons['ac']
DRIVE = cons['drive']

NAVY = 'FF1F3864'; LIGHT = 'FFDCE6F1'; GREY = 'FF808080'; RED = 'FFC00000'
AMBER = 'FFBF8F00'; GREEN = 'FF375623'
hdr_font = Font(bold=True, color='FFFFFFFF', size=9)
hdr_fill = PatternFill('solid', fgColor=NAVY)
sub_fill = PatternFill('solid', fgColor=LIGHT)
thin = Side(style='thin', color='FFB0B0B0')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)

NUM_RE = re.compile(r'^-?\d+(?:[.,]\d+)?$')

def write_value(ws, r, c, raw):
    """Numeric values as real numbers with the certificate's decimal places;
    qualifiers (N.D., <LOQ, Одговара, 1.6x10^4 ...) stay text verbatim."""
    cell = ws.cell(row=r, column=c)
    if raw is None:
        cell.value = '—'; cell.font = Font(color=GREY, size=8)
        cell.alignment = center; cell.border = border
        return
    t = str(raw).strip()
    if NUM_RE.match(t.replace(',', '.')):
        dec = len(t.replace(',', '.').split('.')[1]) if ('.' in t or ',' in t) else 0
        cell.value = float(t.replace(',', '.'))
        cell.number_format = '0' if dec == 0 else '0.' + '0' * dec
    else:
        cell.value = t
    cell.font = Font(size=9)
    cell.alignment = center; cell.border = border

# ---------------------------------------------------------------- Register
wb = openpyxl.Workbook()
ws = wb.active; ws.title = 'Batch QC Register'

META = ['No.', 'Production', 'Tranche', 'Batch No. (cultivation)', 'P-number',
        'Strain', 'Product code', 'CoQ doc. code (proposed)',
        'First result', 'Latest result', 'eCoAs']
ws.cell(row=1, column=1, value='Purely Plant — Batch QC Testing Results Register'
        ).font = Font(bold=True, size=13, color=NAVY)
ws.cell(row=2, column=1, value=('One row per batch · consolidated from the eCoA_DATABASE (RAGFlow, 291 certificates, complete mirror of the '
        'ingestion folder) · CoQ-forming value = latest release-type result from a laboratory scoped for the parameter · '
        'stability-programme results excluded (see Stability sheet) · values exactly as printed on the certificate, '
        'units in headers · built 27.08.2026')).font = Font(size=8, italic=True, color=GREY)

hr = 4
for j, h in enumerate(META, start=1):
    cell = ws.cell(row=hr, column=j, value=h)
    cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = border
for j, (f, label, unit) in enumerate(PARAMS, start=len(META) + 1):
    txt = f'{label}' + (f' [{unit}]' if unit else '')
    cell = ws.cell(row=hr, column=j, value=txt)
    cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = border
    ac = ws.cell(row=hr + 1, column=j, value=f"A.C. {AC.get(f, '')}")
    ac.font = Font(size=7, italic=True, color=GREY); ac.alignment = center; ac.border = border
for j in range(1, len(META) + 1):
    c = ws.cell(row=hr + 1, column=j, value='')
    c.fill = sub_fill; c.border = border

r0 = hr + 2
old_batches = {re.sub(r'[^A-Z0-9*]', '', str(b['batch']).upper().replace('-', '')) for b in ref}
new_flag = []
for i, row in enumerate(rows):
    r = r0 + i
    vals = [i + 1, row.get('production'), int(row['tranche']) if row.get('tranche') else None,
            row['batch'], row.get('p_number'), row.get('strain'),
            row.get('product_code') or row.get('product_name'), row.get('coq'),
            row.get('first_result'), row.get('last_result'), row.get('n_certs')]
    for j, v in enumerate(vals, start=1):
        cell = ws.cell(row=r, column=j, value=v)
        cell.font = Font(size=9, bold=(j == 4)); cell.alignment = center; cell.border = border
    for j, (f, label, unit) in enumerate(PARAMS, start=len(META) + 1):
        write_value(ws, r, j, row.get(f))
        if row.get(f + '__retest'):
            ws.cell(row=r, column=j).font = Font(size=9, bold=True, color=AMBER)
        if row.get(f + '__note'):
            ws.cell(row=r, column=j).font = Font(size=9, bold=True, color=RED)
    k = re.sub(r'[^A-Z0-9*]', '', str(row['batch']).upper())
    if k not in old_batches and (row.get('p_number') is None or
            re.sub(r'[^A-Z0-9*]', '', str(row.get('p_number') or '').upper()) not in old_batches):
        new_flag.append(row['batch'])

ws.column_dimensions['D'].width = 15
for col in ('E', 'F', 'G', 'H'):
    ws.column_dimensions[col].width = 13
for j in range(len(META) + 1, len(META) + len(PARAMS) + 1):
    ws.column_dimensions[get_column_letter(j)].width = 10
ws.freeze_panes = ws.cell(row=r0, column=5)

# ------------------------------------------------------------- Provenance
wp = wb.create_sheet('Result Provenance')
wp.cell(row=1, column=1, value='Certificate behind every CoQ-forming value — code, date of issue, laboratory. '
        'Hyperlinks open the source eCoA in the ingestion folder.').font = Font(size=8, italic=True, color=GREY)
for j, h in enumerate(['Batch', 'P-number'] + [p[1] for p in PARAMS], start=1):
    c = wp.cell(row=2, column=j, value=h); c.font = hdr_font; c.fill = hdr_fill
    c.alignment = center; c.border = border
for i, row in enumerate(rows):
    r = 3 + i
    wp.cell(row=r, column=1, value=row['batch']).font = Font(size=9, bold=True)
    wp.cell(row=r, column=2, value=row.get('p_number')).font = Font(size=9)
    for j, (f, label, unit) in enumerate(PARAMS, start=3):
        src = row.get(f + '__src')
        cell = wp.cell(row=r, column=j, value=src or '—')
        cell.font = Font(size=8, color=GREY if not src else 'FF000000')
        cell.border = border
        fn = row.get(f + '__file')
        if fn and fn in DRIVE:
            cell.hyperlink = f'https://drive.google.com/file/d/{DRIVE[fn]}/view'
            cell.font = Font(size=8, color='FF0563C1', underline='single')
for j in range(3, 3 + len(PARAMS)):
    wp.column_dimensions[get_column_letter(j)].width = 22
wp.freeze_panes = 'C3'

# ------------------------------------------------------------ CoQ Dossier
wd = wb.create_sheet('CoQ Dossier')
wd.cell(row=1, column=1, value='CoQ Compilation Register — one issuance event per batch; the certificates that form it; '
        'retest supersessions per QCSOP 012 v.03 §6.4.2 (superseding CoQ carries the retested values and reuses '
        'time-invariant results from the earlier round).').font = Font(size=8, italic=True, color=GREY)
hdrs = ['No.', 'CoQ doc. code (proposed)', 'Batch', 'P-number', 'Strain',
        'Forming certificates (code · date · lab)', 'Retest supersessions']
for j, h in enumerate(hdrs, start=1):
    c = wd.cell(row=2, column=j, value=h); c.font = hdr_font; c.fill = hdr_fill
    c.alignment = center; c.border = border
retest_by_batch = {}
for t in cons['retests']:
    retest_by_batch.setdefault(t['batch'], []).append(t)
for i, row in enumerate(rows):
    r = 3 + i
    certs_txt = '\n'.join(f'{c[0]} · {c[1]} · {c[2]}' for c in row.get('_certs', []))
    sup = retest_by_batch.get(row['batch'], [])
    sup_txt = '\n'.join(f"{t['parameter']}: {t['superseded_value']} ({t['superseded_src']}) → "
                        f"{t['current_value']} ({t['current_src']}), +{t['gap_days']}d" for t in sup)
    for j, v in enumerate([i + 1, row.get('coq'), row['batch'], row.get('p_number'),
                            row.get('strain'), certs_txt, sup_txt or '—'], start=1):
        c = wd.cell(row=r, column=j, value=v); c.font = Font(size=8)
        c.alignment = Alignment(vertical='top', wrap_text=True); c.border = border
wd.column_dimensions['F'].width = 52; wd.column_dimensions['G'].width = 60
wd.column_dimensions['B'].width = 18; wd.column_dimensions['C'].width = 14

# -------------------------------------------------------------- Stability
wst = wb.create_sheet('Stability (excluded)')
wst.cell(row=1, column=1, value='Stability-programme observations — deliberately aged samples (ICH long-term 25C/60RH and '
        'accelerated 40C/75RH). NEVER CoQ-forming; kept for trend only.').font = Font(size=8, italic=True, color=RED)
for j, h in enumerate(['Batch', 'Parameter', 'Unit', 'Value', 'Certificate', 'Note'], start=1):
    c = wst.cell(row=2, column=j, value=h); c.font = hdr_font; c.fill = hdr_fill; c.border = border
for i, t in enumerate(cons['stability']):
    for j, v in enumerate([t['batch'], t['parameter'], t['unit'], t['value'], t['src'], t['note']], start=1):
        c = wst.cell(row=3 + i, column=j, value=v); c.font = Font(size=8); c.border = border
for col, w in (('A', 12), ('B', 18), ('D', 10), ('E', 28), ('F', 42)):
    wst.column_dimensions[col].width = w

# ----------------------------------------------------------- Verification
wv = wb.create_sheet('Verification')
wv.cell(row=1, column=1, value='Truthfulness check of the shared register/tranche sheets against the eCoA database '
        '(certificate content is the authority).').font = Font(bold=True, size=10, color=NAVY)
r = 3
wv.cell(row=r, column=1, value='A — Tranche-sheet THC values that DISAGREE with the certificates (13)').font = Font(bold=True, size=9); r += 1
for j, h in enumerate(['Cultivation batch', 'P-number', 'Tranche', 'Sheet THC', 'Certificate THC', 'Certificate (code · date · lab)', 'Assessment'], start=1):
    c = wv.cell(row=r, column=j, value=h); c.font = hdr_font; c.fill = hdr_fill; c.border = border
r += 1
ASSESS = {
    'JD012603/02': 'sheet stale — newer CNP retest 30.06.2026 supersedes',
    'JD012603/02V': 'sheet stale — newer CNP retest 30.06.2026 supersedes',
    'FB012603': 'sheet stale — newer CNP retest 30.06.2026 supersedes (20.83 verified)',
    'FB012603V': 'sheet stale — newer CNP retest 30.06.2026 supersedes (18.29 verified)',
    'SCR022601': 'sheet stale — newer CNP retest 06.07.2026 supersedes',
    'CJ052501/02': 'sheet digit swap — certificate prints 14.93 (verified in text)',
    'KC102501': 'sheet digit swap — certificate prints 17.04 (verified in text)',
    'PM112501': 'sheet value 13.00 not on any certificate — cert prints 13.33 (verified)',
    'SJ112501': 'sheet rounded 9.00 — certificate prints 9.20 (verified)',
    'GP062501': 'PP certificate itself prints 24.89 (verified); sheet 22.89 source unknown — QC review',
    'BSS052501': 'sheet 20.47 plausibly the NGP result; NGP scan parse unusable in DB — verify NGP paper cert',
    'J31112501': 'sheet 25.00 matches no certificate in the ingestion folder — source unknown, QC review',
    'J31122501': 'sheet 21.77 matches no certificate in the ingestion folder — source unknown, QC review',
}
for m in verify['mismatches']:
    for j, v in enumerate([m['cult'], m['pp'], m['tranche'], m['sheet'], m['mine'], m['src'], ASSESS.get(m['cult'], '')], start=1):
        c = wv.cell(row=r, column=j, value=v); c.font = Font(size=8); c.border = border
    r += 1
r += 1
wv.cell(row=r, column=1, value='B — Batches on the tranche sheet with NO certificate in the ingestion folder (10)').font = Font(bold=True, size=9); r += 1
for b in verify['missing']:
    wv.cell(row=r, column=1, value=b).font = Font(size=9, color=RED); r += 1
r += 1
wv.cell(row=r, column=1, value='C — Values corrected against certificate text (extraction/print anomalies)').font = Font(bold=True, size=9); r += 1
for row in rows:
    for f, label, unit in PARAMS:
        if row.get(f + '__note'):
            wv.cell(row=r, column=1, value=row['batch']).font = Font(size=8)
            wv.cell(row=r, column=2, value=label).font = Font(size=8)
            wv.cell(row=r, column=3, value=row.get(f)).font = Font(size=8, bold=True)
            wv.cell(row=r, column=4, value=row.get(f + '__note')).font = Font(size=8, color=AMBER)
            r += 1
r += 1
wv.cell(row=r, column=1, value='D — Batches in the database but NOT in the prior Batch-Release register').font = Font(bold=True, size=9); r += 1
for b in new_flag:
    wv.cell(row=r, column=1, value=b).font = Font(size=9, color=GREEN); r += 1
for col, w in (('A', 16), ('B', 12), ('D', 12), ('E', 14), ('F', 30), ('G', 60)):
    wv.column_dimensions[col].width = w

wb.save(f'{S}/PP_Batch_QC_Testing_Results_Register.xlsx')
print('saved workbook:', len(rows), 'batches |', len(cons['retests']), 'retest events |',
      len(cons['stability']), 'stability obs | new-vs-register:', new_flag)
