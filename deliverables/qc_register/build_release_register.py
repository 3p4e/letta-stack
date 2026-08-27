#!/usr/bin/env python3
"""QC Batch Release Results Register — final grade edition (27.08.2026).

Built on the owner's main register structure (1wSJ…, sheet 'Batch Release QC'),
filled from the RAGFlow eCoA_DATABASE consolidation, with the FINAL
mandatory-code potency grades (grade_design_even.json) — plus the
comprehensive strain/grade/results list in the owner's Specifications-by-
Strain format (1YgJz…) and the Batch Assignments sheet.
"""
import json, re, unicodedata
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

Q = '/home/user/letta-stack/deliverables/qc_register'
P = '/home/user/letta-stack/deliverables/potency_study'
SC = '/tmp/claude-0/-home-user-letta-stack/4877ce6e-ae82-551e-bf35-5698c379c3be/scratchpad'

CONS = json.load(open(f'{Q}/consolidated.json'))
RECS = {r['name']: r for r in json.load(open(f'{Q}/extracted_params.json'))}
GD = json.load(open(f'{P}/grade_design_even.json'))
DS = json.load(open(f'{P}/potency_dataset.json'))

def nrm(s):
    s = unicodedata.normalize('NFKC', str(s))
    return re.sub(r'[\s,_/()\-*]+', '', s).lower()

DRIVE = {}
for line in open(f'{Q}/drive_all.tsv', encoding='utf-8'):
    fid, name = line.rstrip('\n').split('\t')
    DRIVE[nrm(name)] = fid
def drive_url(name):
    fid = DRIVE.get(nrm(name))
    return f'https://drive.google.com/file/d/{fid}/view' if fid else None

def base_batch(b): return str(b).split(' (')[0]

LABNAME = {'CNP': 'UKIM Faculty of Pharmacy — Center for Natural Products',
           'IJZ': 'IPH — Institute of Public Health (chemistry)',
           'IJZ-MB': 'IPH — Institute of Public Health (microbiology)',
           'FHM': 'Farmahem', 'PP': 'Purely Plant (in-house)',
           'NGP': 'NGP — in-house GC cross-check', 'DFL': 'DFL Deutsches Institut (DE)'}

# ---- grade lookup (final design) -------------------------------------------
def bkey(b):
    s = str(b)
    return nrm(s) + ('#star' if '*' in s else '')

GRADE = {}     # bkey(batch) -> grade dict + strain
for strain, entry in GD['strains'].items():
    for g in entry:
        for b in g['batches']:
            GRADE[bkey(b['batch'])] = {**g, 'strain': strain, 'anchor': b['v'],
                                       'basis': b['basis']}
BALIAS = {'CJ052501/1': 'CJ052501/01', 'CJ052501/2': 'CJ052501/02',
          'JD012603/2': 'JD012603/02', 'JD012603/2V': 'JD012603/02V',
          'OPM1024_01': 'OMP1024_01', 'BSS1024_01/1': 'BSS1024_01',
          'CJ082501-2': 'CJ082501/2', 'FB012601_1': 'FB012601/1',
          'JD012603_02': 'JD012603/02', 'JD012603_02V': 'JD012603/02V',
          'GRC102501_2': 'GRC102501/2'}
def canonb(b): return BALIAS.get(b, b)

def grade_of(row):
    for k in (row['batch'], base_batch(row['batch']),
              canonb(base_batch(row['batch'])), row.get('p_number') or '§'):
        v = GRADE.get(bkey(k))
        if v: return v
    return None

def gr_range(g):  return f"{g['lower']:.2f} – {g['upper']:.2f} %"
def gr_tol(g):    return f"{g['nominal']}.00% ±{g['tol']:.2f}%"
def gr_label(g):  return f"{g['grade']} · {g['product_code'].replace(':', ' : ')}"

# basis helpers: declared (no analysis at all) vs T2 re-analysis (26.08.2026,
# unofficial Farmahem values — CoQ-forming per R5, formal eCoAs pending)
def is_decl(basis): return basis.startswith('declared')
def is_t2(basis):   return 'T2' in basis
T2_TAG = 'Т2 26.08.2026 (неофиц. | unofficial)'

# ---- styling ----------------------------------------------------------------
HDR = Font(bold=True, size=9, color='FFFFFF'); HDRF = PatternFill('solid', fgColor='1F4E5F')
CELL = Font(size=9); BOLD9 = Font(bold=True, size=9)
TITLE = Font(bold=True, size=12, color='1F4E5F'); SUB = Font(size=9, italic=True, color='555555')
CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
LFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
LINK = Font(size=9, color='0563C1', underline='single')
BATCHF = PatternFill('solid', fgColor='EAF1F5'); REFF = PatternFill('solid', fgColor='FFF4DE')
GAPF = PatternFill('solid', fgColor='FBE4E4'); PENDF = PatternFill('solid', fgColor='FDF3D8')
DEVF = PatternFill('solid', fgColor='F8CBAD')
thin = Side(style='thin', color='C9D2D0'); BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()

def sheet(title, heads, widths, title_txt, sub_txt, ref_row=None):
    ws = wb.active if wb.sheetnames == ['Sheet'] else wb.create_sheet()
    ws.title = title
    ws.cell(1, 1, title_txt).font = TITLE
    ws.cell(2, 1, sub_txt).font = SUB
    for i, h in enumerate(heads, 1):
        c = ws.cell(4, i, h); c.font = HDR; c.fill = HDRF; c.alignment = CTR; c.border = BORDER
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    r = 5
    if ref_row:
        for i, v in enumerate(ref_row, 1):
            c = ws.cell(5, i, v); c.font = Font(size=8, italic=True); c.fill = REFF
            c.alignment = CTR; c.border = BORDER
        r = 6
    ws.freeze_panes = ws.cell(r, 1).coordinate
    return ws, r

def put(ws, r, c, v, font=CELL, align=CTR, fill=None):
    cell = ws.cell(r, c, v); cell.font = font; cell.alignment = align; cell.border = BORDER
    if fill: cell.fill = fill
    return cell

# ============ Sheet 1 — Batch Release QC (per-certificate) ===================
HEADS = ['No.','Batch number','P-number','Strain','THC %','THC spec (grade range)','CBD %','CBN %',
 'Loss on drying %','Foreign matter %','Macroscopic ID','Microscopic ID','HPTLC ID',
 'TAMC CFU/g','TYMC CFU/g','Bile-tolerant GNB /1 g','Salmonella /25 g','E. coli /1 g',
 'Aflatoxins Σ µg/kg','Aflatoxin B1 µg/kg','Ochratoxin A µg/kg','Pb mg/kg','Cd mg/kg',
 'As mg/kg','Hg mg/kg','Pesticides','CoA code','Date of issue','Issuing institution','PDF',
 'Potency grade','Spec. doc. code']
WID = [5,15,10,17,8,15,7,7,9,9,10,10,9,11,11,12,11,11,9,9,9,7,7,7,7,11,16,11,26,7,20,16]
REFROW = ['Ref.','','','','CoQ-forming','nominal ±t → range','< 1.00 %','< 1.00 %','≤ 12.00 (3028)',
 '≤ 2.00 %','Conforms','Conforms','Confirmed','≤ 10⁵','≤ 10⁴','≤ 10⁴','Absent','Absent',
 '≤ 4','≤ 2','≤ 20','≤ 0.5','≤ 0.3','≤ 0.2','≤ 0.1','≤ LOQ 0.01','','','','','','']
FIELD2COL = {'total_thc_pct':5,'total_cbd_pct':7,'total_cbn_pct':8,'loss_on_drying_pct':9,
 'foreign_matter_pct':10,'macroscopic_id':11,'microscopic_id':12,'hptlc_id':13,'tamc':14,'tymc':15,
 'bile_tolerant_gnb':16,'salmonella':17,'e_coli':18,'aflatoxins_total':19,'aflatoxin_b1':20,
 'ochratoxin_a':21,'pb':22,'cd':23,'arsenic':24,'hg':25,'pesticides':26}
LABSC = {'cannabinoid': ('CNP','FHM','NGP','PP'), 'micro': ('IJZ-MB','PP'),
         'chem': ('IJZ','FHM','DFL','PP')}
FAM = {**{k:'cannabinoid' for k in ('total_thc_pct','total_cbd_pct','total_cbn_pct',
       'loss_on_drying_pct','foreign_matter_pct','macroscopic_id','microscopic_id','hptlc_id')},
       **{k:'micro' for k in ('tamc','tymc','bile_tolerant_gnb','salmonella','e_coli')},
       **{k:'chem' for k in ('aflatoxins_total','aflatoxin_b1','ochratoxin_a','pb','cd','arsenic','hg','pesticides')}}
QUAL = re.compile(r'^(n\.?d\.?|н\.?д\.?|nd|absent|отсутна?)$', re.I)
PRINT_ANOM = {nrm('P050042, ППК25117, 06.05.2025, CNP.pdf'): 'printed 1.58 vs components 15.38 — VERIFY ON PAPER',
              nrm('P050062, ППК25154, 24.06.2025, CNP.pdf'): 'printed 1.87 anomalous vs FHM 18.04 — VERIFY ON PAPER'}

ws, r = sheet('Batch Release QC', HEADS, WID,
    'Purely Plant GmbH — QC Batch Release Results Register (final potency grades, 27.08.2026)',
    'One row per certificate · "/" = not tested on that certificate · grouped per batch · '
    'source: RAGFlow eCoA_DATABASE (291 certificates, sanity-checked) · grades: ImB_Potency_Grade_Ranges '
    '(mandatory management codes revised 27.08.2026, symmetric balanced tolerances, anchors = '
    'T1+T2 re-analysis of 26.08.2026 where retested — T2 values unofficial pending formal eCoAs)', REFROW)

def dkey(name):
    d = RECS[name]['meta'].get('date_of_issue') or '99.99.9999'
    p = d.split('.')
    return (p[2], p[1], p[0]) if len(p) == 3 else ('9', '9', '9')

nb = 0
for row in CONS['rows']:
    nb += 1
    g = grade_of(row)
    certs = sorted(CONS['cert_map'].get(row['key'], []), key=dkey)
    first = True
    for name in certs:
        rec = RECS.get(name)
        if rec is None: continue
        p, m = rec.get('params') or {}, rec['meta']
        fill = BATCHF if first else None
        put(ws, r, 1, nb if first else '', BOLD9 if first else CELL, CTR, fill)
        put(ws, r, 2, row['batch'] if first else '', BOLD9, LFT, fill)
        put(ws, r, 3, (row.get('p_number') or '') if first else '', CELL, CTR, fill)
        put(ws, r, 4, (row.get('strain') or '') if first else '', CELL, LFT, fill)
        for col in range(5, 27): put(ws, r, col, '/')
        anom = PRINT_ANOM.get(nrm(name))
        for f, col in FIELD2COL.items():
            v = p.get(f)
            if v is None and f == 'aflatoxins_total': v = p.get('total_aflatoxins')
            if v not in (None, ''):
                lab = m.get('lab')
                if lab not in LABSC.get(FAM[f], ()) and QUAL.match(str(v).strip()):
                    continue
                c = put(ws, r, col, str(v))
                if anom and f == 'total_thc_pct': c.fill = DEVF
        if first and g:
            put(ws, r, 6, gr_range(g), BOLD9)
            put(ws, r, 31, gr_label(g), CELL, LFT, fill)
            put(ws, r, 32, g['spec_code'], CELL, CTR, fill)
        put(ws, r, 27, m.get('cert_code') or '')
        put(ws, r, 28, m.get('date_of_issue') or '')
        put(ws, r, 29, LABNAME.get(m.get('lab'), m.get('lab') or ''), CELL, LFT)
        url = drive_url(name)
        cell = put(ws, r, 30, 'Open' if url else '—', LINK if url else CELL)
        if url: cell.hyperlink = url
        if anom: put(ws, r, 33, anom, Font(size=8, italic=True, color='B05000'), LFT)
        first = False; r += 1

# pending batches (no eCoA on file) with their provisional grades
r += 1
put(ws, r, 1, '', CELL); ws.cell(r, 2, 'PENDING — no eCoA on file (grades provisional: '
    'owner-declared values, or unofficial T2 re-analysis values of 26.08.2026)').font = BOLD9
r += 1
reg_norms = {bkey(base_batch(x['batch'])) for x in CONS['rows']} | \
            {bkey(x.get('p_number') or '§') for x in CONS['rows']}
for bk, g in sorted(GRADE.items(), key=lambda kv: kv[1]['strain']):
    if (is_decl(g['basis']) or is_t2(g['basis'])) and bk not in reg_norms:
        bname = next(b['batch'] for e in GD['strains'].values() for gg in e
                     for b in gg['batches'] if bkey(b['batch']) == bk)
        nb += 1
        vtxt = (f"{g['anchor']:.2f} (declared)" if is_decl(g['basis'])
                else f"{g['anchor']:.2f} ({T2_TAG})")
        put(ws, r, 1, nb, BOLD9, CTR, PENDF)
        put(ws, r, 2, bname, BOLD9, LFT, PENDF)
        put(ws, r, 4, g['strain'], CELL, LFT, PENDF)
        put(ws, r, 5, vtxt, CELL, CTR, PENDF)
        put(ws, r, 6, gr_range(g), CELL, CTR, PENDF)
        put(ws, r, 31, gr_label(g), CELL, LFT, PENDF)
        put(ws, r, 32, g['spec_code'], CELL, CTR, PENDF)
        r += 1
S1_ROWS = r

# ============ Sheet 2 — Specifications by Strain =============================
H2 = ['#','First result','Strain / Cultivar','Specification Code','Grade','Product Code',
      'Range (Total Δ⁹-THC)','Nominal ± tolerance',
      'Batches in range (anchor Total Δ⁹-THC %)','Other historical results in range']
W2 = [4,11,26,16,7,18,17,16,52,44]
ws2, r2 = sheet('Specifications by Strain', H2, W2,
    'Purely Plant GmbH — Product Specifications by Strain (final potency grades, 27.08.2026)',
    'Every strain · every potency grade with its symmetric range · every Total Δ⁹-THC result that falls '
    'within it (anchors = latest/CoQ-forming; historical = superseded pre-retest results) · '
    'documented gaps carry no results by construction')

# all historical results per strain (with reattribution + star names)
STRAIN_FIX = {'JellyDonutz': 'Jelly Donutz', 'Permanent Market': 'Permanent Marker'}
STAR = {'GG012601': 'GG012601*', 'JD012601': 'JD012601*'}
ALL_RES = {}
_seen_res = set()
for res in DS['register_results']:
    b = res['batch']
    if b == 'JD112501' and res.get('cert') == 'ППК26065': b = 'JD112501*'
    b = STAR.get(canonb(b), canonb(b))
    s = STRAIN_FIX.get(res['strain'], res['strain'])
    kk = (s, b, round(float(res['value']), 2), res.get('date'))
    if kk in _seen_res: continue
    _seen_res.add(kk)
    ALL_RES.setdefault(s, []).append({**res, 'batch': b})

def pdate(s):
    try: return datetime.strptime(s, '%d.%m.%Y')
    except Exception: return datetime.min

no = 0
for strain in sorted(GD['strains']):
    entry = GD['strains'][strain]
    no += 1
    sc = entry[0]['strain_code']
    n_b = sum(len(g['batches']) for g in entry)
    results = ALL_RES.get(strain, [])
    latest_per_batch = {}
    for res in results:
        b = res['batch']
        if b not in latest_per_batch or pdate(res['date']) > pdate(latest_per_batch[res['date'] if False else b]['date']):
            latest_per_batch[b] = res
    first_date = min((pdate(res['date']) for res in results), default=None)
    first_txt = first_date.strftime('%b %Y') if first_date and first_date != datetime.min else '—'
    first = True
    for gi, g in enumerate(entry):
        anchors = ', '.join(f"{b['batch']} {b['v']:.2f}"
                            + (' (декл.)' if is_decl(b['basis'])
                               else ' (Т2 неофиц.)' if is_t2(b['basis']) else '')
                            for b in g['batches'])
        hist = []
        for res in results:
            v = float(res['value'])
            is_anchor = any(abs(b['v'] - v) < 0.005 and b['batch'] == res['batch']
                            for b in g['batches'])
            if not is_anchor and g['lower'] - 1e-9 <= v <= g['upper'] + 1e-9:
                hist.append(f"{res['batch']} {v:.2f} ({res['date']})")
        put(ws2, r2, 1, f'{no}' if first else '', BOLD9, CTR)
        put(ws2, r2, 2, first_txt if first else '', CELL, CTR)
        put(ws2, r2, 3, f'{strain}  ({sc} · {n_b} batches)' if first else '', BOLD9, LFT)
        put(ws2, r2, 4, f'QCSP_001_{sc}' if first else '', CELL, CTR)
        put(ws2, r2, 5, g['grade'], BOLD9, CTR)
        put(ws2, r2, 6, g['product_code'].replace(':', ' : '), CELL, CTR)
        put(ws2, r2, 7, gr_range(g), CELL, CTR)
        put(ws2, r2, 8, gr_tol(g), CELL, CTR)
        put(ws2, r2, 9, anchors, CELL, LFT)
        put(ws2, r2, 10, '; '.join(hist) if hist else '—', Font(size=8, italic=True), LFT)
        r2 += 1; first = False
        # documented gap row after this grade?
        for e in GD['exceptions']:
            mgap = re.match(rf'^{re.escape(strain)}: uncovered zone ([\d.]+)-([\d.]+) between '
                            rf'THC{g["nominal"]} and THC(\d+)', e)
            if mgap:
                za, zb, dn = mgap.group(1), mgap.group(2), mgap.group(3)
                for c in range(5, 11):
                    put(ws2, r2, c, '', CELL, CTR, GAPF)
                put(ws2, r2, 5, '⚠', BOLD9, CTR, GAPF)
                put(ws2, r2, 6, 'documented gap', Font(size=8, italic=True), CTR, GAPF)
                put(ws2, r2, 7, f'{za} – {zb} %', Font(size=8, italic=True), CTR, GAPF)
                put(ws2, r2, 9, 'нема резултат | no result on file — a future result here '
                                'triggers a grade-set revision', Font(size=8, italic=True), LFT, GAPF)
                r2 += 1
    # out-of-spec historical (superseded) results of the strain
    outs = []
    for res in results:
        v = float(res['value'])
        if not any(g['lower'] - 1e-9 <= v <= g['upper'] + 1e-9 for g in entry):
            outs.append(f"{res['batch']} {v:.2f} ({res['date']})")
    if outs:
        put(ws2, r2, 5, 'ⓘ', CELL, CTR)
        put(ws2, r2, 6, 'superseded, outside spec', Font(size=8, italic=True), CTR)
        put(ws2, r2, 9, '; '.join(outs) + ' — pre-retest values, replaced per the retest rule (R5)',
            Font(size=8, italic=True), LFT)
        r2 += 1

# ============ Sheet 3 — Batch Assignments ====================================
H3 = ['#','Batch No.','P-number','Strain','Tested Total Δ⁹-THC %','Basis','Tested (date, lab)',
      'Specification Code','Grade','Product Code','Range (Total Δ⁹-THC)','Nominal ± tolerance','Note']
W3 = [4,16,10,20,12,11,22,16,7,18,17,16,34]
ws3, r3 = sheet('Batch Assignments', H3, W3,
    'Purely Plant GmbH — Batch → Potency Grade Assignments (final, 27.08.2026)',
    '86 batches · anchor = CoQ-forming Total Δ⁹-THC (T1+T2 re-analysis of 26.08.2026 where retested; '
    'T2 values unofficial pending formal eCoAs; declared where no analysis) · '
    'mandatory management codes revised 27.08.2026: 42/48 exact, 6 flagged unavoidable deviations')

PNUM = {}
for row in CONS['rows']:
    if row.get('p_number'):
        PNUM[bkey(base_batch(row['batch']))] = row['p_number']
MANUAL_P = {'ACC102501':'P060122','CC012603':'P060372','CF102501':'P060132',
            'JD012603/01':'P060362','PUM102501':'P060112'}
# mandatory-code deviations, derived from the design flags (T2-revised codes)
DEVN = {}
for strain in GD['strains']:
    for g in GD['strains'][strain]:
        for b in g['batches']:
            f = next((f for f in GD['flags']
                      if f.startswith(f"{strain} / {b['batch']}:")
                      and ('infeasible' in f or 'MANDATORY-CODE DEVIATION' in f)), None)
            if f:
                txt = f.split(': ', 1)[1]
                if not txt.startswith('MANDATORY-CODE DEVIATION'):
                    txt = 'MANDATORY-CODE DEVIATION: ' + txt
                DEVN[b['batch']] = txt
i3 = 0
for strain in sorted(GD['strains']):
    for g in GD['strains'][strain]:
        for b in g['batches']:
            i3 += 1
            src = b['src'] if not is_decl(b['basis']) else 'owner-declared value'
            note = DEVN.get(b['batch'], '')
            if is_decl(b['basis']) and not note:
                note = 'declared — no eCoA yet, grade provisional'
            elif is_t2(b['basis']) and not note:
                note = 'T2 re-analysis (unofficial) — formal eCoA pending'
            fill = DEVF if b['batch'] in DEVN else (
                PENDF if is_decl(b['basis']) or is_t2(b['basis']) else None)
            put(ws3, r3, 1, i3, CELL, CTR, fill)
            put(ws3, r3, 2, b['batch'], BOLD9, LFT, fill)
            put(ws3, r3, 3, PNUM.get(bkey(b['batch']), MANUAL_P.get(b['batch'], '')), CELL, CTR, fill)
            put(ws3, r3, 4, strain, CELL, LFT, fill)
            put(ws3, r3, 5, f"{b['v']:.2f}", BOLD9, CTR, fill)
            basis_txt = ('declared' if is_decl(b['basis'])
                         else 'T2 re-analysis' if is_t2(b['basis']) else 'certificate')
            put(ws3, r3, 6, basis_txt, CELL, CTR, fill)
            put(ws3, r3, 7, src[:40], Font(size=8), LFT, fill)
            put(ws3, r3, 8, f"QCSP_001_{g['strain_code']}", CELL, CTR, fill)
            put(ws3, r3, 9, g['grade'], BOLD9, CTR, fill)
            put(ws3, r3, 10, g['product_code'].replace(':', ' : '), CELL, CTR, fill)
            put(ws3, r3, 11, gr_range(g), CELL, CTR, fill)
            put(ws3, r3, 12, gr_tol(g), CELL, CTR, fill)
            put(ws3, r3, 13, note, Font(size=8, italic=True), LFT, fill)
            r3 += 1

# ============ Sheet 4 — Stability Testing Programme (values copy) ============
src_wb = openpyxl.load_workbook(f'{Q}/PP_QC_eCoA_Master_Register.xlsx', data_only=True)
src_ws = src_wb['Stability Programme']
H4 = [src_ws.cell(4, c).value or '' for c in range(1, src_ws.max_column + 1)]
W4 = [5,20,14,10,14,11,11,26,10,7,7,7,8,9,9,10,26]
ws4, r4 = sheet('Stability Testing Programme', H4, W4,
    src_ws.cell(1, 1).value or 'Stability Testing Programme',
    src_ws.cell(2, 1).value or '')
for rr in range(5, src_ws.max_row + 1):
    vals = [src_ws.cell(rr, c).value for c in range(1, src_ws.max_column + 1)]
    if not any(v not in (None, '') for v in vals): continue
    for c, v in enumerate(vals, 1):
        put(ws4, r4, c, v if v is not None else '', CELL, CTR if c != 8 else LFT)
    r4 += 1

out = f'{Q}/QC_Batch_Release_Results_Register.xlsx'
wb.save(out)
print(f'saved {out}: S1 {S1_ROWS} rows, S2 {r2} rows, S3 {i3} batches, S4 {r4} rows')

# ============ Compact stacked CSV (Google-Sheet delivery) ====================
# Part 1 is a batch-level condensation of Sheet 1; Parts 2-3 are taken verbatim
# from the Specifications-by-Strain and Batch-Assignments worksheet cells so the
# CSV can never drift from the workbook.
import csv as _csv
cpath = f'{Q}/release_register_compact.csv'
FIELDS1 = ['total_cbd_pct','total_cbn_pct','loss_on_drying_pct','foreign_matter_pct',
           'tamc','tymc','bile_tolerant_gnb','salmonella','e_coli','aflatoxins_total',
           'aflatoxin_b1','ochratoxin_a','pb','cd','arsenic','hg','pesticides']
with open(cpath, 'w', newline='', encoding='utf-8') as fh:
    w = _csv.writer(fh)
    w.writerow(['QC BATCH RELEASE RESULTS REGISTER — FINAL POTENCY GRADES '
                '(27.08.2026, rev. 2 — T1+T2 re-analysis anchors of 26.08.2026)'])
    w.writerow(['Part 1: batch-level release results (CoQ-forming values) with assigned grade · '
                'Part 2: Specifications by Strain — every grade, range, and the results in it · '
                'Part 3: Batch -> Grade Assignments. Per-certificate detail: '
                'QC_Batch_Release_Results_Register.xlsx (letta-stack) and the eCoA Register sheet rev.5. '
                'Source: RAGFlow eCoA_DATABASE · ImB_Potency_Grade_Ranges · Potency Atlas. '
                'T2 values (Farmahem, 26.08.2026) are unofficial until the formal eCoAs are filed.'])
    w.writerow([])
    w.writerow(['PART 1 — BATCH RELEASE QC (batch level)'])
    w.writerow(['No.','Batch','P-number','Strain','THC %','Grade range %','Potency grade',
                'Spec code','CBD %','CBN %','LoD %','FM %','TAMC','TYMC','BTGNB','Salmonella',
                'E.coli','AflaΣ','AflaB1','OTA','Pb','Cd','As','Hg','Pesticides',
                'No. of certs','CoQ code'])
    n1 = 0
    for row in CONS['rows']:
        n1 += 1
        g = grade_of(row)
        thc = (f"{g['anchor']:.2f} (Т2 неофиц.)" if g and is_t2(g['basis'])
               else row.get('total_thc_pct') or '')
        def _dot(v):
            # bare decimal-comma numbers -> dots so Sheets parses them as numbers
            s = str(v)
            return s.replace(',', '.') if re.fullmatch(r'\d+,\d+', s) else s
        w.writerow([n1, row['batch'], row.get('p_number') or '', row.get('strain') or '', thc,
                    f"{g['lower']:.2f}-{g['upper']:.2f}" if g else '',
                    f"{g['grade']} · {g['product_code']}" if g else '',
                    g['spec_code'] if g else '']
                   + [_dot(row.get(f) or '') for f in FIELDS1]
                   + [row.get('n_certs') or '', row.get('coq') or ''])
    for bk, g in sorted(GRADE.items(), key=lambda kv: (kv[1]['strain'], kv[0])):
        if (is_decl(g['basis']) or is_t2(g['basis'])) and bk not in reg_norms:
            bname = next(b['batch'] for e in GD['strains'].values() for gg in e
                         for b in gg['batches'] if bkey(b['batch']) == bk)
            tag = 'declared' if is_decl(g['basis']) else 'Т2 26.08.2026, неофиц.'
            w.writerow(['', bname, '', g['strain'], f"{g['anchor']:.2f} ({tag})",
                        f"{g['lower']:.2f}-{g['upper']:.2f}",
                        f"{g['grade']} · {g['product_code']}", g['spec_code'],
                        'PENDING — no eCoA on file'])
    w.writerow([]); w.writerow([])
    w.writerow(['PART 2 — SPECIFICATIONS BY STRAIN (all grades · ranges · results within)'])
    w.writerow([ws2.cell(4, c).value or '' for c in range(1, 11)])
    for rr in range(5, r2):
        w.writerow([ws2.cell(rr, c).value if ws2.cell(rr, c).value is not None else ''
                    for c in range(1, 11)])
    w.writerow([]); w.writerow([])
    w.writerow(['PART 3 — BATCH -> GRADE ASSIGNMENTS (86 batches)'])
    w.writerow([ws3.cell(4, c).value or '' for c in range(1, 14)])
    for rr in range(5, r3):
        w.writerow([ws3.cell(rr, c).value if ws3.cell(rr, c).value is not None else ''
                    for c in range(1, 14)])
print(f'saved {cpath}')
