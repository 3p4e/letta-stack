#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Potency by Strain — grade declarations against the release register.

TWO sources only, per owner instruction:

  1. PP_Batch_Release_QC_Register_CORRECTED.xlsx — the outsourced-laboratory
     release register (its "THC by Strain" sheet supplies every Total Δ⁹-THC
     result per batch; its "Batch Release QC" sheet supplies the CoA →
     Drive/RAG document hyperlink map).
  2. potency_dataset.json — the Potency Atlas: the declared grade ladders
     (nominal ± tolerance) per strain, and each batch's tier assignment.

Nothing else from the certificates is carried over: of the CoA credentials
only the DOCUMENT CODE and DATE OF ISSUE appear, the code hyperlinked to the
ingested document. No micro / metals / mycotoxin / pesticide columns.

Layout follows the register's own "THC by Strain" sheet: one strain band,
then its batches, one row per test result. To the RIGHT of every row sits the
RENAMED-name block — where a batch was renamed, its new specification name
and that new name's own declared ladder (which can cluster differently from
the original strain's) are printed adjacent to the original data.
"""
import json
import os
import re
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

HERE = os.path.dirname(os.path.abspath(__file__))
REGISTER = os.path.join(HERE, "incoming",
                        "PP_Batch_Release_QC_Register_CORRECTED.xlsx")
OUT = os.path.join(HERE, "Potency_by_Strain.xlsx")

FONT = "Arial"
INK = "16232B"
NAVY = "1B3A5C"
GREEN = "14532B"
GFILL = "E9F4EC"
ROSE = "8A2E2E"
RFILL = "FDECEA"
GOLD = "8A6D14"
YFILL = "FFF8E7"
BAND = "EDF3F9"
ZEBRA = "F7FAFC"
LINE = "D8E2EC"
MUT = "4A5B6C"

d = json.load(open(os.path.join(HERE, "potency_dataset.json"), encoding="utf-8"))
PM = json.load(open(os.path.join(HERE, "portfolio_master.json"), encoding="utf-8"))

# the Atlas's own tier algorithm — the renamed-name ladders are computed with
# exactly the same rule the Atlas applies, never re-implemented here
sys.path.insert(0, HERE)
import build_potency_dataset as bpd
# the Atlas's canonical batch-key normaliser: the SAME function the dataset
# builder keys its stock and register rows by, so identities cannot diverge
import crosscheck_sources as cc

CANON = {"Cap Junkie": "Cap Junky", "CashCow": "Cash Cow", "GG4": "Gorilla Glue",
         "Grapes and Cream": "Graps & Creme", "Jelly Donuts": "Jelly Donutz",
         "Sleepy Joe": "Sleepy Joy", "Clemosa": "Clemosa a bud",
         "Wedding Crusher": "Wedding Crasher", "Appels & Bananas": "Apple and Banana",
         "Apples and Bananas": "Apple and Banana"}
# register spellings -> Atlas strain names
REG_ALIAS = {"Cup Junky": "Cap Junky", "CashCow": "Cash Cow",
             "OrangePunchMimosa": "Orange Punch Mimosa", "JellyDonutz": "Jelly Donutz",
             "Jokerz31": "Jokerz 31", "FatBastard": "Fat Bastard",
             "GorillaGlue": "Gorilla Glue", "GrapePie": "Grape Pie"}


def normb(b):
    return cc.norm_batch(b)


def num(text):
    """Leading plain percentage of a printed result; None for <LOQ / prose."""
    if not text:
        return None
    head = str(text).split("(")[0]
    if "<" in head or "≤" in head:
        return None
    m = re.search(r"(\d+(?:[.,]\d+)?)", head)
    return round(float(m.group(1).replace(",", ".")), 2) if m else None


# ---------------------------------------------------------------- register --
rwb = openpyxl.load_workbook(REGISTER, data_only=False)

coa2link = {}
for r in rwb["Batch Release QC"].iter_rows(min_row=6):
    code, cell = r[22].value, r[25]
    if code and cell.hyperlink:
        coa2link.setdefault(str(code).strip(), cell.hyperlink.target)

hdr_re = re.compile(r"^\s*(.+?)\s{2,}(\d+)\s+batch")
records, strain, batch, pnum = [], None, None, None
for r in rwb["THC by Strain"].iter_rows(min_row=5):
    v = [c.value for c in r]
    if v[0] is not None and v[1] is None and v[3] is None:
        m = hdr_re.match(str(v[0]))
        raw = m.group(1).strip() if m else str(v[0]).strip()
        strain = REG_ALIAS.get(raw, raw)
        continue
    if v[1] is not None:
        batch, pnum = str(v[1]).strip(), (str(v[2]).strip() if v[2] else None)
    if v[3] is None:
        continue
    coa = str(v[5] or "").strip()
    records.append(dict(strain=strain, batch=batch, key=normb(batch), p=pnum,
                        printed=str(v[3]).strip(), value=num(v[3]),
                        coa=coa, date=str(v[6] or "").strip(),
                        link=coa2link.get(coa)))

assert records, "no records parsed from the register"

# The register books some rows under the PACKAGING P-number (the 07.08.2026
# Farmahem retests) and spells a few suffixes differently (FB012601 vs
# FB012601/1). The certificate code is the unambiguous join key, so every
# record's batch identity is resolved through it before anything is matched
# against the Atlas — never by string-guessing the batch number.
cert2batch = {}
for rr in d["register_results"]:
    cert2batch.setdefault(rr["cert"].strip(), set()).add(normb(rr["batch"]))
resolved = {}
for r in records:
    hit = cert2batch.get(r["coa"])
    if hit and len(hit) == 1:
        resolved.setdefault(r["key"], next(iter(hit)))
atlas_spelling = {}
for rr in d["register_results"]:
    atlas_spelling.setdefault(normb(rr["batch"]), rr["batch"].strip())
for r in records:
    k = resolved.get(r["key"])
    if k is not None and k != r["key"]:
        # the row was booked under the packaging P-number; show the flower
        # batch it certifies and keep the P-number in its own column
        if re.fullmatch(r"P\d{6}", r["batch"].strip().upper()):
            r["p"] = r["p"] or r["batch"].strip()
            r["batch"] = atlas_spelling.get(k, r["batch"])
        r["key"] = k

# ------------------------------------------------------------------- atlas --
tiers_by_strain = d["merged_ranges"]
stock = {normb(b["batch"]): b for b in d["stock"]}

# A handful of batches are spelled differently in the tranche sheet than on
# their certificate (e.g. stock "BSS1024_01/1" vs eCoA "BSS1024_01"). Rather
# than guess from the string, pair them on EVIDENCE: same strain, and the
# stock row's governing value equals one of that batch's register results.
# Only a unique, unambiguous pairing is accepted; anything else is left
# unmatched and shows as "not in stock" rather than being forced.
_unmatched = {r["key"] for r in records} - set(stock)
for uk in sorted(_unmatched):
    rs = [r for r in records if r["key"] == uk]
    vals = {r["value"] for r in rs if r["value"] is not None}
    sname = rs[0]["strain"]
    cands = [k for k, b in stock.items()
             if k not in {r["key"] for r in records}
             and b["strain"] == sname
             and (b["anchor"] if b["anchor"] is not None else b["declared"]) in vals]
    if len(cands) == 1:
        for r in rs:
            r["key"] = cands[0]
        print("  paired by evidence: register %-14s -> stock %s" % (uk, cands[0]))

anchors = {}
for b in d["stock"]:
    a = b["anchor"] if b["anchor"] is not None else b["declared"]
    if a is not None:
        anchors[normb(b["batch"])] = (a, b["anchor"] is not None)

# renamed layer: batch -> new name; new name -> its own Atlas ladder
neu_of, groups = {}, {}
for rec in PM:
    neu = (rec.get("neu") or "").strip()
    if not neu:
        continue
    k = normb(rec["batch"])
    neu_of[k] = dict(neu=neu, brand=(rec.get("brand") or "").strip(),
                     orig=CANON.get((rec.get("original") or "").strip(),
                                    (rec.get("original") or "").strip()))
    groups.setdefault(neu, []).append(k)

override_by_top_anchor = {}
_top = {}
for b in d["stock"]:
    a = b["anchor"] if b["anchor"] is not None else b["declared"]
    if a is not None:
        _top[b["strain"]] = max(_top.get(b["strain"], -1.0), a)
for s, v in d["design"].get("top_nominal_override", {}).items():
    if s in _top:
        override_by_top_anchor[round(_top[s], 2)] = v

neu_tiers = {}          # new name -> list of tier dicts
neu_tier_of_batch = {}  # batch -> (new name, tier index 1-based, tier dict)
for neu, keys in groups.items():
    items = [(anchors[k][0], k) for k in keys if k in anchors]
    if not items:
        continue
    items.sort(key=lambda x: x[0])
    ov = override_by_top_anchor.get(round(max(a for a, _k in items), 2))
    tl = bpd.build_strain_tiers(items, top_override=ov)
    neu_tiers[neu] = tl
    for i, t in enumerate(tl, 1):
        for k in t["payloads"]:
            neu_tier_of_batch[k] = (neu, i, t)

# ------------------------------------------------------------------ layout --
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Potency by Strain"
ws.sheet_view.showGridLines = False

thin = Side(style="thin", color=LINE)
box = Border(left=thin, right=thin, top=thin, bottom=thin)


def put(row, col, value, *, size=10, color=INK, bold=False, italic=False,
        fill=None, align="center", link=None, wrap=False, border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT, size=size, bold=bold, italic=italic,
                  color=("0563C1" if link else color),
                  underline=("single" if link else None))
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if border:
        c.border = box
    if link:
        c.hyperlink = link
    return c


NCOL = 18
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOL)
put(1, 1, "Purely Plant GmbH — Потенција по сорта: декларирани класи наспроти "
          "резултатите од регистарот | Potency by Strain: declared grades vs the "
          "release register", size=13, bold=True, color=NAVY, align="left", border=False)
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOL)
put(2, 1, "Извори: PP_Batch_Release_QC_Register_CORRECTED.xlsx (сите резултати за "
          "Вкупен Δ⁹-THC по серија, шифра и датум на еCoA, врска до документот) + "
          "Атлас на потенција (декларирани класи: номинала ± толеранција). Од "
          "сертификатите се пренесени само шифрата и датумот на издавање. | Sources: "
          "the corrected release register (every Total Δ⁹-THC result per batch, its "
          "eCoA code and issue date, linked to the ingested document) + the Potency "
          "Atlas (declared grades: nominal ± tolerance). Of the certificate "
          "credentials only the code and date of issue are carried over.",
    size=9, italic=True, color=MUT, align="left", wrap=True, border=False)
ws.row_dimensions[2].height = 42

ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=11)
put(4, 1, "ОРИГИНАЛНА СОРТА | ORIGINAL STRAIN", size=10, bold=True,
    color="FFFFFF", fill=NAVY)
ws.merge_cells(start_row=4, start_column=12, end_row=4, end_column=NCOL)
put(4, 12, "ПО ПРЕИМЕНУВАЊЕ | AFTER RENAMING", size=10, bold=True,
    color="FFFFFF", fill=GOLD)

HEAD = ["№", "Серија | Batch", "P-серија | P-number", "Вкупен Δ⁹-THC | Result",
        "Сидро | Anchor", "Шифра на eCoA | eCoA code", "Датум | Date of issue",
        "Класа | Tier", "Номинала | Nominal", "Толеранција | Tolerance",
        "Опсег | Range",
        "Ново име | New name", "Бренд | Brand", "Класа | Tier",
        "Номинала | Nominal", "Толеранција | Tolerance", "Опсег | Range",
        "Во опсег? | In range?"]
for i, h in enumerate(HEAD, 1):
    put(5, i, h, size=8.5, bold=True, color="FFFFFF",
        fill=(NAVY if i <= 11 else GOLD), wrap=True)
ws.row_dimensions[5].height = 30

row = 6
n_rows = n_batches = n_in = n_out = 0
by_strain = {}
for r in records:
    by_strain.setdefault(r["strain"], []).append(r)

for strain in sorted(by_strain):
    recs = by_strain[strain]
    tl = tiers_by_strain.get(strain, [])
    bkeys = []
    for r in recs:
        if r["key"] not in bkeys:
            bkeys.append(r["key"])
    vals = [r["value"] for r in recs if r["value"] is not None]
    band = ("%s — %d серии | batches · %d резултати | results" % (strain, len(bkeys), len(recs)))
    if vals:
        band += " · %.2f%% – %.2f%%" % (min(vals), max(vals))
    if tl:
        band += "  ·  " + " + ".join("Pot.-%d %.2f%% ±%.2f%% (%.2f%%–%.2f%%)"
                                     % (i, t["nominal"], t["tol"], t["range"][0], t["range"][1])
                                     for i, t in enumerate(tl, 1))
    else:
        band += "  ·  нема декларирана класа | no declared grade"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOL)
    put(row, 1, band, size=10, bold=True, color=NAVY, fill=BAND, align="left")
    row += 1

    seq = 0
    for k in bkeys:
        brs = [r for r in recs if r["key"] == k]
        brs.sort(key=lambda r: tuple(reversed((r["date"] or "").split("."))))
        seq += 1
        n_batches += 1
        st = stock.get(k)
        anc = anchors.get(k)
        tier_i = st.get("tier") if st else None
        t = tl[tier_i - 1] if (tl and tier_i and tier_i <= len(tl)) else None
        nt = neu_tier_of_batch.get(k)
        ren = neu_of.get(k)

        for j, r in enumerate(brs):
            zeb = ZEBRA if n_batches % 2 == 0 else "FFFFFF"
            is_anchor = (anc is not None and anc[1] and r["value"] is not None
                         and abs(r["value"] - anc[0]) < 1e-9)
            put(row, 1, seq if j == 0 else None, size=9, color=MUT, fill=zeb)
            put(row, 2, r["batch"] if j == 0 else None, size=9.5,
                bold=(j == 0), fill=zeb)
            put(row, 3, r["p"] if j == 0 else None, size=9, color=MUT, fill=zeb)
            # house numeric rule: two decimals, % hard against the digit. The
            # register's verbatim string (which may carry the expanded
            # uncertainty) is preserved as a cell note so nothing is lost.
            shown = ("%.2f%%" % r["value"]) if r["value"] is not None else r["printed"]
            c = put(row, 4, shown, size=9.5, bold=True, color=GREEN, fill=zeb)
            if r["printed"] and r["printed"] != shown:
                c.comment = Comment("Регистар | Register (verbatim):\n%s" % r["printed"],
                                    "Potency by Strain")
            put(row, 5, "✔" if is_anchor else "", size=10, bold=True,
                color=GREEN, fill=zeb)
            put(row, 6, r["coa"] or "—", size=9, fill=zeb, link=r["link"])
            put(row, 7, r["date"] or "—", size=9, color=MUT, fill=zeb)

            # declared grade — printed once per batch, on its first row
            if j == 0 and t is not None:
                put(row, 8, "Pot.-%d" % tier_i, size=9.5, bold=True, color=GREEN, fill=GFILL)
                put(row, 9, "%.2f%%" % t["nominal"], size=9.5, bold=True, color=GREEN, fill=GFILL)
                put(row, 10, "± %.2f%%" % t["tol"], size=9.5, color=GREEN, fill=GFILL)
                put(row, 11, "%.2f%% – %.2f%%" % (t["range"][0], t["range"][1]),
                    size=9.5, color=GREEN, fill=GFILL)
            elif j == 0:
                put(row, 8, "нема | none", size=8.5, italic=True, color=MUT, fill=YFILL)
                ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=11)
                put(row, 9, "не е во залихата Т1/Т2/Т3 | not in T1/T2/T3 stock",
                    size=8.5, italic=True, color=MUT, fill=YFILL)
            else:
                for c in (8, 9, 10, 11):
                    put(row, c, None, fill=zeb)

            # renamed block
            if j == 0 and ren is not None:
                same = ren["neu"].strip().lower() == strain.strip().lower()
                put(row, 12, ren["neu"], size=9.5, bold=not same,
                    color=(MUT if same else NAVY), fill=(zeb if same else YFILL))
                put(row, 13, ren["brand"] or "—", size=8.5, color=MUT,
                    fill=(zeb if same else YFILL))
                if nt is not None:
                    _neu, ni, ntier = nt
                    put(row, 14, "Pot.-%d" % ni, size=9.5, bold=True, color=GOLD, fill=YFILL)
                    put(row, 15, "%.2f%%" % ntier["nominal"], size=9.5, bold=True,
                        color=GOLD, fill=YFILL)
                    put(row, 16, "± %.2f%%" % ntier["tol"], size=9.5, color=GOLD, fill=YFILL)
                    put(row, 17, "%.2f%% – %.2f%%" % (ntier["lo"], ntier["hi"]),
                        size=9.5, color=GOLD, fill=YFILL)
                else:
                    for c in (14, 15, 16, 17):
                        put(row, c, "—" if c == 14 else None, size=9, color=MUT, fill=YFILL)
            elif j == 0:
                put(row, 12, "— не е во мастерот | not in master", size=8.5,
                    italic=True, color=MUT, fill=zeb)
                for c in (13, 14, 15, 16, 17):
                    put(row, c, None, fill=zeb)
            else:
                for c in range(12, 18):
                    put(row, c, None, fill=zeb)

            # in-range verdict, per result against its batch's declared tier
            if t is not None and r["value"] is not None:
                ok = t["range"][0] - 1e-6 <= r["value"] <= t["range"][1] + 1e-6
                if ok:
                    n_in += 1
                    put(row, 18, "ДА | YES", size=9, bold=True, color=GREEN, fill=GFILL)
                else:
                    n_out += 1
                    put(row, 18, "НАДВОР | OUTSIDE", size=8.5, bold=True,
                        color=ROSE, fill=RFILL)
            else:
                put(row, 18, "—", size=9, color=MUT, fill=zeb)
            row += 1
            n_rows += 1

# legend
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOL)
put(row, 1, "Сидро ✔ = најновиот верификуван резултат — само тој ја одредува "
            "класата на серијата; постарите резултати се прикажани заради историја и "
            "затоа некои паѓаат надвор од тековниот опсег. Шифрата на eCoA е врска до "
            "ингестираниот документ. | Anchor ✔ = the most recent verified result — only "
            "it sets the batch's grade; earlier results are shown for history, which is "
            "why some fall outside the current range. The eCoA code links to the "
            "ingested document.", size=8.5, italic=True, color=MUT, align="left",
    wrap=True, border=False)
ws.row_dimensions[row].height = 30

widths = [4.5, 16, 12, 20, 8, 22, 13, 8.5, 11, 12, 17,
          20, 9, 8.5, 11, 12, 17, 15]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = ws.cell(row=6, column=1)
ws.auto_filter.ref = "A5:%s5" % get_column_letter(NCOL)

# ------------------------------- sheet 2: stock without a register result --
ws2 = wb.create_sheet("Stock without register result")
ws2.sheet_view.showGridLines = False
ws = ws2  # put() writes to the module-level `ws`
ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
put(1, 1, "Серии на залиха БЕЗ резултат во регистарот за пуштање | Stock batches with "
          "NO result in the release register", size=12, bold=True, color=NAVY,
    align="left", border=False)
ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
put(2, 1, "Тука не се губи ниту една серија: овие се декларирани на основа на "
          "декларираната вредност (без сертификат), или носат интерен PP сертификат "
          "што е надвор од опфатот на регистарот за надворешни лаборатории. | No batch "
          "is silently dropped: these are declared on the declared value (no "
          "certificate on file), or carry an in-house PP CoA which is outside the "
          "outsourced-laboratory register's scope.",
    size=9, italic=True, color=MUT, align="left", wrap=True, border=False)
ws2.row_dimensions[2].height = 32

for i, h in enumerate(["Сорта | Strain", "Серија | Batch", "Транша | Tranche",
                       "Основа | Basis", "Вредност | Value", "Класа | Tier",
                       "Номинала ± толер. | Nominal ± tol.", "Опсег | Range"], 1):
    put(4, i, h, size=8.5, bold=True, color="FFFFFF", fill=NAVY, wrap=True)
ws2.row_dimensions[4].height = 26

reg_keys = {r["key"] for r in records}
row2, n_missing = 5, 0
for b in sorted(d["stock"], key=lambda x: (x["strain"], x["batch"])):
    k = normb(b["batch"])
    if k in reg_keys or not b.get("proposed"):
        continue
    n_missing += 1
    zeb = ZEBRA if n_missing % 2 == 0 else "FFFFFF"
    decl = b["anchor"] is None
    v = b["declared"] if decl else b["anchor"]
    put(row2, 1, b["strain"], size=9.5, align="left", fill=zeb)
    put(row2, 2, b["batch"], size=9.5, bold=True, fill=zeb)
    put(row2, 3, "Т%s" % b["tranche"], size=9, color=MUT, fill=zeb)
    put(row2, 4, "декларирана вредност (без сертификат) | declared (no certificate)"
        if decl else "интерен PP сертификат | in-house PP CoA",
        size=8.5, italic=True, color=(GOLD if decl else MUT), align="left",
        fill=(YFILL if decl else zeb))
    put(row2, 5, "%.2f%%" % v, size=9.5, bold=True, fill=zeb)
    put(row2, 6, "Pot.-%s" % b.get("tier", "—"), size=9.5, color=GREEN, fill=GFILL)
    put(row2, 7, "%.2f%% ± %.2f%%" % (b["nominal"], b["tol"]), size=9.5,
        bold=True, color=GREEN, fill=GFILL)
    put(row2, 8, "%.2f%% – %.2f%%" % (b["proposed"][0], b["proposed"][1]),
        size=9.5, color=GREEN, fill=GFILL)
    row2 += 1

for i, w in enumerate([22, 15, 8, 46, 11, 9, 20, 18], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = ws2.cell(row=5, column=1)

# ------------------------------------------------------------------ checks --
reg_batches = {r["key"] for r in records}
atlas_batches = {normb(b["batch"]) for b in d["stock"] if b.get("proposed")}
print("SELF-CHECK")
print("  stock rows w/o register result:", n_missing)
print("  register results        :", n_rows)
print("  batches                 :", n_batches)
print("  strains                 :", len(by_strain))
print("  results in declared tier:", n_in)
print("  results outside (older) :", n_out)
print("  linked eCoA codes       :", sum(1 for r in records if r["link"]), "/", len(records))
print("  register batches not in Atlas stock:", sorted(reg_batches - atlas_batches))
print("  Atlas stock batches not in register:", sorted(atlas_batches - reg_batches))
assert n_rows == len(records), (n_rows, len(records))

wb.save(OUT)
print("wrote", OUT, os.path.getsize(OUT), "bytes")
