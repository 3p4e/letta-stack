#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Export-class distribution — the owner's 8…28 grid set against our own data.

The customer's export list ("THC level,qty by batch.xlsx") declares eleven
export classes (8, 10, 12 … 28), each with a -10% / +10% limit, and books
every tranche batch's quantity into one of them. Those windows OVERLAP
(18 -> 16.20-19.80, 20 -> 18.00-22.00, 22 -> 19.80-24.20), so the class of a
result is not determined by the windows alone. The rule was recovered by
reconciling the list's own per-tranche quantity tables:

    a batch is booked into the HIGHEST class whose +/-10% window
    contains its THC value

Under that rule every one of the 33 per-tranche class quantities reproduces
exactly, with a single documented exception (see BELOW_ALL) — so the rule is
not an assumption, it is the list's own arithmetic.

This module adds two sheets to the register+Atlas workbook:
  * "Export Classes"       — every tranche batch: export list vs our register
                             result vs our declared Atlas grade vs its class.
  * "Export Class Totals"  — the per-class quantity roll-up per tranche, our
                             recomputation set against the list's own figures.
"""
import json
import os

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "incoming", "export_list_2026-08-18.json")

FONT = "Arial"
INK = "16232B"
NAVY = "1B3A5C"
GREEN = "14532B"
GFILL = "E9F4EC"
ROSE = "8A2E2E"
RFILL = "FDECEA"
GOLD = "8A6D14"
YFILL = "FFF8E7"
BAND = "EDF3F9"
ZEBRA = "F7FAFC"
LINE = "D8E2EC"
MUT = "4A5B6C"

TOL = 0.10          # the export grid's own +/-10% on the class nominal
EPS = 1e-9

# The one batch the list books into a class whose window does not reach it:
# Graps & Creme P060142 at 7.05% sits 0.15 pp below class 8's 7.20% floor
# (-2.08% of the nominal) yet is counted in Tranche 3's class-8 quantity.
BELOW_ALL = {"P060142": 8}


def window(c):
    return (round(c * (1 - TOL), 4), round(c * (1 + TOL), 4))


def classify(v):
    """Highest export class whose +/-10% window contains v; None if none does."""
    hits = [c for c in CLASSES if window(c)[0] - EPS <= v <= window(c)[1] + EPS]
    return max(hits) if hits else None


data = json.load(open(SRC, encoding="utf-8"))
CLASSES = data["classes"]


def add_sheets(wb, ctx):
    """Append the two export-class sheets. ctx supplies the verified joins."""
    records = ctx["records"]
    stock = ctx["stock"]
    anchors = ctx["anchors"]
    tiers_by_strain = ctx["tiers_by_strain"]
    normb = ctx["normb"]

    # ---- join keys: the export list books batches by P-number, and a handful
    # by the flower batch itself. Both routes resolve to the Atlas batch key.
    by_p, by_b = {}, {}
    for b, pc in ctx["p_codes"].items():
        by_p.setdefault(str(pc).strip().upper(), normb(b))
    for rr in ctx["register_results"]:
        if rr.get("p_code"):
            by_p.setdefault(str(rr["p_code"]).strip().upper(), normb(rr["batch"]))
    for r in records:
        if r["p"]:
            by_p.setdefault(str(r["p"]).strip().upper(), r["key"])
        by_b.setdefault(normb(r["batch"]), r["key"])
    for b in stock.values():
        by_b.setdefault(normb(b["batch"]), normb(b["batch"]))

    # Third route, for the tranche-2/3 packaging numbers that exist only on the
    # export list: pair on EVIDENCE — same strain, and the list's THC equals a
    # value we hold for exactly one batch of that strain. Ambiguous or absent
    # matches are left unresolved rather than guessed.
    CANON_LIST = {"CashCow": "Cash Cow", "GG4": "Gorilla Glue",
                  "Grapes and Cream": "Graps & Creme",
                  "Jelly Donuts": "Jelly Donutz", "Sleepy Joe": "Sleepy Joy",
                  "Clemosa A Bud": "Clemosa a bud",
                  "Wedding Crusher": "Wedding Crasher",
                  "Appels & Bananas": "Apple and Banana"}
    by_val = {}
    for rr in ctx["register_results"]:
        by_val.setdefault((rr["strain"], round(rr["value"], 2)),
                          set()).add(normb(rr["batch"]))
    for b in ctx["stock_rows"]:
        a = b["anchor"] if b["anchor"] is not None else b["declared"]
        if a is not None:
            by_val.setdefault((b["strain"], round(a, 2)),
                              set()).add(normb(b["batch"]))

    def resolve(code, strain=None, thc=None):
        c = str(code).strip().upper()
        k = by_p.get(c) or by_b.get(normb(code))
        if k is not None:
            return k, "P-број / серија | P-number / batch"
        if strain is not None and thc is not None:
            cands = by_val.get((CANON_LIST.get(strain, strain), round(thc, 2)))
            if cands and len(cands) == 1:
                return next(iter(cands)), "по доказ: сорта + вредност | by evidence"
        return None, None

    thin = Side(style="thin", color=LINE)
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws = wb.create_sheet("Export Classes")
    ws.sheet_view.showGridLines = False

    def put(row, col, value, *, size=10, color=INK, bold=False, italic=False,
            fill=None, align="center", wrap=False, border=True, sheet=None):
        c = (sheet or ws).cell(row=row, column=col, value=value)
        c.font = Font(name=FONT, size=size, bold=bold, italic=italic, color=color)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
        if border:
            c.border = box
        return c

    N = 17
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N)
    put(1, 1, "Purely Plant GmbH — Распределба по извозни класи: транши и серии "
              "наспроти нашите резултати и декларирани опсези | Export-class "
              "distribution: tranches and batches against our results and declared "
              "ranges", size=13, bold=True, color=NAVY, align="left", border=False)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N)
    put(2, 1, "Извори: извозната листа „%s“ (класи, THC на листата, количини) + "
              "PP_Batch_Release_QC_Register_CORRECTED.xlsx (нашиот резултат) + "
              "Атласот на потенција (декларирана класа: номинала ± толеранција). | "
              "Sources: the export list “%s” (classes, list THC, quantities) + the "
              "corrected release register (our result) + the Potency Atlas (declared "
              "grade: nominal ± tolerance)." % (data["source"]["name"],
                                                data["source"]["name"]),
        size=9, italic=True, color=MUT, align="left", wrap=True, border=False)
    ws.row_dimensions[2].height = 30

    # the class rule — stated ONCE, here, and never repeated per batch
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=N)
    put(3, 1, "ПРАВИЛО ЗА ИЗВОЗНАТА КЛАСА | THE EXPORT-CLASS RULE — Прозорците на "
              "класите се преклопуваат (18 → 16,20–19,80; 20 → 18,00–22,00; "
              "22 → 19,80–24,20), па само границите не ја определуваат класата. "
              "Правилото е реконструирано од самите количински табели на листата: "
              "серијата оди во НАЈВИСОКАТА класа чиј прозорец ±10% ја содржи "
              "вредноста. Под тоа правило сите 33 количини по транша и класа се "
              "репродуцираат точно, со еден исклучок (P060142, 7,05% — под подот од "
              "7,20% на класата 8, а сепак е броена таму). | The class windows "
              "overlap, so the limits alone do not fix the class. The rule was "
              "recovered from the list's own quantity tables: a batch goes to the "
              "HIGHEST class whose ±10% window contains its value. Under it all 33 "
              "per-tranche class quantities reproduce exactly, with one exception "
              "(P060142, 7.05% — below class 8's 7.20% floor, yet counted there). "
              "Декларираната класа е секогаш онаа на ОРИГИНАЛНАТА сорта — "
              "преименувањето е ознака кај купувачот, не PP сорта. | The declared "
              "grade is always the ORIGINAL strain's — a renaming is the buyer's "
              "label, not a PP strain.",
        size=8.5, italic=True, color=NAVY, fill=BAND, align="left", wrap=True)
    ws.row_dimensions[3].height = 74

    HEAD = ["№", "Транша | Tranche", "Сорта (листа) | Strain (list)",
            "Ново име кај купувачот | Buyer's new name", "Серија на листата | List batch no.",
            "Наша серија | Our batch", "THC на листата | List THC",
            "Наш резултат | Our result", "Δ (пп | pp)",
            "Наша класа | Our tier", "Наш опсег | Our range",
            "Извозна класа | Export class", "Долна −10% | Lower",
            "Горна +10% | Upper", "Класа по наш резултат | Class on our result",
            "Количина (kg) | Quantity", "Забелешка | Note"]
    for i, h in enumerate(HEAD, 1):
        put(5, i, h, size=8.5, bold=True, color="FFFFFF",
            fill=(NAVY if i <= 11 else GOLD), wrap=True)
    ws.row_dimensions[5].height = 40

    row = 6
    n = 0
    rollup = {t: {c: 0.0 for c in CLASSES} for t in (1, 2, 3)}
    stats = dict(matched=0, unmatched=0, moved=0, mismatch=0)
    rows_out = []

    for tr in (1, 2, 3):
        items = [b for b in data["batches"] if b["tranche"] == tr]
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N)
        put(row, 1, "ТРАНША %d | TRANCHE %d — %d серии | batches · %.3f kg"
            % (tr, tr, len(items), sum(i["qty"] for i in items)),
            size=10.5, bold=True, color=NAVY, fill=BAND, align="left")
        row += 1

        for it in items:
            n += 1
            zeb = ZEBRA if n % 2 == 0 else "FFFFFF"
            key, how = resolve(it["batch"], it["orig"], it["thc"])
            anc = anchors.get(key) if key else None
            ours = anc[0] if anc else None
            st = stock.get(key) if key else None
            tier_i = st.get("tier") if st else None
            tl = tiers_by_strain.get(st["strain"], []) if st else []
            t = tl[tier_i - 1] if (tl and tier_i and tier_i <= len(tl)) else None

            cls = BELOW_ALL.get(str(it["batch"]).strip().upper()) or classify(it["thc"])
            ocls = classify(ours) if ours is not None else None
            rollup[tr][cls] = rollup[tr].get(cls, 0.0) + it["qty"]

            notes = []
            if key is None:
                notes.append("нема во регистарот/Атласот | not in register/Atlas")
                stats["unmatched"] += 1
            else:
                stats["matched"] += 1
                stats.setdefault(how, 0)
                stats[how] += 1
                if how.startswith("по доказ"):
                    notes.append("серијата е спарена по доказ (сорта + вредност) | "
                                 "batch paired by evidence")
            if ours is not None and abs(ours - it["thc"]) > 0.005:
                notes.append("листата: %.2f%% ≠ наш: %.2f%% | list vs ours"
                             % (it["thc"], ours))
                stats["mismatch"] += 1
            if ocls is not None and cls is not None and ocls != cls:
                notes.append("по наш резултат: класа %d | on our result: class %d"
                             % (ocls, ocls))
                stats["moved"] += 1
            if str(it["batch"]).strip().upper() in BELOW_ALL:
                notes.append("под подот на класата 8 (7,20%) | below class 8's floor")

            put(row, 1, n, size=9, color=MUT, fill=zeb)
            put(row, 2, "Т%d" % tr, size=9, color=MUT, fill=zeb)
            put(row, 3, it["orig"], size=9.5, align="left", fill=zeb)
            same = it["new"].strip().lower() == it["orig"].strip().lower()
            put(row, 4, it["new"], size=9.5, bold=not same,
                color=(MUT if same else GOLD), align="left",
                fill=(zeb if same else YFILL))
            put(row, 5, it["batch"], size=9.5, bold=True, fill=zeb)
            put(row, 6, (st["batch"] if st else (key or "—")), size=9, color=MUT, fill=zeb)
            put(row, 7, "%.2f%%" % it["thc"], size=9.5, fill=zeb)
            put(row, 8, ("%.2f%%" % ours) if ours is not None else "—",
                size=9.5, bold=True, color=GREEN, fill=zeb)
            put(row, 9, ("%+.2f" % (ours - it["thc"])) if ours is not None else "—",
                size=9, color=(ROSE if ours is not None and abs(ours - it["thc"]) > 0.005
                               else MUT), fill=zeb)
            if t is not None:
                put(row, 10, "Pot.-%d" % tier_i, size=9.5, bold=True, color=GREEN, fill=GFILL)
                put(row, 11, "%.2f%% ± %.2f%% (%.2f%%–%.2f%%)"
                    % (t["nominal"], t["tol"], t["range"][0], t["range"][1]),
                    size=9, color=GREEN, fill=GFILL)
            else:
                put(row, 10, "—", size=9, color=MUT, fill=YFILL)
                put(row, 11, "нема декларирана класа | no declared grade",
                    size=8.5, italic=True, color=MUT, fill=YFILL)
            lo, hi = window(cls)
            put(row, 12, cls, size=10, bold=True, color=GOLD, fill=YFILL)
            put(row, 13, "%.2f" % lo, size=9, color=GOLD, fill=YFILL)
            put(row, 14, "%.2f" % hi, size=9, color=GOLD, fill=YFILL)
            put(row, 15, (ocls if ocls is not None else "—"), size=9.5,
                bold=(ocls != cls), color=(ROSE if ocls != cls else MUT),
                fill=(RFILL if ocls != cls else zeb))
            put(row, 16, round(it["qty"], 3), size=9.5, fill=zeb)
            put(row, 17, "; ".join(notes) if notes else "", size=8,
                italic=True, color=MUT, align="left", wrap=True, fill=zeb)
            rows_out.append((tr, it, cls, ocls, ours))
            row += 1

    for i, w in enumerate([4.5, 8, 20, 20, 15, 14, 12, 12, 9, 9, 26,
                           9, 9, 9, 11, 11, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=6, column=1)
    ws.auto_filter.ref = "A5:%s5" % get_column_letter(N)

    # ------------------------------------------------ per-class roll-up sheet
    ws2 = wb.create_sheet("Export Class Totals")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    put(1, 1, "Количини по извозна класа: наша пресметка наспроти извозната листа | "
              "Quantity per export class: our recomputation vs the export list",
        size=12.5, bold=True, color=NAVY, align="left", border=False, sheet=ws2)
    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    put(2, 1, "Секоја количина е збир на сериите распоредени по правилото погоре. "
              "Разликите не се израмнуваат — се прикажуваат. | Every quantity is the "
              "sum of the batches assigned by the rule above. Differences are shown, "
              "never silently reconciled.", size=9, italic=True, color=MUT,
        align="left", wrap=True, border=False, sheet=ws2)

    r2 = 4
    grand = {c: 0.0 for c in CLASSES}
    for tr in (1, 2, 3, "all"):
        label = ("ТРАНША %s | TRANCHE %s" % (tr, tr)) if tr != "all" \
            else "СИТЕ ТРАНШИ | ALL TRANCHES"
        ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=7)
        put(r2, 1, label, size=10.5, bold=True, color=NAVY, fill=BAND,
            align="left", sheet=ws2)
        r2 += 1
        for i, h in enumerate(["THC (%)", "Долна −10% | Lower", "Горна +10% | Upper",
                               "Количина (наша) | Quantity (ours)",
                               "Количина (листа) | Quantity (list)",
                               "Разлика | Difference", "Серии | Batches"], 1):
            put(r2, i, h, size=8.5, bold=True, color="FFFFFF", fill=NAVY,
                wrap=True, sheet=ws2)
        ws2.row_dimensions[r2].height = 26
        r2 += 1
        exp = data["class_quantities"][str(tr)]
        tot_o = tot_l = 0.0
        for c in CLASSES:
            if tr == "all":
                mine = grand[c]
                cnt = sum(1 for (t_, it, cl, _o, _v) in rows_out if cl == c)
            else:
                mine = rollup[tr][c]
                grand[c] += mine
                cnt = sum(1 for (t_, it, cl, _o, _v) in rows_out
                          if t_ == tr and cl == c)
            listed = float(exp[str(c)] if str(c) in exp else exp[c])
            diff = mine - listed
            lo, hi = window(c)
            tot_o += mine
            tot_l += listed
            ok = abs(diff) <= 0.05
            put(r2, 1, c, size=10, bold=True, color=GOLD, fill=YFILL, sheet=ws2)
            put(r2, 2, "%.2f" % lo, size=9, color=MUT, sheet=ws2)
            put(r2, 3, "%.2f" % hi, size=9, color=MUT, sheet=ws2)
            put(r2, 4, round(mine, 3), size=9.5, bold=True, color=GREEN,
                fill=GFILL, sheet=ws2)
            put(r2, 5, round(listed, 3), size=9.5, color=INK, sheet=ws2)
            put(r2, 6, ("—" if ok else "%+.3f" % diff), size=9,
                bold=not ok, color=(MUT if ok else ROSE),
                fill=(None if ok else RFILL), sheet=ws2)
            put(r2, 7, cnt, size=9, color=MUT, sheet=ws2)
            r2 += 1
        listed_tot = float(data["tranche_totals"][str(tr)])
        put(r2, 1, "Вкупно | Total", size=9.5, bold=True, color=NAVY,
            fill=BAND, sheet=ws2)
        put(r2, 2, None, fill=BAND, sheet=ws2)
        put(r2, 3, None, fill=BAND, sheet=ws2)
        put(r2, 4, round(tot_o, 3), size=9.5, bold=True, color=NAVY,
            fill=BAND, sheet=ws2)
        put(r2, 5, round(listed_tot, 3), size=9.5, bold=True, color=NAVY,
            fill=BAND, sheet=ws2)
        put(r2, 6, ("—" if abs(tot_o - listed_tot) <= 0.05
                    else "%+.3f" % (tot_o - listed_tot)), size=9, bold=True,
            color=(NAVY if abs(tot_o - listed_tot) <= 0.05 else ROSE),
            fill=BAND, sheet=ws2)
        put(r2, 7, sum(1 for (t_, _i, _c, _o, _v) in rows_out
                       if tr == "all" or t_ == tr), size=9, color=NAVY,
            fill=BAND, sheet=ws2)
        r2 += 2

    ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=7)
    put(r2, 1, data["source"]["note"], size=8.5, italic=True, color=MUT,
        align="left", wrap=True, border=False, sheet=ws2)
    ws2.row_dimensions[r2].height = 28

    for i, w in enumerate([11, 14, 14, 20, 20, 14, 10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    print("EXPORT CLASSES")
    print("  batches on the export list :", n)
    print("  joined to register/Atlas   :", stats["matched"],
          "· unmatched:", stats["unmatched"])
    print("  list THC != our result     :", stats["mismatch"])
    print("  would move class on ours   :", stats["moved"])
    return stats
