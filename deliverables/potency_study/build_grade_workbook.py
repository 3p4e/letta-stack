#!/usr/bin/env python3
"""Render grade_design_even.json into ImB_Potency_Grade_Ranges.xlsx."""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

D = json.load(open('grade_design_even.json'))

INK = '1C2426'; HEAD = '175E63'; HEADFILL = PatternFill('solid', fgColor='175E63')
SUB = PatternFill('solid', fgColor='EEF2F1'); WARN = PatternFill('solid', fgColor='FDEBD0')
FULLF = PatternFill('solid', fgColor='E3EFE7')
thin = Side(style='thin', color='C9D2D0')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
BOLD = Font(bold=True); WHITE = Font(bold=True, color='FFFFFF')
CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
LFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

wb = openpyxl.Workbook()

# ---------------------------------------------------------------- Sheet 1
ws = wb.active; ws.title = 'Grade Boards'
ws.append(['ImB POTENCY GRADE RANGES — EVEN WHOLE-NUMBER NOMINALS '
           '(proposal 27.08.2026, rev. 2 — T1+T2 re-analysis anchors of 26.08.2026)'])
ws['A1'].font = Font(bold=True, size=13)
ws.append(['Rules: whole even nominal preferred (odd only where no even configuration works); '
           'tolerance EQUAL above and below the nominal, at most 10% of it; NO empty grades — '
           'result-free spans stay as documented gaps; touching grades split their budget as '
           'equally as the constraints allow; every batch falls in exactly one grade.'])
ws['A2'].alignment = LFT
ws.append([])
hdr = ['Strain', 'Code', 'Grade', 'Product Code', 'Spec. Doc. Code', 'Nominal %', 'Tolerance %',
        'Lower %', 'Upper %', 'Width pp', 'Potency expression', 'Range status',
        'Batches (latest Total THC %)']
ws.append(hdr)
hr = ws.max_row
for c in range(1, len(hdr) + 1):
    cell = ws.cell(hr, c); cell.font = WHITE; cell.fill = HEADFILL; cell.alignment = CTR; cell.border = BORDER
for strain, entry in D['strains'].items():
    first = True
    for g in entry:
        bl = ', '.join(f"{b['batch']} {b['v']:.2f}" for b in g['batches']) \
             or '— reserve grade (no current batch)'
        status = 'FULL ±10%' if g['full'] else 'balanced (shared budget)'
        if g.get('bridge'): status = 'RESERVE (contiguity bridge)'
        if g.get('odd'): status += ' — ODD nominal (fallback rule 5)'
        tol = (f"±{g['plus_tol']:.2f}" if g['symmetric']
               else f"+{g['plus_tol']:.2f} / −{g['minus_tol']:.2f}")
        ws.append([strain if first else '', g['strain_code'] if first else '', g['grade'],
                   g['product_code'].replace(':', ' : '), g['spec_code'], g['nominal'], tol,
                   g['lower'], g['upper'], g['width'], g['expression'], status, bl])
        r = ws.max_row
        for c in range(1, len(hdr) + 1):
            cell = ws.cell(r, c); cell.border = BORDER
            cell.alignment = LFT if c in (1, 11, 13) else CTR
        for c in (6, 8, 9, 10):
            ws.cell(r, c).number_format = '0.00'
        ws.cell(r, 12).fill = FULLF if g['full'] else SUB
        if first:
            ws.cell(r, 1).font = BOLD
        first = False
widths = [22, 6, 7, 20, 18, 10, 14, 9, 9, 9, 40, 24, 66]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws.freeze_panes = 'A5'

# ---------------------------------------------------------------- Sheet 2
ws2 = wb.create_sheet('Batch Assignments')
hdr2 = ['Batch', 'Strain', 'Grade', 'Product Code', 'Total THC %', 'Basis / source',
        'Owner-requested code', 'Note']
ws2.append(hdr2)
for c in range(1, len(hdr2) + 1):
    cell = ws2.cell(1, c); cell.font = WHITE; cell.fill = HEADFILL; cell.alignment = CTR; cell.border = BORDER
for strain, entry in D['strains'].items():
    for g in entry:
        for b in g['batches']:
            note = ''
            if b['basis'].startswith('declared'):
                note = 'declared value — no eCoA yet, grade provisional until tested'
            elif 'T2' in b['basis']:
                note = 'T2 re-analysis value 26.08.2026 (unofficial) — formal eCoA pending'
            if not b['owner_req'] and 'declared' not in note:
                note = ('code assigned by value (batch not on the tranche list)'
                        + ('; ' + note if note else ''))
            dev = next((f for f in D['flags']
                        if f.startswith(f"{strain} / {b['batch']}:")
                        and ('infeasible' in f or 'MANDATORY-CODE DEVIATION' in f)), None)
            if dev:
                note = dev.split(': ', 1)[1] + ' — infeasible, deviation flagged'
            ws2.append([b['batch'], strain, g['grade'], g['product_code'].replace(':', ' : '),
                        b['v'], b['src'], 'yes' if b['owner_req'] else '—', note])
            r = ws2.max_row
            for c in range(1, len(hdr2) + 1):
                cell = ws2.cell(r, c); cell.border = BORDER
                cell.alignment = CTR if c in (3, 5, 7) else LFT
            ws2.cell(r, 5).number_format = '0.00'
            if note and 'infeasible' in note:
                for c in range(1, len(hdr2) + 1):
                    ws2.cell(r, c).fill = WARN
for i, w in enumerate([16, 20, 7, 20, 11, 46, 12, 52], 1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws2.freeze_panes = 'A2'

# ---------------------------------------------------------------- Sheet 3
ws3 = wb.create_sheet('Notes & Exceptions')
rows = [
 ('DESIGN RULES', ''),
 ('1', 'MANDATORY product codes (management, 27.08.2026 — revised with the 26.08.2026 '
       'T1+T2 re-analysis values): the 48 tranche-list batches carry exactly the THCnn : CBD1 '
       'codes management assigned. The solver treats them as fixed constraints; a deviation '
       'is permitted only where the assigned code is mathematically impossible under rules '
       '2-4, is minimal (fewest batches moved), and is flagged. Result: 42/48 honored; '
       '6 unavoidable deviations: FB012601/1 17.99 > 17.60 ceiling of THC16 -> THC18; '
       'GRC102501/2 9.80 < 10.80 floor of THC12 -> THC10; JD012603/02 14.43 far below the '
       '18.00 floor of THC20 -> THC14; PM112501 10.79 misses the 10.80 floor of THC12 by '
       '0.01 -> THC10; GP092501 25.24 and GP082501/1 25.13 coded THC26 cannot coexist with '
       'the mandatory THC24 holding GP0824_02 22.61 (no-overlap budget) -> THC24. Uncoded '
       'batches take any feasible even nominal — the nominal is NOT chased after the '
       'analysis value; odd only where no even configuration exists.'),
 ('2', 'SYMMETRIC tolerance: every grade is nominal ± t with the SAME t above and below '
       '(window centred on the nominal), t at most 10% of the nominal, bounds two decimals, '
       'minimum meaningful tolerance 0.50.'),
 ('3', 'NO EMPTY GRADES (management rule, 27.08.2026): every grade holds at least one tested '
       'batch. Where real grades cannot reach each other under the 10% cap, the result-free '
       'span between them stays as a DOCUMENTED GAP — never as an empty reserve grade.'),
 ('4', 'BALANCED distribution (management rule, 27.08.2026): where two grades touch, the '
       'shared budget (nominal difference − 0.01) is split as EQUALLY as the batch-containment '
       'and 10% constraints allow — no grade is squeezed for the benefit of its neighbour. '
       'Grades touch wherever the mathematics permits; a grade facing a gap takes its maximum '
       'coverage.'),
 ('5', 'Every batch falls in exactly one grade at its CoQ-forming value: the 26.08.2026 '
       'T1+T2 re-analysis (Farmahem; T2 values unofficial until the formal eCoAs are '
       'filed), else the latest certificate, else the declared value. Retest = CoQ-forming '
       '(rule R5); superseded pre-retest results are out of specification scope.'),
 ('6', 'Gaps carry no results by construction (each gap lies between the windows of grades '
       'that jointly contain every result of the strain); typical at the very lowest grades '
       '(GRC 7.71-8.99; OPM 8.81-8.99) and in result-free mid-spans (CJ 22.01-22.99 & '
       '17.61-17.99; CC 15.41-16.19; FB 13.21-14.85; GP 22.01-22.60 & 17.61-17.99; JD '
       '17.10-17.99; OPM 15.41-16.19 & 11.01-12.59). A future result landing in a gap '
       'triggers a grade-set revision, not an ad-hoc stretch.'),
 ('', ''),
 ('FLAGS', ''),
]
for f in D['flags']: rows.append(('•', f))
if not D['flags']: rows.append(('•', 'none'))
rows.append(('', ''))
rows.append(('EXCEPTIONS', ''))
for e in D['exceptions']: rows.append(('•', e))
if not D['exceptions']: rows.append(('•', 'none — the odd-nominal fallback (rule 5) resolves all former dead-zone cases'))
rows += [
 ('', ''),
 ('OPEN VERIFICATIONS CARRIED OVER', ''),
 ('•', 'The 29 Tranche-2 batches are graded on the 26.08.2026 Farmahem re-analysis values, '
       'which are UNOFFICIAL until the formal eCoAs are filed; those grades are provisional '
       'and the design re-runs automatically once the certificates arrive.'),
 ('•', 'Resolved by T2 (26.08.2026): the BSS052501 20.39-vs-20.47 discrepancy (now 20.52, '
       'THC20 holds), the OMP1024_01 ППК25117 printing anomaly (now 18.99 -> THC18) and the '
       'PM112501 ППК26030 13.33-vs-13.00 question (now 10.79 -> THC10) are all superseded.'),
 ('•', 'GG1024_02: 15.95 correction (cert ППК25140 prints 15.59) — VERIFY ON PAPER ORIGINAL; '
       'grade THC16 holds under either value.'),
 ('•', 'GP062501: PP cert prints 24.89 vs owner sheet 22.89 — grade THC24 holds under either value.'),
 ('•', 'J31122501 graded on the machine-trimmed CoQ-forming preparation (21.84, CoQ-PP-2026-0054); '
       'the hand-trimmed 19.84 result is experimental only.'),
 ('•', 'Truth-check finding (27.08.2026): certificate ППК26065 (13.93, CNP, 11.05.2026) prints '
       'серија JD112501* — the milled Jelly Donutz presentation, a separate register batch — and was '
       'misattributed to JD112501 (whole flower, retest 20.32) in the source dataset. Reattributed; '
       'JD112501* carries its own grade JD_THC14:CBD1.'),
 ('•', 'Batches still on owner-declared values (no analysis on file): GRC102501/1 7.05, '
       'SCR012601 18.07, WED102501 22.05 — their grades are provisional.'),
]
for a, b in rows:
    ws3.append([a, b])
    r = ws3.max_row
    ws3.cell(r, 1).font = BOLD if a and a not in '•' else Font()
    ws3.cell(r, 2).alignment = LFT
ws3.column_dimensions['A'].width = 26
ws3.column_dimensions['B'].width = 130

wb.save('ImB_Potency_Grade_Ranges.xlsx')
n_g = sum(len(e) for e in D['strains'].values())
n_b = sum(len(g['batches']) for e in D['strains'].values() for g in e)
print(f"saved: {len(D['strains'])} strains, {n_g} grades, {n_b} batches")
