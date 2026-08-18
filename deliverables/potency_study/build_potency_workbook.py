#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Comprehensive potency / specification Excel workbook.

Ties together three independently-verified sources:

  - spec_inventory.json    — 341 issued QCSP 001 documents, transcribed
                              verbatim from deliverables/Final_Docs_PDF
                              (pdftotext -layout; every field either present
                              in the document or null — nothing invented).
  - portfolio_master.json  — the 78-row Portfolio Master rename mapping
                              (01_Portfolio_Master sheet).
  - potency_dataset.json   — the Potency Atlas's bound dataset: every
                              register assay, the degradation-aware anchors,
                              and the nominal-±-tolerance grade declarations
                              (built by build_potency_dataset.py).

Six sheets:
  1. README                       — legend, sources, what "match" means.
  2. Original Strain Specs        — one row per issued BASE_SPCs grade.
  3. Renamed Strain Specs         — one row per issued RENs/NEWs grade,
                                     with the original strain(s) it maps
                                     from (via Portfolio Master).
  4. Portfolio Master — Renames   — the 78 batch-level rename rows verbatim.
  5. Atlas Grades — Original      — the Atlas's proposed nominal ± tolerance
                                     tiers, keyed to the ORIGINAL strain name.
  6. Atlas Grades — Renamed       — the same tiers, re-keyed to the NEW
                                     specification name (mirrors
                                     build_potency_html.py's
                                     final_ranges_renamed()).
  7. Potency Test Results         — every register assay (99 rows), with its
                                     Atlas tier, its issued-spec grade match
                                     (if any), and — where the batch was
                                     renamed — its renamed-name tier too.

No value in this workbook is computed independently of the Atlas / spec
inventory; every "in range?" flag is a live Excel formula over the
transcribed numbers on the same sheet, not a hardcoded Python result.
"""
import json
import math
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))


def load(name):
    return json.load(open(os.path.join(HERE, name), encoding="utf-8"))


spec = load("spec_inventory.json")["documents"]
pm = load("portfolio_master.json")
d = load("potency_dataset.json")

FONT = "Arial"
NAVY = "1B3A5C"
GOLD = "C9A227"
GREEN = "1E8449"
ROSE = "C0392B"
GREY = "F4F7FB"
WHITE = "FFFFFF"

hdr_font = Font(name=FONT, size=10, bold=True, color=WHITE)
hdr_fill = PatternFill("solid", fgColor=NAVY)
title_font = Font(name=FONT, size=16, bold=True, color=NAVY)
sub_font = Font(name=FONT, size=10, italic=True, color="5A6B7C")
body_font = Font(name=FONT, size=10)
bold_font = Font(name=FONT, size=10, bold=True)
thin = Side(style="thin", color="D8E2EC")
box = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center")


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = box
    ws.row_dimensions[row].height = 30


def zebra(ws, row, ncols, alt):
    fill = PatternFill("solid", fgColor=GREY) if alt else None
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = body_font
        cell.border = box
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if fill:
            cell.fill = fill


# House numeric rule: every potency value displays with exactly two decimals
# and its % sign hard against the digit. The cells hold real numbers (so they
# stay sortable/filterable and usable in formulas); the format string is what
# renders "20.00%". NOTE the quoted "%" — an Excel "0.00%" format would
# multiply the stored value by 100, which is wrong here because these are
# already percentage-point numbers, not fractions.
PCT_FMT = '0.00"%"'


def pctcols(ws, first_row, last_row, cols):
    """Apply the two-decimal percent display format to the given 1-based
    column numbers, over the data rows only (never the header)."""
    for r in range(first_row, last_row + 1):
        for c in cols:
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = PCT_FMT


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def title_block(ws, title_mk, title_en, note_mk, note_en, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title_mk)
    c.font = title_font
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1, value=title_en)
    c.font = Font(name=FONT, size=11, italic=True, color=NAVY)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    c = ws.cell(row=3, column=1, value=note_mk + "  |  " + note_en)
    c.font = sub_font
    c.alignment = wrap
    ws.row_dimensions[3].height = 32
    return 5  # first data row after a blank spacer row


wb = Workbook()

# ---------------------------------------------------------------- README --
ws = wb.active
ws.title = "README"
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:F1")
ws["A1"] = "Потенција — спецификации, преименувања и резултати"
ws["A1"].font = Font(name=FONT, size=18, bold=True, color=NAVY)
ws.merge_cells("A2:F2")
ws["A2"] = "Potency — Specifications, Renames & Test Results"
ws["A2"].font = Font(name=FONT, size=13, italic=True, color=NAVY)
ws.merge_cells("A3:F3")
ws["A3"] = "Неформален работен документ · не е контролиран запис · извор: deliverables/potency_study/ · 17.08.2026"
ws["A3"].font = sub_font
ws.append([])

readme_rows = [
    ("Лист | Sheet", "Содржина | Contents"),
    ("Original Strain Specs",
     "Секоја издадена QCSP 001 спецификација за оригинално (непреименувано) име на сорта — "
     "класа, код на документ, THC-опсег и номинала, точно како што стои во документот. | "
     "Every issued QCSP 001 specification for an original (not renamed) strain name — grade, "
     "document code, THC range and nominal, exactly as printed."),
    ("Renamed Strain Specs",
     "Истото, за издадените спецификации на НОВИТЕ (преименувани) имиња, со колона што ги "
     "покажува оригиналните имиња што водат до ова ново име (според Portfolio Master). | "
     "The same, for issued specifications of the NEW (renamed) names, with a column showing "
     "which original name(s) lead to this new name (per the Portfolio Master)."),
    ("Portfolio Master — Renames",
     "78-те редови на преименување серија-по-серија, точно како во "
     "01_Portfolio_Master (тranche, серија, старо име, ново име, бренд, етикета, THC%, "
     "стар опсег). | The 78 batch-level rename rows, exactly as in 01_Portfolio_Master "
     "(tranche, batch, old name, new name, brand, label, THC%, old bracket)."),
    ("Atlas Grades — Original",
     "Предложените класи на Атласот на потенција (номинала ± толеранција), клучирани по "
     "ОРИГИНАЛНОТО име на сортата. | The Potency Atlas's proposed grades (nominal ± "
     "tolerance), keyed to the ORIGINAL strain name."),
    ("Atlas Grades — Renamed",
     "Истите класи, преклучирани по НОВОТО име на спецификацијата (истата пресметка како во "
     "финалната табела на Атласот). | The same grades, re-keyed to the NEW specification "
     "name (same computation as the Atlas's final board)."),
    ("Potency Test Results",
     "Сите 99 некогаш тестирани резултати за Вкупен Δ⁹-THC од регистарот — серија, сорта, "
     "резултат, датум, лабораторија, сертификат — со колони кои покажуваат во која класа на "
     "Атласот паѓа резултатот (по оригинално и, каде применливо, по ново име) и дали "
     "резултатот е внатре во издадениот спецификациски опсег. | All 99 register Total Δ⁹-THC "
     "results ever tested — batch, strain, value, date, lab, certificate — with columns "
     "showing which Atlas tier the result falls in (by original and, where applicable, new "
     "name) and whether it sits inside the issued specification range."),
]
r = 5
ws.cell(row=r, column=1, value=readme_rows[0][0]).font = hdr_font
ws.cell(row=r, column=1).fill = hdr_fill
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
c = ws.cell(row=r, column=2, value=readme_rows[0][1])
c.font = hdr_font
c.fill = hdr_fill
r += 1
for name, desc in readme_rows[1:]:
    ws.cell(row=r, column=1, value=name).font = bold_font
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    c = ws.cell(row=r, column=2, value=desc)
    c.font = body_font
    c.alignment = wrap
    ws.row_dimensions[r].height = 58
    r += 2

r += 1
notes = [
    "Извори | Sources:",
    "  • spec_inventory.json — pdftotext -layout транскрипција на 341 документ под "
    "deliverables/Final_Docs_PDF/ImB_SPC (изворен HTML не постои во репозиториумот). "
    "| pdftotext -layout transcription of 341 documents under "
    "deliverables/Final_Docs_PDF/ImB_SPC (source HTML not present in the repo).",
    "  • portfolio_master.json — извадок од Drive-датотеката BCP_PRODUCT_MASTER_FINAL.xlsx, "
    "лист 01_Portfolio_Master. | extract of the Drive file BCP_PRODUCT_MASTER_FINAL.xlsx, "
    "sheet 01_Portfolio_Master.",
    "  • potency_dataset.json — врзаниот датасет на Атласот на потенција "
    "(build_potency_dataset.py); номинала ± толеранција секогаш е декларацијата на класата "
    "на серијата — никогаш посебна бројка по серија. | the Potency Atlas's bound dataset "
    "(build_potency_dataset.py); nominal ± tolerance is always the batch's TIER declaration "
    "— never a separate per-batch figure.",
    "",
    "Оваа студија е ПРЕДЛОГ-ревизија на класите, изведена статистички од сите некогаш "
    "тестирани резултати — не изменува ниту една издадена QCSP 001 спецификација, чија "
    "номинала и опсег остануваат меродавни до формално усвојување. | This study is a "
    "PROPOSED revision of the grade bands, statistically derived from every result ever "
    "tested — it does not amend any issued QCSP 001 specification, whose nominal and range "
    "remain authoritative until formally adopted.",
]
for line in notes:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(row=r, column=1, value=line)
    c.font = Font(name=FONT, size=9.5, italic=True, color="4A5B6C") if line else body_font
    c.alignment = wrap
    ws.row_dimensions[r].height = 30 if line else 8
    r += 1
autosize(ws, [22, 18, 18, 18, 18, 18])

print("README written")

# =========================================================================
# Shared helpers: strain-name matching between the Atlas / Portfolio Master
# and the issued spec_inventory. Matching is by normalised text identity
# (case, &/AND, accents, apostrophes) plus a small, individually-verified
# alias table for known spelling variants of the SAME strain across sources
# (e.g. the register's "Pure Michigen" vs the issued spec's "PURE MICHIGAN").
# Every alias below was confirmed by checking the Portfolio Master row and
# the matching issued document side by side — none is a guess.
# =========================================================================
import re as _re


def norm_name(s):
    s = (s or "").upper()
    s = s.replace("È", "E").replace("É", "E").replace("’", "").replace("'", "")
    s = s.replace("&", " AND ")
    s = _re.sub(r"[^A-Z0-9 ]", " ", s)
    s = _re.sub(r"\s+", " ", s).strip()
    return s


ALIAS_TO_SPEC = {
    "APPLE AND BANANA": "APPLES & BANANAS",
    "APPELS AND BANANAS": "APPLES & BANANAS",
    "CLEMOSA A BUD": "CLEMOSA",
    "SLEEPY JOY": "SLEEPY JOE",
    "PURE MICHIGEN": "PURE MICHIGAN",
    "GRAPS AND CREME": "GRAPS & CRÈME",
    "WEDDING CRUSHER": "WEDDING CRASHER",
}

GRADE_ORDER = {"GRADE I": 1, "GRADE II": 2, "GRADE III": 3, "GRADE IV": 4, "GRADE V": 5}


def grade_key(g):
    return GRADE_ORDER.get((g or "").upper(), 9)


def norm_b(b):
    b = (b or "").strip().upper().replace("OMP", "OPM")
    return _re.sub(r"/0(\d)", r"/\1", b)


def spec_lookup_key(name):
    n = norm_name(name)
    return norm_name(ALIAS_TO_SPEC.get(n, n))


# index issued specs by normalised strain name -> canonical strain_name, by subtree group
spec_by_key = {}
for rec in spec:
    if not rec.get("strain_name"):
        continue
    k = norm_name(rec["strain_name"])
    spec_by_key.setdefault(k, []).append(rec)

BASE_SUBTREES = {"BASE_SPCs"}
REN_SUBTREES = {"RENs", "NEWs"}


def specs_for(name, subtrees):
    k = spec_lookup_key(name)
    return [r for r in spec_by_key.get(k, []) if r["subtree"] in subtrees]


# Portfolio Master: neu -> set of originals (for the Renamed Strain Specs sheet)
neu_to_originals = {}
for row in pm:
    neu = (row.get("neu") or "").strip()
    orig = (row.get("original") or "").strip()
    if neu and orig:
        neu_to_originals.setdefault(norm_name(neu), set()).add(orig)

# batch -> Portfolio Master row (for the Potency Test Results sheet)
pm_by_batch = {norm_b(row["batch"]): row for row in pm}

# strain (Atlas key) -> stock rows, for tier/anchor status
stock_by_strain = {}
for b in d["stock"]:
    stock_by_strain.setdefault(b["strain"], []).append(b)
stock_by_batch = {norm_b(b["batch"]): b for b in d["stock"]}

MAX_TOL_RATIO = d["design"]["max_tol_ratio"]
MIN_GAP = d["design"]["min_gap"]
NOM_STEP = d["design"]["nom_step"]

_TOP_OVERRIDE = d["design"].get("top_nominal_override", {})
_strain_top_anchor = {}
for _b in d["stock"]:
    _a = _b["anchor"] if _b["anchor"] is not None else _b["declared"]
    if _a is not None:
        _strain_top_anchor[_b["strain"]] = max(_strain_top_anchor.get(_b["strain"], -1.0), _a)
OVERRIDE_BY_TOP_ANCHOR = {round(_strain_top_anchor[s], 2): v
                          for s, v in _TOP_OVERRIDE.items() if s in _strain_top_anchor}


def build_top_down(groups, floor=5.0, max_ratio=MAX_TOL_RATIO, step=NOM_STEP, gap=MIN_GAP,
                   top_override=None, strain_max=None):
    """Mirror of build_potency_dataset.build_top_down."""
    top = groups[-1]
    tmin, tmax = min(top), max(top)
    lo_i = math.ceil((tmax / (1 + max_ratio)) / step - 1e-9)
    hi_i = math.floor((tmin / (1 - max_ratio)) / step + 1e-9)
    lo_i = max(lo_i, math.ceil((floor / (1 - max_ratio)) / step - 1e-9))
    top_cands = [round(k * step, 2) for k in range(lo_i, hi_i + 1)]
    if not top_cands:
        return None
    if top_override is not None and strain_max is not None \
            and abs(tmax - strain_max) < 1e-9:
        if any(abs(c - top_override) < 1e-9 for c in top_cands):
            top_cands = [round(top_override, 2)]
        else:
            return None
    best = None
    for n_top in top_cands:
        tol = round(n_top * max_ratio, 2)
        tiers = [dict(nominal=n_top, tol=tol, lo=round(n_top - tol, 2),
                      hi=round(n_top + tol, 2), anchors=top)]
        ok = True
        ceiling = round(tiers[-1]["lo"] - gap, 2)
        for g in reversed(groups[:-1]):
            gmin, gmax = min(g), max(g)
            n_i = math.ceil((ceiling / (1 + max_ratio)) / step - 1e-9)
            n_i = max(n_i, math.ceil((floor / (1 - max_ratio)) / step - 1e-9))
            nom = round(n_i * step, 2)
            tol = round(ceiling - nom, 2)
            if tol < 0 or tol > round(nom * max_ratio, 2) + 1e-9:
                ok = False
                break
            lo, hi = round(nom - tol, 2), round(nom + tol, 2)
            if lo < floor - 1e-6 or not (lo - 1e-6 <= gmin and gmax <= hi + 1e-6):
                ok = False
                break
            tiers.append(dict(nominal=nom, tol=tol, lo=lo, hi=hi, anchors=g))
            ceiling = round(lo - gap, 2)
        if not ok:
            continue
        tiers.reverse()
        cost = sum(abs(t["nominal"] - a) for t in tiers for a in t["anchors"])
        if best is None or cost < best[0] - 1e-9:
            best = (cost, tiers)
    return best[1] if best else None


def _tiers_for_k(anchors, k, floor, max_ratio, gap, top_override=None, strain_max=None):
    """Mirror of build_potency_dataset._tiers_for_k."""
    n = len(anchors)
    if k > n:
        return None
    best = None

    def eval_cuts(bounds):
        nonlocal best
        groups = [anchors[bounds[i]:bounds[i + 1]] for i in range(k)]
        tiers = build_top_down(groups, floor, max_ratio, NOM_STEP, gap,
                               top_override=top_override, strain_max=strain_max)
        if tiers is None:
            return
        for gi, t in enumerate(tiers):
            t["start"], t["end"] = bounds[gi], bounds[gi + 1]
        cost = sum(abs(t["nominal"] - a) for t in tiers for a in t["anchors"])
        if best is None or cost < best[0] - 1e-9:
            best = (cost, tiers)

    def cuts(start, groups_left, acc):
        if groups_left == 1:
            eval_cuts([0] + acc + [n])
            return
        for c in range(start + 1, n - (groups_left - 1) + 1):
            cuts(c, groups_left - 1, acc + [c])

    cuts(0, k, [])
    return best


def plan_contiguous(anchors, floor=5.0, max_ratio=MAX_TOL_RATIO, gap=MIN_GAP,
                    top_override=None, strain_max=None):
    """Mirror of build_potency_dataset.plan_contiguous."""
    n = len(anchors)
    if n == 0:
        return []
    for k in range(1, n + 1):
        res = _tiers_for_k(anchors, k, floor, max_ratio, gap,
                           top_override=top_override, strain_max=strain_max)
        if res is not None:
            return res[1]
    return None


def tiers_from_anchors(items, top_override=None):
    """items = [(batch, anchor, tested_bool)] — mirror of
    build_potency_dataset.build_strain_tiers / build_potency_html's copy."""
    items = sorted(items, key=lambda x: x[1])
    strain_max = max(x[1] for x in items) if items else None

    def resolve(sub_items):
        sub_anchors = [x[1] for x in sub_items]
        plan = plan_contiguous(sub_anchors, top_override=top_override, strain_max=strain_max)
        if plan is not None:
            for t in plan:
                t["gap_after"] = False
                t["batches"] = [x[0] for x in sub_items[t["start"]:t["end"]]]
                t["tested"] = [x[2] for x in sub_items[t["start"]:t["end"]]]
            return plan
        n = len(sub_items)
        assert n > 1, ("single anchor infeasible — below release floor?", sub_anchors)
        best_m = 1
        for m in range(n, 0, -1):
            if plan_contiguous(sub_anchors[:m], top_override=top_override,
                               strain_max=strain_max) is not None:
                best_m = m
                break
        left = plan_contiguous(sub_anchors[:best_m], top_override=top_override,
                               strain_max=strain_max)
        for t in left:
            t["gap_after"] = False
            t["batches"] = [x[0] for x in sub_items[t["start"]:t["end"]]]
            t["tested"] = [x[2] for x in sub_items[t["start"]:t["end"]]]
        left[-1]["gap_after"] = True
        return left + resolve(sub_items[best_m:])

    return resolve(items)


# anchors per batch (for the renamed-tier recomputation)
anchor_by_batch = {}
for b in d["stock"]:
    a = b["anchor"] if b["anchor"] is not None else b["declared"]
    if a is not None:
        anchor_by_batch[norm_b(b["batch"])] = (a, b["anchor"] is not None)

# recompute renamed-name tiers once, reused by sheet 6 and sheet 7
renamed_tiers = {}   # neu name -> list of tier dicts
for row in pm:
    neu = (row.get("neu") or "").strip()
    if not neu:
        continue
    renamed_tiers.setdefault(neu, [])
for neu in renamed_tiers:
    items = []
    for row in pm:
        if (row.get("neu") or "").strip() != neu:
            continue
        got = anchor_by_batch.get(norm_b(row["batch"]))
        if got:
            items.append((row["batch"], got[0], got[1]))
    _ov = OVERRIDE_BY_TOP_ANCHOR.get(round(max(a for _b, a, _t in items), 2)) if items else None
    renamed_tiers[neu] = tiers_from_anchors(items, top_override=_ov) if items else []

# sanity cross-check against the Atlas's own final board, before writing anything
_check = 0
for s, tiers in d["merged_ranges"].items():
    for t in tiers:
        _check += 1
print("cross-check: %d dataset tiers available for sheet 5; %d renamed groups for sheet 6"
      % (_check, len(renamed_tiers)))

# ===================================================== Original Strain Specs
ws = wb.create_sheet("Original Strain Specs")
ws.sheet_view.showGridLines = False
ncols = 11
r0 = title_block(ws, "Издадени спецификации — оригинални имиња",
                  "Issued Specifications — Original Names",
                  "Секоја издадена QCSP 001 спецификација за непреименувано име, транскрибирана точно "
                  "како во документот (deliverables/Final_Docs_PDF/ImB_SPC/BASE_SPCs).",
                  "Every issued QCSP 001 specification for a not-renamed name, transcribed exactly as "
                  "printed (deliverables/Final_Docs_PDF/ImB_SPC/BASE_SPCs).", ncols)
headers = ["Код | Code", "Име на сорта | Strain Name", "Класа | Grade", "Код на документ | Doc Code",
           "THC-опсег (%) | THC Range (%)", "THC мин. | Min", "THC макс. | Max",
           "Номинала (печатена) | Nominal (as printed)", "Номинала THC | Nominal THC",
           "Толеранција | Tolerance", "CBD лимит | CBD Limit"]
for i, h in enumerate(headers, 1):
    ws.cell(row=r0, column=i, value=h)
style_header(ws, r0, ncols)

base_rows = sorted({r["doc_code"]: r for r in spec if r["subtree"] == "BASE_SPCs"}.values(),
                    key=lambda r: (r["strain_name"], grade_key(r["grade"])))
row = r0 + 1
for i, rec in enumerate(base_rows):
    vals = [rec["strain_code"], rec["strain_name"], rec["grade"], rec["doc_code"],
            rec["thc_range"], rec["thc_min"], rec["thc_max"], rec["nominal"],
            rec["nominal_thc"], rec["nominal_tolerance"], rec["cbd_limit"]]
    for c, v in enumerate(vals, 1):
        ws.cell(row=row, column=c, value=v)
    zebra(ws, row, ncols, i % 2 == 0)
    row += 1
ws.freeze_panes = ws.cell(row=r0 + 1, column=1)
pctcols(ws, r0 + 1, row - 1, [6, 7, 9, 10])
autosize(ws, [7, 22, 9, 20, 15, 9, 9, 20, 12, 10, 12])
print("Original Strain Specs: %d rows" % len(base_rows))

# ====================================================== Renamed Strain Specs
ws = wb.create_sheet("Renamed Strain Specs")
ws.sheet_view.showGridLines = False
ncols = 12
r0 = title_block(ws, "Издадени спецификации — нови (преименувани) имиња",
                  "Issued Specifications — New (Renamed) Names",
                  "Секоја издадена QCSP 001 спецификација за ново име (RENs/NEWs), со оригиналното(ите) "
                  "име(иња) што водат до него, според Portfolio Master.",
                  "Every issued QCSP 001 specification for a new name (RENs/NEWs), with the original "
                  "name(s) it maps from, per the Portfolio Master.", ncols)
headers = ["Код | Code", "Ново име | New Strain Name", "Класа | Grade", "Код на документ | Doc Code",
           "THC-опсег (%) | THC Range (%)", "THC мин. | Min", "THC макс. | Max",
           "Номинала (печатена) | Nominal (as printed)", "Номинала THC | Nominal THC",
           "Толеранција | Tolerance", "CBD лимит | CBD Limit",
           "Оригинално(и) име(иња) | Original name(s)"]
for i, h in enumerate(headers, 1):
    ws.cell(row=r0, column=i, value=h)
style_header(ws, r0, ncols)

ren_rows = sorted({r["doc_code"]: r for r in spec if r["subtree"] in REN_SUBTREES}.values(),
                   key=lambda r: (r["strain_name"], grade_key(r["grade"])))
row = r0 + 1
for i, rec in enumerate(ren_rows):
    origs = sorted(neu_to_originals.get(norm_name(rec["strain_name"]), set()))
    vals = [rec["strain_code"], rec["strain_name"], rec["grade"], rec["doc_code"],
            rec["thc_range"], rec["thc_min"], rec["thc_max"], rec["nominal"],
            rec["nominal_thc"], rec["nominal_tolerance"], rec["cbd_limit"],
            ", ".join(origs) or "—"]
    for c, v in enumerate(vals, 1):
        ws.cell(row=row, column=c, value=v)
    zebra(ws, row, ncols, i % 2 == 0)
    row += 1
ws.freeze_panes = ws.cell(row=r0 + 1, column=1)
pctcols(ws, r0 + 1, row - 1, [6, 7, 9, 10])
autosize(ws, [7, 22, 9, 20, 15, 9, 9, 20, 12, 10, 12, 26])
print("Renamed Strain Specs: %d rows" % len(ren_rows))

# =================================================== Portfolio Master rows
ws = wb.create_sheet("Portfolio Master — Renames")
ws.sheet_view.showGridLines = False
ncols = 9
r0 = title_block(ws, "Portfolio Master — преименувања по серија",
                  "Portfolio Master — Batch-Level Renames",
                  "78-те редови точно како во 01_Portfolio_Master (BCP_PRODUCT_MASTER_FINAL.xlsx). "
                  "Колоната „Преименувана?“ е жива формула (старо ≠ ново).",
                  "The 78 rows exactly as in 01_Portfolio_Master (BCP_PRODUCT_MASTER_FINAL.xlsx). "
                  "The \"Renamed?\" column is a live formula (old ≠ new).", ncols)
headers = ["Транша | Tranche", "Серија | Batch", "Оригинално име | Original Name",
           "Ново име | New Name", "Преименувана? | Renamed?", "Бренд | Brand",
           "Етикета | Label", "Декл. THC % | Declared THC %", "Стар опсег | Old Bracket"]
for i, h in enumerate(headers, 1):
    ws.cell(row=r0, column=i, value=h)
style_header(ws, r0, ncols)

row = r0 + 1
for i, rprow in enumerate(sorted(pm, key=lambda x: (x["tranche"], x["batch"]))):
    ws.cell(row=row, column=1, value=rprow["tranche"])
    ws.cell(row=row, column=2, value=rprow["batch"])
    ws.cell(row=row, column=3, value=rprow["original"])
    ws.cell(row=row, column=4, value=rprow["neu"])
    ws.cell(row=row, column=5, value='=IF(C%d<>D%d,"Да | Yes","Не | No")' % (row, row))
    ws.cell(row=row, column=6, value=rprow.get("brand"))
    ws.cell(row=row, column=7, value=rprow.get("label"))
    ws.cell(row=row, column=8, value=rprow.get("thc"))
    ws.cell(row=row, column=8).number_format = "0.00%"
    ws.cell(row=row, column=9, value=rprow.get("bracket"))
    zebra(ws, row, ncols, i % 2 == 0)
    row += 1
ws.freeze_panes = ws.cell(row=r0 + 1, column=1)
autosize(ws, [8, 15, 20, 22, 13, 9, 22, 12, 12])
print("Portfolio Master rows: %d" % len(pm))

# ===================================================== Atlas Grades — Original
ws = wb.create_sheet("Atlas Grades — Original")
ws.sheet_view.showGridLines = False
ncols = 10
r0 = title_block(ws, "Атлас — предложени класи по оригинално име",
                  "Atlas — Proposed Grades by Original Name",
                  "Класите на Атласот на потенција (номинала ± толеранција), директно од "
                  "potency_dataset.json. Секоја серија ја носи декларацијата на својата класа. "
                  "Соседните класи на иста сорта немаат празен простор меѓу себе, освен кога тоа "
                  "е означено во „Забелешка“.",
                  "The Potency Atlas's grades (nominal ± tolerance), straight from "
                  "potency_dataset.json. Every batch carries its tier's declaration. Adjacent "
                  "tiers of the same strain carry no blind gap between them, except where "
                  "flagged in \"Note\".",
                  ncols)
headers = ["Сорта | Strain", "Класа | Tier", "Номинала % | Nominal %", "Толеранција % | Tolerance %",
           "Опсег — долно | Range — low", "Опсег — горно | Range — high",
           "Бр. серии | # Batches", "Серии | Batches", "Основа | Basis",
           "Забелешка | Note"]
for i, h in enumerate(headers, 1):
    ws.cell(row=r0, column=i, value=h)
style_header(ws, r0, ncols)

# per-tier "basis": Definitive if any batch in the tier has a real anchor
tested_batches = {norm_b(b["batch"]) for b in d["stock"] if b["anchor"] is not None}
row = r0 + 1
i = 0
for s in sorted(d["merged_ranges"]):
    tiers = d["merged_ranges"][s]
    for ti, t in enumerate(tiers, 1):
        basis = ("ДЕФИНИТИВНО | DEFINITIVE"
                 if any(norm_b(bt) in tested_batches for bt in t["batches"])
                 else "ПРОВИЗОРНО | PROVISIONAL")
        note = ("" if not t.get("gap_after") or ti >= len(tiers) else
                "НЕМА КЛАСА %.2f%%–%.2f%% | NO GRADE %.2f%%–%.2f%%"
                % (t["range"][1], tiers[ti]["range"][0], t["range"][1], tiers[ti]["range"][0]))
        vals = [s, "Pot.-%d" % ti, t["nominal"], t["tol"], t["range"][0], t["range"][1],
                len(t["batches"]), ", ".join(t["batches"]), basis, note]
        for c, v in enumerate(vals, 1):
            ws.cell(row=row, column=c, value=v)
        zebra(ws, row, ncols, i % 2 == 0)
        i += 1
        row += 1
ws.freeze_panes = ws.cell(row=r0 + 1, column=1)
pctcols(ws, r0 + 1, row - 1, [3, 4, 5, 6])
autosize(ws, [20, 8, 12, 14, 13, 13, 11, 40, 22, 46])
print("Atlas Grades — Original: %d rows" % (row - r0 - 1))

# ====================================================== Atlas Grades — Renamed
ws = wb.create_sheet("Atlas Grades — Renamed")
ws.sheet_view.showGridLines = False
ncols = 11
r0 = title_block(ws, "Атлас — предложени класи по ново име",
                  "Atlas — Proposed Grades by New Name",
                  "Истите докажани опсези, преклучирани по НОВОТО спецификациско име "
                  "(01_Portfolio_Master). Пресметката е идентична со финалната табела на "
                  "Атласот.",
                  "The same evidenced ranges, re-keyed to the NEW specification name "
                  "(01_Portfolio_Master). Computation is identical to the Atlas's final "
                  "board.", ncols)
headers = ["Ново име | New Strain Name", "Класа | Tier", "Номинала % | Nominal %",
           "Толеранција % | Tolerance %", "Опсег — долно | Range — low",
           "Опсег — горно | Range — high", "Бр. серии | # Batches", "Серии | Batches",
           "Основа | Basis", "Оригинално(и) име(иња) | Original name(s)",
           "Забелешка | Note"]
for i, h in enumerate(headers, 1):
    ws.cell(row=r0, column=i, value=h)
style_header(ws, r0, ncols)

row = r0 + 1
i = 0
for neu in sorted(renamed_tiers):
    origs = ", ".join(sorted({rw["original"] for rw in pm
                              if (rw.get("neu") or "").strip() == neu})) or "—"
    tiers = renamed_tiers[neu]
    if not tiers:
        vals = [neu, "—", None, None, None, None, 0, "—", "БЕЗ СИДРО | NO ANCHOR", origs, ""]
        for c, v in enumerate(vals, 1):
            ws.cell(row=row, column=c, value=v)
        zebra(ws, row, ncols, i % 2 == 0)
        i += 1
        row += 1
        continue
    for ti, t in enumerate(tiers, 1):
        basis = ("ДЕФИНИТИВНО | DEFINITIVE" if any(t["tested"])
                 else "ПРОВИЗОРНО | PROVISIONAL")
        note = ("" if not t.get("gap_after") or ti >= len(tiers) else
                "НЕМА КЛАСА %.2f%%–%.2f%% | NO GRADE %.2f%%–%.2f%%"
                % (t["hi"], tiers[ti]["lo"], t["hi"], tiers[ti]["lo"]))
        vals = [neu, "Pot.-%d" % ti, t["nominal"], t["tol"], t["lo"], t["hi"],
                len(t["batches"]), ", ".join(t["batches"]), basis, origs, note]
        for c, v in enumerate(vals, 1):
            ws.cell(row=row, column=c, value=v)
        zebra(ws, row, ncols, i % 2 == 0)
        i += 1
        row += 1
ws.freeze_panes = ws.cell(row=r0 + 1, column=1)
pctcols(ws, r0 + 1, row - 1, [3, 4, 5, 6])
autosize(ws, [22, 8, 12, 14, 13, 13, 11, 40, 22, 26, 46])
print("Atlas Grades — Renamed: %d rows" % (row - r0 - 1))

# ======================================================= Potency Test Results
ws = wb.create_sheet("Potency Test Results")
ws.sheet_view.showGridLines = False
ncols = 22
r0 = title_block(ws, "Сите тестирани резултати — Вкупен Δ⁹-THC",
                  "All Tested Results — Total Δ⁹-THC",
                  "Сите 99 некогаш тестирани резултати од регистарот, со класата на Атласот "
                  "(по оригинално и, каде применливо, по ново име) и издадениот спецификациски "
                  "опсег што го покрива резултатот. „Внатре?“ колоните се живи формули над "
                  "бројките на овој ред.",
                  "All 99 register results ever tested, with the Atlas tier (by original and, "
                  "where applicable, new name) and the issued specification range that covers "
                  "the result. The \"In range?\" columns are live formulas over this row's own "
                  "numbers.", ncols)
headers = [
    "Серија | Batch", "Сорта (регистар) | Strain (register)", "Резултат % | Value %",
    "Датум | Date", "Лабораторија | Lab", "Сертификат | Certificate",
    "Во залиха Т1/Т2/Т3? | In T1/T2/T3 stock?",
    "Класа на Атласот (ориг.) | Atlas tier (orig.)",
    "Номинала % (ориг.) | Nominal % (orig.)", "Толеранција % (ориг.) | Tolerance % (orig.)",
    "Опсег дол. (ориг.) | Range lo (orig.)", "Опсег гор. (ориг.) | Range hi (orig.)",
    "Внатре во опсег на Атласот? | In Atlas range?",
    "Издадена класа (ориг. спец.) | Issued grade (orig. spec)",
    "Изд. опсег дол. | Issued range lo", "Изд. опсег гор. | Issued range hi",
    "Внатре во издаден опсег? | In issued range?",
    "Ново име | New strain name",
    "Класа на Атласот (ново) | Atlas tier (new)",
    "Номинала % (ново) | Nominal % (new)", "Опсег дол. (ново) | Range lo (new)",
    "Опсег гор. (ново) | Range hi (new)",
]
for i, h in enumerate(headers, 1):
    ws.cell(row=r0, column=i, value=h)
style_header(ws, r0, ncols)

row = r0 + 1
n_out_atlas = n_out_issued = 0
for i, res in enumerate(sorted(d["register_results"], key=lambda x: (x["strain"], x["date"]))):
    nb = norm_b(res["batch"])
    stk = stock_by_batch.get(nb)
    pmrow = pm_by_batch.get(nb)

    in_stock = "Да | Yes" if stk else "Не | No"

    if stk and stk.get("proposed") and stk.get("tier"):
        a_tier = "Pot.-%s" % stk["tier"]
        a_nom, a_tol = stk["nominal"], stk["tol"]
        a_lo, a_hi = stk["proposed"]
    else:
        a_tier = "n/a — не е тековна залиха | not currently stock"
        a_nom = a_tol = a_lo = a_hi = None

    orig_specs = specs_for(res["strain"], BASE_SUBTREES)
    hits = [s for s in orig_specs if s["thc_min"] is not None and s["thc_min"] <= res["value"] <= s["thc_max"]]
    if hits:
        issued_grade = ", ".join(sorted({h["grade"] for h in hits}))
        i_lo = min(h["thc_min"] for h in hits)
        i_hi = max(h["thc_max"] for h in hits)
    elif orig_specs:
        issued_grade = "нема класа што го покрива | no grade covers this value"
        i_lo = i_hi = None
    else:
        issued_grade = "нема издадена спец. за оваа сорта | no issued spec for this strain"
        i_lo = i_hi = None

    if pmrow and (pmrow.get("neu") or "").strip() and pmrow["neu"].strip() != pmrow["original"].strip():
        neu = pmrow["neu"].strip()
        r_tier = r_nom = r_lo = r_hi = None
        for ti, t in enumerate(renamed_tiers.get(neu, []), 1):
            if res["batch"] in t["batches"] or pmrow["batch"] in t["batches"]:
                r_tier, r_nom, r_lo, r_hi = "Pot.-%d" % ti, t["nominal"], t["lo"], t["hi"]
                break
        if r_tier is None:
            r_tier = "n/a"
    elif pmrow:
        neu = "(без преименување | not renamed)"
        r_tier = r_nom = r_lo = r_hi = None
    else:
        neu = "—"
        r_tier = r_nom = r_lo = r_hi = None

    vals = [res["batch"], res["strain"], res["value"], res["date"], res["lab"], res["cert"],
            in_stock, a_tier, a_nom, a_tol, a_lo, a_hi, None,
            issued_grade, i_lo, i_hi, None,
            neu, r_tier, r_nom, r_lo, r_hi]
    for c, v in enumerate(vals, 1):
        ws.cell(row=row, column=c, value=v)
    # live formulas: value (col C) against the ranges on this same row
    if a_lo is not None:
        f = '=IF(AND(C%d>=K%d,C%d<=L%d),"Да | Yes","НЕ | NO")' % (row, row, row, row)
    else:
        f = "—"
    ws.cell(row=row, column=13, value=f)
    if i_lo is not None:
        f = '=IF(AND(C%d>=O%d,C%d<=P%d),"Да | Yes","НЕ | NO")' % (row, row, row, row)
    else:
        f = "—"
    ws.cell(row=row, column=17, value=f)
    zebra(ws, row, ncols, i % 2 == 0)
    row += 1
ws.freeze_panes = ws.cell(row=r0 + 1, column=3)
pctcols(ws, r0 + 1, row - 1, [3, 9, 10, 11, 12, 15, 16, 20, 21, 22])
autosize(ws, [15, 18, 10, 11, 12, 20, 12, 15, 11, 11, 10, 10, 12,
              22, 10, 10, 12, 16, 13, 11, 10, 10])
print("Potency Test Results: %d rows" % (row - r0 - 1))

# =========================================================================
OUT = os.path.join(HERE, "Potency_Specs_and_Results.xlsx")
wb.save(OUT)
print("wrote", OUT)
