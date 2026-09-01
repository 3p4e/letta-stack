# -*- coding: utf-8 -*-
"""
MASTER potency table + manufacturing (harvest) and packaging dates.
Where a date is given as a range or a pair, the FIRST date is taken.
"""
import datetime as dt, openpyxl
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
NAVY="2B547E";LGREY="F7FAFC";WHITE="FFFFFF";AMBER="FFF2CC";GREEN="E2EFDA";ROSE="FCE4D6"
OUT="PP_Potency_MASTER_T1_T2_dates.xlsx";CAP=0.10;U=0.05;EPS=1e-9
TODAY=dt.date(2026,8,31)
def band(n): t=n*CAP; return t,n-t,n+t-0.01
LAD=list(range(28,7,-2));ROMAN=["I","II","III","IV","V","VI","VII","VIII","IX","X","XI"]
GRADE={n:ROMAN[i] for i,n in enumerate(LAD)}

# PP batch -> (harvest first date, packaging first date, source text harvest, source text pkg)
# '*' marks a year inferred from the harvest date because the source omitted it
D={
"BG1024":((2025,1,21),(2025,3,7),"21.01.2025","07.03.2025"),
"BSS1024":((2025,1,22),(2025,3,4),"22.01.2025","04.03.2025"),
"GG1024":((2025,1,20),(2025,3,11),"20.01.2025","11.03.2025"),
"OPM1024":((2025,1,21),(2025,4,24),"21.01.2025","24/25.04.2025"),
"HPA1024":((2025,1,22),(2025,4,11),"22.01.2025","11.04.2025"),
"P050092":((2025,3,17),(2025,7,31),"17/18.03.2025","31.07-05.08*"),
"P050042":((2025,3,26),(2025,6,20),"26/27.03.2025","20.06.2025"),
"P050022":((2025,4,22),(2025,6,3),"22/23.04.2025","03-11.06.2025"),
"P050052":((2025,5,27),(2025,6,27),"27/28.05.2025","27.06-04.07.2025"),
"P050062":((2025,5,29),(2025,7,4),"29/30.05.2025","04.07-10.07.2025"),
"P050072":((2025,6,9),(2025,7,14),"09/10.06.2025","14-17.07*"),
"P050082":((2025,7,1),(2025,7,22),"01/02.07.2025","22-30.07.2025"),
"P050112":((2025,7,8),(2025,8,25),"08/09.07.2025","25-31.08*"),
"P050152":(None,(2025,9,25),"— (source malformed)","25.09.2025"),
"P050162":((2025,8,20),(2025,9,30),"20.08.2025","30.09-01.10.2025"),
"P050182":((2025,8,21),(2025,10,3),"21-23.08.2025","03-09.10.2025"),
"P050192":((2025,9,1),(2025,10,10),"01-03.09.2025","10-21.10.2025"),
"P050222":((2025,9,23),(2025,11,24),"23.09.2025","24.11-25.11.2025"),
"P050282":((2025,10,28),(2025,12,3),"28.10.2025","03.12-04.12.2025"),
"P050302":((2025,11,13),(2025,12,22),"13.11.2025","22.12.2025"),
"P050312":((2025,11,19),(2025,12,23),"19-20.11.2025","23-26.12.2025"),
"P050322":((2025,11,20),(2025,12,26),"20.11.2025","26.12.2025"),
"P060012":((2025,12,9),(2026,1,13),"09.12.2025","13-14.01.2026"),
"P060022":((2025,12,9),(2026,1,16),"9-10.12.2025","16.01.2026"),
"P060032":((2025,12,10),(2026,1,29),"10.12.2025","29.01.2026"),
"P060042":((2025,12,12),(2026,1,29),"12.12.2025","29.01.2026"),
"P060062":((2025,12,11),(2026,1,30),"11.12.2025","30.01.2026"),
"P060072":((2025,12,12),(2026,2,2),"12.12.2025","02.02.2026"),
"P060082":((2025,12,11),(2026,2,2),"11-12.12.2025","02.02.2026"),
"P060092":((2025,12,16),(2026,2,3),"16-17.12.2025","03-09.02.2026"),
"P060112":((2026,1,8),(2026,2,17),"08-09.01.2026","17.02.2026"),
"P060122":((2026,1,9),(2026,2,18),"09.01.2026","18-19.02.2026"),
"P060132":((2026,1,9),(2026,2,19),"09.01.2026","19-20.02.2026"),
"P060152":((2026,1,21),(2026,3,5),"21.01.2026","05.03.2026"),
"P060172":((2026,1,20),(2026,3,12),"20-21.01.2026","12.03.2026"),
"P060182":((2026,1,20),(2026,3,14),"20.01.2026","14.03.2026"),
"P060212":((2026,3,25),(2026,4,21),"25.03.2026","21.04.2026"),
"P060232":((2026,2,5),(2026,4,27),"05.02.2026","27.04.2026"),
"P060242":((2026,3,11),(2026,4,28),"11.03.2026","28.04.2026"),
"P060282":((2026,3,25),(2026,5,18),"25.03.2026","18.05.2026"),
"P060322":((2026,3,27),(2026,5,22),"27.03.2026","22.05.2026"),
"P060352":((2026,4,16),(2026,5,23),"16.04.2026","23.05.2026"),
"P060362":((2026,4,22),(2026,5,23),"22.04.2026","23.05.2026"),
"P060372":((2026,4,22),(2026,5,24),"22.04.2026","24.05.2026"),
"P060382":((2026,4,22),(2026,5,25),"22.04.2026","25.05.2026"),
"P060402":((2026,5,4),(2026,5,24),"04.05.2026","24.05.2026"),
"P060412":((2026,5,4),(2026,5,24),"04.05.2026","24.05.2026"),
"P060422":((2026,4,29),(2026,5,23),"29.04.2026","23.05.2026"),
}
B=[("Blue Gelato","BG1024","BG1024",26.14,26,1),
("Blue Sunset Sherbet","BSS1024","BSS1024",25.01,24,1),
("Blue Sunset Sherbet","BSS052501","P050192",20.52,20,2),
("Cap Junky","CJ052501/01","P050162",24.05,24,1),
("Cap Junky","CJ082501/2","P060032",18.29,20,1),
("Cap Junky","CJ062501/1","P050222",24.80,24,2),
("Cap Junky","CJ082501/1","P060022",28.34,28,2),
("Cap Junky","CJ092501","P060072",25.65,26,2),
("Amnesia Core Cut","ACC102501","P060122",12.09,12,2),
("Cash Cow","CC012603","P060372",14.76,14,2),
("Chem Flyer","CF102501","P060132",9.55,10,2),
("Clemosa A Bud","CLE072501","P050282",8.65,8,2),
("Fat Bastard","FB012602","P060352",18.86,18,1),
("Fat Bastard","FB012601/1","P060322",17.99,18,2),
("Gorilla Glue","GG012603","P060402",16.70,16,1),
("Gorilla Glue","GG1024_01","P050092",18.67,18,1),
("Gorilla Glue","GG1024","GG1024",15.51,16,2),
("Grape Pie","GP052501","P050152",18.52,18,1),
("Grape Pie","GP0824_02","P050022",22.61,24,1),
("Grape Pie","GP082501/2","P050322",15.70,16,1),
("Grape Pie","GP072501/2","P050302",18.29,18,2),
("Grape Pie","GP0824_03","P050072",28.30,28,2),
("Grape Pie","GP082501/1","P050312",25.13,26,2),
("Grape Pie","GP092501","P060092",25.24,26,2),
("Grapes and Cream","GRC102501/2","P060182",9.80,10,2),
("High Pro Amnesia","HPA1024","HPA1024",17.31,18,1),
("High Pro Amnesia","HPA1024_01","P050052",21.61,22,1),
("High Pro Amnesia","HPA052501","P050182",19.69,20,2),
("Jelly Donuts","JD112501","P060212",20.32,20,1),
("Jelly Donuts","JD012603/01","P060362",21.01,22,2),
("Jelly Donuts","JD012603/02","P060412",14.43,14,2),
("Jelly Donuts","JD012603/02V","P060422",17.09,16,2),
("Jokerz 31","J31102501","P060152",17.32,18,1),
("Kush Crasher","KC102501","P060172",17.06,16,2),
("Motor Breath","MB0824_05","P050112",17.93,18,2),
("Orange Punch Mimosa","OPM1024","OPM1024",20.03,22,1),
("Orange Punch Mimosa","OPM1024_02","P050062",18.04,18,1),
("Orange Punch Mimosa","OPM122501","P060242",7.91,8,1),
("Orange Punch Mimosa","OMP1024_01","P050042",18.99,18,2),
("Orange Punch Mimosa","OPM092501","P060042",10.18,10,2),
("Orange Punch Mimosa","OPM1024_03","P050082",19.30,20,2),
("Permanent Marker","PM092501","P060062",12.25,12,1),
("Permanent Marker","PM112501","P060232",10.79,10,2),
("Pure Michigen","PUM102501","P060112",13.95,14,2),
("Scrambler","SCR012603","P060382",17.84,18,1),
("Scrambler","SCR112501","P060282",19.73,20,2),
("Sleepy Joe","SJ092501","P060082",10.39,10,2),
("Wedding Crusher","WC082501","P060012",23.48,24,2)]

rows=[]
for s,cul,pp,a,n,tr in B:
    t,lo,hi=band(n)
    assert lo-EPS<=a<=hi+EPS
    assert pp in D, f"no dates for {pp}"
    h,pk,hs,ps=D[pp]
    hd=dt.date(*h) if h else None; pd_=dt.date(*pk)
    rows.append(dict(strain=s,cultiv=cul,pp=pp,assay=a,nom=n,tol=t,lo=lo,hi=hi,tr=tr,
                     g=GRADE[n],ml=a-lo,mu=hi-a,h=hd,p=pd_,hs=hs,ps=ps,
                     cure=(pd_-hd).days if hd else None,age=(TODAY-pd_).days))
assert len(rows)==48

thin=Side(style="thin",color="B7C3CF");BD=Border(left=thin,right=thin,top=thin,bottom=thin)
def hdr(ws,nc):
    for c in range(1,nc+1):
        x=ws.cell(row=1,column=c);x.font=Font(name="Calibri",size=10,bold=True,color=WHITE)
        x.fill=PatternFill("solid",fgColor=NAVY);x.border=BD
        x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
def sc(cell,bold=False,fill=None,align="center",size=10):
    cell.font=Font(name="Calibri",size=size,bold=bold);cell.border=BD
    cell.alignment=Alignment(horizontal=align,vertical="center",wrap_text=True)
    if fill: cell.fill=PatternFill("solid",fgColor=fill)
def fd(d): return d.strftime("%d.%m.%Y") if d else "—"

wb=openpyxl.Workbook()
ws=wb.active;ws.title="Master - All Batches"
h=["#","Tranche","Strain","Grade","Class","Nominal value %","Tolerance","Potency range",
   "Exact THC %","Above floor (pp)","Below ceiling (pp)","Cultiv. Batch No.","PP Batch No.",
   "Manufacturing (harvest) date","Packaging date","Harvest→pack (days)","Age since packaging (days)"]
ws.append(h);hdr(ws,len(h));ws.freeze_panes="A2"
r=2
for i,x in enumerate(sorted(rows,key=lambda r:(-r["nom"],-r["assay"])),start=1):
    ws.append([i,f"T{x['tr']}",x["strain"],f"Grade {x['g']}",f"THC {x['nom']} / CBD 1",
               f"{x['nom']:.2f}%",f"± {x['tol']:.2f}%",f"{x['lo']:.2f}% – {x['hi']:.2f}%",
               f"{x['assay']:.2f}%",round(x["ml"],2),round(x["mu"],2),x["cultiv"],x["pp"],
               fd(x["h"]),fd(x["p"]),x["cure"] if x["cure"] is not None else "—",x["age"]])
    for c in range(1,len(h)+1):
        sc(ws.cell(row=r,column=c),fill=LGREY if r%2==0 else None,
           align="left" if c==3 else "center",bold=(c in (4,6,7,8,14,15)))
    sc(ws.cell(row=r,column=8),fill=GREEN,bold=True)
    if x["ml"]<x["nom"]*U: sc(ws.cell(row=r,column=10),fill=ROSE,bold=True)
    if x["mu"]<x["nom"]*U: sc(ws.cell(row=r,column=11),fill=AMBER,bold=True)
    if x["h"] is None: sc(ws.cell(row=r,column=14),fill=ROSE,bold=True)
    r+=1
for c,w in enumerate([5,9,20,10,16,14,11,20,12,16,17,15,13,26,16,18,22],start=1):
    ws.column_dimensions[get_column_letter(c)].width=w
ws.auto_filter.ref=f"A1:Q{r-1}"

ws2=wb.create_sheet("Master - Per Strain")
h2=["Strain","Grade","Class","Nominal value %","Tolerance","Potency range","PP Batch No.",
    "Exact THC %","Manufacturing (harvest) date","Packaging date","Tranche"]
ws2.append(h2);hdr(ws2,len(h2));ws2.freeze_panes="A2"
bs={}
for x in rows: bs.setdefault(x["strain"],{}).setdefault(x["nom"],[]).append(x)
r=2
for s in sorted(bs,key=str.lower):
    firsts=True
    for n in sorted(bs[s],reverse=True):
        mem=sorted(bs[s][n],key=lambda m:-m["assay"]);t,lo,hi=band(n);firstg=True
        for m in mem:
            ws2.append([s if firsts else "",f"Grade {GRADE[n]}" if firstg else "",
                        f"THC {n} / CBD 1" if firstg else "",f"{n:.2f}%" if firstg else "",
                        f"± {t:.2f}%" if firstg else "",
                        f"{lo:.2f}% – {hi:.2f}%" if firstg else "",
                        m["pp"],f"{m['assay']:.2f}%",fd(m["h"]),fd(m["p"]),f"T{m['tr']}"])
            for c in range(1,len(h2)+1):
                sc(ws2.cell(row=r,column=c),fill=LGREY if r%2==0 else None,
                   align="left" if c==1 else "center",bold=(c in (1,2,6)))
            if firstg: sc(ws2.cell(row=r,column=6),fill=GREEN,bold=True)
            if m["h"] is None: sc(ws2.cell(row=r,column=9),fill=ROSE,bold=True)
            firsts=False;firstg=False;r+=1
for c,w in enumerate([20,10,16,14,11,20,13,12,26,16,9],start=1):
    ws2.column_dimensions[get_column_letter(c)].width=w

ws3=wb.create_sheet("Notes")
ws3.append(["#","Note"]);hdr(ws3,2)
inf=[x["pp"] for x in rows if "*" in x["ps"]]
txt=[("1","Where the source gives a date range or a pair, the FIRST date is taken — for both "
      "manufacturing (harvest) and packaging."),
 ("2","Grade numbering is absolute across the company ladder: Grade I = THC 28 / CBD 1 (the "
      "strongest, 25.20% – 30.79%) down to Grade XI = THC 8 / CBD 1."),
 ("3",f"P050152 (GP052501) has no usable harvest date — the source cell contains ']'. "
      "Manufacturing date is left blank pending the real value."),
 ("4",f"Three packaging dates omitted the year in the source; it is inferred from the harvest "
      f"date and marked in the source-text column: {', '.join(inf)}."),
 ("5","'Harvest→pack' is the cure/processing interval in days. 'Age since packaging' is counted "
      f"to {TODAY.strftime('%d.%m.%Y')} and is the figure that matters for THC degradation: the "
      "older the batch, the more its assay will have drifted down from the released value."),
 ("6","The date table lists P060082 as cultivation batch SJ0925021; it is carried here as "
      "SJ092501, matching the potency data. Confirm which is correct."),
]
for i,(k,t) in enumerate(txt,start=1):
    ws3.append([k,t]);sc(ws3.cell(row=i+1,column=1),bold=True)
    sc(ws3.cell(row=i+1,column=2),align="left");ws3.row_dimensions[i+1].height=48
ws3.column_dimensions["A"].width=6;ws3.column_dimensions["B"].width=118
wb.save(OUT)

print(f"{len(rows)} batches | dates matched for {sum(1 for x in rows if x['h'])}/48 harvest, 48/48 packaging")
print("missing harvest:",[x["pp"] for x in rows if not x["h"]])
cure=[x["cure"] for x in rows if x["cure"] is not None]
print(f"harvest->pack: min {min(cure)} d, median {sorted(cure)[len(cure)//2]} d, max {max(cure)} d")
age=[x["age"] for x in rows]
print(f"age since packaging: min {min(age)} d, max {max(age)} d")
