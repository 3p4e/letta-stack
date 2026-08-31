# -*- coding: utf-8 -*-
"""
MASTER potency table — Tranche 1 + Tranche 2, all 48 batches.

Grade numbering is ABSOLUTE across the whole ladder (not per strain):
  Grade I  = the strongest class, THC 28 / CBD 1 (25.20% - 30.79%)
  Grade II = THC 26, Grade III = THC 24, ... down to Grade XI = THC 8.
So a strain's Grade I is Grade I on the company scale, and two strains carrying
"Grade V" carry the same declared potency.
"""
import openpyxl
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
NAVY="2B547E";LGREY="F7FAFC";WHITE="FFFFFF";AMBER="FFF2CC";GREEN="E2EFDA";ROSE="FCE4D6"
OUT="PP_Potency_MASTER_T1_T2.xlsx";CAP=0.10;U=0.05;EPS=1e-9
def band(n): t=n*CAP; return t,n-t,n+t-0.01
LAD=list(range(28,7,-2))                       # 28 down to 8
ROMAN=["I","II","III","IV","V","VI","VII","VIII","IX","X","XI"]
GRADE={n:ROMAN[i] for i,n in enumerate(LAD)}   # THC28 -> I ... THC8 -> XI

# (strain, cultiv batch, PP batch, THC %, class, tranche)
B=[("Blue Gelato","BG1024","BG1024",26.14,26,1),
("Blue Sunset Sherbet","BSS1024","BSS1024",25.01,24,1),
("Blue Sunset Sherbet","BSS052501","P050192",20.52,20,2),
("Cap Junky","CJ052501/01","P050162",24.05,24,1),
("Cap Junky","CJ082501/2","P060032",18.29,20,1),
("Cap Junky","CJ062501/1","P050222",24.80,24,2),
("Cap Junky","CJ082501/1","P060022",28.34,28,2),
("Cap Junky","CJ092501","P060072",25.65,26,2),
("Amnesia Core Cut","ACC102501","P060122",12.09,12,2),
("Cash Cow","CC012603","P060372",14.76,14,2),
("Chem Flyer","CF102501","P060132",9.55,10,2),
("Clemosa A Bud","CLE072501","P050282",8.65,8,2),
("Fat Bastard","FB012602","P060352",18.86,18,1),
("Fat Bastard","FB012601/1","P060322",17.99,18,2),
("Gorilla Glue","GG012603","P060402",16.70,16,1),
("Gorilla Glue","GG1024_01","P050092",18.67,18,1),
("Gorilla Glue","GG1024","GG1024",15.51,16,2),
("Grape Pie","GP052501","P050152",18.52,18,1),
("Grape Pie","GP0824_02","P050022",22.61,24,1),
("Grape Pie","GP082501/2","P050322",15.70,16,1),
("Grape Pie","GP072501/2","P050302",18.29,18,2),
("Grape Pie","GP0824_03","P050072",28.30,28,2),
("Grape Pie","GP082501/1","P050312",25.13,26,2),
("Grape Pie","GP092501","P060092",25.24,26,2),
("Grapes and Cream","GRC102501/2","P060182",9.80,10,2),
("High Pro Amnesia","HPA1024","HPA1024",17.31,18,1),
("High Pro Amnesia","HPA1024_01","P050052",21.61,22,1),
("High Pro Amnesia","HPA052501","P050182",19.69,20,2),
("Jelly Donuts","JD112501","P060212",20.32,20,1),
("Jelly Donuts","JD012603/01","P060362",21.01,22,2),
("Jelly Donuts","JD012603/02","P060412",14.43,14,2),
("Jelly Donuts","JD012603/02V","P060422",17.09,16,2),
("Jokerz 31","J31102501","P060152",17.32,18,1),
("Kush Crasher","KC102501","P060172",17.06,16,2),
("Motor Breath","MB0824_05","P050112",17.93,18,2),
("Orange Punch Mimosa","OPM1024","OPM1024",20.03,22,1),
("Orange Punch Mimosa","OPM1024_02","P050062",18.04,18,1),
("Orange Punch Mimosa","OPM122501","P060242",7.91,8,1),
("Orange Punch Mimosa","OMP1024_01","P050042",18.99,18,2),
("Orange Punch Mimosa","OPM092501","P060042",10.18,10,2),
("Orange Punch Mimosa","OPM1024_03","P050082",19.30,20,2),
("Permanent Marker","PM092501","P060062",12.25,12,1),
("Permanent Marker","PM112501","P060232",10.79,10,2),
("Pure Michigen","PUM102501","P060112",13.95,14,2),
("Scrambler","SCR012603","P060382",17.84,18,1),
("Scrambler","SCR112501","P060282",19.73,20,2),
("Sleepy Joe","SJ092501","P060082",10.39,10,2),
("Wedding Crusher","WC082501","P060012",23.48,24,2)]

rows=[]
for s,cul,pp,a,n,tr in B:
    t,lo,hi=band(n)
    assert lo-EPS<=a<=hi+EPS,f"{pp} outside THC{n}"
    rows.append(dict(strain=s,cultiv=cul,pp=pp,assay=a,nom=n,tol=t,lo=lo,hi=hi,tr=tr,
                     g=GRADE[n],ml=a-lo,mu=hi-a))
assert len(rows)==48

thin=Side(style="thin",color="B7C3CF");BD=Border(left=thin,right=thin,top=thin,bottom=thin)
def hdr(ws,nc,row=1):
    for c in range(1,nc+1):
        x=ws.cell(row=row,column=c);x.font=Font(name="Calibri",size=10,bold=True,color=WHITE)
        x.fill=PatternFill("solid",fgColor=NAVY);x.border=BD
        x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
def sc(cell,bold=False,fill=None,align="center",size=10):
    cell.font=Font(name="Calibri",size=size,bold=bold);cell.border=BD
    cell.alignment=Alignment(horizontal=align,vertical="center",wrap_text=True)
    if fill: cell.fill=PatternFill("solid",fgColor=fill)

wb=openpyxl.Workbook()

# ---------- TABLE 1: all batches ----------
ws=wb.active;ws.title="Master - All Batches"
h=["#","Tranche","Strain","Grade","Class","Nominal value %","Tolerance","Potency range",
   "Exact THC %","Above floor (pp)","Below ceiling (pp)","Cultiv. Batch No.","PP Batch No."]
ws.append(h);hdr(ws,len(h));ws.freeze_panes="A2"
r=2
for i,x in enumerate(sorted(rows,key=lambda r:(-r["nom"],-r["assay"])),start=1):
    ws.append([i,f"T{x['tr']}",x["strain"],f"Grade {x['g']}",f"THC {x['nom']} / CBD 1",
               f"{x['nom']:.2f}%",f"± {x['tol']:.2f}%",f"{x['lo']:.2f}% – {x['hi']:.2f}%",
               f"{x['assay']:.2f}%",round(x["ml"],2),round(x["mu"],2),x["cultiv"],x["pp"]])
    for c in range(1,len(h)+1):
        sc(ws.cell(row=r,column=c),fill=LGREY if r%2==0 else None,
           align="left" if c==3 else "center",bold=(c in (4,6,7,8)))
    sc(ws.cell(row=r,column=8),fill=GREEN,bold=True)
    if x["ml"]<x["nom"]*U: sc(ws.cell(row=r,column=10),fill=ROSE,bold=True)
    if x["mu"]<x["nom"]*U: sc(ws.cell(row=r,column=11),fill=AMBER,bold=True)
    r+=1
for c,w in enumerate([5,9,20,10,16,14,11,20,12,16,17,15,13],start=1):
    ws.column_dimensions[get_column_letter(c)].width=w
ws.auto_filter.ref=f"A1:M{r-1}"

# ---------- TABLE 2: per strain ----------
ws2=wb.create_sheet("Master - Per Strain")
h2=["Strain","Grade","Class","Nominal value %","Tolerance","Potency range",
    "Batches in this grade","Exact THC % values","Tranche"]
ws2.append(h2);hdr(ws2,len(h2));ws2.freeze_panes="A2"
bs={}
for x in rows: bs.setdefault(x["strain"],{}).setdefault(x["nom"],[]).append(x)
r=2
for s in sorted(bs,key=str.lower):
    first=True
    for n in sorted(bs[s],reverse=True):
        mem=sorted(bs[s][n],key=lambda m:-m["assay"]);t,lo,hi=band(n)
        ws2.append([s if first else "",f"Grade {GRADE[n]}",f"THC {n} / CBD 1",f"{n:.2f}%",
                    f"± {t:.2f}%",f"{lo:.2f}% – {hi:.2f}%",
                    ", ".join(m["pp"] for m in mem),
                    ", ".join(f"{m['assay']:.2f}%" for m in mem),
                    ", ".join(sorted({f"T{m['tr']}" for m in mem}))])
        for c in range(1,len(h2)+1):
            sc(ws2.cell(row=r,column=c),fill=LGREY if r%2==0 else None,
               align="left" if c in (1,7,8) else "center",bold=(c in (1,2,6)))
        sc(ws2.cell(row=r,column=6),fill=GREEN,bold=True)
        first=False;r+=1
for c,w in enumerate([20,10,16,14,11,20,34,26,10],start=1):
    ws2.column_dimensions[get_column_letter(c)].width=w

# ---------- TABLE 3: the grade scale ----------
ws3=wb.create_sheet("Grade Scale")
h3=["Grade","Class","Nominal value %","Tolerance","Potency range","Band width (pp)",
    "Batches","T1","T2","Strains represented"]
ws3.append(h3);hdr(ws3,len(h3))
byc={}
for x in rows: byc.setdefault(x["nom"],[]).append(x)
r=2
for n in LAD:
    t,lo,hi=band(n);mem=byc.get(n,[])
    ws3.append([f"Grade {GRADE[n]}",f"THC {n} / CBD 1",f"{n:.2f}%",f"± {t:.2f}%",
                f"{lo:.2f}% – {hi:.2f}%",f"{hi-lo+0.01:.2f}",len(mem),
                sum(1 for m in mem if m["tr"]==1),sum(1 for m in mem if m["tr"]==2),
                ", ".join(sorted({m["strain"] for m in mem})) or "—"])
    for c in range(1,len(h3)+1):
        sc(ws3.cell(row=r,column=c),fill=LGREY if r%2==0 else None,
           align="left" if c==10 else "center",bold=(c in (1,5)))
    sc(ws3.cell(row=r,column=5),fill=GREEN,bold=True)
    if not mem: sc(ws3.cell(row=r,column=10),fill=AMBER,align="left")
    r+=1
for c,w in enumerate([10,16,14,11,20,14,9,6,6,60],start=1):
    ws3.column_dimensions[get_column_letter(c)].width=w
wb.save(OUT)

print("GRADE SCALE (absolute — Grade I is the strongest class on the company ladder)\n")
for n in LAD:
    t,lo,hi=band(n);mem=byc.get(n,[])
    print(f"Grade {GRADE[n]:<5s} THC{n} : CBD1,  {n:.2f}% ±{t:.2f}%  ({lo:.2f}% – {hi:.2f}%)   "
          f"{len(mem)} batch(es)")
print(f"\n{len(rows)} batches | 19 Tranche 1 | 29 Tranche 2 | all inside their grade")
