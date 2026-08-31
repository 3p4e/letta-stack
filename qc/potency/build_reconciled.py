# -*- coding: utf-8 -*-
"""
RECONCILED potency grades.

Tranche 1 (19 batches) keeps its stated values and its declared classes — all 19
conform at nominal ±10%, nothing changes.

29 batches carry newly valid THC results. Each was re-checked against the class it
was originally declared against, at the full ±10% tolerance:
  * 13 still conform  -> class unchanged
  * 16 no longer conform. In every one of those cases the tolerance needed to hold
    the batch in its original class exceeds the 10% cap (10.1% to 31.4%), so the
    class cannot be kept -- the declared class is corrected to the class the new
    result actually falls in (nearest declared value where more than one fits).

Band = [nominal - 10%*nominal, nominal + 10%*nominal - 0.01].
"""
import openpyxl
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
NAVY="2B547E";LGREY="F7FAFC";WHITE="FFFFFF";AMBER="FFF2CC";GREEN="E2EFDA";ROSE="FCE4D6"
OUT="PP_Potency_Grades_RECONCILED.xlsx"; CAP=0.10; EPS=1e-9
ROMAN=["I","II","III","IV","V","VI","VII","VIII"]
def band(n): t=n*CAP; return t,n-t,n+t-0.01
LAD=list(range(8,27,2))
def fits(a): return [n for n in LAD if band(n)[1]-EPS<=a<=band(n)[2]+EPS]
def nearest(a):
    f=fits(a); return min(f,key=lambda n:(abs(n-a),n)) if f else None

T1=[("Blue Gelato","BG1024","BG1024",26.14,26),("Blue Sunset Sherbet","BSS1024","BSS1024",25.01,24),
("Cap Junky","CJ052501/01","P050162",24.05,24),("Cap Junky","CJ082501/2","P060032",18.29,20),
("Fat Bastard","FB012602","P060352",18.86,18),("Gorilla Glue","GG012603","P060402",16.70,16),
("Gorilla Glue","GG1024_01","P050092",18.67,18),("Grape Pie","GP052501","P050152",18.52,18),
("Grape Pie","GP0824_02","P050022",22.61,24),("Grape Pie","GP082501/2","P050322",15.70,16),
("High Pro Amnesia","HPA1024","HPA1024",17.31,18),("High Pro Amnesia","HPA1024_01","P050052",21.61,22),
("Jelly Donuts","JD112501","P060212",20.32,20),("Jokerz 31","J31102501","P060152",17.32,18),
("Orange Punch Mimosa","OPM1024","OPM1024",20.03,22),("Orange Punch Mimosa","OPM1024_02","P050062",18.04,18),
("Orange Punch Mimosa","OPM122501","P060242",7.91,8),("Permanent Marker","PM092501","P060062",12.25,12),
("Scrambler","SCR012603","P060382",17.84,18)]

NEW=[("Amnesia Core Cut","ACC102501","P060122",12.09,12),("Blue Sunset Sherbet","BSS052501","P050192",20.52,20),
("Cap Junky","CJ062501/1","P050222",24.80,22),("Cap Junky","CJ082501/1","P060022",28.34,24),
("Cap Junky","CJ092501","P060072",25.65,24),("Cash Cow","CC012603","P060372",14.76,12),
("Chem Flyer","CF102501","P060132",9.55,10),("Clemosa A Bud","CLE072501","P050282",8.65,8),
("Fat Bastard","FB012601/1","P060322",17.99,14),("Gorilla Glue","GG1024","GG1024",15.51,14),
("Grape Pie","GP072501/2","P050302",18.29,20),("Grape Pie","GP0824_03","P050072",28.30,24),
("Grape Pie","GP082501/1","P050312",25.13,20),("Grape Pie","GP092501","P060092",25.24,24),
("Grapes and Cream","GRC102501/2","P060182",9.80,12),("High Pro Amnesia","HPA052501","P050182",19.69,22),
("Jelly Donuts","JD012603/01","P060362",21.01,16),("Jelly Donuts","JD012603/02","P060412",14.43,20),
("Jelly Donuts","JD012603/02V","P060422",17.09,16),("Kush Crasher","KC102501","P060172",17.06,18),
("Motor Breath","MB0824_05","P050112",17.93,16),("Orange Punch Mimosa","OMP1024_01","P050042",18.99,16),
("Orange Punch Mimosa","OPM092501","P060042",10.18,10),("Orange Punch Mimosa","OPM1024_03","P050082",19.30,16),
("Permanent Marker","PM112501","P060232",10.79,12),("Pure Michigen","PUM102501","P060112",13.95,16),
("Scrambler","SCR112501","P060282",19.73,18),("Sleepy Joe","SJ092501","P060082",10.39,10),
("Wedding Crusher","WC082501","P060012",23.48,22)]

rows=[]
for s,cul,pp,a,n in T1:
    t,lo,hi=band(n)
    rows.append(dict(strain=s,cultiv=cul,pp=pp,assay=a,decl=n,final=n,src="Tranche 1",
                     changed=False,need=None,tol=t,lo=lo,hi=hi))
for s,cul,pp,a,n in NEW:
    t,lo,hi=band(n)
    if lo-EPS<=a<=hi+EPS:
        rows.append(dict(strain=s,cultiv=cul,pp=pp,assay=a,decl=n,final=n,src="new result",
                         changed=False,need=None,tol=t,lo=lo,hi=hi))
    else:
        need=((a-n+0.01) if a>=n else (n-a))/n*100
        f=nearest(a); tf,lof,hif=band(f)
        rows.append(dict(strain=s,cultiv=cul,pp=pp,assay=a,decl=n,final=f,src="new result",
                         changed=True,need=need,tol=tf,lo=lof,hi=hif))
assert all(r["lo"]-EPS<=r["assay"]<=r["hi"]+EPS for r in rows)
assert len(rows)==48

gr={}
by={}
for r in rows: by.setdefault(r["strain"],set()).add(r["final"])
for s,ns in by.items():
    for i,n in enumerate(sorted(ns,reverse=True)): gr[(s,n)]=ROMAN[i]

thin=Side(style="thin",color="B7C3CF");BD=Border(left=thin,right=thin,top=thin,bottom=thin)
def hdr(ws,nc):
    for c in range(1,nc+1):
        x=ws.cell(row=1,column=c);x.font=Font(name="Calibri",size=10,bold=True,color=WHITE)
        x.fill=PatternFill("solid",fgColor=NAVY);x.border=BD
        x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
def sc(cell,bold=False,fill=None,size=10,align="center"):
    cell.font=Font(name="Calibri",size=size,bold=bold);cell.border=BD
    cell.alignment=Alignment(horizontal=align,vertical="center",wrap_text=True)
    if fill: cell.fill=PatternFill("solid",fgColor=fill)

wb=openpyxl.Workbook()
ws=wb.active;ws.title="Reconciled Grades"
h=["Strain","Grade","PP Batch No.","THC/CBD ratio","Nominal value %","Tolerance","Range",
   "THC result %","Source","Originally declared","Changed?","Tol. needed to keep original"]
ws.append(h);hdr(ws,len(h));ws.freeze_panes="A2"
r=2
for x in sorted(rows,key=lambda r:(r["strain"].lower(),-r["final"],-r["assay"])):
    ws.append([x["strain"],f"Grade {gr[(x['strain'],x['final'])]}",x["pp"],
               f"THC {x['final']} / CBD 1",f"{x['final']:.2f}%",f"± {x['tol']:.2f}",
               f"{x['lo']:.2f}% – {x['hi']:.2f}%",f"{x['assay']:.2f}%",x["src"],
               f"THC {x['decl']} / CBD 1","CHANGED" if x["changed"] else "no",
               f"{x['need']:.1f}%  (> 10% cap)" if x["changed"] else "—"])
    for c in range(1,len(h)+1):
        sc(ws.cell(row=r,column=c),fill=LGREY if r%2==0 else None,
           align="left" if c==1 else "center",bold=(c in (2,5,6,7)))
    sc(ws.cell(row=r,column=7),fill=GREEN,bold=True)
    if x["changed"]:
        for c in (10,11,12): sc(ws.cell(row=r,column=c),fill=ROSE,bold=(c==11))
    r+=1
for c,w in enumerate([20,9,13,16,14,11,20,13,12,18,11,26],start=1):
    ws.column_dimensions[get_column_letter(c)].width=w
ws.auto_filter.ref=f"A1:L{r-1}"

ws2=wb.create_sheet("Class Changes")
h2=["PP Batch No.","Strain","New THC result %","Originally declared","Its range",
    "Off by (pp)","Tolerance needed","Verdict","Reconciled class","New range"]
ws2.append(h2);hdr(ws2,len(h2));ws2.freeze_panes="A2"
r=2
for x in sorted([r_ for r_ in rows if r_["changed"]],key=lambda r_:-r_["need"]):
    td,lod,hid=band(x["decl"]); off=(x["assay"]-hid) if x["assay"]>hid else (lod-x["assay"])
    ws2.append([x["pp"],x["strain"],f"{x['assay']:.2f}%",f"THC {x['decl']} / CBD 1",
                f"{lod:.2f}% – {hid:.2f}%",f"{off:.2f}",f"{x['need']:.1f}%",
                "exceeds the 10% cap — class cannot be kept",
                f"THC {x['final']} / CBD 1",f"{x['lo']:.2f}% – {x['hi']:.2f}%"])
    for c in range(1,len(h2)+1):
        sc(ws2.cell(row=r,column=c),fill=LGREY if r%2==0 else None,
           align="left" if c in (2,8) else "center",bold=(c in (1,9)))
    sc(ws2.cell(row=r,column=7),fill=ROSE,bold=True);sc(ws2.cell(row=r,column=9),fill=GREEN,bold=True)
    r+=1
for c,w in enumerate([13,20,16,18,18,12,16,38,18,18],start=1):
    ws2.column_dimensions[get_column_letter(c)].width=w

ws3=wb.create_sheet("Class Ladder")
h3=["Class","Nominal","Tolerance","Range","Batches"]
ws3.append(h3);hdr(ws3,len(h3))
pop={}
for x in rows: pop.setdefault(x["final"],[]).append(x["pp"])
r=2
for n in range(26,7,-2):
    t,lo,hi=band(n)
    ws3.append([f"THC {n} / CBD 1",f"{n:.2f}%",f"± {t:.2f}",f"{lo:.2f}% – {hi:.2f}%",
                ", ".join(sorted(pop.get(n,[]))) or "— (no batch)"])
    for c in range(1,len(h3)+1):
        sc(ws3.cell(row=r,column=c),fill=LGREY if r%2==0 else None,
           align="left" if c==5 else "center",bold=(c in (1,4)))
    sc(ws3.cell(row=r,column=4),fill=GREEN,bold=True)
    if n not in pop: sc(ws3.cell(row=r,column=5),fill=AMBER,align="left")
    r+=1
for c,w in enumerate([16,12,11,20,80],start=1): ws3.column_dimensions[get_column_letter(c)].width=w
wb.save(OUT)
ch=[x for x in rows if x["changed"]]
print(f"48 batches | Tranche 1: 19 unchanged | new results: {29-len(ch)} unchanged, {len(ch)} re-classed")
print("all 48 now conform at nominal +/-10%")
print("empty classes:",[n for n in range(8,27,2) if n not in pop])
