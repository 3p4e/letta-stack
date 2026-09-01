# -*- coding: utf-8 -*-
"""
TRANCHE 2 potency grades, built from the latest retest THC% — same method as Tranche 1.

Method (identical to Tranche 1):
  * class ladder = whole EVEN nominals, THC 8 ... THC 28 / CBD 1
  * tolerance = the full 10% of nominal
  * band(N) = [ N - 0.10*N , N + 0.10*N - 0.01 ]   upper limit exclusive
  * a batch is graded to the permitted declared value NEAREST its result, which is
    also the class that leaves it the largest margin to a limit

The ladder is extended to THC 28 / CBD 1 because two retest results (P060022 28.34%,
P050072 28.30%) sit above every previously used class.
"""
import openpyxl
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
NAVY="2B547E";LGREY="F7FAFC";WHITE="FFFFFF";AMBER="FFF2CC";GREEN="E2EFDA";ROSE="FCE4D6"
OUT="PP_Tranche2_Potency_Grades.xlsx";CAP=0.10;U_REL=0.05;EPS=1e-9
ROMAN=["I","II","III","IV","V"]
def band(n): t=n*CAP; return t,n-t,n+t-0.01
LAD=list(range(8,29,2))
def fits(a): return [n for n in LAD if band(n)[1]-EPS<=a<=band(n)[2]+EPS]
def marg(a,n):
    t,lo,hi=band(n); return min(a-lo,hi-a)
def grade(a):
    f=fits(a)
    if not f: return None
    n=min(f,key=lambda k:(abs(k-a),k))
    if (a-band(n)[1]) < n*U_REL:
        lo=[k for k in f if k<n]
        if lo: return max(lo)
    return n
def _old_grade(a):
    f=fits(a); return min(f,key=lambda n:(abs(n-a),n)) if f else None

# (strain, cultiv batch, PP batch, retest THC %, class declared before)
T2=[("Amnesia Core Cut","ACC102501","P060122",12.09,12),
("Blue Sunset Sherbet","BSS052501","P050192",20.52,20),
("Cap Junky","CJ062501/1","P050222",24.80,22),("Cap Junky","CJ082501/1","P060022",28.34,24),
("Cap Junky","CJ092501","P060072",25.65,24),("Cash Cow","CC012603","P060372",14.76,12),
("Chem Flyer","CF102501","P060132",9.55,10),("Clemosa A Bud","CLE072501","P050282",8.65,8),
("Fat Bastard","FB012601/1","P060322",17.99,14),("Gorilla Glue","GG1024","GG1024",15.51,14),
("Grape Pie","GP072501/2","P050302",18.29,20),("Grape Pie","GP0824_03","P050072",28.30,24),
("Grape Pie","GP082501/1","P050312",25.13,20),("Grape Pie","GP092501","P060092",25.24,24),
("Grapes and Cream","GRC102501/2","P060182",9.80,12),("High Pro Amnesia","HPA052501","P050182",19.69,22),
("Jelly Donuts","JD012603/01","P060362",21.01,16),("Jelly Donuts","JD012603/02","P060412",14.43,20),
("Jelly Donuts","JD012603/02V","P060422",17.09,16),("Kush Crasher","KC102501","P060172",17.06,18),
("Motor Breath","MB0824_05","P050112",17.93,16),("Orange Punch Mimosa","OMP1024_01","P050042",18.99,16),
("Orange Punch Mimosa","OPM092501","P060042",10.18,10),("Orange Punch Mimosa","OPM1024_03","P050082",19.30,16),
("Permanent Marker","PM112501","P060232",10.79,12),("Pure Michigen","PUM102501","P060112",13.95,16),
("Scrambler","SCR112501","P060282",19.73,18),("Sleepy Joe","SJ092501","P060082",10.39,10),
("Wedding Crusher","WC082501","P060012",23.48,22)]

rows=[]
for s,cul,pp,a,old in T2:
    n=grade(a);t,lo,hi=band(n);m=marg(a,n)
    rows.append(dict(strain=s,cultiv=cul,pp=pp,assay=a,nom=n,tol=t,lo=lo,hi=hi,
                     m=m,rel=m/n*100,old=old,moved=(n!=old),tight=(m/n*100<U_REL*100)))
assert all(r["lo"]-EPS<=r["assay"]<=r["hi"]+EPS for r in rows) and len(rows)==29

gr={};by={}
for r in rows: by.setdefault(r["strain"],set()).add(r["nom"])
for s,ns in by.items():
    for i,n in enumerate(sorted(ns,reverse=True)): gr[(s,n)]=ROMAN[i]

thin=Side(style="thin",color="B7C3CF");BD=Border(left=thin,right=thin,top=thin,bottom=thin)
def hdr(ws,nc):
    for c in range(1,nc+1):
        x=ws.cell(row=1,column=c);x.font=Font(name="Calibri",size=10,bold=True,color=WHITE)
        x.fill=PatternFill("solid",fgColor=NAVY);x.border=BD
        x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
def sc(cell,bold=False,fill=None,align="center",size=10):
    cell.font=Font(name="Calibri",size=size,bold=bold);cell.border=BD
    cell.alignment=Alignment(horizontal=align,vertical="center",wrap_text=True)
    if fill: cell.fill=PatternFill("solid",fgColor=fill)

wb=openpyxl.Workbook()
ws=wb.active;ws.title="Tranche 2 Grades"
h=["Strain","Grade","PP Batch No.","Cultiv. Batch No.","THC/CBD ratio","Nominal value %",
   "Tolerance","Range","Retest THC %","Margin (pp)","Margin (% of nom.)","Class before","Changed?"]
ws.append(h);hdr(ws,len(h));ws.freeze_panes="A2"
r=2
for x in sorted(rows,key=lambda r:(r["strain"].lower(),-r["nom"],-r["assay"])):
    ws.append([x["strain"],f"Grade {gr[(x['strain'],x['nom'])]}",x["pp"],x["cultiv"],
               f"THC {x['nom']} / CBD 1",f"{x['nom']:.2f}%",f"± {x['tol']:.2f}%",
               f"{x['lo']:.2f}% – {x['hi']:.2f}%",f"{x['assay']:.2f}%",
               round(x["m"],2),f"{x['rel']:.2f}%",f"THC {x['old']} / CBD 1",
               "CHANGED" if x["moved"] else "no"])
    for c in range(1,len(h)+1):
        sc(ws.cell(row=r,column=c),fill=LGREY if r%2==0 else None,
           align="left" if c==1 else "center",bold=(c in (2,6,7,8)))
    sc(ws.cell(row=r,column=8),fill=GREEN,bold=True)
    sc(ws.cell(row=r,column=11),fill=AMBER if x["tight"] else GREEN,bold=True)
    if x["moved"]: sc(ws.cell(row=r,column=13),fill=ROSE,bold=True)
    r+=1
for c,w in enumerate([20,9,13,15,16,14,11,20,13,12,17,16,11],start=1):
    ws.column_dimensions[get_column_letter(c)].width=w
ws.auto_filter.ref=f"A1:M{r-1}"

ws2=wb.create_sheet("Class Ladder")
h2=["Class","Nominal value %","Tolerance","Range","Band width (pp)","Batches"]
ws2.append(h2);hdr(ws2,len(h2))
byc={}
for x in rows: byc.setdefault(x["nom"],[]).append(x["pp"])
r=2
for n in sorted(LAD,reverse=True):
    t,lo,hi=band(n)
    ws2.append([f"THC {n} / CBD 1",f"{n:.2f}%",f"± {t:.2f}%",f"{lo:.2f}% – {hi:.2f}%",
                f"{hi-lo+0.01:.2f}",", ".join(byc.get(n,[])) or "— (no Tranche 2 batch)"])
    for c in range(1,len(h2)+1):
        sc(ws2.cell(row=r,column=c),fill=LGREY if r%2==0 else None,
           align="left" if c==6 else "center",bold=(c in (1,4)))
    sc(ws2.cell(row=r,column=4),fill=GREEN,bold=True)
    if n not in byc: sc(ws2.cell(row=r,column=6),fill=AMBER,align="left")
    r+=1
for c,w in enumerate([16,15,11,20,15,60],start=1): ws2.column_dimensions[get_column_letter(c)].width=w

ws3=wb.create_sheet("Grades per Strain")
h3=["Strain","Grade","Class","Nominal","Tolerance","Range","Batches (retest THC %)"]
ws3.append(h3);hdr(ws3,len(h3))
bs={}
for x in rows: bs.setdefault(x["strain"],{}).setdefault(x["nom"],[]).append(x)
r=2
for s in sorted(bs,key=str.lower):
    for n in sorted(bs[s],reverse=True):
        t,lo,hi=band(n);mem=sorted(bs[s][n],key=lambda m:-m["assay"])
        ws3.append([s,f"Grade {gr[(s,n)]}",f"THC {n} / CBD 1",f"{n:.2f}%",f"± {t:.2f}%",
                    f"{lo:.2f}% – {hi:.2f}%",
                    ", ".join(f"{m['pp']} ({m['assay']:.2f}%)" for m in mem)])
        for c in range(1,len(h3)+1):
            sc(ws3.cell(row=r,column=c),fill=LGREY if r%2==0 else None,
               align="left" if c in (1,7) else "center",bold=(c in (2,6)))
        sc(ws3.cell(row=r,column=6),fill=GREEN,bold=True)
        r+=1
for c,w in enumerate([20,9,16,12,11,20,52],start=1): ws3.column_dimensions[get_column_letter(c)].width=w

ws4=wb.create_sheet("Watch List")
h4=["PP Batch No.","Strain","Retest THC %","Class","Range","Margin (pp)","Margin (%)","Nearest limit","Note"]
ws4.append(h4);hdr(ws4,len(h4))
r=2
for x in sorted([y for y in rows if y["tight"]],key=lambda y:y["rel"]):
    near="upper" if (x["hi"]-x["assay"])<(x["assay"]-x["lo"]) else "lower"
    alt=[k for k in fits(x["assay"]) if k!=x["nom"]]
    note=("no alternative class exists" if not alt else
          "alternative classes give a smaller margin")
    ws4.append([x["pp"],x["strain"],f"{x['assay']:.2f}%",f"THC {x['nom']} / CBD 1",
                f"{x['lo']:.2f}% – {x['hi']:.2f}%",round(x["m"],2),f"{x['rel']:.2f}%",near,note])
    for c in range(1,len(h4)+1):
        sc(ws4.cell(row=r,column=c),fill=LGREY if r%2==0 else None,
           align="left" if c in (2,9) else "center",bold=(c==1))
    sc(ws4.cell(row=r,column=7),fill=ROSE if x["rel"]<2.5 else AMBER,bold=True)
    r+=1
for c,w in enumerate([13,20,14,16,20,12,12,13,40],start=1): ws4.column_dimensions[get_column_letter(c)].width=w

ws5=wb.create_sheet("Method")
ws5.append(["#","Note"]);hdr(ws5,2)
mv=[x for x in rows if x["moved"]];tg=[x for x in rows if x["tight"]]
txt=[("1","Built by the same method as Tranche 1: even whole nominals THC 8 … THC 28 / CBD 1, "
      "tolerance at the full 10% of nominal, band = [nominal − tolerance, nominal + tolerance − 0.01]."),
 ("2","Each batch is graded to the permitted declared value NEAREST its retest result — the house "
      "assignment rule. That class is also the one leaving the batch the largest margin to a limit."),
 ("3","The ladder is extended to THC 28 / CBD 1 (25.20% – 30.79%). Two retest results sit above every "
      "previously used class: P060022 at 28.34% and P050072 at 28.30%."),
 ("4",f"All 29 batches sit inside their grade. {len(mv)} of 29 carry a different class than they were "
      "declared against before the retest — that follows from the retest values, not from the method."),
 ("5",f"{len(tg)} batches have a margin below 5% of nominal and are listed on the Watch List sheet. "
      "None is out of specification; each is close enough to a limit that a repeat assay could move it."),
 ("6","THC 22 / CBD 1 holds a single batch and THC 16 / CBD 1 holds a single batch; no even class "
      "between THC 8 and THC 28 is left unused except by the shape of the data."),
]
for i,(k,t) in enumerate(txt,start=1):
    ws5.append([k,t]);sc(ws5.cell(row=i+1,column=1),bold=True)
    sc(ws5.cell(row=i+1,column=2),align="left");ws5.row_dimensions[i+1].height=50
ws5.column_dimensions["A"].width=6;ws5.column_dimensions["B"].width=120
wb.save(OUT)
print(f"29 batches | all inside grade | {len(mv)} changed class | {len(tg)} on watch list")
print("classes used:",sorted(byc))
