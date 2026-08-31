# -*- coding: utf-8 -*-
"""
Potency grades derived from the actual THC distribution WITHIN each strain.

Rules
-----
R1  Band semantics (from the worked example 20.00% ±2.0 -> 18.00% – 21.99%):
        band(N, t) = [ N - t , N + t - 0.01 ]
    upper limit exclusive, so consecutive bands butt together with no gap and
    no overlap.

R2  NO EMPTY GRADES. Every grade must contain at least one batch. Bridge /
    placeholder classes are not allowed.

R3  NO GAPS between grades. With a symmetric ±t per grade, two consecutive
    nominals tile exactly when t_i + t_(i+1) = N_(i+1) - N_i. Holding the
    tolerance constant across a strain, that means the ladder must have
    UNIFORM SPACING s, with t = s/2.

R4  Nominals are whole numbers. EVEN is preferred. Where an even-only ladder
    cannot close a gap, one nominal may be pushed up or down onto an ODD whole
    number so the ladder tiles — this is the explicit trade the brief allows:
        either leave a gap and keep every nominal even,
        or shift a nominal to odd and close the gap.

R5  Hard cap: t <= 10% of that grade's nominal.

R6  Exception: where the cap makes tiling impossible (the low-THC end, and
    genuinely bimodal strains), the ladder is split and the gap is left open
    with even nominals on both sides — per R4's first branch.

R7  Grade numbering runs within the strain: strongest = Grade I, descending.

Search: for every candidate uniform ladder (integer spacing s, integer base b)
the batches are binned; a ladder is valid when the populated bins form a
contiguous run (R2 + R3) and the cap holds (R5). Valid ladders are ranked by
   (number of odd nominals, |s - 2|, -number of grades)
so an all-even ladder always beats an odd one, and the house's existing 2 pp
band width is preferred. When no uniform ladder is valid the strain is split at
its widest gap and each part solved independently (R6).
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY="2B547E"; LGREY="F7FAFC"; WHITE="FFFFFF"; AMBER="FFF2CC"; GREEN="E2EFDA"; ROSE="FCE4D6"
OUT=("/tmp/claude-0/-home-user-letta-stack/fa8a4b28-563d-5957-9984-7e34a8196007/"
     "scratchpad/batch_excel/PP_Potency_Grades_Distribution.xlsx")
CAP=0.10
ROMAN=["I","II","III","IV","V","VI","VII","VIII","IX","X"]

BATCHES=[
 ("Blue Gelato","BG1024","BG1024",26.14),
 ("Blue Sunset Sherbet","BSS1024","BSS1024",25.01),
 ("Blue Sunset Sherbet","BSS052501","P050192",20.47),
 ("Cap Junky","CJ052501/01","P050162",24.05),
 ("Cap Junky","CJ082501/2","P060032",18.29),
 ("Cap Junky","CJ062501/1","P050222",21.51),
 ("Cap Junky","CJ082501/1","P060022",24.96),
 ("Cap Junky","CJ092501","P060072",22.30),
 ("Fat Bastard","FB012602","P060352",18.86),
 ("Fat Bastard","FB012601/1","P060322",14.68),
 ("Gorilla Glue","GG012603","P060402",16.70),
 ("Gorilla Glue","GG1024_01","P050092",18.67),
 ("Gorilla Glue","GG1024","GG1024",13.34),
 ("Grape Pie","GP052501","P050152",18.52),
 ("Grape Pie","GP0824_02","P050022",22.61),
 ("Grape Pie","GP082501/2","P050322",15.70),
 ("Grape Pie","GP072501/2","P050302",19.81),
 ("Grape Pie","GP0824_03","P050072",25.45),
 ("Grape Pie","GP082501/1","P050312",21.29),
 ("Grape Pie","GP092501","P060092",25.73),
 ("Grapes and Cream","GRC102501/2","P060182",11.53),
 ("High Pro Amnesia","HPA1024","HPA1024",17.31),
 ("High Pro Amnesia","HPA1024_01","P050052",21.61),
 ("High Pro Amnesia","HPA052501","P050182",20.39),
 ("Jelly Donuts","JD112501","P060212",20.32),
 ("Jelly Donuts","JD012603/01","P060362",16.71),
 ("Jelly Donuts","JD012603/02","P060412",20.54),
 ("Jelly Donuts","JD012603/02V","P060422",15.16),
 ("Jokerz 31","J31102501","P060152",17.32),
 ("Orange Punch Mimosa","OPM1024","OPM1024",20.03),
 ("Orange Punch Mimosa","OPM1024_02","P050062",18.04),
 ("Orange Punch Mimosa","OPM122501","P060242",7.91),
 ("Orange Punch Mimosa","OMP1024_01","P050042",15.38),
 ("Orange Punch Mimosa","OPM092501","P060042",9.43),
 ("Orange Punch Mimosa","OPM1024_03","P050082",16.55),
 ("Permanent Marker","PM092501","P060062",12.25),
 ("Permanent Marker","PM112501","P060232",13.00),
 ("Scrambler","SCR012603","P060382",17.84),
 ("Scrambler","SCR112501","P060282",17.13),
 ("Amnesia Core Cut","ACC102501","P060122",12.91),
 ("Cash Cow","CC012603","P060372",12.32),
 ("Chem Flyer","CF102501","P060132",9.83),
 ("Clemosa A Bud","CLE072501","P050282",8.02),
 ("Kush Crasher","KC102501","P060172",17.40),
 ("Motor Breath","MB0824_05","P050112",16.82),
 ("Pure Michigen","PUM102501","P060112",15.63),
 ("Sleepy Joe","SJ092501","P060082",10.96),
 ("Wedding Crusher","WC082501","P060012",21.67),
]

EPS=1e-9


def try_ladder(items, s, b):
    """items = [(cultiv, pp, assay)]. Uniform ladder base b, spacing s, t=s/2.
    Returns list of grades or None."""
    t = s / 2.0
    bins = {}
    for cul, pp, a in items:
        k = int((a - (b - t)) // s)
        n = b + k * s
        if n <= 0:
            return None
        bins.setdefault(k, []).append((cul, pp, a))
    ks = sorted(bins)
    if ks != list(range(ks[0], ks[-1] + 1)):       # R2 + R3: populated bins must be contiguous
        return None
    grades = []
    for k in ks:
        n = b + k * s
        if t > n * CAP + EPS:                      # R5 cap
            return None
        lo, hi = n - t, n + t - 0.01
        for _, _, a in bins[k]:
            if not (lo - EPS <= a <= hi + EPS):
                return None
        grades.append(dict(nom=n, tol=t, lo=lo, hi=hi, members=bins[k]))
    return grades


def solve_uniform(items):
    """Best uniform ladder for these items, or None."""
    lo_a = min(a for _, _, a in items)
    hi_a = max(a for _, _, a in items)
    best = None
    for s in range(1, 13):
        for b in range(max(1, int(lo_a) - s - 2), int(hi_a) + s + 3):
            g = try_ladder(items, float(s), float(b))
            if not g:
                continue
            n_odd = sum(1 for x in g if int(x["nom"]) % 2 == 1)
            key = (n_odd, abs(s - 2), -len(g), s)
            if best is None or key < best[0]:
                best = (key, g)
    return best[1] if best else None


def solve(items):
    """Solve a strain: uniform ladder if possible, else split at the widest gap (R6)."""
    g = solve_uniform(items)
    if g:
        return [g]
    order = sorted(items, key=lambda x: x[2])
    gaps = [(order[i+1][2] - order[i][2], i) for i in range(len(order) - 1)]
    if not gaps:
        return None
    _, idx = max(gaps)
    left, right = order[:idx+1], order[idx+1:]
    ls, rs = solve(left), solve(right)
    if ls is None or rs is None:
        return None
    return ls + rs


def fallback_single(items):
    """Last resort for a lone batch: nearest even nominal inside the cap."""
    a = items[0][2]
    for n in sorted(range(2, 40), key=lambda n: (n % 2, abs(n - a))):
        t = min(1.0, n * CAP)
        if n - t - EPS <= a <= n + t - 0.01 + EPS:
            return [[dict(nom=float(n), tol=t, lo=n - t, hi=n + t - 0.01, members=items)]]
    return None


thin=Side(style="thin",color="B7C3CF"); BD=Border(left=thin,right=thin,top=thin,bottom=thin)
def hdr(ws,nc):
    for c in range(1,nc+1):
        x=ws.cell(row=1,column=c); x.font=Font(name="Calibri",size=10,bold=True,color=WHITE)
        x.fill=PatternFill("solid",fgColor=NAVY)
        x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); x.border=BD
def sc(cell,bold=False,fill=None,size=10,align="center"):
    cell.font=Font(name="Calibri",size=size,bold=bold); cell.border=BD
    cell.alignment=Alignment(horizontal=align,vertical="center",wrap_text=True)
    if fill: cell.fill=PatternFill("solid",fgColor=fill)


def main():
    strains={}
    for s,cul,pp,a in BATCHES: strains.setdefault(s,[]).append((cul,pp,a))

    solved={}
    for st,items in strains.items():
        blocks = solve(items) or fallback_single(items)
        assert blocks, f"no ladder for {st}"
        flat=[g for blk in blocks for g in blk]
        flat.sort(key=lambda g:-g["nom"])
        for i,g in enumerate(flat): g["grade"]=ROMAN[i]
        solved[st]=dict(blocks=blocks, grades=flat)

    # verify
    seen=0
    for st,S in solved.items():
        for g in S["grades"]:
            assert g["members"], f"{st} empty grade"          # R2
            assert float(g["nom"]).is_integer()               # R4
            assert g["tol"] <= g["nom"]*CAP+EPS               # R5
            for _,_,a in g["members"]:
                assert g["lo"]-EPS<=a<=g["hi"]+EPS
                seen+=1
        asc=sorted(S["grades"],key=lambda g:g["nom"])
        for x,y in zip(asc,asc[1:]):
            d=y["lo"]-(x["hi"]+0.01)
            assert d>-EPS, f"{st} overlap"                     # never overlap
    assert seen==len(BATCHES), f"{seen} != {len(BATCHES)}"

    wb=openpyxl.Workbook()

    # ---- Sheet 1: final answer, requested format ----
    ws=wb.active; ws.title="Potency Grades"
    h=["Strain","Grade","PP Batch No.","Cultiv. Batch No.","Nominal value %","Tolerance",
       "Range","Assay THC %","Nominal type"]
    ws.append(h); hdr(ws,len(h)); ws.freeze_panes="A2"
    r=2
    for st in sorted(solved,key=str.lower):
        for g in solved[st]["grades"]:
            odd = int(g["nom"])%2==1
            for cul,pp,a in sorted(g["members"],key=lambda m:-m[2]):
                ws.append([st,f"Grade {g['grade']}",pp,cul,f"{g['nom']:.2f}%",
                           f"± {g['tol']:.2f}",f"{g['lo']:.2f}% – {g['hi']:.2f}%",
                           f"{a:.2f}%","ODD (gap closed)" if odd else "even"])
                for c in range(1,len(h)+1):
                    sc(ws.cell(row=r,column=c),fill=LGREY if (r%2==0) else None,
                       align="left" if c==1 else "center",bold=(c in (2,5,7)))
                if odd: sc(ws.cell(row=r,column=9),fill=AMBER,bold=True)
                sc(ws.cell(row=r,column=7),fill=GREEN,bold=True)
                r+=1
    for c,w in enumerate([20,9,13,15,14,11,20,12,17],start=1):
        ws.column_dimensions[get_column_letter(c)].width=w
    ws.auto_filter.ref=f"A1:I{r-1}"

    # ---- Sheet 2: ladder per strain + gap report ----
    ws2=wb.create_sheet("Ladder per Strain")
    h2=["Strain","Grade","Nominal","Tolerance","Range","Batches","Assays",
        "Boundary to grade below"]
    ws2.append(h2); hdr(ws2,len(h2)); ws2.freeze_panes="A2"
    r=2; gapcount=0; oddcount=0
    for st in sorted(solved,key=str.lower):
        gl=solved[st]["grades"]
        for i,g in enumerate(gl):
            below = gl[i+1] if i+1<len(gl) else None
            if below is None: note="lowest grade in strain"
            else:
                d=g["lo"]-(below["hi"]+0.01)
                note=("exact — no gap" if abs(d)<1e-9
                      else f"GAP {d:.2f} pp ({below['hi']+0.01:.2f}–{g['lo']-0.01:.2f})")
                if abs(d)>=1e-9: gapcount+=1
            odd=int(g["nom"])%2==1
            if odd: oddcount+=1
            ws2.append([st,f"Grade {g['grade']}",
                        f"{g['nom']:.2f}%"+(" (odd)" if odd else ""),
                        f"± {g['tol']:.2f}",f"{g['lo']:.2f}% – {g['hi']:.2f}%",
                        ", ".join(pp for _,pp,_ in g["members"]),
                        ", ".join(f"{a:.2f}" for _,_,a in sorted(g["members"],key=lambda m:-m[2])),
                        note])
            for c in range(1,len(h2)+1):
                sc(ws2.cell(row=r,column=c),fill=LGREY if r%2==0 else None,
                   align="left" if c in (1,6,7,8) else "center",bold=(c in (2,5)))
            if odd: sc(ws2.cell(row=r,column=3),fill=AMBER,bold=True)
            if "GAP" in note: sc(ws2.cell(row=r,column=8),fill=ROSE,align="left")
            elif below is not None: sc(ws2.cell(row=r,column=8),fill=GREEN,align="left")
            r+=1
    for c,w in enumerate([20,9,15,11,20,26,24,34],start=1):
        ws2.column_dimensions[get_column_letter(c)].width=w

    # ---- Sheet 3: rules ----
    ws3=wb.create_sheet("Rules Applied"); ws3.append(["#","Rule / outcome"]); hdr(ws3,2)
    ng=sum(len(S["grades"]) for S in solved.values())
    txt=[
     ("R1","Band = [nominal − tolerance, nominal + tolerance − 0.01]; the upper limit is exclusive so "
           "consecutive bands butt together. Per the worked example: 20.00% ±2.0 gives 18.00% – 21.99%."),
     ("R2","NO EMPTY GRADES — every grade holds at least one batch. Bridge/placeholder classes are gone; "
           f"the ladder is built from the actual THC distribution in each strain ({ng} grades over 48 batches)."),
     ("R3","NO GAPS — with a symmetric ±t, two nominals tile exactly when t_i + t_(i+1) equals their spacing. "
           "Holding tolerance constant across a strain, the ladder must have UNIFORM SPACING s with t = s/2."),
     ("R4","Nominals are whole numbers, EVEN preferred. Where an even-only ladder cannot close a gap, one "
           "nominal is pushed onto an ODD whole number so the ladder tiles — the explicit trade allowed: "
           f"either leave a gap with even nominals, or shift to odd and close it. {oddcount} odd nominal(s) used."),
     ("R5","Tolerance never exceeds 10% of its own nominal."),
     ("R6","Where the cap makes tiling impossible — the low-THC end, and genuinely bimodal strains — the "
           f"ladder is split and the gap left open with even nominals either side. {gapcount} such gap(s)."),
     ("R7","Grade numbering runs within the strain: strongest = Grade I, descending."),
     ("R8","Search: every uniform ladder (integer spacing, integer base) is tested; valid ones are ranked by "
           "(count of odd nominals, |s − 2|, −number of grades), so an all-even ladder always beats an odd "
           "one and the house's existing 2 pp band width is preferred. Where none is valid the strain is "
           "split at its widest gap and each part solved independently."),
     ("R9","All 48 batches sit inside their grade. No grade is empty and no two bands overlap anywhere."),
    ]
    for i,(k,t) in enumerate(txt,start=1):
        ws3.append([k,t]); sc(ws3.cell(row=i+1,column=1),bold=True)
        sc(ws3.cell(row=i+1,column=2),align="left"); ws3.row_dimensions[i+1].height=50
    ws3.column_dimensions["A"].width=6; ws3.column_dimensions["B"].width=120

    wb.save(OUT)

    print(f"{len(solved)} strains | {ng} grades | 48 batches | "
          f"{oddcount} odd nominal(s) | {gapcount} open gap(s)\n")
    for st in sorted(solved,key=str.lower):
        gl=solved[st]["grades"]
        print(f"{st}")
        for i,g in enumerate(gl):
            odd=" ODD" if int(g['nom'])%2==1 else ""
            tag=",".join(pp for _,pp,_ in g["members"])
            note=""
            if i+1<len(gl):
                d=g["lo"]-(gl[i+1]["hi"]+0.01)
                if abs(d)>=1e-9: note=f"   <-- GAP {d:.2f}pp"
            print(f"   Grade {g['grade']:<4} nom {g['nom']:5.2f}%{odd:4s} ±{g['tol']:.2f}  "
                  f"[{g['lo']:6.2f}–{g['hi']:6.2f}]  {tag}{note}")
    print("\nwrote",OUT)


if __name__=="__main__":
    main()
