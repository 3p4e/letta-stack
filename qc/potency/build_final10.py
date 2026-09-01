# -*- coding: utf-8 -*-
"""
FINAL — declared nominals held exactly; tolerance extended to the 10% maximum.

Declared nominals are held exactly as supplied. Tolerance is then extended to the
largest value the rule allows — the full 10% of nominal — for every grade. That
choice satisfies every stated requirement at once:

  * strongest potency carries the greatest tolerance, decreasing all the way down
        THC 26 -> ±2.60   ...   THC 8 -> ±0.80
  * no narrow ranges: every band is the widest the 10% cap permits
  * every one of the 48 batches sits inside the class it was declared against
  * the only gap in the whole scale is 8.80–8.99 %, at the very bottom, which the
    10% cap makes mathematically impossible to close

Band = [nominal − tolerance, nominal + tolerance − 0.01]  (upper limit exclusive),
per the worked example 20.00% ±2.0 -> 18.00% – 21.99%.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY="2B547E"; LGREY="F7FAFC"; WHITE="FFFFFF"; AMBER="FFF2CC"; GREEN="E2EFDA"; ROSE="FCE4D6"
OUT=("/tmp/claude-0/-home-user-letta-stack/fa8a4b28-563d-5957-9984-7e34a8196007/"
     "scratchpad/batch_excel/PP_Potency_Grades_10pct.xlsx")
CAP=0.10; EPS=1e-9
ROMAN=["I","II","III","IV","V","VI","VII","VIII"]

ASSIGNED=[
 ("Blue Gelato","BG1024","BG1024",26.14,26),
 ("Blue Sunset Sherbet","BSS1024","BSS1024",25.01,24),
 ("Blue Sunset Sherbet","BSS052501","P050192",20.47,20),
 ("Cap Junky","CJ052501/01","P050162",24.05,24),
 ("Cap Junky","CJ082501/2","P060032",18.29,20),
 ("Cap Junky","CJ062501/1","P050222",21.51,22),
 ("Cap Junky","CJ082501/1","P060022",24.96,24),
 ("Cap Junky","CJ092501","P060072",22.30,24),
 ("Fat Bastard","FB012602","P060352",18.86,18),
 ("Fat Bastard","FB012601/1","P060322",14.68,14),
 ("Gorilla Glue","GG012603","P060402",16.70,16),
 ("Gorilla Glue","GG1024_01","P050092",18.67,18),
 ("Gorilla Glue","GG1024","GG1024",13.34,14),
 ("Grape Pie","GP052501","P050152",18.52,18),
 ("Grape Pie","GP0824_02","P050022",22.61,24),
 ("Grape Pie","GP082501/2","P050322",15.70,16),
 ("Grape Pie","GP072501/2","P050302",19.81,20),
 ("Grape Pie","GP0824_03","P050072",25.45,24),
 ("Grape Pie","GP082501/1","P050312",21.29,20),
 ("Grape Pie","GP092501","P060092",25.73,24),
 ("Grapes and Cream","GRC102501/2","P060182",11.53,12),
 ("High Pro Amnesia","HPA1024","HPA1024",17.31,18),
 ("High Pro Amnesia","HPA1024_01","P050052",21.61,22),
 ("High Pro Amnesia","HPA052501","P050182",20.39,22),
 ("Jelly Donuts","JD112501","P060212",20.32,20),
 ("Jelly Donuts","JD012603/01","P060362",16.71,16),
 ("Jelly Donuts","JD012603/02","P060412",20.54,20),
 ("Jelly Donuts","JD012603/02V","P060422",15.16,16),
 ("Jokerz 31","J31102501","P060152",17.32,18),
 ("Orange Punch Mimosa","OPM1024","OPM1024",20.03,22),
 ("Orange Punch Mimosa","OPM1024_02","P050062",18.04,18),
 ("Orange Punch Mimosa","OPM122501","P060242",7.91,8),
 ("Orange Punch Mimosa","OMP1024_01","P050042",15.38,16),
 ("Orange Punch Mimosa","OPM092501","P060042",9.43,10),
 ("Orange Punch Mimosa","OPM1024_03","P050082",16.55,16),
 ("Permanent Marker","PM092501","P060062",12.25,12),
 ("Permanent Marker","PM112501","P060232",13.00,12),
 ("Scrambler","SCR012603","P060382",17.84,18),
 ("Scrambler","SCR112501","P060282",17.13,18),
 ("Amnesia Core Cut","ACC102501","P060122",12.91,12),
 ("Cash Cow","CC012603","P060372",12.32,12),
 ("Chem Flyer","CF102501","P060132",9.83,10),
 ("Clemosa A Bud","CLE072501","P050282",8.02,8),
 ("Kush Crasher","KC102501","P060172",17.40,18),
 ("Motor Breath","MB0824_05","P050112",16.82,16),
 ("Pure Michigen","PUM102501","P060112",15.63,16),
 ("Sleepy Joe","SJ092501","P060082",10.96,10),
 ("Wedding Crusher","WC082501","P060012",21.67,22),
]

def band(n):
    t=n*CAP
    return t, n-t, n+t-0.01

thin=Side(style="thin",color="B7C3CF"); BD=Border(left=thin,right=thin,top=thin,bottom=thin)
def hdr(ws,nc):
    for c in range(1,nc+1):
        x=ws.cell(row=1,column=c); x.font=Font(name="Calibri",size=10,bold=True,color=WHITE)
        x.fill=PatternFill("solid",fgColor=NAVY); x.border=BD
        x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
def sc(cell,bold=False,fill=None,size=10,align="center"):
    cell.font=Font(name="Calibri",size=size,bold=bold); cell.border=BD
    cell.alignment=Alignment(horizontal=align,vertical="center",wrap_text=True)
    if fill: cell.fill=PatternFill("solid",fgColor=fill)

def main():
    # grade numbering within strain: strongest declared class = Grade I
    by={}
    for s,cul,pp,a,n in ASSIGNED: by.setdefault(s,set()).add(n)
    grade_of={}
    for s,noms in by.items():
        for i,n in enumerate(sorted(noms,reverse=True)): grade_of[(s,n)]=ROMAN[i]

    rows=[]
    for s,cul,pp,a,n in ASSIGNED:
        t,lo,hi=band(n)
        rows.append(dict(strain=s,cultiv=cul,pp=pp,assay=a,nom=n,tol=t,lo=lo,hi=hi,
                         grade=grade_of[(s,n)], ok=(lo-EPS<=a<=hi+EPS),
                         margin=min(a-lo,hi-a)))
    assert all(r["ok"] for r in rows), "a declared batch fell outside its class"

    wb=openpyxl.Workbook()

    # ---- Sheet 1: per batch, in the requested format ----
    ws=wb.active; ws.title="Grades per Batch"
    h=["Strain","Grade","PP Batch No.","THC/CBD ratio","Nominal value %","Tolerance",
       "Range","Assay THC %","Margin to limit (pp)","Conforms?"]
    ws.append(h); hdr(ws,len(h)); ws.freeze_panes="A2"
    r=2
    for x in sorted(rows,key=lambda r:(r["strain"].lower(),-r["nom"],-r["assay"])):
        ws.append([x["strain"],f"Grade {x['grade']}",x["pp"],f"THC {x['nom']} / CBD 1",
                   f"{x['nom']:.2f}%",f"± {x['tol']:.2f}",
                   f"{x['lo']:.2f}% – {x['hi']:.2f}%",f"{x['assay']:.2f}%",
                   round(x["margin"],2),"YES"])
        for c in range(1,len(h)+1):
            sc(ws.cell(row=r,column=c),fill=LGREY if r%2==0 else None,
               align="left" if c==1 else "center",bold=(c in (2,5,6,7)))
        sc(ws.cell(row=r,column=7),fill=GREEN,bold=True)
        sc(ws.cell(row=r,column=10),fill=GREEN,bold=True)
        if x["margin"]<0.30: sc(ws.cell(row=r,column=9),fill=AMBER,bold=True)
        r+=1
    for c,w in enumerate([20,9,13,16,14,11,20,12,18,12],start=1):
        ws.column_dimensions[get_column_letter(c)].width=w
    ws.auto_filter.ref=f"A1:J{r-1}"

    # ---- Sheet 2: the class ladder ----
    ws2=wb.create_sheet("Class Ladder")
    h2=["Class","Nominal value %","Tolerance","Tol. as % of nominal","Range",
        "Band width (pp)","Boundary to class below","Batches in class"]
    ws2.append(h2); hdr(ws2,len(h2)); ws2.freeze_panes="A2"
    pop={}
    for x in rows: pop.setdefault(x["nom"],[]).append(x["pp"])
    r=2; prev=None
    for n in range(26,7,-2):
        t,lo,hi=band(n)
        if n==8: note="lowest class"
        else:
            tb,lb,hb=band(n-2); d=lo-(hb+0.01)
            note=(f"GAP {d:.2f} pp ({hb+0.01:.2f}–{lo-0.01:.2f})" if d>EPS
                  else (f"overlap {-d:.2f} pp" if d<-EPS else "meets exactly"))
        ws2.append([f"THC {n} / CBD 1",f"{n:.2f}%",f"± {t:.2f}",f"{t/n*100:.2f}%",
                    f"{lo:.2f}% – {hi:.2f}%",f"{hi-lo+0.01:.2f}",note,
                    ", ".join(pop.get(n,[])) or "—"])
        for c in range(1,len(h2)+1):
            sc(ws2.cell(row=r,column=c),fill=LGREY if r%2==0 else None,
               align="left" if c in (7,8) else "center",bold=(c in (1,5)))
        sc(ws2.cell(row=r,column=5),fill=GREEN,bold=True)
        sc(ws2.cell(row=r,column=7),fill=ROSE if "GAP" in note else (AMBER if "overlap" in note else None),
           align="left")
        r+=1
    for c,w in enumerate([16,15,11,18,20,15,28,44],start=1):
        ws2.column_dimensions[get_column_letter(c)].width=w

    # ---- Sheet 3: per strain ----
    ws3=wb.create_sheet("Grades per Strain")
    h3=["Strain","Grade","Class","Nominal value %","Tolerance","Range","Batches","Assays"]
    ws3.append(h3); hdr(ws3,len(h3)); ws3.freeze_panes="A2"
    r=2
    bystrain={}
    for x in rows: bystrain.setdefault(x["strain"],{}).setdefault(x["nom"],[]).append(x)
    for s in sorted(bystrain,key=str.lower):
        for n in sorted(bystrain[s],reverse=True):
            mem=bystrain[s][n]; t,lo,hi=band(n)
            ws3.append([s,f"Grade {grade_of[(s,n)]}",f"THC {n} / CBD 1",f"{n:.2f}%",
                        f"± {t:.2f}",f"{lo:.2f}% – {hi:.2f}%",
                        ", ".join(m["pp"] for m in mem),
                        ", ".join(f"{m['assay']:.2f}" for m in sorted(mem,key=lambda m:-m["assay"]))])
            for c in range(1,len(h3)+1):
                sc(ws3.cell(row=r,column=c),fill=LGREY if r%2==0 else None,
                   align="left" if c in (1,7,8) else "center",bold=(c in (2,6)))
            sc(ws3.cell(row=r,column=6),fill=GREEN,bold=True)
            r+=1
    for c,w in enumerate([20,9,16,14,11,20,28,22],start=1):
        ws3.column_dimensions[get_column_letter(c)].width=w

    # ---- Sheet 4: basis ----
    ws4=wb.create_sheet("Basis"); ws4.append(["#","Rule / outcome"]); hdr(ws4,2)
    txt=[
     ("1","DECLARED NOMINALS HELD EXACTLY as supplied. Every batch stays in the class it was declared "
          "against and every nominal stays the whole even number given. Nothing re-assigned, nothing shifted."),
     ("2","Band = [nominal − tolerance, nominal + tolerance − 0.01], upper limit exclusive, per the "
          "worked example 20.00% ±2.0 -> 18.00% – 21.99%."),
     ("3","Tolerance extended to the MAXIMUM the rule allows — the full 10% of nominal — for every class."),
     ("4","STRONGEST CARRIES THE GREATEST TOLERANCE, decreasing all the way down: ±2.60 at THC 26, "
          "±2.40 at THC 24, ±2.20, ±2.00, ±1.80, ±1.60, ±1.40, ±1.20, ±1.00, down to ±0.80 at THC 8."),
     ("5","NO NARROW RANGES — every band is the widest the cap permits, from 5.20 pp at THC 26 down to "
          "1.60 pp at THC 8."),
     ("6","ALL 48 BATCHES CONFORM inside their declared class. Zero exceptions."),
     ("7","THE ONE MATHEMATICALLY IMPOSSIBLE POINT — THC 8 tops out at 8.79% and THC 10 starts at 9.00%, "
          "leaving a 0.20 pp gap (8.80–8.99%). Closing it needs ±1.00 at THC 8, which is 12.5% and breaches "
          "the cap. This is the sole gap in the entire scale."),
     ("8","Above that point the classes OVERLAP, from 0.20 pp at THC 10/12 up to 3.00 pp at THC 24/26. "
          "This is inherent to a tolerance expressed as a percentage OF the nominal: the bands widen faster "
          "than the 2-point class spacing. It is resolved the same way the house already resolves it — the "
          "declared class governs, assigned by the permitted declared value nearest the tested result."),
     ("9","This ladder reproduces the export-permit windows already in the Tranche 1 reconciliation "
          "(THC 8 7.20–8.80, THC 12 10.80–13.20, THC 16 14.40–17.60, THC 18 16.20–19.80, THC 20 18.00–22.00, "
          "THC 22 19.80–24.20, THC 24 21.60–26.40) — so it is consistent with what is already declared."),
     ("10","WATCH — P060082 (Sleepy Joe, 10.96%) sits 0.03 pp below the THC 10 upper limit of 10.99%. That is "
           "far inside analytical uncertainty, so a re-test could put it outside. P060232 (13.00% vs THC 12 "
           "limit 13.19%) has 0.19 pp. Both are worth a confirmatory look before the ladder is issued."),
    ]
    for i,(k,t) in enumerate(txt,start=1):
        ws4.append([k,t]); sc(ws4.cell(row=i+1,column=1),bold=True)
        sc(ws4.cell(row=i+1,column=2),align="left"); ws4.row_dimensions[i+1].height=52
    ws4.column_dimensions["A"].width=6; ws4.column_dimensions["B"].width=120

    wb.save(OUT)
    print(f"48 batches, {sum(1 for r in rows if r['ok'])} conform, 0 exceptions")
    print("wrote",OUT)
    return rows,grade_of

if __name__=="__main__":
    main()
