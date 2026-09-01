# -*- coding: utf-8 -*-
"""
Purely Plant — cultivation (CU) batch <-> production (P) batch cross-reference.

Authoritative source: the 87-row PP manufacturing/packaging register
(cultivation batch | P batch | date of manufacture | date of packaging),
cross-checked against the eCoA master table
(ingestion/coa_track/letta-imb-coas/exports/master_coa_table.tsv).
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY="2B547E"; LBL="EDF2F7"; LGREY="F7FAFC"; WHITE="FFFFFF"
AMBER="FFF2CC"; GREEN="E2EFDA"; ROSE="FCE4D6"
OUT="/home/user/letta-stack/qc/batch_xref/PP_CU_P_Batch_CrossReference.xlsx"

# seq, cultivation batch (as written in the register), P batch, manuf, packaging
REG = [
(1,"CJ1024 - R&D","CJ1024","20.01.2025","03.03.2025"),
(2,"GG1024 - R&D","GG1024","20.01.2025","11.03.2025"),
(3,"OPM1024 - R&D","OPM1024","21.01.2025","24/25.04.2025"),
(4,"BG1024 - R&D","BG1024","21.01.2025","07.03.2025"),
(5,"BSS1024 - R&D","BSS1024","22.01.2025","04.03.2025"),
(6,"HPA1024 - R&D","HPA1024","22.01.2025","11.04.2025"),
(7,"GG1024_01","P050092","17/18.03.2025","31.07-05.08.2025"),
(8,"GP0824_01","P050102","20/21.03.2025","05-07.08.2025"),
(9,"OMP1024_01","P050042","26/27.03.2025","20.06.2025"),
(10,"MB0824_04","P050032","02/03.04.2025","16-19.06.2025"),
(11,"GG1024_02","P050012","14/15.04.2025","11-16.06.2025"),
(12,"GP0824_02","P050022","22/23.04.2025","03-11.06.2025"),
(13,"HPA1024_01","P050052","27/28.05.2025","27.06-04.07.2025"),
(14,"OPM1024_02","P050062","29/30.05.2025","04.07-10.07.2025"),
(15,"BSS1024_01/1","P050122","04/05.06.2025","01.09-09.09.2025"),
(16,"BSS1024_01/2","P050142","","30.09.2025"),
(17,"GP0824_03","P050072","09/10.06.2025","14-17.07.2025"),
(18,"OPM1024_03","P050082","01/02.07.2025","22-30.07.2025"),
(19,"MB0824_05","P050112","08/09.07.2025","25-31.08.2025"),
(20,"OPM052501","P050132","11-13.08.2025","09-15.09.2025"),
(21,"GP052501","P050152","","25.09.2025"),
(22,"CJ052501/01","P050162","20.08.2025","30.09-01.10.2025"),
(23,"CJ052501/02","P050172","20.08.2025","01-02.10.2025"),
(24,"HPA052501","P050182","21-23.08.2025","03-09.10.2025"),
(25,"BSS052501","P050192","01-03.09.2025","10-21.10.2025"),
(26,"GP062501","P050202","15-17.09.2025","30.10-07.11.2025"),
(27,"SC062501","P050242","17.09.2025","27.11.2025"),
(28,"GOG062501","P050232","17.09.2025","27.11.2025"),
(29,"CJ062501/1","P050222","23.09.2025","24-25.11.2025"),
(30,"CJ062501/2","P050212","23-24.09.2025","18-24.11.2025"),
(31,"CJ072501","P050252","27.10.2025","27-28.11.2025"),
(32,"PM072501","P050272","27.10.2025","02-03.12.2025"),
(33,"WC072501","P050262","27-28.10.2025","01.12.2025"),
(34,"CLE072501","P050282","28.10.2025","03-04.12.2025"),
(35,"GP072501/1","P050292","11-13.11.2025","15-21.12.2025"),
(36,"GP072501/2","P050302","13.11.2025","22.12.2025"),
(37,"GP082501/1","P050312","19-20.11.2025","23-26.12.2025"),
(38,"GP082501/2","P050322","20.11.2025","26.12.2025"),
(39,"WC082501","P060012","09.12.2025","13-14.01.2026"),
(40,"CJ082501/1","P060022","09-10.12.2025","16.01.2026"),
(41,"CJ082501/2","P060032","10.12.2025","29.01.2026"),
(42,"PM092501","P060062","11.12.2025","30.01.2026"),
(43,"AB092501","P060052","11.12.2025","30.01.2026"),
(44,"SJ0925021","P060082","11-12.12.2025","02.02.2026"),
(45,"CJ092501","P060072","12.12.2025","02.02.2026"),
(46,"OPM092501","P060042","12.12.2025","29.01.2026"),
(47,"GP092501","P060092","16-17.12.2025","03-09.02.2026"),
(48,"ACC102501","P060122","09.01.2026","18-19.02.2026"),
(49,"PUM102501","P060112","08-09.01.2026","17.02.2026"),
(50,"GRC102501/1","P060142","09.01.2026","23.02.2026"),
(51,"WED102501","P060102","08.01.2026","16.02.2026"),
(52,"CF102501","P060132","09.01.2026","19-20.02.2026"),
(53,"GRC102501/2","P060182","20.01.2026","14.03.2026"),
(54,"SJ102501","P060162","20.01.2026","09.03.2026"),
(55,"KC102501","P060172","20-21.01.2026","12.03.2026"),
(56,"J31102501","P060152","21.01.2026","05.03.2026"),
(57,"J31112501","P060202","26.01.2026","17.04.2026"),
(58,"SJ112501","P060192","26-27.01.2026","31.03.2026"),
(59,"OPM112501","P060222","05.02.2026","24.04.2026"),
(60,"PM112501","P060232","05.02.2026","27.04.2026"),
(61,"JD112501","P060212","25.03.2026","21.04.2026"),
(62,"OPM122501","P060242","11.03.2026","28.04.2026"),
(63,"GG112501","P060252","25.03.2026","05.05.2026"),
(64,"J31122501","P060262","04.03.2026","14.05.2026"),
(65,"CC112501","P060272","25.03.2026","18.05.2026"),
(66,"SCR112501","P060282","25.03.2026","18.05.2026"),
(67,"FB112501","P060292","25.03.2026","19.05.2026"),
(68,"GG012601*","P060302","27.03.2026","20.05.2026"),
(69,"JD012601*","P060312","27.03.2026","21.05.2026"),
(70,"FB012601/1","P060322","27.03.2026","22.05.2026"),
(71,"CC012601/1","P060332","16.04.2026","22.05.2026"),
(72,"SCR012601*","P060342","16.04.2026","25.05.2026"),
(73,"FB012602*","P060352","16.04.2026","23.05.2026"),
(74,"JD012603/01","P060362","22.04.2026","23.05.2026"),
(75,"CC012603","P060372","22.04.2026","24.05.2026"),
(76,"SCR012603","P060382","22.04.2026","25.05.2026"),
(77,"FB012603V","P060392","29.04.2026","23.05.2026"),
(78,"GG012603","P060402","04.05.2026","24.05.2026"),
(79,"JD012603/02","P060412","04.05.2026","24.05.2026"),
(80,"JD012603/02V","P060422","29.04.2026","23.05.2026"),
(81,"FB012603","P060432","04.05.2026","15.06.2026"),
(82,"SCR022601","P060442","13.05.2026","18.06.2026"),
(83,"FB032601","P060452","02.06.2026","29.06.2026"),
(84,"GG032601","P060462","16.06.2026","16.07.2026"),
(85,"JD032601","P060472","24.06.2026","21.07.2026"),
(86,"JD022601","P060482","18.05.2026","05.08.2026"),
(87,"JD042601","P060492","21.07.2026","13.08.2026"),
(88,"","P160012","",""),
(89,"","P160022","",""),
(90,"","P160032","",""),
]

STRAIN = {
 "AB":"Apples & Bananas","ACC":"Amnesia Core Cut","BG":"Blue Gelato",
 "BSS":"Blue Sunset Sherbet","CC":"Cash Cow","CF":"Chem Flyer","CJ":"Cap Junky",
 "CLE":"Clemosa A Bud","FB":"Fat Bastard","GG":"Gorilla Glue","GP":"Grape Pie",
 "GRC":"Grapes and Cream","HPA":"High Pro Amnesia","J31":"Jokerz 31",
 "JD":"Jelly Donuts","KC":"Kush Crasher","MB":"Motor Breath",
 "OPM":"Orange Punch Mimosa","OMP":"Orange Punch Mimosa","PM":"Permanent Marker",
 "PUM":"Pure Michigen","SCR":"Scrambler","SJ":"Sleepy Joe",
 "WC":"Wedding Crusher","WED":"Wedding Cake",
}
def strain_of(cu, p):
    if not cu:
        return "Grape Pie" if p.startswith("P16") else ""
    base = cu.split(" ")[0].replace("*","")
    for n in (3,2,1):
        pre = base[:n]
        if pre in STRAIN and (len(base) == n or not base[n].isalpha()):
            return STRAIN[pre]
    for pre in sorted(STRAIN, key=len, reverse=True):
        if base.startswith(pre): return STRAIN[pre]
    return ""

def clean_cu(cu):
    return cu.replace(" - R&D","").strip()

# ---------------- what the user asked for ----------------
ASK_CU = ["BG1024","BSS1024","CC112501","CJ1024","FB012601","FB012603","FB012603V",
 "FB032601","FB112501","GG012601*","GG012603","GG032601","GG1024","GG112501",
 "GRC102501","HPA1024","J31102501","J31112501","J31122501","JD012601*","JD012603",
 "JD022601","JD112501","JD112501*","KC102501","OPM1024","OPM112501","OPM122501",
 "PM112501","SCR022601","SCR112501","SJ102501","SJ112501"]

ASK_P = ["P050012","P050022","P050032","P050042","P050052","P050062","P050072",
 "P050082","P050092","P050102","P050112","P050122","P050132","P050152","P050162",
 "P050172","P050182","P050192","P050202","P050212","P050222","P050252","P050262",
 "P050272","P050282","P050292","P050302","P050312","P050322","P060012","P060022",
 "P060032","P060042","P060052","P060062","P060072","P060082","P060092","P060152",
 "P060212","P060242","P060332","P060352","P060382","P060402","P160012","P160022",
 "P160032"]

by_cu, by_p = {}, {}
for seq, cu, p, mf, pk in REG:
    c = clean_cu(cu)
    if c: by_cu.setdefault(c, []).append((seq, c, p, mf, pk))
    by_p[p] = (seq, c, p, mf, pk)

def resolve_cu(q):
    """Return list of matching register rows for a queried cultivation code."""
    if q in by_cu: return by_cu[q], "exact"
    # sub-lot roll-up: FB012601 -> FB012601/1 ; JD012603 -> /01 /02 /02V
    subs = [r for c, rows in by_cu.items() for r in rows
            if c.startswith(q) and len(c) > len(q) and c[len(q)] in "/_"]
    if subs: return sorted(subs, key=lambda r: r[0]), "sub-lots"
    return [], "none"

NOTE_CU = {
 "BG1024":"R&D batch — no separate P-number issued; the cultivation code is the batch identifier.",
 "BSS1024":"R&D batch — no separate P-number issued; the cultivation code is the batch identifier.",
 "CJ1024":"R&D batch — no separate P-number issued; the cultivation code is the batch identifier.",
 "GG1024":"R&D batch — no separate P-number issued; the cultivation code is the batch identifier.",
 "HPA1024":"R&D batch — no separate P-number issued; the cultivation code is the batch identifier.",
 "OPM1024":"R&D batch — no separate P-number issued; the cultivation code is the batch identifier.",
 "FB012601":"Register carries the sub-lot code FB012601/1.",
 "GRC102501":"Two sub-lots packed under separate P-numbers.",
 "JD012603":"Three sub-lots packed under separate P-numbers (/01, /02, /02V).",
 "J31122501":"One P-number covers both the hand-trimmed and machine-trimmed presentations.",
 "JD112501*":"OPEN — no P-number on record in the manufacturing register or the eCoA master table. "
             "JD112501 (unstarred) is P060212; confirm whether the starred entry is a separate lot or a duplicate.",
}
NOTE_P = {
 "P050042":"Register writes the cultivation code as OMP1024_01 — letter transposition of OPM1024_01.",
 "P060082":"Register writes the cultivation code as SJ0925021 — trailing-digit typo for SJ092501.",
 "P160012":"OPEN — no cultivation code on record. eCoA master table records strain Grape Pie; batch identified by P-number only.",
 "P160022":"OPEN — no cultivation code on record. eCoA master table records strain Grape Pie; batch identified by P-number only.",
 "P160032":"OPEN — no cultivation code on record. eCoA master table records strain Grape Pie; batch identified by P-number only.",
}

DISCREP = [
 ("Letter transposition","P050042","Register: OMP1024_01 · correct: OPM1024_01",
  "Same batch, mis-keyed strain abbreviation. Correct in the register."),
 ("Digit typo","P060082","Register: SJ0925021 · correct: SJ092501",
  "Trailing '1' appended. Correct in the register."),
 ("Separator inconsistency","P050162 / P050172 / P060412 / P060422",
  "CJ052501/01 vs CJ052501/1 · JD012603/02 vs JD012603/2",
  "Manufacturing register and eCoA master table write sub-lot suffixes differently "
  "(zero-padded vs not). Pick one convention before the codes are used as keys."),
 ("eCoA master table incomplete","P060432 / P060442 / P060452 / P060462 / P060482",
  "FB012603, SCR022601, FB032601, GG032601, JD022601 carry a blank P-Number in "
  "exports/master_coa_table.tsv and PP_Coverage_By_Batch.tsv",
  "The manufacturing register does assign P-numbers to all five. Back-fill the "
  "P-Number column in the CoA ingestion so certificates key correctly."),
 ("Batch missing from eCoA extract","P060102 / P060112 / P060122 / P060132 / P060142 / P060342 / P060362 / P060372",
  "WED102501, PUM102501, ACC102501, CF102501, GRC102501/1, SCR012601*, JD012603/01, CC012603",
  "Present in the manufacturing register and in the potency ledger, absent from the "
  "77-row CoA coverage export. Confirm whether certificates exist for them."),
 ("No cultivation code","P160012 / P160022 / P160032",
  "Batch identifier is the P-number itself",
  "Strain recorded as Grape Pie in the eCoA master table. Supply the cultivation "
  "batch code, or record formally that these were packed without one."),
 ("No P-number","JD112501*",
  "Starred variant of JD112501 (= P060212)",
  "Not in the manufacturing register or the eCoA master table. Confirm whether it is "
  "a distinct lot awaiting a P-number, or a duplicate line."),
 ("Strain-name variants","(all sources)",
  "Cup Junky / Cap Junky / Cap Junkie · Sleepy Joy / Sleepy Joe · Permanent Market / "
  "Permanent Marker · Graps & Creme / Grapes and Cream · Apple and Banana / Apples & Bananas",
  "The eCoA master table carries several spellings per strain. Normalise to one "
  "controlled strain name before these tables feed a specification or label."),
]

thin=Side(style="thin",color="B7C3CF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
def hdr(ws,row,n,fill=NAVY):
    for c in range(1,n+1):
        x=ws.cell(row=row,column=c)
        x.font=Font(name="Calibri",size=10,bold=True,color=WHITE)
        x.fill=PatternFill("solid",fgColor=fill)
        x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        x.border=BORDER
def cel(c,bold=False,fill=None,size=10,align="center",color="000000"):
    c.font=Font(name="Calibri",size=size,bold=bold,color=color)
    c.alignment=Alignment(horizontal=align,vertical="center",wrap_text=True)
    c.border=BORDER
    if fill: c.fill=PatternFill("solid",fgColor=fill)
def widths(ws,ws_widths):
    for i,w in enumerate(ws_widths,start=1): ws.column_dimensions[get_column_letter(i)].width=w

wb=openpyxl.Workbook()

# ---------- Sheet 1: CU -> P ----------
ws=wb.active; ws.title="CU to P"
H=["#","CU Batch No. (asked)","P Batch No. (found)","Strain","Register code",
   "Date of manufacture","Date of packaging","Status","Note"]
ws.append(H); hdr(ws,1,len(H)); ws.freeze_panes="A2"
n=0; open_cu=0
for q in ASK_CU:
    rows,kind=resolve_cu(q)
    if not rows:
        n+=1; open_cu+=1
        ws.append([n,q,"— not found —","","","","","OPEN",NOTE_CU.get(q,"")])
        for c in range(1,len(H)+1): cel(ws.cell(row=n+1,column=c),
            fill=ROSE,align="left" if c in (2,3,4,5,9) else "center")
        continue
    for seq,cu,p,mf,pk in rows:
        n+=1
        rnd = (p==cu)
        ws.append([n,q,p,strain_of(cu,p),cu,mf,pk,
                   "R&D — code is the batch id" if rnd else ("resolved" if kind=="exact" else "resolved via sub-lot"),
                   NOTE_CU.get(q,"")])
        fill = AMBER if rnd else (LGREY if n%2==0 else None)
        for c in range(1,len(H)+1):
            cel(ws.cell(row=n+1,column=c),fill=fill,
                bold=(c==3),align="left" if c in (2,3,4,5,9) else "center")
widths(ws,[5,20,18,22,16,20,20,24,60])

# ---------- Sheet 2: P -> CU ----------
ws=wb.create_sheet("P to CU")
H=["#","P Batch No. (asked)","CU Batch No. (found)","Strain",
   "Date of manufacture","Date of packaging","Status","Note"]
ws.append(H); hdr(ws,1,len(H)); ws.freeze_panes="A2"
open_p=0
for i,q in enumerate(ASK_P,start=1):
    r=by_p.get(q)
    if r is None or not r[1]:
        open_p+=1
        ws.append([i,q,"— not found —","Grape Pie" if q.startswith("P16") else "","","","OPEN",NOTE_P.get(q,"")])
        for c in range(1,len(H)+1): cel(ws.cell(row=i+1,column=c),
            fill=ROSE,align="left" if c in (2,3,4,8) else "center")
        continue
    seq,cu,p,mf,pk=r
    ws.append([i,q,cu,strain_of(cu,p),mf,pk,"resolved",NOTE_P.get(q,"")])
    fill=LGREY if i%2==0 else None
    for c in range(1,len(H)+1):
        cel(ws.cell(row=i+1,column=c),fill=fill,bold=(c==3),
            align="left" if c in (2,3,4,8) else "center")
widths(ws,[5,20,20,22,20,20,14,60])

# ---------- Sheet 3: full register ----------
ws=wb.create_sheet("Master Register")
H=["Seq","CU Batch No.","P Batch No.","Strain","Date of manufacture","Date of packaging"]
ws.append(H); hdr(ws,1,len(H)); ws.freeze_panes="A2"
for i,(seq,cu,p,mf,pk) in enumerate(REG,start=1):
    c_cu=clean_cu(cu)
    ws.append([seq,c_cu if c_cu else "— none —",p,strain_of(c_cu,p),mf,pk])
    rnd=(p==c_cu)
    fill=AMBER if rnd else (ROSE if not c_cu else (LGREY if i%2==0 else None))
    for c in range(1,len(H)+1):
        cel(ws.cell(row=i+1,column=c),fill=fill,align="left" if c in (2,3,4) else "center")
widths(ws,[6,20,16,22,22,22])
ws.auto_filter.ref=f"A1:F{len(REG)+1}"

# ---------- Sheet 4: discrepancies ----------
ws=wb.create_sheet("Open Items")
H=["#","Type","Batch(es)","What it is","Action"]
ws.append(H); hdr(ws,1,len(H)); ws.freeze_panes="A2"
for i,(t,b,w,a) in enumerate(DISCREP,start=1):
    ws.append([i,t,b,w,a])
    fill=LGREY if i%2==0 else None
    for c in range(1,len(H)+1):
        cel(ws.cell(row=i+1,column=c),fill=fill,align="left" if c>1 else "center",
            size=9,bold=(c==2))
widths(ws,[5,28,42,58,72])
for r in range(2,len(DISCREP)+2): ws.row_dimensions[r].height=58

wb.save(OUT)
print(f"saved {OUT}")
print(f"CU->P asked {len(ASK_CU)}  unresolved {open_cu}")
print(f"P->CU asked {len(ASK_P)}  unresolved {open_p}")
