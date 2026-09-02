#!/usr/bin/env python3
"""Rebuild the "CoQ Parameter Tracker" workbook sheet from the eCoA database.

Why this exists (Head of QC, 02.09.2026)
----------------------------------------
The v7 tracker was built by an Apps Script from the "eCOA Document Index" sheet,
whose PARAMETER VALUES column was itself a text export of the desk. Every place
where that export and the database disagreed rendered as the certificate code
next to the words "no result on file" - i.e. an accredited certificate shown as
silent. Audit of the 291 such cells in v7:

    44   #6 Total CBN     CNP prints CBN but no CBNA/total  -> compiler key gap (fixed in build_coq.derive_totals)
    96   #3-#8            certificate matched by CODE TEXT   -> Cyrillic/Latin code, "/26" vs "-26", or code not
                          extracted at all (42 certificates have cert_code NULL) -> looked "not in database"
    29   #4,#5,#8,#12     value IS confirmed in the database -> the Index text export said "no result on the desk"
    11   #12 Pesticides   per-compound panel rows            -> builder expected one value
    ~72  #2 Ident. B      credited to CNP, nothing on page   -> a CREDIT question, not a result question
    ~10  #8 LoD           credited to a Farmahem K cert      -> LoD is on the separate ГС/LoD certificate

Rules this builder applies
--------------------------
1. Certificates are matched to the database by FILENAME (document), never by the
   printed code - the code is what the extractor read, the filename is what QC filed.
2. Values come from the database (confidence 'ok'), through build_coq.derive_totals
   so a certificate printing CBN supplies row 6, and through the pesticide panel
   logic so a per-compound panel reads as one result.
3. A certificate is NEVER shown as "no result on file". A cell is one of:
      <value>                     - confirmed result on that certificate
      <value> ᴰ                   - derived per build_coq (note in the eCOA ref line)
      held for review             - the two reads disagreed; a human must confirm
      not on this certificate     - the certificate is credited on the tracker for
                                    this parameter but carries no such row: the CREDIT
                                    is wrong or the extraction missed the row - listed
                                    on the "Credit Audit" sheet either way
      not ingested                - the Index names a file the database does not hold
      — MISSING —                 - no certificate credited for the parameter at all
4. Stability-timepoint certificates are shown but never counted as release results.
5. Nothing here judges conformity beyond the database's own exceeds/outside flags.

Usage
-----
    python3 build_tracker_workbook.py --db ecoa.sqlite --in CoQ_Analysis_Master_v7.xlsx \
                                      --out CoQ_Analysis_Master_v8.xlsx
"""
import os, sys, re, sqlite3, argparse, datetime, unicodedata, collections
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import build_coq as CQ                                   # noqa: E402
from common.controlled import is_panel_statement, is_not_found   # noqa: E402

TRACKER_OUT = 'CoQ Parameter Tracker v8'
INDEX_SHEET = 'eCOA Document Index'
PARAMS_SHEET = 'Parameters'
AUDIT_SHEET = 'Credit Audit'

# One entry per specification parameter, in tracker order. `keys` are database
# parameter keys; groups 9/10/11 carry sub-determinations (one column each).
PARAMS = [
    dict(n=1,  group='IDENTIFICATION  1–3',    title='#1 Identification A',         method='Appearance · Ph. Eur. mon. 3028',            keys=['identification_a_macroscopic']),
    dict(n=2,  group='IDENTIFICATION  1–3',    title='#2 Identification B',         method='Microscopy · Ph. Eur. 2.8.23',               keys=['identification_b_microscopic']),
    dict(n=3,  group='IDENTIFICATION  1–3',    title='#3 Identification C',         method='HPLC · Ph. Eur. 2.2.29',                     keys=['identification_c_hplc']),
    dict(n=4,  group='CANNABINOID ASSAY  4–6', title='#4 Assay — Total Δ⁹-THC*',    method='Ph. Eur. 2.2.29 (HPLC)',                     keys=['total_thc']),
    dict(n=5,  group='CANNABINOID ASSAY  4–6', title='#5 Assay — Total CBD',        method='Ph. Eur. 2.2.29 · CBD + CBDA×0.877',          keys=['total_cbd']),
    dict(n=6,  group='CANNABINOID ASSAY  4–6', title='#6 Total CBN',                method='Ph. Eur. 2.2.29 · CBN + CBNA×0.876',          keys=['total_cbn']),
    dict(n=7,  group='PHYSICAL  7–8',          title='#7 Foreign Matter',           method='Ph. Eur. 2.8.2 / in-house',                  keys=['foreign_matter']),
    dict(n=8,  group='PHYSICAL  7–8',          title='#8 Loss on Drying',           method='Ph. Eur. 2.2.32 · 40 °C, 24 h',              keys=['loss_on_drying']),
    dict(n=9,  group='MICROBIOLOGY  9',        title='#9 Microbiological Purity',   method='Ph. Eur. 2.6.12 / 2.6.13 / 2.6.31 · cat. C',
         subs=['TAMC', 'TYMC', 'GNB', 'Salm.', 'E. coli'], det=['9.1', '9.2', '9.3', '9.4', '9.5'],
         keys=['tamc', 'tymc', 'bile_tolerant_gram_negative', 'salmonella', 'escherichia_coli']),
    dict(n=10, group='CONTAMINANTS  10–12',    title='#10 Mycotoxins',              method='Ph. Eur. 2.8.18 / 2.8.22 (HPLC-FLD)',
         subs=['AfB₁', 'ΣAf', 'OTA'], det=['10.1', '10.2', '10.3'],
         keys=['aflatoxin_b1', 'aflatoxins_total', 'ochratoxin_a']),
    dict(n=11, group='CONTAMINANTS  10–12',    title='#11 Heavy Metals',            method='Ph. Eur. 2.4.27 (ICP-MS)',
         subs=['Pb', 'Cd', 'As', 'Hg'], det=['11.1', '11.2', '11.3', '11.4'],
         keys=['lead', 'cadmium', 'arsenic', 'mercury']),
    dict(n=12, group='CONTAMINANTS  10–12',    title='#12 Pesticide Residues',      method='Ph. Eur. 2.8.13 · CUMCS equivalency',        keys=['pesticide_residues']),
]

# colours (v7 palette kept so the sheet reads the same)
C = dict(navy='1F3864', white='FFFFFF', subhdr='FFF2CC', grey='EFEFEF',
         green='C6EFCE', orange='FCE5CD', amber='FDE9D9', red='F4CCCC', blue='DDEBF7',
         greenFill='6AA84F', orangeFill='E69138', amberFill='F6B26B', redFill='CC0000', blueFill='3D85C6',
         statusRed='CC0000', statusOrange='E69138', statusGreen='38761D', oos='9C0006', held='7F6000')
THIN = Side(style='thin', color='404040'); THICK = Side(style='medium', color='000000')


def fill(h): return PatternFill('solid', fgColor=h)


def fold(s):
    if s is None: return ''
    s = unicodedata.normalize('NFKD', str(s)).lower()
    return re.sub(r'[^0-9a-zа-я]+', '', s)


# ---------------------------------------------------------------- database
SELECT = """SELECT r.parameter, r.result_printed, r.result_numeric, r.unit, r.method,
              r.date_iso, r.cert_code, r.lab, r.confidence, r.exceeds_criterion,
              r.outside_range, c.document, c.doc_id, r.parameter_printed,
              r.method_accredited, r.test_type
         FROM result r JOIN certificate c ON c.doc_id = r.doc_id"""


def load_db(dbpath):
    db = sqlite3.connect(dbpath)
    rows = CQ.derive_totals(db.execute(SELECT).fetchall())
    by_doc = collections.defaultdict(list)
    for r in rows:
        by_doc[r[11]].append(r)
    meta = {}
    for doc, code, date, lab, ttype in db.execute(
            "SELECT document, cert_code, date_iso, lab, test_type FROM certificate"):
        meta[doc] = dict(code=code, date=date, lab=lab, test_type=ttype)
    return by_doc, meta


def _pest_value(rows):
    ok = [r for r in rows if r[8] == 'ok' and r[1] not in (None, '')]
    if not ok:
        return None
    comps = [r for r in ok if not is_panel_statement(r[13] or '')]
    if comps:
        finds = [(r[13], r[1]) for r in comps if not is_not_found(r[1])]
        if finds:
            return 'FINDS: ' + '; '.join('%s %s' % f for f in finds[:4])
        return '≤ LOQ (all %d compounds)' % len(comps)
    return ok[-1][1]


def cert_values(rows):
    """{key: (value, state, note, flags)} for one certificate's database rows.

    state: 'ok' | 'held' | 'absent'. flags: set of 'oos','derived','stability'."""
    by = collections.defaultdict(list)
    for r in rows:
        by[r[0]].append(r)
    out = {}
    all_keys = {k for p in PARAMS for k in p['keys']}
    for key in all_keys:
        rs = by.get(key, [])
        if key == 'pesticide_residues':
            v = _pest_value(rs)
            if v is not None:
                out[key] = (v, 'ok', '', set())
            elif rs:
                out[key] = ('held for review', 'held', '', set())
            continue
        ok = [r for r in rs if r[8] == 'ok' and r[1] not in (None, '')]
        if ok:
            r = ok[-1]
            flags = set(); note = ''
            if r[9] == 1 or r[10] == 1:
                flags.add('oos')
            d = CQ._derivation(r)
            if d:
                flags.add('derived'); note = d['working']
            if (r[15] or '') == 'stability':
                flags.add('stability')
            out[key] = (str(r[1]), 'ok', note, flags)
        elif rs:
            out[key] = ('held for review', 'held', '', set())
    # Identification C: the CNP/Farmahem HPLC profile IS the identification by
    # HPLC (Ph. Eur. 3028 identification C = the assay chromatogram). Keep the
    # owner's convention, but say that it is inferred.
    if 'identification_c_hplc' not in out and 'total_thc' in out and out['total_thc'][1] == 'ok':
        out['identification_c_hplc'] = ('Conforms', 'ok', 'inferred from the HPLC cannabinoid profile on this certificate', {'inferred'})
    return out


# ---------------------------------------------------------------- index
def read_index(ws):
    hdr = [str(c.value or '').strip() for c in ws[1]]
    col = {h: i for i, h in enumerate(hdr)}
    need = ['CU Batch', 'Lab', 'Kind', 'Certificate', 'Date', 'Parameters covered', 'Filename']
    for n in need:
        if n not in col:
            raise SystemExit('Index sheet lacks column %r; headers: %s' % (n, hdr))
    key_col = col.get('BATCH KEY')
    credit_col = col.get('CREDITED FOR', col['Parameters covered'])
    batches, order = {}, []
    last_cu = ''
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = str(row[col['Certificate']] or '').strip()
        fn = str(row[col['Filename']] or '').strip()
        if not code and not fn:
            continue
        cu = str(row[col['CU Batch']] or '').strip() or last_cu
        last_cu = cu
        key = (str(row[key_col] or '').strip() if key_col is not None else '') or cu
        if key not in batches:
            batches[key] = dict(key=key, cu=cu, p=str(row[col.get('P Batch', col['CU Batch'])] or '').strip(), certs=[])
            order.append(key)
        d = row[col['Date']]
        date = d.strftime('%d.%m.%Y') if isinstance(d, (datetime.date, datetime.datetime)) else str(d or '').strip()
        covers = [int(x) for x in re.findall(r'#(\d+)', str(row[credit_col] or ''))]
        batches[key]['certs'].append(dict(
            lab=str(row[col['Lab']] or '').strip(), kind=str(row[col['Kind']] or '').strip(),
            code=code, date=date, sortkey=_sortkey(date), covers=covers, filename=fn))
    for b in batches.values():
        b['certs'].sort(key=lambda c: (c['sortkey'], c['code']))
    return [batches[k] for k in order]


def _sortkey(d):
    m = re.match(r'(\d{1,2})\.(\d{1,2})\.(\d{4})', d or '')
    return '%s%02d%02d' % (m.group(3), int(m.group(2)), int(m.group(1))) if m else '99999999'


def read_ac(wb):
    ac = {}
    if PARAMS_SHEET not in wb.sheetnames:
        return ac
    ws = wb[PARAMS_SHEET]
    hdr = [str(c.value or '').strip() for c in ws[1]]
    try:
        i_no = hdr.index('#'); i_ac = [i for i, h in enumerate(hdr) if 'Acceptance' in h][0]
    except (ValueError, IndexError):
        return ac
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[i_no] and row[i_ac]:
            ac[str(row[i_no]).strip()] = str(row[i_ac]).strip()
    return ac


# ---------------------------------------------------------------- matching
def match_doc(cert, by_doc, meta):
    fn = cert['filename']
    if fn in meta:
        return fn
    f = fold(fn)
    for doc in meta:
        if fold(doc) == f:
            return doc
    # last resort: same batch prefix + same code digits
    digits = re.sub(r'\D', '', cert['code'])
    if digits:
        cands = [doc for doc in meta if digits in re.sub(r'\D', '', doc.split('_', 1)[-1])
                 and fold(doc).startswith(fold(fn.split('_')[0]))]
        if len(cands) == 1:
            return cands[0]
    return None


# ---------------------------------------------------------------- layout
def layout():
    cols, col = [], 4
    for p in PARAMS:
        p['start'] = col
        if p.get('subs'):
            for i, s in enumerate(p['subs']):
                cols.append(dict(col=col, p=p, sub=s, det=p['det'][i])); col += 1
        else:
            cols.append(dict(col=col, p=p, kind='result')); col += 1
            cols.append(dict(col=col, p=p, kind='ecoa')); col += 1
        cols.append(dict(col=col, p=p, kind='check')); col += 1
        p['end'] = col - 1; p['width'] = p['end'] - p['start'] + 1
    return cols, col - 1


def write_headers(ws, ac, lastcol):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3); ws.cell(1, 1, 'BATCH IDENTIFICATION')
    g, gs = None, 0
    for i, p in enumerate(PARAMS):
        if p['group'] != g:
            if g:
                ws.merge_cells(start_row=1, start_column=gs, end_row=1, end_column=p['start'] - 1); ws.cell(1, gs, g)
            g, gs = p['group'], p['start']
        if i == len(PARAMS) - 1:
            ws.merge_cells(start_row=1, start_column=gs, end_row=1, end_column=p['end']); ws.cell(1, gs, g)
    for c, t in ((1, 'CU Batch'), (2, 'P Batch'), (3, 'STATUS')):
        ws.merge_cells(start_row=2, start_column=c, end_row=3, end_column=c); ws.cell(2, c, t)
    for p in PARAMS:
        ws.merge_cells(start_row=2, start_column=p['start'], end_row=2, end_column=p['end'])
        ws.cell(2, p['start'], p['title'] + '\n' + p['method'])
        if p.get('subs'):
            for i, s in enumerate(p['subs']):
                ws.cell(3, p['start'] + i, 'A.C.: ' + ac.get(p['det'][i], 'not on the Parameters sheet'))
                ws.cell(4, p['start'] + i, s)
        else:
            ws.merge_cells(start_row=3, start_column=p['start'], end_row=3, end_column=p['start'] + 1)
            ws.cell(3, p['start'], 'A.C.: ' + ac.get(str(p['n']), 'not on the Parameters sheet'))
            ws.cell(4, p['start'], 'Result (as reported)')
            ws.cell(4, p['start'] + 1, 'eCOA ref, (date) [Lab] — one certificate per line')
        ws.cell(4, p['end'], '✓/✗')
    for r in (1, 2):
        for c in range(1, lastcol + 1):
            ws.cell(r, c).fill = fill(C['navy']); ws.cell(r, c).font = Font(bold=True, color=C['white'], size=8)
    for r in (2, 3):
        for c in (1, 2, 3):
            ws.cell(r, c).fill = fill(C['navy']); ws.cell(r, c).font = Font(bold=True, color=C['white'], size=8)
    for c in range(4, lastcol + 1):
        ws.cell(3, c).fill = fill(C['grey']); ws.cell(3, c).font = Font(italic=True, size=6.5)
        ws.cell(4, c).fill = fill(C['subhdr']); ws.cell(4, c).font = Font(bold=True, size=7)
    for r in range(1, 5):
        for c in range(1, lastcol + 1):
            ws.cell(r, c).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def ecoa_line(c, note=''):
    s = '%s, (%s) [%s]' % (c['code'] or 'NO-DOC-CODE', c['date'] or '—', c['lab'])
    if c['kind'].lower().startswith('stab'):
        s += ' · stability timepoint'
    if note:
        s += ' — ' + note
    return s


def build(dbpath, src, dst):
    by_doc, meta = load_db(dbpath)
    wb = openpyxl.load_workbook(src)
    batches = read_index(wb[INDEX_SHEET])
    ac = read_ac(wb)
    if TRACKER_OUT in wb.sheetnames:
        del wb[TRACKER_OUT]
    ws = wb.create_sheet(TRACKER_OUT)
    cols, lastcol = layout()
    write_headers(ws, ac, lastcol)

    audit = []            # (batch, cert code, lab, filename, parameter, finding)
    tally = collections.Counter()
    row = 5
    for b in batches:
        top, bot = row, row + 1
        # resolve every certificate against the database once
        for c in b['certs']:
            doc = match_doc(c, by_doc, meta)
            c['doc'] = doc
            c['vals'] = cert_values(by_doc.get(doc, [])) if doc else {}
            c['stab'] = c['kind'].lower().startswith('stab') or (doc and (meta[doc]['test_type'] or '') == 'stability')
            if not doc:
                tally['not ingested'] += 1
        ws.merge_cells(start_row=top, start_column=1, end_row=bot, end_column=1); ws.cell(top, 1, b['cu'] or 'CU CODE NOT RECORDED — TBC')
        ws.merge_cells(start_row=top, start_column=2, end_row=bot, end_column=2)
        ws.cell(top, 2, b['p'] if b['p'] and b['p'] != b['cu'] and 'not assigned' not in b['p'] else 'N/A — no P batch assigned')
        for c in (1, 2):
            for r in (top, bot):
                ws.cell(r, c).fill = fill(C['navy']); ws.cell(r, c).font = Font(bold=True, color=C['white'], size=7)

        n_missing = n_credit = n_held = n_ingest = 0
        oos = []
        for p in PARAMS:
            lines = []          # (result text per sub or single, ecoa text, glyph, state)
            for c in b['certs']:
                credited = p['n'] in c['covers']
                if p.get('subs'):
                    vals = [c['vals'].get(k) for k in p['keys']]
                    have = [v for v in vals if v]
                    if not have and not credited:
                        continue
                    if c['doc'] is None:
                        lines.append((['not ingested'] * len(p['keys']), ecoa_line(c, 'file not in the eCoA database'), '✗', 'ingest'))
                        audit.append((b['cu'], c['code'], c['lab'], c['filename'], p['title'], 'certificate not ingested'))
                        continue
                    if not have:
                        lines.append((['not on this certificate'] * len(p['keys']), ecoa_line(c, 'credited, no such rows on the certificate'), '✗', 'credit'))
                        audit.append((b['cu'], c['code'], c['lab'], c['filename'], p['title'], 'credited but parameter not on certificate'))
                        continue
                    txt, st = [], 'ok'
                    for v in vals:
                        if v is None:
                            txt.append('n.r.')
                        else:
                            txt.append(v[0] + (' ᴰ' if 'derived' in v[3] else ''))
                            if v[1] == 'held': st = 'held'
                            if 'oos' in v[3] and not c['stab']: oos.append('%s %s' % (p['title'].split()[0], p['subs'][vals.index(v)]))
                    lines.append((txt, ecoa_line(c), '✓' if st == 'ok' else '✗', 'stab' if c['stab'] else st))
                else:
                    v = c['vals'].get(p['keys'][0])
                    if v is None and not credited:
                        continue
                    if c['doc'] is None:
                        lines.append((['not ingested'], ecoa_line(c, 'file not in the eCoA database'), '✗', 'ingest'))
                        audit.append((b['cu'], c['code'], c['lab'], c['filename'], p['title'], 'certificate not ingested'))
                        continue
                    if v is None:
                        lines.append((['not on this certificate'], ecoa_line(c, 'credited, no such row on the certificate'), '✗', 'credit'))
                        audit.append((b['cu'], c['code'], c['lab'], c['filename'], p['title'], 'credited but parameter not on certificate'))
                        continue
                    val, st, note, flags = v
                    txt = val + (' ᴰ' if 'derived' in flags else '')
                    if 'oos' in flags and not c['stab']:
                        oos.append(p['title'].split()[0])
                    lines.append(([txt], ecoa_line(c, note), '✓' if st == 'ok' else '✗', 'stab' if c['stab'] else st))
            states = [l[3] for l in lines]
            if 'ok' in states:            state = 'green'
            elif 'stab' in states:        state = 'orange'
            elif 'held' in states:        state = 'held';   n_held += 1
            elif 'credit' in states:      state = 'amber';  n_credit += 1
            elif 'ingest' in states:      state = 'blue';   n_ingest += 1
            else:                         state = 'red';    n_missing += 1
            bg = dict(green=C['green'], orange=C['orange'], held=C['amber'], amber=C['amber'], blue=C['blue'], red=C['red'])[state]
            gbg = dict(green=C['greenFill'], orange=C['orangeFill'], held=C['amberFill'], amber=C['amberFill'], blue=C['blueFill'], red=C['redFill'])[state]
            ecoa_txt = '\n'.join(l[1] for l in lines) or '— no certificate —'
            glyph_txt = '\n'.join(l[2] for l in lines) or '✗'
            if p.get('subs'):
                for i in range(len(p['subs'])):
                    cell = ws.cell(top, p['start'] + i, '\n'.join(l[0][i] for l in lines) or '— MISSING —')
                    cell.font = Font(bold=True, size=7, color=(C['oos'] if any(('#%d %s' % (p['n'], p['subs'][i])) in o for o in oos) else None))
                ws.merge_cells(start_row=bot, start_column=p['start'], end_row=bot, end_column=p['start'] + len(p['subs']) - 1)
                ws.cell(bot, p['start'], ecoa_txt).font = Font(size=6)
                for r in (top, bot):
                    for i in range(len(p['subs'])):
                        ws.cell(r, p['start'] + i).fill = fill(bg)
            else:
                ws.merge_cells(start_row=top, start_column=p['start'], end_row=bot, end_column=p['start'])
                cell = ws.cell(top, p['start'], '\n'.join(l[0][0] for l in lines) or '— MISSING —')
                cell.font = Font(bold=True, size=7, color=(C['oos'] if p['title'].split()[0] in oos else (C['held'] if state == 'held' else None)))
                ws.merge_cells(start_row=top, start_column=p['start'] + 1, end_row=bot, end_column=p['start'] + 1)
                ws.cell(top, p['start'] + 1, ecoa_txt).font = Font(size=6)
                for r in (top, bot):
                    for cc in (p['start'], p['start'] + 1):
                        ws.cell(r, cc).fill = fill(bg)
            ws.merge_cells(start_row=top, start_column=p['end'], end_row=bot, end_column=p['end'])
            g = ws.cell(top, p['end'], glyph_txt); g.fill = fill(gbg); g.font = Font(bold=True, color=C['white'], size=7)

        gaps = n_missing + n_credit + n_held + n_ingest
        if gaps == 0:
            st, colour = '✓ COMPLETE', C['statusGreen']
        else:
            parts = []
            if n_missing: parts.append('%d no cert' % n_missing)
            if n_credit:  parts.append('%d credit to confirm' % n_credit)
            if n_held:    parts.append('%d held' % n_held)
            if n_ingest:  parts.append('%d not ingested' % n_ingest)
            st = ('⚠' if gaps <= 3 else '✗') + ' %d OPEN\n(%s)' % (gaps, ' / '.join(parts))
            colour = C['statusOrange'] if gaps <= 3 else C['statusRed']
        if oos:
            st += '\n✗ OUT OF SPECIFICATION: ' + ', '.join(sorted(set(oos))); colour = C['statusRed']
        ws.merge_cells(start_row=top, start_column=3, end_row=bot, end_column=3)
        sc = ws.cell(top, 3, st); sc.fill = fill(colour); sc.font = Font(bold=True, color=C['white'], size=7)
        tally['batches'] += 1; tally['complete'] += (gaps == 0); tally['credit'] += n_credit
        tally['held'] += n_held; tally['missing'] += n_missing
        for r in (top, bot):
            for c in range(1, lastcol + 1):
                cell = ws.cell(r, c)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(left=THIN, right=THIN, top=THICK if r == top else None, bottom=THICK if r == bot else None)
        row += 2

    key = ('KEY — ✓ green: certificate on file and its confirmed result in the eCoA database. ✓ orange: stability-timepoint '
           'certificate — not a release result. ᴰ: value derived per build_coq.derive_totals (free form + acid form × factor, or the '
           'free-form figure where the acid form is not reported; the working is stated on the eCOA ref line). "held for review": the '
           'two independent reads of the page disagreed — a human confirms. "not on this certificate" (amber): the certificate is credited '
           'for the parameter but carries no such row — the credit or the extraction is wrong; every case is listed on the sheet "%s". '
           '"not ingested" (blue): the Index names a file the database does not hold. — MISSING — (red): no certificate credited. '
           'A certificate is never shown as "no result on file". RED result = exceeds its criterion per the database flags; every one needs '
           'an investigation record. Built %s by build_tracker_workbook.py from %s.'
           % (AUDIT_SHEET, datetime.date.today().strftime('%d.%m.%Y'), os.path.basename(dbpath)))
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=lastcol)
    kc = ws.cell(row + 1, 1, key); kc.font = Font(size=6.5); kc.alignment = Alignment(wrap_text=True, vertical='top'); kc.fill = fill(C['grey'])
    ws.row_dimensions[row + 1].height = 80
    ws.freeze_panes = 'D5'
    ws.column_dimensions['A'].width = 13; ws.column_dimensions['B'].width = 15; ws.column_dimensions['C'].width = 22
    for c in cols:
        w = 9
        if c.get('kind') == 'result': w = 11
        if c.get('kind') == 'ecoa': w = 30
        if c.get('kind') == 'check': w = 4
        if c.get('sub'): w = 9
        ws.column_dimensions[get_column_letter(c['col'])].width = w

    # ---- Credit Audit sheet
    if AUDIT_SHEET in wb.sheetnames:
        del wb[AUDIT_SHEET]
    wa = wb.create_sheet(AUDIT_SHEET)
    wa.append(['CU Batch', 'Certificate', 'Lab', 'Filename', 'Parameter', 'Finding', 'Action'])
    for r in audit:
        action = ('Confirm the credit against the page. If the page carries the determination, the extraction missed it: '
                  'add the row to the record. If not, remove the credit on the Index.' if 'credited' in r[5]
                  else 'File is on the Index but not in the eCoA database: ingest it (or correct the Index filename).')
        wa.append(list(r) + [action])
    for c in wa[1]:
        c.font = Font(bold=True, color=C['white']); c.fill = fill(C['navy'])
    for col, w in zip('ABCDEFG', (13, 16, 8, 44, 24, 40, 70)):
        wa.column_dimensions[col].width = w
    wa.freeze_panes = 'A2'

    # ---- Read Me note
    if 'Read Me' in wb.sheetnames:
        rm = wb['Read Me']
        rm.append([])
        rm.append(['v8', 'Sheet "%s" rebuilt from the eCoA database by build_tracker_workbook.py (letta-stack). Certificates are matched by filename, '
                        'cannabinoid totals are derived where a certificate prints components only (ᴰ), pesticide panels read as one result, '
                        'and a certificate is never shown as "no result on file": a cell is a value, "held for review", "not on this certificate" '
                        '(listed on "%s"), "not ingested", or — MISSING —.' % (TRACKER_OUT, AUDIT_SHEET)])
    wb.save(dst)
    return tally, audit


if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('--db', default=HERE + '/ecoa.sqlite')
    ap.add_argument('--in', dest='src', required=True)
    ap.add_argument('--out', dest='dst', required=True)
    a = ap.parse_args()
    tally, audit = build(a.db, a.src, a.dst)
    print('wrote %s: %d batches, %d complete; open credits %d, held %d, missing %d; audit rows %d'
          % (a.dst, tally['batches'], tally['complete'], tally['credit'], tally['held'], tally['missing'], len(audit)))
