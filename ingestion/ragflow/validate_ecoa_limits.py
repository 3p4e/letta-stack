#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Compare every recorded result against its own acceptance criterion.

Nothing in the pipeline compared a number to the limit sitting beside it, so
nothing objected when a certificate reported a mould count over the Ph. Eur.
criterion and concluded ОДГОВАРА. This is that check.

**Corrected 31.08.2026, and the correction is the point of the file.** The first
version of this script asked `magnitude()` for the limit as well as for the
result. `magnitude()` answers *what number is this measurement*; an acceptance
criterion asks *what is the largest result that still conforms*, and for a
microbial enumeration criterion written as a bare power of ten Ph. Eur. 5.1.4
says the two answers differ by a factor of two — 10⁴ CFU/g means a maximum
acceptable count of 20 000, not 10 000. Comparing against 10 000 reported nine
TYMC results as out of specification. **Five are. Four are not**, and they stood
in a release register flagged as failures.

That is this campaign's own recurring defect turned on the campaign: the right
value taken from the wrong column. See `acceptance_limit()` for the rule, the
authority and the three different multipliers that were in play — ×5 on the
register's specification row and on the Purely Plant in-house CoA form it was
copied from, ×2 in the pharmacopoeia, ×1 in this file.

Three rules, and they catch different failures:

  R1  a **release** result above its acceptance criterion is reported, whatever
      the certificate concluded. A laboratory verdict is not a substitute for
      the arithmetic. Five findings, all TYMC, all concluding ОДГОВАРА.

  R5  a result above the criterion on a sample the register's own Stability
      Testing Programme sheet lists is reported **separately and is not a
      release failure**. Four accelerated 40 °C / 75 % RH timepoints exceed the
      CBN release limit; the sample was never a release sample. Which
      certificates those are is read off the workbook, not hard-coded.

  R2  a result recorded at exactly one power of ten below the criterion's
      printed exponent is reported as SUSPECT. That is the shape the superscript
      misread produces — 4,2×10⁴ read as 4,2×10³ — and the only shape that
      silently turns a fail into a pass. Verify those against the rendered page,
      never the text layer.

One limitation to know about: R1 compares against the **register's column
criterion**, taken from the specification row. Some certificates print a tighter
criterion of their own — 1220/2171/25 states TYMC ≤ 10² where the column says
10⁴, and its result of 200 therefore fails on the paper while passing here.
Comparing against the criterion printed on each certificate is the right
long-term shape, and needs the limit captured per result at extraction time.
That is what the structured extraction in ECOA_RAG_PIPELINE_2026-08-30.md is for.

Run it against the register, or against any extraction that produces the same
columns:

    python3 ingestion/ragflow/validate_ecoa_limits.py REGISTER.xlsx

Exit status is 1 if any R1 finding exists, so it can gate a pipeline. R5 and R2
do not gate: a stability observation and a suspect notation both need a person,
not a blocked pipeline.

(R3 and R4 are numbered in `ingestion/ragflow/config/ecoa_extraction_agent.json`:
R3 is a verdict contradicting its own value, which needs the certificate's
ЗАКЛУЧОК and so belongs to extraction rather than to this register; R4 is the
Total-THC arithmetic, implemented here as `total_thc_consistent`.)
"""
import re
import sys
from collections import namedtuple
from openpyxl import load_workbook

SHEET = "Batch Release QC"
HEADER_ROW, SPEC_ROW, FIRST_DATA = 4, 5, 6

SUP = {"⁰": "0", "¹": "1", "²": "2", "³": "3", "⁴": "4", "⁵": "5", "⁶": "6"}

# Δ9-THCA decarboxylates to Δ9-THC on heating, losing the carboxyl group. The mass
# ratio of the two molecules is 314.46/358.47 = 0.877, which is why every certificate
# in this corpus prints the same conversion in its own footnote.
THCA_TO_THC = 0.877


def total_thc_consistent(d9_thc, thca, total_thc, tol=0.06):
    """R4 — check a potency certificate against its own arithmetic.

    Every CNP certificate prints all three of Δ9-THC, Δ9-THCA and Вкупно Δ9-THC,
    and states the relation between them in a footnote:

        Вкупно Δ9-THC = Δ9-THC + Δ9-THCA × 0.877

    So each certificate carries its own proof, and a single corrupted digit stops
    being invisible. Run over eCoA_DATABASE on 30.08.2026 it flagged two of the
    fifteen certificates where all three values could be read. **Both were then
    checked against the source PDF in Drive, and in both the certificate was correct
    and the corpus was wrong:**

      ППК25117  corpus total 1.58 beside 0.46 and 17.01. The page prints 15.38,
                which is what 0.46 + 17.01 x 0.877 gives, and what the register
                already held.
      ППК25139  corpus THCA 0.52 against a total of 23.79. The page prints 26.52,
                and 0.53 + 26.52 x 0.877 = 23.79 exactly. The same corpus record
                also holds "Satre Pie" for Grape Pie and "GF0824_02" for GP0824_02
                — three corruptions in one document.

    So the measured rate is **2 of 15 potency certificates carrying a value that is
    wrong rather than missing — 13%** — and nothing but the certificates' own
    footnote formula detected it. That is a second, independent measurement of the
    defect first found in the mould counts, on a different parameter and a different
    laboratory, and it says the corpus is unreliable for numbers in general and not
    merely for superscripts.

    The rule raises the page; it does not decide it. Both times the page had a clear
    answer and the laboratory was not at fault — which is the outcome to expect, and
    the reason a flag here must never be reported as a laboratory finding before
    someone has looked.

    Returns None when any input is missing (nothing to check), True when the printed
    total matches the computed one within `tol`, False when it does not.

    >>> total_thc_consistent(0.46, 17.01, 15.38)
    True
    >>> total_thc_consistent(0.46, 17.01, 1.58)
    False
    >>> total_thc_consistent(0.46, None, 15.38) is None
    True
    """
    if d9_thc is None or thca is None or total_thc is None:
        return None
    return abs(total_thc - (d9_thc + thca * THCA_TO_THC)) <= tol


def magnitude(text):
    """Parse '4,2×10⁴ CFU/g', '4.9×10^3', '200', '< 10' to a float, or None.

    Certificates and the register between them use Cyrillic 'х' and Latin 'x',
    superscripts and caret notation, decimal commas and decimal points. All of
    them mean the same number and all of them appear in this corpus.
    """
    if text is None:
        return None
    s = str(text).strip()
    if not s or s in {"/", "-", "—", "n/a", "not tested"}:
        return None
    # A cell that is mostly prose is a note, not a measurement. Without this,
    # "COMPLIES (numeric value not present in captured source excerpt for report
    # 1625/2026 …)" parses to 1625 and is reported as 406x over an aflatoxin
    # limit of 4 — a false positive that would train people to ignore the check.
    letters = sum(ch.isalpha() for ch in s)
    if letters > 12 and letters > len(s) * 0.35:
        return None
    for k, v in SUP.items():
        s = s.replace(k, "^" + v)
    s = s.replace("^^", "^").replace("х", "x").replace("Х", "x")
    s = s.replace("×", "x").replace(" ", "")

    m = re.search(r"([\d.,]*)x?10\^?(\d)", s)
    if m:
        mant = m.group(1).replace(",", ".").rstrip(".")
        try:
            return (float(mant) if mant else 1.0) * (10 ** int(m.group(2)))
        except ValueError:
            return None
    m = re.search(r"([\d]+[.,]?[\d]*)", s)
    if m:
        try:
            return float(m.group(1).replace(",", "."))
        except ValueError:
            return None
    return None


class Limit(namedtuple("Limit", "value basis printed conflict power")):
    """An acceptance criterion resolved to the largest conforming result."""
    __slots__ = ()


# Ph. Eur. 5.1.4 / 2.6.12 and USP <1111> / <61>: an enumeration criterion of
# 10^n CFU is interpreted as a maximum acceptable count of 2 x 10^n. The factor
# is a rounding convention on the log scale — the criteria are order-of-magnitude
# limits and plate counting is intrinsically imprecise — not a safety margin to
# be stacked with anything else.
PH_EUR_ENUMERATION_FACTOR = 2

# Columns whose criterion is a microbial enumeration. The interpretation rule
# travels with the parameter, not with the notation: a bare power of ten in some
# other column is just a number.
COUNTED = re.compile(r"tamc|tymc|cfu|gnb|gram-negative|aerobic|yeast|mould|mold", re.I)

# A criterion that is a bare power of ten: an optional relational operator, then
# 10^n with no mantissa in front of it. `4,2x10^4` is a measurement and must not
# match; `<= 10^4`, `<10^4 CFU/g` and `10^5` must.
BARE_POWER = re.compile(r"^[<≤]?\s*10\^(\d)(?!\d)")


def acceptance_limit(spec, parameter=""):
    """Return the largest result that still conforms, and the basis for it.

    **This is not `magnitude()`, and must never be replaced by it.** `magnitude()`
    answers *what number is this measurement* — a measurement of `4,2 x 10⁴` is
    42 000 and nothing else. An acceptance criterion answers a different question,
    *what is the largest result that still conforms*, and for a microbial
    enumeration criterion written as a bare power of ten the pharmacopoeia says
    the two numbers are not the same:

        Ph. Eur. 5.1.4, and in identical PDG-harmonised wording USP <1111>,
        repeated in the "Interpretation of results" text of Ph. Eur. 2.6.12 /
        USP <61>:

            "The following interpretation should be applied:
             10¹ CFU: maximum acceptable count = 20;
             10² CFU: maximum acceptable count = 200;
             10³ CFU: maximum acceptable count = 2000, and so forth."

    The series is ×2 per decade, so 10⁴ CFU/g means **20 000** and 10⁵ means
    **200 000**. It applies to enumeration criteria only — TAMC, TYMC and
    bile-tolerant Gram-negative bacteria in CFU/g — and never to the "absence in
    1 g / 25 g" criteria for specified micro-organisms, which are absolute.

    Reading a criterion with `magnitude()` was this file's own version of the
    defect it exists to catch: **the right value taken from the wrong column.**
    It compared 42 000 CFU/g against 10 000 and reported nine TYMC results as
    out of specification. Five of them are. Four are not, and they were sitting
    in a release register flagged as failures for three days.

    A third number is printed on the paper. The register's specification row
    reads `≤ 10⁵ (max 500 000)` and `≤ 10⁴ (max 50 000)` — a ×5 reading. It is
    not a transcription error: the same `<10^5, max 500 000 CFU/g` is printed on
    the Purely Plant in-house CoA form itself, on every batch it covers. No
    pharmacopoeial text uses a ×5 multiplier for microbial limits. This function
    does not follow the parenthetical; it returns the compendial value and puts
    the disagreement in `conflict`, so the wrong number on the paper stays
    visible instead of being silently overwritten. **Correcting the in-house CoA
    form is a QA action and not a register edit** — the register may only record
    what the document says, and say what is wrong with it.

    Returns a `Limit`:

      value    the maximum acceptable count, or the plain ceiling for an
               ordinary numeric criterion, or None when the criterion states no
               number
      basis    ``pharmacopoeial`` when the ×2 interpretation was applied,
               ``numeric`` for an ordinary ceiling such as ``≤ 4``,
               ``unparsed`` when the criterion states no number at all
      printed  the criterion exactly as the register prints it
      conflict a sentence naming the disagreement, or ``""``

    >>> acceptance_limit("≤ 10⁴ (max 50 000)", "TYMC CFU/g").value
    20000.0
    >>> acceptance_limit("≤ 10⁵ (max 500 000)", "TAMC CFU/g").value
    200000.0
    >>> acceptance_limit("≤ 10⁴ CFU/g", "Bile-tolerant GNB /1 g").value
    20000.0
    >>> acceptance_limit("≤ 10⁴ (max 50 000)", "TYMC CFU/g").basis
    'pharmacopoeial'

    The register's own parenthetical is reported, not obeyed:

    >>> "50 000" in acceptance_limit("≤ 10⁴ (max 50 000)", "TYMC CFU/g").conflict
    True
    >>> acceptance_limit("≤ 10⁴ CFU/g", "Bile-tolerant GNB /1 g").conflict
    ''

    An ordinary numeric ceiling is returned unchanged — the ×2 rule is for
    enumeration criteria and would be nonsense on a percentage or a µg/kg limit:

    >>> acceptance_limit("< 1.00 %", "CBN %").value
    1.0
    >>> acceptance_limit("≤ 4", "Aflatoxins Σ µg/kg").basis
    'numeric'
    >>> acceptance_limit("≤ 0.1", "Hg mg/kg").value
    0.1

    A power of ten outside a counted column gets no multiplier, because nothing
    says the convention travels:

    >>> acceptance_limit("≤ 10⁴", "Some other column").basis
    'numeric'
    >>> acceptance_limit("≤ 10⁴", "Some other column").value
    10000.0

    A criterion that states no ceiling states no number:

    >>> acceptance_limit("Absent", "Salmonella /25 g").value is None
    True
    >>> acceptance_limit("per PP spec", "Ochratoxin A µg/kg").basis
    'unparsed'
    >>> acceptance_limit("", "").basis
    'unparsed'
    """
    printed = "" if spec is None else str(spec).strip()
    plain = printed
    for k, val in SUP.items():
        plain = plain.replace(k, "^" + val)
    plain = plain.replace("^^", "^").replace("х", "x").replace("Х", "x")
    plain = plain.replace("×", "x").replace(" ", "")

    m = BARE_POWER.match(plain)
    if m and COUNTED.search(str(parameter) + " " + printed):
        value = float(PH_EUR_ENUMERATION_FACTOR * 10 ** int(m.group(1)))
        stated = re.search(r"max[^\d]*([\d\s .,]+)", printed, re.I)
        conflict = ""
        if stated:
            claimed = magnitude(stated.group(1).replace(" ", " "))
            if claimed is not None and abs(claimed - value) > 1e-9:
                conflict = (
                    f"the criterion prints 'max {stated.group(1).strip()}' beside "
                    f"10^{m.group(1)}; Ph. Eur. 5.1.4 gives {value:,.0f}. The printed "
                    f"parenthetical has no pharmacopoeial authority and is not used."
                ).replace(",", " ")
        return Limit(value, "pharmacopoeial", printed, conflict,
                     float(10 ** int(m.group(1))))

    plain_value = magnitude(printed)
    return Limit(plain_value, "numeric" if plain_value is not None else "unparsed",
                 printed, "", None)


_CYR = "АВЕКМНОРСТУХЈЅІ"
_LAT = "ABEKMHOPCTYXJSI"

STABILITY_SHEET = "Stability Testing Programme"


def fold(s):
    """The comparison key used across this campaign: homoglyphs onto Latin, a
    trailing bracketed note dropped, then everything but A-Z0-9 removed.

    Certificate codes in this corpus are written in both alphabets, sometimes
    within one string — `ППК26033` mixes Cyrillic П with Latin-looking К — so
    two spellings of the same code must fold to one key before they can be
    compared.

    >>> fold("ППК26033") == fold("ППК 26033 (month 6)")
    True
    """
    u = re.sub(r"\s*\([^)]*\)\s*$", "", str(s or "")).upper()
    for a, b in zip(_CYR, _LAT):
        u = u.replace(a, b)
    return re.sub(r"[^A-Z0-9]", "", u)


def stability_codes(wb):
    """Every certificate the register itself files under stability, folded.

    The register carries a second sheet, `Stability Testing Programme`, whose
    own subtitle says what it is for:

        "These are stability-study results only — they are NOT batch-release
         results and must not be used as release or CoA-register values."

    Four of the results this file flags sit on accelerated 40 °C / 75 % RH
    timepoints, where Δ9-THCA has decarboxylated and the neutral cannabinoids
    have oxidised on to CBN. Reporting those beside a genuine release failure
    says something false about the batch — the sample was never a release
    sample. Which certificates those are is not hard-coded here, because the
    register already answers it; this reads the answer off the workbook.

    Returns an empty set when the sheet is absent, so the rule degrades to
    arithmetic rather than failing.
    """
    if STABILITY_SHEET not in wb.sheetnames:
        return set()
    ws = wb[STABILITY_SHEET]
    return {fold(row[5].value) for row in ws.iter_rows(min_row=6)
            if row[5].value} - {""}


def is_power_notation(text):
    """True when a value is written as a power of ten, which is what R2 is about.

    The rule used to ask `"10" in text`, which is a substring test on a decimal
    and matched `0.10`. Four CBD percentages were reported as suspect
    superscript misreads on the strength of it.

    >>> is_power_notation("4,2×10⁴ CFU/g")
    True
    >>> is_power_notation("< 10^3 and > 10^2")
    True
    >>> is_power_notation("0.10")
    False
    >>> is_power_notation("200")
    False
    """
    s = str(text or "")
    for k, v in SUP.items():
        s = s.replace(k, "^" + v)
    return bool(re.search(r"10\^\d", s.replace("^^", "^")))


def is_range(text):
    """True for a two-bound statement such as `< 10³ and > 10²`.

    That is IJZ-MB's standard phrasing for a count between 100 and 1000. It is
    not a measurement, and `magnitude()` reads only its first bound, so any
    comparison built on it is comparing against half a sentence.

    >>> is_range("< 10^3 and > 10^2")
    True
    >>> is_range("< 10³ и >10²")
    True
    >>> is_range("<10²>10³")
    True
    >>> is_range("4,2×10⁴")
    False
    """
    s = str(text or "")
    if re.search(r"\b(and|и)\b", s, re.I) and len(re.findall(r"[<>≤≥]", s)) >= 2:
        return True
    return len(re.findall(r"[<>≤≥]", s)) >= 2


def main(path):
    ws = load_workbook(path, data_only=True)[SHEET]
    v = lambda x: "" if x is None else str(x).strip()
    hdr = [v(c.value) for c in ws[HEADER_ROW]]
    spec = [v(c.value) for c in ws[SPEC_ROW]]

    # Which certificates are stability timepoints rather than release results.
    # The register says so itself, on its own second sheet, so this is read from
    # the workbook and not hard-coded here: a code that appears in the Stability
    # Testing Programme is not a release sample, and its result must never be
    # reported as a release failure.
    stability = stability_codes(load_workbook(path, data_only=True))

    # Columns whose specification cell states a ceiling. The ceiling is resolved
    # with acceptance_limit(), never with magnitude() — see that function for why
    # the two are not the same question.
    cols = []
    for j, s in enumerate(spec):
        if not (hdr[j] and "spec" not in hdr[j].lower()):
            continue
        lim = acceptance_limit(s, hdr[j])
        if lim.value:
            cols.append((j, hdr[j], lim))

    over, held, suspect, batch = [], [], [], ""
    for row in ws.iter_rows(min_row=FIRST_DATA):
        cells = [v(c.value) for c in row]
        if not any(cells):
            continue
        if cells[1]:
            batch = cells[1]
        code = cells[22] or ""
        if not re.match(r"^[\w/().\- ]{3,}$", code):
            continue
        for j, name, lim in cols:
            txt = cells[j] if j < len(cells) else ""
            got = magnitude(txt)
            if got is None:
                continue
            if got > lim.value:
                finding = (row[0].row, batch, code, name, txt, lim, got / lim.value)
                (held if fold(code) in stability else over).append(finding)

        # R2: a value exactly one decade below the limit, written as a power of
        # ten, is the shape a misread ⁴->³ leaves behind. Two guards, both
        # of which this rule needed and did not have:
        #   · the value must actually be written as a power of ten. Testing
        #     `"10" in text` matched the substring in "0.10" and reported four
        #     CBD percentages as suspect superscripts.
        #   · a two-bound range such as "< 10³ and > 10²" is a statement that
        #     the count lies between 100 and 1000, not a measurement. magnitude()
        #     reads only its first bound, so the comparison is meaningless.
        for j, name, lim in cols:
            txt = cells[j] if j < len(cells) else ""
            if not is_power_notation(txt) or is_range(txt):
                continue
            got = magnitude(txt)
            if got is None:
                continue
            # Against the PRINTED power of ten, not the maximum acceptable
            # count. R2 is about notation: the misread turns 4,2×10⁴ into
            # 4,2×10³, one decade below the exponent the certificate prints.
            # The ×2 interpretation belongs to the disposition, not to this.
            ceiling = lim.power or lim.value
            if ceiling / 100 < got < ceiling / 10 * 1.0000001 and got < lim.value:
                suspect.append((row[0].row, batch, code, name, txt, lim))

    print(f"limit-bearing columns: {len(cols)}")
    for _, n, lim in cols:
        print(f"   {n:<26} {lim.printed!r} -> {lim.value:g}  [{lim.basis}]")
        if lim.conflict:
            print(f"       !! {lim.conflict}")

    print(f"\nR1  release results ABOVE the maximum acceptable value: {len(over)}")
    for r, b, code, name, txt, lim, ratio in over:
        print(f"   r{r:<4} {b:<13} {code:<15} {name:<22} {txt:<10} vs "
              f"{lim.value:g} ({lim.printed}) {ratio:.2f}x")
    if not over:
        print("   none")

    print(f"\nR5  stability timepoints above the release limit — NOT release "
          f"failures: {len(held)}")
    for r, b, code, name, txt, lim, ratio in held:
        print(f"   r{r:<4} {b:<13} {code:<15} {name:<22} {txt:<10} vs "
              f"{lim.value:g} ({lim.printed}) {ratio:.2f}x")
    if held:
        print("   These rows carry a certificate listed on the register's own "
              "Stability Testing Programme sheet, which states that its results "
              "are not batch-release results. An accelerated-condition sample "
              "over a release limit is a stability observation; whether it is a "
              "failure is a question for the stability protocol.")
    else:
        print("   none")

    print(f"\nR2  one decade below the limit — verify against the page: {len(suspect)}")
    for r, b, code, name, txt, lim in suspect[:40]:
        print(f"   r{r:<4} {b:<13} {code:<15} {name:<22} {txt}")
    if len(suspect) > 40:
        print(f"   … and {len(suspect)-40} more")
    if not suspect:
        print("   none")

    return 1 if over else 0


if __name__ == "__main__":
    if len(sys.argv) < 2:
        raise SystemExit(__doc__)
    sys.exit(main(sys.argv[1]))
