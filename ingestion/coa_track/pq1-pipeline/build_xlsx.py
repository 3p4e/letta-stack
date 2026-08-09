#!/usr/bin/env python3
"""
Extract water-test parameters from 91 PDFs and emit a multi-sheet xlsx.

Sheets:
  - Raw          : long-format (one row per sample × parameter)
  - Summary      : wide-format (sample × parameter matrix) — easiest to scan
  - <SP>         : per-sampling-point pivot (date × parameter), one sheet per SP
  - OOS          : out-of-spec results only
  - Limits       : the spec (MinDK / MaxDK / unit) from the lab template
  - Coverage     : which sampling points were sampled on which dates
"""
import re, subprocess, pathlib, json, sys
from collections import defaultdict
import pandas as pd

PDFS_DIR = pathlib.Path("/tmp/pq1_pdfs")
OUT_XLSX = pathlib.Path("/tmp/PQ1_water_results.xlsx")

# Cyrillic→Latin transliteration for look-alike chars (codes only — names stay Cyrillic)
TRANS = str.maketrans({"Т":"T","Е":"E","О":"O","А":"A","С":"C","Р":"P","Х":"X","В":"B","М":"M","Н":"H","К":"K","У":"U"})

# Param catalogue: short English label, Cyrillic search anchor (line prefix), category, unit-hint
# The order here is the canonical column order in the output.
PARAMS = [
    # (key,           cat,           anchor regex (extended to swallow "во N ml" tail),             unit_default)
    ("Faecal coliforms",         "Micro",  r"Колиформни бактерии од фекално потекло во \d+ ml",     "CFU/100ml"),
    ("Total bact. 37°C",         "Micro",  r"Вкупен број на бактерии на 37[°\s]*C во \d+ ml",        "CFU/ml"),
    ("Enterococci",              "Micro",  r"Ентерококи во \d+ ml",                                  "CFU/100ml"),
    ("Sulf.-red. clostridia",    "Micro",  r"Сулфито редуцирачки клостридии во \d+ ml",              "CFU/100ml"),
    ("Aerobes 22°C",             "Micro",  r"Вкупен број на аеробни бактерии на 22[°\s]*С во \d+ ml","CFU/ml"),
    ("Pseudomonas aeruginosa",   "Micro",  r"Pseudomonas aeruginosa во \d+ ml",                      "CFU/100ml"),
    ("E. coli",                  "Micro",  r"Е\.coli",                                               "CFU/100ml"),
    ("Colour",                   "Phys",   r"Боја",                                                  "Pt-Co"),
    ("Turbidity",                "Phys",   r"Матност",                                               "NTU"),
    ("Temperature",              "Phys",   r"температура",                                           "°C"),
    ("pH",                       "PhysChem", r"^\s*рН\b",                                            "-"),
    ("KMnO4 consumption",        "PhysChem", r"Потрошувачка на KMnO4",                               "mg/l"),
    ("Conductivity 20°C",        "PhysChem", r"Електролитска спроводливост на 20[°\s]*[CС]",         "µS/cm"),
    ("Residual chlorine",        "PhysChem", r"резидуален хлор",                                     "mg/l"),
    ("Ammonia",                  "Chem",   r"Амонијак",                                              "mg/l"),
    ("Nitrites",                 "Chem",   r"Нитрити",                                               "mg/l"),
    ("Nitrates",                 "Chem",   r"Нитрати",                                               "mg/l"),
    ("Chlorides",                "Chem",   r"Хлориди",                                               "mg/l"),
    ("Iron",                     "Chem",   r"Железо",                                                "µg/l"),
    ("Manganese",                "Chem",   r"Манган",                                                "µg/l"),
]

# Hand-coded limit dictionary (from the report template; same for all samples).
# (min, max) — None = not specified
LIMITS = {
    "Faecal coliforms":        (None, 0),       # MUST be 0 (н.д.)
    "Total bact. 37°C":        (None, 20),
    "Enterococci":             (None, 0),
    "Sulf.-red. clostridia":   (None, 0),
    "Aerobes 22°C":            (None, 100),
    "Pseudomonas aeruginosa":  (None, 0),
    "E. coli":                 (None, 0),
    "Colour":                  (None, 20),
    "Turbidity":               (None, 1.5),
    "Temperature":             (None, 30),       # per A03/A05/A06 spec (was 25 from lab)
    "pH":                      (6.5, 9.5),
    "KMnO4 consumption":       (0, 8),
    "Conductivity 20°C":       (None, 2500),    # DEFAULT for TW — overridden per SL below
    "Residual chlorine":       (0.2, 0.5),       # only meaningful for chlorinated water (TW)
    "Ammonia":                 (None, 0.0005),  # per spec (lab LOD is 0.016 → not verifiable!)
    "Nitrites":                (None, 0.1),
    "Nitrates":                (None, 50),
    "Chlorides":               (None, 250),
    "Iron":                    (None, 200),     # 0.2 mg/L = 200 µg/L
    "Manganese":               (None, 50),      # 0.05 mg/L = 50 µg/L
}

# Per-water-class overrides (from A03/A05/A06)
# Key: (SL_prefix, parameter) → (min, max)
LIMITS_BY_SL = {
    ("TW","Conductivity 20°C"): (None, 2500),
    ("TR","Conductivity 20°C"): (None, 650),
    ("RO","Conductivity 20°C"): (None, 40),
}

# Sampling-point parser (re-use proven regex from prior step)
RE_DATE = re.compile(r"Датум на земање:\s*(\d{2})\.(\d{2})\.(\d{4})")
RE_MM   = re.compile(r"Мерно место:\s*(.*?)(?=Вид\s|Странка\s|Датум на земање|Хигиено|$)", re.S)
RE_CODE = re.compile(r"\b([A-Z]{2})\s*[_-]\s*([A-Z]\d{1,4})\s*[_-]\s*(\d{2,4})(?!\d)")

# Verdict
RE_VERDICT_PASS = re.compile(r"примерокот\s+ОДГОВАРА\s+НА")
RE_VERDICT_FAIL = re.compile(r"примерокот\s+НЕ\s+ОДГОВАРА")

# Number token: 1060, 6,76, 0,449, "н.д." (not detected)
RE_NUM = re.compile(r"(?:н\.д\.|\d+(?:[.,]\d+)?)")

def text(p):
    return subprocess.run(["pdftotext","-layout",str(p),"-"],capture_output=True,text=True,timeout=30).stdout

def parse_value(s):
    """Return (numeric_value or None, raw_string).
    'н.д.' → 0 numeric, 'ND' raw."""
    s = s.strip()
    if s == "н.д.":
        return 0.0, "ND"
    s2 = s.replace(",", ".")
    try:
        return float(s2), s
    except ValueError:
        return None, s

def extract_params(t):
    """For each param, find its line then capture the first numeric token after the param name."""
    lines = t.splitlines()
    out = {}
    for key, cat, pat, unit in PARAMS:
        rx = re.compile(pat)
        # find the first line matching the pattern (after potentially scanning the whole text)
        idx = None
        for i, ln in enumerate(lines):
            if rx.search(ln):
                idx = i
                break
        if idx is None:
            out[key] = {"value_num": None, "value_raw": "", "unit": unit, "found": False}
            continue
        # Capture numbers from the current line and a couple of following lines
        # (because pdftotext layout sometimes wraps values to next line)
        scope = " ".join(lines[idx:idx+3])
        # strip out the parameter name region (everything up to "ml" or end of name string)
        # but simpler: look for first number AFTER the param-name match position
        m_pat = rx.search(scope)
        post = scope[m_pat.end():]
        # The first token is the result; for params with unit text inline (like "br.bakt./100 ml")
        # we want the FIRST numeric/ND token, not the digits inside "100".
        # Most params: result comes before any of "100 ml" etc. — but for micro the unit text comes AFTER result.
        # For micro, result is "н.д." which RE_NUM picks up. For "температура" → "20  °C".
        # For pH the layout gives "    6,76                    0,244     6,5    9,5".
        m = RE_NUM.search(post)
        if not m:
            out[key] = {"value_num": None, "value_raw": "", "unit": unit, "found": False}
            continue
        val, raw = parse_value(m.group(0))
        out[key] = {"value_num": val, "value_raw": raw, "unit": unit, "found": True}
    return out

# Per user clarification:
# - C71 is a typo for C171 (same physical sampling point)
# - T161_001..005 are 5 interconnected vessels of ONE reservoir asset
CPH_FIXUPS = {"C71_007": "C171_007", "C71_008": "C171_008"}
RESERVOIR_GROUPS = {
    "RO_T161_001": "RO_T161_RES (5-vessel reservoir)",
    "RO_T161_002": "RO_T161_RES (5-vessel reservoir)",
    "RO_T161_003": "RO_T161_RES (5-vessel reservoir)",
    "RO_T161_004": "RO_T161_RES (5-vessel reservoir)",
    "RO_T161_005": "RO_T161_RES (5-vessel reservoir)",
}

def parse_sampling_point(t):
    d = RE_DATE.search(t); mm = RE_MM.search(t)
    if not (d and mm): return None, None, None
    date = f"{d.group(3)}-{d.group(2)}-{d.group(1)}"   # ISO for sorting
    seg = mm.group(1).translate(TRANS).replace("\n"," ")
    m = RE_CODE.search(seg)
    if not m: return date, None, None
    sl, p1, p2 = m.groups()
    cph = f"{p1}_{p2}"
    cph = CPH_FIXUPS.get(cph, cph)   # normalise C71→C171
    return date, sl, cph

def main():
    pdfs = sorted(PDFS_DIR.glob("*.pdf"))
    rows = []
    coverage = []
    for p in pdfs:
        sid = p.name.split()[0]
        if p.stat().st_size < 5000:
            coverage.append({"sample_id":sid, "date":None, "sl":None, "cph":None,
                             "report_size_kb":round(p.stat().st_size/1024,1),
                             "status":"STUB (empty PDF)"})
            continue
        t = text(p)
        date, sl, cph = parse_sampling_point(t)
        sp = f"{sl}_{cph}" if sl and cph else None
        verdict = "PASS" if RE_VERDICT_PASS.search(t) else ("FAIL" if RE_VERDICT_FAIL.search(t) else "?")
        params = extract_params(t)
        coverage.append({"sample_id":sid, "date":date, "sl":sl, "cph":cph,
                         "report_size_kb":round(p.stat().st_size/1024,1),
                         "verdict":verdict,
                         "params_found":sum(1 for v in params.values() if v["found"]),
                         "status":"OK"})
        for key, cat, pat, unit in PARAMS:
            v = params[key]
            # apply per-SL override if applicable
            mn, mx = LIMITS_BY_SL.get((sl, key), LIMITS[key])
            num = v["value_num"]
            # Pass/fail per parameter
            pf = ""
            if v["found"] and num is not None:
                violates = False
                if mn is not None and num < mn: violates = True
                if mx is not None and num > mx: violates = True
                pf = "FAIL" if violates else "OK"
            rows.append({
                "sample_id": sid,
                "date": date,
                "sampling_point": sp,
                "asset_group": RESERVOIR_GROUPS.get(sp, sp),
                "SL": sl,
                "CPH": cph,
                "category": cat,
                "parameter": key,
                "result": v["value_raw"] or "",
                "result_num": num,
                "unit": v["unit"],
                "min_limit": mn,
                "max_limit": mx,
                "per_param_status": pf,
                "overall_verdict": verdict,
            })
    df = pd.DataFrame(rows)
    cov = pd.DataFrame(coverage)

    # === Sheet 1: Raw long-format ===
    df_raw = df.sort_values(["date","sampling_point","category","parameter"], na_position="last")

    # === Sheet 2: Summary wide (sample × parameter) ===
    df_wide = df.pivot_table(
        index=["date","sampling_point","sample_id"],
        columns="parameter", values="result", aggfunc="first"
    )
    # Re-order columns to canonical PARAM order
    canonical = [k for k,_,_,_ in PARAMS]
    df_wide = df_wide.reindex(columns=[c for c in canonical if c in df_wide.columns])
    df_wide = df_wide.sort_index()

    # === Sheet 3: per-SP pivot tables (date × parameter) ===
    sp_pivots = {}
    for sp in sorted(df["sampling_point"].dropna().unique()):
        sub = df[df["sampling_point"]==sp]
        piv = sub.pivot_table(index=["date","sample_id"], columns="parameter",
                              values="result", aggfunc="first")
        piv = piv.reindex(columns=[c for c in canonical if c in piv.columns])
        piv = piv.sort_index()
        sp_pivots[sp] = piv

    # === Sheet 4: OOS (out-of-spec) only ===
    df_oos = df[df["per_param_status"]=="FAIL"].copy()
    df_oos = df_oos[["date","sampling_point","sample_id","parameter","result","result_num","unit","min_limit","max_limit"]]

    # === Sheet 5: Limits reference (per water class, from A03/A05/A06) ===
    lim_rows = []
    for k,c,_,u in PARAMS:
        for cls in ["TW","TR","RO"]:
            mn, mx = LIMITS_BY_SL.get((cls, k), LIMITS[k])
            lim_rows.append({"parameter":k, "category":c, "unit":u,
                             "water_class":cls, "min_limit":mn, "max_limit":mx})
    df_lim = pd.DataFrame(lim_rows)
    # Pivot to wide so each row is param and cols are TW/TR/RO limits
    df_lim_wide = df_lim.pivot_table(index=["category","parameter","unit"],
                                     columns="water_class", values=["min_limit","max_limit"],
                                     aggfunc="first")
    df_lim_wide.columns = [f"{a}_{b}" for a,b in df_lim_wide.columns]
    df_lim_wide = df_lim_wide.reset_index()

    # === Sheet 6: Coverage ===
    cov_sorted = cov.sort_values(["date","sample_id"], na_position="last")

    # === Sheet 7: Reservoir rollup (T161_001..005 → one logical asset) ===
    # For each date, for each parameter: aggregate min/avg/max across the 5 vessels.
    df_res = df[df["asset_group"]=="RO_T161_RES (5-vessel reservoir)"].copy()
    df_res["result_num_f"] = pd.to_numeric(df_res["result_num"], errors="coerce")
    res_rollup = (df_res
        .groupby(["date","parameter"], as_index=False)
        .agg(vessels_sampled=("sample_id","nunique"),
             min_val=("result_num_f","min"),
             avg_val=("result_num_f","mean"),
             max_val=("result_num_f","max"),
             any_OOS=("per_param_status", lambda s: "YES" if (s=="FAIL").any() else ""))
    )
    # Sort and reorder cols
    res_rollup = res_rollup.sort_values(["date","parameter"])
    res_rollup = res_rollup[["date","parameter","vessels_sampled","min_val","avg_val","max_val","any_OOS"]]

    # Write the workbook
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as xw:
        df_raw.to_excel(xw, sheet_name="Raw", index=False)
        df_wide.to_excel(xw, sheet_name="Summary (wide)")
        df_oos.to_excel(xw, sheet_name="OOS", index=False)
        df_lim_wide.to_excel(xw, sheet_name="Limits", index=False)
        cov_sorted.to_excel(xw, sheet_name="Coverage", index=False)
        res_rollup.to_excel(xw, sheet_name="Reservoir T161 rollup", index=False)
        for sp, piv in sp_pivots.items():
            # Excel sheet names: max 31 chars, no [ ] : * ? / \
            safe = sp.replace("/","-")[:31]
            piv.to_excel(xw, sheet_name=safe)

    # Apply some formatting: highlight OOS cells in Summary
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    wb = load_workbook(OUT_XLSX)
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF")

    # Build a lookup of OOS cells for each "summary-like" sheet
    oos_cells = {(r["date"], r["sampling_point"], r["sample_id"], r["parameter"]) for _,r in df_oos.iterrows()}

    # Style headers + freeze panes
    for sh_name in wb.sheetnames:
        ws = wb[sh_name]
        ws.freeze_panes = "A2"
        for c in ws[1]:
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Reasonable column widths
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max(8, max_len+1), 24)

    # Color OOS cells in the Summary (wide) sheet
    ws = wb["Summary (wide)"]
    # Header layout: row 1 has param names starting at col D (after date, sampling_point, sample_id)
    headers = [c.value for c in ws[1]]
    for r in range(2, ws.max_row+1):
        date = ws.cell(r, 1).value
        sp   = ws.cell(r, 2).value
        sid  = ws.cell(r, 3).value
        for ci, h in enumerate(headers[3:], start=4):
            key = (date, sp, sid, h)
            if key in oos_cells:
                ws.cell(r, ci).fill = red_fill
                ws.cell(r, ci).font = bold

    wb.save(OUT_XLSX)

    # Console summary
    print(f"✓ Wrote {OUT_XLSX}  ({OUT_XLSX.stat().st_size//1024} KB)")
    print(f"  Samples (OK):    {(cov['status']=='OK').sum()}")
    print(f"  Samples (stubs): {(cov['status']!='OK').sum()}")
    print(f"  Sampling points: {sorted(df['sampling_point'].dropna().unique())}")
    print(f"  Dates: {sorted(df['date'].dropna().unique())}")
    print(f"  OOS rows: {len(df_oos)}")
    if len(df_oos):
        print("  OOS detail:")
        for _, r in df_oos.iterrows():
            print(f"    {r['date']} {r['sampling_point']} #{r['sample_id']} : {r['parameter']} = {r['result']} {r['unit']} (limits {r['min_limit']}–{r['max_limit']})")

main()
