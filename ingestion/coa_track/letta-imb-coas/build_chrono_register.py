#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""QC results in production order — the THC-by-strain approach extended to every
parameter, listed chronologically.

One table. Each batch is a band in chronological (Seq) order; under it, one row per
parameter per testing, ordered by certificate date. A row holds the bare result value
alone; the unit sits in its own column, the acceptance criterion in the next, and every
row names the CoA code, date of issue and issuing institution it came from.

Release results only — stability timepoints, coverage notes and conclusions stay in the
master workbook. The pesticide screen collapses to a single row carrying the
certificate's own N.D. notation; a compound would get its own row only if a laboratory
reported an actual finding (none has). On a certificate that assayed only total
aflatoxins, Aflatoxin B1 and Ochratoxin A appear as "not tested" per house convention.
"""
import os
import re
from collections import Counter, OrderedDict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import coa_pivot as cp

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "exports", "PP_QC_Results_Chronological.xlsx")
FONT = "Arial"
INK = "16232B"
ACCENT = "0E6E6E"
BAND = "F6F8F7"
CRIT_BG, CRIT_FG = "F9DEDB", "AE2318"
WARN_BG, WARN_FG = "FAEDD4", "9A6300"

ND_RE = re.compile(r"^(n\.?\s?d\.?|н\.?\s?д\.?|not detected|≤\s*loq|<\s*loq|nd)\b", re.I)
UNIT_RE = re.compile(
    r"(%\s*w\s*/\s*w|%w/w|%|µg/kg|μg/kg|mg/kg|CFU\s*/?\s*g|ppm)", re.I)
DATE_KEY_RE = re.compile(r"(\d{2})\.(\d{2})\.(\d{4})")
MYCO_KEYS = ("aflatoxins", "afla_b1", "ochratoxin")


def unit_of(raw):
    """The unit printed with the result, for the Unit column; '' for verdict text."""
    m = UNIT_RE.search(raw or "")
    if not m:
        return ""
    u = m.group(1)
    u = re.sub(r"\s+", "", u)
    return {"%w/w": "% w/w", "μg/kg": "µg/kg", "CFUg": "CFU/g", "CFU/g": "CFU/g"}.get(u, u)


def date_key(d):
    m = DATE_KEY_RE.search(d or "")
    return (m.group(3), m.group(2), m.group(1)) if m else ("9999", "99", "99")


def main():
    rows, batches = cp.build()
    by_seq = cp.group_by_seq(rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QC Results (chronological)"
    ws.sheet_view.showGridLines = False

    hdr = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    base = Font(name=FONT, size=10)
    small = Font(name=FONT, size=9, color="5A6E75")
    faint = Font(name=FONT, size=9, color="9AA9AD", italic=True)
    link_f = Font(name=FONT, size=9, color=ACCENT, underline="single")
    crit = (PatternFill("solid", fgColor=CRIT_BG), Font(name=FONT, size=10, bold=True, color=CRIT_FG))
    warn = (PatternFill("solid", fgColor=WARN_BG), Font(name=FONT, size=10, color=WARN_FG))
    hair = Side(style="thin", color="C6D0CD")

    ws["A1"] = "Purely Plant GmbH — QC Results in Production Order"
    ws["A1"].font = Font(name=FONT, size=15, bold=True, color=INK)
    ws["A2"] = ("Every outsourced-laboratory release result for every batch, chronologically. "
                "One row per parameter per testing; the value stands alone, its unit beside it, "
                "and each row names the certificate it came from.")
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color="5A6E75")

    heads = ["Seq", "Parameter", "Result", "Unit", "Acceptance criterion",
             "CoA code", "Date of issue", "Issuing institution", "PDF"]
    widths = [6, 34, 18, 10, 26, 20, 14, 40, 7]
    HROW = 4
    for ci, h in enumerate(heads, start=1):
        c = ws.cell(row=HROW, column=ci, value=h)
        c.font = hdr
        c.fill = PatternFill("solid", fgColor=INK)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = widths[ci - 1]
    ws.row_dimensions[HROW].height = 22
    ws.freeze_panes = "A%d" % (HROW + 1)

    r = HROW + 1
    n_rows = 0
    for b in batches:
        # ---- batch band ----
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(heads))
        label = "   %02d      %s" % (b["seq"], b["batch"])
        if b["p_number"]:
            label += "   (%s)" % b["p_number"]
        label += "      %s" % (b["strain"] or "strain not stated")
        h = ws.cell(row=r, column=1, value=label)
        h.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
        h.fill = PatternFill("solid", fgColor=ACCENT)
        h.alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 19
        r += 1

        # ---- group records by certificate, ordered by date of issue ----
        groups = OrderedDict()
        for rec in by_seq[b["seq"]]:
            if cp.STABILITY_RE.search(rec["Parameter"] or ""):
                continue
            k = cp.classify(rec["Parameter"])
            if k not in cp.COLS:
                continue
            code = rec["Certificate Code"].strip() or "(not numbered)"
            gkey = (code, cp.issue_date(rec["Issue Date"]))
            g = groups.setdefault(gkey, {"inst": (rec["Issuing Institution"] or "").strip(),
                                         "link": cp.clean_link(rec["Drive File Link"]),
                                         "recs": []})
            g["recs"].append((k, rec))
        ordered = sorted(groups.items(), key=lambda kv: date_key(kv[0][1]))

        for (code, date), g in ordered:
            keys_here = [k for k, _ in g["recs"]]
            # order parameters in house order; pesticides collapse to one row
            pest_recs = [rec for k, rec in g["recs"] if k == "pesticides"]
            emit = []
            for col_key in cp.COLS:
                if col_key == "pesticides":
                    if pest_recs:
                        hits = [x for x in pest_recs if not ND_RE.match(x["Result"].strip())]
                        if hits:
                            for x in hits:
                                emit.append((cp.short_label(x["Parameter"]),
                                             x["Result"], x["Acceptance Criterion"]))
                        else:
                            tok = Counter(x["Result"].strip() for x in pest_recs).most_common(1)[0][0]
                            emit.append(("Pesticide screen (%d compounds)" % len(pest_recs),
                                         tok, "≤ LOQ 0.01 mg/kg"))
                    continue
                for k, rec in g["recs"]:
                    if k == col_key:
                        emit.append((cp.COLS[k], rec["Result"], rec["Acceptance Criterion"]))
            # mycotoxin sub-parameters not analysed on this certificate
            if any(k in MYCO_KEYS for k in keys_here):
                for mk in ("afla_b1", "ochratoxin"):
                    if mk not in keys_here:
                        emit.append((cp.COLS[mk], "not tested", ""))

            cert_start = r
            for pname, raw, ac in emit:
                ws.cell(row=r, column=1, value=b["seq"]).font = faint
                pc = ws.cell(row=r, column=2, value=pname)
                pc.font = base
                pc.alignment = Alignment(wrap_text=True, vertical="center")
                if raw == "not tested":
                    vc = ws.cell(row=r, column=3, value="not tested")
                    vc.font = faint
                    ws.cell(row=r, column=4, value="").font = faint
                else:
                    vc = ws.cell(row=r, column=3, value=cp.clean_value(raw))
                    sev = cp.severity(raw)
                    if sev == "crit":
                        vc.fill, vc.font = crit
                    elif sev == "warn":
                        vc.fill, vc.font = warn
                    else:
                        vc.font = base
                    vc.alignment = Alignment(vertical="center")
                    ws.cell(row=r, column=4, value=unit_of(raw)).font = small
                acell = ws.cell(row=r, column=5, value=(ac or "").strip() or "/")
                acell.font = small
                acell.alignment = Alignment(wrap_text=True, vertical="center")
                r += 1
                n_rows += 1
            # certificate reference merged down its rows
            if r > cert_start:
                for col, val in ((6, code), (7, date), (8, g["inst"])):
                    if r - 1 > cert_start:
                        ws.merge_cells(start_row=cert_start, start_column=col,
                                       end_row=r - 1, end_column=col)
                    cc = ws.cell(row=cert_start, column=col, value=val)
                    cc.font = small if col == 8 else base
                    cc.alignment = Alignment(vertical="center", wrap_text=True)
                if r - 1 > cert_start:
                    ws.merge_cells(start_row=cert_start, start_column=9,
                                   end_row=r - 1, end_column=9)
                if g["link"]:
                    lc = ws.cell(row=cert_start, column=9, value="Open")
                    lc.hyperlink = g["link"]
                    lc.font = link_f
                    lc.alignment = Alignment(horizontal="center", vertical="center")
                for ci in range(1, len(heads) + 1):
                    ws.cell(row=r - 1, column=ci).border = Border(bottom=Side(style="thin", color="DBE3E1"))
        for ci in range(1, len(heads) + 1):
            ws.cell(row=r - 1, column=ci).border = Border(bottom=hair)

    # legend
    r += 1
    ws.cell(row=r, column=1, value="LEGEND").font = Font(name=FONT, size=10, bold=True, color=ACCENT)
    r += 1
    for txt, fill in [
        ("Result cells hold the bare value; the unit is beside it and expanded measurement "
         "uncertainty is excluded (it remains on the certificate).", None),
        ("\"not tested\" — mycotoxin sub-parameter not analysed on a certificate that assayed "
         "only total aflatoxins.", None),
        ("Pesticide screen: one row with the certificate's own N.D. notation; a compound is "
         "named only where a laboratory reported an actual finding.", None),
        ("Red — result the issuing laboratory declared out of specification.", CRIT_BG),
        ("Amber — a laboratory finding or data-integrity flag on the certificate.", WARN_BG),
    ]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(heads))
        c = ws.cell(row=r, column=1, value=txt)
        c.font = Font(name=FONT, size=9)
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(vertical="center", indent=1)
        r += 1

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print("Wrote %s" % OUT)
    print("batches=%d result-rows=%d" % (len(batches), n_rows))


if __name__ == "__main__":
    main()
