#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Batch Release QC Register — ONE sheet, modelled on the Purely Plant
"Production Batches QC Result Table".

Layout follows the example: a header row, a single "Laboratory reference values"
row stating each acceptance criterion once, then one row per batch — with a
further row for each additional laboratory/certificate, "/" where that report did
not cover a parameter. A legend sits at the foot of the sheet.

Extended only where the example leaves a gap and the batch-release / CoQ use
needs it: the per-batch Total-THC spec, the three Ph. Eur. mycotoxin parameters,
arsenic, and — so every value can be traced when a CoQ is compiled — the CoA
code, date of issue and issuing institution behind each row.
"""
import os
import re
from collections import Counter, OrderedDict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import coa_pivot as cp

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "exports", "PP_Batch_Release_QC_Register.xlsx")
FONT = "Arial"
INK = "16232B"
ACCENT = "0E6E6E"
REFBG = "E7ECEA"
BAND = "F6F8F7"
CRIT_BG, CRIT_FG = "F9DEDB", "AE2318"
WARN_BG, WARN_FG = "FAEDD4", "9A6300"

# column key -> (header, reference-value, width)
COLUMNS = [
    ("seq",          "No.",                       "",                              5),
    ("batch",        "Batch number",              "",                              20),
    ("pnum",         "P-number",                  "",                              11),
    ("strain",       "Strain",                    "",                              17),
    ("thc",          "THC %",                     "per-batch label range",         16),
    ("thc_spec",     "THC spec",                  "% w/w (label claim)",           18),
    ("cbd",          "CBD %",                     "< 1.00 %",                      11),
    ("cbn",          "CBN %",                     "< 1.00 %",                      11),
    ("loss_on_drying", "Loss on drying %",        "≤ 12.00 (3028)",           14),
    ("tamc",         "TAMC CFU/g",                "≤ 10⁵ (max 500 000)", 15),
    ("tymc",         "TYMC CFU/g",                "≤ 10⁴ (max 50 000)",  15),
    ("bile",         "Bile-tolerant GNB /1 g",    "≤ 10⁴ CFU/g",         18),
    ("salmonella",   "Salmonella /25 g",          "Absent",                        13),
    ("ecoli",        "E. coli /1 g",              "Absent",                        13),
    ("aflatoxins",   "Aflatoxins Σ µg/kg",   "≤ 4",                      14),
    ("afla_b1",      "Aflatooxin B1 µg/kg" if False else "Aflatoxin B1 µg/kg",  "≤ 2",  13),
    ("ochratoxin",   "Ochratoxin A µg/kg",        "per PP spec",                   13),
    ("pb",           "Pb mg/kg",                  "≤ 0.5",                    11),
    ("cd",           "Cd mg/kg",                  "≤ 0.3",                    11),
    ("as_",          "As mg/kg",                  "≤ 0.2",                    11),
    ("hg",           "Hg mg/kg",                  "≤ 0.1",                    11),
    ("pesticides",   "Pesticides",                "≤ LOQ 0.01 mg/kg",         30),
    ("coa",          "CoA code",                  "",                              20),
    ("date",         "Date of issue",             "",                              14),
    ("inst",         "Issuing institution",       "",                              34),
    ("pdf",          "PDF",                       "",                              7),
]
PARAM_KEYS = ["thc", "thc_spec", "cbd", "cbn", "loss_on_drying", "tamc", "tymc", "bile",
              "salmonella", "ecoli", "aflatoxins", "afla_b1", "ochratoxin",
              "pb", "cd", "as_", "hg", "pesticides"]
MYCO_KEYS = ("aflatoxins", "afla_b1", "ochratoxin")
ND_RE = re.compile(r"^(n\.?\s?d\.?|н\.?\s?д\.?|not detected|≤\s*loq|<\s*loq|nd)\b", re.I)


def lab_of(inst):
    for name, rx, _acc in cp.LAB_PATTERNS:
        if rx.search(inst or ""):
            return name
    if "purely plant" in (inst or "").lower() or "in-house" in (inst or "").lower():
        return "Purely Plant GmbH (in-house)"
    return (inst or "").strip()


def main():
    rows, batches = cp.build()
    by_seq = cp.group_by_seq(rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Batch Release QC"
    ws.sheet_view.showGridLines = False

    hdr = Font(name=FONT, size=9, bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor=INK)
    base = Font(name=FONT, size=10)
    bold = Font(name=FONT, size=10, bold=True)
    small = Font(name=FONT, size=9, color="5A6E75")
    faint = Font(name=FONT, size=9, color="9AA9AD", italic=True)
    link_f = Font(name=FONT, size=9, color=ACCENT, underline="single")
    ref_font = Font(name=FONT, size=8, bold=True, color=INK)
    crit_fill = PatternFill("solid", fgColor=CRIT_BG)
    crit_font = Font(name=FONT, size=10, bold=True, color=CRIT_FG)
    warn_fill = PatternFill("solid", fgColor=WARN_BG)
    warn_font = Font(name=FONT, size=10, color=WARN_FG)
    band_fill = PatternFill("solid", fgColor=BAND)
    ref_fill = PatternFill("solid", fgColor=REFBG)
    hair = Side(style="thin", color="C6D0CD")

    ncol = len(COLUMNS)

    # title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    ws["A1"] = "Purely Plant GmbH — Production Batches QC Result Register"
    ws["A1"].font = Font(name=FONT, size=15, bold=True, color=INK)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    ws["A2"] = ("Outsourced-laboratory release results, one row per certificate. "
                "Reference values below are the current Ph. Eur. 07/2024:3028 release "
                "specification; each row is traceable to its CoA code, date and institution.")
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color="5A6E75")

    HROW, REFROW, DATA0 = 4, 5, 6

    for ci, (key, head, ref, w) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=HROW, column=ci, value=head)
        c.font = hdr
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
        rc = ws.cell(row=REFROW, column=ci, value=ref)
        rc.font = ref_font
        rc.fill = ref_fill
        rc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(row=REFROW, column=1, value="Ref.").font = Font(name=FONT, size=8, bold=True, color=INK)
    ws.cell(row=REFROW, column=1).fill = ref_fill
    ws.row_dimensions[HROW].height = 40
    ws.row_dimensions[REFROW].height = 26
    ws.freeze_panes = ws.cell(row=DATA0, column=5)

    r = DATA0
    for bi, b in enumerate(batches):
        # group this batch's records by certificate (code + date)
        groups = OrderedDict()
        for rec in by_seq[b["seq"]]:
            code = rec["Certificate Code"].strip() or "(not numbered)"
            gkey = (code, cp.issue_date(rec["Issue Date"]))
            g = groups.setdefault(gkey, {"inst": "", "link": "", "vals": {}})
            if not g["inst"]:
                g["inst"] = lab_of(rec["Issuing Institution"] or "")
            if not g["link"]:
                g["link"] = cp.clean_link(rec["Drive File Link"])
            k = cp.classify(rec["Parameter"])
            if k in PARAM_KEYS:
                g["vals"].setdefault(k, []).append(rec["Result"].strip())
                if k == "thc":
                    ac = (rec["Acceptance Criterion"] or "").strip()
                    if ac and ac not in ("/", "—"):
                        g["vals"].setdefault("thc_spec", []).append(ac)
        if not groups:
            groups[("(none)", "")] = {"inst": "", "link": "", "vals": {}}

        first = True
        block_start = r
        shade = band_fill if (bi % 2 == 1) else None
        for (code, date), g in groups.items():
            myco_present = any(g["vals"].get(k) for k in MYCO_KEYS)
            # identity
            idfont = base if first else faint
            ws.cell(row=r, column=1, value=b["seq"] if first else "").font = idfont
            ws.cell(row=r, column=2, value=b["batch"] if first else "").font = (bold if first else faint)
            ws.cell(row=r, column=3, value=(b["pnum"] if False else b["p_number"]) if first else "").font = idfont
            ws.cell(row=r, column=4, value=b["strain"] if first else "").font = idfont

            for key in PARAM_KEYS:
                ci = key_col(key)
                vals = g["vals"].get(key)
                if not vals:
                    blank = "not tested" if (key in MYCO_KEYS and myco_present) else "/"
                    cc = ws.cell(row=r, column=ci, value=blank)
                    cc.font = faint
                    cc.alignment = Alignment(horizontal="center")
                    continue
                if key == "thc_spec":
                    text = cp.clean_thc_spec(vals[0])
                elif key == "pesticides":
                    hits = [v for v in vals if not ND_RE.match(v)]
                    text = ("; ".join(cp.clean_value(v) for v in hits) if hits
                            else cp.clean_value(Counter(vals).most_common(1)[0][0]))
                elif len(vals) == 1:
                    text = cp.clean_value(vals[0])
                else:
                    text = " ; ".join(dict.fromkeys(cp.clean_value(v) for v in vals))
                cc = ws.cell(row=r, column=ci, value=text)
                cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=(key == "pesticides"))
                sev = cp.severity(" ".join(vals))
                if sev == "crit":
                    cc.fill, cc.font = crit_fill, crit_font
                elif sev == "warn":
                    cc.fill, cc.font = warn_fill, warn_font
                else:
                    cc.font = base

            # provenance
            ws.cell(row=r, column=key_col("coa"), value=code).font = base
            ws.cell(row=r, column=key_col("date"), value=date).font = base
            ic = ws.cell(row=r, column=key_col("inst"), value=g["inst"])
            ic.font = small
            ic.alignment = Alignment(wrap_text=True, vertical="center")
            if g["link"]:
                lc = ws.cell(row=r, column=key_col("pdf"), value="Open")
                lc.hyperlink = g["link"]
                lc.font = link_f
                lc.alignment = Alignment(horizontal="center")

            if shade:
                for ci in range(1, ncol + 1):
                    cell = ws.cell(row=r, column=ci)
                    if cell.fill.fgColor.rgb in (None, "00000000"):
                        cell.fill = shade
            first = False
            r += 1
        # rule under the batch block
        for ci in range(1, ncol + 1):
            ws.cell(row=r - 1, column=ci).border = Border(bottom=hair)

    ws.auto_filter.ref = "A%d:%s%d" % (HROW, get_column_letter(ncol), r - 1)

    # legend
    r += 1
    ws.cell(row=r, column=1, value="LEGEND").font = Font(name=FONT, size=10, bold=True, color=ACCENT)
    r += 1
    legend = [
        ("n.d — not detected    ·    BLQ — below limit of quantification    ·    "
         "LOQ — limit of quantification", None),
        ("\"/\" — parameter not covered by that certificate", None),
        ("\"not tested\" — that mycotoxin sub-parameter (Aflatoxin B1 / Ochratoxin A) was not "
         "analysed on a certificate that assayed only total aflatoxins", None),
        ("Red — result the issuing laboratory declared out of specification", CRIT_BG),
        ("Amber — a laboratory finding or a data-integrity flag on the certificate", WARN_BG),
        ("One row per certificate: a batch tested by several laboratories has one row each; "
         "batch identity is shown once and left blank on the continuation rows.", None),
    ]
    for txt, fill in legend:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
        c = ws.cell(row=r, column=1, value=txt)
        c.font = Font(name=FONT, size=9)
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(vertical="center", indent=1)
        r += 1

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print("Wrote %s" % OUT)
    print("columns=%d certificate-rows=%d batches=%d" % (ncol, r, len(batches)))


_COLINDEX = {key: i + 1 for i, (key, _h, _r, _w) in enumerate(COLUMNS)}


def key_col(key):
    return _COLINDEX[key]


if __name__ == "__main__":
    main()
