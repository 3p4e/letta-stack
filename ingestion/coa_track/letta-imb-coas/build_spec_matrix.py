#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Wide-format spec matrix: one chronological row-block per batch, parameters as columns.

The transposition of exports/PP_Spec_Parameter_Listing.xlsx requested by the QC Manager:
the same fact-checked data spread across the columns to the right, row-per-batch style.
Every QCSP 001 parameter owns a dedicated column pair — Result | eCoA reference — under
frozen header rows that carry the parameter name (with unit) and its acceptance criterion
in dedicated cells. Each result's provenance is rendered compactly as "code, date, lab"
(e.g. ППК25105, 17.04.2025, CNP) and that text is hyperlinked to the certificate scan.

Where one batch was tested in more than one campaign — release testing in one season, a
re-test a year later — the batch row splits into testing-round sub-rows, one per campaign,
so results from different dates never share a cell. Rounds are clustered on certificate
issue dates (a gap of more than ROUND_GAP_DAYS opens a new round; observed campaigns span
at most 50 days internally and the nearest distinct re-test sits 88 days out); each
sub-row is labelled with its round and month span. The batch identity cells span the
sub-rows. A parameter never tested for the batch reads "Missing / not tested" once,
spanning the sub-rows; a parameter tested in one round but not another reads "—" in the
round that lacks it.

The specification-document and grade identity columns are gone: the grade sits in its own
cell immediately right of each THC assay, as classification of that assay against the
cultivar's QCSP 001 v.02 ladder — "Grade III, 15.00 – 20.00" — never a disposition.

Column C carries the proposed CoQ document code of each TESTING ROUND — one issuance
event per sub-row. The initial round's CoQ issues on completion of the initial QC
sampling; a resampling round produces a superseding CoQ with a new issue date that
carries the resampled parameters' new results and reuses the non-resampled results from
the earlier sampling's eCoAs (time-invariant parameters such as heavy metals and
pesticides). QCSOP 012 v.03 §6.4.2 sets the convention — CoQ-PP-[YYYY]-[NNNN], the next
sequential number for the certificate type within the calendar year, drawn from the
Certificate Issuance Register (QCLB 020) — so codes are proposed in results-complete
chronology across all events, pending the register entries the QC Manager makes at
final approval. The 'CoQ issuance plan' sheet spells out each event's basis date,
resampled vs carried-forward parameters and the supersession chain (§6.7).

Workbook layout: a "Read me" cover sheet carries the reading guide, laboratory legend,
house-rule notes and a per-parameter coverage table, so the matrix itself stays clean; the
"Spec matrix" sheet opens active; the "eCoA index" sheet lists every certificate with its
own scan link (a matrix cell holds one hyperlink, so a cell stacking the paired same-day
certificates links to the first scan). Clean numeric results are stored as real numbers
with number formats that preserve the certificate's printed decimals (0.010 stays 0.010),
so columns sort and aggregate without any transcription drift.

Nothing is recomputed from the certificates: the rows come from
build_spec_param_listing.build_rows(), the dataset the four-layer fact check verified,
and the build asserts the matrix holds the same totals. Conformity is never colour-coded
or derived — a failure exists only where the laboratory declared one.

Laboratory abbreviations follow the QC Manager's referencing convention:
CNP (UKIM Faculty of Pharmacy — Center for Natural Products), FARM (Farmahem),
IPH (Institute of Public Health), SPL (State Phytosanitary Laboratory), PP (in-house).
"""
import csv
import datetime
import os
import re
import sys
from collections import OrderedDict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

import build_house_specs as hs
import build_spec_param_listing as bl

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_XLSX = os.path.join(HERE, "exports", "PP_Spec_Parameter_Matrix.xlsx")
OUT_TSV = os.path.join(HERE, "exports", "PP_Spec_Parameter_Matrix.tsv")

BUILD_DATE = "13.08.2026"

# The reference convention writes UKIM's certificates under the Center's own initials.
REF_LAB = {"UKIM": "CNP"}
LAB_LEGEND = [
    ("CNP", "УКИМ Фармацевтски факултет — Центар за природни производи / UKIM Faculty "
            "of Pharmacy — Center for Natural Products (LT-083)"),
    ("FARM", "Фармахем / Farmahem Laboratory for Environmental Sciences (LT-017)"),
    ("IPH", "Институт за јавно здравје / Institute of Public Health (LT-005)"),
    ("SPL", "Државна фитосанитарна лабораторија / State Phytosanitary Laboratory "
            "(LT-036)"),
    ("PP", "Purely Plant — in-house CoA (three R&D batches ingested on instruction)"),
]

COQ_YEAR = 2026          # calendar year the CoQs would be created in
ICOA_YEAR = 2026         # calendar year the Foreign-Matter iCoAs would be issued in
ICOA_DATE = BUILD_DATE   # proposed issue date for those iCoAs
FM_ITEM = "7"            # Foreign Matter — the spec item covered by declared iCoAs
ROUND_GAP_DAYS = 60      # a longer gap between certificate dates opens a new round

# (item №, short column title, unit shown in the header, acceptance criterion)
MATRIX_PANEL = [
    ("1",    "Macroscopic ID (Test A)",    "",      "Conforms to description"),
    ("2",    "Microscopic ID (Test B)",    "",      "Conforms to description"),
    ("3",    "HPTLC ID (Test C)",          "",      "Identity confirmed"),
    ("4",    "Total Δ⁹-THC",               "%",     "Per target grade (QCSP 001 §01)"),
    ("5",    "Total CBD",                  "%",     "≤ 1.0"),
    ("6",    "Total CBN",                  "%",     "≤ 1.0"),
    ("7",    "Foreign Matter",             "%",     "≤ 2.0"),
    ("8",    "Loss on Drying",             "%",     "≤ 12.0"),
    ("9.1",  "TAMC",                       "CFU/g", "≤ 10⁵"),
    ("9.2",  "TYMC",                       "CFU/g", "≤ 10⁴"),
    ("9.3",  "Bile-tol. gram-negative",    "CFU/g", "≤ 10⁴"),
    ("9.4",  "Salmonella",                 "",      "Absence /25 g"),
    ("9.5",  "Escherichia coli",           "",      "Absence /1 g"),
    ("10.1", "Aflatoxin B1",               "µg/kg", "≤ 2"),
    ("10.2", "Total Aflatoxins",           "µg/kg", "≤ 4"),
    ("10.3", "Ochratoxin A",               "µg/kg", "≤ 20"),
    ("11.1", "Lead (Pb)",                  "mg/kg", "≤ 0.5"),
    ("11.2", "Cadmium (Cd)",               "mg/kg", "≤ 0.3"),
    ("11.3", "Arsenic (As)",               "mg/kg", "≤ 0.2"),
    ("11.4", "Mercury (Hg)",               "mg/kg", "≤ 0.1"),
    ("12",   "Pesticide Residues",         "",      "≤ LOQ (Ph. Eur. 2.8.13)"),
]
THC_ITEM = "4"           # the one parameter that carries the extra Grade column

IDENTITY = [("#", 4), ("Production", 9), ("CoQ doc. code\n(proposed)", 16),
            ("Batch", 13), ("P-Number", 10), ("Strain", 17), ("Product code", 19)]
N_ID = len(IDENTITY)
ROUND_COL = N_ID + 1     # per-sub-row round label, first column after the identity block
ROUND_W = 13

RESULT_W = {"12": 26, "1": 11, "2": 11, "3": 11}
GRADE_W = 18
REF_W = 20

NAVY = "1B3A5C"
GOLD = "C9A227"
GOLD_TINT = "FBF6E9"
SUB_GREY = "E9EEF5"
BAND = "F4F7FB"
MISS = "9AA7B4"
INK = "16232B"
LINK = "0563C1"

MATRIX_TITLE = ("Purely Plant — QCSP 001 specification matrix · one row-block per "
                "batch, sub-rows = testing rounds · reading guide, legend and coverage "
                "on the 'Read me' sheet · informal working export, %s" % BUILD_DATE)
READ_TITLE = "Purely Plant — QCSP 001 Specification Matrix"
READ_SUBTITLE = ("eCoA results per specification parameter · all batches, "
                 "chronological · informal working export, not a controlled "
                 "document · built %s" % BUILD_DATE)
# Optional extra cover-sheet table: (heading, headers, column widths, data rows).
# A derived workbook (e.g. a tranche subset) injects its overview here.
EXTRA_COVER = None

DATE_RE = re.compile(r"^(\d{2})\.(\d{2})\.(\d{4})$")
NUM_CELL_RE = re.compile(r"^\d+(?:\.\d+)?$")
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def parse_date(s):
    m = DATE_RE.match((s or "").strip())
    if not m:
        return None
    try:
        return datetime.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    except ValueError:
        return None


def month_span(d1, d2):
    a = "%s %d" % (MONTHS[d1.month - 1], d1.year)
    b = "%s %d" % (MONTHS[d2.month - 1], d2.year)
    if a == b:
        return a
    if d1.year == d2.year:
        return "%s–%s %d" % (MONTHS[d1.month - 1], MONTHS[d2.month - 1], d1.year)
    return "%s – %s" % (a, b)


def matrix_value(rec):
    """The result line for a matrix cell: bare value, unit stripped into the header."""
    v = (rec["value"] or "").strip()
    u = rec["unit"]
    if u not in ("—", "") and v.endswith(u):
        head = v[: -len(u)].strip()
        if head:
            v = head
    return v.replace(" — no analyte detected", "")


def ref_line(rec):
    lab = REF_LAB.get(rec["lab"], rec["lab"])
    return "%s, %s, %s" % (rec["cert"], rec["date"], lab)


_LADDERS = {}


def ladder_for(strain):
    if strain not in _LADDERS:
        pub = hs.PUBLISHED.get(strain)
        _LADDERS[strain] = hs.extend(pub[3])[0] if pub else None
    return _LADDERS[strain]


def grade_line(strain, value):
    """'Grade III, 15.00 – 20.00' for a numeric assay; '—' for qualifiers/no ladder.

    Tolerance mirrors build_spec_param_listing.batch_grade exactly — qualified values
    (<, ≤, LOQ, BLQ) carry no grade; otherwise the first number in the string is the
    assay — so the per-assay grades here always union to the listing's batch grade.
    """
    ladder = ladder_for(strain)
    if ladder is None:
        return "—"
    v = value.strip()
    if not v or any(t in v.lower() for t in ("<", "≤", "loq", "blq")):
        return "—"
    m = re.search(r"\d+(?:\.\d+)?", v)
    if not m:
        return "—"
    g = hs.grade_of(ladder, float(m.group(0)))
    if g is None:
        return "outside every grade"
    for label, _nominal, floor, cap in ladder:
        if label == g:
            return "Grade %s, %.2f – %.2f" % (g, floor, cap)
    return "Grade %s" % g


ID_ITEMS = ("1", "2")    # macroscopic + microscopic ID — one combined iCoA per batch
FM_SCOPE = "Foreign Matter (prior to final QC sampling)"
ID_SCOPE = "Macroscopic + Microscopic ID (after final QC sampling)"


def icoa_assignments(full_listing):
    """batch -> {'fm': code|None, 'id': code|None} — the proposed iCoA issuance plan.

    Per the QC Manager's instructions of 13.08.2026: Foreign Matter was determined
    in-house for every batch per SOP prior to final QC sampling and packaging, and
    macroscopic + microscopic identification per specification was performed in-house
    after final QC sampling — all with passing results. Those determinations are to
    be issued as iCoAs so the CoQ can cite a source certificate: one FM iCoA where no
    Foreign Matter is on file, and one combined ID iCoA covering both identification
    tests wherever either is not already on file.

    QCSOP 012 v.03 gives the iCoA type ONE sequential series per calendar year (from
    the Certificate Issuance Register, QCLB 020), so both kinds share a single
    proposed sequence, chronological by batch, FM (pre-sampling) before ID
    (post-sampling) within a batch. Pass the FULL listing here even when building a
    subset workbook, so a batch keeps the same iCoA numbers everywhere.
    """
    per = OrderedDict()
    for rec in full_listing:
        f = per.setdefault(rec["batch"], {"fm": False, "macro": False, "micro": False})
        if not rec["value"].startswith("Missing"):
            if rec["no"] == FM_ITEM:
                f["fm"] = True
            elif rec["no"] == "1":
                f["macro"] = True
            elif rec["no"] == "2":
                f["micro"] = True
    out = OrderedDict()
    seq = 0
    for batch, f in per.items():
        codes = {"fm": None, "id": None}
        if not f["fm"]:
            seq += 1
            codes["fm"] = "iCoA-PP-%d-%04d" % (ICOA_YEAR, seq)
        if not (f["macro"] and f["micro"]):
            seq += 1
            codes["id"] = "iCoA-PP-%d-%04d" % (ICOA_YEAR, seq)
        out[batch] = codes
    return out


def declared_ref(code):
    return "%s (proposed), %s, PP" % (code, ICOA_DATE)


def proposed_icoa_rows(rows):
    """The proposed-iCoA issuance list for the batches in `rows`, in sequence order.

    One entry per proposed iCoA: the FM certificate covers spec item 7; the combined
    ID certificate covers the identification cells it was injected into (both items
    for most batches; only the microscopic cell where a macroscopic record is already
    on file).
    """
    by_code = OrderedDict()
    for b in rows:
        for rnd in b["rounds"]:
            for no, _t, _u, _ac in MATRIX_PANEL:
                cell = rnd[no]
                if not cell.get("declared"):
                    continue
                for v, ref in zip(cell["values"], cell["refs"]):
                    if "(proposed)" not in ref:
                        continue
                    code = ref.split(" (proposed)")[0]
                    e = by_code.setdefault(code, {
                        "cert": code, "date": ICOA_DATE, "lab": "PP", "n": 0,
                        "batches": b["batch"], "items": [], "link": "",
                        "proposed": True, "p": b["p"], "strain": b["strain"],
                        "prod": b["prod"],
                        "scope": FM_SCOPE if no == FM_ITEM else ID_SCOPE,
                    })
                    e["n"] += 1
                    if no not in e["items"]:
                        e["items"].append(no)
    out = []
    for e in by_code.values():
        e["items"] = ", ".join(e["items"])
        out.append(e)
    out.sort(key=lambda e: e["cert"])
    return out


def pivot(listing, icoa=None):
    """listing rows -> ordered batches, each with testing rounds of per-item cells.

    icoa: batch -> {'fm': code, 'id': code} proposed-iCoA plan for the declared
    in-house entries. When None it is derived from `listing` itself — correct for
    the master build; subset builds must pass icoa_assignments(full_listing) to
    keep master numbering.
    """
    if icoa is None:
        icoa = icoa_assignments(listing)
    batches = OrderedDict()
    for rec in listing:
        b = batches.setdefault(rec["batch"], {
            "chrono": rec["chrono"], "prod": rec["prod"], "batch": rec["batch"],
            "p": rec["p"], "strain": rec["strain"], "code": rec["code"],
            "entries": [],
        })
        if not rec["value"].startswith("Missing"):
            b["entries"].append(rec)

    for b in batches.values():
        # cluster certificate issue dates into testing rounds
        dates = sorted({d for d in (parse_date(r["date"]) for r in b["entries"]) if d})
        cluster_of = {}
        start = None
        for d in dates:
            if start is None or (d - prev).days > ROUND_GAP_DAYS:
                start = d
            cluster_of[d] = start
            prev = d
        starts = sorted(set(cluster_of.values()))
        idx_of = {s: i for i, s in enumerate(starts)}
        n_rounds = max(1, len(starts))

        rounds = [
            {no: {"values": [], "grades": [], "refs": [], "links": []}
             for no, _t, _u, _ac in MATRIX_PANEL}
            for _ in range(n_rounds)
        ]
        spans = [[] for _ in range(n_rounds)]
        for rec in b["entries"]:
            d = parse_date(rec["date"])
            ri = idx_of[cluster_of[d]] if d else 0
            if d:
                spans[ri].append(d)
            cell = rounds[ri][rec["no"]]
            val = matrix_value(rec)
            cell["values"].append(val)
            cell["refs"].append(ref_line(rec))
            cell["links"].append(rec["link"])
            if rec["no"] == THC_ITEM:
                cell["grades"].append(grade_line(b["strain"], val))
        b["rounds"] = rounds
        b["round_labels"] = []
        for ri in range(n_rounds):
            tag = "R%d" % (ri + 1) if n_rounds > 1 else "R1"
            if spans[ri]:
                tag += " · " + month_span(min(spans[ri]), max(spans[ri]))
            b["round_labels"].append(tag)
        # declared in-house determinations, entered as Conforms citing the proposed
        # iCoA: Foreign Matter (per SOP prior to final QC sampling and packaging) and
        # macroscopic + microscopic ID (per specification after final QC sampling —
        # one combined iCoA backs both identification cells)
        codes = icoa.get(b["batch"], {})
        plan = [(FM_ITEM, codes.get("fm"))] + [(i, codes.get("id")) for i in ID_ITEMS]
        for item, code in plan:
            if code and not any(rnd[item]["values"] for rnd in rounds):
                cell = rounds[0][item]
                cell["values"].append("Conforms")
                cell["refs"].append(declared_ref(code))
                cell["links"].append("")
                cell["declared"] = True
    return list(batches.values())


def coq_plan(master_rows):
    """One CoQ issuance event per testing round of every batch, in true chronology.

    Per the QC Manager's instruction (13.08.2026, BSS1024 example): a CoQ is issued
    on completion of the initial QC sampling round; where resampling (R) was later
    requested and performed for certain parameters, a second CoQ is issued with a
    new issue date — it carries the resampled parameters' new results and REUSES the
    non-resampled parameters' results from the earlier sampling's eCoAs (heavy
    metals, pesticides and the like do not change with time) — and supersedes the
    earlier CoQ per QCSOP 012 v.03 §6.7.

    Events are ordered by the date each round's results became complete (the latest
    certificate issue date within the round, declared/proposed iCoAs excluded) and
    numbered CoQ-PP-YYYY-NNNN in that order — one sequential series, proposed
    pending the Certificate Issuance Register (QCLB 020). Always compute this on
    the MASTER rows so every subset workbook cites identical codes.

    Returns (events, map) with map keyed by (batch, round_index).
    """
    panel = [no for no, _t, _u, _ac in MATRIX_PANEL]
    events = []
    for b in master_rows:
        for ri, rnd in enumerate(b["rounds"]):
            dates = [parse_date(ref.split(", ")[-2])
                     for no in panel for ref in rnd[no]["refs"]
                     if "(proposed)" not in ref]
            dates = [d for d in dates if d]
            new_items = [no for no in panel if rnd[no]["values"]]
            prior = {no for rj in range(ri) for no in panel
                     if b["rounds"][rj][no]["values"]}
            carried = [no for no in panel if no in prior and no not in new_items]
            covered = set(new_items) | prior
            events.append({
                "batch": b["batch"], "prod": b["prod"], "strain": b["strain"],
                "chrono": b["chrono"], "ri": ri,
                "round_label": b["round_labels"][ri],
                "ready": max(dates) if dates else None,
                "new": new_items, "carried": carried,
                "missing": [no for no in panel if no not in covered],
            })
    far = datetime.date(9999, 12, 31)
    events.sort(key=lambda e: (e["ready"] or far, e["chrono"], e["ri"]))
    for i, e in enumerate(events, start=1):
        e["code"] = "CoQ-PP-%d-%04d" % (COQ_YEAR, i)
    cmap = {(e["batch"], e["ri"]): e for e in events}
    for e in events:
        e["supersedes"] = cmap[(e["batch"], e["ri"] - 1)]["code"] if e["ri"] else ""
    return events, cmap


def apply_coq(rows, cmap):
    """Attach each sub-row's proposed CoQ code; batches outside the plan get '—'."""
    for b in rows:
        b["coq_codes"] = [
            cmap.get((b["batch"], ri), {}).get("code", "—")
            for ri in range(len(b["rounds"]))
        ]


def cert_index(listing):
    """Unique certificates with the batches, spec items and result count of each."""
    idx = OrderedDict()
    for rec in listing:
        if rec["cert"] == "—" or rec["value"].startswith("Missing"):
            continue
        key = (rec["cert"], rec["date"], rec["lab"], rec["link"])
        e = idx.setdefault(key, {"batches": [], "items": [], "n": 0})
        if rec["batch"] not in e["batches"]:
            e["batches"].append(rec["batch"])
        if rec["no"] not in e["items"]:
            e["items"].append(rec["no"])
        e["n"] += 1
    rows = []
    for (cert, date, lab, link), e in idx.items():
        rows.append({"cert": cert, "date": date,
                     "lab": REF_LAB.get(lab, lab), "link": link, "n": e["n"],
                     "batches": ", ".join(e["batches"]),
                     "items": ", ".join(e["items"])})
    rows.sort(key=lambda r: (r["date"] or "9999", r["cert"]))
    return rows


def coverage(rows):
    """Per spec parameter: batches with at least one result vs batches missing."""
    out = []
    for no, title, unit, ac in MATRIX_PANEL:
        tested = sum(1 for b in rows
                     if any(rnd[no]["values"] for rnd in b["rounds"]))
        out.append((no, title, unit, ac, tested, len(rows) - tested))
    return out


def group_cols():
    """(item no -> (first column, width in columns)) for the parameter groups."""
    out, col = {}, ROUND_COL + 1
    for no, _t, _u, _ac in MATRIX_PANEL:
        w = 3 if no == THC_ITEM else 2
        out[no] = (col, w)
        col += w
    return out, col - 1


ARIAL = dict(name="Arial", size=8)


def set_result_cell(c, text, color=INK):
    """Store a clean numeric as a real number, preserving the printed decimals."""
    if "\n" not in text and NUM_CELL_RE.match(text):
        c.value = float(text)
        dec = len(text.split(".")[1]) if "." in text else 0
        c.number_format = ("0." + "0" * dec) if dec else "0"
        c.font = Font(color=color, **ARIAL)
        c.alignment = Alignment(horizontal="right", vertical="top")
        return
    c.value = text
    c.font = Font(color=color, **ARIAL)
    c.alignment = Alignment(vertical="top", wrap_text=True)


def write_readme(wb, rows, index_rows, n_lines, n_missing):
    ws = wb.create_sheet("Read me", 0)
    ws.sheet_properties.tabColor = GOLD
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin", color="D6DEEA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def line(r, text, size=8, bold=False, italic=False, color=INK, col=1):
        c = ws.cell(r, col, text)
        c.font = Font(name="Arial", size=size, bold=bold, italic=italic, color=color)
        c.alignment = Alignment(vertical="top", wrap_text=(col == 1))
        return c

    ws.column_dimensions["A"].width = 118
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 13

    line(1, READ_TITLE, 15, bold=True, color=NAVY)
    line(2, READ_SUBTITLE, 9, italic=True, color="555555")

    n_declared = sum(1 for b in rows for rnd in b["rounds"]
                     for no, _t, _u, _ac in MATRIX_PANEL
                     if rnd[no].get("declared"))
    r = 4
    line(r, "SOURCE AND SCOPE", 9, bold=True, color=NAVY); r += 1
    for t in [
        "Source: %s." % bl.SOURCE_NOTE,
        "%d batches · %d sub-rows (testing rounds) · %d result lines "
        "(%d transcribed + %d declared in-house Conforms: FM and macro/micro ID) · "
        "%d parameter cells Missing / not tested · %d certificates indexed."
        % (len(rows), sum(len(b["rounds"]) for b in rows), n_lines,
           n_lines - n_declared, n_declared, n_missing, len(index_rows)),
        "Values are transcribed from the certificates verbatim; measurement "
        "uncertainty is excluded from result cells; units live in the column header, "
        "not the cells. Clean numeric results are stored as numbers whose format "
        "preserves the certificate's printed decimals.",
        "'Conforms' entries whose reference reads '(proposed)' are declared by the "
        "QC Manager (13.08.2026): Foreign Matter was determined in-house per SOP "
        "prior to final QC sampling and packaging, and macroscopic + microscopic "
        "identification per specification was performed in-house after final QC "
        "sampling — all results pass. Each cites the iCoA proposed for issuance "
        "(iCoA-PP-[YYYY]-[NNNN] per QCSOP 012 v.03, one sequential series for the "
        "iCoA type per calendar year; codes and issue dates pending the Certificate "
        "Issuance Register, QCLB 020). One combined ID iCoA backs both "
        "identification cells of a batch. The full chronological plan is on the "
        "'iCoA issuance list' sheet; batches whose Foreign Matter or macroscopic ID "
        "was tested at CNP or already recorded in-house keep that reference.",
        "Grade cells classify each THC assay against the cultivar's QCSP 001 v.02 "
        "ladder. Classification is not a disposition: no conformity is derived or "
        "colour-coded anywhere in this workbook — a failure exists only where the "
        "laboratory declared one.",
    ]:
        line(r, "•  " + t); ws.row_dimensions[r].height = 24; r += 1

    if EXTRA_COVER:
        heading, headers, widths, data = EXTRA_COVER[:4]
        cover_note = EXTRA_COVER[4] if len(EXTRA_COVER) > 4 else None
        r += 1
        line(r, heading, 9, bold=True, color=NAVY); r += 1
        for ci, (h, wd) in enumerate(zip(headers, widths), start=1):
            c = ws.cell(r, ci, h)
            c.font = Font(name="Arial", size=8, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", start_color=NAVY)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
            if ci > 1:
                ws.column_dimensions[get_column_letter(ci)].width = wd
        r += 1
        for row_vals in data:
            for ci, v in enumerate(row_vals, start=1):
                c = ws.cell(r, ci, v)
                c.font = Font(name="Arial", size=8,
                              bold=(row_vals[0] or "").startswith("TOTAL"), color=INK)
                c.alignment = Alignment(
                    horizontal="right" if isinstance(v, (int, float)) else "left",
                    vertical="top")
                c.border = border
            r += 1
        if cover_note:
            line(r, cover_note, italic=True, color="8A6D1C")
            ws.row_dimensions[r].height = 24
            r += 1

    r += 1
    line(r, "HOW TO READ THE MATRIX", 9, bold=True, color=NAVY); r += 1
    for t in [
        "One row-block per batch. Batches tested in more than one campaign split "
        "into sub-rows, one per testing round (certificate issue dates more than %d "
        "days apart open a new round); the Round column labels each sub-row with its "
        "month span." % ROUND_GAP_DAYS,
        "Per parameter: Result cell, then the certificate reference 'code, date, "
        "lab', hyperlinked to the scan on Drive. Same-day paired certificates stack "
        "one line each inside a sub-row; the cell's hyperlink opens the first listed "
        "scan, and the Certificate index sheet carries every certificate's own link.",
        "'Missing / not tested' spans all sub-rows when the knowledgebase holds no "
        "result for that parameter; '—' marks a parameter tested in another round of "
        "the same batch, but not this one.",
        "Column C proposes the CoQ document code of each TESTING ROUND — one CoQ "
        "issuance event per sub-row, per QCSOP 012 v.03 §6.4.2 (CoQ-PP-[YYYY]-"
        "[NNNN], one sequential series per calendar year). The initial round's CoQ "
        "issues on completion of the initial QC sampling; a resampling (R) round "
        "produces a superseding CoQ with a new issue date that carries the "
        "resampled parameters' new results and reuses the non-resampled results "
        "from the earlier sampling's eCoAs (time-invariant parameters such as "
        "heavy metals and pesticides). Codes run in results-complete chronology — "
        "see the 'CoQ issuance plan' sheet for each event's basis date, resampled "
        "vs carried-forward parameters and supersession chain. Final numbers are "
        "assigned from the Certificate Issuance Register (QCLB 020) at approval — "
        "if any %d CoQ numbers are already consumed there, the sequence starts "
        "after the last used one." % COQ_YEAR,
    ]:
        line(r, "•  " + t); ws.row_dimensions[r].height = 32; r += 1

    r += 1
    line(r, "LABORATORY ABBREVIATIONS", 9, bold=True, color=NAVY); r += 1
    for ab, name in LAB_LEGEND:
        c = ws.cell(r, 1, "%s — %s" % (ab, name))
        c.font = Font(name="Arial", size=8, color=INK)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 13
        r += 1

    r += 1
    line(r, "COVERAGE PER SPECIFICATION PARAMETER (of %d batches)" % len(rows),
         9, bold=True, color=NAVY); r += 1
    heads = ["№", "Parameter", "Unit", "A.C.", "Tested", "Missing"]
    widths = [6, 34, 8, 30, 8, 8]
    for ci, (h, w) in enumerate(zip(heads, widths), start=1):
        c = ws.cell(r, ci, h)
        c.font = Font(name="Arial", size=8, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
        if ci > 1:
            ws.column_dimensions[get_column_letter(ci)].width = w
    r += 1
    for no, title, unit, ac, tested, missing in coverage(rows):
        vals = [no, title, unit or "—", ac, tested, missing]
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(r, ci, v)
            c.font = Font(name="Arial", size=8,
                          color=MISS if (ci == 6 and missing == 0) else INK)
            c.alignment = Alignment(
                horizontal="center" if ci in (1, 3, 5, 6) else "left",
                vertical="top")
            c.border = border
        r += 1

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.oddFooter.left.text = "Purely Plant — QC · informal working export"
    ws.oddFooter.right.text = "Page &P of &N"
    return ws


def write_xlsx(rows, index_rows, n_lines, n_missing, coq_events=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Spec matrix"
    ws.sheet_properties.tabColor = NAVY
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85
    thin = Side(style="thin", color="D6DEEA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    groups, n_cols = group_cols()

    def put(r, c, v, **font):
        cell = ws.cell(r, c)
        cell.value = v
        cell.font = Font(**dict(ARIAL, **font))
        cell.border = border
        return cell

    def fill_range(r1, c1, r2, c2, color):
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                ws.cell(r, c).fill = PatternFill("solid", start_color=color)
                ws.cell(r, c).border = border

    # row 1 — compact title line (the full reading guide lives on the Read me sheet)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = put(1, 1, MATRIX_TITLE, italic=True, color="555555")
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 16

    # rows 2-4 — frozen headers: parameter (with №+unit), A.C., sub-labels
    for i, (title, _w) in enumerate(IDENTITY + [("Round", ROUND_W)]):
        col = i + 1
        ws.merge_cells(start_row=2, start_column=col, end_row=4, end_column=col)
        c = put(2, col, title, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        fill_range(2, col, 4, col, NAVY)

    for no, title, unit, ac in MATRIX_PANEL:
        col, w = groups[no]
        head = "%s — %s" % (no, title) + (" (%s)" % unit if unit else "")
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + w - 1)
        c = put(2, col, head, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        fill_range(2, col, 2, col + w - 1, NAVY)

        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + w - 1)
        c = put(3, col, "A.C.  %s" % ac, bold=True, color=INK)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        fill_range(3, col, 3, col + w - 1, GOLD_TINT)

        labels = (["Result", "Grade (v.02)", "eCoA (code, date, lab)"] if w == 3
                  else ["Result", "eCoA (code, date, lab)"])
        for off, label in enumerate(labels):
            c = put(4, col + off, label, bold=True, color="4A5B6C")
            c.fill = PatternFill("solid", start_color=SUB_GREY)
            c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = 26
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 13

    # data — one row-block per batch, one sub-row per testing round
    r = 5
    for bi, b in enumerate(rows):
        n_r = len(b["rounds"])
        band = (bi % 2 == 1)
        fill = BAND if band else None
        r_top, r_bot = r, r + n_r - 1

        ident = [b["chrono"], b["prod"], None, b["batch"], b["p"], b["strain"],
                 b["code"]]
        for cidx, v in enumerate(ident, start=1):
            if cidx == 3:
                # the CoQ code belongs to the testing round, not the batch — one
                # issuance event per sub-row, so no vertical merge here
                for ri in range(n_r):
                    c = put(r_top + ri, cidx,
                            b.get("coq_codes", ["—"] * n_r)[ri], color=INK)
                    c.alignment = Alignment(vertical="top", wrap_text=True)
                if fill:
                    fill_range(r_top, cidx, r_bot, cidx, fill)
                continue
            if n_r > 1:
                ws.merge_cells(start_row=r_top, start_column=cidx,
                               end_row=r_bot, end_column=cidx)
            c = put(r_top, cidx, v, bold=(cidx == 4), color=INK)
            c.alignment = Alignment(vertical="top", wrap_text=(cidx >= 6),
                                    horizontal="center" if cidx <= 2 else "left")
            if fill:
                fill_range(r_top, cidx, r_bot, cidx, fill)
            else:
                for rr in range(r_top, r_bot + 1):
                    ws.cell(rr, cidx).border = border

        # any wrapped fixed text ("Missing / not tested") needs two lines of height
        line_counts = [2 if any(not any(rnd[no]["values"] for rnd in b["rounds"])
                                for no, _t, _u, _ac in MATRIX_PANEL) else 1
                       for rnd in b["rounds"]]

        for ri in range(n_r):
            c = put(r_top + ri, ROUND_COL, b["round_labels"][ri], color="4A5B6C",
                    italic=True)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                c.fill = PatternFill("solid", start_color=fill)

        for no, _t, _u, _ac in MATRIX_PANEL:
            col, w = groups[no]
            tested_rounds = [ri for ri in range(n_r)
                             if b["rounds"][ri][no]["values"]]
            if not tested_rounds:
                # never tested for this batch — one statement spanning the sub-rows
                for off in range(w):
                    if n_r > 1:
                        ws.merge_cells(start_row=r_top, start_column=col + off,
                                       end_row=r_bot, end_column=col + off)
                    txt = "Missing / not tested" if off == 0 else "—"
                    c = put(r_top, col + off, txt, color=MISS, italic=True)
                    c.alignment = Alignment(vertical="top", wrap_text=(off == 0),
                                            horizontal="left" if off == 0 else "center")
                    if fill:
                        fill_range(r_top, col + off, r_bot, col + off, fill)
                    else:
                        for rr in range(r_top, r_bot + 1):
                            ws.cell(rr, col + off).border = border
                continue

            for ri in range(n_r):
                cell = b["rounds"][ri][no]
                rr = r_top + ri
                if not cell["values"]:
                    # tested in another round, not this one
                    for off in range(w):
                        c = put(rr, col + off, "—", color=MISS)
                        c.alignment = Alignment(vertical="top", horizontal="center")
                        if fill:
                            c.fill = PatternFill("solid", start_color=fill)
                    continue
                line_counts[ri] = max(line_counts[ri], len(cell["values"]))
                c = ws.cell(rr, col)
                c.border = border
                set_result_cell(c, "\n".join(cell["values"]))
                off = 1
                if w == 3:
                    c = put(rr, col + 1, "\n".join(cell["grades"]), color=INK)
                    c.alignment = Alignment(vertical="top", wrap_text=True)
                    off = 2
                if cell.get("declared"):
                    c = put(rr, col + off, "\n".join(cell["refs"]),
                            color="8A6D1C", italic=True)
                else:
                    c = put(rr, col + off, "\n".join(cell["refs"]), color=LINK,
                            underline="single")
                c.alignment = Alignment(vertical="top", wrap_text=True)
                first_link = next((l for l in cell["links"] if l), None)
                if first_link:
                    c.hyperlink = first_link
                if fill:
                    for o in range(w):
                        ws.cell(rr, col + o).fill = PatternFill("solid",
                                                                start_color=fill)

        for ri in range(n_r):
            ws.row_dimensions[r_top + ri].height = max(24, 11 * line_counts[ri] + 3)
        r = r_bot + 1

    for i, (_t, w) in enumerate(IDENTITY):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.column_dimensions[get_column_letter(ROUND_COL)].width = ROUND_W
    for no, _t, _u, _ac in MATRIX_PANEL:
        col, w = groups[no]
        ws.column_dimensions[get_column_letter(col)].width = RESULT_W.get(no, 11)
        if w == 3:
            ws.column_dimensions[get_column_letter(col + 1)].width = GRADE_W
        ws.column_dimensions[get_column_letter(col + w - 1)].width = REF_W
    ws.freeze_panes = "G5"

    # print: A3 landscape, headers and identity repeat on every page
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.print_title_rows = "2:4"
    ws.print_title_cols = "A:F"
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.5
    ws.oddFooter.left.text = "PP Spec Parameter Matrix — informal working export"
    ws.oddFooter.center.text = "data as transcribed from the eCoA scans"
    ws.oddFooter.right.text = "Page &P of &N"

    # sheet 3 — the proposed CoQ issuance plan: one event per testing round
    if coq_events:
        wsq = wb.create_sheet("CoQ issuance plan")
        wsq.sheet_properties.tabColor = NAVY
        wsq.sheet_view.showGridLines = False
        wsq.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
        c = wsq.cell(1, 1, (
            "Proposed CoQ issuance plan — one Certificate of Quality per testing "
            "round of each batch, in results-complete chronology (the latest "
            "certificate issue date within the round; declared iCoAs excluded from "
            "the basis date). The initial round's CoQ issues on completion of the "
            "initial QC sampling; a resampling (R) round produces a superseding CoQ "
            "with a new issue date carrying the resampled parameters' new results "
            "and reusing the non-resampled results from the earlier sampling's "
            "eCoAs (time-invariant parameters such as heavy metals and pesticides "
            "— 'Carried forward' below). Supersession per QCSOP 012 v.03 §6.7 "
            "('Supersedes [original] — reason: resampling'). Codes are proposed "
            "pending the Certificate Issuance Register (QCLB 020)."))
        c.font = Font(italic=True, color="555555", **ARIAL)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        wsq.row_dimensions[1].height = 50
        heads = [("№", 5), ("CoQ code (proposed)", 19), ("Results complete", 13),
                 ("Production", 10), ("Batch", 13), ("Strain", 17), ("Round", 13),
                 ("Newly tested / resampled (items)", 26),
                 ("Carried forward (items)", 22), ("Not covered (items)", 20),
                 ("Supersedes", 19)]
        for cidx, (h, w) in enumerate(heads, start=1):
            c = wsq.cell(2, cidx, h)
            c.font = Font(bold=True, color="FFFFFF", **ARIAL)
            c.fill = PatternFill("solid", start_color=NAVY)
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
            c.border = border
            wsq.column_dimensions[get_column_letter(cidx)].width = w
        for i, e in enumerate(coq_events, start=1):
            ready = (e["ready"].strftime("%d.%m.%Y") if e["ready"]
                     else "date not printed")
            resample = e["ri"] > 0
            vals = [int(e["code"][-4:]), e["code"], ready, e["prod"], e["batch"],
                    e["strain"], e["round_label"], ", ".join(e["new"]) or "—",
                    ", ".join(e["carried"]) or ("—" if not resample else "—"),
                    ", ".join(e["missing"]) or "—", e["supersedes"] or "—"]
            for cidx, v in enumerate(vals, start=1):
                c = wsq.cell(2 + i, cidx, v)
                c.font = Font(color=INK, bold=(cidx == 2 and resample), **ARIAL)
                c.alignment = Alignment(
                    vertical="top", wrap_text=(cidx in (8, 9, 10)),
                    horizontal="center" if cidx in (1, 3, 4, 7) else "left")
                c.border = border
            if resample:
                for cidx in range(1, len(heads) + 1):
                    wsq.cell(2 + i, cidx).fill = PatternFill(
                        "solid", start_color=GOLD_TINT)
        wsq.freeze_panes = "A3"
        wsq.auto_filter.ref = "A2:K%d" % (len(coq_events) + 2)
        wsq.page_setup.orientation = "landscape"
        wsq.page_setup.paperSize = wsq.PAPERSIZE_A4
        wsq.page_setup.fitToWidth = 1
        wsq.page_setup.fitToHeight = 0
        wsq.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        wsq.print_title_rows = "2:2"

    # sheet 4 — the proposed iCoA issuance plan, chronological
    icoa_rows = proposed_icoa_rows(rows)
    if icoa_rows:
        wsi = wb.create_sheet("iCoA issuance list")
        wsi.sheet_properties.tabColor = GOLD
        wsi.sheet_view.showGridLines = False
        wsi.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
        c = wsi.cell(1, 1, (
            "Proposed iCoA issuance list — internal Certificates of Analysis to be "
            "issued for in-house determinations performed per SOP: Foreign Matter "
            "prior to final QC sampling and packaging; macroscopic + microscopic "
            "identification (one combined iCoA) after final QC sampling — all "
            "results pass, declared by the QC Manager %s. Numbering per QCSOP 012 "
            "v.03: one sequential series for the iCoA type per calendar year; codes "
            "and issue dates are proposed pending assignment from the Certificate "
            "Issuance Register (QCLB 020). Chronological by the production date "
            "encoded in the batch number." % ICOA_DATE))
        c.font = Font(italic=True, color="555555", **ARIAL)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        wsi.row_dimensions[1].height = 44
        heads = [("№", 5), ("iCoA code (proposed)", 20), ("Scope", 44),
                 ("Spec items", 10), ("Production", 10), ("Batch", 13),
                 ("P-Number", 10), ("Strain", 18), ("Proposed issue date", 16)]
        for cidx, (h, w) in enumerate(heads, start=1):
            c = wsi.cell(2, cidx, h)
            c.font = Font(bold=True, color="FFFFFF", **ARIAL)
            c.fill = PatternFill("solid", start_color=NAVY)
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
            c.border = border
            wsi.column_dimensions[get_column_letter(cidx)].width = w
        for i, e in enumerate(icoa_rows, start=1):
            vals = [i, e["cert"], e["scope"], e["items"], e["prod"], e["batches"],
                    e["p"], e["strain"], e["date"]]
            for cidx, v in enumerate(vals, start=1):
                c = wsi.cell(2 + i, cidx, v)
                c.font = Font(color=INK, **ARIAL)
                c.alignment = Alignment(
                    vertical="top", wrap_text=(cidx == 3),
                    horizontal="center" if cidx in (1, 4, 5, 9) else "left")
                c.border = border
        wsi.freeze_panes = "A3"
        wsi.auto_filter.ref = "A2:I%d" % (len(icoa_rows) + 2)
        wsi.page_setup.orientation = "landscape"
        wsi.page_setup.paperSize = wsi.PAPERSIZE_A4
        wsi.page_setup.fitToWidth = 1
        wsi.page_setup.fitToHeight = 0
        wsi.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        wsi.print_title_rows = "2:2"

    # sheet 4 — every certificate, on file or proposed, with its own link
    wsx = wb.create_sheet("Certificate index")
    wsx.sheet_properties.tabColor = SUB_GREY
    wsx.sheet_view.showGridLines = False
    for cidx, (h, w) in enumerate([("Certificate code", 22), ("Type", 14),
                                   ("Issued", 11), ("Lab", 6), ("Results", 8),
                                   ("Batches", 34), ("Spec items", 22),
                                   ("Scan", 8)], start=1):
        c = wsx.cell(1, cidx, h)
        c.font = Font(bold=True, color="FFFFFF", **ARIAL)
        c.fill = PatternFill("solid", start_color=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
        wsx.column_dimensions[get_column_letter(cidx)].width = w
    for i, e in enumerate(index_rows, start=2):
        proposed = e.get("proposed", False)
        ctype = ("iCoA (proposed)" if proposed
                 else "iCoA / in-house" if e["lab"] == "PP" else "eCoA")
        vals = [e["cert"], ctype, e["date"], e["lab"], e["n"], e["batches"],
                e["items"]]
        for cidx, v in enumerate(vals, start=1):
            c = wsx.cell(i, cidx, v)
            c.font = Font(color="8A6D1C" if proposed else INK,
                          italic=proposed, **ARIAL)
            c.alignment = Alignment(vertical="top",
                                    wrap_text=(cidx in (6, 7)),
                                    horizontal="center" if cidx in (2, 3, 4, 5)
                                    else "left")
            c.border = border
        c = wsx.cell(i, 8)
        if e["link"]:
            c.value = "open"
            c.hyperlink = e["link"]
            c.font = Font(color=LINK, underline="single", **ARIAL)
        elif proposed:
            c.value = "pending"
            c.font = Font(color="8A6D1C", italic=True, **ARIAL)
        else:
            c.value = "—"
            c.font = Font(color=MISS, **ARIAL)
        c.alignment = Alignment(horizontal="center", vertical="top")
        c.border = border
    wsx.freeze_panes = "A2"
    wsx.auto_filter.ref = "A1:H%d" % (len(index_rows) + 1)
    wsx.page_setup.orientation = "landscape"
    wsx.page_setup.paperSize = wsx.PAPERSIZE_A4
    wsx.page_setup.fitToWidth = 1
    wsx.page_setup.fitToHeight = 0
    wsx.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    wsx.print_title_rows = "1:1"

    write_readme(wb, rows, index_rows, n_lines, n_missing)
    wb.active = wb.sheetnames.index("Spec matrix")
    wb.save(OUT_XLSX)


def write_tsv(rows):
    with open(OUT_TSV, "w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh, delimiter="\t", lineterminator="\n")
        head1 = [t.replace("\n", " ") for t, _w in IDENTITY] + ["Round"]
        head2 = [""] * (N_ID + 1)
        for no, title, unit, ac in MATRIX_PANEL:
            head1.append("%s — %s%s" % (no, title, " (%s)" % unit if unit else ""))
            head2.append("A.C. %s" % ac)
            if no == THC_ITEM:
                head1.append("Grade (v.02)")
                head2.append("")
            head1.append("eCoA reference")
            head2.append("")
        w.writerow(head1)
        w.writerow(head2)
        for b in rows:
            n_r = len(b["rounds"])
            for ri in range(n_r):
                coq = b.get("coq_codes", ["—"] * n_r)[ri]
                if ri == 0:
                    line = [b["chrono"], b["prod"], coq, b["batch"], b["p"],
                            b["strain"], b["code"]]
                else:
                    line = ["", "", coq, "", "", "", ""]
                line.append(b["round_labels"][ri])
                for no, _t, _u, _ac in MATRIX_PANEL:
                    cell = b["rounds"][ri][no]
                    tested_any = any(b["rounds"][x][no]["values"] for x in range(n_r))
                    if cell["values"]:
                        line.append(" ; ".join(cell["values"]))
                        if no == THC_ITEM:
                            line.append(" ; ".join(cell["grades"]))
                        line.append(" ; ".join(cell["refs"]))
                    else:
                        txt = "—" if tested_any else (
                            "Missing / not tested" if ri == 0 else "")
                        line.append(txt)
                        if no == THC_ITEM:
                            line.append("—" if txt else "")
                        line.append("—" if txt else "")
                w.writerow(line)


def main():
    listing, stats, n_batches = bl.build_rows()
    rows = pivot(listing)
    coq_events, cmap = coq_plan(rows)
    apply_coq(rows, cmap)
    index_rows = cert_index(listing) + proposed_icoa_rows(rows)

    # the matrix must hold exactly the fact-checked totals — nothing lost, nothing
    # added beyond the declared Foreign-Matter entries, which are counted explicitly
    assert len(rows) == n_batches, (len(rows), n_batches)
    n_declared = sum(1 for b in rows for rnd in b["rounds"]
                     for no, _t, _u, _ac in MATRIX_PANEL
                     if rnd[no].get("declared"))
    n_lines = sum(len(rnd[no]["values"])
                  for b in rows for rnd in b["rounds"]
                  for no, _t, _u, _ac in MATRIX_PANEL)
    n_missing = sum(1 for b in rows for no, _t, _u, _ac in MATRIX_PANEL
                    if not any(rnd[no]["values"] for rnd in b["rounds"]))
    assert n_lines == stats["results"] + n_declared, \
        (n_lines, stats["results"], n_declared)
    assert n_missing == stats["missing"] - n_declared, \
        (n_missing, stats["missing"], n_declared)
    n_rounds = sum(len(b["rounds"]) for b in rows)
    multi = [(b["batch"], len(b["rounds"])) for b in rows if len(b["rounds"]) > 1]

    write_xlsx(rows, index_rows, n_lines, n_missing, coq_events=coq_events)
    write_tsv(rows)

    print("batches: %d   sub-rows: %d   multi-round batches: %d   "
          "result lines: %d (%d transcribed + %d declared in-house)   "
          "missing cells: %d   certs: %d   CoQ events: %d (%d superseding)"
          % (len(rows), n_rounds, len(multi), n_lines, n_lines - n_declared,
             n_declared, n_missing, len(index_rows), len(coq_events),
             sum(1 for e in coq_events if e["supersedes"])))
    for name, k in multi:
        print("  %s: %d rounds" % (name, k))
    print("wrote %s (%d KB)" % (os.path.relpath(OUT_XLSX, HERE),
                                os.path.getsize(OUT_XLSX) // 1024))
    print("wrote %s (%d KB)" % (os.path.relpath(OUT_TSV, HERE),
                                os.path.getsize(OUT_TSV) // 1024))
    return 0


if __name__ == "__main__":
    sys.exit(main())
