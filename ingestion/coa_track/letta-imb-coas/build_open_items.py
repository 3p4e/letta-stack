#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Write the open reconciliation items — the ones only the owner can close.

Everything the certificates could settle has been settled and is already in the register.
What is left divides into two kinds of question, and neither is a script's to answer:

  * a batch appears on one side and not the other, or
  * a name or a released potency differs between a control sheet and the certificate.

Each row states both sides and, where a certificate is involved, names it, so the answer
can be checked against the document rather than taken on trust. The "Where it stands"
column says what is known, not what should be done.
"""
import csv
import os
import re
import sys

import coa_pivot as cp
import crosscheck_sources as cc
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "exports", "PP_Open_Reconciliation_Items.xlsx")

HEAD = PatternFill("solid", fgColor="1F3864")
BAND = PatternFill("solid", fgColor="EEF2F8")
GREY = PatternFill("solid", fgColor="F5F5F5")
THIN = Side(style="thin", color="B4C6E7")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = [("#", 5), ("Class", 7), ("Batch", 15), ("Question", 30),
        ("Control sheet says", 26), ("Knowledgebase / certificate says", 34),
        ("Certificate", 20), ("Where it stands", 62)]


def sheets():
    tr = {cc.norm_batch(r["batch"]): r for r in cc.read_tsv("tranches_overview.tsv")}
    qt = {cc.norm_batch(r["batch_label"]): r for r in cc.read_tsv("quantities_table.tsv")}
    lc = {}
    for r in cc.read_tsv("list_of_coas.tsv"):
        if (r.get("batch") or "").strip():
            lc.setdefault(cc.norm_batch(r["batch"]), {}).update(
                {k: v for k, v in r.items() if (v or "").strip()})
    return tr, qt, lc


def main():
    tr, qt, lc = sheets()
    rows, batches = cp.build()
    by_seq = cp.group_by_seq(rows)

    rag = {}
    for b in batches:
        rag[cc.norm_batch(b["batch"])] = {
            "seq": b["seq"], "strain": b["strain"],
            "p": (b["p_number"] or "").strip(),
            "thc": cc.thc_values(by_seq[b["seq"]]),
        }

    out = []

    # --- A: on the sheets, no certificate in the knowledgebase ---------------
    for k in sorted(set(tr) | set(qt) | set(lc)):
        if k in rag:
            continue
        lot = ((qt.get(k) or {}).get("p_number") or "").strip().upper()
        if lot and any(r["p"].upper() == lot for r in rag.values()):
            continue
        src = ", ".join(n for n, d in (("Tranches", tr), ("Quantities", qt),
                                       ("List of COAs", lc)) if k in d)
        thc = (tr.get(k) or lc.get(k) or {}).get("thc_pct", "")
        note = ("No certificate for this batch is in the knowledgebase under either the "
                "batch number or its packaging lot. Either it was never tested by an "
                "outsourced lab, or its CoA has not been filed into the Drive folder.")
        out.append(["A", k, "Batch has no certificate on file",
                    "listed on %s%s" % (src, (", THC %s" % thc) if thc else ""),
                    "absent", "", note])

    # --- B: in the knowledgebase, on no sheet -------------------------------
    for k in sorted(rag):
        if k in tr or k in qt or k in lc:
            continue
        certs = [c for c in {r["Certificate Code"].strip()
                             for r in by_seq[rag[k]["seq"]]} if c][:2]
        note = ("Certificates exist for this batch but it is on none of the three control "
                "sheets. If it is a real released batch the sheets are short a row.")
        out.append(["B", k, "Batch is on no control sheet", "absent",
                    "seq %s, %s" % (rag[k]["seq"], rag[k]["strain"]),
                    ", ".join(certs), note])

    # --- D: strain name ------------------------------------------------------
    for k in sorted(rag):
        want = None
        for d, col in ((tr, "strain"), (lc, "strain")):
            if k in d and (d[k].get(col) or "").strip():
                want = d[k][col].strip()
                break
        if not want or cc.norm_strain(want) == cc.norm_strain(rag[k]["strain"]):
            continue
        note = ("Two spellings of one cultivar. The certificate spelling is what an "
                "outsourced lab printed; the sheet spelling is the house name. A CoQ "
                "should carry one of them consistently.")
        out.append(["D", k, "Strain name spelled differently", want, rag[k]["strain"],
                    "", note])

    # --- F: released THC % ---------------------------------------------------
    for k in sorted(rag):
        want = cc.num((tr.get(k) or {}).get("thc_pct"))
        src = "Tranches Overview"
        if want is None:
            want = cc.num((lc.get(k) or {}).get("thc_pct"))
            src = "List of COAs"
        if want is None:
            continue
        vals = [v for v, _c, _d in rag[k]["thc"]]
        if not vals or any(abs(v - want) <= 0.005 for v in vals):
            continue
        near = min(vals, key=lambda v: abs(v - want))
        cert = next((c for v, c, _d in rag[k]["thc"] if v == near), "")
        ongoing = ((qt.get(k) or {}).get("thc_int") or "").strip().lower()
        if ongoing.startswith("anal"):
            note = ("The Quantities table marks this batch 'anal.ong.' — analysis ongoing "
                    "— so the sheet figure was entered before the certificate arrived. "
                    "The certificate figure is the later one.")
        elif abs(want - round(want)) < 1e-9:
            note = ("The sheet figure is a whole number while the certificate carries two "
                    "decimals, which reads as a provisional entry rather than a "
                    "transcription.")
        elif src == "List of COAs":
            note = ("The List of COAs figure is rounded; the certificate value is the "
                    "precise one.")
        else:
            note = ("Both sides are precise and they disagree. Worth checking which "
                    "certificate was used to release the batch.")
        out.append(["F", k, "Released THC %% differs by %.2f" % abs(near - want),
                    "%.2f %% (%s)" % (want, src),
                    " / ".join("%.2f %%" % v for v in sorted(set(vals))), cert, note])

    wb = Workbook()
    ws = wb.active
    ws.title = "Open Items"

    ws["A1"] = "Purely Plant — Open Reconciliation Items"
    ws["A1"].font = Font(bold=True, size=14, color="1F3864")
    ws["A2"] = ("Control sheets vs. the CoA knowledgebase. Everything the certificates "
                "could settle has been settled in the register; these are the questions "
                "that need the owner.")
    ws["A2"].font = Font(size=9, italic=True, color="555555")
    ws["A3"] = ("A = batch has no certificate    B = batch on no sheet    "
                "D = strain name    F = released THC %")
    ws["A3"].font = Font(size=9, color="555555")

    for i, (name, width) in enumerate(COLS, start=1):
        c = ws.cell(row=5, column=i, value=name)
        c.fill, c.font = HEAD, Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[5].height = 30

    out.sort(key=lambda r: (r[0], r[1]))
    for n, rec in enumerate(out, start=1):
        r = 5 + n
        for i, v in enumerate([n] + rec, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BOX
            c.font = Font(size=9)
            c.alignment = Alignment(vertical="top", wrap_text=(i >= 4))
            if n % 2 == 0:
                c.fill = BAND
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row=r, column=3).font = Font(size=9, bold=True)
        ws.cell(row=r, column=8).fill = GREY if n % 2 else PatternFill()

    ws.freeze_panes = "D6"
    ws.auto_filter.ref = "A5:H%d" % (5 + len(out))
    wb.save(OUT)

    print("Wrote %s" % OUT)
    for cl in ("A", "B", "D", "F"):
        print("   class %s: %d" % (cl, sum(1 for r in out if r[0] == cl)))
    print("   total open items: %d" % len(out))
    return 0


if __name__ == "__main__":
    sys.exit(main())
