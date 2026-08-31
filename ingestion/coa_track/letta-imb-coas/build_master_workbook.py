#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build the Purely Plant eCoA master workbook.

Ten sheets, every fact drawn from the ImB_QC_COAs corpus (the Letta source of
that name was retired 19.08.2026; the certificates are now in RAGFlow):

  Read Me            conventions, legend and the data-integrity rules this file follows
  Batch Register     one row per tested batch: identity, disposition, coverage, labs
  Release Panel      one row per batch across the release-critical parameters
  Coverage Matrix    batch x analytical panel, so a missing panel is visible at a glance
  QC Exceptions      every declared non-conformance, laboratory finding and data flag
  Certificates       one row per certificate: code, date of issue, lab, what it covers
  Laboratories       the accredited laboratories, their scope and volume
  Strain Summary     potency range and batch count per strain
  Stability Studies  ICH storage results, kept apart from release disposition
  Full Data          all parameter records, one row each

Usage:  python3 build_master_workbook.py [input.tsv] [output.xlsx]
"""
import os
import re
import sys
from collections import Counter, OrderedDict, defaultdict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import coa_pivot as cp

HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_OUT = os.path.join(HERE, "exports", "PP_eCoA_Master_Database.xlsx")
FONT = "Arial"

INK = "16232B"
PASS_BG, PASS_FG = "E3F3E9", "1B7F4B"
WARN_BG, WARN_FG = "FAEDD4", "9A6300"
CRIT_BG, CRIT_FG = "F9DEDB", "AE2318"
BAND = "F4F6F5"
ACCENT = "0E6E6E"

# Release-critical reads for the Release Panel sheet.
RELEASE = [
    ("identification", "Identification", 17),
    ("foreign_matter", "Foreign matter", 17),
    ("loss_on_drying", "Loss on drying", 15),
    ("thc", "Total Δ9-THC", 24),
    ("cbd", "Total CBD", 19),
    ("cbn", "Total CBN", 17),
    ("aflatoxins", "Aflatoxins", 21),
    ("ochratoxin", "Ochratoxin A", 22),
    ("pb", "Lead", 13), ("cd", "Cadmium", 13), ("as_", "Arsenic", 13), ("hg", "Mercury", 13),
    ("pesticides", "Pesticide screen", 30),
    ("tamc", "TAMC", 15), ("tymc", "TYMC", 15), ("bile", "Bile-tol. GNB", 16),
    ("ecoli", "E. coli", 15), ("salmonella", "Salmonella", 15),
]

# Analytical panels for the coverage matrix.
PANELS = OrderedDict([
    ("Identity & physical", ["appearance", "identification", "foreign_matter", "loss_on_drying"]),
    ("Potency", ["thc", "cbd", "cbn", "cbda", "cbd_n", "cbn_n", "thc_n", "thca"]),
    ("Mycotoxins", ["aflatoxins", "afla_b1", "afla_b2", "afla_g1", "afla_g2", "ochratoxin"]),
    ("Heavy metals", ["pb", "cd", "as_", "hg", "cu", "metals_panel"]),
    ("Pesticides", ["pesticides"]),
    ("Microbiology", ["tamc", "tymc", "bile", "ecoli", "salmonella", "micro_other"]),
    ("Packaging", ["packaging"]),
])

STATUS_TEXT = {"complete": "Complete", "flag": "Complete (flag)",
               "partial": "Partial", "open": "OPEN QC ISSUE"}

ACCRED_RE = re.compile(r"\b(?:LT|ЛТ|JT)\s*-\s*(\d{3})\b", re.I)
NUM_RE = re.compile(r"(\d+(?:[.,]\d+)?)\s*%")


class Style(object):
    """Fonts and fills built once and reused across sheets."""

    def __init__(self):
        self.hdr = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        self.hdr_fill = PatternFill("solid", fgColor=INK)
        self.base = Font(name=FONT, size=10)
        self.bold = Font(name=FONT, size=10, bold=True)
        self.mono = Font(name="Consolas", size=10)
        self.small = Font(name=FONT, size=9, color="5A6E75")
        self.faint = Font(name=FONT, size=10, color="8B9EA3")
        self.link = Font(name=FONT, size=9, color=ACCENT, underline="single")
        self.title = Font(name=FONT, size=16, bold=True, color=INK)
        self.sub = Font(name=FONT, size=10, color="5A6E75", italic=True)
        self.section = Font(name=FONT, size=11, bold=True, color=ACCENT)
        self.pass_ = (PatternFill("solid", fgColor=PASS_BG), Font(name=FONT, size=10, color=PASS_FG, bold=True))
        self.warn = (PatternFill("solid", fgColor=WARN_BG), Font(name=FONT, size=10, color=WARN_FG, bold=True))
        self.crit = (PatternFill("solid", fgColor=CRIT_BG), Font(name=FONT, size=10, color=CRIT_FG, bold=True))
        self.band = PatternFill("solid", fgColor=BAND)
        thin = Side(style="thin", color="DBE3E1")
        self.box = Border(left=thin, right=thin, top=thin, bottom=thin)

    def status(self, key):
        return {"complete": self.pass_, "flag": self.pass_,
                "partial": self.warn, "open": self.crit}[key]


def header_row(ws, row, headers, widths, S, height=26):
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = S.hdr
        c.fill = S.hdr_fill
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = height
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_title(ws, title, subtitle, S, width_note=None):
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = S.title
    ws["A2"] = subtitle
    ws["A2"].font = S.sub
    if width_note:
        ws["A3"] = width_note
        ws["A3"].font = Font(name=FONT, size=9, color="8B9EA3")


def put_link(ws, row, col, url, label, S):
    c = ws.cell(row=row, column=col, value=label if url else "")
    if url:
        c.hyperlink = url
        c.font = S.link
    return c


def parse_pct(text):
    """Leading percentage as a float, for sorting and ranges only.

    Returns None when the value is not a plain percentage (BLQ, <LOQ, a range, prose),
    so nothing is ever coerced into a number it does not state.
    """
    if not text:
        return None
    head = text.split("|")[0].strip()
    if head.startswith("<") or head.startswith("≤") or "loq" in head.lower():
        return None
    m = NUM_RE.search(head)
    if not m:
        return None
    try:
        return float(m.group(1).replace(",", "."))
    except ValueError:
        return None


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else None
    out_path = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUT
    rows, batches = cp.build(src)
    by_seq = cp.group_by_seq(rows)
    S = Style()
    wb = openpyxl.Workbook()

    # =================================================== 1. Read Me
    ws = wb.active
    ws.title = "Read Me"
    sheet_title(ws, "Purely Plant GmbH — eCoA Master Database",
                "Outsourced laboratory certificates for every tested batch, compiled from "
                "the ImB_QC_COAs corpus (Letta source retired 19.08.2026; the same "
                "certificates are now the RAGFlow eCoA dataset).", S)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 104
    r = 4
    facts = [("Batches", len(batches)), ("Parameter records", len(rows)),
             ("Accredited laboratories", cp.count_labs(rows)),
             ("Certificates referenced", len({(x["Certificate Code"].strip(),
                                               x["Issuing Institution"].strip())
                                              for x in rows if x["Certificate Code"].strip()})),
             ("Drive folder", "16oMK_j0FUusjveV61B5rxWi6Sl5kQsn5")]
    for k, v in facts:
        ws.cell(row=r, column=1, value=k).font = S.bold
        ws.cell(row=r, column=2, value=v).font = S.base
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="SHEETS").font = S.section
    r += 1
    for name, desc in [
        ("Batch Register", "One row per tested batch: identity, disposition, panel coverage, laboratories."),
        ("Release Panel", "One row per batch across the release-critical parameters."),
        ("Coverage Matrix", "Batch against analytical panel — a missing panel is visible immediately."),
        ("QC Exceptions", "Every declared non-conformance, laboratory finding and data-integrity flag."),
        ("Certificates", "One row per certificate: code, date of issue, issuing laboratory, coverage."),
        ("Laboratories", "The accredited laboratories, their accreditation number and volume of work."),
        ("Strain Summary", "Batches and potency range per strain."),
        ("Stability Studies", "ICH storage results. Never a substitute for release values."),
        ("Full Data", "Every parameter record, one row each, with its certificate."),
    ]:
        ws.cell(row=r, column=1, value=name).font = S.bold
        c = ws.cell(row=r, column=2, value=desc)
        c.font = S.base
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="DISPOSITION").font = S.section
    r += 1
    for label, desc in [
        ("Complete", "All core Ph. Eur. 3028 panels are on file for the batch."),
        ("Complete (flag)", "All panels on file; a certificate carries a data-integrity flag."),
        ("Partial", "One or more analytical panels have no certificate on file."),
        ("OPEN QC ISSUE", "The issuing laboratory declared a non-conformance that is unresolved."),
    ]:
        c = ws.cell(row=r, column=1, value=label)
        key = {"Complete": "complete", "Complete (flag)": "flag",
               "Partial": "partial", "OPEN QC ISSUE": "open"}[label]
        fill, fnt = S.status(key)
        c.fill, c.font = fill, fnt
        d = ws.cell(row=r, column=2, value=desc)
        d.font = S.base
        d.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="DATA INTEGRITY").font = S.section
    r += 1
    for line in [
        "Values appear exactly as printed on the certificate. N.D., <LOQ, BLQ and the Macedonian "
        "Одговара / Не одговара verdicts are preserved verbatim.",
        "A pass or fail is never derived from a value. A result is called a failure only where the "
        "issuing laboratory declared one, and highlighted as a finding only where the laboratory "
        "itself wrote that something was detected.",
        "Where a laboratory prints results without a limit column, the acceptance criterion is drawn "
        "from the Purely Plant release specification / Ph. Eur. 3028 and marked as added for review.",
        "Where a certificate contradicts itself, the row carries a data-integrity flag rather than a "
        "silent correction, and the discrepancy is raised with the issuing laboratory.",
        "Date of issue only. Analysis and receipt dates printed alongside it are not reproduced.",
    ]:
        c = ws.cell(row=r, column=2, value="• " + line)
        c.font = S.base
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30
        r += 1

    # =================================================== 2. Batch Register
    ws = wb.create_sheet("Batch Register")
    sheet_title(ws, "Batch Register", "Every tested batch, in chronological sequence.", S)
    heads = ["Seq", "Cultivation batch", "P-number", "Strain", "Disposition",
             "Panels on file", "Parameter records", "Certificates", "Laboratories",
             "Notes", "Batch folder"]
    widths = [6, 20, 12, 20, 17, 14, 12, 12, 40, 62, 12]
    header_row(ws, 4, heads, widths, S)
    ws.freeze_panes = "A5"

    batch_folder = {}
    for b in batches:
        for rec in by_seq[b["seq"]]:
            raw = rec["Drive File Link"] or ""
            if "/folders/" in raw:
                batch_folder[b["seq"]] = cp.clean_link(raw)
                break

    r = 5
    for b in batches:
        brows = by_seq[b["seq"]]
        panels_present = sum(1 for _n, keys in PANELS.items()
                             if any(cp.classify(x["Parameter"]) in keys for x in brows))
        certs = {x["Certificate Code"].strip() for x in brows if x["Certificate Code"].strip()
                 and x["Certificate Code"].strip().lower() not in ("n/a", "na")}
        labs = set()
        for x in brows:
            for name, rx, _acc in cp.LAB_PATTERNS:
                if rx.search(x["Issuing Institution"] or ""):
                    labs.add(name.split("—")[0].strip())
        ws.cell(row=r, column=1, value=b["seq"]).font = S.base
        ws.cell(row=r, column=2, value=b["batch"]).font = S.bold
        ws.cell(row=r, column=3, value=b["p_number"]).font = S.base
        ws.cell(row=r, column=4, value=b["strain"]).font = S.base
        sc = ws.cell(row=r, column=5, value=STATUS_TEXT[b["status"]])
        fill, fnt = S.status(b["status"])
        sc.fill, sc.font = fill, fnt
        ws.cell(row=r, column=6, value="%d of %d" % (panels_present, len(PANELS))).font = S.base
        ws.cell(row=r, column=7, value=len(brows)).font = S.base
        ws.cell(row=r, column=8, value=len(certs)).font = S.base
        ws.cell(row=r, column=9, value=", ".join(sorted(labs))).font = S.small
        nc = ws.cell(row=r, column=10, value=b["notes"])
        nc.font = Font(name=FONT, size=9, color=CRIT_FG if b["status"] == "open" else "5A6E75")
        put_link(ws, r, 11, batch_folder.get(b["seq"], ""), "Folder", S)
        if b["seq"] % 2 == 0:
            for c in range(1, len(heads) + 1):
                if c != 5:
                    ws.cell(row=r, column=c).fill = S.band
        r += 1
    ws.auto_filter.ref = "A4:%s%d" % (get_column_letter(len(heads)), r - 1)


    # =================================================== 2b. QC Results
    # Modelled on the Purely Plant "Production Batches QC Result Table": acceptance criteria
    # stated once in a reference row under the header, parameters as columns with their units,
    # one row per batch, and a further row per additional laboratory. "/" marks a parameter
    # that report did not cover. A pesticide screen collapses to the house notation - the
    # limit when nothing was found, the named compounds when something was.
    ws = wb.create_sheet("QC Results")
    sheet_title(ws, "Production Batches — QC Results",
                "Every tested batch in chronological order. Acceptance criteria are stated once "
                "in the reference row; where a certificate prints a different criterion, Full "
                "Data holds it verbatim.", S)

    GROUPS = [
        ("", ["Seq", "Batch number", "P-number", "Strain",
              "CoA code", "Date of issue", "Issuing institution", "PDF"]),
        ("IDENTITY & PHYSICAL", ["Appearance", "Identification", "Foreign matter",
                                 "Loss on drying %", "LoD vs 3028"]),
        ("POTENCY (release)", ["Total THC spec", "Total THC %", "vs spec",
                               "Total CBD %", "Total CBN %"]),
        ("CANNABINOID PROFILE", ["CBDA", "Δ⁹-THCA", "CBD", "CBN", "Δ⁹-THC"]),
        ("MYCOTOXINS", ["Aflatoxin B1", "Aflatoxins Σ", "Ochratoxin A"]),
        ("HEAVY METALS", ["Pb", "Cd", "As", "Hg", "Cu"]),
        ("PESTICIDES", ["Pesticide screen"]),
        ("MICROBIOLOGY", ["TAMC CFU/g", "TYMC CFU/g", "Bile-tolerant GNB /1 g",
                          "E. coli /1 g", "Salmonella /25 g"]),
        ("PACKAGING", ["Packaging"]),
    ]
    FIELDS = ["appearance", "identification", "foreign_matter", "loss_on_drying", "lod_check",
              "thc_spec", "thc", "thc_check", "cbd", "cbn",
              "cbda", "thca", "cbd_n", "cbn_n", "thc_n",
              "afla_b1", "aflatoxins", "ochratoxin",
              "pb", "cd", "as_", "hg", "cu",
              "pesticides",
              "tamc", "tymc", "bile", "ecoli", "salmonella",
              "packaging"]
    MYCO_KEYS = ("afla_b1", "aflatoxins", "ochratoxin")
# The microbial enumeration criteria are stated as the maximum acceptable count
    # the pharmacopoeia gives, not as the bare power of ten and not as the figure the
    # Purely Plant in-house CoA form prints.
    #
    # Ph. Eur. 5.1.4, in identical PDG-harmonised wording USP <1111>, and again in the
    # "Interpretation of results" text of Ph. Eur. 2.6.12 / USP <61>:
    #
    #   "10^1 CFU: maximum acceptable count = 20; 10^2 CFU: maximum acceptable
    #    count = 200; 10^3 CFU: maximum acceptable count = 2000, and so forth."
    #
    # x2 per decade: 10^4 -> 20 000, 10^5 -> 200 000.
    #
    # CORRECTED 31.08.2026. These cells read "max 500 000" and "max 50 000" — a x5
    # reading, taken verbatim from the in-house CoA form (see add_gg1024_rows.py,
    # which transcribes "<10^5, max 500 000 CFU/g" off the GG1024, HPA1024 and
    # OPM1024 release CoAs). No pharmacopoeial text uses a x5 multiplier for
    # microbial limits. The transcription in add_gg1024_rows.py stays as the paper
    # has it; this dict states the criterion and is corrected.
    #
    # ACTION FOR QA, outside this file: the in-house CoA form states a maximum
    # acceptable count 2.5 times the compendial one and needs correcting at source.
    # Record: review/OOS_RECTIFICATION_2026-08-31.md
    SPEC = {
        "appearance": "Characteristic flower & odour",
        "identification": "Conforms (Ph. Eur. 2.8.23)",
        "foreign_matter": "≤ 2.00 %, no seed, no leaves > 1 cm (2.8.2)",
        "loss_on_drying": "≤ 12.00 % (Ph. Eur. 07/2024:3028)",
        "lod_check": "computed vs 3028",
        "thc_spec": "per batch, % w/w (label claim)",
        "thc": "per batch (see Total THC spec)",
        "thc_check": "computed",
        "cbd": "< 1.00 %", "cbn": "< 1.00 %",
        "cbda": "profile", "thca": "profile", "cbd_n": "profile",
        "cbn_n": "profile", "thc_n": "profile",
        "afla_b1": "≤ 2 µg/kg",
        "aflatoxins": "≤ 4 µg/kg (Σ B1+B2+G1+G2)",
        "ochratoxin": "per PP release spec",
        "pb": "≤ 0.5 mg/kg", "cd": "≤ 0.3 mg/kg",
        "as_": "≤ 0.2 mg/kg", "hg": "≤ 0.1 mg/kg",
        "cu": "informational (no limit)",
        "pesticides": "≤ LOQ 0.01 mg/kg (Ph. Eur. 2.8.13)",
        "tamc": "≤ 10⁵ CFU/g   (max 200 000)",
        "tymc": "≤ 10⁴ CFU/g   (max 20 000)",
        "bile": "≤ 10⁴ CFU/g", "ecoli": "Absent", "salmonella": "Absent",
        "packaging": "Conforms",
    }
    ID_N = len(GROUPS[0][1])
    widths = ([5, 20, 11, 16, 18, 13, 20, 7]
              + [16, 15, 15, 15, 11]
              + [20, 16, 10, 12, 12]
              + [11, 11, 11, 11, 11]
              + [13, 14, 14]
              + [14, 14, 14, 14, 14]
              + [34]
              + [15, 15, 20, 14, 14]
              + [22])
    heads = [h for _g, hs in GROUPS for h in hs]

    # group band
    col = 1
    for gname, hs in GROUPS:
        if gname:
            ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + len(hs) - 1)
            gc = ws.cell(row=4, column=col, value=gname)
            gc.font = Font(name=FONT, size=8, bold=True, color="FFFFFF")
            gc.fill = PatternFill("solid", fgColor=ACCENT)
            gc.alignment = Alignment(horizontal="center", vertical="center")
        col += len(hs)
    ws.row_dimensions[4].height = 14
    header_row(ws, 5, heads, widths, S, height=42)
    for ci in range(1, len(heads) + 1):
        ws.cell(row=5, column=ci).alignment = Alignment(
            horizontal="left", vertical="bottom", wrap_text=True)

    # reference-values row
    rv = ws.cell(row=6, column=1, value="Laboratory reference values")
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=ID_N)
    rv.font = Font(name=FONT, size=9, bold=True, color=INK)
    rv.fill = PatternFill("solid", fgColor="E7ECEA")
    rv.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    for j, key in enumerate(FIELDS):
        c = ws.cell(row=6, column=ID_N + 1 + j, value=SPEC.get(key, "/"))
        c.font = Font(name=FONT, size=8, bold=True, color=INK)
        c.fill = PatternFill("solid", fgColor="E7ECEA")
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[6].height = 30
    ws.freeze_panes = "I7"

    ND_RE = re.compile(r"^(n\.?\s?d\.?|н\.?\s?д\.?|not detected|≤\s*loq|<\s*loq|nd)\b", re.I)

    r = 7
    qc_rows = 0
    for b in batches:
        groups = OrderedDict()
        for rec in by_seq[b["seq"]]:
            code = rec["Certificate Code"].strip() or "(not numbered)"
            key = (code, cp.issue_date(rec["Issue Date"]))
            g = groups.setdefault(key, {"link": "", "lab": "", "vals": defaultdict(list)})
            if not g["link"]:
                g["link"] = cp.clean_link(rec["Drive File Link"])
            if not g["lab"]:
                inst = rec["Issuing Institution"] or ""
                for name, rx, _acc in cp.LAB_PATTERNS:
                    if rx.search(inst):
                        g["lab"] = name.split("—")[0].strip()
                        break
            k = cp.classify(rec["Parameter"])
            if k in FIELDS:
                g["vals"][k].append((cp.short_label(rec["Parameter"]), rec["Result"].strip()))
                if k == "thc":
                    ac = (rec["Acceptance Criterion"] or "").strip()
                    if ac and ac not in ("/", "—"):
                        g["vals"]["thc_spec"].append(("", ac))
        if not groups:
            groups[("(none)", "")] = {"link": "", "lab": "", "vals": defaultdict(list)}

        first = True
        block_start = r
        for (code, date), g in groups.items():
            myco_present = any(g["vals"].get(k) for k in MYCO_KEYS)
            idf = S.base if first else S.faint
            ws.cell(row=r, column=1, value=b["seq"]).font = idf
            ws.cell(row=r, column=2, value=b["batch"]).font = S.bold if first else S.faint
            ws.cell(row=r, column=3, value=b["p_number"]).font = idf
            ws.cell(row=r, column=4, value=b["strain"]).font = idf
            ws.cell(row=r, column=5, value=code).font = S.base
            ws.cell(row=r, column=6, value=date).font = S.base
            ws.cell(row=r, column=7, value=g["lab"]).font = S.small
            put_link(ws, r, 8, g["link"], "Open", S)

            for j, key in enumerate(FIELDS):
                vals = g["vals"].get(key) or []
                if not vals:
                    blank = ("not tested" if (key in MYCO_KEYS and myco_present) else "/")
                    c = ws.cell(row=r, column=ID_N + 1 + j, value=blank)
                    c.font = S.faint
                    c.alignment = Alignment(
                        horizontal="left" if blank == "not tested" else "center")
                    continue
                if key in ("thc_check", "lod_check"):
                    continue
                if key == "pesticides":
                    hits = [(lb, v) for lb, v in vals if not ND_RE.match(v)]
                    if hits:
                        text = "; ".join("%s %s" % (lb, v) for lb, v in hits)
                    else:
                        text = Counter(v for _, v in vals).most_common(1)[0][0]
                elif key == "thc_spec":
                    text = cp.clean_thc_spec(vals[0][1])
                elif len(vals) == 1:
                    text = vals[0][1]
                else:
                    text = "; ".join("%s %s" % (lb, v) for lb, v in vals)
                c = ws.cell(row=r, column=ID_N + 1 + j, value=text)
                sev = cp.severity(text)
                if sev == "crit":
                    c.fill, c.font = S.crit
                elif sev == "warn":
                    c.fill, c.font = S.warn
                else:
                    c.font = S.base
                c.alignment = Alignment(vertical="center", wrap_text=True)
            thc_vals = g["vals"].get("thc") or []
            spec_vals = g["vals"].get("thc_spec") or []
            verdict, ok = "", None
            if thc_vals and spec_vals:
                got = parse_pct(thc_vals[0][1])
                lo, hi = cp.spec_range(spec_vals[0][1])
                if got is not None and lo is not None:
                    ok = lo <= got <= hi
                    verdict = "within" if ok else "OUTSIDE"
                elif got is not None and cp.spec_minimum(spec_vals[0][1]) is not None:
                    mn = cp.spec_minimum(spec_vals[0][1])
                    ok = got >= mn
                    verdict = "within" if ok else "OUTSIDE"
            cc = ws.cell(row=r, column=ID_N + 1 + FIELDS.index("thc_check"),
                         value=verdict or "/")
            if verdict == "within":
                cc.fill, cc.font = S.pass_
            elif verdict == "OUTSIDE":
                cc.fill, cc.font = S.crit
            else:
                cc.font = S.faint
            cc.alignment = Alignment(horizontal="center")
            lod_vals = g["vals"].get("loss_on_drying") or []
            lv = parse_pct(lod_vals[0][1]) if lod_vals else None
            lverdict = ("within" if lv <= 12.0 else "OUTSIDE") if lv is not None else ""
            lc2 = ws.cell(row=r, column=ID_N + 1 + FIELDS.index("lod_check"),
                          value=lverdict or "/")
            if lverdict == "within":
                lc2.fill, lc2.font = S.pass_
            elif lverdict == "OUTSIDE":
                lc2.fill, lc2.font = S.crit
            else:
                lc2.font = S.faint
            lc2.alignment = Alignment(horizontal="center")
            first = False
            r += 1
            qc_rows += 1
        for ci in range(1, len(heads) + 1):
            ws.cell(row=r - 1, column=ci).border = Border(
                bottom=Side(style="thin", color="C6D0CD"))

    ws.auto_filter.ref = "A5:%s%d" % (get_column_letter(len(heads)), r - 1)

    # legend, in the house style of the source table
    r += 1
    ws.cell(row=r, column=1, value="LEGEND").font = S.section
    r += 1
    for txt, fill in [
        ("n.d — not detected;   BLQ — below limit of quantification;   "
         "LOQ — limit of quantification;   \"/\" — parameter not covered by that report", None),
        ("Red — the issuing laboratory declared a deviation from specification", CRIT_BG),
        ("Amber — a laboratory finding or a data-integrity flag on the certificate", WARN_BG),
        ("A batch tested by more than one laboratory has one row per certificate; "
         "identity is repeated in grey on the continuation rows.", None),
    ]:
        c = ws.cell(row=r, column=1, value=txt)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        c.font = Font(name=FONT, size=9)
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(vertical="center", indent=1)
        r += 1

    # =================================================== 4. Coverage Matrix
    ws = wb.create_sheet("Coverage Matrix")
    sheet_title(ws, "Coverage Matrix",
                "Records held per analytical panel. An empty cell means no certificate on file "
                "for that panel — not a failing result.", S)
    heads = ["Seq", "Batch", "P-number", "Strain"] + list(PANELS.keys()) + ["Panels on file"]
    widths = [6, 20, 12, 18] + [19] * len(PANELS) + [14]
    header_row(ws, 4, heads, widths, S, height=34)
    ws.freeze_panes = "E5"
    r = 5
    for b in batches:
        brows = by_seq[b["seq"]]
        counts = Counter(cp.classify(x["Parameter"]) for x in brows)
        ws.cell(row=r, column=1, value=b["seq"]).font = S.base
        ws.cell(row=r, column=2, value=b["batch"]).font = S.bold
        ws.cell(row=r, column=3, value=b["p_number"]).font = S.base
        ws.cell(row=r, column=4, value=b["strain"]).font = S.base
        present = 0
        for j, (_name, keys) in enumerate(PANELS.items()):
            n = sum(counts.get(k, 0) for k in keys)
            c = ws.cell(row=r, column=5 + j, value=n if n else "—")
            c.alignment = Alignment(horizontal="center")
            if n:
                present += 1
                c.fill, c.font = S.pass_
            else:
                c.font = S.faint
        tot = ws.cell(row=r, column=5 + len(PANELS), value="%d of %d" % (present, len(PANELS)))
        tot.font = S.bold if present == len(PANELS) else S.base
        tot.alignment = Alignment(horizontal="center")
        r += 1
    ws.auto_filter.ref = "A4:%s%d" % (get_column_letter(len(heads)), r - 1)

    # =================================================== 5. QC Exceptions
    ws = wb.create_sheet("QC Exceptions")
    sheet_title(ws, "QC Exceptions",
                "Declared non-conformances, laboratory findings and data-integrity flags. "
                "Nothing here is a judgement made by this workbook.", S)
    heads = ["Type", "Seq", "Batch", "P-number", "Strain", "Parameter",
             "Acceptance criterion", "Result", "Certificate code", "Date of issue",
             "Certificate"]
    widths = [20, 6, 18, 12, 18, 44, 32, 50, 20, 15, 12]
    header_row(ws, 4, heads, widths, S)
    ws.freeze_panes = "A5"
    r = 5
    exc = 0
    for b in batches:
        for rec in by_seq[b["seq"]]:
            sev = cp.severity(rec["Result"])
            gap = cp.GAP_RE.match(rec["Parameter"])
            if not sev and not gap:
                continue
            if gap:
                kind, style = "Coverage gap", S.warn
            elif sev == "crit":
                kind, style = "Non-conformance", S.crit
            elif "data integrity flag" in rec["Result"].lower():
                kind, style = "Data-integrity flag", S.warn
            else:
                kind, style = "Laboratory finding", S.warn
            tc = ws.cell(row=r, column=1, value=kind)
            tc.fill, tc.font = style
            ws.cell(row=r, column=2, value=b["seq"]).font = S.base
            ws.cell(row=r, column=3, value=b["batch"]).font = S.bold
            ws.cell(row=r, column=4, value=b["p_number"]).font = S.base
            ws.cell(row=r, column=5, value=b["strain"]).font = S.base
            pc = ws.cell(row=r, column=6, value=rec["Parameter"])
            pc.font = S.base
            pc.alignment = Alignment(wrap_text=True, vertical="top")
            ac = ws.cell(row=r, column=7, value=rec["Acceptance Criterion"])
            ac.font = S.small
            ac.alignment = Alignment(wrap_text=True, vertical="top")
            rc = ws.cell(row=r, column=8, value=rec["Result"])
            rc.font = S.base
            rc.alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=r, column=9, value=rec["Certificate Code"]).font = S.base
            ws.cell(row=r, column=10, value=cp.issue_date(rec["Issue Date"])).font = S.base
            put_link(ws, r, 11, cp.clean_link(rec["Drive File Link"]), "Open", S)
            r += 1
            exc += 1
    ws.auto_filter.ref = "A4:%s%d" % (get_column_letter(len(heads)), max(r - 1, 5))

    # =================================================== 6. Certificates
    ws = wb.create_sheet("Certificates")
    sheet_title(ws, "Certificates",
                "One row per certificate referenced by the register.", S)
    heads = ["Certificate code", "Date of issue", "Issuing laboratory", "Accreditation",
             "Batches covered", "Parameters", "Panels", "Certificate"]
    widths = [24, 14, 44, 14, 34, 12, 40, 12]
    header_row(ws, 4, heads, widths, S)
    ws.freeze_panes = "A5"

    certs = OrderedDict()
    for rec in rows:
        code = rec["Certificate Code"].strip()
        if not code or code.lower() in ("n/a", "na"):
            continue
        inst = rec["Issuing Institution"].strip()
        key = (code, inst)
        e = certs.setdefault(key, {"date": cp.issue_date(rec["Issue Date"]),
                                   "batches": set(), "params": 0, "panels": set(), "link": ""})
        e["batches"].add(rec["Cultivation Batch"])
        e["params"] += 1
        k = cp.classify(rec["Parameter"])
        for pname, keys in PANELS.items():
            if k in keys:
                e["panels"].add(pname)
        if not e["link"]:
            e["link"] = cp.clean_link(rec["Drive File Link"])
    r = 5
    for (code, inst), e in sorted(certs.items(), key=lambda kv: (kv[0][1], kv[0][0])):
        ws.cell(row=r, column=1, value=code).font = S.bold
        ws.cell(row=r, column=2, value=e["date"]).font = S.base
        ws.cell(row=r, column=3, value=inst).font = S.base
        accred = ""
        for _n, _rx, _a in cp.LAB_PATTERNS:
            if _rx.search(inst) and _a:
                accred = _a
                break
        if not accred:
            m = ACCRED_RE.search(inst)
            accred = ("LT-%s" % m.group(1)) if m else ""
        ws.cell(row=r, column=4, value=accred).font = S.small
        ws.cell(row=r, column=5, value=", ".join(sorted(e["batches"]))).font = S.small
        ws.cell(row=r, column=6, value=e["params"]).font = S.base
        ws.cell(row=r, column=7, value=", ".join(sorted(e["panels"]))).font = S.small
        put_link(ws, r, 8, e["link"], "Open", S)
        r += 1
    ws.auto_filter.ref = "A4:%s%d" % (get_column_letter(len(heads)), max(r - 1, 5))
    n_certs = len(certs)

    # =================================================== 7. Laboratories
    ws = wb.create_sheet("Laboratories")
    sheet_title(ws, "Laboratories", "Accredited laboratories issuing the certificates in this register.", S)
    heads = ["Laboratory", "Accreditation", "Certificates", "Parameter records",
             "Batches served", "Panels performed"]
    widths = [52, 14, 14, 18, 16, 46]
    header_row(ws, 4, heads, widths, S)
    ws.freeze_panes = "A5"
    labstat = OrderedDict()
    lab_accred = {}
    for name, rx, accred in cp.LAB_PATTERNS:
        lab_accred[name] = accred
        labstat[name] = {"accred": set(), "certs": set(), "params": 0,
                         "batches": set(), "panels": set()}
    for rec in rows:
        inst = rec["Issuing Institution"] or ""
        for name, rx, accred in cp.LAB_PATTERNS:
            if rx.search(inst):
                e = labstat[name]
                if accred:
                    e["accred"].add(accred)
                e["params"] += 1
                e["batches"].add(rec["Cultivation Batch"])
                code = rec["Certificate Code"].strip()
                if code and code.lower() not in ("n/a", "na"):
                    e["certs"].add(code)
                m = ACCRED_RE.search(inst)
                if m:
                    e["accred"].add("LT-%s" % m.group(1))
                k = cp.classify(rec["Parameter"])
                for pname, keys in PANELS.items():
                    if k in keys:
                        e["panels"].add(pname)
                break
    r = 5
    for name, e in labstat.items():
        if not e["params"]:
            continue
        ws.cell(row=r, column=1, value=name).font = S.bold
        ws.cell(row=r, column=2, value=", ".join(sorted(e["accred"]))).font = S.base
        ws.cell(row=r, column=3, value=len(e["certs"])).font = S.base
        ws.cell(row=r, column=4, value=e["params"]).font = S.base
        ws.cell(row=r, column=5, value=len(e["batches"])).font = S.base
        c = ws.cell(row=r, column=6, value=", ".join(sorted(e["panels"])))
        c.font = S.small
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    # =================================================== 8. Strain Summary
    ws = wb.create_sheet("Strain Summary")
    sheet_title(ws, "Strain Summary",
                "Per strain. Potency figures are parsed from the Total Δ9-THC result for ranking "
                "only; values that are not a plain percentage (BLQ, <LOQ) are excluded from the range.", S)
    heads = ["Strain", "Batches", "Batch codes", "Total Δ9-THC lowest",
             "Total Δ9-THC highest", "Batches with an open issue", "Batches with partial coverage"]
    widths = [24, 10, 56, 20, 20, 22, 24]
    header_row(ws, 4, heads, widths, S)
    ws.freeze_panes = "A5"
    strains = defaultdict(lambda: {"b": [], "thc": [], "open": 0, "partial": 0})
    for b in batches:
        e = strains[b["strain"] or "(not stated)"]
        e["b"].append(b["batch"])
        v = parse_pct(b["vals"]["thc"]["value"])
        if v is not None:
            e["thc"].append(v)
        if b["status"] == "open":
            e["open"] += 1
        if b["status"] == "partial":
            e["partial"] += 1
    r = 5
    for name in sorted(strains):
        e = strains[name]
        ws.cell(row=r, column=1, value=name).font = S.bold
        ws.cell(row=r, column=2, value=len(e["b"])).font = S.base
        c = ws.cell(row=r, column=3, value=", ".join(e["b"]))
        c.font = S.small
        ws.cell(row=r, column=4, value=("%.2f %%" % min(e["thc"])) if e["thc"] else "—").font = S.base
        ws.cell(row=r, column=5, value=("%.2f %%" % max(e["thc"])) if e["thc"] else "—").font = S.base
        oc = ws.cell(row=r, column=6, value=e["open"] or "")
        if e["open"]:
            oc.fill, oc.font = S.crit
        pc = ws.cell(row=r, column=7, value=e["partial"] or "")
        if e["partial"]:
            pc.fill, pc.font = S.warn
        r += 1
    ws.auto_filter.ref = "A4:%s%d" % (get_column_letter(len(heads)), max(r - 1, 5))

    # =================================================== 9. Stability Studies
    stab = [x for x in rows if cp.STABILITY_RE.search(x["Parameter"])]
    ws = wb.create_sheet("Stability Studies")
    sheet_title(ws, "Stability Studies",
                "Batches re-tested after storage (month 3 / 6 / 9 at 25 °C/60 % RH and 40 °C/75 % RH).",
                S, "These are not release results and must never be substituted for release values.")
    detail_heads = ["Seq", "Batch", "P-number", "Strain", "Parameter", "Acceptance criterion",
                    "Result", "Certificate code", "Date of issue", "Issuing institution",
                    "Certificate"]
    detail_widths = [6, 18, 11, 18, 50, 32, 38, 20, 15, 34, 12]
    header_row(ws, 5, detail_heads, detail_widths, S)
    ws.freeze_panes = "A6"
    write_detail(ws, 6, stab, S)
    if stab:
        ws.auto_filter.ref = "A5:K%d" % (len(stab) + 5)

    # =================================================== 10. Full Data
    ws = wb.create_sheet("Full Data")
    sheet_title(ws, "Full Data", "Every parameter record held for every batch.", S)
    header_row(ws, 4, detail_heads, detail_widths, S)
    ws.freeze_panes = "E5"
    write_detail(ws, 5, rows, S)
    ws.auto_filter.ref = "A4:K%d" % (len(rows) + 4)

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print("Wrote %s" % out_path)
    print("sheets=%d batches=%d records=%d certificates=%d exceptions=%d stability=%d qc_rows=%d"
          % (len(wb.sheetnames), len(batches), len(rows), n_certs, exc, len(stab), qc_rows))


def write_detail(ws, start, records, S):
    r = start
    for rec in records:
        ws.cell(row=r, column=1, value=rec["Seq"]).font = S.base
        ws.cell(row=r, column=2, value=rec["Cultivation Batch"]).font = S.base
        ws.cell(row=r, column=3, value=rec["P-Number"]).font = S.base
        ws.cell(row=r, column=4, value=rec["Strain"]).font = S.base
        pc = ws.cell(row=r, column=5, value=rec["Parameter"])
        pc.font = S.base
        pc.alignment = Alignment(wrap_text=True, vertical="top")
        ac = ws.cell(row=r, column=6, value=rec["Acceptance Criterion"])
        ac.font = S.small
        ac.alignment = Alignment(wrap_text=True, vertical="top")
        rc = ws.cell(row=r, column=7, value=rec["Result"])
        rc.alignment = Alignment(wrap_text=True, vertical="top")
        sev = cp.severity(rec["Result"])
        if sev == "crit":
            rc.fill, rc.font = S.crit
        elif sev == "warn":
            rc.fill, rc.font = S.warn
        else:
            rc.font = S.base
        ws.cell(row=r, column=8, value=rec["Certificate Code"]).font = S.base
        ws.cell(row=r, column=9, value=cp.issue_date(rec["Issue Date"])).font = S.base
        ic = ws.cell(row=r, column=10, value=rec["Issuing Institution"])
        ic.font = S.small
        ic.alignment = Alignment(wrap_text=True, vertical="top")
        put_link(ws, r, 11, cp.clean_link(rec["Drive File Link"]), "Open", S)
        r += 1
    return r


if __name__ == "__main__":
    main()
