#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Production Batches QC Result Table — Purely Plant.

A direct copy of the house example table (Replek-era "Production Batches QC Result
Table"): the same columns in the same order, the "Laboratory Reference values" row
directly under the header, one row per batch with a further row per additional
laboratory report, "/" where a report did not cover a parameter, and the legend at
the foot. Populated from the ImB_QC_COAs corpus (Letta source retired 19.08.2026;
the certificates are now in RAGFlow).

Adjustments the owner has directed, and nothing else:
  - result values stand alone in their cells; units live in the column headers
  - mycotoxins are three columns (Aflatoxin B1 / total aflatoxins / ochratoxin A),
    with "not tested" where a certificate assayed only the total
  - arsenic joins Pb, Hg, Cd — four heavy metals are tested
  - loss on drying reference is the current Ph. Eur. 07/2024:3028 limit (≤ 12 %)
  - every row carries the CoA code and date of issue beside the laboratory,
    with a link to the certificate PDF
"""
import os
import re
from collections import Counter, OrderedDict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import coa_pivot as cp

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "exports", "PP_Production_Batches_QC_Result_Table.xlsx")
FONT = "Arial"
HDR_BG = "16232B"
REF_BG = "E7ECEA"
CRIT_BG, CRIT_FG = "F9DEDB", "AE2318"
WARN_BG, WARN_FG = "FAEDD4", "9A6300"

ND_RE = re.compile(r"^(n\.?\s?d\.?|н\.?\s?д\.?|not detected|≤\s*loq|<\s*loq|nd)\b", re.I)
MYCO_KEYS = ("afla_b1", "aflatoxins", "ochratoxin")

# (key, header, reference value, width) — the example's columns, in its order.
COLUMNS = [
    (None,        "Seeding/Harvest",                          "/",          16),
    (None,        "Batch number",                             "/",          12),
    (None,        "Strain",                                   "/",          16),
    ("thc",       "THC %",                                    "≥ 5.0 %",    10),
    ("cbd",       "CBD %",                                    "< 1.0 %",    10),
    ("cbn",       "CBN %",                                    "< 1.0 %",    10),
    ("loss_on_drying", "Loss on drying %",                    "≤ 12 % (Ph. Eur. 3028)", 13),
    ("tamc",      "TAMC CFU/g",                               "≤ 10⁵ CFU/g — max acceptable count 500 000", 15),
    ("tymc",      "TYMC CFU/g",                               "≤ 10⁴ CFU/g — max acceptable count 50 000",  15),
    ("bile",      "Bile tolerant gram-negative bacteria /1 g", "≤ 10⁴ CFU/g", 17),
    ("salmonella", "Salmonella /25 g",                        "Absent",     12),
    ("ecoli",     "Escherichia coli /1 g",                    "Absent",     12),
    ("afla_b1",   "Aflatoxin B1 µg/kg",                       "< 2 µg/kg",  12),
    ("aflatoxins", "Total aflatoxins (B1,B2,G1,G2) µg/kg",    "< 4 µg/kg",  15),
    ("ochratoxin", "Ochratoxin A µg/kg",                      "per PP spec", 12),
    ("pb",        "Pb mg/kg",                                 "≤ 0.5",      10),
    ("hg",        "Hg mg/kg",                                 "≤ 0.1",      10),
    ("cd",        "Cd mg/kg",                                 "≤ 0.3",      10),
    ("as_",       "As mg/kg",                                 "≤ 0.2",      10),
    ("pesticides", "Absence of Pesticides",                   "LOQ (limit of quantification) 0.01 mg/kg", 22),
    (None,        "Laboratory",                               "/",          30),
    (None,        "CoA code",                                 "/",          20),
    (None,        "Date of issue",                            "/",          13),
    (None,        "PDF",                                      "/",          7),
]
PARAM_KEYS = [k for k, _h, _r, _w in COLUMNS if k]
COL_OF = {}
for _i, (_k, _h, _r, _w) in enumerate(COLUMNS, start=1):
    if _k:
        COL_OF[_k] = _i
    COL_OF[_h] = _i


def lab_of(inst):
    for name, rx, _acc in cp.LAB_PATTERNS:
        if rx.search(inst or ""):
            return name.split("—")[0].strip()
    if "purely plant" in (inst or "").lower() or "in-house" in (inst or "").lower():
        return "Purely Plant (in-house)"
    return (inst or "").strip()[:36]


def main():
    rows, batches = cp.build()
    by_seq = cp.group_by_seq(rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QC Result Table"
    ws.sheet_view.showGridLines = False

    hdr_font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor=HDR_BG)
    ref_font = Font(name=FONT, size=8, bold=True, color=HDR_BG)
    ref_fill = PatternFill("solid", fgColor=REF_BG)
    base = Font(name=FONT, size=10)
    bold = Font(name=FONT, size=10, bold=True)
    small = Font(name=FONT, size=9, color="5A6E75")
    faint = Font(name=FONT, size=9, color="9AA9AD", italic=True)
    link_f = Font(name=FONT, size=9, color="0E6E6E", underline="single")
    crit = (PatternFill("solid", fgColor=CRIT_BG), Font(name=FONT, size=10, bold=True, color=CRIT_FG))
    warn = (PatternFill("solid", fgColor=WARN_BG), Font(name=FONT, size=10, color=WARN_FG))
    hair = Side(style="thin", color="C6D0CD")

    ncol = len(COLUMNS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    ws["A1"] = "Purely Plant GmbH — Production Batches QC Result Table"
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color=HDR_BG)

    HROW, REFROW, DATA0 = 3, 4, 5
    for ci, (_k, head, ref, w) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=HROW, column=ci, value=head)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
        rc = ws.cell(row=REFROW, column=ci, value=ref)
        rc.font = ref_font
        rc.fill = ref_fill
        rc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(row=REFROW, column=1, value="Laboratory Reference values")
    ws.row_dimensions[HROW].height = 44
    ws.row_dimensions[REFROW].height = 34
    ws.freeze_panes = "D%d" % DATA0

    r = DATA0
    for b in batches:
        groups = OrderedDict()
        for rec in by_seq[b["seq"]]:
            if cp.STABILITY_RE.search(rec["Parameter"] or ""):
                continue
            k = cp.classify(rec["Parameter"])
            if k not in PARAM_KEYS and k != "cbn_n":
                continue
            code = rec["Certificate Code"].strip() or "(not numbered)"
            gkey = (code, cp.issue_date(rec["Issue Date"]))
            g = groups.setdefault(gkey, {"inst": lab_of(rec["Issuing Institution"] or ""),
                                         "link": cp.clean_link(rec["Drive File Link"]),
                                         "vals": {}})
            g["vals"].setdefault(k, []).append(rec["Result"].strip())
        if not groups:
            continue

        first = True
        for (code, date), g in groups.items():
            myco_present = any(g["vals"].get(k) for k in MYCO_KEYS)
            idfont = bold if first else faint
            ws.cell(row=r, column=1, value=b["batch"] if first else "").font = idfont
            ws.cell(row=r, column=2, value=(b["p_number"] or "/") if first else "").font = (base if first else faint)
            ws.cell(row=r, column=3, value=b["strain"] if first else "").font = (base if first else faint)

            for key in PARAM_KEYS:
                ci = COL_OF[key]
                vals = g["vals"].get(key)
                # A PP/UKIM certificate prints "Cannabinol" / "CBN content" without the word
                # "Total"; that reading classifies to the profile bucket. The CBN % column
                # takes it when no Total-CBN value exists on that certificate.
                if key == "cbn" and not vals:
                    vals = g["vals"].get("cbn_n")
                if not vals:
                    txt = "not tested" if (key in MYCO_KEYS and myco_present) else "/"
                    cc = ws.cell(row=r, column=ci, value=txt)
                    cc.font = faint
                    cc.alignment = Alignment(horizontal="center")
                    continue
                if key == "pesticides":
                    hits = [v for v in vals if not ND_RE.match(v)]
                    text = ("; ".join(cp.clean_value(v) for v in hits) if hits
                            else cp.clean_value(Counter(vals).most_common(1)[0][0]))
                elif len(vals) == 1:
                    text = cp.clean_value(vals[0])
                else:
                    text = " ; ".join(dict.fromkeys(cp.clean_value(v) for v in vals))
                cc = ws.cell(row=r, column=ci, value=text)
                cc.alignment = Alignment(horizontal="center", vertical="center")
                sev = cp.severity(" ".join(vals))
                if sev == "crit":
                    cc.fill, cc.font = crit
                elif sev == "warn":
                    cc.fill, cc.font = warn
                else:
                    cc.font = base

            ws.cell(row=r, column=COL_OF["Laboratory"], value=g["inst"]).font = small
            ws.cell(row=r, column=COL_OF["CoA code"], value=code).font = base
            ws.cell(row=r, column=COL_OF["Date of issue"], value=date).font = base
            if g["link"]:
                lc = ws.cell(row=r, column=COL_OF["PDF"], value="Open")
                lc.hyperlink = g["link"]
                lc.font = link_f
                lc.alignment = Alignment(horizontal="center")
            first = False
            r += 1
        for ci in range(1, ncol + 1):
            ws.cell(row=r - 1, column=ci).border = Border(bottom=hair)

    ws.auto_filter.ref = "A%d:%s%d" % (HROW, get_column_letter(ncol), r - 1)

    # legend — the example's own notes, extended only by the owner's conventions
    r += 1
    for txt, fill in [
        ("*BLQ — под лимит на квантификација (below limit of quantification)", None),
        ("*LOQ — лимит на квантификација (limit of quantification)", None),
        ("*n.d — not detected", None),
        ("*\"/\" — параметарот не е опфатен со тој извештај (not covered by that report)", None),
        ("*\"not tested\" — микотоксинскиот под-параметар не е анализиран на сертификат со само вкупни афлатоксини", None),
        ("*црвена боја — означува отстапка од спецификација (deviation from specification)", CRIT_BG),
        ("*портокалова боја — лабораториски наод или забелешка за интегритет на податоци", WARN_BG),
    ]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
        c = ws.cell(row=r, column=1, value=txt)
        c.font = Font(name=FONT, size=9)
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
        r += 1

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print("Wrote %s" % OUT)
    print("rows=%d batches=%d" % (r, len(batches)))


if __name__ == "__main__":
    main()
