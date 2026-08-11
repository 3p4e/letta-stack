#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Potency (Total Δ9-THC) by strain.

One table. Every product batch grouped under its strain, and for each batch every
Total Δ9-THC result on file — release assay and any later retest — each referenced to
the certificate, date and issuing laboratory it came from. A strain header band gives the
batch count and the THC range across all of that strain's production batches.
"""
import os
import re

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import coa_pivot as cp

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "exports", "PP_THC_by_Strain.xlsx")
FONT = "Arial"
INK = "16232B"
ACCENT = "0E6E6E"
BAND = "F4F6F5"

PCT = re.compile(r"(\d+(?:[.,]\d+)?)\s*(?:%|%\s*w/w|\s*%w/w)")


def thc_number(result):
    """Leading plain percentage for range/mean only; None for BLQ / <LOQ / prose."""
    if not result:
        return None
    head = result.split("(")[0]
    if "<" in head or "≤" in head or "loq" in head.lower() or "blq" in head.lower():
        return None
    m = PCT.search(head)
    if not m:
        return None
    try:
        return float(m.group(1).replace(",", "."))
    except ValueError:
        return None


def main():
    rows, batches = cp.build()
    by_seq = cp.group_by_seq(rows)

    # strain -> ordered list of batches; each batch -> its THC test rows
    strains = {}
    order = []
    for b in batches:
        strain = (b["strain"] or "(strain not stated)").strip()
        if strain not in strains:
            strains[strain] = []
            order.append(strain)
        thc_tests = []
        for rec in by_seq[b["seq"]]:
            if cp.classify(rec["Parameter"]) != "thc":
                continue
            thc_tests.append({
                "result": rec["Result"].strip(),
                "spec": cp.clean_thc_spec((rec["Acceptance Criterion"] or "").strip()),
                "code": rec["Certificate Code"].strip(),
                "date": cp.issue_date(rec["Issue Date"]),
                "lab": lab_of(rec["Issuing Institution"] or ""),
                "link": cp.clean_link(rec["Drive File Link"]),
                "num": thc_number(rec["Result"]),
            })
        strains[strain].append({"b": b, "tests": thc_tests})

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "THC by Strain"
    ws.sheet_view.showGridLines = False

    hdr = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor=INK)
    base = Font(name=FONT, size=10)
    bold = Font(name=FONT, size=10, bold=True)
    small = Font(name=FONT, size=9, color="5A6E75")
    faint = Font(name=FONT, size=10, color="8B9EA3")
    link_f = Font(name=FONT, size=9, color=ACCENT, underline="single")
    thin = Side(style="thin", color="DBE3E1")

    ws["A1"] = "Purely Plant GmbH — Total Δ⁹-THC by strain"
    ws["A1"].font = Font(name=FONT, size=16, bold=True, color=INK)
    ws["A2"] = ("Every production batch grouped by strain, with every Total Δ⁹-THC result on "
                "file, referenced to its certificate. BLQ / <LOQ values are shown but excluded "
                "from the numeric range.")
    ws["A2"].font = Font(name=FONT, size=10, italic=True, color="5A6E75")

    heads = ["Seq", "Batch number", "P-number", "Total Δ⁹-THC % w/w", "Batch THC spec",
             "CoA code", "Date of issue", "Issuing institution", "Certificate"]
    widths = [6, 22, 12, 20, 22, 20, 14, 34, 12]
    for ci, h in enumerate(heads, start=1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = hdr
        c.fill = hdr_fill
        c.alignment = Alignment(vertical="center", wrap_text=True)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 24
    ws.freeze_panes = "A5"

    r = 5
    for strain in order:
        entries = strains[strain]
        nums = [t["num"] for e in entries for t in e["tests"] if t["num"] is not None]
        rng = ""
        if nums:
            lo, hi = min(nums), max(nums)
            rng = ("%.2f %%" % lo) if lo == hi else ("%.2f – %.2f %%" % (lo, hi))
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(heads))
        h = ws.cell(row=r, column=1, value="   %s      %d batch(es)%s" % (
            strain, len(entries), ("      Total Δ⁹-THC across batches: " + rng) if rng else ""))
        h.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
        h.fill = PatternFill("solid", fgColor=ACCENT)
        h.alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 20
        r += 1

        for e in entries:
            b = e["b"]
            tests = e["tests"] or [None]
            block_start = r
            for t in tests:
                ws.cell(row=r, column=1, value=b["seq"]).font = base
                ws.cell(row=r, column=2, value=b["batch"]).font = bold
                ws.cell(row=r, column=3, value=b["p_number"]).font = base
                if t is None:
                    nc = ws.cell(row=r, column=4, value="no THC result on file")
                    nc.font = faint
                else:
                    vc = ws.cell(row=r, column=4, value=cp.clean_value(t["result"]))
                    vc.font = base
                    vc.alignment = Alignment(wrap_text=True, vertical="center")
                    ws.cell(row=r, column=5, value=t["spec"] or "").font = small
                    ws.cell(row=r, column=6, value=t["code"]).font = base
                    ws.cell(row=r, column=7, value=t["date"]).font = base
                    ic = ws.cell(row=r, column=8, value=t["lab"])
                    ic.font = small
                    ic.alignment = Alignment(wrap_text=True, vertical="center")
                    if t["link"]:
                        lc = ws.cell(row=r, column=9, value="Open")
                        lc.hyperlink = t["link"]
                        lc.font = link_f
                if len(order) and (order.index(strain) % 2 == 0):
                    for ci in range(1, len(heads) + 1):
                        if ws.cell(row=r, column=ci).fill.fgColor.rgb in (None, "00000000"):
                            ws.cell(row=r, column=ci).fill = PatternFill("solid", fgColor=BAND)
                r += 1
            # merge batch identity down its testings
            if r - 1 > block_start:
                for col in (1, 2, 3):
                    ws.merge_cells(start_row=block_start, start_column=col,
                                   end_row=r - 1, end_column=col)
                    ws.cell(row=block_start, column=col).alignment = Alignment(vertical="center")
        for ci in range(1, len(heads) + 1):
            ws.cell(row=r - 1, column=ci).border = Border(bottom=Side(style="thin", color="C6D0CD"))

    ws.auto_filter.ref = "A4:%s%d" % (get_column_letter(len(heads)), r - 1)
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print("Wrote %s" % OUT)
    print("strains=%d batches=%d rows=%d" % (len(order), len(batches), r - 5))


def lab_of(inst):
    for name, rx, _acc in cp.LAB_PATTERNS:
        if rx.search(inst or ""):
            return name.split("—")[0].strip()
    if "purely plant" in (inst or "").lower() or "in-house" in (inst or "").lower():
        return "Purely Plant (in-house)"
    return inst.strip()[:40]


if __name__ == "__main__":
    main()
