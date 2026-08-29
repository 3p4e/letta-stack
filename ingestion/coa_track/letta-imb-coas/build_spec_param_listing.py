#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Chronological per-batch, per-spec-parameter listing of every eCoA result.

One row per (batch, QCSP 001 specification parameter, certificate result), for all 77
batches, ordered by production date as encoded in the batch number (prefix + MMYY +
serial: GP0824 is August 2024, CC112501 is November 2025 batch 01). The panel is the
specification's own — the twelve numbered parameters of QCSP 001, with the grouped ones
expanded (five microbiology, three mycotoxin, four heavy-metal sub-parameters), so the
listing answers "what does the spec require and what does the file hold" line by line.

Each row carries the result value (the bare number where the result is a clean numeric,
verbatim otherwise), the unit in its own column, the eCoA document code, its issue date,
the issuing laboratory's abbreviation, and the Drive link to the certificate scan. Where
the knowledgebase holds no result for a spec parameter the row says so explicitly —
"Missing / not tested" — rather than being omitted, because an absent row reads as an
oversight and an explicit gap reads as a fact.

Data source: exports/master_coa_table.tsv, the structured extraction of the ImB_QC_COAs
knowledgebase on the Letta host (source-271bc3be…, re-verified live before that build:
263 files, all processing_status=completed — the Letta source has since been retired,
19.08.2026, and the certificates now live in RAGFlow). Nothing is re-derived or judged: results are
transcribed values with their certificates; the only computed column is the specification
grade, which is classification against the QCSP 001 v.02 ladder, not a pass/fail verdict.
Measurement uncertainty is stripped from result cells per the house rule; multiple assays
of one parameter appear as separate rows, one per certificate, never reconciled.

Spec parameters the outsourced laboratories never test (macroscopic/microscopic/HPTLC
identification, foreign matter — determined in-house) will read Missing on most batches;
the three R&D batches whose in-house CoAs were ingested on instruction carry what those
printed. The "Source parameter (as printed)" column preserves the certificate's own
wording wherever a spec row was matched, so the mapping is auditable.
"""
import csv
import os
import re
import sys
from collections import OrderedDict, defaultdict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import build_house_specs as hs
import coa_pivot as cp

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_XLSX = os.path.join(HERE, "exports", "PP_Spec_Parameter_Listing.xlsx")
OUT_TSV = os.path.join(HERE, "exports", "PP_Spec_Parameter_Listing.tsv")

# A dated verification record, not a live pointer: this is what was checked, and
# when. The Letta source was retired 19.08.2026 — repointing the note at RAGFlow
# would claim a verification that was never run there.
SOURCE_NOTE = ("ImB_QC_COAs · source-271bc3be-10d1-4541-8a5b-be3f6fab7c97 · 263 files, "
               "all completed (verified live 13.08.2026; that Letta source was "
               "retired 19.08.2026 — the certificates are now in RAGFlow)")

# The QCSP 001 panel, expanded. (item №, spec parameter, unit, pivot keys that feed it)
SPEC_PANEL = [
    ("1",    "Macroscopic Identification (Test A)", "—",      ("appearance",)),
    ("2",    "Microscopic Identification (Test B)", "—",      ()),
    ("3",    "HPTLC / Chromatographic Identification (Test C)", "—", ("identification",)),
    ("4",    "Assay — Total Δ⁹-THC",                "%",      ("thc",)),
    ("5",    "Assay — Total CBD",                   "%",      ("cbd",)),
    ("6",    "Total CBN",                           "%",      ("cbn",)),
    ("7",    "Foreign Matter",                      "%",      ("foreign_matter",)),
    ("8",    "Loss on Drying",                      "%",      ("loss_on_drying",)),
    ("9.1",  "Microbiology — TAMC",                 "CFU/g",  ("tamc",)),
    ("9.2",  "Microbiology — TYMC",                 "CFU/g",  ("tymc",)),
    ("9.3",  "Microbiology — Bile-tolerant gram-negative", "CFU/g", ("bile",)),
    ("9.4",  "Microbiology — Salmonella",           "/25 g",  ("salmonella",)),
    ("9.5",  "Microbiology — Escherichia coli",     "/1 g",   ("ecoli",)),
    ("10.1", "Mycotoxins — Aflatoxin B1",           "µg/kg",  ("afla_b1",)),
    ("10.2", "Mycotoxins — Total Aflatoxins (B1+B2+G1+G2)", "µg/kg",
             ("aflatoxins", "afla_b2", "afla_g1", "afla_g2")),
    ("10.3", "Mycotoxins — Ochratoxin A",           "µg/kg",  ("ochratoxin",)),
    ("11.1", "Heavy Metals — Lead (Pb)",            "mg/kg",  ("pb",)),
    ("11.2", "Heavy Metals — Cadmium (Cd)",         "mg/kg",  ("cd",)),
    ("11.3", "Heavy Metals — Arsenic (As)",         "mg/kg",  ("as_",)),
    ("11.4", "Heavy Metals — Mercury (Hg)",         "mg/kg",  ("hg",)),
    ("12",   "Pesticide Residues — multi-residue panel", "—", ("pesticides",)),
]
KEY_TO_ITEM = {}
for _no, _name, _unit, _keys in SPEC_PANEL:
    for _k in _keys:
        KEY_TO_ITEM[_k] = (_no, _name, _unit)

STRAIN_ABBR = {s: v[0] for s, v in hs.PUBLISHED.items()}

LAB_ABBR = [
    ("farmahem", "FARM"),
    ("ukim", "UKIM"), ("faculty of pharmacy", "UKIM"), ("natural products", "UKIM"),
    ("institute of public health", "IPH"), ("институт за јавно здравје", "IPH"),
    ("iph", "IPH"), ("ijz", "IPH"),
    ("phytosanitary", "SPL"), ("фитосанитарна", "SPL"),
    ("purely plant", "PP"), ("in-house", "PP"),
]

UNC_RE = re.compile(r"\(\s*[Uu]\s*[:=]?\s*±?\s*[\d.,]+\s*%?\s*(?:w/w)?"
                    r"\s*(?:,\s*k\s*=\s*\d+)?\s*\)")
NUM_RE = re.compile(r"^\s*([<≤>≥]?\s*\d+(?:[.,]\d+)?)\s*"
                    r"(%\s*w/w|%|µg/kg|ug/kg|mg/kg|CFU/g|cfu/g)?\s*$")
# batch codes embed MMYY; J31 is the one prefix that itself contains digits
DATE_RE = re.compile(r"^(J31|[A-Z]+)[\s_]*(\d{2})(\d{2})(\d*)")
NO_FINDING_RE = re.compile(r"^\s*(?:[<≤]\s*[\d.,]+.*|N\.?D\.?|ND|BLQ|n\.?d\.?|"
                           r"<\s*LOQ.*|не е откриен.*)\s*$", re.I)


def lab_abbr(inst):
    low = (inst or "").strip().lower()
    if not low or low == "n/a":
        return "—"
    for token, ab in LAB_ABBR:
        if token in low:
            return ab
    return (inst or "").strip()[:12]


def prod_key(batch):
    """(year, month, serial) parsed from the batch code; unparseable sorts last."""
    b = (batch or "").strip().upper()
    m = DATE_RE.match(b.replace("/", "").replace("-", ""))
    if not m:
        return (9999, 99, 999, b)
    mm, yy = int(m.group(2)), int(m.group(3))
    if not 1 <= mm <= 12:
        return (9999, 99, 999, b)
    serial = int(m.group(4)[:2]) if m.group(4)[:2].isdigit() and m.group(4) else 0
    return (2000 + yy, mm, serial, b)


def clean_value(raw, spec_unit):
    """Result cell: the bare number where the result is one, verbatim otherwise.

    Uncertainty is stripped (house rule: it does not belong in the result cell); a value
    that is qualifier+number+unit collapses to qualifier+number because the unit has its
    own column. Anything else — 'Одговара (absent)', dual readings, text findings — is
    kept verbatim so nothing the laboratory wrote is lost.
    """
    txt = UNC_RE.sub("", (raw or "")).strip().rstrip("·").strip()
    m = NUM_RE.match(txt)
    if m:
        return m.group(1).replace(" ", "").replace(",", ".")
    if spec_unit != "—" and txt.endswith(spec_unit):
        head = txt[: -len(spec_unit)].strip()
        if NUM_RE.match(head):
            return NUM_RE.match(head).group(1).replace(" ", "").replace(",", ".")
    return txt


def batch_grade(strain, thc_rows):
    """Classification against the QCSP 001 v.02 ladder — never a pass/fail verdict."""
    pub = hs.PUBLISHED.get(strain)
    if not pub:
        return "—", "—", "—"
    abbr = pub[0]
    ladder, _added = hs.extend(pub[3])
    spec_doc = "QCSP 001_%s_v.02" % abbr
    vals = []
    for r in thc_rows:
        head = UNC_RE.sub("", r["Result"].split("(")[0])
        if any(t in head.lower() for t in ("<", "≤", "loq", "blq")):
            continue
        m = re.search(r"(\d+(?:[.,]\d+)?)", head)
        if m:
            vals.append(float(m.group(1).replace(",", ".")))
    if not vals:
        return spec_doc, "no assay on file", "—"
    grades = []
    for v in vals:
        g = hs.grade_of(ladder, v)
        if g and g not in grades:
            grades.append(g)
    if not grades:
        return spec_doc, "outside every grade", "—"
    codes = []
    for g in grades:
        for label, nominal, _f, _c in ladder:
            if label == g:
                codes.append(hs.code(abbr, nominal))
    return spec_doc, " / ".join("Grade " + g for g in grades), " / ".join(codes)


def build_rows():
    rows, batches = cp.build()
    by_seq = cp.group_by_seq(rows)

    listing = []
    stats = {"results": 0, "missing": 0}

    ordered = sorted(batches, key=lambda b: prod_key(b["batch"]))
    for chrono, b in enumerate(ordered, start=1):
        recs = [r for r in by_seq[b["seq"]]
                if not cp.STABILITY_RE.search(r["Parameter"] or "")]
        strain = (b["strain"] or "").strip()

        by_item = defaultdict(list)
        for r in recs:
            kind = cp.classify(r["Parameter"])
            if kind in KEY_TO_ITEM:
                by_item[KEY_TO_ITEM[kind][0]].append(r)

        thc_rows = by_item.get("4", [])
        spec_doc, grade, prod_code = batch_grade(strain, thc_rows)

        y, mth = prod_key(b["batch"])[:2]
        prod = "%04d-%02d" % (y, mth) if y != 9999 else "—"

        base = {
            "chrono": chrono, "prod": prod, "batch": b["batch"],
            "p": (b["p_number"] or "").strip() or "—", "strain": strain,
            "spec_doc": spec_doc, "grade": grade, "code": prod_code,
        }

        for no, name, unit, keys in SPEC_PANEL:
            hits = by_item.get(no, [])
            if not hits:
                listing.append(dict(base, no=no, param=name,
                                    value="Missing / not tested", unit="—",
                                    cert="—", date="—", lab="—", link="",
                                    printed="—"))
                stats["missing"] += 1
                continue

            if no == "12":
                # the spec has one pesticide row; the certificates have one row per
                # analyte, so summarise per certificate and spell out any finding
                per_cert = OrderedDict()
                for r in hits:
                    c = (r["Certificate Code"] or "").strip() or "—"
                    per_cert.setdefault(c, []).append(r)
                for c, group in per_cert.items():
                    findings = [g for g in group
                                if not NO_FINDING_RE.match((g["Result"] or "").strip())]
                    if findings:
                        val = "; ".join("%s: %s" % (cp.short_label(g["Parameter"]),
                                                    UNC_RE.sub("", g["Result"]).strip())
                                        for g in findings)
                    else:
                        val = "ND / <LOQ — no analyte detected (%d analytes)" % len(group)
                    g0 = group[0]
                    listing.append(dict(base, no=no, param=name, value=val, unit="—",
                                        cert=c, date=cp.issue_date(g0["Issue Date"]),
                                        lab=lab_abbr(g0["Issuing Institution"]),
                                        link=cp.clean_link(g0["Drive File Link"]),
                                        printed="multi-residue screen, per analyte"))
                    stats["results"] += 1
                continue

            hits.sort(key=lambda r: cp.issue_date(r["Issue Date"]) or "9999")
            for r in hits:
                cert = (r["Certificate Code"] or "").strip() or "—"
                listing.append(dict(
                    base, no=no, param=name,
                    value=clean_value(r["Result"], unit), unit=unit,
                    cert=cert, date=cp.issue_date(r["Issue Date"]) or "—",
                    lab=lab_abbr(r["Issuing Institution"]),
                    link=cp.clean_link(r["Drive File Link"]),
                    printed=(r["Parameter"] or "").strip() or "—"))
                stats["results"] += 1

    return listing, stats, len(ordered)


HEADERS = ["#", "Production", "Batch", "P-Number", "Strain", "Spec document",
           "Grade (v.02)", "Product code", "№", "Specification parameter",
           "Result", "Unit", "eCoA code", "Issued", "Lab", "eCoA scan",
           "Source parameter (as printed)"]
FIELDS = ["chrono", "prod", "batch", "p", "strain", "spec_doc", "grade", "code",
          "no", "param", "value", "unit", "cert", "date", "lab", "link", "printed"]
WIDTHS = [5, 10, 13, 10, 19, 17, 22, 26, 6, 38, 34, 8, 20, 11, 6, 10, 34]

NAVY = "1B3A5C"
GOLD_TINT = "FBF6E9"
BAND = "F4F7FB"
MISS = "9AA7B4"


def write_xlsx(listing):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Spec parameter listing"
    thin = Side(style="thin", color="D6DEEA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(["Purely Plant — eCoA results per QCSP 001 specification parameter, "
               "all batches, chronological by production date encoded in the batch "
               "number. Source: %s. Values as transcribed; uncertainty excluded; "
               "multiple assays are separate rows; grade is classification against "
               "the v.02 ladder, not a disposition." % SOURCE_NOTE])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    ws.cell(1, 1).font = Font(name="Arial", size=8, italic=True, color="555555")
    ws.cell(1, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[1].height = 30

    ws.append(HEADERS)
    for cidx in range(1, len(HEADERS) + 1):
        c = ws.cell(2, cidx)
        c.font = Font(name="Arial", size=8, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    prev_batch = None
    band = False
    for rec in listing:
        if rec["batch"] != prev_batch:
            band = not band
            prev_batch = rec["batch"]
        ridx = ws.max_row + 1
        for cidx, f in enumerate(FIELDS, start=1):
            c = ws.cell(ridx, cidx)
            v = rec[f]
            if f == "link":
                if v:
                    c.value = "open"
                    c.hyperlink = v
                    c.font = Font(name="Arial", size=8, color="0563C1", underline="single")
                else:
                    c.value = "—"
                    c.font = Font(name="Arial", size=8, color=MISS)
            else:
                c.value = v
                missing = rec["value"].startswith("Missing")
                c.font = Font(name="Arial", size=8,
                              color=MISS if missing and f in ("value", "unit", "cert",
                                                              "date", "lab") else "16232B",
                              bold=(f == "batch"))
            if band and f != "link":
                c.fill = PatternFill("solid", start_color=BAND)
            if f in ("no", "unit", "lab", "date", "chrono", "prod"):
                c.alignment = Alignment(horizontal="center", vertical="top")
            else:
                c.alignment = Alignment(vertical="top", wrap_text=(f in ("value", "param",
                                                                         "printed", "code",
                                                                         "grade")))
            c.border = border

    for cidx, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(cidx)].width = w
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = "A2:%s%d" % (get_column_letter(len(HEADERS)), ws.max_row)
    wb.save(OUT_XLSX)


def main():
    listing, stats, n_batches = build_rows()

    with open(OUT_TSV, "w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh, delimiter="\t", lineterminator="\n")
        w.writerow(HEADERS)
        for rec in listing:
            w.writerow([rec[f] for f in FIELDS])

    write_xlsx(listing)

    print("batches: %d (chronological)   rows: %d   results: %d   missing: %d"
          % (n_batches, len(listing), stats["results"], stats["missing"]))
    print("wrote %s (%d KB)" % (os.path.relpath(OUT_XLSX, HERE),
                                os.path.getsize(OUT_XLSX) // 1024))
    print("wrote %s (%d KB)" % (os.path.relpath(OUT_TSV, HERE),
                                os.path.getsize(OUT_TSV) // 1024))
    return 0


if __name__ == "__main__":
    sys.exit(main())
